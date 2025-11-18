# filepath: apps/inventario/views.py
from datetime import date, timedelta

from django.db.models import Sum, Min, F, Q, Count
from django.http import HttpResponse

from rest_framework import viewsets, filters
from rest_framework.decorators import api_view, permission_classes, action
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from rest_framework.exceptions import MethodNotAllowed, PermissionDenied

from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

from openpyxl import load_workbook

from .models import Producto, Lote, Movimiento
from .serializers import ProductoSerializer, LoteSerializer, MovimientoSerializer
from .permissions import (
    IsCatalogEditorOrReadOnly,
    IsInventoryEditorOrReadOnly,
)


def home(request):
    return HttpResponse("✅ Módulo Inventario funcionando correctamente.")


# --------------------------------------------------------------------
# PRODUCTOS
# --------------------------------------------------------------------
class ProductoViewSet(viewsets.ModelViewSet):
    queryset = Producto.objects.all().annotate(
        stock_total=Sum("lotes__existencias"),
        proximo_vencer=Min("lotes__fecha_caducidad"),
    )
    serializer_class = ProductoSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ["clave", "descripcion"]
    ordering_fields = ["clave", "descripcion"]
    permission_classes = [IsCatalogEditorOrReadOnly]


# --------------------------------------------------------------------
# LOTES
# --------------------------------------------------------------------
class LoteViewSet(viewsets.ModelViewSet):
    queryset = Lote.objects.select_related("producto").all()
    serializer_class = LoteSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ["producto__clave", "numero_lote"]
    ordering_fields = ["fecha_caducidad", "existencias"]
    permission_classes = [IsInventoryEditorOrReadOnly]

    @action(detail=True, methods=["get"], url_path="cardex_pdf")
    def cardex_pdf(self, request, pk=None):
        """
        GET /api/lotes/<pk>/cardex_pdf/
        Genera el Kárdex en PDF para el lote indicado.
        Acepta filtros por query params:
          - tipo = entrada|salida|ajuste|all
          - fecha_desde (YYYY-MM-DD)
          - fecha_hasta (YYYY-MM-DD)
        """
        # 1) Recuperar lote
        try:
            lote = Lote.objects.select_related("producto").get(pk=pk)
        except Lote.DoesNotExist:
            return Response({"detail": "Lote no encontrado."}, status=404)

        # 2) Query de movimientos del lote (solo activos)
        qs = (
            Movimiento.objects.filter(lote_id=pk, activo=True)
            .order_by("fecha", "id")
            .select_related("lote", "lote__producto")
        )

        # Filtro por tipo (desde el front llega en minúsculas)
        tipo = request.query_params.get("tipo")
        if tipo and tipo.lower() != "all":
            tipo_map = {
                "entrada": "entrada",
                "salida": "salida",
                "ajuste": "ajuste",
            }
            db_value = tipo_map.get(tipo.lower())
            if db_value:
                qs = qs.filter(tipo=db_value)

        # Filtros por fecha
        fecha_desde = request.query_params.get("fecha_desde")
        fecha_hasta = request.query_params.get("fecha_hasta")

        if fecha_desde:
            qs = qs.filter(fecha__date__gte=fecha_desde)
        if fecha_hasta:
            qs = qs.filter(fecha__date__lte=fecha_hasta)

        movimientos = list(qs)

        # 3) Calcular saldos desde las existencias actuales del lote
        def delta_mov(m):
            if m.tipo == "entrada":
                return m.cantidad
            if m.tipo == "salida":
                return -m.cantidad
            return m.cantidad  # ajuste

        running_saldo = lote.existencias
        deltas = [delta_mov(m) for m in movimientos]
        saldos = [0] * len(movimientos)

        # Recorremos del final hacia el inicio
        for i in range(len(movimientos) - 1, -1, -1):
            saldos[i] = running_saldo  # saldo DESPUÉS de ese movimiento
            running_saldo -= deltas[i]

        # 4) Construir PDF
        response = HttpResponse(content_type="application/pdf")
        filename = f"cardex_lote_{lote.numero_lote}.pdf"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        p = canvas.Canvas(response, pagesize=letter)
        width, height = letter
        y = height - 50

        # Encabezado
        p.setFont("Helvetica-Bold", 13)
        p.drawString(40, y, "KÁRDEX DE INVENTARIO POR LOTE")
        y -= 22

        p.setFont("Helvetica", 10)
        p.drawString(
            40,
            y,
            f"Producto: {lote.producto.clave} - {lote.producto.descripcion[:80]}",
        )
        y -= 14
        p.drawString(40, y, f"Lote: {lote.numero_lote}")
        y -= 14
        p.drawString(40, y, f"Existencias actuales: {lote.existencias}")
        y -= 22

        # Encabezado de tabla
        def draw_header():
            nonlocal y
            p.setFont("Helvetica-Bold", 9)
            p.drawString(40, y, "Fecha")
            p.drawString(110, y, "Tipo")
            p.drawString(180, y, "Cant.")
            p.drawString(230, y, "Saldo")
            p.drawString(280, y, "Unidad médica")
            p.drawString(420, y, "Observaciones")
            y -= 12
            p.line(40, y + 10, width - 40, y + 10)
            p.setFont("Helvetica", 8)

        draw_header()

        # Filas
        for m, saldo in zip(movimientos, saldos):
            if y < 60:
                p.showPage()
                y = height - 50
                draw_header()

            fecha_str = m.fecha.strftime("%d/%m/%Y")
            p.drawString(40, y, fecha_str)
            p.drawString(110, y, m.get_tipo_display())
            p.drawRightString(215, y, str(m.cantidad))
            p.drawRightString(265, y, str(saldo))
            p.drawString(280, y, (m.unidad_medica or "")[:25])
            p.drawString(420, y, (m.observaciones or "")[:60])
            y -= 11

        p.showPage()
        p.save()
        return response


# --------------------------------------------------------------------
# MOVIMIENTOS
# --------------------------------------------------------------------
class MovimientoViewSet(viewsets.ModelViewSet):
    # Solo movimientos activos
    queryset = Movimiento.objects.select_related("lote", "lote__producto").filter(
        activo=True
    )
    serializer_class = MovimientoSerializer
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    # 🔍 campos que se buscan con ?search=
    search_fields = [
        "lote__producto__clave",        # clave del medicamento
        "lote__producto__descripcion",  # descripción
        "lote__numero_lote",            # número de lote
        "unidad_medica",                # unidad médica
        "tipo",                         # tipo movimiento
    ]
    ordering_fields = ["fecha", "cantidad"]
    permission_classes = [IsInventoryEditorOrReadOnly]

    # 🔒 No se permite editar movimientos: solo crear/listar.
    #     Para correcciones usar movimientos de tipo "ajuste".
    def update(self, request, *args, **kwargs):
        raise MethodNotAllowed(
            "PUT",
            detail=(
                "La edición de movimientos no está permitida. "
                'Para corregir el stock, registra un movimiento de tipo "ajuste".'
            ),
        )

    def partial_update(self, request, *args, **kwargs):
        raise MethodNotAllowed(
            "PATCH",
            detail=(
                "La edición parcial de movimientos no está permitida. "
                'Para corregir el stock, registra un movimiento de tipo "ajuste".'
            ),
        )

    # 🔒 Borrado lógico: solo superusuario y se marca activo=False
    def destroy(self, request, *args, **kwargs):
        user = request.user
        if not user.is_superuser:
            raise PermissionDenied(
                "Solo el superusuario puede eliminar movimientos. "
                'Para corregir existencias, usa un movimiento de tipo "ajuste".'
            )

        instance = self.get_object()
        instance.activo = False
        instance.save(update_fields=["activo"])
        # No tocamos existencias: cualquier corrección debe hacerse con "ajuste".
        return Response(status=204)


# --------------------------------------------------------------------
# ENDPOINT DE REPORTES POR PERIODO
# --------------------------------------------------------------------
@api_view(["GET"])
@permission_classes([IsInventoryEditorOrReadOnly])
def reportes_movimientos(request):
    """
    GET /api/reportes/movimientos/

    Parámetros de consulta:
      - fecha_desde: YYYY-MM-DD
      - fecha_hasta: YYYY-MM-DD
      - tipo: entrada|salida|ajuste|all
      - search: texto libre (clave, descripción, unidad, tipo)
      - unidad_medica
      - producto_clave
      - lote_id
      - formato: json (default) | pdf
    """
    # Solo movimientos activos
    qs = Movimiento.objects.select_related("lote", "lote__producto").filter(
        activo=True
    )

    tipo = request.query_params.get("tipo")
    search = request.query_params.get("search")
    fecha_desde = request.query_params.get("fecha_desde")
    fecha_hasta = request.query_params.get("fecha_hasta")
    unidad_medica = request.query_params.get("unidad_medica")
    producto_clave = request.query_params.get("producto_clave")
    lote_id = request.query_params.get("lote_id")
    formato = (request.query_params.get("formato") or "json").lower()

    # Filtro por tipo
    if tipo and tipo.lower() != "all":
        qs = qs.filter(tipo=tipo.lower())

    # Filtros por fecha
    if fecha_desde:
        qs = qs.filter(fecha__date__gte=fecha_desde)
    if fecha_hasta:
        qs = qs.filter(fecha__date__lte=fecha_hasta)

    # Unidad médica
    if unidad_medica:
        qs = qs.filter(unidad_medica__icontains=unidad_medica)

    # Producto por clave
    if producto_clave:
        qs = qs.filter(lote__producto__clave__icontains=producto_clave)

    # Lote específico
    if lote_id:
        qs = qs.filter(lote_id=lote_id)

    # Búsqueda general
    if search:
        qs = qs.filter(
            Q(lote__producto__clave__icontains=search)
            | Q(lote__producto__descripcion__icontains=search)
            | Q(unidad_medica__icontains=search)
            | Q(tipo__icontains=search)
        )

    qs = qs.order_by("fecha", "id")

    # Resumen agregados
    agg = qs.aggregate(
        total_movimientos=Count("id"),
        total_entradas=Sum("cantidad", filter=Q(tipo="entrada")),
        total_salidas=Sum("cantidad", filter=Q(tipo="salida")),
        total_ajustes=Sum("cantidad", filter=Q(tipo="ajuste")),
    )

    for key in ["total_movimientos", "total_entradas", "total_salidas", "total_ajustes"]:
        if agg[key] is None:
            agg[key] = 0

    # Agrupado por producto
    por_producto_qs = (
        qs.values("lote__producto__clave", "lote__producto__descripcion")
        .annotate(
            total_entradas=Sum("cantidad", filter=Q(tipo="entrada")),
            total_salidas=Sum("cantidad", filter=Q(tipo="salida")),
            total_ajustes=Sum("cantidad", filter=Q(tipo="ajuste")),
        )
        .order_by("lote__producto__clave")
    )

    por_producto = [
        {
            "clave": row["lote__producto__clave"],
            "descripcion": row["lote__producto__descripcion"],
            "total_entradas": row["total_entradas"] or 0,
            "total_salidas": row["total_salidas"] or 0,
            "total_ajustes": row["total_ajustes"] or 0,
        }
        for row in por_producto_qs
    ]

    # Si piden PDF, devolvemos PDF
    if formato == "pdf":
        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = 'attachment; filename="reporte_movimientos.pdf"'

        p = canvas.Canvas(response, pagesize=letter)
        width, height = letter
        y = height - 50

        p.setFont("Helvetica-Bold", 13)
        p.drawString(40, y, "REPORTE DE MOVIMIENTOS DE INVENTARIO")
        y -= 20

        p.setFont("Helvetica", 10)
        p.drawString(40, y, f"Total movimientos: {agg['total_movimientos']}")
        y -= 14
        p.drawString(40, y, f"Total entradas: {agg['total_entradas']}")
        y -= 14
        p.drawString(40, y, f"Total salidas: {agg['total_salidas']}")
        y -= 14
        saldo_neto = (agg["total_entradas"] or 0) - (agg["total_salidas"] or 0)
        p.drawString(40, y, f"Saldo neto (entradas - salidas): {saldo_neto}")
        y -= 24

        # Encabezado tabla
        def draw_header():
            nonlocal y
            p.setFont("Helvetica-Bold", 9)
            p.drawString(40, y, "Clave")
            p.drawString(120, y, "Descripción")
            p.drawRightString(380, y, "Entradas")
            p.drawRightString(450, y, "Salidas")
            p.drawRightString(520, y, "Ajustes")
            y -= 12
            p.line(40, y + 10, width - 40, y + 10)
            p.setFont("Helvetica", 8)

        draw_header()

        for row in por_producto:
            if y < 60:
                p.showPage()
                y = height - 50
                draw_header()

            p.drawString(40, y, (row["clave"] or "")[:15])
            p.drawString(120, y, (row["descripcion"] or "")[:40])
            p.drawRightString(380, y, str(row["total_entradas"]))
            p.drawRightString(450, y, str(row["total_salidas"]))
            p.drawRightString(520, y, str(row["total_ajustes"]))
            y -= 11

        p.showPage()
        p.save()
        return response

    # Respuesta JSON normal
    detalle = MovimientoSerializer(qs, many=True).data

    return Response(
        {
            "filters": {
                "tipo": tipo or "all",
                "search": search or "",
                "fecha_desde": fecha_desde,
                "fecha_hasta": fecha_hasta,
                "unidad_medica": unidad_medica,
                "producto_clave": producto_clave,
                "lote_id": lote_id,
            },
            "resumen": agg,
            "por_producto": por_producto,
            "detalle": detalle,
        }
    )


# --------------------------------------------------------------------
# DASHBOARD
# --------------------------------------------------------------------
@api_view(["GET"])
def dashboard_resumen(request):
    hoy = date.today()
    en_30 = hoy + timedelta(days=30)

    stock_total = Lote.objects.aggregate(s=Sum("existencias"))["s"] or 0
    por_agotar = (
        Producto.objects.annotate(stock=Sum("lotes__existencias"))
        .filter(stock__lt=F("stock_minimo"))
        .count()
    )
    por_vencer = Lote.objects.filter(
        fecha_caducidad__isnull=False, fecha_caducidad__lte=en_30
    ).count()

    # Solo movimientos activos
    ultimos_movs = (
        Movimiento.objects.filter(activo=True)
        .order_by("-fecha")[:10]
    )
    ultimos_movs = MovimientoSerializer(ultimos_movs, many=True).data

    return Response(
        {
            "kpi": {
                "stock_total": stock_total,
                "productos_bajo_stock": por_agotar,
                "lotes_por_vencer_30d": por_vencer,
            },
            "ultimos_movimientos": ultimos_movs,
        }
    )


# --------------------------------------------------------------------
# INFO DE USUARIO PARA REACT
# --------------------------------------------------------------------
@api_view(["GET"])
@permission_classes([AllowAny])  # nunca 403 aquí
def user_info(request):
    user = request.user

    if not user.is_authenticated:
        return Response(
            {
                "is_authenticated": False,
                "username": None,
                "groups": [],
                "is_superuser": False,
                "can_edit_catalog": False,
                "can_edit_inventory": False,
            }
        )

    groups = list(user.groups.values_list("name", flat=True))

    def in_group(names):
        return any(name in groups for name in names)

    can_edit_catalog = user.is_superuser or in_group(
        ["ADMIN_SISTEMA", "FARMACIA_CENTRAL"]
    )
    can_edit_inventory = user.is_superuser or in_group(
        ["ADMIN_SISTEMA", "FARMACIA_CENTRAL", "FARMACIA_UM"]
    )

    return Response(
        {
            "is_authenticated": True,
            "username": user.username,
            "groups": groups,
            "is_superuser": user.is_superuser,
            "can_edit_catalog": can_edit_catalog,
            "can_edit_inventory": can_edit_inventory,
        }
    )


# --------------------------------------------------------------------
# IMPORTACIÓN DE PRODUCTOS DESDE EXCEL
# --------------------------------------------------------------------
@api_view(["POST"])
@permission_classes([IsCatalogEditorOrReadOnly])
def importar_productos_excel(request):
    """
    Espera un archivo Excel (.xlsx) con columnas:
    clave, descripcion, unidad_medida, precio_unitario, stock_minimo
    en la primera fila.
    """
    file = request.FILES.get("file")
    if not file:
        return Response({"detail": "No se recibió archivo."}, status=400)

    try:
        wb = load_workbook(file)
        sheet = wb.active

        # Leer encabezados
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        header_index = {h: i for i, h in enumerate(headers) if h}

        required = [
            "clave",
            "descripcion",
            "unidad_medida",
            "precio_unitario",
            "stock_minimo",
        ]
        faltantes = [h for h in required if h not in header_index]
        if faltantes:
            return Response(
                {"detail": f"Faltan columnas requeridas: {', '.join(faltantes)}"},
                status=400,
            )

        creados = 0
        actualizados = 0

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or all(v is None for v in row):
                continue

            clave = row[header_index["clave"]]
            if not clave:
                continue

            descripcion = row[header_index["descripcion"]] or ""
            unidad = row[header_index["unidad_medida"]] or ""
            precio = row[header_index["precio_unitario"]] or 0
            stock_minimo = row[header_index["stock_minimo"]] or 0

            obj, created = Producto.objects.update_or_create(
                clave=str(clave).strip(),
                defaults={
                    "descripcion": str(descripcion).strip(),
                    "unidad_medida": str(unidad).strip(),
                    "precio_unitario": precio,
                    "stock_minimo": stock_minimo,
                },
            )
            if created:
                creados += 1
            else:
                actualizados += 1

        return Response(
            {
                "detail": "Importación completada.",
                "creados": creados,
                "actualizados": actualizados,
            }
        )
    except Exception as e:
        return Response({"detail": f"Error procesando el archivo: {e}"}, status=400)

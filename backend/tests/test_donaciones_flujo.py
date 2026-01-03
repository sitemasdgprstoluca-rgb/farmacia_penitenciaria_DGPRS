# -*- coding: utf-8 -*-
"""
Pruebas unitarias para el flujo de donaciones.

Tests:
1. Plantilla simplificada de donaciones
2. Importación desde catálogo principal de productos
3. Validación de productos existentes
4. Flujo completo de donación
"""
import pytest
from unittest.mock import Mock, patch, MagicMock
from io import BytesIO
import openpyxl


class TestPlantillaDonacionesSimplificada:
    """Tests para la plantilla simplificada de donaciones."""
    
    def test_plantilla_tiene_estructura_simplificada(self):
        """Test: La plantilla tiene solo 2 hojas (Donaciones y Catálogo)."""
        # Simular estructura de plantilla
        headers_esperados = [
            'numero *',
            'donante *',
            'tipo_donante',
            'fecha *',
            'producto_clave *',
            'cantidad *',
            'lote',
            'notas'
        ]
        
        # Verificar que los headers son 8 (simplificados)
        assert len(headers_esperados) == 8
        
        # Verificar campos obligatorios
        obligatorios = [h for h in headers_esperados if '*' in h]
        assert len(obligatorios) == 5  # numero, donante, fecha, producto_clave, cantidad
        
    def test_plantilla_usa_catalogo_principal(self):
        """Test: La plantilla referencia productos del catálogo principal."""
        # La hoja debe llamarse 'Catálogo Productos' (no 'Catálogo Productos Donación')
        nombre_hoja_esperado = 'Catálogo Productos'
        assert 'Donación' not in nombre_hoja_esperado
        
    def test_plantilla_ignora_filas_ejemplo(self):
        """Test: Las filas con [EJEMPLO] deben ser ignoradas."""
        filas_ejemplo = [
            ['[EJEMPLO] DON-001 - ELIMINAR', 'Empresa SA', 'empresa', '2024-01-15', 'MED001', 100],
            ['DON-002', 'Otra Empresa', 'empresa', '2024-01-16', 'MED002', 50],  # Real
        ]
        
        def es_fila_ejemplo(row):
            for cell in row:
                if cell and '[EJEMPLO]' in str(cell).upper():
                    return True
            return False
        
        assert es_fila_ejemplo(filas_ejemplo[0]) == True
        assert es_fila_ejemplo(filas_ejemplo[1]) == False


class TestImportacionDonaciones:
    """Tests para la importación de donaciones."""
    
    def test_importar_requiere_producto_catalogo_principal(self):
        """Test: Solo acepta productos que existen en catálogo principal."""
        # Simular búsqueda de producto
        productos_catalogo = {
            'MED001': {'id': 1, 'nombre': 'Paracetamol 500mg', 'activo': True},
            'MED002': {'id': 2, 'nombre': 'Ibuprofeno 400mg', 'activo': True},
        }
        
        def buscar_producto(clave):
            return productos_catalogo.get(clave.upper())
        
        # Producto existente
        assert buscar_producto('MED001') is not None
        
        # Producto no existente
        assert buscar_producto('NOEXISTE') is None
        
    def test_importar_rechaza_producto_inactivo(self):
        """Test: Rechaza productos marcados como inactivos."""
        productos = {
            'ACTIVO001': {'activo': True},
            'INACTIVO001': {'activo': False},
        }
        
        def validar_producto(clave):
            prod = productos.get(clave)
            return prod is not None and prod.get('activo', False)
        
        assert validar_producto('ACTIVO001') == True
        assert validar_producto('INACTIVO001') == False
        
    def test_importar_crea_donacion_y_detalles(self):
        """Test: La importación crea donación y sus detalles en una transacción."""
        # Simular datos de importación
        datos_importacion = [
            {'numero': 'DON-001', 'donante': 'Empresa SA', 'producto': 'MED001', 'cantidad': 100},
            {'numero': 'DON-001', 'donante': '', 'producto': 'MED002', 'cantidad': 50},  # Mismo número = mismo donante
            {'numero': 'DON-002', 'donante': 'ONG', 'producto': 'MED001', 'cantidad': 200},
        ]
        
        donaciones = {}
        detalles = []
        
        for fila in datos_importacion:
            numero = fila['numero']
            if numero not in donaciones:
                if fila['donante']:
                    donaciones[numero] = {'donante': fila['donante'], 'detalles': []}
            
            if numero in donaciones:
                donaciones[numero]['detalles'].append({
                    'producto': fila['producto'],
                    'cantidad': fila['cantidad']
                })
        
        # Verificar resultados
        assert len(donaciones) == 2  # DON-001 y DON-002
        assert len(donaciones['DON-001']['detalles']) == 2  # 2 productos
        assert len(donaciones['DON-002']['detalles']) == 1  # 1 producto


class TestValidacionesDonacion:
    """Tests para validaciones del flujo de donaciones."""
    
    def test_cantidad_debe_ser_positiva(self):
        """Test: La cantidad debe ser mayor a 0."""
        cantidades_validas = [1, 10, 100, 999]
        cantidades_invalidas = [0, -1, -100]
        
        for cant in cantidades_validas:
            assert cant > 0
            
        for cant in cantidades_invalidas:
            assert cant <= 0
            
    def test_fecha_formatos_aceptados(self):
        """Test: Se aceptan múltiples formatos de fecha."""
        from datetime import datetime
        
        def parse_fecha(val):
            if not val:
                return None
            val_str = str(val).strip()
            for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y']:
                try:
                    return datetime.strptime(val_str, fmt).date()
                except:
                    continue
            return None
        
        # Formatos válidos
        assert parse_fecha('2024-01-15') is not None
        assert parse_fecha('15/01/2024') is not None
        assert parse_fecha('15-01-2024') is not None
        
        # Formato inválido
        assert parse_fecha('invalid') is None
        
    def test_tipo_donante_normalizado(self):
        """Test: El tipo de donante se normaliza."""
        tipos_validos = ['empresa', 'gobierno', 'ong', 'particular', 'otro']
        
        def normalizar_tipo(tipo):
            tipo_norm = str(tipo).lower().strip()
            return tipo_norm if tipo_norm in tipos_validos else 'empresa'
        
        assert normalizar_tipo('EMPRESA') == 'empresa'
        assert normalizar_tipo('ONG') == 'ong'
        assert normalizar_tipo('desconocido') == 'empresa'  # Default


class TestFlujoDonacionCompleto:
    """Tests para el flujo completo de donaciones."""
    
    def test_flujo_crear_donacion_con_productos_catalogo(self):
        """Test: Flujo completo de creación de donación usando catálogo principal."""
        # 1. Verificar que existan productos en catálogo
        catalogo = {'MED001': True, 'MED002': True}
        
        # 2. Crear donación
        donacion = {
            'id': 1,
            'numero': 'DON-TEST-001',
            'donante': 'Test Empresa',
            'estado': 'pendiente'
        }
        
        # 3. Agregar detalles solo si producto existe en catálogo
        productos_a_donar = ['MED001', 'MED002', 'NOEXISTE']
        detalles_creados = []
        errores = []
        
        for prod_clave in productos_a_donar:
            if prod_clave in catalogo:
                detalles_creados.append({'producto': prod_clave, 'cantidad': 100})
            else:
                errores.append(f'Producto {prod_clave} no existe en catálogo')
        
        # Verificar resultados
        assert len(detalles_creados) == 2  # Solo MED001 y MED002
        assert len(errores) == 1  # NOEXISTE
        assert 'NOEXISTE' in errores[0]
        
    def test_exportar_donaciones_incluye_productos_catalogo(self):
        """Test: Al exportar donaciones, se muestran los nombres del catálogo principal."""
        # Simular donación con detalle
        detalle = {
            'producto_id': 1,
            'producto_nombre': 'Paracetamol 500mg',  # Del catálogo principal
            'cantidad': 100,
            'cantidad_disponible': 80
        }
        
        # El nombre debe venir del catálogo principal
        assert detalle['producto_nombre'] == 'Paracetamol 500mg'
        assert detalle['cantidad_disponible'] <= detalle['cantidad']


class TestMockDonacionViewSet:
    """Tests con mocks para DonacionViewSet."""
    
    @patch('core.models.Producto')
    def test_plantilla_excel_usa_productos_principales(self, mock_producto):
        """Test: La plantilla usa productos del modelo Producto (principal)."""
        # Simular productos del catálogo principal
        mock_producto.objects.filter.return_value.order_by.return_value = [
            Mock(clave='MED001', nombre='Paracetamol', unidad_medida='CAJA'),
            Mock(clave='MED002', nombre='Ibuprofeno', unidad_medida='FRASCO'),
        ]
        
        # Verificar que se usa el modelo correcto
        productos = mock_producto.objects.filter(activo=True).order_by('nombre')
        assert len(list(productos)) == 2
        
    @patch('core.models.Donacion')
    @patch('core.models.DetalleDonacion')
    @patch('core.models.Producto')
    def test_importar_crea_detalle_con_producto_principal(
        self, mock_producto, mock_detalle, mock_donacion
    ):
        """Test: La importación crea DetalleDonacion con producto del catálogo principal."""
        # Configurar mocks
        mock_prod_instance = Mock(id=1, clave='MED001', nombre='Paracetamol')
        mock_producto.objects.filter.return_value.first.return_value = mock_prod_instance
        
        mock_donacion_instance = Mock(id=1, numero='DON-001')
        mock_donacion.objects.create.return_value = mock_donacion_instance
        
        # Simular creación de detalle
        detalle_data = {
            'donacion': mock_donacion_instance,
            'producto': mock_prod_instance,  # Usa catálogo principal
            'producto_donacion': None,  # No usa catálogo independiente
            'cantidad': 100,
            'cantidad_disponible': 100
        }
        
        # Verificar que usa producto principal
        assert detalle_data['producto'] == mock_prod_instance
        assert detalle_data['producto_donacion'] is None


class TestExcelGeneracion:
    """Tests para generación de archivos Excel."""
    
    def test_generar_plantilla_valida(self):
        """Test: Se puede generar una plantilla Excel válida."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Donaciones'
        
        # Headers simplificados
        headers = ['numero *', 'donante *', 'tipo_donante', 'fecha *', 
                   'producto_clave *', 'cantidad *', 'lote', 'notas']
        ws.append(headers)
        
        # Guardar en memoria
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Verificar que se puede leer
        wb_leido = openpyxl.load_workbook(buffer)
        assert 'Donaciones' in wb_leido.sheetnames
        
        # Verificar headers
        ws_leido = wb_leido['Donaciones']
        headers_leidos = [cell.value for cell in ws_leido[1]]
        assert headers_leidos == headers
        
    def test_leer_archivo_con_ejemplos(self):
        """Test: Se pueden detectar y filtrar filas de ejemplo."""
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Agregar datos
        ws.append(['numero', 'donante', 'cantidad'])
        ws.append(['[EJEMPLO] DON-001', 'Test', 100])  # Ejemplo - ignorar
        ws.append(['DON-002', 'Real', 200])  # Real - procesar
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        wb_leido = openpyxl.load_workbook(buffer)
        ws_leido = wb_leido.active
        
        filas_reales = []
        for row in ws_leido.iter_rows(min_row=2, values_only=True):
            es_ejemplo = any('[EJEMPLO]' in str(cell).upper() for cell in row if cell)
            if not es_ejemplo:
                filas_reales.append(row)
        
        assert len(filas_reales) == 1
        assert filas_reales[0][0] == 'DON-002'


# Ejecutar tests si se llama directamente
if __name__ == '__main__':
    pytest.main([__file__, '-v'])

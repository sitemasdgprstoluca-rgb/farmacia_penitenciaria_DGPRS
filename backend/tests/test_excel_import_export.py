"""
Pruebas unitarias para importación/exportación de Excel.

Verifica que:
1. Las plantillas se generan con las columnas correctas
2. La estructura de los Excel es correcta

Nota: Estas pruebas NO crean datos en BD porque los modelos son managed=False
(tablas externas en Supabase). Se prueban las estructuras de los archivos Excel.

Ejecutar con: pytest tests/test_excel_import_export.py -v
"""
import pytest
import io
import openpyxl
from datetime import date, timedelta
from unittest.mock import Mock, patch, MagicMock
from django.test import TestCase, RequestFactory
from django.contrib.auth import get_user_model
from rest_framework.test import APIRequestFactory
from rest_framework import status

# Importar los ViewSets
from inventario.views_legacy import ProductoViewSet, LoteViewSet

User = get_user_model()


class TestProductoPlantillaStructure(TestCase):
    """Pruebas de estructura de plantilla de productos."""
    
    def setUp(self):
        """Configurar factory y mock user."""
        self.factory = APIRequestFactory()
        self.mock_user = Mock()
        self.mock_user.is_authenticated = True
        self.mock_user.is_superuser = True
        self.mock_user.is_staff = True
        self.mock_user.rol = 'admin_sistema'
    
    def test_plantilla_productos_tiene_columnas_correctas(self):
        """Verifica que la plantilla de productos tenga todas las columnas."""
        # Crear request mock
        request = self.factory.get('/api/productos/plantilla/')
        request.user = self.mock_user
        
        # Ejecutar el método
        viewset = ProductoViewSet()
        viewset.request = request
        viewset.format_kwarg = None
        
        response = viewset.plantilla_productos(request)
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Leer el archivo Excel
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Obtener headers (fila 1)
        headers = [cell.value for cell in ws[1]]
        
        # Columnas esperadas (con nombre_comercial)
        expected_columns = [
            'Clave', 'Nombre', 'Nombre Comercial', 'Unidad Medida', 'Stock Minimo', 
            'Categoria', 'Sustancia Activa', 'Presentacion', 'Concentracion',
            'Via Admin', 'Requiere Receta', 'Controlado', 'Estado'
        ]
        
        for col in expected_columns:
            self.assertIn(col, headers, f"Columna '{col}' falta en plantilla de productos")
        
        # Verificar que tiene hoja de instrucciones
        self.assertIn('INSTRUCCIONES', wb.sheetnames)
        
        wb.close()
    
    def test_plantilla_productos_tiene_ejemplos(self):
        """Verifica que la plantilla tenga filas de ejemplo."""
        request = self.factory.get('/api/productos/plantilla/')
        request.user = self.mock_user
        
        viewset = ProductoViewSet()
        viewset.request = request
        viewset.format_kwarg = None
        
        response = viewset.plantilla_productos(request)
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Verificar que hay filas de ejemplo (filas 2, 3, 4)
        self.assertGreater(ws.max_row, 1, "Plantilla debe tener filas de ejemplo")
        
        # Verificar que las filas de ejemplo tienen datos
        ejemplo_fila_2 = [cell.value for cell in ws[2]]
        self.assertTrue(any(ejemplo_fila_2), "Fila 2 debe tener datos de ejemplo")
        
        wb.close()
    
    def test_plantilla_productos_nombre_comercial_en_ejemplos(self):
        """Verifica que los ejemplos incluyan nombre comercial."""
        request = self.factory.get('/api/productos/plantilla/')
        request.user = self.mock_user
        
        viewset = ProductoViewSet()
        viewset.request = request
        viewset.format_kwarg = None
        
        response = viewset.plantilla_productos(request)
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Encontrar índice de Nombre Comercial
        headers = [cell.value for cell in ws[1]]
        nombre_comercial_idx = headers.index('Nombre Comercial')
        
        # Verificar que al menos un ejemplo tiene nombre comercial
        ejemplos_con_nombre_comercial = False
        for row_num in range(2, min(5, ws.max_row + 1)):
            valor = ws.cell(row=row_num, column=nombre_comercial_idx + 1).value
            if valor and valor.strip():
                ejemplos_con_nombre_comercial = True
                break
        
        self.assertTrue(ejemplos_con_nombre_comercial, 
                       "Al menos un ejemplo debe tener nombre comercial")
        
        wb.close()


class TestLotePlantillaStructure(TestCase):
    """Pruebas de estructura de plantilla de lotes."""
    
    def setUp(self):
        """Configurar factory y mock user."""
        self.factory = APIRequestFactory()
        self.mock_user = Mock()
        self.mock_user.is_authenticated = True
        self.mock_user.is_superuser = True
        self.mock_user.is_staff = True
        self.mock_user.rol = 'admin_sistema'
    
    def test_plantilla_lotes_tiene_columnas_correctas(self):
        """Verifica que la plantilla de lotes tenga todas las columnas."""
        request = self.factory.get('/api/lotes/plantilla/')
        request.user = self.mock_user
        
        viewset = LoteViewSet()
        viewset.request = request
        viewset.format_kwarg = None
        
        response = viewset.plantilla_lotes(request)
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        
        # Leer el archivo Excel
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Obtener headers (fila 1)
        headers = [cell.value for cell in ws[1]]
        
        # Columnas esperadas (con nombre_comercial, sin ubicacion)
        expected_columns = [
            'Clave Producto', 'Nombre Producto', 'Nombre Comercial', 'Numero Lote',
            'Fecha Caducidad', 'Cantidad Inicial', 'Fecha Fabricacion', 
            'Precio Unitario', 'Numero Contrato', 'Marca', 'Activo'
        ]
        
        for col in expected_columns:
            self.assertIn(col, headers, f"Columna '{col}' falta en plantilla de lotes")
        
        # Verificar que NO tiene ubicacion
        self.assertNotIn('Ubicacion', headers)
        self.assertNotIn('Ubicación', headers)
        
        # Verificar que tiene hoja de instrucciones
        self.assertIn('INSTRUCCIONES', wb.sheetnames)
        
        wb.close()
    
    def test_plantilla_lotes_sin_ubicacion(self):
        """Verifica que la plantilla de lotes NO tenga columna ubicación."""
        request = self.factory.get('/api/lotes/plantilla/')
        request.user = self.mock_user
        
        viewset = LoteViewSet()
        viewset.request = request
        viewset.format_kwarg = None
        
        response = viewset.plantilla_lotes(request)
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        headers = [cell.value for cell in ws[1]]
        
        # Verificar que NO hay ubicacion en ninguna variante
        ubicacion_variantes = ['Ubicacion', 'Ubicación', 'ubicacion', 'ubicación', 
                              'UBICACION', 'UBICACIÓN', 'Almacen', 'Almacén']
        
        for variante in ubicacion_variantes:
            self.assertNotIn(variante, headers, 
                           f"La columna '{variante}' NO debe estar en plantilla de lotes")
        
        wb.close()
    
    def test_plantilla_lotes_tiene_nombre_comercial(self):
        """Verifica que la plantilla de lotes tenga columna Nombre Comercial."""
        request = self.factory.get('/api/lotes/plantilla/')
        request.user = self.mock_user
        
        viewset = LoteViewSet()
        viewset.request = request
        viewset.format_kwarg = None
        
        response = viewset.plantilla_lotes(request)
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        headers = [cell.value for cell in ws[1]]
        
        self.assertIn('Nombre Comercial', headers, 
                     "Plantilla de lotes debe tener columna 'Nombre Comercial'")
        
        wb.close()
    
    def test_plantilla_lotes_tiene_ejemplos_con_nombre_comercial(self):
        """Verifica que los ejemplos de lotes incluyan nombre comercial."""
        request = self.factory.get('/api/lotes/plantilla/')
        request.user = self.mock_user
        
        viewset = LoteViewSet()
        viewset.request = request
        viewset.format_kwarg = None
        
        response = viewset.plantilla_lotes(request)
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Encontrar índice de Nombre Comercial
        headers = [cell.value for cell in ws[1]]
        nombre_comercial_idx = headers.index('Nombre Comercial')
        
        # Verificar que al menos un ejemplo tiene nombre comercial
        ejemplos_con_nombre_comercial = False
        for row_num in range(2, min(5, ws.max_row + 1)):
            valor = ws.cell(row=row_num, column=nombre_comercial_idx + 1).value
            if valor and str(valor).strip():
                ejemplos_con_nombre_comercial = True
                break
        
        self.assertTrue(ejemplos_con_nombre_comercial, 
                       "Al menos un ejemplo de lote debe tener nombre comercial")
        
        wb.close()


class TestExportStructure(TestCase):
    """Pruebas de estructura de exportación (sin datos reales)."""
    
    def setUp(self):
        """Configurar factory y mock user."""
        self.factory = APIRequestFactory()
        self.mock_user = Mock()
        self.mock_user.is_authenticated = True
        self.mock_user.is_superuser = True
        self.mock_user.is_staff = True
        self.mock_user.rol = 'admin_sistema'
    
    @patch.object(ProductoViewSet, 'get_queryset')
    def test_exportar_productos_columnas_correctas(self, mock_queryset):
        """Verifica que exportación de productos tenga columnas correctas."""
        # Mock de queryset vacío
        mock_queryset.return_value = []
        
        request = self.factory.get('/api/productos/exportar-excel/')
        request.user = self.mock_user
        request.query_params = {}
        
        viewset = ProductoViewSet()
        viewset.request = request
        viewset.format_kwarg = None
        
        response = viewset.exportar_excel(request)
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Headers en fila 1
        headers = [cell.value for cell in ws[1]]
        
        # Columnas esperadas en exportación
        expected_columns = [
            '#', 'Clave', 'Nombre', 'Nombre Comercial', 'Categoria', 'Unidad Medida',
            'Stock Minimo', 'Stock Actual', 'Sustancia Activa', 'Presentacion',
            'Concentracion', 'Via Admin', 'Requiere Receta', 'Controlado',
            'Lotes Activos', 'Estado'
        ]
        
        for col in expected_columns:
            self.assertIn(col, headers, f"Columna '{col}' falta en exportación de productos")
        
        wb.close()
    
    @patch.object(LoteViewSet, 'get_queryset')
    def test_exportar_lotes_columnas_correctas_sin_ubicacion(self, mock_queryset):
        """Verifica que exportación de lotes tenga columnas correctas sin ubicación."""
        # Mock de queryset vacío
        mock_queryset.return_value = []
        
        request = self.factory.get('/api/lotes/exportar-excel/')
        request.user = self.mock_user
        request.query_params = {}
        
        viewset = LoteViewSet()
        viewset.request = request
        viewset.format_kwarg = None
        
        response = viewset.exportar_excel(request)
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        
        # Headers en fila 3 (hay título y fila vacía)
        headers = [cell.value for cell in ws[3]]
        
        # Columnas esperadas (sin ubicacion)
        expected_columns = [
            '#', 'Clave', 'Nombre Producto', 'Nombre Comercial', 'Número Lote',
            'Fecha Fabricación', 'Fecha Caducidad', 'Cantidad Inicial', 'Cantidad Actual',
            'Precio Unitario', 'Número Contrato', 'Marca', 'Centro', 'Activo'
        ]
        
        for col in expected_columns:
            self.assertIn(col, headers, f"Columna '{col}' falta en exportación de lotes")
        
        # Verificar que NO tiene ubicacion
        self.assertNotIn('Ubicacion', headers)
        self.assertNotIn('Ubicación', headers)
        
        wb.close()


class TestImportColumnAliases(TestCase):
    """Pruebas para verificar que los aliases de columnas son correctos."""
    
    def test_producto_column_aliases_incluyen_nombre_comercial(self):
        """Verifica que el importador de productos reconoce nombre_comercial."""
        # Verificar que el código tiene los aliases correctos
        from inventario.views_legacy import ProductoViewSet
        import inspect
        
        # Obtener el código fuente del método importar_excel
        source = inspect.getsource(ProductoViewSet.importar_excel)
        
        # Verificar que nombre_comercial está en COLUMN_ALIASES
        self.assertIn("'nombre_comercial'", source, 
                     "importar_excel debe tener alias para nombre_comercial")
        self.assertIn("'nombre comercial'", source.lower(),
                     "importar_excel debe reconocer 'nombre comercial' como alias")
    
    def test_lote_column_aliases_no_incluyen_ubicacion(self):
        """Verifica que el importador de lotes NO tiene ubicación en aliases."""
        from inventario.views_legacy import LoteViewSet
        import inspect
        
        source = inspect.getsource(LoteViewSet.importar_excel)
        
        # El código no debe tener 'ubicacion' como clave principal en COLUMN_ALIASES
        # Nota: puede mencionarse en otros contextos pero no como alias principal
        lines = source.split('\n')
        in_column_aliases = False
        ubicacion_as_key = False
        
        for line in lines:
            if 'COLUMN_ALIASES' in line and '{' in line:
                in_column_aliases = True
            if in_column_aliases:
                if "'ubicacion'" in line and ':' in line.split("'ubicacion'")[0][-5:]:
                    # 'ubicacion': [...] - es una clave
                    ubicacion_as_key = True
                if '}' in line and in_column_aliases:
                    if line.strip().startswith('}'):
                        in_column_aliases = False
        
        self.assertFalse(ubicacion_as_key,
                        "importar_excel de lotes NO debe tener 'ubicacion' como alias de columna")


class TestConsistenciaPlantilaExport(TestCase):
    """Pruebas de consistencia entre plantilla y exportación."""
    
    def setUp(self):
        """Configurar factory y mock user."""
        self.factory = APIRequestFactory()
        self.mock_user = Mock()
        self.mock_user.is_authenticated = True
        self.mock_user.is_superuser = True
        self.mock_user.is_staff = True
        self.mock_user.rol = 'admin_sistema'
    
    def test_lotes_consistencia_sin_ubicacion(self):
        """Verifica que ni plantilla ni export de lotes tengan ubicación."""
        request = self.factory.get('/api/lotes/plantilla/')
        request.user = self.mock_user
        
        viewset = LoteViewSet()
        viewset.request = request
        viewset.format_kwarg = None
        
        # Obtener plantilla
        response_plantilla = viewset.plantilla_lotes(request)
        wb_plantilla = openpyxl.load_workbook(io.BytesIO(response_plantilla.content))
        headers_plantilla = [cell.value for cell in wb_plantilla.active[1]]
        wb_plantilla.close()
        
        # Verificar que NO hay ubicacion en plantilla
        self.assertNotIn('Ubicacion', headers_plantilla)
        self.assertNotIn('Ubicación', headers_plantilla)
        
        # Verificar que Nombre Comercial SÍ está
        self.assertIn('Nombre Comercial', headers_plantilla)
    
    def test_productos_consistencia_nombre_comercial(self):
        """Verifica que plantilla de productos tenga nombre_comercial."""
        request = self.factory.get('/api/productos/plantilla/')
        request.user = self.mock_user
        
        viewset = ProductoViewSet()
        viewset.request = request
        viewset.format_kwarg = None
        
        response = viewset.plantilla_productos(request)
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        headers = [cell.value for cell in wb.active[1]]
        wb.close()
        
        self.assertIn('Nombre Comercial', headers,
                     "Plantilla de productos debe tener 'Nombre Comercial'")


class TestAlineacionEsquemaBD(TestCase):
    """
    Pruebas de alineación con el esquema real de la base de datos Supabase.
    
    Esquema BD productos: clave, nombre, nombre_comercial, descripcion, unidad_medida,
                          categoria, stock_minimo, stock_actual, sustancia_activa, 
                          presentacion, concentracion, via_administracion, requiere_receta,
                          es_controlado, activo, imagen, created_at, updated_at
    
    Esquema BD lotes: numero_lote, producto_id, cantidad_inicial, cantidad_actual,
                      fecha_fabricacion, fecha_caducidad, precio_unitario, numero_contrato,
                      marca, ubicacion, centro_id, activo, created_at, updated_at
    """
    
    def test_producto_importador_mapea_campos_bd(self):
        """Verifica que el importador de productos mapee todos los campos relevantes de BD."""
        from core.utils.excel_importer import importar_productos_desde_excel
        
        # Campos de la BD que el importador debe reconocer
        campos_bd_productos = [
            'clave', 'nombre', 'nombre_comercial', 'unidad_medida', 'categoria',
            'sustancia_activa', 'presentacion', 'concentracion', 'via_administracion',
            'stock_minimo', 'requiere_receta', 'es_controlado', 'activo'
        ]
        
        # El importador usa estos sinónimos internamente
        # Verificar que el código del importador contiene estos campos
        import inspect
        source = inspect.getsource(importar_productos_desde_excel)
        
        for campo in campos_bd_productos:
            # Verificar en SINONIMOS o en el código
            self.assertTrue(
                campo in source or campo.replace('_', ' ') in source.lower(),
                f"Campo BD '{campo}' debe estar mapeado en importador de productos"
            )
    
    def test_lote_importador_mapea_campos_bd(self):
        """Verifica que el importador de lotes mapee todos los campos relevantes de BD."""
        from core.utils.excel_importer import importar_lotes_desde_excel
        
        # Campos de la BD que el importador debe reconocer
        campos_bd_lotes = [
            'numero_lote', 'cantidad_inicial', 'fecha_fabricacion', 'fecha_caducidad',
            'precio_unitario', 'numero_contrato', 'marca', 'activo'
        ]
        
        import inspect
        source = inspect.getsource(importar_lotes_desde_excel)
        
        for campo in campos_bd_lotes:
            # Verificar en SINONIMOS o en el código
            self.assertTrue(
                campo in source or campo.replace('_', ' ') in source.lower(),
                f"Campo BD '{campo}' debe estar mapeado en importador de lotes"
            )
    
    def test_ubicacion_se_asigna_automaticamente_en_importacion(self):
        """
        Verifica que ubicacion se asigna automáticamente como 'Almacén Central'.
        
        La BD tiene campo 'ubicacion' pero NO se pide al usuario, 
        se asigna automáticamente.
        """
        from core.utils.excel_importer import importar_lotes_desde_excel
        import inspect
        source = inspect.getsource(importar_lotes_desde_excel)
        
        # Debe asignar ubicacion automáticamente
        self.assertIn("ubicacion='Almacén Central'", source,
                     "Importador de lotes debe asignar ubicacion='Almacén Central' automáticamente")
    
    def test_exportar_productos_incluye_nombre_comercial(self):
        """Verifica que el exportador de productos incluya nombre_comercial."""
        from inventario.views.productos import ProductoViewSet
        import inspect
        source = inspect.getsource(ProductoViewSet.exportar_excel)
        
        self.assertIn('Nombre Comercial', source,
                     "Exportador de productos debe incluir columna 'Nombre Comercial'")
        self.assertIn('nombre_comercial', source,
                     "Exportador de productos debe acceder al campo nombre_comercial")
    
    def test_exportar_lotes_incluye_nombre_comercial_sin_ubicacion(self):
        """Verifica que el exportador de lotes incluya nombre_comercial y NO ubicacion."""
        from inventario.views.lotes import LoteViewSet
        import inspect
        source = inspect.getsource(LoteViewSet.exportar_excel)
        
        self.assertIn('Nombre Comercial', source,
                     "Exportador de lotes debe incluir columna 'Nombre Comercial'")
        self.assertNotIn("'Ubicación'", source,
                        "Exportador de lotes NO debe incluir columna 'Ubicación'")
        self.assertNotIn("'Ubicacion'", source,
                        "Exportador de lotes NO debe incluir columna 'Ubicacion'")


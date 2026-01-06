"""
Tests para la exportación del Inventario de Donaciones.

Este módulo prueba:
1. Exportación a Excel con formato trazabilidad
2. Exportación a PDF
3. Respeto de filtros (search, disponible, caducidad, estado_producto)
4. Permisos de acceso
"""

import pytest
from django.test import TestCase
from django.contrib.auth import get_user_model
from rest_framework.test import APIClient
from rest_framework import status
from datetime import date, timedelta
from decimal import Decimal
import io

User = get_user_model()


def crear_usuario_farmacia(username):
    """Crea un usuario con permisos de farmacia."""
    user = User.objects.create_user(
        username=username,
        password='test123',
        is_staff=True,
        is_superuser=True
    )
    return user


class TestExportarInventarioExcel(TestCase):
    """Pruebas de exportación a Excel con formato trazabilidad."""
    
    def setUp(self):
        self.user = crear_usuario_farmacia('inv_export_admin')
        self.client = APIClient()
        self.client.force_authenticate(user=self.user)
        
        from core.models import ProductoDonacion, Donacion, DetalleDonacion
        
        # Crear producto de donación
        self.producto = ProductoDonacion.objects.create(
            clave='INV-EXP-001',
            nombre='Producto Inventario Export',
            unidad_medida='CAJA',
            presentacion='Caja x 10'
        )
        
        # Crear donación procesada
        self.donacion = Donacion.objects.create(
            numero='DON-INV-EXP-001',
            donante_nombre='Donante Export Test',
            fecha_donacion=date.today(),
            estado='procesada'
        )
        
        # Crear varios detalles para probar
        self.detalle1 = DetalleDonacion.objects.create(
            donacion=self.donacion,
            producto_donacion=self.producto,
            cantidad=100,
            cantidad_disponible=80,
            numero_lote='LOTE-001',
            fecha_caducidad=date.today() + timedelta(days=120),  # Vigente
            estado_producto='bueno'
        )
        
        self.detalle2 = DetalleDonacion.objects.create(
            donacion=self.donacion,
            producto_donacion=self.producto,
            cantidad=50,
            cantidad_disponible=0,  # Agotado
            numero_lote='LOTE-002',
            fecha_caducidad=date.today() + timedelta(days=20),  # Crítico
            estado_producto='regular'
        )
    
    def test_exportar_excel_retorna_archivo(self):
        """Exportar Excel retorna archivo válido."""
        response = self.client.get('/api/detalle-donaciones/exportar-excel/')
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn('application/vnd.openxmlformats', response['Content-Type'])
        self.assertIn('attachment', response['Content-Disposition'])
        self.assertIn('.xlsx', response['Content-Disposition'])
    
    def test_exportar_excel_nombre_archivo_con_trazabilidad(self):
        """El nombre del archivo indica que es formato trazabilidad."""
        response = self.client.get('/api/detalle-donaciones/exportar-excel/')
        
        self.assertIn('trazabilidad', response['Content-Disposition'])
    
    def test_exportar_excel_respeta_filtro_disponible(self):
        """Exportar con filtro disponible=true solo incluye con stock."""
        import openpyxl
        from io import BytesIO
        
        response = self.client.get('/api/detalle-donaciones/exportar-excel/?disponible=true')
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        
        # Leer el Excel y verificar contenido
        wb = openpyxl.load_workbook(BytesIO(response.content))
        ws = wb.active
        
        # Buscar filas de datos (después de encabezados)
        data_rows = []
        for row in ws.iter_rows(min_row=6, values_only=True):
            if row[0] and isinstance(row[0], int):  # Fila con número
                data_rows.append(row)
        
        # Solo debe haber 1 registro (detalle1 con stock disponible)
        self.assertEqual(len(data_rows), 1)
    
    def test_exportar_excel_respeta_filtro_agotado(self):
        """Exportar con filtro disponible=agotado solo incluye sin stock."""
        import openpyxl
        from io import BytesIO
        
        response = self.client.get('/api/detalle-donaciones/exportar-excel/?disponible=agotado')
        
        wb = openpyxl.load_workbook(BytesIO(response.content))
        ws = wb.active
        
        data_rows = []
        for row in ws.iter_rows(min_row=6, values_only=True):
            if row[0] and isinstance(row[0], int):
                data_rows.append(row)
        
        # Solo debe haber 1 registro (detalle2 agotado)
        self.assertEqual(len(data_rows), 1)
    
    def test_exportar_excel_respeta_filtro_caducidad(self):
        """Exportar con filtro caducidad respeta el semáforo."""
        import openpyxl
        from io import BytesIO
        
        # Filtrar solo críticos (<=30 días)
        response = self.client.get('/api/detalle-donaciones/exportar-excel/?caducidad=critico')
        
        wb = openpyxl.load_workbook(BytesIO(response.content))
        ws = wb.active
        
        data_rows = []
        for row in ws.iter_rows(min_row=6, values_only=True):
            if row[0] and isinstance(row[0], int):
                data_rows.append(row)
        
        # Solo debe haber 1 registro (detalle2 con 20 días)
        self.assertEqual(len(data_rows), 1)
    
    def test_exportar_excel_respeta_filtro_estado(self):
        """Exportar con filtro estado_producto."""
        import openpyxl
        from io import BytesIO
        
        response = self.client.get('/api/detalle-donaciones/exportar-excel/?estado_producto=regular')
        
        wb = openpyxl.load_workbook(BytesIO(response.content))
        ws = wb.active
        
        data_rows = []
        for row in ws.iter_rows(min_row=6, values_only=True):
            if row[0] and isinstance(row[0], int):
                data_rows.append(row)
        
        # Solo debe haber 1 registro (detalle2 con estado regular)
        self.assertEqual(len(data_rows), 1)
    
    def test_exportar_excel_respeta_filtro_search(self):
        """Exportar con búsqueda de texto."""
        # Crear producto con nombre diferente
        from core.models import ProductoDonacion, DetalleDonacion
        
        producto_nuevo = ProductoDonacion.objects.create(
            clave='BUSCAR-001',
            nombre='Producto Específico Búsqueda'
        )
        DetalleDonacion.objects.create(
            donacion=self.donacion,
            producto_donacion=producto_nuevo,
            cantidad=25,
            cantidad_disponible=25,
            numero_lote='BUSCAR-LOTE'
        )
        
        response = self.client.get('/api/detalle-donaciones/exportar-excel/?search=BUSCAR')
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)


class TestExportarInventarioPdf(TestCase):
    """Pruebas de exportación a PDF."""
    
    def setUp(self):
        self.user = crear_usuario_farmacia('pdf_export_admin')
        self.client = APIClient()
        self.client.force_authenticate(user=self.user)
        
        from core.models import ProductoDonacion, Donacion, DetalleDonacion
        
        self.producto = ProductoDonacion.objects.create(
            clave='PDF-001',
            nombre='Producto PDF Export'
        )
        
        self.donacion = Donacion.objects.create(
            numero='DON-PDF-001',
            donante_nombre='Donante PDF',
            fecha_donacion=date.today(),
            estado='procesada'
        )
        
        self.detalle = DetalleDonacion.objects.create(
            donacion=self.donacion,
            producto_donacion=self.producto,
            cantidad=50,
            cantidad_disponible=50,
            numero_lote='PDF-LOTE'
        )
    
    def test_exportar_pdf_retorna_archivo(self):
        """Exportar PDF retorna archivo válido."""
        response = self.client.get('/api/detalle-donaciones/exportar-pdf/')
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response['Content-Type'], 'application/pdf')
        self.assertIn('attachment', response['Content-Disposition'])
        self.assertIn('.pdf', response['Content-Disposition'])
    
    def test_exportar_pdf_respeta_filtros(self):
        """Exportar PDF respeta filtros aplicados."""
        # Con filtro que excluye todo
        response = self.client.get('/api/detalle-donaciones/exportar-pdf/?disponible=agotado')
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        # El PDF debe generarse aunque esté vacío


class TestPermisosExportacion(TestCase):
    """Pruebas de permisos para exportación."""
    
    def setUp(self):
        # Usuario autenticado normal
        self.reader_user = User.objects.create_user(
            username='reader_export',
            password='test123',
            is_staff=False
        )
        
        self.client = APIClient()
    
    def test_usuario_autenticado_puede_exportar_excel(self):
        """Cualquier usuario autenticado puede exportar Excel."""
        self.client.force_authenticate(user=self.reader_user)
        
        response = self.client.get('/api/detalle-donaciones/exportar-excel/')
        
        # Debe tener acceso (aunque no haya datos)
        self.assertIn(response.status_code, [status.HTTP_200_OK])
    
    def test_usuario_autenticado_puede_exportar_pdf(self):
        """Cualquier usuario autenticado puede exportar PDF."""
        self.client.force_authenticate(user=self.reader_user)
        
        response = self.client.get('/api/detalle-donaciones/exportar-pdf/')
        
        self.assertIn(response.status_code, [status.HTTP_200_OK])
    
    def test_sin_autenticar_no_puede_exportar(self):
        """Sin autenticación no puede exportar."""
        # Sin autenticar
        response = self.client.get('/api/detalle-donaciones/exportar-excel/')
        
        self.assertEqual(response.status_code, status.HTTP_401_UNAUTHORIZED)


class TestFiltrosQueryset(TestCase):
    """Pruebas de los filtros del queryset."""
    
    def setUp(self):
        self.user = crear_usuario_farmacia('filtros_admin')
        self.client = APIClient()
        self.client.force_authenticate(user=self.user)
        
        from core.models import ProductoDonacion, Donacion, DetalleDonacion
        
        self.producto = ProductoDonacion.objects.create(
            clave='FILTRO-001',
            nombre='Producto Filtros'
        )
        
        self.donacion = Donacion.objects.create(
            numero='DON-FILTRO-001',
            donante_nombre='Donante Filtros',
            fecha_donacion=date.today(),
            estado='procesada'
        )
        
        # Crear detalles con diferentes caducidades
        hoy = date.today()
        
        # Vencido
        DetalleDonacion.objects.create(
            donacion=self.donacion,
            producto_donacion=self.producto,
            cantidad=10,
            cantidad_disponible=10,
            numero_lote='VENCIDO',
            fecha_caducidad=hoy - timedelta(days=10)
        )
        
        # Crítico (<=30 días)
        DetalleDonacion.objects.create(
            donacion=self.donacion,
            producto_donacion=self.producto,
            cantidad=10,
            cantidad_disponible=10,
            numero_lote='CRITICO',
            fecha_caducidad=hoy + timedelta(days=15)
        )
        
        # Próximo (31-90 días)
        DetalleDonacion.objects.create(
            donacion=self.donacion,
            producto_donacion=self.producto,
            cantidad=10,
            cantidad_disponible=10,
            numero_lote='PROXIMO',
            fecha_caducidad=hoy + timedelta(days=60)
        )
        
        # Normal (>90 días)
        DetalleDonacion.objects.create(
            donacion=self.donacion,
            producto_donacion=self.producto,
            cantidad=10,
            cantidad_disponible=10,
            numero_lote='NORMAL',
            fecha_caducidad=hoy + timedelta(days=180)
        )
    
    def test_filtro_caducidad_vencido(self):
        """Filtro caducidad=vencido funciona."""
        response = self.client.get('/api/detalle-donaciones/?caducidad=vencido')
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        # Verificar que solo trae vencidos
        for item in response.data.get('results', []):
            self.assertIn('VENCIDO', item.get('numero_lote', ''))
    
    def test_filtro_caducidad_critico(self):
        """Filtro caducidad=critico funciona."""
        response = self.client.get('/api/detalle-donaciones/?caducidad=critico')
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        for item in response.data.get('results', []):
            self.assertIn('CRITICO', item.get('numero_lote', ''))
    
    def test_filtro_caducidad_proximo(self):
        """Filtro caducidad=proximo funciona."""
        response = self.client.get('/api/detalle-donaciones/?caducidad=proximo')
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        for item in response.data.get('results', []):
            self.assertIn('PROXIMO', item.get('numero_lote', ''))
    
    def test_filtro_caducidad_normal(self):
        """Filtro caducidad=normal funciona."""
        response = self.client.get('/api/detalle-donaciones/?caducidad=normal')
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        for item in response.data.get('results', []):
            self.assertIn('NORMAL', item.get('numero_lote', ''))

"""
Tests unitarios para verificar la importación de productos con nombre_comercial.
Valida que la nueva estructura de plantilla funcione correctamente.
"""
import pytest
from io import BytesIO
from unittest.mock import MagicMock, patch
from django.test import TestCase
import openpyxl


class TestImportacionNombreComercialLogica(TestCase):
    """Tests para la lógica de importación de productos con campo nombre_comercial."""
    
    def _crear_excel_con_nombre_comercial(self, productos_data):
        """
        Crea un archivo Excel en memoria con la nueva estructura incluyendo nombre_comercial.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Productos"
        
        # Headers con la nueva estructura (incluyendo Nombre Comercial)
        headers = [
            "Clave", "Nombre", "Nombre Comercial", "Unidad", "Stock Minimo",
            "Categoria", "Sustancia Activa", "Presentacion", "Concentracion",
            "Via Admin", "Requiere Receta", "Controlado", "Estado"
        ]
        
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Agregar datos
        for row_num, prod in enumerate(productos_data, 2):
            ws.cell(row=row_num, column=1, value=prod.get('clave', ''))
            ws.cell(row=row_num, column=2, value=prod.get('nombre', ''))
            ws.cell(row=row_num, column=3, value=prod.get('nombre_comercial', ''))
            ws.cell(row=row_num, column=4, value=prod.get('unidad', 'CAJA'))
            ws.cell(row=row_num, column=5, value=prod.get('stock_minimo', 10))
            ws.cell(row=row_num, column=6, value=prod.get('categoria', 'medicamento'))
            ws.cell(row=row_num, column=7, value=prod.get('sustancia_activa', ''))
            ws.cell(row=row_num, column=8, value=prod.get('presentacion', ''))
            ws.cell(row=row_num, column=9, value=prod.get('concentracion', ''))
            ws.cell(row=row_num, column=10, value=prod.get('via_admin', ''))
            ws.cell(row=row_num, column=11, value=prod.get('requiere_receta', 'No'))
            ws.cell(row=row_num, column=12, value=prod.get('controlado', 'No'))
            ws.cell(row=row_num, column=13, value=prod.get('estado', 'Activo'))
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    @patch('core.utils.excel_importer.Producto')
    def test_importar_producto_con_nombre_comercial(self, MockProducto):
        """Verifica que se importe correctamente un producto con nombre_comercial."""
        from core.utils.excel_importer import importar_productos_desde_excel
        
        # Configurar mock
        mock_obj = MagicMock()
        MockProducto.objects.update_or_create.return_value = (mock_obj, True)
        
        productos_data = [{
            'clave': 'TEST_NC_001',
            'nombre': 'Paracetamol 500mg',
            'nombre_comercial': 'Tylenol',
            'unidad': 'CAJA',
            'stock_minimo': 50,
        }]
        
        archivo = self._crear_excel_con_nombre_comercial(productos_data)
        usuario_mock = MagicMock()
        usuario_mock.id = 1
        
        resultado = importar_productos_desde_excel(archivo, usuario_mock)
        
        # Verificar que se llamó update_or_create con nombre_comercial
        MockProducto.objects.update_or_create.assert_called()
        call_kwargs = MockProducto.objects.update_or_create.call_args
        defaults = call_kwargs.kwargs.get('defaults', call_kwargs[1].get('defaults', {}))
        
        assert 'nombre_comercial' in defaults, f"nombre_comercial no está en defaults: {defaults}"
        assert defaults['nombre_comercial'] == 'Tylenol'
    
    @patch('core.utils.excel_importer.Producto')
    def test_importar_producto_sin_nombre_comercial(self, MockProducto):
        """Verifica que un producto sin nombre_comercial se importe correctamente."""
        from core.utils.excel_importer import importar_productos_desde_excel
        
        mock_obj = MagicMock()
        MockProducto.objects.update_or_create.return_value = (mock_obj, True)
        
        productos_data = [{
            'clave': 'TEST_NC_002',
            'nombre': 'Ibuprofeno 400mg',
            'nombre_comercial': '',  # Sin nombre comercial
        }]
        
        archivo = self._crear_excel_con_nombre_comercial(productos_data)
        usuario_mock = MagicMock()
        usuario_mock.id = 1
        
        resultado = importar_productos_desde_excel(archivo, usuario_mock)
        
        # Verificar que nombre_comercial es None cuando está vacío
        call_kwargs = MockProducto.objects.update_or_create.call_args
        defaults = call_kwargs.kwargs.get('defaults', call_kwargs[1].get('defaults', {}))
        
        assert defaults.get('nombre_comercial') is None or defaults.get('nombre_comercial') == ''
    
    @patch('core.utils.excel_importer.Producto')
    def test_importar_multiples_productos_con_nombre_comercial(self, MockProducto):
        """Verifica importación masiva con nombre_comercial."""
        from core.utils.excel_importer import importar_productos_desde_excel
        
        mock_obj = MagicMock()
        MockProducto.objects.update_or_create.return_value = (mock_obj, True)
        
        productos_data = [
            {'clave': 'TEST_NC_003', 'nombre': 'Aspirina', 'nombre_comercial': 'Bayer Aspirin'},
            {'clave': 'TEST_NC_004', 'nombre': 'Omeprazol', 'nombre_comercial': 'Prilosec'},
            {'clave': 'TEST_NC_005', 'nombre': 'Jeringa', 'nombre_comercial': ''},
        ]
        
        archivo = self._crear_excel_con_nombre_comercial(productos_data)
        usuario_mock = MagicMock()
        usuario_mock.id = 1
        
        resultado = importar_productos_desde_excel(archivo, usuario_mock)
        
        # Verificar que se llamó 3 veces
        assert MockProducto.objects.update_or_create.call_count == 3
    
    @patch('core.utils.excel_importer.Producto')
    def test_nombre_comercial_largo_se_trunca(self, MockProducto):
        """Verifica que nombre_comercial largo se trunca a 200 caracteres."""
        from core.utils.excel_importer import importar_productos_desde_excel
        
        mock_obj = MagicMock()
        MockProducto.objects.update_or_create.return_value = (mock_obj, True)
        
        nombre_comercial_largo = 'A' * 250  # 250 caracteres
        
        productos_data = [{
            'clave': 'TEST_NC_007',
            'nombre': 'Producto Test',
            'nombre_comercial': nombre_comercial_largo,
        }]
        
        archivo = self._crear_excel_con_nombre_comercial(productos_data)
        usuario_mock = MagicMock()
        usuario_mock.id = 1
        
        resultado = importar_productos_desde_excel(archivo, usuario_mock)
        
        call_kwargs = MockProducto.objects.update_or_create.call_args
        defaults = call_kwargs.kwargs.get('defaults', call_kwargs[1].get('defaults', {}))
        
        assert len(defaults['nombre_comercial']) <= 200


class TestSinonimosNombreComercial(TestCase):
    """Tests para verificar que los sinónimos de nombre_comercial funcionan."""
    
    def _crear_excel_con_header_alternativo(self, header_nombre_comercial, valor):
        """Crea Excel con un header alternativo para nombre comercial."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Productos"
        
        headers = ["Clave", "Nombre", header_nombre_comercial]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        ws.cell(row=2, column=1, value='TEST_SIN_001')
        ws.cell(row=2, column=2, value='Producto Test')
        ws.cell(row=2, column=3, value=valor)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
    
    @patch('core.utils.excel_importer.Producto')
    def test_sinonimo_marca_comercial(self, MockProducto):
        """Verifica que 'Marca Comercial' se reconozca como nombre_comercial."""
        from core.utils.excel_importer import importar_productos_desde_excel
        
        mock_obj = MagicMock()
        MockProducto.objects.update_or_create.return_value = (mock_obj, True)
        
        archivo = self._crear_excel_con_header_alternativo('Marca Comercial', 'TestBrand')
        usuario_mock = MagicMock()
        usuario_mock.id = 1
        
        resultado = importar_productos_desde_excel(archivo, usuario_mock)
        
        call_kwargs = MockProducto.objects.update_or_create.call_args
        defaults = call_kwargs.kwargs.get('defaults', call_kwargs[1].get('defaults', {}))
        
        assert defaults.get('nombre_comercial') == 'TestBrand'
    
    @patch('core.utils.excel_importer.Producto')
    def test_sinonimo_nombre_de_marca(self, MockProducto):
        """Verifica que 'Nombre de Marca' se reconozca como nombre_comercial."""
        from core.utils.excel_importer import importar_productos_desde_excel
        
        mock_obj = MagicMock()
        MockProducto.objects.update_or_create.return_value = (mock_obj, True)
        
        archivo = self._crear_excel_con_header_alternativo('Nombre de Marca', 'BrandTest')
        usuario_mock = MagicMock()
        usuario_mock.id = 1
        
        resultado = importar_productos_desde_excel(archivo, usuario_mock)
        
        call_kwargs = MockProducto.objects.update_or_create.call_args
        defaults = call_kwargs.kwargs.get('defaults', call_kwargs[1].get('defaults', {}))
        
        assert defaults.get('nombre_comercial') == 'BrandTest'


class TestPlantillaProductosEstructura(TestCase):
    """Tests para verificar la estructura de la plantilla generada."""
    
    def test_plantilla_incluye_nombre_comercial_en_headers(self):
        """Verifica que la plantilla generada incluya Nombre Comercial en headers."""
        from core.utils.excel_templates import generar_plantilla_productos
        
        response = generar_plantilla_productos()
        
        # Leer el Excel generado
        wb = openpyxl.load_workbook(BytesIO(response.content))
        ws = wb['Productos']
        
        # Obtener headers (fila 1)
        headers = [cell.value for cell in ws[1]]
        
        assert 'Nombre Comercial' in headers, f"Headers: {headers}"
        
        # Verificar posición correcta (después de Nombre)
        nombre_idx = headers.index('Nombre')
        nombre_comercial_idx = headers.index('Nombre Comercial')
        assert nombre_comercial_idx == nombre_idx + 1, "Nombre Comercial debe estar después de Nombre"
    
    def test_plantilla_incluye_ejemplos_con_nombre_comercial(self):
        """Verifica que los ejemplos en la plantilla incluyan nombre_comercial."""
        from core.utils.excel_templates import generar_plantilla_productos
        
        response = generar_plantilla_productos()
        
        wb = openpyxl.load_workbook(BytesIO(response.content))
        ws = wb['Productos']
        
        # Obtener índice de Nombre Comercial
        headers = [cell.value for cell in ws[1]]
        nc_idx = headers.index('Nombre Comercial') + 1  # +1 porque openpyxl es 1-based
        
        # Verificar fila de ejemplo (fila 2)
        ejemplo_nc = ws.cell(row=2, column=nc_idx).value
        assert ejemplo_nc is not None and ejemplo_nc != '', f"Ejemplo de nombre comercial vacío: {ejemplo_nc}"
    
    def test_plantilla_instrucciones_incluyen_nombre_comercial(self):
        """Verifica que las instrucciones mencionen Nombre Comercial."""
        from core.utils.excel_templates import generar_plantilla_productos
        
        response = generar_plantilla_productos()
        
        wb = openpyxl.load_workbook(BytesIO(response.content))
        ws_inst = wb['INSTRUCCIONES']
        
        # Buscar mención de Nombre Comercial en instrucciones
        instrucciones_text = ' '.join([
            str(cell.value or '') 
            for row in ws_inst.iter_rows() 
            for cell in row
        ])
        
        assert 'Nombre Comercial' in instrucciones_text, "Instrucciones deben mencionar Nombre Comercial"


# Para ejecutar: pytest backend/tests/test_importacion_nombre_comercial.py -v

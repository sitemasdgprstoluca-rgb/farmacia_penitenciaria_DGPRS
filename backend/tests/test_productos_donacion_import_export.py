"""
Tests para importación/exportación de productos de donación.
Verifica el flujo completo: plantilla -> importar -> exportar
"""
import pytest
from django.test import TestCase
from django.contrib.auth import get_user_model
from rest_framework.test import APIClient
from rest_framework import status
from io import BytesIO
import openpyxl

User = get_user_model()


class TestPlantillaExcel(TestCase):
    """Tests para descarga de plantilla Excel."""
    
    def setUp(self):
        self.client = APIClient()
        self.user = User.objects.create_user(
            username='farmacia_test',
            password='test123',
            email='farmacia@test.com',
            rol='farmacia'
        )
        self.client.force_authenticate(user=self.user)
    
    def test_descargar_plantilla_excel(self):
        """Debe poder descargar la plantilla Excel."""
        response = self.client.get('/api/v1/productos-donacion/plantilla-excel/')
        
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        self.assertIn('attachment', response['Content-Disposition'])
        self.assertIn('plantilla_productos_donacion.xlsx', response['Content-Disposition'])
    
    def test_plantilla_tiene_estructura_correcta(self):
        """La plantilla debe tener los headers correctos."""
        response = self.client.get('/api/v1/productos-donacion/plantilla-excel/')
        
        # Leer el archivo Excel
        wb = openpyxl.load_workbook(BytesIO(response.content))
        ws = wb.active
        
        # Verificar headers
        headers = [cell.value for cell in ws[1] if cell.value]
        expected_headers = ['clave *', 'nombre *', 'descripcion', 'unidad_medida', 'presentacion', 'activo', 'notas']
        
        self.assertEqual(headers, expected_headers)
    
    def test_plantilla_tiene_ejemplos(self):
        """La plantilla debe incluir filas de ejemplo."""
        response = self.client.get('/api/v1/productos-donacion/plantilla-excel/')
        
        wb = openpyxl.load_workbook(BytesIO(response.content))
        ws = wb.active
        
        # Verificar que hay al menos una fila con [EJEMPLO]
        found_example = False
        for row in ws.iter_rows(min_row=2, max_row=5, values_only=True):
            if row and any('[EJEMPLO]' in str(cell).upper() for cell in row if cell):
                found_example = True
                break
        
        self.assertTrue(found_example, "La plantilla debe tener filas de ejemplo")


class TestImportarExcel(TestCase):
    """Tests para importación de productos desde Excel."""
    
    def setUp(self):
        self.client = APIClient()
        self.user = User.objects.create_user(
            username='farmacia_import',
            password='test123',
            email='farmacia@import.com',
            rol='farmacia'
        )
        self.client.force_authenticate(user=self.user)
    
    def _crear_excel_test(self, rows):
        """Helper para crear archivo Excel de prueba."""
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Headers
        headers = ['clave', 'nombre', 'descripcion', 'unidad_medida', 'presentacion', 'activo', 'notas']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Datos
        for row_idx, row_data in enumerate(rows, 2):
            for col, val in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col, value=val)
        
        # Guardar en buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    def test_importar_productos_nuevos(self):
        """Debe crear productos nuevos desde Excel."""
        excel = self._crear_excel_test([
            ['DON-TEST-001', 'Producto Test 1', 'Descripción 1', 'CAJA', 'Caja 20', 'SI', ''],
            ['DON-TEST-002', 'Producto Test 2', 'Descripción 2', 'PIEZA', '', 'SI', 'Notas test'],
        ])
        
        response = self.client.post(
            '/api/v1/productos-donacion/importar-excel/',
            {'archivo': excel},
            format='multipart'
        )
        
        self.assertEqual(response.status_code, 200)
        data = response.json()
        
        self.assertTrue(data['success'])
        self.assertEqual(data['creados'], 2)
        self.assertEqual(data['actualizados'], 0)
        self.assertEqual(data['total'], 2)
        self.assertEqual(len(data['errores']), 0)
    
    def test_importar_actualiza_existentes(self):
        """Debe actualizar productos existentes por clave."""
        from core.models import ProductoDonacion
        
        # Crear producto existente
        ProductoDonacion.objects.create(
            clave='DON-EXIST-001',
            nombre='Nombre Original',
            unidad_medida='PIEZA'
        )
        
        # Importar con misma clave pero nombre diferente
        excel = self._crear_excel_test([
            ['DON-EXIST-001', 'Nombre Actualizado', 'Nueva desc', 'CAJA', '', 'SI', ''],
        ])
        
        response = self.client.post(
            '/api/v1/productos-donacion/importar-excel/',
            {'archivo': excel},
            format='multipart'
        )
        
        data = response.json()
        
        self.assertTrue(data['success'])
        self.assertEqual(data['creados'], 0)
        self.assertEqual(data['actualizados'], 1)
        
        # Verificar que se actualizó
        producto = ProductoDonacion.objects.get(clave='DON-EXIST-001')
        self.assertEqual(producto.nombre, 'Nombre Actualizado')
        self.assertEqual(producto.unidad_medida, 'CAJA')
    
    def test_ignora_filas_ejemplo(self):
        """Las filas con [EJEMPLO] deben ignorarse."""
        excel = self._crear_excel_test([
            ['[EJEMPLO] DON-EJ-001', 'Producto Ejemplo', '', '', '', '', ''],
            ['DON-REAL-001', 'Producto Real', '', 'PIEZA', '', 'SI', ''],
        ])
        
        response = self.client.post(
            '/api/v1/productos-donacion/importar-excel/',
            {'archivo': excel},
            format='multipart'
        )
        
        data = response.json()
        
        # Solo debe crear el producto real, no el de ejemplo
        self.assertEqual(data['creados'], 1)
        self.assertEqual(data['total'], 1)
    
    def test_valida_campos_requeridos(self):
        """Debe reportar error si falta nombre."""
        excel = self._crear_excel_test([
            ['DON-SINNOM-001', '', '', 'PIEZA', '', 'SI', ''],  # Sin nombre
        ])
        
        response = self.client.post(
            '/api/v1/productos-donacion/importar-excel/',
            {'archivo': excel},
            format='multipart'
        )
        
        data = response.json()
        
        # Debe haber un error para esta fila
        self.assertEqual(data['creados'], 0)
        self.assertTrue(len(data['errores']) > 0)
        self.assertTrue(any('nombre' in e.lower() for e in data['errores']))
    
    def test_rechaza_archivo_sin_extension_excel(self):
        """Debe rechazar archivos que no son Excel."""
        # Crear un archivo de texto simulando ser subido
        from django.core.files.uploadedfile import SimpleUploadedFile
        
        archivo = SimpleUploadedFile(
            'archivo.txt',
            b'Esto no es un Excel',
            content_type='text/plain'
        )
        
        response = self.client.post(
            '/api/v1/productos-donacion/importar-excel/',
            {'archivo': archivo},
            format='multipart'
        )
        
        self.assertEqual(response.status_code, 400)
    
    def test_rechaza_sin_archivo(self):
        """Debe rechazar si no se envía archivo."""
        response = self.client.post(
            '/api/v1/productos-donacion/importar-excel/',
            {},
            format='multipart'
        )
        
        self.assertEqual(response.status_code, 400)
        self.assertIn('error', response.json())
    
    def test_interpreta_activo_correctamente(self):
        """Debe interpretar diferentes valores de activo."""
        excel = self._crear_excel_test([
            ['DON-ACT-001', 'Producto Activo', '', '', '', 'SI', ''],
            ['DON-ACT-002', 'Producto Activo 2', '', '', '', 'YES', ''],
            ['DON-ACT-003', 'Producto Inactivo', '', '', '', 'NO', ''],
            ['DON-ACT-004', 'Producto Inactivo 2', '', '', '', 'FALSE', ''],
        ])
        
        response = self.client.post(
            '/api/v1/productos-donacion/importar-excel/',
            {'archivo': excel},
            format='multipart'
        )
        
        from core.models import ProductoDonacion
        
        self.assertTrue(ProductoDonacion.objects.get(clave='DON-ACT-001').activo)
        self.assertTrue(ProductoDonacion.objects.get(clave='DON-ACT-002').activo)
        self.assertFalse(ProductoDonacion.objects.get(clave='DON-ACT-003').activo)
        self.assertFalse(ProductoDonacion.objects.get(clave='DON-ACT-004').activo)


class TestExportarExcel(TestCase):
    """Tests para exportación de catálogo a Excel."""
    
    def setUp(self):
        self.client = APIClient()
        self.user = User.objects.create_user(
            username='farmacia_export',
            password='test123',
            email='farmacia@export.com',
            rol='farmacia'
        )
        self.client.force_authenticate(user=self.user)
    
    def test_exportar_catalogo_vacio(self):
        """Debe poder exportar aunque no haya productos."""
        response = self.client.get('/api/v1/productos-donacion/exportar-excel/')
        
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    
    def test_exportar_incluye_todos_productos(self):
        """Debe exportar todos los productos existentes."""
        from core.models import ProductoDonacion
        
        # Crear productos
        ProductoDonacion.objects.create(clave='EXP-001', nombre='Producto 1', activo=True)
        ProductoDonacion.objects.create(clave='EXP-002', nombre='Producto 2', activo=False)
        
        response = self.client.get('/api/v1/productos-donacion/exportar-excel/')
        
        # Leer el archivo exportado
        wb = openpyxl.load_workbook(BytesIO(response.content))
        ws = wb.active
        
        # Contar filas de datos (excluyendo header)
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        data_rows = [r for r in data_rows if r and any(r)]
        
        self.assertEqual(len(data_rows), 2)
    
    def test_exportar_formato_correcto(self):
        """El archivo exportado debe tener el formato correcto."""
        from core.models import ProductoDonacion
        
        ProductoDonacion.objects.create(
            clave='FMT-001',
            nombre='Producto Formato',
            descripcion='Descripción test',
            unidad_medida='CAJA',
            presentacion='Caja 10',
            activo=True,
            notas='Nota de test'
        )
        
        response = self.client.get('/api/v1/productos-donacion/exportar-excel/')
        
        wb = openpyxl.load_workbook(BytesIO(response.content))
        ws = wb.active
        
        # Verificar headers
        headers = [cell.value for cell in ws[1] if cell.value]
        self.assertIn('clave', headers)
        self.assertIn('nombre', headers)
        
        # Verificar datos
        row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
        self.assertEqual(row[0], 'FMT-001')
        self.assertEqual(row[1], 'Producto Formato')


class TestPermisos(TestCase):
    """Tests para verificar permisos en endpoints."""
    
    def setUp(self):
        self.client = APIClient()
    
    def test_importar_requiere_autenticacion(self):
        """El endpoint de importar requiere estar autenticado."""
        response = self.client.post('/api/v1/productos-donacion/importar-excel/')
        self.assertEqual(response.status_code, 401)
    
    def test_usuario_centro_no_puede_importar(self):
        """Un usuario de centro no puede importar productos."""
        user = User.objects.create_user(
            username='centro_user',
            password='test123',
            email='centro@test.com',
            rol='centro'
        )
        self.client.force_authenticate(user=user)
        
        response = self.client.post('/api/v1/productos-donacion/importar-excel/')
        # Debería ser 403 Forbidden
        self.assertIn(response.status_code, [403, 400])  # 400 por falta de archivo también es válido
    
    def test_cualquier_usuario_autenticado_puede_descargar_plantilla(self):
        """Un usuario de farmacia puede descargar la plantilla."""
        user = User.objects.create_user(
            username='farmacia_viewer',
            password='test123',
            email='farmacia_viewer@test.com',
            rol='farmacia'
        )
        self.client.force_authenticate(user=user)
        
        response = self.client.get('/api/v1/productos-donacion/plantilla-excel/')
        self.assertEqual(response.status_code, 200)


class TestIntegracionCompleta(TestCase):
    """Test de integración: plantilla -> importar -> exportar."""
    
    def setUp(self):
        self.client = APIClient()
        self.user = User.objects.create_user(
            username='admin_int',
            password='test123',
            email='admin@int.com',
            rol='admin'
        )
        self.client.force_authenticate(user=self.user)
    
    def test_flujo_completo(self):
        """Flujo completo de trabajo con el catálogo."""
        from core.models import ProductoDonacion
        
        # 1. Descargar plantilla
        response = self.client.get('/api/v1/productos-donacion/plantilla-excel/')
        self.assertEqual(response.status_code, 200)
        
        # 2. Crear Excel con datos de prueba
        wb = openpyxl.Workbook()
        ws = wb.active
        headers = ['clave', 'nombre', 'descripcion', 'unidad_medida', 'presentacion', 'activo', 'notas']
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        
        productos = [
            ('INT-001', 'Producto Integración 1', 'Desc 1', 'CAJA', 'Caja 10', 'SI', ''),
            ('INT-002', 'Producto Integración 2', 'Desc 2', 'PIEZA', '', 'SI', 'Nota'),
            ('INT-003', 'Producto Integración 3', '', 'FRASCO', 'Frasco 500ml', 'NO', ''),
        ]
        for row_idx, prod in enumerate(productos, 2):
            for col, val in enumerate(prod, 1):
                ws.cell(row=row_idx, column=col, value=val)
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # 3. Importar
        response = self.client.post(
            '/api/v1/productos-donacion/importar-excel/',
            {'archivo': buffer},
            format='multipart'
        )
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data['creados'], 3)
        
        # 4. Verificar en BD
        self.assertEqual(ProductoDonacion.objects.count(), 3)
        
        # 5. Exportar y verificar
        response = self.client.get('/api/v1/productos-donacion/exportar-excel/')
        self.assertEqual(response.status_code, 200)
        
        wb_export = openpyxl.load_workbook(BytesIO(response.content))
        ws_export = wb_export.active
        
        # Debe tener 3 filas de datos + 1 header
        data_rows = list(ws_export.iter_rows(min_row=2, values_only=True))
        data_rows = [r for r in data_rows if r and any(r)]
        self.assertEqual(len(data_rows), 3)

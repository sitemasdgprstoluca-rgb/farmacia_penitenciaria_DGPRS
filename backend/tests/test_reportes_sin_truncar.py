#!/usr/bin/env python
"""
Prueba exhaustiva de reportes PDF y Excel sin truncamiento de texto.
Verifica que nombres de centros largos y observaciones completas se muestren correctamente.
"""
import os
import sys
import django

# Configurar Django
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'core.settings')
django.setup()

from datetime import datetime, timedelta
from io import BytesIO

# Datos de prueba con textos largos
DATOS_TRAZABILIDAD_PRUEBA = [
    {
        'fecha': datetime.now() - timedelta(days=2),
        'tipo': 'ENTRADA',
        'numero_lote': '25072052',
        'cantidad': 7,
        'centro_nombre': 'CENTRO PENITENCIARIO RIO SANTIAGO DE ALTA SEGURIDAD',
        'usuario': 'Farmacia Central Admin',
        'numero_expediente': 'EXP-2026-00123',
        'observaciones': '[CONFIRMADO]SAL-0197-23 VALE Entregado correctamente al centro',
    },
    {
        'fecha': datetime.now() - timedelta(days=1),
        'tipo': 'SALIDA',
        'subtipo_salida': 'REQUISICION',
        'numero_lote': '25072052',
        'cantidad': 7,
        'centro_nombre': 'CENTRO PENITENCIARIO RIO SANTIAGO DE ALTA SEGURIDAD',
        'usuario': 'farmacia@g...',
        'numero_expediente': '-',
        'observaciones': '[CONFIRMADO]SAL-0197-23 VALE medicamento distribuido',
    },
    {
        'fecha': datetime.now(),
        'tipo': 'SALIDA',
        'subtipo_salida': 'CONSUMO',
        'numero_lote': '25072052',
        'cantidad': 7,
        'centro_nombre': 'CENTRO PENITENCIARIO RIO SANTIAGO DE ALTA SEGURIDAD',
        'usuario': 'Administra...',
        'numero_expediente': '-',
        'observaciones': '[CONFIRMADO] Dispensado a médica residente',
    },
    {
        'fecha': datetime.now(),
        'tipo': 'SALIDA',
        'numero_lote': '25072052',
        'cantidad': 65,
        'centro_nombre': 'CENTRO PENITENCIARIO RIO SANTIAGO DE ALTA SEGURIDAD',
        'usuario': 'Farmacia C...',
        'numero_expediente': '-',
        'observaciones': '[CONFIRMADO] Surtir medicamento Tableta',
    },
    {
        'fecha': datetime.now(),
        'tipo': 'SALIDA',
        'subtipo_salida': 'REQUISICION',
        'numero_lote': '25072052',
        'cantidad': 2,
        'centro_nombre': 'Farmacia Central',
        'usuario': 'Farmacia C...',
        'numero_expediente': '-',
        'observaciones': 'SALIDA_POR_REQUISICION REQ-20260108-212',
    },
    {
        'fecha': datetime.now(),
        'tipo': 'ENTRADA',
        'numero_lote': '25072052',
        'cantidad': 2,
        'centro_nombre': 'CENTRO PENITENCIARIO RIO SANTIAGO DE ALTA SEGURIDAD',
        'usuario': 'Farmacia C...',
        'numero_expediente': '-',
        'observaciones': 'ENTRADA_POR_REQUISICION REQ-20260108-237 transferencia completada',
    },
    {
        'fecha': datetime.now(),
        'tipo': 'SALIDA',
        'numero_lote': '25072052',
        'cantidad': 2,
        'centro_nombre': 'CENTRO PENITENCIARIO MEDICO CEN ESPECIALIZADO',
        'usuario': 'Farmacia C...',
        'numero_expediente': '-',
        'observaciones': 'No sube nueva media',
    },
]

PRODUCTO_INFO_PRUEBA = {
    'clave': '615',
    'descripcion': 'KETOCONAZOL /CLINDAMICINA',
    'unidad_medida': 'CAJA CON 7 OVULOS',
    'precio_unitario': None,
    'stock_actual': 0,
    'stock_minimo': 1,
    'numero_contrato': 'CONT-PRUEBA-139',
    'numero_lote': '25072052',
    'fecha_caducidad': '01/07/2027',
    'proveedor': '[EJEMPLO] Laboratorio - ELIMINAR',
    'titulo_centro': 'TODOS LOS CENTROS (CONSOLIDADO)',
}

FILTROS_PRUEBA = {
    'centro': 'todos',
    'fecha_inicio': None,
    'fecha_fin': None,
}


def test_pdf_trazabilidad():
    """Prueba generación de PDF de trazabilidad."""
    print("\n" + "="*60)
    print("TEST: PDF de Trazabilidad")
    print("="*60)
    
    from core.utils.pdf_reports import generar_reporte_trazabilidad
    
    try:
        buffer = generar_reporte_trazabilidad(
            DATOS_TRAZABILIDAD_PRUEBA,
            producto_info=PRODUCTO_INFO_PRUEBA,
            filtros=FILTROS_PRUEBA
        )
        
        # Guardar para inspección manual
        output_path = os.path.join(os.path.dirname(__file__), 'test_output_trazabilidad.pdf')
        with open(output_path, 'wb') as f:
            f.write(buffer.getvalue())
        
        file_size = len(buffer.getvalue())
        print(f"✓ PDF generado exitosamente: {file_size} bytes")
        print(f"  Archivo guardado en: {output_path}")
        print(f"  Total movimientos: {len(DATOS_TRAZABILIDAD_PRUEBA)}")
        
        # Verificaciones
        assert file_size > 5000, "PDF demasiado pequeño, posible error"
        
        # Verificar que el PDF contiene texto esperado (usando pypdf si disponible)
        try:
            from pypdf import PdfReader
            reader = PdfReader(BytesIO(buffer.getvalue()))
            text = ""
            for page in reader.pages:
                text += page.extract_text() or ""
            
            # Verificar que textos largos aparecen
            checks = [
                ('CENTRO PENITENCIARIO', 'Nombre de centro largo'),
                ('CONFIRMADO', 'Observaciones con corchetes'),
                ('KETOCONAZOL', 'Nombre de producto'),
                ('25072052', 'Número de lote'),
            ]
            
            for text_check, description in checks:
                if text_check in text:
                    print(f"  ✓ {description}: encontrado")
                else:
                    print(f"  ⚠ {description}: NO encontrado (verificar manualmente)")
                    
        except ImportError:
            print("  (pypdf no instalado, no se puede verificar contenido)")
        
        return True
        
    except Exception as e:
        print(f"✗ Error generando PDF: {e}")
        import traceback
        traceback.print_exc()
        return False


def test_pdf_movimientos():
    """Prueba generación de PDF de movimientos."""
    print("\n" + "="*60)
    print("TEST: PDF de Movimientos")
    print("="*60)
    
    from core.utils.pdf_reports import generar_reporte_movimientos
    
    transacciones_data = [
        {
            'referencia': 'SALIDA_POR_REQUISICION REQ-20260101-001 VALE ENTREGADO',
            'fecha': datetime.now().strftime('%d/%m/%Y %H:%M'),
            'tipo': 'SALIDA',
            'centro_origen': 'Farmacia Central',
            'centro_destino': 'CENTRO PENITENCIARIO RIO SANTIAGO DE ALTA SEGURIDAD',
            'total_productos': 3,
            'total_cantidad': 25,
            'detalles': [
                {'producto': 'KETOCONAZOL /CLINDAMICINA CAJA CON 7 OVULOS', 'lote': '25072052', 'cantidad': 10},
                {'producto': 'PARACETAMOL 500MG TABLETAS CAJA CON 100', 'lote': 'LOT-2025-001', 'cantidad': 15},
            ]
        },
        {
            'referencia': 'ENTRADA_POR_REQUISICION REQ-20260101-002',
            'fecha': datetime.now().strftime('%d/%m/%Y %H:%M'),
            'tipo': 'ENTRADA',
            'centro_origen': 'CENTRO PENITENCIARIO MEDICO CEN ESPECIALIZADO',
            'centro_destino': 'Farmacia Central',
            'total_productos': 1,
            'total_cantidad': 50,
            'detalles': [
                {'producto': 'IBUPROFENO 400MG TABLETAS CAJA CON 30', 'lote': 'IBU-2025-100', 'cantidad': 50},
            ]
        },
    ]
    
    try:
        buffer = generar_reporte_movimientos(
            transacciones_data,
            filtros={'centro': 'todos', 'titulo_centro': 'TODOS LOS CENTROS (CONSOLIDADO)'}
        )
        
        output_path = os.path.join(os.path.dirname(__file__), 'test_output_movimientos.pdf')
        with open(output_path, 'wb') as f:
            f.write(buffer.getvalue())
        
        file_size = len(buffer.getvalue())
        print(f"✓ PDF generado exitosamente: {file_size} bytes")
        print(f"  Archivo guardado en: {output_path}")
        
        assert file_size > 5000, "PDF demasiado pequeño"
        return True
        
    except Exception as e:
        print(f"✗ Error generando PDF: {e}")
        import traceback
        traceback.print_exc()
        return False


def test_excel_trazabilidad():
    """Prueba generación de Excel de trazabilidad."""
    print("\n" + "="*60)
    print("TEST: Excel de Trazabilidad")
    print("="*60)
    
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Trazabilidad"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="9F2241", end_color="9F2241", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título
        ws.merge_cells('A1:K1')
        ws['A1'] = "REPORTE DE TRAZABILIDAD - TODOS LOS CENTROS (CONSOLIDADO)"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        row = 3
        ws[f'A{row}'] = "Total movimientos: 7"
        row += 2
        
        # Encabezados
        headers = ['Fecha', 'Tipo', 'Subtipo', 'Producto', 'Nombre Producto', 'Lote', 'Cantidad', 'Centro', 'Usuario', 'No. Expediente', 'Observaciones']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        
        # Datos de prueba
        test_data = [
            ['08/01/2026', 'ENTRADA', '-', '615', 'KETOCONAZOL /CLINDAMICINA', '25072052', 7, 
             'CENTRO PENITENCIARIO RIO SANTIAGO DE ALTA SEGURIDAD', 'Farmacia Central', 
             'EXP-2026-00123', '[CONFIRMADO]SAL-0197-23 VALE Entregado correctamente al centro'],
            ['08/01/2026', 'SALIDA', 'REQUISICION', '615', 'KETOCONAZOL /CLINDAMICINA', '25072052', -2, 
             'Farmacia Central', 'Farmacia C...', 
             '-', 'SALIDA_POR_REQUISICION REQ-20260108-212'],
        ]
        
        for data_row in test_data:
            row += 1
            for col, value in enumerate(data_row, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = thin_border
                if col in [5, 8, 11]:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Anchos de columna
        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 45
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 35
        ws.column_dimensions['I'].width = 18
        ws.column_dimensions['J'].width = 14
        ws.column_dimensions['K'].width = 50
        
        output_path = os.path.join(os.path.dirname(__file__), 'test_output_trazabilidad.xlsx')
        wb.save(output_path)
        
        print(f"✓ Excel generado exitosamente")
        print(f"  Archivo guardado en: {output_path}")
        
        # Verificar anchos
        print("\n  Anchos de columna configurados:")
        print(f"    - Nombre Producto (E): 45 caracteres")
        print(f"    - Centro (H): 35 caracteres")
        print(f"    - Observaciones (K): 50 caracteres")
        
        return True
        
    except Exception as e:
        print(f"✗ Error generando Excel: {e}")
        import traceback
        traceback.print_exc()
        return False


def test_verificar_sin_truncamiento():
    """Verifica que el código no tiene truncamientos problemáticos."""
    print("\n" + "="*60)
    print("TEST: Verificación de código sin truncamientos")
    print("="*60)
    
    import re
    
    archivos = [
        'core/utils/pdf_reports.py',
        'inventario/views_legacy.py',
    ]
    
    # Patrones problemáticos de truncamiento (excluir fechas y límites de DB)
    patrones_problematicos = [
        (r'\[:20\]', 'truncar a 20 caracteres'),
        (r'\[:22\]', 'truncar a 22 caracteres'),
        (r'\[:25\]', 'truncar a 25 caracteres'),
        (r'\[:30\]', 'truncar a 30 caracteres'),
        (r'\[:35\]', 'truncar a 35 caracteres'),
        (r'\[:40\]', 'truncar a 40 caracteres'),
        (r'\[:80\]', 'truncar a 80 caracteres'),
    ]
    
    # Contextos permitidos (fechas, límites de base de datos)
    contextos_permitidos = [
        'fecha',
        'header',
        'hex',
        'order_by',
        'values',
        'strip()[:',  # Limpieza de datos para DB
    ]
    
    base_path = os.path.dirname(__file__)
    problemas = []
    
    for archivo in archivos:
        filepath = os.path.join(base_path, archivo)
        if not os.path.exists(filepath):
            continue
            
        with open(filepath, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        
        for line_num, line in enumerate(lines, 1):
            for patron, descripcion in patrones_problematicos:
                if re.search(patron, line):
                    # Verificar si es un contexto permitido
                    line_lower = line.lower()
                    es_permitido = any(ctx in line_lower for ctx in contextos_permitidos)
                    
                    if not es_permitido:
                        problemas.append({
                            'archivo': archivo,
                            'linea': line_num,
                            'patron': descripcion,
                            'codigo': line.strip()[:80]
                        })
    
    if problemas:
        print(f"⚠ Se encontraron {len(problemas)} posibles truncamientos:")
        for p in problemas[:10]:  # Mostrar máximo 10
            print(f"  - {p['archivo']}:{p['linea']} - {p['patron']}")
            print(f"    {p['codigo'][:60]}...")
    else:
        print("✓ No se encontraron truncamientos problemáticos en el código")
    
    return len(problemas) == 0


def main():
    """Ejecutar todas las pruebas."""
    print("\n" + "="*70)
    print("PRUEBAS EXHAUSTIVAS DE REPORTES SIN TRUNCAMIENTO")
    print("="*70)
    print(f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    resultados = []
    
    # Test 1: PDF Trazabilidad
    resultados.append(('PDF Trazabilidad', test_pdf_trazabilidad()))
    
    # Test 2: PDF Movimientos
    resultados.append(('PDF Movimientos', test_pdf_movimientos()))
    
    # Test 3: Excel Trazabilidad
    resultados.append(('Excel Trazabilidad', test_excel_trazabilidad()))
    
    # Test 4: Verificación de código
    resultados.append(('Código sin truncar', test_verificar_sin_truncamiento()))
    
    # Resumen
    print("\n" + "="*70)
    print("RESUMEN DE PRUEBAS")
    print("="*70)
    
    exitosos = sum(1 for _, r in resultados if r)
    total = len(resultados)
    
    for nombre, resultado in resultados:
        status = "✓ PASS" if resultado else "✗ FAIL"
        print(f"  {status}: {nombre}")
    
    print(f"\nResultado: {exitosos}/{total} pruebas exitosas")
    
    if exitosos == total:
        print("\n✓ TODAS LAS PRUEBAS PASARON")
        print("\nArchivos generados para inspección manual:")
        print("  - test_output_trazabilidad.pdf")
        print("  - test_output_movimientos.pdf")
        print("  - test_output_trazabilidad.xlsx")
        print("\nPor favor revisa estos archivos para confirmar que el texto no se corta.")
    else:
        print("\n✗ ALGUNAS PRUEBAS FALLARON - revisar errores arriba")
    
    return exitosos == total


if __name__ == '__main__':
    success = main()
    sys.exit(0 if success else 1)

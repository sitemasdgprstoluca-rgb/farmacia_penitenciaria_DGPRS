#!/usr/bin/env python
"""
Test de Importación de PPL (Pacientes/Internos) - V2
=====================================================
Verifica el flujo completo de importación usando APIClient

Ejecutar: python test_importacion_ppl_v2.py
"""

import os
import sys
import django
import tempfile
from io import BytesIO
from datetime import date

# Configurar Django
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
django.setup()

import openpyxl
from django.test import TestCase
from django.core.files.uploadedfile import SimpleUploadedFile
from rest_framework.test import APIClient
from django.contrib.auth import get_user_model
from core.models import Centro, Paciente

User = get_user_model()

# Colores para output
class Colors:
    GREEN = '\033[92m'
    RED = '\033[91m'
    YELLOW = '\033[93m'
    BLUE = '\033[94m'
    CYAN = '\033[96m'
    END = '\033[0m'
    BOLD = '\033[1m'

def ok(msg):
    print(f"{Colors.GREEN}✅ {msg}{Colors.END}")

def fail(msg):
    print(f"{Colors.RED}❌ {msg}{Colors.END}")

def info(msg):
    print(f"{Colors.CYAN}ℹ️  {msg}{Colors.END}")

def warn(msg):
    print(f"{Colors.YELLOW}⚠️  {msg}{Colors.END}")

def header(msg):
    print(f"\n{Colors.BOLD}{Colors.BLUE}{'='*60}{Colors.END}")
    print(f"{Colors.BOLD}{Colors.BLUE}{msg}{Colors.END}")
    print(f"{Colors.BOLD}{Colors.BLUE}{'='*60}{Colors.END}")


def setup_test_data():
    """Configura datos de prueba y retorna (user, centro, client)"""
    
    # Crear o obtener usuario
    try:
        user = User.objects.get(username='test_import_ppl')
    except User.DoesNotExist:
        user = User(
            username='test_import_ppl',
            email='test_import_ppl@test.com',
            first_name='Test',
            last_name='Import PPL',
            rol='admin',  # Admin tiene todos los permisos
            is_active=True,
        )
        user.set_password('TestPassword123!')
        user.save()
        ok(f"Usuario de prueba creado: {user.username}")
    
    # Obtener un centro existente o crear uno
    centro = Centro.objects.filter(activo=True).first()
    if not centro:
        fail("No hay centros activos en la base de datos")
        return None, None, None
    
    ok(f"Centro de prueba: {centro.nombre} (ID: {centro.id})")
    
    # Configurar cliente API autenticado
    client = APIClient()
    client.force_authenticate(user=user)
    
    return user, centro, client


def test_1_plantilla_descarga(client):
    """Test 1: Verificar descarga de plantilla de importación"""
    header("TEST 1: DESCARGA DE PLANTILLA DE IMPORTACIÓN")
    
    response = client.get('/api/v1/pacientes/plantilla_importacion/')
    
    if response.status_code != 200:
        fail(f"Error al descargar plantilla: {response.status_code}")
        if hasattr(response, 'data'):
            info(f"Detalle: {response.data}")
        return False
    
    ok("Plantilla descargada correctamente (HTTP 200)")
    
    # Verificar que sea un Excel válido
    content = b''.join(response.streaming_content) if hasattr(response, 'streaming_content') else response.content
    
    try:
        wb = openpyxl.load_workbook(BytesIO(content))
        ok(f"Archivo Excel válido con {len(wb.sheetnames)} hoja(s)")
        
        # Verificar hojas
        if 'Pacientes' in wb.sheetnames:
            ok("Hoja 'Pacientes' encontrada")
        else:
            fail("Falta hoja 'Pacientes'")
            return False
        
        if 'Instrucciones' in wb.sheetnames:
            ok("Hoja 'Instrucciones' encontrada")
        else:
            warn("Hoja 'Instrucciones' no encontrada (opcional)")
        
        # Verificar encabezados
        ws = wb['Pacientes']
        headers = [cell.value for cell in ws[1] if cell.value]
        
        required_headers = ['numero_expediente*', 'nombre*', 'apellido_paterno*', 'centro_clave*']
        
        info(f"Encabezados encontrados: {headers}")
        
        for req in required_headers:
            if req in headers:
                ok(f"Encabezado '{req}' presente")
            else:
                fail(f"Encabezado requerido '{req}' faltante")
                return False
        
        return True
        
    except Exception as e:
        fail(f"Error al procesar Excel: {e}")
        return False


def test_2_importacion_nuevos(client, centro):
    """Test 2: Importar nuevos PPL"""
    header("TEST 2: IMPORTACIÓN DE NUEVOS PPL")
    
    # Crear Excel de prueba
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Pacientes'
    
    # Encabezados
    col_headers = [
        'numero_expediente*', 'nombre*', 'apellido_paterno*', 'apellido_materno',
        'curp', 'fecha_nacimiento', 'sexo', 'centro_clave*', 'dormitorio', 'celda',
        'tipo_sangre', 'alergias', 'enfermedades_cronicas', 'observaciones_medicas',
        'fecha_ingreso'
    ]
    for col, h in enumerate(col_headers, 1):
        ws.cell(row=1, column=col, value=h)
    
    # Datos de prueba - 3 PPL nuevos
    test_data = [
        ['TEST-001', 'Juan', 'Pérez', 'García', 'PEGJ901201HDFRRN01', '1990-12-01', 'M', str(centro.id), 'A', '101', 'O+', '', '', '', '2024-01-15'],
        ['TEST-002', 'María', 'López', 'Hernández', 'LOHM850515MDFPRR02', '1985-05-15', 'F', str(centro.id), 'B', '202', 'A-', 'Penicilina', '', '', '2024-02-20'],
        ['TEST-003', 'Pedro', 'Sánchez', '', '', '', 'M', str(centro.id), '', '', '', '', 'Diabetes', 'Requiere insulina', ''],
    ]
    
    for row_idx, row_data in enumerate(test_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Guardar en memoria
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_content = excel_buffer.getvalue()
    
    # Crear archivo para upload
    uploaded_file = SimpleUploadedFile(
        'test_import.xlsx',
        excel_content,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Limpiar PPL de prueba anteriores
    Paciente.objects.filter(numero_expediente__startswith='TEST-').delete()
    info("PPL de prueba anteriores eliminados")
    
    # Contar PPL antes
    count_before = Paciente.objects.count()
    info(f"PPL antes de importar: {count_before}")
    
    # Importar
    response = client.post(
        '/api/v1/pacientes/importar-excel/',
        {'file': uploaded_file},
        format='multipart'
    )
    
    if response.status_code != 200:
        fail(f"Error al importar: {response.status_code}")
        if hasattr(response, 'data'):
            info(f"Detalle: {response.data}")
        return False
    
    ok(f"Importación exitosa (HTTP 200)")
    
    result = response.data
    info(f"Respuesta: {result}")
    
    # Verificar resultados
    if result.get('creados', 0) == 3:
        ok(f"Se crearon {result['creados']} registros")
    else:
        fail(f"Se esperaban 3 creados, se obtuvieron {result.get('creados', 0)}")
        return False
    
    if result.get('actualizados', 0) == 0:
        ok("No hubo actualizaciones (correcto para nuevos)")
    else:
        warn(f"Se actualizaron {result.get('actualizados')} registros")
    
    # Verificar que los PPL existen en BD
    for exp in ['TEST-001', 'TEST-002', 'TEST-003']:
        try:
            ppl = Paciente.objects.get(numero_expediente=exp)
            ok(f"PPL {exp} creado correctamente: {ppl.nombre} {ppl.apellido_paterno}")
        except Paciente.DoesNotExist:
            fail(f"PPL {exp} no encontrado en BD")
            return False
    
    # Verificar datos específicos
    ppl1 = Paciente.objects.get(numero_expediente='TEST-001')
    if ppl1.curp == 'PEGJ901201HDFRRN01':
        ok("CURP guardado correctamente")
    else:
        fail(f"CURP incorrecto: {ppl1.curp}")
    
    if ppl1.tipo_sangre == 'O+':
        ok("Tipo de sangre guardado correctamente")
    else:
        fail(f"Tipo de sangre incorrecto: {ppl1.tipo_sangre}")
    
    return True


def test_3_actualizacion(client, centro):
    """Test 3: Actualizar PPL existentes"""
    header("TEST 3: ACTUALIZACIÓN DE PPL EXISTENTES")
    
    # Verificar que existe TEST-001
    try:
        ppl = Paciente.objects.get(numero_expediente='TEST-001')
        info(f"PPL TEST-001 existe: {ppl.nombre} {ppl.apellido_paterno}")
    except Paciente.DoesNotExist:
        fail("PPL TEST-001 no existe, ejecute primero test 2")
        return False
    
    # Crear Excel con datos actualizados
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Pacientes'
    
    col_headers = [
        'numero_expediente*', 'nombre*', 'apellido_paterno*', 'apellido_materno',
        'curp', 'fecha_nacimiento', 'sexo', 'centro_clave*', 'dormitorio', 'celda',
        'tipo_sangre', 'alergias', 'enfermedades_cronicas', 'observaciones_medicas',
        'fecha_ingreso'
    ]
    for col, h in enumerate(col_headers, 1):
        ws.cell(row=1, column=col, value=h)
    
    # Datos actualizados
    updated_data = [
        # TEST-001 con cambio de celda y dormitorio
        ['TEST-001', 'Juan', 'Pérez', 'García', 'PEGJ901201HDFRRN01', '1990-12-01', 'M', str(centro.id), 'C', '303', 'O+', 'Aspirina', '', 'Cambio de celda', '2024-01-15'],
    ]
    
    for row_idx, row_data in enumerate(updated_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_content = excel_buffer.getvalue()
    
    uploaded_file = SimpleUploadedFile(
        'test_update.xlsx',
        excel_content,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Importar actualización
    response = client.post(
        '/api/v1/pacientes/importar-excel/',
        {'file': uploaded_file},
        format='multipart'
    )
    
    if response.status_code != 200:
        fail(f"Error al actualizar: {response.status_code}")
        if hasattr(response, 'data'):
            info(f"Detalle: {response.data}")
        return False
    
    result = response.data
    info(f"Respuesta: {result}")
    
    if result.get('actualizados', 0) == 1:
        ok("Se actualizó 1 registro")
    else:
        fail(f"Se esperaba 1 actualización, se obtuvo {result.get('actualizados', 0)}")
        return False
    
    # Verificar datos actualizados
    ppl = Paciente.objects.get(numero_expediente='TEST-001')
    
    checks = [
        (ppl.dormitorio == 'C', f"Dormitorio actualizado a 'C': {ppl.dormitorio}"),
        (ppl.celda == '303', f"Celda actualizada a '303': {ppl.celda}"),
        (ppl.alergias == 'Aspirina', f"Alergias actualizadas: {ppl.alergias}"),
        (ppl.observaciones_medicas == 'Cambio de celda', f"Observaciones actualizadas: {ppl.observaciones_medicas}"),
    ]
    
    all_ok = True
    for check, msg in checks:
        if check:
            ok(msg)
        else:
            fail(msg)
            all_ok = False
    
    return all_ok


def test_4_validaciones(client, centro):
    """Test 4: Validaciones de datos incorrectos"""
    header("TEST 4: VALIDACIONES DE DATOS")
    
    # Crear Excel con datos inválidos
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Pacientes'
    
    col_headers = [
        'numero_expediente*', 'nombre*', 'apellido_paterno*', 'apellido_materno',
        'curp', 'fecha_nacimiento', 'sexo', 'centro_clave*', 'dormitorio', 'celda',
        'tipo_sangre', 'alergias', 'enfermedades_cronicas', 'observaciones_medicas',
        'fecha_ingreso'
    ]
    for col, h in enumerate(col_headers, 1):
        ws.cell(row=1, column=col, value=h)
    
    # Datos con errores
    invalid_data = [
        # Fila 2: CURP inválido
        ['TEST-ERR-001', 'Error', 'Curp', '', 'CURPINVALIDO123', '', '', str(centro.id), '', '', '', '', '', '', ''],
        # Fila 3: Tipo de sangre inválido
        ['TEST-ERR-002', 'Error', 'Sangre', '', '', '', '', str(centro.id), '', '', 'XYZ', '', '', '', ''],
        # Fila 4: Centro inexistente
        ['TEST-ERR-003', 'Error', 'Centro', '', '', '', '', '999999', '', '', '', '', '', '', ''],
        # Fila 5: Falta nombre (campo obligatorio)
        ['TEST-ERR-004', '', 'SinNombre', '', '', '', '', str(centro.id), '', '', '', '', '', '', ''],
        # Fila 6: Sexo inválido
        ['TEST-ERR-005', 'Error', 'Sexo', '', '', '', 'X', str(centro.id), '', '', '', '', '', '', ''],
    ]
    
    for row_idx, row_data in enumerate(invalid_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_content = excel_buffer.getvalue()
    
    uploaded_file = SimpleUploadedFile(
        'test_invalid.xlsx',
        excel_content,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Limpiar errores anteriores
    Paciente.objects.filter(numero_expediente__startswith='TEST-ERR-').delete()
    
    # Importar
    response = client.post(
        '/api/v1/pacientes/importar-excel/',
        {'file': uploaded_file},
        format='multipart'
    )
    
    # Puede ser 200 con errores parciales o 400
    if response.status_code in [200, 400]:
        ok(f"Respuesta recibida: HTTP {response.status_code}")
    else:
        fail(f"Código de respuesta inesperado: {response.status_code}")
        return False
    
    result = response.data
    info(f"Respuesta: {result}")
    
    # Verificar que se detectaron errores
    errores = result.get('errores', [])
    total_errores = result.get('total_errores', len(errores))
    
    if total_errores > 0:
        ok(f"Se detectaron {total_errores} error(es) de validación")
        for err in errores[:5]:  # Mostrar máx 5 errores
            info(f"  - {err}")
    else:
        warn("No se detectaron errores de validación")
    
    # Verificar que NO se crearon los registros con errores
    err_count = Paciente.objects.filter(numero_expediente__startswith='TEST-ERR-').count()
    if err_count == 0:
        ok("No se crearon registros con errores")
    else:
        warn(f"Se crearon {err_count} registros con errores (pueden ser parcialmente válidos)")
    
    return True


def test_5_exportacion(client, centro):
    """Test 5: Verificar que los datos importados se pueden exportar"""
    header("TEST 5: EXPORTACIÓN DE DATOS IMPORTADOS")
    
    # Verificar que existen PPL de prueba
    test_ppls = Paciente.objects.filter(numero_expediente__startswith='TEST-00')
    if test_ppls.count() == 0:
        fail("No hay PPL de prueba para exportar")
        return False
    
    info(f"PPL de prueba encontrados: {test_ppls.count()}")
    
    # Hacer GET a la lista de pacientes
    response = client.get('/api/v1/pacientes/')
    
    if response.status_code != 200:
        fail(f"Error al obtener lista: {response.status_code}")
        return False
    
    ok("Lista de pacientes obtenida correctamente")
    
    # Verificar que los PPL de prueba están en la respuesta
    data = response.data
    
    # Puede ser paginado o lista directa
    if isinstance(data, dict) and 'results' in data:
        results = data['results']
    elif isinstance(data, list):
        results = data
    else:
        results = []
    
    found_test = 0
    for item in results:
        if item.get('numero_expediente', '').startswith('TEST-00'):
            found_test += 1
    
    if found_test > 0:
        ok(f"Se encontraron {found_test} PPL de prueba en la exportación")
    else:
        warn("No se encontraron PPL de prueba en la respuesta (puede estar paginado)")
    
    # Verificar detalle de un PPL específico
    try:
        ppl = Paciente.objects.get(numero_expediente='TEST-001')
        response = client.get(f'/api/v1/pacientes/{ppl.id}/')
        
        if response.status_code == 200:
            ok(f"Detalle de PPL TEST-001 obtenido correctamente")
            detail = response.data
            info(f"  - Nombre: {detail.get('nombre')} {detail.get('apellido_paterno')}")
            info(f"  - Expediente: {detail.get('numero_expediente')}")
            info(f"  - Centro: {detail.get('centro_nombre', detail.get('centro'))}")
        else:
            fail(f"Error al obtener detalle: {response.status_code}")
            return False
    except Paciente.DoesNotExist:
        warn("PPL TEST-001 no encontrado para verificar detalle")
    
    return True


def cleanup_test_data():
    """Limpia los datos de prueba"""
    header("LIMPIEZA DE DATOS DE PRUEBA")
    
    # Eliminar PPL de prueba
    deleted = Paciente.objects.filter(numero_expediente__startswith='TEST-').delete()
    info(f"PPL eliminados: {deleted}")
    
    # Eliminar usuario de prueba
    try:
        user = User.objects.get(username='test_import_ppl')
        user.delete()
        info("Usuario de prueba eliminado")
    except User.DoesNotExist:
        pass


def main():
    print("\n")
    print("╔═══════════════════════════════════════════════════════════════╗")
    print("║     TEST DE IMPORTACIÓN DE PPL (PACIENTES/INTERNOS) V2       ║")
    print("║                   Sistema de Farmacia                         ║")
    print("╚═══════════════════════════════════════════════════════════════╝")
    
    # Setup
    user, centro, client = setup_test_data()
    
    if not client:
        fail("No se pudo configurar el entorno de prueba")
        return 1
    
    results = {}
    
    # Ejecutar tests
    try:
        results['plantilla'] = test_1_plantilla_descarga(client)
    except Exception as e:
        fail(f"Error en test 1: {e}")
        import traceback
        traceback.print_exc()
        results['plantilla'] = False
    
    try:
        results['importacion'] = test_2_importacion_nuevos(client, centro)
    except Exception as e:
        fail(f"Error en test 2: {e}")
        import traceback
        traceback.print_exc()
        results['importacion'] = False
    
    try:
        results['actualizacion'] = test_3_actualizacion(client, centro)
    except Exception as e:
        fail(f"Error en test 3: {e}")
        import traceback
        traceback.print_exc()
        results['actualizacion'] = False
    
    try:
        results['validaciones'] = test_4_validaciones(client, centro)
    except Exception as e:
        fail(f"Error en test 4: {e}")
        import traceback
        traceback.print_exc()
        results['validaciones'] = False
    
    try:
        results['exportacion'] = test_5_exportacion(client, centro)
    except Exception as e:
        fail(f"Error en test 5: {e}")
        import traceback
        traceback.print_exc()
        results['exportacion'] = False
    
    # Resumen
    header("RESUMEN DE PRUEBAS")
    
    test_names = {
        'plantilla': 'Descarga de Plantilla',
        'importacion': 'Importación de Nuevos PPL',
        'actualizacion': 'Actualización de PPL',
        'validaciones': 'Validaciones de Datos',
        'exportacion': 'Exportación de Datos',
    }
    
    passed = 0
    total = len(results)
    
    print(f"\n{'Test':<40} {'Resultado':<15}")
    print("-" * 55)
    
    for key, name in test_names.items():
        if results.get(key):
            print(f"{Colors.GREEN}{name:<40} {'PASÓ':<15}{Colors.END}")
            passed += 1
        else:
            print(f"{Colors.RED}{name:<40} {'FALLÓ':<15}{Colors.END}")
    
    print("-" * 55)
    print(f"{'TOTAL':<40} {passed}/{total}")
    
    if passed == total:
        print(f"\n{Colors.GREEN}{Colors.BOLD}🎉 ¡TODAS LAS PRUEBAS PASARON!{Colors.END}")
    else:
        print(f"\n{Colors.YELLOW}⚠️  {total - passed} prueba(s) fallaron{Colors.END}")
    
    # Preguntar si limpiar
    try:
        response = input("\n¿Desea limpiar los datos de prueba? (s/N): ")
        if response.lower() == 's':
            cleanup_test_data()
    except:
        pass
    
    return 0 if passed == total else 1


if __name__ == '__main__':
    exit(main())

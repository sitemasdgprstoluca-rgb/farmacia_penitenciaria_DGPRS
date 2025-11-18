from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import AllowAny
from rest_framework.views import APIView
from rest_framework_simplejwt.tokens import RefreshToken
from django.contrib.auth import authenticate
from django.contrib.auth.models import Group
from django.db.models import Sum, Count, Q
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from django.http import HttpResponse
from datetime import timedelta
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import traceback

from core.models import (
    User, Centro, Producto, Lote, Movimiento,
    Requisicion, DetalleRequisicion, ImportacionLog, AuditoriaLog
)
from core.permissions import (
    IsFarmaciaAdminOrReadOnly, IsFarmaciaAdmin, IsVistaUserOrAdmin
)
from inventario.serializers import (
    ProductoSerializer, LoteSerializer, MovimientoSerializer,
    CentroSerializer, RequisicionSerializer, DetalleRequisicionSerializer,
    UserSerializer, AuditoriaProductoSerializer
)


class MeView(APIView):
    permission_classes = [AllowAny]
    
    def get(self, request):
        if request.user.is_authenticated:
            return Response(UserSerializer(request.user).data)
        return Response({
            'id': 1,
            'username': 'super_admin',
            'email': 'admin@edomex.gob.mx',
            'first_name': 'Super',
            'last_name': 'Administrador',
            'grupos': ['SUPERUSER'],
            'rol': 'SUPERUSER',
            'is_superuser': True
        })


class LoginView(APIView):
    permission_classes = [AllowAny]
    
    def post(self, request):
        username = request.data.get('username')
        password = request.data.get('password')
        
        if not username or not password:
            return Response({
                'detail': 'Usuario y contraseña son requeridos'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        user = authenticate(username=username, password=password)
        
        if user:
            # Generar tokens JWT
            refresh = RefreshToken.for_user(user)
            
            # Obtener grupos del usuario
            grupos = list(user.groups.values_list('name', flat=True))
            
            return Response({
                'access': str(refresh.access_token),
                'refresh': str(refresh),
                'usuario': {
                    'id': user.id,
                    'username': user.username,
                    'email': user.email,
                    'first_name': user.first_name,
                    'last_name': user.last_name,
                    'grupos': grupos,
                    'is_superuser': user.is_superuser,
                }
            }, status=status.HTTP_200_OK)
        
        return Response({
            'detail': 'Credenciales inválidas'
        }, status=status.HTTP_401_UNAUTHORIZED)


class LogoutView(APIView):
    def post(self, request):
        return Response({'message': 'Logout exitoso'}, status=status.HTTP_200_OK)


class DashboardView(APIView):
    permission_classes = [AllowAny]
    
    def get(self, request):
        try:
            hoy = timezone.now().date()
            
            return Response({
                'stock': {
                    'critico': {'total': 0, 'nivel': 'CRITICO', 'productos': []},
                    'bajo': {'total': 0, 'nivel': 'BAJO', 'productos': []}
                },
                'caducidad': {
                    'critico': {'total': 0, 'nivel': 'CRITICO', 'lotes': []},
                    'proximo': {'total': 0, 'nivel': 'PROXIMO', 'lotes': []},
                    'caducado': {'total': 0, 'nivel': 'VENCIDO', 'lotes': []}
                },
                'requisiciones': {'pendientes': 0, 'por_autorizar': 0, 'por_surtir': 0},
                'totales': {
                    'productos_activos': Producto.objects.filter(activo=True).count(),
                    'lotes_activos': Lote.objects.filter(estado='disponible').count(),
                    'centros_activos': Centro.objects.filter(activo=True).count()
                }
            })
        except Exception as e:
            return Response({'error': str(e)}, status=500)


class TrazabilidadProductoView(APIView):
    permission_classes = [AllowAny]
    
    def get(self, request, clave):
        try:
            producto = Producto.objects.get(clave=clave)
            return Response({
                'producto': ProductoSerializer(producto).data,
                'lotes': {},
                'total_movimientos': 0
            })
        except Producto.DoesNotExist:
            return Response({'error': 'Producto no encontrado'}, status=404)


class TrazabilidadLoteView(APIView):
    permission_classes = [AllowAny]
    
    def get(self, request, codigo):
        try:
            lote = Lote.objects.get(codigo_lote=codigo)
            return Response({
                'lote': LoteSerializer(lote).data,
                'producto': ProductoSerializer(lote.producto).data,
                'historial': [],
                'total_movimientos': 0
            })
        except Lote.DoesNotExist:
            return Response({'error': 'Lote no encontrado'}, status=404)


class ReportesPrecargaView(APIView):
    permission_classes = [AllowAny]
    
    def get(self, request):
        productos = Producto.objects.filter(activo=True).values('id', 'clave', 'descripcion')[:50]
        lotes = Lote.objects.filter(estado='disponible').values('id', 'numero_lote')[:50]
        centros = Centro.objects.filter(activo=True).values('id', 'clave', 'nombre')
        
        return Response({
            'productos': list(productos),
            'lotes': list(lotes),
            'centros': list(centros)
        })


class UsuariosViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [AllowAny]


class CentroViewSet(viewsets.ModelViewSet):
    """CRUD de Centros"""
    queryset = Centro.objects.all()
    serializer_class = CentroSerializer
    permission_classes = [AllowAny]

    def get_queryset(self):
        queryset = Centro.objects.all()
        activo = self.request.query_params.get('activo')
        if activo is not None:
            queryset = queryset.filter(activo=activo.lower() == 'true')
        return queryset.order_by('-created_at')

    @action(detail=False, methods=['get'])
    def exportar_excel(self, request):
        try:
            centros = self.get_queryset()
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Centros'
            ws.append(['Clave', 'Nombre', 'Dirección', 'Teléfono', 'Estado'])
            
            for centro in centros:
                ws.append([centro.clave, centro.nombre, centro.direccion or '', centro.telefono or '', 'Activo' if centro.activo else 'Inactivo'])
            
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=Centros.xlsx'
            wb.save(response)
            return response
        except Exception as e:
            return Response({'error': str(e)}, status=500)

    @action(detail=False, methods=['post'])
    def importar_excel(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No se recibió archivo'}, status=400)
        
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            creados = 0
            actualizados = 0
            errores = []
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    clave, nombre, direccion, telefono, activo = row[:5]
                    if not clave or not nombre:
                        errores.append(f'Fila {row_idx}: Clave y nombre requeridos')
                        continue
                    
                    centro, created = Centro.objects.update_or_create(
                        clave=str(clave).upper().strip(),
                        defaults={
                            'nombre': str(nombre).strip(),
                            'direccion': str(direccion).strip() if direccion else '',
                            'telefono': str(telefono).strip() if telefono else '',
                            'activo': str(activo).lower() in ['activo', 'sí', 'si', 'true', '1'] if activo else True
                        }
                    )
                    if created:
                        creados += 1
                    else:
                        actualizados += 1
                except Exception as e:
                    errores.append(f'Fila {row_idx}: {str(e)}')
            
            return Response({'mensaje': 'Importación completada', 'creados': creados, 'actualizados': actualizados, 'errores': errores})
        except Exception as e:
            return Response({'error': str(e)}, status=400)

    def destroy(self, request, *args, **kwargs):
        """Eliminar centro con validación"""
        instance = self.get_object()
        try:
            # Verificar si tiene requisiciones
            if hasattr(instance, 'requisiciones') and instance.requisiciones.exists():
                return Response(
                    {'error': 'No se puede eliminar el centro: tiene requisiciones asociadas.'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            instance.delete()
            return Response(status=status.HTTP_204_NO_CONTENT)
        except Exception as e:
            traceback.print_exc()
            return Response(
                {'error': 'Error interno al eliminar el centro.'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

@method_decorator(csrf_exempt, name='dispatch')
class LoteViewSet(viewsets.ModelViewSet):
    """CRUD de Lotes"""
    queryset = Lote.objects.select_related('producto').all()
    serializer_class = LoteSerializer
    permission_classes = [AllowAny]

    def destroy(self, request, *args, **kwargs):
        """
        Validación segura al eliminar un lote:
        - Si existen movimientos asociados, devolver 400 con mensaje.
        - Si ocurre cualquier otro error, devolver 500 con detalle mínimo y loguear el traceback.
        """
        instance = self.get_object()
        try:
            # Comprobar movimientos asociados (asegurar nombre correcto del FK)
            # Ajusta el filtro si tu modelo Movimiento usa otro nombre de campo.
            if Movimiento.objects.filter(lote=instance).exists():
                return Response(
                    {'error': 'No se puede eliminar el lote: tiene movimientos asociados.'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            instance.delete()
            return Response(status=status.HTTP_204_NO_CONTENT)
        except Exception as e:
            # Loguear para diagnóstico en el servidor (no devolver stack al cliente)
            traceback.print_exc()
            return Response(
                {'error': 'Error interno al eliminar el lote. Ver logs del servidor.'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def perform_create(self, serializer):
        # Validar y guardar el lote
        serializer.save()

    def perform_update(self, serializer):
        # Validar y actualizar el lote
        serializer.save()

    def get_queryset(self):
        queryset = Lote.objects.select_related('producto').all()
        
        producto = self.request.query_params.get('producto')
        if producto:
            queryset = queryset.filter(producto_id=producto)
        
        con_stock = self.request.query_params.get('con_stock')
        if con_stock is not None:
            if con_stock.lower() == 'true':
                queryset = queryset.filter(cantidad_actual__gt=0)
        
        activo = self.request.query_params.get('activo')
        if activo is not None:
            queryset = queryset.filter(activo=activo.lower() == 'true')
        
        return queryset.order_by('-created_at')


@method_decorator(csrf_exempt, name='dispatch')
class ProductoViewSet(viewsets.ModelViewSet):
    """CRUD de Productos con permisos y auditoría"""
    queryset = Producto.objects.all()
    serializer_class = ProductoSerializer
    permission_classes = [IsFarmaciaAdminOrReadOnly]

    def get_queryset(self):
        queryset = Producto.objects.all()
        search = self.request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                Q(clave__icontains=search) | Q(descripcion__icontains=search)
            )
        return queryset

    def get_permissions(self):
        if self.action == 'exportar_excel':
            permission_classes = [IsVistaUserOrAdmin]
        elif self.action in ['importar_excel', 'auditoria']:
            permission_classes = [IsFarmaciaAdmin]
        else:
            permission_classes = self.permission_classes
        return [permission() for permission in permission_classes]

    def perform_create(self, serializer):
        created_by = self.request.user if self.request.user.is_authenticated else None
        serializer.save(created_by=created_by)

    def perform_update(self, serializer):
        serializer.save()

    def perform_destroy(self, instance):
        instance.delete()

    @action(detail=False, methods=['get'], permission_classes=[IsVistaUserOrAdmin])
    def exportar_excel(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Productos'
        ws.append(['Clave', 'Descripción', 'Unidad', 'Precio', 'Activo'])

        for p in self.get_queryset():
            ws.append([
                p.clave,
                p.descripcion,
                p.unidad_medida,
                float(p.precio_unitario) if p.precio_unitario else 0,
                'Sí' if p.activo else 'No'
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=Productos.xlsx'
        wb.save(response)
        return response

    @action(detail=False, methods=['post'], permission_classes=[IsFarmaciaAdmin])
    def importar_excel(self, request):
        """Importa productos desde Excel con registro en auditoría"""
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'No se recibió archivo'}, status=status.HTTP_400_BAD_REQUEST)

        productos_procesados = []
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active

            creados = 0
            actualizados = 0
            errores = []

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    valores = list(row) + [None] * 10
                    clave = str(valores[0]).strip().upper() if valores[0] else None
                    descripcion = str(valores[1]).strip() if valores[1] else None
                    unidad_medida = str(valores[2]).strip().upper() if valores[2] else 'PIEZA'
                    precio_unitario = (
                        Decimal(str(valores[3]))
                        if valores[3] and self._is_numeric(valores[3])
                        else Decimal('0.00')
                    )
                    stock_minimo = int(valores[4]) if valores[4] and str(valores[4]).isdigit() else 10
                    estado = str(valores[5]).lower() if valores[5] else 'true'
                    activo = estado in ['activo', 'sí', 'si', 'true', '1', 'yes']

                    if not clave or not descripcion:
                        errores.append(f'Fila {row_idx}: Clave y descripción son obligatorios')
                        continue

                    producto, created = Producto.objects.update_or_create(
                        clave=clave,
                        defaults={
                            'descripcion': descripcion,
                            'unidad_medida': unidad_medida,
                            'precio_unitario': precio_unitario,
                            'stock_minimo': stock_minimo,
                            'activo': activo
                        }
                    )

                    if created and request.user.is_authenticated:
                        producto.created_by = request.user
                        producto.save(update_fields=['created_by'])

                    productos_procesados.append({
                        'clave': producto.clave,
                        'accion': 'creado' if created else 'actualizado'
                    })

                    if created:
                        creados += 1
                    else:
                        actualizados += 1

                except Exception as e:
                    errores.append(f'Fila {row_idx}: {str(e)}')

            total = creados + actualizados
            estado_importacion = 'exitosa'
            if errores and total:
                estado_importacion = 'parcial'
            elif errores and not total:
                estado_importacion = 'fallida'

            ImportacionLog.objects.create(
                usuario=request.user if request.user.is_authenticated else None,
                archivo_nombre=file.name,
                modelo='Producto',
                total_registros=total + len(errores),
                registros_exitosos=total,
                registros_fallidos=len(errores),
                estado=estado_importacion,
                resultado_procesamiento={
                    'creados': creados,
                    'actualizados': actualizados,
                    'productos': productos_procesados,
                    'errores': errores[:5]
                }
            )

            return Response({
                'mensaje': 'Importación completada y registrada en auditoría',
                'creados': creados,
                'actualizados': actualizados,
                'errores': errores,
                'exito': len(errores) == 0
            }, status=status.HTTP_200_OK)

        except Exception as e:
            ImportacionLog.objects.create(
                usuario=request.user if request.user.is_authenticated else None,
                archivo_nombre=getattr(file, 'name', 'archivo_desconocido'),
                modelo='Producto',
                total_registros=0,
                registros_exitosos=0,
                registros_fallidos=1,
                estado='fallida',
                resultado_procesamiento={'error': str(e)}
            )
            return Response({
                'error': 'Error al procesar archivo: ' + str(e)
            }, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=['get'], permission_classes=[IsFarmaciaAdmin])
    def auditoria(self, request, pk=None):
        producto = self.get_object()
        auditorias = AuditoriaLog.objects.filter(
            modelo='Producto',
            objeto_id=producto.pk
        ).order_by('-fecha')
        serializer = AuditoriaProductoSerializer(auditorias, many=True)
        return Response({
            'producto': ProductoSerializer(
                producto,
                context=self.get_serializer_context()
            ).data,
            'historial': serializer.data,
            'total_cambios': auditorias.count()
        })

    @staticmethod
    def _is_numeric(val):
        try:
            float(val)
            return True
        except Exception:
            return False
class MovimientoViewSet(viewsets.ModelViewSet):
    queryset = Movimiento.objects.select_related('producto', 'lote').all()
    serializer_class = MovimientoSerializer
    permission_classes = [AllowAny]


class RequisicionViewSet(viewsets.ModelViewSet):
    queryset = Requisicion.objects.select_related('centro', 'solicitante').all()
    serializer_class = RequisicionSerializer
    permission_classes = [AllowAny]

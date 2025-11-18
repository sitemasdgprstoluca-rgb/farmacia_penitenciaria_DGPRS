from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
import qrcode
from io import BytesIO


def crear_estilo_titulo():
    """Estilo para títulos de reportes"""
    return ParagraphStyle(
        'CustomTitle',
        parent=getSampleStyleSheet()['Title'],
        fontSize=16,
        textColor=colors.HexColor('#1f2937'),
        spaceAfter=12,
        alignment=TA_CENTER
    )


def crear_estilo_subtitulo():
    """Estilo para subtítulos"""
    return ParagraphStyle(
        'CustomSubtitle',
        parent=getSampleStyleSheet()['Normal'],
        fontSize=12,
        textColor=colors.HexColor('#6b7280'),
        spaceAfter=12,
        alignment=TA_CENTER
    )


def crear_tabla_pdf(data, col_widths=None):
    """Crea una tabla con estilo para PDF"""
    tabla = Table(data, colWidths=col_widths)
    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e5e7eb')),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')])
    ]))
    return tabla


def generar_qr(data):
    """Genera código QR"""
    qr = qrcode.QRCode(version=1, box_size=10, border=5)
    qr.add_data(data)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    
    buffer = BytesIO()
    img.save(buffer, format='PNG')
    buffer.seek(0)
    return buffer


def configurar_hoja_excel(ws, titulo):
    """Configura estilos básicos de hoja Excel"""
    # Título
    ws['A1'] = titulo
    ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Fecha de generación
    ws['A2'] = f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A2'].font = Font(size=10, italic=True)
    ws['A2'].alignment = Alignment(horizontal='center')
    
    return ws


def aplicar_estilo_encabezados(ws, row, max_col):
    """Aplica estilo a los encabezados de columna"""
    header_fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment


def aplicar_bordes(ws, max_row, max_col):
    """Aplica bordes a toda la tabla"""
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            ws.cell(row=row, column=col).border = thin_border


def ajustar_ancho_columnas(ws):
    """Ajusta automáticamente el ancho de las columnas"""
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

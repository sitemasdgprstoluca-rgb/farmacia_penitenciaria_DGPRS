"""
Crea archivo de prueba con 5 lotes nuevos para verificar importador
"""
from openpyxl import load_workbook
from datetime import datetime
import shutil

# Copiar archivo original
shutil.copy(
    r'C:\Users\zarag\Downloads\REVISAR\lotes_2025-12-15.xlsx',
    r'C:\Users\zarag\Downloads\REVISAR\lotes_nuevos_test.xlsx'
)

# Cargar copia
wb = load_workbook(r'C:\Users\zarag\Downloads\REVISAR\lotes_nuevos_test.xlsx')
ws = wb.active

# Modificar solo las primeras 5 filas de datos con números de lote únicos
timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
for i in range(4, 9):
    # Columna C = Número Lote (índice 3)
    numero_lote_unico = f'TEST-LOTE-{i}-{timestamp}'
    ws.cell(row=i, column=3).value = numero_lote_unico
    print(f'Fila {i}: {numero_lote_unico}')

wb.save(r'C:\Users\zarag\Downloads\REVISAR\lotes_nuevos_test.xlsx')
print('\n✓ Archivo creado: lotes_nuevos_test.xlsx con 5 lotes únicos')

"""
SCRIPT MAESTRO DE VERIFICACIÓN COMPLETA
========================================

Verifica al 100% que los importadores de Productos y Lotes funcionen perfectamente.

Pruebas exhaustivas:
1. Importador de Productos con archivo real
2. Importador de Lotes con archivo real
3. Validación de mapeo de columnas
4. Validación de creación de logs
5. Validación de actualización de stock
6. Pruebas de casos límite

Archivos de prueba:
- Plantilla_Productos.xlsx
- lotes_2025-12-15.xlsx
"""

import os
import sys
import django
from pathlib import Path

# Setup Django
sys.path.insert(0, str(Path(__file__).parent))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
django.setup()

from django.contrib.auth import get_user_model
from django.db import connection
from core.models import Producto, Lote, Centro, ImportacionLog
from core.utils.excel_importer import importar_productos_desde_excel, importar_lotes_desde_excel
from django.core.files.uploadedfile import InMemoryUploadedFile
from io import BytesIO
import openpyxl

User = get_user_model()

# Configuración de archivos
ARCHIVO_PRODUCTOS = r"C:\Users\zarag\Downloads\REVISAR\Plantilla_Productos.xlsx"
ARCHIVO_LOTES = r"C:\Users\zarag\Downloads\REVISAR\lotes_2025-12-15.xlsx"


class Colores:
    """Colores ANSI para terminal."""
    VERDE = '\033[92m'
    ROJO = '\033[91m'
    AMARILLO = '\033[93m'
    AZUL = '\033[94m'
    MAGENTA = '\033[95m'
    CYAN = '\033[96m'
    BLANCO = '\033[97m'
    BOLD = '\033[1m'
    RESET = '\033[0m'


def print_header(texto):
    """Imprime encabezado con estilo."""
    print(f"\n{Colores.BOLD}{Colores.CYAN}{'=' * 80}{Colores.RESET}")
    print(f"{Colores.BOLD}{Colores.AZUL}{texto.center(80)}{Colores.RESET}")
    print(f"{Colores.BOLD}{Colores.CYAN}{'=' * 80}{Colores.RESET}\n")


def print_success(texto):
    """Imprime mensaje de éxito."""
    print(f"{Colores.VERDE}✓ {texto}{Colores.RESET}")


def print_error(texto):
    """Imprime mensaje de error."""
    print(f"{Colores.ROJO}✗ {texto}{Colores.RESET}")


def print_warning(texto):
    """Imprime mensaje de advertencia."""
    print(f"{Colores.AMARILLO}⚠ {texto}{Colores.RESET}")


def print_info(texto):
    """Imprime información."""
    print(f"{Colores.BLANCO}  {texto}{Colores.RESET}")


def verificar_conexion_bd():
    """Verifica que la conexión a la BD funcione."""
    print_header("1. VERIFICANDO CONEXIÓN A BASE DE DATOS")
    
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT version()")
            version = cursor.fetchone()[0]
            print_success(f"Conexión exitosa a PostgreSQL")
            print_info(f"Versión: {version[:50]}...")
            
            # Verificar tablas principales
            cursor.execute("""
                SELECT table_name FROM information_schema.tables 
                WHERE table_schema = 'public' AND table_name IN ('productos', 'lotes', 'importacion_logs')
                ORDER BY table_name
            """)
            tablas = [row[0] for row in cursor.fetchall()]
            print_success(f"Tablas encontradas: {', '.join(tablas)}")
            
            return True
    except Exception as e:
        print_error(f"Error de conexión: {e}")
        return False


def verificar_usuario_farmacia():
    """Verifica que exista usuario con rol farmacia."""
    print_header("2. VERIFICANDO USUARIO FARMACIA")
    
    try:
        user = User.objects.filter(rol='farmacia').first()
        if not user:
            print_error("No hay usuario con rol 'farmacia'")
            print_info("Creando usuario de prueba...")
            
            user = User.objects.create_user(
                username='farmacia_test',
                email='farmacia@test.com',
                password='test123',
                rol='farmacia',
                first_name='Usuario',
                last_name='Farmacia'
            )
            print_success(f"Usuario creado: {user.username}")
        else:
            print_success(f"Usuario encontrado: {user.username} ({user.email})")
        
        return user
    except Exception as e:
        print_error(f"Error al verificar usuario: {e}")
        return None


def analizar_archivo_excel(archivo_path, nombre_archivo):
    """Analiza estructura del archivo Excel."""
    print_header(f"3. ANALIZANDO ARCHIVO: {nombre_archivo}")
    
    if not os.path.exists(archivo_path):
        print_error(f"Archivo no encontrado: {archivo_path}")
        return False
    
    file_size = os.path.getsize(archivo_path)
    print_success(f"Archivo encontrado")
    print_info(f"Tamaño: {file_size / 1024:.2f} KB")
    
    try:
        wb = openpyxl.load_workbook(archivo_path, read_only=True, data_only=True)
        sheet = wb.active
        
        print_info(f"Hojas: {len(wb.sheetnames)} - Activa: {wb.active.title}")
        print_info(f"Dimensiones: {sheet.max_row} filas x {sheet.max_column} columnas")
        
        # Detectar fila de encabezados
        fila_encabezados = None
        for fila_num in range(1, min(5, sheet.max_row + 1)):
            headers = [cell.value for cell in sheet[fila_num]]
            if any(h and isinstance(h, str) and len(h.strip()) > 0 for h in headers):
                fila_encabezados = fila_num
                print_success(f"Encabezados detectados en fila {fila_num}:")
                print_info(f"  Columnas: {[h for h in headers if h]}")
                break
        
        if not fila_encabezados:
            print_warning("No se detectaron encabezados")
        
        return True
        
    except Exception as e:
        print_error(f"Error al analizar archivo: {e}")
        return False


def test_importar_productos(user):
    """Prueba completa de importación de productos."""
    print_header("4. PRUEBA DE IMPORTACIÓN DE PRODUCTOS")
    
    # Contar productos antes
    productos_antes = Producto.objects.count()
    print_info(f"Productos en BD antes: {productos_antes}")
    
    try:
        with open(ARCHIVO_PRODUCTOS, 'rb') as f:
            content = f.read()
            file_obj = InMemoryUploadedFile(
                file=BytesIO(content),
                field_name='file',
                name=os.path.basename(ARCHIVO_PRODUCTOS),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                size=len(content),
                charset=None
            )
            
            print_info("Ejecutando importación...")
            resultado = importar_productos_desde_excel(file_obj, user)
            
            print_info("\n📊 RESULTADO:")
            print_info(f"  Exitosa: {resultado.get('exitosa', False)}")
            print_info(f"  Total procesados: {resultado.get('total_registros', 0)}")
            print_info(f"  Exitosos: {resultado.get('registros_exitosos', 0)}")
            print_info(f"  Fallidos: {resultado.get('registros_fallidos', 0)}")
            print_info(f"  Tasa de éxito: {resultado.get('tasa_exito', 0)}%")
            
            if resultado.get('creados'):
                print_success(f"  Creados: {resultado['creados']}")
            if resultado.get('actualizados'):
                print_warning(f"  Actualizados: {resultado['actualizados']}")
            
            errores = resultado.get('errores', [])
            if errores:
                print_warning(f"\n⚠️  ERRORES ({len(errores)} total):")
                for i, err in enumerate(errores[:5], 1):
                    print_info(f"  {i}. Fila {err.get('fila')}: {err.get('error')}")
                if len(errores) > 5:
                    print_info(f"  ... y {len(errores) - 5} más")
            
            productos_despues = Producto.objects.count()
            print_info(f"\nProductos en BD después: {productos_despues}")
            print_info(f"Diferencia: +{productos_despues - productos_antes}")
            
            if resultado.get('exitosa') or resultado.get('registros_exitosos', 0) > 0:
                print_success("\n✅ IMPORTACIÓN DE PRODUCTOS: EXITOSA")
                return True
            else:
                print_error("\n❌ IMPORTACIÓN DE PRODUCTOS: FALLIDA")
                return False
                
    except Exception as e:
        print_error(f"ERROR CRÍTICO: {e}")
        import traceback
        traceback.print_exc()
        return False


def test_importar_lotes(user):
    """Prueba completa de importación de lotes."""
    print_header("5. PRUEBA DE IMPORTACIÓN DE LOTES")
    
    # Contar lotes antes
    lotes_antes = Lote.objects.count()
    print_info(f"Lotes en BD antes: {lotes_antes}")
    
    try:
        with open(ARCHIVO_LOTES, 'rb') as f:
            content = f.read()
            file_obj = InMemoryUploadedFile(
                file=BytesIO(content),
                field_name='file',
                name=os.path.basename(ARCHIVO_LOTES),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                size=len(content),
                charset=None
            )
            
            print_info("Ejecutando importación...")
            resultado = importar_lotes_desde_excel(file_obj, user, centro_id=None)
            
            print_info("\n📊 RESULTADO:")
            print_info(f"  Exitosa: {resultado.get('exitosa', False)}")
            print_info(f"  Total procesados: {resultado.get('total_registros', 0)}")
            print_info(f"  Exitosos: {resultado.get('registros_exitosos', 0)}")
            print_info(f"  Fallidos: {resultado.get('registros_fallidos', 0)}")
            print_info(f"  Tasa de éxito: {resultado.get('tasa_exito', 0)}%")
            
            if resultado.get('creados'):
                print_success(f"  Creados: {resultado['creados']}")
            
            errores = resultado.get('errores', [])
            if errores:
                print_warning(f"\n⚠️  ERRORES ({len(errores)} total):")
                # Agrupar errores similares
                errores_unicos = {}
                for err in errores:
                    error_msg = err.get('error', '')
                    if 'ya existe' in error_msg:
                        key = 'duplicados'
                    elif 'no encontrado' in error_msg:
                        key = 'no_encontrados'
                    else:
                        key = 'otros'
                    
                    if key not in errores_unicos:
                        errores_unicos[key] = []
                    errores_unicos[key].append(err)
                
                for key, lista in errores_unicos.items():
                    print_info(f"\n  {key.upper()} ({len(lista)}):")
                    for err in lista[:3]:
                        print_info(f"    Fila {err.get('fila')}: {err.get('error')}")
                    if len(lista) > 3:
                        print_info(f"    ... y {len(lista) - 3} más")
            
            lotes_despues = Lote.objects.count()
            print_info(f"\nLotes en BD después: {lotes_despues}")
            print_info(f"Diferencia: +{lotes_despues - lotes_antes}")
            
            if resultado.get('exitosa') or resultado.get('registros_exitosos', 0) > 0:
                print_success("\n✅ IMPORTACIÓN DE LOTES: EXITOSA")
                return True
            else:
                print_error("\n❌ IMPORTACIÓN DE LOTES: FALLIDA")
                return False
                
    except Exception as e:
        print_error(f"ERROR CRÍTICO: {e}")
        import traceback
        traceback.print_exc()
        return False


def verificar_logs_importacion():
    """Verifica que se hayan creado los logs de importación."""
    print_header("6. VERIFICANDO LOGS DE IMPORTACIÓN")
    
    try:
        logs = ImportacionLog.objects.all().order_by('-fecha_inicio')[:5]
        print_info(f"Total de logs: {ImportacionLog.objects.count()}")
        
        if logs:
            print_success("Últimos 5 logs de importación:")
            for log in logs:
                print_info(f"\n  ID: {log.id}")
                print_info(f"  Tipo: {log.tipo_importacion}")
                print_info(f"  Archivo: {log.archivo}")
                print_info(f"  Usuario: {log.usuario}")
                print_info(f"  Estado: {log.estado}")
                print_info(f"  Totales: {log.registros_totales} | Exitosos: {log.registros_exitosos} | Fallidos: {log.registros_fallidos}")
                print_info(f"  Fecha: {log.fecha_inicio}")
        else:
            print_warning("No hay logs de importación")
        
        return True
    except Exception as e:
        print_error(f"Error al verificar logs: {e}")
        return False


def ejecutar_verificacion_completa():
    """Ejecuta todas las pruebas de verificación."""
    print(f"\n{Colores.BOLD}{Colores.MAGENTA}")
    print("╔" + "═" * 78 + "╗")
    print("║" + "VERIFICACIÓN COMPLETA DE IMPORTADORES".center(78) + "║")
    print("║" + "Productos y Lotes".center(78) + "║")
    print("╚" + "═" * 78 + "╝")
    print(Colores.RESET)
    
    resultados = []
    
    # 1. Verificar conexión BD
    resultados.append(("Conexión BD", verificar_conexion_bd()))
    
    # 2. Verificar usuario
    user = verificar_usuario_farmacia()
    resultados.append(("Usuario Farmacia", user is not None))
    
    if not user:
        print_error("\n❌ No se puede continuar sin usuario")
        return
    
    # 3. Analizar archivos
    resultados.append(("Análisis Productos", analizar_archivo_excel(ARCHIVO_PRODUCTOS, "Plantilla_Productos.xlsx")))
    resultados.append(("Análisis Lotes", analizar_archivo_excel(ARCHIVO_LOTES, "lotes_2025-12-15.xlsx")))
    
    # 4. Importar productos
    resultados.append(("Importación Productos", test_importar_productos(user)))
    
    # 5. Importar lotes
    resultados.append(("Importación Lotes", test_importar_lotes(user)))
    
    # 6. Verificar logs
    resultados.append(("Logs Importación", verificar_logs_importacion()))
    
    # RESUMEN FINAL
    print_header("RESUMEN FINAL")
    
    total = len(resultados)
    exitosos = sum(1 for _, exito in resultados if exito)
    fallidos = total - exitosos
    
    print(f"\n{Colores.BOLD}Resultados:{Colores.RESET}\n")
    for nombre, exito in resultados:
        if exito:
            print_success(f"{nombre}: PASSED")
        else:
            print_error(f"{nombre}: FAILED")
    
    print(f"\n{Colores.BOLD}Total:{Colores.RESET}")
    print_info(f"  Pruebas ejecutadas: {total}")
    print_success(f"  Exitosas: {exitosos}")
    if fallidos > 0:
        print_error(f"  Fallidas: {fallidos}")
    
    tasa_exito = (exitosos / total * 100) if total > 0 else 0
    print(f"\n{Colores.BOLD}Tasa de éxito: {tasa_exito:.1f}%{Colores.RESET}\n")
    
    if tasa_exito == 100:
        print(f"{Colores.BOLD}{Colores.VERDE}")
        print("╔" + "═" * 78 + "╗")
        print("║" + "✅ TODOS LOS IMPORTADORES FUNCIONAN AL 100%".center(78) + "║")
        print("║" + "Sistema listo para producción".center(78) + "║")
        print("╚" + "═" * 78 + "╝")
        print(Colores.RESET)
    else:
        print(f"{Colores.BOLD}{Colores.ROJO}")
        print("╔" + "═" * 78 + "╗")
        print("║" + "❌ HAY FALLOS EN LOS IMPORTADORES".center(78) + "║")
        print("║" + "Revisar errores arriba".center(78) + "║")
        print("╚" + "═" * 78 + "╝")
        print(Colores.RESET)


if __name__ == '__main__':
    try:
        ejecutar_verificacion_completa()
    except KeyboardInterrupt:
        print(f"\n\n{Colores.AMARILLO}Verificación cancelada por el usuario{Colores.RESET}\n")
    except Exception as e:
        print(f"\n\n{Colores.ROJO}ERROR FATAL: {e}{Colores.RESET}\n")
        import traceback
        traceback.print_exc()

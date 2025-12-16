"""
Script para probar la importación de productos desde Excel.
Uso: python test_importacion_productos.py <ruta_archivo_excel>
"""
import os
import sys
import django

# Configurar Django
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
django.setup()

from core.utils.excel_importer import importar_productos_desde_excel
from core.models import User
import openpyxl


def analizar_plantilla(ruta_archivo):
    """Analiza la estructura de la plantilla Excel."""
    print(f"\n{'='*80}")
    print(f"ANÁLISIS DE PLANTILLA: {ruta_archivo}")
    print(f"{'='*80}\n")
    
    try:
        wb = openpyxl.load_workbook(ruta_archivo, read_only=True, data_only=True)
        ws = wb.active
        
        # Encabezados
        headers = [cell.value for cell in ws[1]]
        print(f"📋 COLUMNAS DETECTADAS ({len(headers)}):")
        for i, h in enumerate(headers, 1):
            print(f"  {i:2}. {h}")
        
        # Filas de datos
        total_filas = ws.max_row - 1
        print(f"\n📊 FILAS DE DATOS: {total_filas}")
        
        if total_filas > 0:
            print(f"\n🔍 PRIMERAS 3 FILAS DE EJEMPLO:")
            for fila_num in range(2, min(5, ws.max_row + 1)):
                fila = list(ws[fila_num])
                print(f"\n  Fila {fila_num}:")
                for i, cell in enumerate(fila):
                    if i < len(headers):
                        valor = cell.value if cell.value else "(vacío)"
                        print(f"    {headers[i]}: {valor}")
        
        return True
    except Exception as e:
        print(f"❌ Error al analizar plantilla: {e}")
        return False


def probar_importacion(ruta_archivo):
    """Prueba la importación de productos."""
    print(f"\n{'='*80}")
    print(f"PRUEBA DE IMPORTACIÓN")
    print(f"{'='*80}\n")
    
    # Usar usuario administrador
    try:
        usuario = User.objects.filter(is_superuser=True).first()
        if not usuario:
            usuario = User.objects.filter(rol='admin').first()
        
        if not usuario:
            print("❌ No se encontró usuario administrador para la prueba")
            return False
        
        print(f"👤 Usuario: {usuario.username} ({usuario.get_rol_display()})")
    except Exception as e:
        print(f"❌ Error al obtener usuario: {e}")
        return False
    
    # Usar la ruta del archivo directamente - openpyxl puede abrirlo
    try:
        print(f"\n⏳ Importando productos...")
        resultado = importar_productos_desde_excel(ruta_archivo, usuario)
        
        print(f"\n{'='*80}")
        print(f"RESULTADO DE LA IMPORTACIÓN")
        print(f"{'='*80}\n")
        
        print(f"✅ Exitosa: {'SÍ' if resultado['exitosa'] else 'NO'}")
        print(f"📊 Total registros: {resultado['total_registros']}")
        print(f"✓  Exitosos: {resultado['registros_exitosos']}")
        print(f"✗  Fallidos: {resultado['registros_fallidos']}")
        print(f"📈 Tasa de éxito: {resultado['tasa_exito']}%")
        
        if 'creados' in resultado:
            print(f"🆕 Creados: {resultado['creados']}")
        if 'actualizados' in resultado:
            print(f"🔄 Actualizados: {resultado['actualizados']}")
        
        if resultado['errores']:
            print(f"\n⚠️  ERRORES ({len(resultado['errores'])}):")
            for i, error in enumerate(resultado['errores'][:10], 1):
                print(f"\n  Error #{i}:")
                print(f"    Fila: {error.get('fila', 'N/A')}")
                print(f"    Campo: {error.get('campo', 'N/A')}")
                print(f"    Detalle: {error.get('error', 'N/A')}")
            
            if len(resultado['errores']) > 10:
                print(f"\n  ... y {len(resultado['errores']) - 10} errores más")
        
        return resultado['exitosa'] or resultado['registros_exitosos'] > 0
        
    except Exception as e:
        print(f"\n❌ Error durante la importación: {e}")
        import traceback
        traceback.print_exc()
        return False


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Uso: python test_importacion_productos.py <ruta_archivo_excel>")
        print("\nEjemplo:")
        print("  python test_importacion_productos.py C:\\Users\\zarag\\Downloads\\REVISAR\\Plantilla_Productos.xlsx")
        sys.exit(1)
    
    ruta_archivo = sys.argv[1]
    
    if not os.path.exists(ruta_archivo):
        print(f"❌ Archivo no encontrado: {ruta_archivo}")
        sys.exit(1)
    
    print("\n" + "="*80)
    print(" PRUEBA DE IMPORTACIÓN DE PRODUCTOS DESDE EXCEL")
    print("="*80)
    
    # Paso 1: Analizar plantilla
    if not analizar_plantilla(ruta_archivo):
        sys.exit(1)
    
    # Paso 2: Probar importación
    if probar_importacion(ruta_archivo):
        print(f"\n✅ PRUEBA EXITOSA")
        sys.exit(0)
    else:
        print(f"\n❌ PRUEBA FALLIDA")
        sys.exit(1)

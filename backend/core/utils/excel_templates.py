"""
Generador de plantillas Excel actualizadas para carga masiva.
Basado en el esquema real de la base de datos (public.productos, public.lotes, public.usuarios).

HALLAZGO #6 - Manejo de Memoria:
- Los archivos Excel se generan completamente en memoria (sin archivos temporales)
- Para volúmenes actuales (<1000 filas de ejemplo), el consumo de RAM es aceptable (~2-5 MB por plantilla)
- Si se requieren plantillas con miles de filas o validaciones complejas (listas desplegables),
  considerar usar openpyxl en modo write_only para reducir consumo de memoria
- Las plantillas actuales no dejan archivos temporales en el servidor
- El workbook se serializa directamente al HttpResponse y se libera automáticamente

Límites recomendados por plantilla:
- Filas de ejemplo: < 100
- Hojas: < 5
- Tamaño máximo del archivo generado: < 10 MB
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.http import HttpResponse
from datetime import date


def aplicar_estilos_header(ws, num_columnas):
    """Aplica estilos consistentes a los encabezados."""
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="9F2241", end_color="9F2241", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col_num in range(1, num_columnas + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border_style
    
    ws.freeze_panes = 'A2'


def generar_plantilla_productos():
    """
    Genera plantilla Excel para carga masiva de productos.
    Basado en tabla: public.productos
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Productos"

    # Columnas EXACTAS que coinciden con la plantilla descargable
    # IMPORTANTE: Estos nombres deben coincidir con los sinónimos en excel_importer.py
    headers = [
        "Clave",
        "Nombre",
        "Nombre Comercial",
        "Unidad",
        "Stock Minimo",
        "Categoria",
        "Sustancia Activa",
        "Presentacion",
        "Concentracion",
        "Via Admin",
        "Requiere Receta",
        "Controlado",
        "Estado"
    ]

    # Anchos personalizados por columna para mejor legibilidad
    column_widths = [15, 45, 30, 18, 15, 20, 25, 25, 20, 20, 18, 15, 12]
    
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)
        if col_num <= len(column_widths):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = column_widths[col_num - 1]
    
    aplicar_estilos_header(ws, len(headers))

    # ============================================================
    # FILAS DE EJEMPLO - ELIMINAR ANTES DE USAR CON DATOS REALES
    # Texto gris itálico para que sea obvio
    # ============================================================
    ejemplos = [
        ["PRUEBA001", "[EJEMPLO] Paracetamol 500mg - ELIMINAR", "Tylenol", "CAJA", 50, "medicamento",
         "Paracetamol", "Tableta", "500 mg", "oral", "No", "No", "Activo"],
        ["PRUEBA002", "[EJEMPLO] Ibuprofeno 400mg - ELIMINAR", "Advil", "FRASCO", 30, "medicamento",
         "Ibuprofeno", "Cápsula", "400 mg", "oral", "No", "No", "Activo"],
        ["PRUEBA003", "[EJEMPLO] Jeringa 10ml - ELIMINAR", "", "PIEZA", 100, "material_curacion",
         "", "", "", "", "No", "No", "Activo"],
    ]
    
    for ejemplo in ejemplos:
        ws.append(ejemplo)
    
    # Estilo para filas de ejemplo (gris claro, itálica)
    example_font = Font(italic=True, color="888888")
    for row_num in range(2, 2 + len(ejemplos)):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = example_font

    # Instrucciones en hoja separada
    ws_inst = wb.create_sheet("INSTRUCCIONES")
    instrucciones = [
        ["╔════════════════════════════════════════════════════════════════════╗"],
        ["║    INSTRUCCIONES PARA IMPORTACIÓN DE PRODUCTOS                    ║"],
        ["╚════════════════════════════════════════════════════════════════════╝"],
        [""],
        ["⚠️  IMPORTANTE: Las filas grises con [EJEMPLO] en la hoja 'Productos' son de muestra."],
        ["    ELIMÍNELAS antes de cargar sus datos reales."],
        [""],
        ["────────────────────────────────────────────────────────────────────────"],
        ["COLUMNAS REQUERIDAS (obligatorias):"],
        ["────────────────────────────────────────────────────────────────────────"],
        ["• Clave      - Código único del producto (ej: 001, MED001, ABC123)"],
        ["• Nombre     - Nombre completo del producto (nombre genérico)"],
        [""],
        ["────────────────────────────────────────────────────────────────────────"],
        ["COLUMNAS OPCIONALES:"],
        ["────────────────────────────────────────────────────────────────────────"],
        ["• Nombre Comercial - Nombre de marca del producto (ej: Tylenol, Aspirina)"],
        ["• Unidad         - CAJA, PIEZA, FRASCO, SOBRE, TABLETA, etc. (default: PIEZA)"],
        ["• Stock Minimo   - Cantidad mínima para alertas (default: 10)"],
        ["• Categoria      - medicamento, material_curacion, insumo (default: medicamento)"],
        ["• Sustancia Activa - Principio activo del medicamento"],
        ["• Presentacion   - Forma farmacéutica (tableta, cápsula, jarabe, etc.)"],
        ["• Concentracion  - Dosis (ej: 500 mg, 10 ml)"],
        ["• Via Admin      - oral, intravenosa, tópica, etc."],
        ["• Requiere Receta - Sí / No (default: No)"],
        ["• Controlado     - Sí / No (default: No)"],
        ["• Estado         - Activo / Inactivo (default: Activo)"],
        [""],
        ["────────────────────────────────────────────────────────────────────────"],
        ["NOTAS:"],
        ["────────────────────────────────────────────────────────────────────────"],
        ["• Si la CLAVE ya existe, el producto se ACTUALIZARÁ con los nuevos datos."],
        ["• Si la CLAVE no existe, se CREARÁ un nuevo producto."],
        ["• La Unidad acepta texto libre: 'CAJA CON 7 OVULOS', 'GOTERO CON 15 ML'"],
        ["• Valores booleanos: Sí/Si/Yes/True/1/Activo o No/False/0/Inactivo"],
        ["• Máximo 5000 productos por archivo."],
        ["• Tamaño máximo de archivo: 10 MB."],
        ["• Formatos aceptados: .xlsx, .xls"],
        [""],
    ]
    
    for row_data in instrucciones:
        ws_inst.append(row_data)
    
    ws_inst.column_dimensions['A'].width = 80

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=Plantilla_Productos.xlsx'
    wb.save(response)
    wb.close()  # HALLAZGO #6: Liberar recursos explícitamente
    return response


def generar_plantilla_lotes(centro=None):
    """
    Genera plantilla Excel para carga masiva de lotes/inventario inicial.
    Basado en tabla: public.lotes
    
    Versión de plantilla: 2.1.0 (Febrero 2026)
    
    Columnas OBLIGATORIAS:
    - Clave Producto*: Código único del producto en el catálogo
    - Nombre Producto*: Nombre del producto (debe coincidir con la clave)
    - Presentación*: Presentación del producto (DEBE COINCIDIR EXACTAMENTE)
    - Número Lote*: Identificador único del lote
    - Fecha Caducidad*: Fecha de vencimiento (YYYY-MM-DD)
    - Cantidad Inicial*: Unidades recibidas/surtidas
    
    Columnas OPCIONALES:
    - Cantidad Contrato Lote: Cantidad según contrato para este lote específico
    - Cantidad Contrato Global: Total contratado para toda la clave del producto
    - Fecha Recepción: Fecha de recepción del lote
    - Precio Unitario: Precio por unidad (default 0)
    - Número Contrato: Referencia del contrato
    - Marca: Laboratorio o fabricante
    - Activo: Estado del lote (default: Activo)
    
    NOTA: La ubicación se asigna automáticamente como "Almacén Central"
    y el centro queda NULL (representa Farmacia Central).
    """
    from datetime import date, timedelta
    
    PLANTILLA_VERSION = "2.1.0"
    FECHA_GENERACION = date.today().strftime('%Y-%m-%d')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lotes Inventario"

    headers = [
        "Clave Producto",
        "Nombre Producto",
        "Presentación",
        "Número Lote",
        "Fecha Recepción",
        "Fecha Caducidad",
        "Cantidad Inicial",
        "Cantidad Contrato Lote",
        "Cantidad Contrato Global",
        "Precio Unitario",
        "Número Contrato",
        "Marca",
        "Activo"
    ]

    column_widths = [15, 40, 25, 20, 18, 18, 16, 20, 22, 15, 18, 35, 12]
    
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)
        if col_num <= len(column_widths):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = column_widths[col_num - 1]
    
    aplicar_estilos_header(ws, len(headers))

    # ============================================================
    # FILAS DE EJEMPLO - ELIMINAR ANTES DE USAR CON DATOS REALES
    # ISS-INV-003: Ejemplos con contrato global
    # ============================================================
    fecha_cad = (date.today() + timedelta(days=365)).strftime('%Y-%m-%d')
    fecha_cad2 = (date.today() + timedelta(days=730)).strftime('%Y-%m-%d')
    fecha_fab = date.today().strftime('%Y-%m-%d')
    
    ejemplos = [
        # EJEMPLO CONTRATO GLOBAL: Total contratado 1000, llega en 3 lotes
        # Lote 1: 300 unidades, caducidad 1 año
        ["615", "[EJEMPLO] PARACETAMOL 500MG - ELIMINAR", "CAJA CON 20 TABLETAS", "LOTE-2026-001",
         fecha_fab, fecha_cad, 300, "", 1000, 15.50, "CONT-2026-PAR-001",
         "[EJEMPLO] Laboratorio A - ELIMINAR", "Activo"],
        # Lote 2: 250 unidades, misma caducidad
        ["615", "[EJEMPLO] PARACETAMOL 500MG - ELIMINAR", "CAJA CON 20 TABLETAS", "LOTE-2026-002",
         fecha_fab, fecha_cad, 250, "", 1000, 15.50, "CONT-2026-PAR-001",
         "[EJEMPLO] Laboratorio A - ELIMINAR", "Activo"],
        # Lote 3: 200 unidades, caducidad 2 años
        ["615", "[EJEMPLO] PARACETAMOL 500MG - ELIMINAR", "CAJA CON 20 TABLETAS", "LOTE-2026-003",
         fecha_fab, fecha_cad2, 200, "", 1000, 15.50, "CONT-2026-PAR-001",
         "[EJEMPLO] Laboratorio A - ELIMINAR", "Activo"],
        # El sistema calculará: Pendiente Global = 1000 - (300+250+200) = 250
        
        # Ejemplo con presentación diferente (CAJA CON 14 TABLETAS)
        ["702.2", "[EJEMPLO] TRIMETOPRIMA/SULFAMETOXAZOL - ELIMINAR", "CAJA CON 14 TABLETAS", "LOTE-TRI-001",
         fecha_fab, fecha_cad, 80, 100, "", 18.75, "CONT-2026-TRI-001",
         "[EJEMPLO] Farmacéutica B - ELIMINAR", "Activo"],
        # ⚠️ IMPORTANTE: La presentación DEBE coincidir con el catálogo de productos
    ]
    
    for ejemplo in ejemplos:
        ws.append(ejemplo)
    
    # Estilo para filas de ejemplo (gris claro, itálica - sin fondo de color)
    example_font = Font(italic=True, color="888888")
    for row_num in range(2, 2 + len(ejemplos)):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = example_font

    # Instrucciones
    ws_inst = wb.create_sheet("INSTRUCCIONES")
    instrucciones = [
        ["╔════════════════════════════════════════════════════════════════════╗"],
        ["║    PLANTILLA DE IMPORTACIÓN DE LOTES                              ║"],
        ["╚════════════════════════════════════════════════════════════════════╝"],
        [""],
        [f"📋 Versión: {PLANTILLA_VERSION}"],
        [f"📅 Generada: {FECHA_GENERACION}"],
        [""],
        ["════════════════════════════════════════════════════════════════════════"],
        ["⚠️  IMPORTANTE: ELIMINE LAS FILAS DE EJEMPLO"],
        ["════════════════════════════════════════════════════════════════════════"],
        ["Las filas marcadas con [EJEMPLO] y texto gris son de muestra."],
        ["ELIMÍNELAS antes de cargar sus datos reales."],
        [""],
        ["════════════════════════════════════════════════════════════════════════"],
        ["⚠️  VERIFICACIÓN DE TRIPLE CAMPO: CLAVE + NOMBRE + PRESENTACIÓN"],
        ["════════════════════════════════════════════════════════════════════════"],
        ["El sistema verifica que los TRES campos coincidan con el producto en"],
        ["la base de datos. Si hay discrepancia en cualquiera, se reportará error."],
        [""],
        ["⚠️ EJEMPLO CRÍTICO: Productos con misma clave base:"],
        ["  - Clave 702: TRIMETOPRIMA/SULFAMETOXAZOL, Presentación: CAJA CON 20 TABLETAS"],
        ["  - Clave 702.2: TRIMETOPRIMA/SULFAMETOXAZOL, Presentación: CAJA CON 14 TABLETAS"],
        [""],
        ["Si importa con Clave=702 pero Presentación='CAJA CON 14 TABLETAS', el"],
        ["sistema buscará el producto correcto (702.2) automáticamente."],
        [""],
        ["────────────────────────────────────────────────────────────────────────"],
        ["COLUMNAS REQUERIDAS (obligatorias):"],
        ["────────────────────────────────────────────────────────────────────────"],
        ["• Clave Producto* - OBLIGATORIA: Clave única del producto en el sistema"],
        ["• Nombre Producto* - OBLIGATORIO: Debe coincidir con la clave"],
        ["• Presentación*   - OBLIGATORIA: Ej: CAJA CON 14 TABLETAS"],
        ["                    ⚠️ DEBE COINCIDIR EXACTAMENTE con el catálogo"],
        ["• Número Lote*    - Identificador único del lote"],
        ["• Fecha Caducidad* - Formato: YYYY-MM-DD (ej: 2026-12-31)"],
        ["• Cantidad Inicial* - Cantidad de unidades RECIBIDAS/SURTIDAS"],
        [""],
        ["────────────────────────────────────────────────────────────────────────"],
        ["COLUMNAS OPCIONALES:"],
        ["────────────────────────────────────────────────────────────────────────"],
        ["• Cantidad Contrato Lote - Cantidad según contrato para ESTE lote"],
        ["                     Si llegan 80 de 100 contratados: Inicial=80, Contrato Lote=100"],
        [""],
        ["• Cantidad Contrato Global - Total contratado para TODA LA CLAVE de producto"],
        ["                     ⚠️ IMPORTANTE: Debe ser EL MISMO VALOR en todas las filas"],
        ["                     del mismo producto + número de contrato."],
        ["                     "],
        ["                     Ejemplo: Si contratas 1000 unidades de Paracetamol (clave 615)"],
        ["                     con el contrato CONT-2026-001, pon 1000 en TODAS las filas de"],
        ["                     ese producto y contrato, aunque lleguen en múltiples lotes."],
        ["                     "],
        ["                     El sistema calculará automáticamente cuánto falta recibir."],
        [""],
        ["• Fecha Recepción   - Formato: YYYY-MM-DD"],
        ["• Precio Unitario   - Precio por unidad (default: 0)"],
        ["• Número Contrato   - Referencia del contrato de adquisición (REQUERIDO si usas Contrato Global)"],
        ["• Marca             - Laboratorio o fabricante"],
        ["• Activo            - Activo/Inactivo (default: Activo)"],
        [""],
        ["════════════════════════════════════════════════════════════════════════"],
        ["📦 ENTREGAS PARCIALES Y REIMPORTACIÓN (ISS-INV-001):"],
        ["════════════════════════════════════════════════════════════════════════"],
        ["Si el contrato establece 100 unidades PARA ESTE LOTE pero solo llegan 80:"],
        ["  • Cantidad Inicial = 80 (lo que realmente llegó)"],
        ["  • Cantidad Contrato Lote = 100 (lo esperado del lote según contrato)"],
        ["  • El sistema calculará: Pendiente Lote = 100 - 80 = 20"],
        [""],
        ["CONTRATO GLOBAL POR CLAVE (ISS-INV-003):"],
        ["Si el contrato total para la clave 615 es de 1000 unidades:"],
        ["  • Cantidad Contrato Global = 1000 (en TODOS los lotes de esa clave+contrato)"],
        ["  • El Número de Contrato DEBE ser el mismo para agrupar los lotes"],
        ["  • El sistema sumará las cantidades iniciales de todos los lotes"],
        ["  • Pendiente Global = 1000 - (suma de lo recibido en todos los lotes)"],
        [""],
        ["⚠️ IMPORTANTE CONTRATO GLOBAL:"],
        ["  • Ponga el MISMO valor en todas las filas del mismo producto+contrato"],
        ["  • El sistema NO suma este valor, lo reemplaza con el más reciente"],
        ["  • Si pone 1000 en una fila y 500 en otra, el sistema usará 500"],
        [""],
        ["ALERTAS AUTOMÁTICAS:"],
        ["  • Total recibido < contrato: ⏳ FALTAN unidades (naranja)"],
        ["  • Total recibido > contrato: ⚠️ EXCESO detectado (rojo)"],
        ["  • Total recibido = contrato: ✅ COMPLETO (verde)"],
        [""],
        ["Cuando llegue el resto, puede REIMPORTAR el Excel con:"],
        ["  - Misma Clave, Lote, Contrato, Marca y Fecha Caducidad"],
        ["  - Cantidad Inicial = 20 (las nuevas unidades)"],
        ["  → El sistema SUMARÁ las cantidades al lote existente."],["  → La Cantidad Contrato original se PRESERVA (no se sobreescribe)."],
        [""],
        ["────────────────────────────────────────────────────────────────────────"],
        ["NOTAS:"],
        ["────────────────────────────────────────────────────────────────────────"],
        ["• Los lotes se asignan automáticamente al Almacén Central (FARMACIA)."],
        ["• El PRODUCTO debe existir antes de importar lotes."],
        ["• Verifique la CLAVE y NOMBRE del producto en el catálogo."],
        ["• Si el lote ya existe (mismo producto + número de lote), se reporta error."],
        ["• La cantidad_actual se inicializa igual a cantidad_inicial."],
        ["• El stock del producto se actualiza automáticamente."],
        ["• Fechas aceptadas: YYYY-MM-DD, DD/MM/YYYY, DD-MM-YYYY"],
        ["• Máximo 5000 lotes por archivo."],
        ["• Tamaño máximo de archivo: 10 MB."],
        [""],
    ]
    
    for row_data in instrucciones:
        ws_inst.append(row_data)
    
    ws_inst.column_dimensions['A'].width = 80

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=Plantilla_Lotes_Inventario.xlsx'
    wb.save(response)
    wb.close()  # HALLAZGO #6: Liberar recursos explícitamente
    return response


def generar_plantilla_usuarios():
    """
    Genera plantilla Excel para carga masiva de usuarios.
    Basado en tabla: public.usuarios (User model extendido)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Usuarios"

    headers = [
        "Username*\n(Único, 3-150 chars)",
        "Email*\n(Único)",
        "Nombre*\n(first_name)",
        "Apellidos*\n(last_name)",
        "Password*\n(Mín 8 chars)",
        "Rol*\n(Ver instrucciones)",
        "Centro ID\n(Número o vacío)",
        "Adscripción",
        "Teléfono",
        "Activo\n(Si/No, default: Si)"
    ]

    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20
    
    aplicar_estilos_header(ws, len(headers))

    # Ejemplos
    ws.append([
        "medico01",
        "medico@centro.gob.mx",
        "Juan",
        "Pérez García",
        "Temporal123!",
        "medico",
        1,
        "Área Médica",
        "5551234567",
        "Si"
    ])
    
    ws.append([
        "admin_farmacia",
        "farmacia@sistema.gob.mx",
        "María",
        "López Hernández",
        "Admin2024!",
        "farmacia",
        "",
        "Farmacia Central",
        "5559876543",
        "Si"
    ])

    # Instrucciones
    ws_inst = wb.create_sheet("Instrucciones")
    instrucciones = [
        ["INSTRUCCIONES DE LLENADO DE PLANTILLA - USUARIOS"],
        [""],
        ["Campos obligatorios marcados con asterisco (*)"],
        [""],
        ["USERNAME*:", "Nombre de usuario único (3-150 caracteres, sin espacios)"],
        ["EMAIL*:", "Correo electrónico único y válido"],
        ["NOMBRE*:", "Nombre(s) del usuario"],
        ["APELLIDOS*:", "Apellido(s) del usuario"],
        ["PASSWORD*:", "Contraseña temporal (mínimo 8 caracteres, se debe cambiar al primer ingreso)"],
        ["ROL*:", "Ver tabla de roles permitidos abajo"],
        ["CENTRO ID:", "ID numérico del centro penitenciario (dejar vacío para Farmacia Central)"],
        ["ADSCRIPCIÓN:", "Área de trabajo o departamento (opcional)"],
        ["TELÉFONO:", "Número telefónico (opcional)"],
        ["ACTIVO:", "Si/No - Estado del usuario (default: Si)"],
        [""],
        ["ROLES DISPONIBLES:"],
        [""],
        ["ROL", "DESCRIPCIÓN", "REQUIERE CENTRO"],
        ["admin", "Administrador del sistema", "No"],
        ["farmacia", "Personal de Farmacia Central", "No"],
        ["vista", "Solo lectura (Farmacia)", "No"],
        ["medico", "Médico del centro (crea requisiciones)", "Sí*"],
        ["administrador_centro", "Administrador del centro penitenciario", "Sí*"],
        ["director_centro", "Director del centro penitenciario", "Sí*"],
        ["centro", "Personal general del centro", "Sí"],
        [""],
        ["* Si el rol requiere centro, debe proporcionar un Centro ID válido"],
        [""],
        ["NOTAS IMPORTANTES:"],
        ["• El username y email deben ser únicos en todo el sistema"],
        ["• Los usuarios deben cambiar su contraseña en el primer acceso"],
        ["• Los permisos se asignan automáticamente según el rol"],
        ["• Para rol 'admin' use un email institucional verificado"],
        ["• Elimine esta hoja antes de subir el archivo"],
    ]
    
    for row_data in instrucciones:
        ws_inst.append(row_data)
    
    ws_inst.column_dimensions['A'].width = 25
    ws_inst.column_dimensions['B'].width = 70
    ws_inst.column_dimensions['C'].width = 20

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=Plantilla_Usuarios.xlsx'
    wb.save(response)
    wb.close()  # HALLAZGO #6: Liberar recursos explícitamente
    return response

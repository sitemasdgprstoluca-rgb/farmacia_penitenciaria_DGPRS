"""
Utilidades para exportación a Excel.

Este módulo proporciona funciones reutilizables para generar archivos Excel
con formato profesional, evitando duplicación de código en los ViewSets.

Uso:
    from core.utils.excel_export import ExcelExporter
    
    exporter = ExcelExporter('Productos')
    exporter.add_header(['Clave', 'Descripción', 'Stock'])
    exporter.add_row([prod.clave, prod.descripcion, prod.stock])
    return exporter.get_response('productos')
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.utils import timezone
from io import BytesIO
from typing import List, Any, Optional, Dict, Tuple
import logging

logger = logging.getLogger(__name__)

# Colores corporativos del sistema
COLORS = {
    'primary': '632842',      # Guinda institucional
    'primary_light': '8A3D5C',
    'secondary': '424242',
    'success': '28A745',
    'warning': 'FFC107',
    'danger': 'DC3545',
    'info': '17A2B8',
    'white': 'FFFFFF',
    'light_gray': 'F5F5F5',
    'yellow_light': 'FFF4E6',  # Para alertas de stock
    'red_light': 'FFEBEE',     # Para alertas críticas
}


class ExcelExporter:
    """
    Clase utilitaria para generar archivos Excel con formato profesional.
    
    Ejemplo de uso:
        exporter = ExcelExporter('Inventario')
        exporter.set_title('REPORTE DE INVENTARIO')
        exporter.set_subtitle(f'Generado el {timezone.now().strftime("%d/%m/%Y")}')
        exporter.add_header(['Clave', 'Descripción', 'Stock', 'Estado'])
        
        for producto in productos:
            row_style = 'warning' if producto.stock < producto.stock_minimo else None
            exporter.add_row(
                [producto.clave, producto.descripcion, producto.stock, producto.estado],
                style=row_style
            )
        
        return exporter.get_response('inventario')
    """
    
    def __init__(self, sheet_name: str = 'Datos'):
        """
        Inicializa el exportador.
        
        Args:
            sheet_name: Nombre de la hoja de Excel
        """
        self.workbook = openpyxl.Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = sheet_name[:31]  # Excel limita a 31 caracteres
        self.current_row = 1
        self.column_widths: Dict[int, int] = {}
        self._header_row = None
        
        # Estilos predefinidos
        self._init_styles()
    
    def _init_styles(self):
        """Inicializa estilos reutilizables."""
        # Bordes
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Fills
        self.header_fill = PatternFill(
            start_color=COLORS['primary'],
            end_color=COLORS['primary'],
            fill_type='solid'
        )
        self.warning_fill = PatternFill(
            start_color=COLORS['yellow_light'],
            end_color=COLORS['yellow_light'],
            fill_type='solid'
        )
        self.danger_fill = PatternFill(
            start_color=COLORS['red_light'],
            end_color=COLORS['red_light'],
            fill_type='solid'
        )
        self.alt_row_fill = PatternFill(
            start_color=COLORS['light_gray'],
            end_color=COLORS['light_gray'],
            fill_type='solid'
        )
        
        # Fonts
        self.header_font = Font(bold=True, color=COLORS['white'], size=11)
        self.title_font = Font(bold=True, size=14, color=COLORS['primary'])
        self.subtitle_font = Font(size=10, italic=True)
        self.normal_font = Font(size=10)
    
    def set_title(self, title: str, merge_cols: int = 0):
        """
        Agrega un título al reporte.
        
        Args:
            title: Texto del título
            merge_cols: Número de columnas a fusionar (0 = auto)
        """
        cell = self.sheet.cell(row=self.current_row, column=1)
        cell.value = title
        cell.font = self.title_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        if merge_cols > 1:
            self.sheet.merge_cells(
                start_row=self.current_row,
                start_column=1,
                end_row=self.current_row,
                end_column=merge_cols
            )
        
        self.sheet.row_dimensions[self.current_row].height = 25
        self.current_row += 1
    
    def set_subtitle(self, subtitle: str, merge_cols: int = 0):
        """
        Agrega un subtítulo al reporte.
        
        Args:
            subtitle: Texto del subtítulo
            merge_cols: Número de columnas a fusionar
        """
        cell = self.sheet.cell(row=self.current_row, column=1)
        cell.value = subtitle
        cell.font = self.subtitle_font
        cell.alignment = Alignment(horizontal='center')
        
        if merge_cols > 1:
            self.sheet.merge_cells(
                start_row=self.current_row,
                start_column=1,
                end_row=self.current_row,
                end_column=merge_cols
            )
        
        self.current_row += 1
    
    def add_blank_row(self):
        """Agrega una fila en blanco."""
        self.current_row += 1
    
    def add_header(self, headers: List[str], freeze: bool = True):
        """
        Agrega la fila de encabezados.
        
        Args:
            headers: Lista de nombres de columnas
            freeze: Si congela la fila de encabezados
        """
        self._header_row = self.current_row
        
        for col_idx, header in enumerate(headers, start=1):
            cell = self.sheet.cell(row=self.current_row, column=col_idx)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.thin_border
            
            # Actualizar ancho de columna
            self._update_column_width(col_idx, str(header))
        
        self.sheet.row_dimensions[self.current_row].height = 20
        
        if freeze:
            self.sheet.freeze_panes = f'A{self.current_row + 1}'
        
        self.current_row += 1
    
    def add_row(
        self,
        data: List[Any],
        style: Optional[str] = None,
        alternate: bool = False
    ):
        """
        Agrega una fila de datos.
        
        Args:
            data: Lista de valores para la fila
            style: Estilo de la fila ('warning', 'danger', None)
            alternate: Si aplica color alternado para filas pares
        """
        for col_idx, value in enumerate(data, start=1):
            cell = self.sheet.cell(row=self.current_row, column=col_idx)
            cell.value = value
            cell.font = self.normal_font
            cell.border = self.thin_border
            
            # Alineación según tipo de dato
            if isinstance(value, (int, float)):
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.alignment = Alignment(horizontal='left')
            
            # Aplicar estilo de fila
            if style == 'warning':
                cell.fill = self.warning_fill
            elif style == 'danger':
                cell.fill = self.danger_fill
            elif alternate and self.current_row % 2 == 0:
                cell.fill = self.alt_row_fill
            
            # Actualizar ancho de columna
            self._update_column_width(col_idx, str(value) if value is not None else '')
        
        self.current_row += 1
    
    def add_summary_row(self, label: str, value: Any, col_span: int = 1):
        """
        Agrega una fila de resumen (ej: totales).
        
        Args:
            label: Etiqueta del resumen
            value: Valor del resumen
            col_span: Columnas que ocupa la etiqueta
        """
        self.add_blank_row()
        
        # Etiqueta
        cell = self.sheet.cell(row=self.current_row, column=1)
        cell.value = label
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal='right')
        
        if col_span > 1:
            self.sheet.merge_cells(
                start_row=self.current_row,
                start_column=1,
                end_row=self.current_row,
                end_column=col_span
            )
        
        # Valor
        value_cell = self.sheet.cell(row=self.current_row, column=col_span + 1)
        value_cell.value = value
        value_cell.font = Font(bold=True, size=11, color=COLORS['primary'])
        value_cell.alignment = Alignment(horizontal='left')
        
        self.current_row += 1
    
    def _update_column_width(self, col_idx: int, value: str):
        """Actualiza el ancho de columna basado en el contenido."""
        current_width = self.column_widths.get(col_idx, 0)
        new_width = min(len(value) + 2, 50)  # Máximo 50 caracteres
        self.column_widths[col_idx] = max(current_width, new_width)
    
    def auto_fit_columns(self):
        """Ajusta automáticamente el ancho de todas las columnas."""
        for col_idx, width in self.column_widths.items():
            col_letter = get_column_letter(col_idx)
            self.sheet.column_dimensions[col_letter].width = width
    
    def set_column_width(self, col_idx: int, width: int):
        """
        Establece un ancho específico para una columna.
        
        Args:
            col_idx: Índice de columna (1-based)
            width: Ancho en caracteres
        """
        col_letter = get_column_letter(col_idx)
        self.sheet.column_dimensions[col_letter].width = width
    
    def get_workbook(self) -> openpyxl.Workbook:
        """Retorna el workbook para manipulación adicional."""
        return self.workbook
    
    def get_response(
        self,
        filename: str,
        include_timestamp: bool = True
    ) -> HttpResponse:
        """
        Genera la respuesta HTTP con el archivo Excel.
        
        Args:
            filename: Nombre base del archivo (sin extensión)
            include_timestamp: Si incluye fecha/hora en el nombre
            
        Returns:
            HttpResponse con el archivo Excel
        """
        # Auto-ajustar columnas antes de generar
        self.auto_fit_columns()
        
        # Generar nombre de archivo
        if include_timestamp:
            timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
            full_filename = f'{filename}_{timestamp}.xlsx'
        else:
            full_filename = f'{filename}.xlsx'
        
        # Crear respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename={full_filename}'
        
        # Guardar workbook en respuesta
        self.workbook.save(response)
        
        logger.info(f'Excel generado: {full_filename}')
        return response
    
    def get_buffer(self) -> BytesIO:
        """
        Retorna el archivo como buffer (para enviar por email, etc.).
        
        Returns:
            BytesIO con el contenido del archivo
        """
        self.auto_fit_columns()
        buffer = BytesIO()
        self.workbook.save(buffer)
        buffer.seek(0)
        return buffer


def export_queryset_to_excel(
    queryset,
    fields: List[Tuple[str, str]],
    filename: str,
    title: Optional[str] = None,
    row_style_callback=None
) -> HttpResponse:
    """
    Función utilitaria para exportar un queryset a Excel rápidamente.
    
    Args:
        queryset: QuerySet de Django
        fields: Lista de tuplas (nombre_campo, titulo_columna)
        filename: Nombre del archivo
        title: Título opcional del reporte
        row_style_callback: Función que recibe un objeto y retorna estilo ('warning', 'danger', None)
        
    Returns:
        HttpResponse con el archivo Excel
        
    Ejemplo:
        return export_queryset_to_excel(
            Producto.objects.filter(activo=True),
            [
                ('clave', 'Clave'),
                ('descripcion', 'Descripción'),
                ('stock_actual', 'Stock'),
            ],
            'productos',
            title='LISTADO DE PRODUCTOS',
            row_style_callback=lambda p: 'warning' if p.stock_actual < p.stock_minimo else None
        )
    """
    exporter = ExcelExporter(filename.replace('_', ' ').title())
    
    # Headers
    headers = [field[1] for field in fields]
    field_names = [field[0] for field in fields]
    
    # Título y subtítulo
    if title:
        exporter.set_title(title, merge_cols=len(headers))
        exporter.set_subtitle(
            f'Generado el {timezone.now().strftime("%d/%m/%Y %H:%M")}',
            merge_cols=len(headers)
        )
        exporter.add_blank_row()
    
    exporter.add_header(headers)
    
    # Datos
    for obj in queryset:
        row_data = []
        for field_name in field_names:
            # Soportar campos anidados con notación de punto
            value = obj
            for attr in field_name.split('.'):
                if value is None:
                    break
                if callable(getattr(value, attr, None)):
                    value = getattr(value, attr)()
                else:
                    value = getattr(value, attr, None)
            row_data.append(value if value is not None else '')
        
        style = row_style_callback(obj) if row_style_callback else None
        exporter.add_row(row_data, style=style)
    
    return exporter.get_response(filename)

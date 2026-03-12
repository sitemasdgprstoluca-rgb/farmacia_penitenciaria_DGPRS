"""
Utilidades para importacion masiva desde Excel.
Valida datos fila por fila y genera logs de importacion.

FORMATO EXCEL PRODUCTOS:
- Clave, Nombre, Unidad, Stock Minimo, Categoria, Sustancia Activa,
  Presentacion, Concentracion, Via Admin, Requiere Receta, Controlado, Estado

FORMATO EXCEL LOTES:
- Clave, Lote, Cantidad, Caducidad, Precio, Marca, Ubicacion
"""

import logging
import re
from datetime import datetime, date, timezone as dt_timezone, timedelta
from decimal import Decimal

import openpyxl
from django.db import transaction, connection
from django.db.models import F
from django.utils import timezone

from core.models import Producto, Lote, Centro, ImportacionLog
from core.utils.producto_variante import extraer_codigo_base

logger = logging.getLogger(__name__)


def _parse_fecha_excel(fecha_raw, nombre_campo='fecha'):
    """
    Convierte una fecha de openpyxl a objeto date de Python sin conversión de timezone.
    Normaliza a mediodía (12:00:00) para evitar desfase por timezone.
    Valida rango 1900-2100.
    """
    if not fecha_raw:
        return None
    
    try:
        resultado = None
        
        if isinstance(fecha_raw, datetime):
            fecha_normalizada = fecha_raw.replace(hour=12, minute=0, second=0, microsecond=0)
            if fecha_normalizada.tzinfo is not None:
                fecha_local = timezone.localtime(fecha_normalizada)
                resultado = fecha_local.date()
            else:
                resultado = fecha_normalizada.date()
            
        elif isinstance(fecha_raw, date):
            resultado = fecha_raw
        else:
            fecha_str = str(fecha_raw).strip()
            for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y', '%Y/%m/%d']:
                try:
                    dt = datetime.strptime(fecha_str, fmt)
                    dt = dt.replace(hour=12, minute=0, second=0, microsecond=0)
                    resultado = dt.date()
                    break
                except:
                    continue
            
            if resultado is None:
                raise ValueError(f'Formato de fecha no reconocido: {fecha_raw}')
        
        # Validar rango razonable
        if resultado:
            if resultado.year < 1900 or resultado.year > 2100:
                raise ValueError(f'{nombre_campo} fuera de rango (1900-2100): {resultado}')
        
        return resultado
        
    except Exception as e:
        raise ValueError(f'Error al parsear fecha: {e}')


class ResultadoImportacion:
    """Contenedor simple para acumular resultados de importacion."""

    def __init__(self, tipo_modelo: str):
        self.tipo_modelo = tipo_modelo
        self.total_procesados = 0
        self.exitosos = 0
        self.fallidos = 0
        self.actualizados = 0  # Para consolidaciones/actualizaciones
        self.omitidos = 0      # Para filas omitidas (ej: otro centro)
        self.errores = []
        self.detalle_actualizados = []  # Para tracking detallado de actualizaciones

    def agregar_error(self, fila, campo, error):
        self.errores.append({
            'fila': fila,
            'campo': campo,
            'error': str(error),
        })
        self.fallidos += 1

    def agregar_exito(self, es_actualizacion=False, info_actualizacion=None):
        self.exitosos += 1
        if es_actualizacion:
            self.actualizados += 1
            if info_actualizacion:
                self.detalle_actualizados.append(info_actualizacion)
    
    def agregar_omitido(self):
        self.omitidos += 1

    def incrementar_procesados(self):
        self.total_procesados += 1

    def get_dict(self):
        creados = self.exitosos - self.actualizados
        return {
            'exitosa': self.fallidos == 0 and self.exitosos > 0,
            'mensaje': self._generar_mensaje(),
            'total_registros': self.total_procesados,
            'registros_exitosos': self.exitosos,
            'registros_fallidos': self.fallidos,
            'creados': creados,
            'actualizados': self.actualizados,
            'omitidos': self.omitidos,
            'total_errores': self.fallidos,
            'tasa_exito': round((self.exitosos / self.total_procesados * 100) if self.total_procesados else 0, 2),
            'errores': self.errores[:100],
            'detalle_errores': self.errores[:100],  # Alias para compatibilidad con frontend
            'detalle_actualizados': self.detalle_actualizados[:100],  # Info de productos actualizados
        }
    
    def _generar_mensaje(self):
        """Genera mensaje descriptivo del resultado."""
        if self.fallidos == 0 and self.exitosos > 0:
            partes = []
            creados = self.exitosos - self.actualizados
            if creados > 0:
                partes.append(f"{creados} creados")
            if self.actualizados > 0:
                partes.append(f"{self.actualizados} actualizados")
            if self.omitidos > 0:
                partes.append(f"{self.omitidos} omitidos")
            return f"Importación exitosa: {', '.join(partes)}"
        elif self.exitosos > 0:
            return f"Importación parcial: {self.exitosos} exitosos, {self.fallidos} errores"
        else:
            return f"Error en importación: {self.fallidos} errores"


def cargar_excel(archivo):
    """Carga un archivo Excel con validaciones de seguridad."""
    try:
        # FIX: Usar read_only=False para evitar problemas con max_row=None
        # y acceso aleatorio a filas. El límite de 10K filas controla el uso de memoria.
        workbook = openpyxl.load_workbook(archivo, read_only=False, data_only=True)
        sheet = workbook.active
        filas_totales = sheet.max_row
        
        if filas_totales is None:
            logger.warning("Archivo rechazado: no se pudo determinar número de filas")
            return None, 0, False
        
        filas_totales = filas_totales - 1  # Restar encabezado

        if filas_totales > 10000:
            logger.warning(f"Archivo rechazado: {filas_totales} filas exceden límite")
            return None, 0, False
        
        if filas_totales <= 0:
            logger.warning("Archivo rechazado: sin filas de datos")
            return None, 0, False

        return workbook, filas_totales, True
    except Exception as exc:
        logger.error(f"Error al cargar Excel: {exc}")
        return None, 0, False


def normalizar_header(h):
    """Normaliza encabezados para mapeo robusto."""
    if not h:
        return ''
    texto = str(h).lower().strip()
    texto = texto.replace('*', '').replace('\n', ' ')
    texto = re.sub(r'\([^)]*\)', '', texto)
    # Remover acentos
    acentos = {'á':'a', 'é':'e', 'í':'i', 'ó':'o', 'ú':'u', 'ñ':'n', 'ü':'u'}
    for ac, rep in acentos.items():
        texto = texto.replace(ac, rep)
    texto = re.sub(r'[^a-z0-9]+', ' ', texto)
    return texto.strip()


def extraer_unidad_base(valor):
    """
    Extrae la unidad base de un valor de presentación.
    "CAJA CON 7 OVULOS" -> "CAJA"
    "FRASCO CON 120 ML" -> "FRASCO"
    """
    if not valor:
        return 'PIEZA'
    
    valor_upper = str(valor).upper().strip()
    
    # Mapeo de palabras a unidades estándar
    unidades_map = {
        'CAJA': 'CAJA', 'CAJAS': 'CAJA',
        'FRASCO': 'FRASCO', 'FRASCOS': 'FRASCO', 'FCO': 'FRASCO',
        'AMPOLLETA': 'AMPOLLETA', 'AMPOLLETAS': 'AMPOLLETA', 'AMP': 'AMPOLLETA',
        'SOBRE': 'SOBRE', 'SOBRES': 'SOBRE',
        'TABLETA': 'TABLETA', 'TABLETAS': 'TABLETA', 'TAB': 'TABLETA',
        'CAPSULA': 'CAPSULA', 'CAPSULAS': 'CAPSULA', 'CAP': 'CAPSULA',
        'PIEZA': 'PIEZA', 'PIEZAS': 'PIEZA', 'PZA': 'PIEZA', 'PZ': 'PIEZA',
        'TUBO': 'TUBO', 'TUBOS': 'TUBO',
        'BOLSA': 'BOLSA', 'BOLSAS': 'BOLSA',
        'ENVASE': 'FRASCO', 'ENVASES': 'FRASCO',
        'OVULO': 'PIEZA', 'OVULOS': 'PIEZA',
        'COMPRIMIDO': 'TABLETA', 'COMPRIMIDOS': 'TABLETA',
        'UNIDAD': 'PIEZA', 'UNIDADES': 'PIEZA',
        'ML': 'ML', 'MILILITROS': 'ML',
        'GR': 'GR', 'GRAMOS': 'GR', 'G': 'GR',
    }
    
    # Buscar la primera palabra que sea una unidad conocida
    palabras = valor_upper.split()
    for palabra in palabras:
        palabra_limpia = re.sub(r'[^A-Z]', '', palabra)
        if palabra_limpia in unidades_map:
            return unidades_map[palabra_limpia]
    
    return 'PIEZA'


def _parse_bool(valor):
    """Convierte valores variados a booleano."""
    if valor is None:
        return False
    if isinstance(valor, bool):
        return valor
    valor_str = str(valor).strip().lower()
    return valor_str in ['si', 'sí', 'yes', 'true', '1', 'verdadero', 'v', 'activo']


def importar_productos_desde_excel(archivo, usuario):
    """
    Importa productos desde Excel.
    
    Columnas soportadas:
    - Clave (REQUERIDO): código único del producto (615, 616, etc.)
    - Nombre (REQUERIDO): nombre del producto
    - Unidad: unidad de medida o presentación
    - Stock Minimo: inventario mínimo
    - Categoria: tipo de producto
    - Sustancia Activa: principio activo
    - Presentacion: forma farmacéutica
    - Concentracion: dosis
    - Via Admin: vía de administración
    - Requiere Receta: Si/No
    - Controlado: Si/No
    - Estado: Activo/Inactivo
    """
    resultado = ResultadoImportacion('Producto')
    workbook, _, valido = cargar_excel(archivo)
    if not valido:
        resultado.agregar_error(0, 'archivo', 'Archivo Excel invalido o vacio')
        return resultado.get_dict()

    sheet = workbook.active
    encabezados = [cell.value for cell in sheet[1]]
    encabezados_norm = [normalizar_header(e) for e in encabezados]
    
    logger.info(f"Productos - Encabezados: {encabezados}")
    logger.info(f"Productos - Normalizados: {encabezados_norm}")
    
    # Mapeo de columnas con sinónimos
    SINONIMOS = {
        'clave': ['clave', 'codigo', 'code', 'id', 'cve', 'sku', 'key', 'producto id', 
                  'codigo barras', 'codigo de barras', 'barcode'],
        'nombre': ['nombre', 'descripcion', 'nombre generico', 'medicamento', 'producto', 
                   'nombre del medicamento', 'nombre generico del medicamento', 'articulo'],
        'nombre_comercial': ['nombre comercial', 'marca comercial', 'nombre de marca', 'brand'],
        'unidad_medida': ['unidad', 'unidad medida', 'um', 'unidad de medida'],
        'categoria': ['categoria', 'tipo', 'clasificacion', 'clase', 'grupo'],
        'presentacion': ['presentacion', 'forma farmaceutica', 'forma', 'envase'],
        'sustancia_activa': ['sustancia activa', 'principio activo', 'formula', 'activo', 
                             'composicion', 'ingrediente'],
        'concentracion': ['concentracion', 'dosis', 'potencia', 'gramaje'],
        'via_administracion': ['via admin', 'via administracion', 'via', 'ruta', 'administracion'],
        'stock_minimo': ['stock minimo', 'minimo', 'stock min', 'inv minimo', 'inventario minimo'],
        'requiere_receta': ['requiere receta', 'receta', 'prescripcion'],
        'es_controlado': ['controlado', 'es controlado', 'control'],
        'activo': ['estado', 'activo', 'estatus', 'status'],
        'marca': ['marca', 'laboratorio', 'fabricante'],
    }
    
    def buscar_columna(sinonimos_lista):
        for i, h in enumerate(encabezados_norm):
            if not h:
                continue
            for sinonimo in sinonimos_lista:
                if sinonimo == h or (len(sinonimo) > 2 and sinonimo in h):
                    return i
        return -1
    
    col_map = {}
    for campo, sinonimos in SINONIMOS.items():
        idx = buscar_columna(sinonimos)
        if idx >= 0:
            col_map[campo] = idx
    
    logger.info(f"Productos - Mapeo: {col_map}")
    
    # ========================================================================
    # VALIDACIÓN: Detectar si se está usando la PLANTILLA DE LOTES por error
    # ========================================================================
    # Si encontramos columnas únicas de lotes, el usuario subió archivo incorrecto
    COLUMNAS_UNICAS_LOTES = ['lote', 'numero lote', 'num lote', 'batch', 'caducidad', 
                             'fecha caducidad', 'vencimiento', 'cantidad inicial',
                             'cantidad recibida', 'cantidad surtida', 
                             'cantidad contrato', 'cantidad contrato global']
    
    columnas_lotes_encontradas = []
    for h in encabezados_norm:
        if h:
            for col_lote in COLUMNAS_UNICAS_LOTES:
                if col_lote in h or h in col_lote:
                    columnas_lotes_encontradas.append(h)
                    break
    
    if len(columnas_lotes_encontradas) >= 2:
        # Alta probabilidad de ser plantilla de Lotes
        resultado.agregar_error(1, 'plantilla_incorrecta', 
            f'⚠️ PLANTILLA INCORRECTA: Parece que está usando una plantilla de LOTES '
            f'en el importador de PRODUCTOS. Columnas de lotes detectadas: {columnas_lotes_encontradas}. '
            f'Por favor use la plantilla de Productos (botón "Plantilla" en la página de Productos).')
        return resultado.get_dict()
    
    # Validar columnas mínimas
    if 'clave' not in col_map:
        resultado.agregar_error(1, 'encabezados', 
            f'No se encontró columna "Clave". Columnas: {encabezados}')
        return resultado.get_dict()
    
    if 'nombre' not in col_map:
        resultado.agregar_error(1, 'encabezados', 
            f'No se encontró columna "Nombre". Columnas: {encabezados}')
        return resultado.get_dict()
    
    creados = 0
    actualizados = 0

    with transaction.atomic():
        for fila_num in range(2, sheet.max_row + 1):
            resultado.incrementar_procesados()
            fila = list(sheet[fila_num])
            
            try:
                def get_val(col_name, default=None):
                    idx = col_map.get(col_name, -1)
                    if idx >= 0 and idx < len(fila):
                        val = fila[idx].value
                        if val is not None and str(val).strip():
                            return str(val).strip()
                    return default
                
                # ========== CLAVE (requerido) ==========
                clave_raw = get_val('clave')
                if not clave_raw:
                    resultado.total_procesados -= 1
                    continue
                
                # La clave se guarda tal cual viene (615, 616, 1A, etc.)
                clave = str(clave_raw).strip().upper()[:50]
                
                # ========== NOMBRE (requerido) ==========
                nombre_raw = get_val('nombre')
                if not nombre_raw:
                    resultado.agregar_error(fila_num, 'nombre', 'Nombre vacío')
                    continue
                nombre = nombre_raw[:500]
                
                # ========== NOMBRE COMERCIAL ==========
                nombre_comercial = get_val('nombre_comercial', '')
                
                # ========== UNIDAD DE MEDIDA ==========
                # Guardar texto libre completo (ej: "CAJA CON 7 OVULOS")
                unidad_raw = get_val('unidad_medida', 'PIEZA')
                unidad_medida = str(unidad_raw).strip().upper()[:100] if unidad_raw else 'PIEZA'
                
                # ========== PRESENTACION (obligatoria) ==========
                presentacion = get_val('presentacion', '')
                if not presentacion or not presentacion.strip():
                    resultado.agregar_error(fila_num, 'presentacion',
                        f'Presentación es OBLIGATORIA para producto "{nombre}" (clave: {clave}). '
                        f'Ejemplo: "CAJA CON 14 TABLETAS", "FRASCO 120ML"')
                    continue
                presentacion = presentacion.strip().upper()
                
                # ========== CATEGORIA ==========
                categoria_raw = get_val('categoria', '')
                if not categoria_raw or categoria_raw.upper() in ['N/A', 'NA', '-', 'NINGUNA', '']:
                    categoria = 'medicamento'
                else:
                    categoria = categoria_raw.lower().replace(' ', '_')
                    categorias_validas = ['medicamento', 'material_curacion', 'insumo']
                    if categoria not in categorias_validas:
                        categoria = 'medicamento'
                
                # ========== SUSTANCIA ACTIVA ==========
                sustancia_activa = get_val('sustancia_activa', '')
                
                # ========== CONCENTRACION ==========
                concentracion = get_val('concentracion', '')
                
                # ========== VIA ADMINISTRACION ==========
                via_administracion = get_val('via_administracion', '')
                
                # ========== MARCA ==========
                marca = get_val('marca', '')
                
                # ========== STOCK MINIMO ==========
                stock_raw = get_val('stock_minimo', '1')
                try:
                    stock_minimo = max(0, int(float(stock_raw)))
                except (ValueError, TypeError):
                    stock_minimo = 1
                
                # ========== REQUIERE RECETA ==========
                requiere_receta = _parse_bool(get_val('requiere_receta', 'No'))
                
                # ========== ES CONTROLADO (OBLIGATORIO) ==========
                es_controlado_raw = get_val('es_controlado')
                if es_controlado_raw is None or str(es_controlado_raw).strip() == '':
                    resultado.agregar_error(fila_num, 'es_controlado', 'Campo obligatorio: Medicamento Controlado no definido')
                    continue
                es_controlado = _parse_bool(es_controlado_raw)
                
                # ========== ACTIVO/ESTADO ==========
                # ISS-FIX: Por defecto activo=True, solo desactivar si explícitamente dice 'No' o 'Inactivo'
                estado_raw = get_val('activo', None)
                if estado_raw is None or str(estado_raw).strip() == '':
                    activo = True  # Default: activo
                else:
                    valor_str = str(estado_raw).strip().lower()
                    # Solo es inactivo si explícitamente dice no/inactivo/false/0
                    activo = valor_str not in ['no', 'inactivo', 'false', '0', 'n', 'inactive']
                
                # ========== DESCRIPCION ==========
                desc_parts = [p for p in [presentacion, concentracion, marca] if p]
                descripcion = ', '.join(desc_parts)[:500] if desc_parts else None

                # ISS-PROD-VAR: Crear o reusar variante por presentación
                from core.utils.producto_variante import obtener_o_crear_variante
                defaults_prod = {
                    'nombre_comercial': nombre_comercial[:200] if nombre_comercial else None,
                    'descripcion': descripcion,
                    'unidad_medida': unidad_medida,
                    'categoria': categoria,
                    'sustancia_activa': sustancia_activa[:200] if sustancia_activa else None,
                    'concentracion': concentracion[:100] if concentracion else None,
                    'via_administracion': via_administracion[:50] if via_administracion else None,
                    'stock_minimo': stock_minimo,
                    'requiere_receta': requiere_receta,
                    'es_controlado': es_controlado,
                    'activo': activo,
                }
                obj, created, var_info = obtener_o_crear_variante(
                    clave_input=clave,
                    nombre=nombre,
                    presentacion=presentacion[:200] if presentacion else '',
                    defaults=defaults_prod,
                )

                if created:
                    creados += 1
                    resultado.agregar_exito(es_actualizacion=False)
                else:
                    actualizados += 1
                    codigo_asignado = var_info.get('codigo_asignado', clave)
                    motivo = var_info.get('motivo', '')
                    if var_info.get('es_variante'):
                        msg = (
                            f'Variante existente reutilizada: {codigo_asignado} '
                            f'(presentación equivalente a clave {clave})'
                        )
                    else:
                        msg = f'Producto ya existía con clave {codigo_asignado}, datos actualizados'
                    resultado.agregar_exito(es_actualizacion=True, info_actualizacion={
                        'fila': fila_num,
                        'clave': codigo_asignado,
                        'nombre': nombre[:100],
                        'mensaje': msg,
                    })
                
            except Exception as exc:
                logger.exception(f"Error fila {fila_num}: {exc}")
                resultado.agregar_error(fila_num, 'general', str(exc))

    result = resultado.get_dict()
    result['creados'] = creados
    result['actualizados'] = actualizados
    return result


def _propagar_contrato_global(filas_consolidadas, centro):
    """
    Propaga cantidad_contrato_global a todos los lotes activos del mismo
    producto + numero_contrato cuando alguno de los lotes importados lo define.
    """
    # Recopilar valores de ccg por (producto_id, numero_contrato)
    ccg_map = {}
    for fila in filas_consolidadas:
        ccg = fila.get('cantidad_contrato_global')
        nc = fila.get('numero_contrato')
        pid = fila.get('producto_id')
        if ccg is not None and nc and pid:
            key = (pid, str(nc).strip().upper())
            ccg_map[key] = ccg
    
    if not ccg_map:
        return
    
    for (pid, nc_upper), ccg_val in ccg_map.items():
        qs = Lote.objects.filter(
            producto_id=pid,
            activo=True,
        ).filter(numero_contrato__iexact=nc_upper)
        if centro:
            qs = qs.filter(centro=centro)
        else:
            qs = qs.filter(centro__isnull=True)
        qs.update(cantidad_contrato_global=ccg_val)
        logger.info(f"Propagado contrato global {ccg_val} a lotes de producto {pid} contrato {nc_upper}")


def _verificar_contrato_global_excedido(filas_consolidadas, centro):
    """
    Después de crear lotes, verifica si algún contrato global fue excedido.
    Retorna lista de alertas (strings) o lista vacía.
    
    NOTA: Se basa en cantidad_inicial (recibido), NO en cantidad_actual (stock).
    Las salidas a centros NO afectan esta verificación.
    """
    from django.db.models import Sum
    
    alertas = []
    # Recopilar combinaciones únicas de (producto_id, numero_contrato)
    pares_verificados = set()
    for fila in filas_consolidadas:
        nc = fila.get('numero_contrato')
        pid = fila.get('producto_id')
        clave = fila.get('clave_producto', '?')
        if nc and pid:
            key = (pid, str(nc).strip().upper())
            if key in pares_verificados:
                continue
            pares_verificados.add(key)
            
            # Buscar el ccg vigente en BD
            qs = Lote.objects.filter(
                producto_id=pid,
                numero_contrato__iexact=key[1],
                activo=True,
            )
            if centro:
                qs = qs.filter(centro=centro)
            else:
                qs = qs.filter(centro__isnull=True)
            
            lote_ref = qs.filter(cantidad_contrato_global__isnull=False).first()
            if not lote_ref or lote_ref.cantidad_contrato_global is None:
                continue
            
            ccg = lote_ref.cantidad_contrato_global
            total_recibido = qs.aggregate(total=Sum('cantidad_inicial'))['total'] or 0
            
            if total_recibido > ccg:
                excedente = total_recibido - ccg
                alertas.append(
                    f'⚠️ Clave {clave} (contrato {nc}): se excede el contrato global por '
                    f'{excedente} unidades. Contratado: {ccg}, total recibido: {total_recibido}.'
                )
    
    return alertas


def _validar_ccg_antes_de_importar(filas_consolidadas, centro):
    """
    BLOQUEO DURO PRE-CREACIÓN: verifica dentro de transaction.atomic() que ningún
    contrato global (cantidad_contrato_global) será excedido por la importación.
    Lanza ValueError si alguna combinación (producto, numero_contrato) superaría su CCG.

    Se debe llamar ANTES del bucle de creación/actualización de lotes, dentro del
    transaction.atomic(), para garantizar que ningún lote se persista si hay violación.

    A diferencia de _verificar_contrato_global_excedido (post-creación/soft-alert),
    esta función bloquea la importación completa antes de escribir nada.

    NOTA: No usa activo=True para no subestimar el total ya recibido.
    """
    from django.db.models import Sum

    # Agrupar la cantidad a importar por (producto_id, numero_contrato)
    incoming: dict = {}
    ccg_por_grupo: dict = {}
    for fila in filas_consolidadas:
        nc = fila.get('numero_contrato')
        pid = fila.get('producto_id')
        ccg = fila.get('cantidad_contrato_global')
        if not nc or not pid or ccg is None:
            continue
        key = (pid, str(nc).strip().upper())
        incoming[key] = incoming.get(key, 0) + fila.get('cantidad_inicial', 0)
        # mantener el valor más reciente (CCG es fijo, no se acumula)
        ccg_por_grupo[key] = ccg

    if not ccg_por_grupo:
        return  # ninguna fila tiene CCG definido → nada que verificar

    for (pid, nc_upper), ccg in ccg_por_grupo.items():
        incoming_total = incoming.get((pid, nc_upper), 0)
        if incoming_total == 0:
            continue

        # CCG aplica exclusivamente a Farmacia Central (centro=None).
        # Los centros penitenciarios NO manejan contratos — solo hacen requisiciones,
        # reciben de farmacia y registran salidas. Si un centro hace una compra de
        # emergencia, su Excel no traerá cantidad_contrato_global, por lo que
        # ccg_por_grupo estará vacío y se regresará arriba sin llegar aquí.
        # El filtro por centro es una salvaguarda defensiva.
        # SELECT FOR UPDATE para evitar race conditions entre importaciones concurrentes.
        if centro is not None:
            qs_centro = Lote.objects.filter(centro=centro)
        else:
            qs_centro = Lote.objects.filter(centro__isnull=True)
        list(
            qs_centro.select_for_update().filter(
                producto_id=pid,
                numero_contrato__iexact=nc_upper,
            ).values_list('id', flat=True)
        )
        total_ya_en_bd = (
            qs_centro.filter(
                producto_id=pid,
                numero_contrato__iexact=nc_upper,
            ).aggregate(total=Sum('cantidad_inicial'))['total'] or 0
        )
        proyectado = total_ya_en_bd + incoming_total

        if proyectado > ccg:
            exceso = proyectado - ccg
            disponible = max(0, ccg - total_ya_en_bd)
            raise ValueError(
                f'Importación rechazada: el contrato global "{nc_upper}" para el producto '
                f'(ID {pid}) sería excedido. '
                f'Contrato global: {ccg}, ya recibido en BD: {total_ya_en_bd}, '
                f'a importar: {incoming_total}, exceso: {exceso}. '
                f'Máximo permitido en esta importación: {disponible}.'
            )


def _consolidar_filas_importacion(filas_parseadas):
    """
    ISS-IMPORT-CONSOLIDATION: Pre-procesa filas parseadas y las consolida.
    
    Reglas:
    1. Filas con mismo (lote_base, producto_id, fecha_caducidad) → se suman cantidades.
    2. Filas con mismo (lote_base, producto_id) pero DIFERENTE fecha_caducidad 
       → se genera sufijo .2, .3, etc. para diferenciar.
    
    DECISIÓN DE DISEÑO sobre sufijos existentes:
    Si un lote ya viene con sufijo desde el proveedor (ej. "ABC123.2"), se RESPETA
    como su lote_base (se trata como identidad distinta a "ABC123").
    Así el proveedor mantiene control de su propia nomenclatura.
    
    Parámetros:
        filas_parseadas: lista de dicts con al menos:
            lote_base, producto_id, fecha_caducidad, cantidad_inicial,
            cantidad_contrato (puede ser None), y demás campos opcionales.
    
    Retorna:
        lista consolidada de dicts listos para crear/actualizar lotes.
    """
    from collections import OrderedDict

    # Paso 1: Agrupar por (lote_base, producto_id, fecha_caducidad) y sumar cantidades
    grupo_completo = OrderedDict()  # (lote_base, producto_id, fecha_cad) → fila acumulada
    for fila in filas_parseadas:
        key = (fila['lote_base'], fila['producto_id'], fila['fecha_caducidad'])
        if key not in grupo_completo:
            grupo_completo[key] = dict(fila)  # copia
        else:
            grupo_completo[key]['cantidad_inicial'] += fila['cantidad_inicial']
            # cantidad_contrato: si ambos tienen valor, sumar; si alguno es None, mantener el existente
            existing_cc = grupo_completo[key].get('cantidad_contrato')
            new_cc = fila.get('cantidad_contrato')
            if existing_cc is not None and new_cc is not None:
                grupo_completo[key]['cantidad_contrato'] = existing_cc + new_cc
            elif new_cc is not None:
                grupo_completo[key]['cantidad_contrato'] = new_cc
            # cantidad_contrato_global: mantener el valor más reciente (no sumar, es global)
            new_ccg = fila.get('cantidad_contrato_global')
            if new_ccg is not None:
                grupo_completo[key]['cantidad_contrato_global'] = new_ccg
            # fecha_fabricacion/recepción: mantener la más reciente si hay múltiples entregas parciales
            existing_fecha_fab = grupo_completo[key].get('fecha_fabricacion')
            new_fecha_fab = fila.get('fecha_fabricacion')
            if new_fecha_fab is not None:
                if existing_fecha_fab is None or new_fecha_fab > existing_fecha_fab:
                    grupo_completo[key]['fecha_fabricacion'] = new_fecha_fab
            # filas_origen: acumular para auditoría
            grupo_completo[key].setdefault('filas_origen', [fila.get('fila_num', '?')])
            grupo_completo[key]['filas_origen'].append(fila.get('fila_num', '?'))

    consolidadas = list(grupo_completo.values())

    # Paso 2: Para cada (lote_base, producto_id), detectar si hay múltiples caducidades
    # y asignar sufijos .2, .3 al segundo, tercer grupo con distinta caducidad.
    lote_producto_caducidades = OrderedDict()  # (lote_base, producto_id) → [fecha_cad1, fecha_cad2, ...]
    for fila in consolidadas:
        lp_key = (fila['lote_base'], fila['producto_id'])
        lote_producto_caducidades.setdefault(lp_key, [])
        if fila['fecha_caducidad'] not in lote_producto_caducidades[lp_key]:
            lote_producto_caducidades[lp_key].append(fila['fecha_caducidad'])

    for fila in consolidadas:
        lp_key = (fila['lote_base'], fila['producto_id'])
        caducidades = lote_producto_caducidades[lp_key]
        if len(caducidades) > 1:
            # Múltiples caducidades para este lote+producto: asignar sufijo
            idx = caducidades.index(fila['fecha_caducidad'])
            if idx == 0:
                fila['numero_lote'] = fila['lote_base']  # el primero se queda sin sufijo
            else:
                fila['numero_lote'] = f"{fila['lote_base']}.{idx + 1}"
        else:
            fila['numero_lote'] = fila['lote_base']

    return consolidadas


def importar_lotes_desde_excel(archivo, usuario, centro_id=None):
    """
    Importa lotes desde Excel.
    
    COLUMNAS OBLIGATORIAS:
    - Clave Producto (REQUERIDO): código/clave del producto
    - Nombre Producto (REQUERIDO): nombre del producto (debe coincidir con clave)
    - Lote (REQUERIDO): número de lote
    - Cantidad Inicial / Cantidad Recibida (REQUERIDO): cantidad que llegó/se surtió
    - Fecha Caducidad (REQUERIDO): fecha de vencimiento
    
    COLUMNAS OPCIONALES:
    - Cantidad Contrato: total según contrato (para cuando llega menos de lo pactado).
                         Si viene vacía o la columna no existe, se guarda como NULL.
    - Precio Unitario: precio unitario (default 0)
    - Número Contrato: número de contrato
    - Marca: laboratorio (opcional)
    - Fecha Entrega: fecha de entrega del lote
    - Activo: estado del lote (default Activo)
    
    ISS-INV-001: SOPORTE PARA CONTRATOS PARCIALES
    =============================================
    Si el contrato establece 100 unidades pero solo llegan 80:
    - Cantidad Contrato = 100 (lo esperado según contrato)
    - Cantidad Inicial = 80 (lo que realmente llegó)
    
    Cuando llegan más unidades del contrato, se puede re-importar el Excel
    con las mismas clave/lote/contrato/marca/caducidad y el sistema SUMARÁ
    la cantidad a la existente (pero mantiene la cantidad_contrato original).
    
    ISS-IMPORT-CONSOLIDATION: CONSOLIDACIÓN DE PARCIALIDADES
    ========================================================
    Si en el mismo archivo hay filas con mismo lote+producto+caducidad,
    se suman las cantidades en un solo registro (evita duplicados).
    
    Si hay filas con mismo lote+producto pero DIFERENTE caducidad,
    se crean lotes diferenciados con sufijo: Lote, Lote.2, Lote.3, etc.
    
    IMPORTANTE: El sistema verifica que CLAVE y NOMBRE coincidan con el producto
    en la base de datos. Si hay discrepancia, se reporta error para evitar
    sumar cantidades a productos incorrectos.
    
    Detecta automáticamente la fila de encabezados.
    """
    resultado = ResultadoImportacion('Lote')
    workbook, _, valido = cargar_excel(archivo)
    if not valido:
        resultado.agregar_error(0, 'archivo', 'Archivo Excel invalido o vacio')
        return resultado.get_dict()

    # P0-1: Calcular checksum del archivo para idempotencia de reimportaciones
    from core.utils.parcialidad_merge import calcular_file_checksum
    file_checksum = calcular_file_checksum(archivo)
    archivo_nombre = getattr(archivo, 'name', str(archivo)) if archivo else 'unknown'
    logger.info(f"[IMPORT-CHECKSUM] Archivo: {archivo_nombre}, Checksum: {file_checksum[:16]}...")

    sheet = workbook.active
    
    # Detectar fila de encabezados (puede estar en fila 1, 2 o 3)
    fila_inicio_datos = 2
    encabezados = []
    for fila_num in range(1, min(5, sheet.max_row + 1)):
        temp_headers = [cell.value for cell in sheet[fila_num]]
        # Si encuentra columnas con "lote" o "producto", esa es la fila de encabezados
        if any(h and ('lote' in str(h).lower() or 'producto' in str(h).lower()) for h in temp_headers):
            encabezados = temp_headers
            fila_inicio_datos = fila_num + 1
            logger.info(f"Lotes - Encabezados detectados en fila {fila_num}: {encabezados}")
            break
    
    if not encabezados or not any(h for h in encabezados):
        encabezados = [cell.value for cell in sheet[1]]
        fila_inicio_datos = 2
    
    encabezados_norm = [normalizar_header(e) for e in encabezados]
    
    logger.info(f"Lotes - Encabezados: {encabezados}")
    logger.info(f"Lotes - Normalizados: {encabezados_norm}")
    
    # IMPORTANTE: Los sinónimos más específicos primero para evitar conflictos
    # 'nombre producto' debe mapearse a producto_nombre, NO a producto_clave
    # ISS-FORMATO-OFICIAL: Agregar sinónimos del formato "Control de Inventarios" oficial
    SINONIMOS_LOTE = {
        # Producto - nombre y clave son OBLIGATORIOS
        # NOTA: "NO. PARTIDA" es número de fila, NO clave de producto - no incluir
        'producto_nombre': ['nombre producto', 'nombre del producto', 'producto nombre', 'descripcion',
                            'articulo', 'nombre articulo'],  # Formato oficial usa ARTÍCULO
        'producto_clave': ['clave producto', 'clave', 'codigo producto', 'codigo', 'sku', 'key'],
        'producto_id': ['id producto', 'producto id', 'id_producto'],
        # Lote
        'numero_lote': ['numero lote', 'lote', 'num lote', 'no lote', 'numero de lote', 
                        'n lote', 'nro lote', 'batch'],
        # ISS-INV-001: Separar cantidad_contrato (total esperado por lote) de cantidad_inicial (recibido)
        'cantidad_contrato': ['cantidad contrato', 'cantidad contrato lote', 'cant contrato',
                              'cant contrato lote', 'total contrato lote',
                              'cantidad esperada', 'qty contrato', 'cantidad por contrato',
                              'contratado', 'pactado', 'cant pactada'],
        'cantidad_contrato_global': ['cantidad contrato global', 'contrato global',
                                     'total contrato global', 'cantidad global contrato',
                                     'cant contrato global', 'qty contrato global',
                                     'ccg', 'cant ccg', 'total global', 'global contrato',
                                     'cant global', 'cantidad global'],
        'cantidad_inicial': ['cantidad inicial', 'cantidad', 'cant inicial', 'stock', 'existencia', 
                             'qty', 'unidades', 'piezas', 'cant', 'cantidad recibida', 
                             'cantidad surtida', 'cant recibida', 'cant surtida'],
        'cantidad_actual': ['cantidad actual', 'cant actual', 'stock actual'],
        # Fechas - ISS-FORMATO-OFICIAL: "VENCIMIENTO/FECHA DE CADUCIDAD" del formato oficial
        'caducidad': ['fecha caducidad', 'caducidad', 'vencimiento', 'fecha vencimiento', 
                      'expira', 'fec cad', 'expiracion', 'fecha expiracion',
                      'vencimiento fecha de caducidad'],  # Formato oficial combinado
        # Fecha de Entrega: Sinónimos exhaustivos para máxima compatibilidad
        'recepcion': ['fecha entrega', 'entrega', 'fecha de entrega', 'f entrega', 'fec entrega',
                      'fecha recepcion', 'recepcion', 'fecha de recepcion', 'fec recepcion',
                      'fecha fabricacion', 'fabricacion', 'fecha de fabricacion', 'fec fab',
                      'elaboracion', 'fecha elaboracion', 'fecha de elaboracion',
                      'f recepcion', 'fec rec', 'fecha rec', 'fecha ent',
                      'fecha recep'],  # ISS-FIX: más variantes
        'fecha_ingreso': ['fecha ingreso', 'fecha de ingreso', 'ingreso', 'fecha entrada', 
                          'fec ing'],  # Formato oficial
        # Otros campos
        'precio': ['precio unitario', 'precio', 'costo', 'valor', 'precio unit', 'pu'],
        'contrato': ['numero contrato', 'contrato', 'no contrato', 'num contrato'],
        # ISS-FORMATO-OFICIAL: "NOMBRE COMERCIAL O GENÉRICO" se usa como marca/laboratorio
        'marca': ['marca', 'laboratorio', 'fabricante', 'proveedor', 'lab',
                  'nombre comercial', 'nombre comercial o generico', 'nombre generico'],
        'concentracion': ['concentracion', 'conc', 'dosis'],  # Campo informativo
        'presentacion': ['presentacion', 'pres', 'envase', 'empaque'],  # Campo informativo
        'ubicacion': ['ubicacion', 'almacen', 'bodega', 'estante', 'localizacion'],
        'centro': ['centro', 'centro nombre', 'nombre centro', 'centro destino'],
        'activo': ['activo', 'estado', 'status', 'active'],
    }
    
    def buscar_columna(sinonimos_lista):
        # Primero buscar coincidencia EXACTA
        for i, h in enumerate(encabezados_norm):
            if not h:
                continue
            if h in sinonimos_lista:
                return i
        return -1
    
    col_map = {}
    for campo, sinonimos in SINONIMOS_LOTE.items():
        idx = buscar_columna(sinonimos)
        if idx >= 0:
            col_map[campo] = idx
    
    # Log de columnas detectadas (nivel info para diagnóstico)
    logger.info(f"[IMPORTADOR] Columnas detectadas: {list(col_map.keys())}")
    # Log específico para fecha de entrega
    if 'recepcion' in col_map:
        logger.info(f"[IMPORTADOR] Columna FECHA ENTREGA mapeada a índice {col_map['recepcion']} (header: {encabezados[col_map['recepcion']]})")
    else:
        logger.warning(f"[IMPORTADOR] Columna FECHA ENTREGA NO encontrada. Headers: {encabezados}")
    
    # FIX: Validar columnas mínimas - CLAVE, NOMBRE y PRESENTACIÓN son OBLIGATORIAS
    # Todos deben coincidir con el producto en la base de datos
    tiene_clave = ('producto_clave' in col_map or 'producto_id' in col_map)
    tiene_nombre = 'producto_nombre' in col_map
    tiene_presentacion = 'presentacion' in col_map
    tiene_lote = 'numero_lote' in col_map
    tiene_cantidad = 'cantidad_inicial' in col_map
    tiene_caducidad = 'caducidad' in col_map
    
    # ========================================================================
    # VALIDACIÓN: Detectar si se está usando la PLANTILLA DE PRODUCTOS por error
    # ========================================================================
    # Si encontramos columnas únicas de productos, el usuario subió archivo incorrecto
    COLUMNAS_UNICAS_PRODUCTOS = ['sustancia activa', 'principio activo', 'via admin',
                                  'via administracion', 'requiere receta', 'receta',
                                  'controlado', 'es controlado', 'stock minimo',
                                  'inventario minimo', 'categoria clasificacion']
    
    columnas_productos_encontradas = []
    for h in encabezados_norm:
        if h:
            for col_prod in COLUMNAS_UNICAS_PRODUCTOS:
                if col_prod in h or h == col_prod:
                    columnas_productos_encontradas.append(h)
                    break
    
    # Verificar también que NO tiene columnas de lotes (lote, caducidad, cantidad)
    tiene_columnas_lotes = tiene_lote and tiene_caducidad and tiene_cantidad
    
    if len(columnas_productos_encontradas) >= 2 and not tiene_columnas_lotes:
        # Alta probabilidad de ser plantilla de Productos
        resultado.agregar_error(1, 'plantilla_incorrecta', 
            f'⚠️ PLANTILLA INCORRECTA: Parece que está usando una plantilla de PRODUCTOS '
            f'en el importador de LOTES. Columnas de productos detectadas: {columnas_productos_encontradas}. '
            f'Por favor use la plantilla de Lotes (botón "Plantilla" en la página de Lotes).')
        return resultado.get_dict()
    
    if not (tiene_clave and tiene_nombre and tiene_presentacion and tiene_lote and tiene_cantidad and tiene_caducidad):
        faltantes = []
        if not tiene_clave:
            faltantes.append('Clave Producto (obligatoria)')
        if not tiene_nombre:
            faltantes.append('Nombre Producto (obligatorio)')
        if not tiene_presentacion:
            faltantes.append('Presentación (OBLIGATORIA - ej: CAJA CON 14 TABLETAS)')
        if not tiene_lote:
            faltantes.append('Número Lote')
        if not tiene_cantidad:
            faltantes.append('Cantidad Inicial')
        if not tiene_caducidad:
            faltantes.append('Fecha Caducidad')
        
        resultado.agregar_error(1, 'encabezados', 
            f'Columnas faltantes: {", ".join(faltantes)}. Detectadas: {encabezados}. '
            f'NOTA: Clave, Nombre y Presentación del producto son OBLIGATORIOS y deben coincidir exactamente con el catálogo.')
        return resultado.get_dict()
    
    # Centro: NULL = Almacén Central (FARMACIA)
    # Solo se usa centro_id si se pasa explícitamente como parámetro
    centro = None
    if centro_id:
        try:
            centro = Centro.objects.get(id=centro_id)
            logger.info(f"Lotes - Centro especificado: {centro.nombre} (ID: {centro.id})")
        except Centro.DoesNotExist:
            resultado.agregar_error(0, 'centro', f'Centro ID {centro_id} no existe')
            return resultado.get_dict()
    else:
        # NULL = Almacén Central (FARMACIA) - comportamiento por defecto
        logger.info("Lotes - Centro: NULL (Almacén Central/FARMACIA)")

    creados = 0
    consolidados_archivo = 0  # Filas consolidadas dentro del archivo
    omitidos_centro = 0  # ISS-FIX: Contar filas omitidas por pertenecer a otro centro
    
    # ========================================================================
    # FASE 1: PARSEO - Leer todas las filas y validar datos individuales
    # ========================================================================
    filas_parseadas = []
    
    for fila_num in range(fila_inicio_datos, sheet.max_row + 1):
        resultado.incrementar_procesados()
        fila = list(sheet[fila_num])
        
        try:
            # ISS-IMPORT-FIX: Lista de valores que se consideran como "vacío" en Excel
            VALORES_VACIOS = {'-', '--', '---', 'n/a', 'na', 'null', 'none', '', ' ', 'sin dato', 'sin datos', 's/d'}
            
            def get_val(col_name, default=None):
                idx = col_map.get(col_name, -1)
                if idx >= 0 and idx < len(fila):
                    val = fila[idx].value
                    if val is not None:
                        val_str = str(val).strip()
                        # Tratar valores como "-", "N/A" como vacíos
                        if val_str.lower() in VALORES_VACIOS:
                            return default
                        if val_str:
                            return val_str
                return default
            
            # ========== DETECCIÓN DE FILAS VACÍAS O INCOMPLETAS ==========
            # Si los campos esenciales están TODOS vacíos, omitir silenciosamente
            clave_test = get_val('producto_clave')
            nombre_test = get_val('producto_nombre')
            lote_test = get_val('numero_lote')
            cantidad_test = get_val('cantidad_inicial', '0')
            
            # Fila completamente vacía: omitir sin error
            if not clave_test and not nombre_test and not lote_test and (not cantidad_test or cantidad_test == '0'):
                logger.debug(f"Fila {fila_num}: OMITIDA - fila vacía o incompleta")
                resultado.agregar_omitido()
                continue
            
            # Fila con solo cantidad (sin producto/lote): probablemente fila residual
            if not clave_test and not nombre_test and not lote_test:
                logger.debug(f"Fila {fila_num}: OMITIDA - sin datos de producto/lote")
                resultado.agregar_omitido()
                continue
            
            # ========== PRODUCTO (requerido - CLAVE, NOMBRE Y PRESENTACIÓN son OBLIGATORIOS) ==========
            producto = None
            clave_producto = None
            nombre_producto = None
            presentacion_producto = None
            
            if 'producto_clave' in col_map:
                clave_producto = get_val('producto_clave')
            if 'producto_nombre' in col_map:
                nombre_producto = get_val('producto_nombre')
            if 'presentacion' in col_map:
                presentacion_producto = get_val('presentacion')
            
            if not clave_producto:
                resultado.agregar_error(fila_num, 'producto', 
                    f'Clave de producto es OBLIGATORIA. Nombre proporcionado: {nombre_producto or "N/A"}')
                continue
            
            if not nombre_producto:
                resultado.agregar_error(fila_num, 'producto', 
                    f'Nombre de producto es OBLIGATORIO. Clave proporcionada: {clave_producto}')
                continue
            
            if not presentacion_producto:
                resultado.agregar_error(fila_num, 'producto', 
                    f'Presentación es OBLIGATORIA. Clave: {clave_producto}, Nombre: {nombre_producto}. '
                    f'Ejemplo: "CAJA CON 14 TABLETAS"')
                continue
            
            # ISS-PRESENTACION-FIX: Normalizar presentación para comparación
            def normalizar_presentacion(pres):
                """Normaliza presentación para comparación tolerante."""
                if not pres:
                    return ''
                p = pres.strip().upper()
                # Normalizar espacios múltiples
                p = re.sub(r'\s+', ' ', p)
                return p
            
            presentacion_excel_norm = normalizar_presentacion(presentacion_producto)
            
            # Buscar producto por CLAVE exacta primero
            try:
                producto = Producto.objects.get(clave__iexact=clave_producto)
            except Producto.DoesNotExist:
                # ISS-VARIANTE-FIX: Usar extraer_codigo_base para manejar claves tipo 010.000.0001.2
                # donde los puntos son parte de la clave, no solo sufijos de variante
                clave_base = extraer_codigo_base(clave_producto)
                # Buscar productos con clave base exacta o variantes (base, base.2, base.3...)
                patron_variante = re.compile(r'^' + re.escape(clave_base) + r'(\.\d+)?$', re.IGNORECASE)
                productos_posibles = [p for p in Producto.objects.filter(clave__istartswith=clave_base) 
                                      if patron_variante.match(p.clave)]
                
                if productos_posibles:
                    # Buscar el que tenga la presentación correcta
                    for p in productos_posibles:
                        pres_bd_norm = normalizar_presentacion(p.presentacion)
                        if pres_bd_norm == presentacion_excel_norm:
                            producto = p
                            logger.info(f"Fila {fila_num}: Producto encontrado por presentación: {p.clave} - {p.presentacion}")
                            break
                
                if not producto:
                    pres_disponibles = [f"{p.clave}: {p.presentacion}" for p in productos_posibles] if productos_posibles else []
                    resultado.agregar_error(fila_num, 'producto', 
                        f'Clave "{clave_producto}" no encontrada. '
                        f'Presentación buscada: "{presentacion_producto}". '
                        f'Productos similares disponibles: {pres_disponibles or "ninguno"}. '
                        f'Verifique clave y presentación en el catálogo.')
                    continue
            
            # Verificar que la presentación del producto coincida
            presentacion_bd_norm = normalizar_presentacion(producto.presentacion)
            if presentacion_bd_norm != presentacion_excel_norm:
                # ISS-VARIANTE-FIX: Buscar producto con misma clave base y presentación correcta
                clave_base = extraer_codigo_base(clave_producto)
                patron_variante = re.compile(r'^' + re.escape(clave_base) + r'(\.\d+)?$', re.IGNORECASE)
                productos_con_misma_base = [p for p in Producto.objects.filter(clave__istartswith=clave_base)
                                            if patron_variante.match(p.clave)]
                
                producto_correcto = None
                for p in productos_con_misma_base:
                    if normalizar_presentacion(p.presentacion) == presentacion_excel_norm:
                        producto_correcto = p
                        break
                
                if producto_correcto:
                    logger.info(f"Fila {fila_num}: Redirigiendo de {producto.clave} a {producto_correcto.clave} por presentación")
                    producto = producto_correcto
                else:
                    pres_disponibles = [f"{p.clave}: {p.presentacion or 'SIN PRES'}" for p in productos_con_misma_base]
                    resultado.agregar_error(fila_num, 'producto', 
                        f'⚠️ PRESENTACIÓN NO COINCIDE: Clave "{clave_producto}" tiene presentación '
                        f'"{producto.presentacion or "N/A"}" en BD, pero Excel dice "{presentacion_producto}". '
                        f'Productos disponibles con clave similar: {pres_disponibles}. '
                        f'Use la clave y presentación exactas del catálogo.')
                    continue
            
            # ISS-NOMBRES-TOLERANTES: Normalizar nombres para comparación flexible
            # Esto maneja variaciones como "KETOCONAZOL /CLINDAMICINA" vs "KETOCONAZOL / CLINDAMICINA"
            def normalizar_nombre_comparacion(nombre):
                """Normaliza nombre para comparación tolerante."""
                n = nombre.strip().lower()
                # Normalizar separadores con espacios inconsistentes
                n = re.sub(r'\s*/\s*', '/', n)  # "/ " → "/"
                n = re.sub(r'\s*-\s*', '-', n)  # " - " → "-"
                n = re.sub(r'\s+', ' ', n)      # espacios múltiples → uno
                return n
            
            nombre_bd_normalizado = normalizar_nombre_comparacion(producto.nombre)
            nombre_excel_normalizado = normalizar_nombre_comparacion(nombre_producto)
            if not (nombre_bd_normalizado == nombre_excel_normalizado or 
                    nombre_bd_normalizado.startswith(nombre_excel_normalizado) or
                    nombre_excel_normalizado.startswith(nombre_bd_normalizado)):
                resultado.agregar_error(fila_num, 'producto', 
                    f'DISCREPANCIA: Clave "{clave_producto}" corresponde a '
                    f'"{producto.nombre}" en BD, pero Excel dice "{nombre_producto}". '
                    f'Verifique que clave y nombre sean correctos.')
                continue
            
            # ========== NUMERO LOTE (requerido) ==========
            numero_lote = get_val('numero_lote')
            if not numero_lote:
                resultado.agregar_error(fila_num, 'lote', 
                    'Producto y numero de lote son obligatorios')
                continue
            
            # ========== FILTRO DE CENTRO/UBICACIÓN ==========
            ubicacion_excel = get_val('ubicacion', '')
            centro_excel = get_val('centro', '')
            ubicacion_or_centro = (ubicacion_excel or centro_excel or '').strip().lower()
            
            if not centro and ubicacion_or_centro:
                es_almacen_central = (
                    not ubicacion_or_centro or
                    'almac' in ubicacion_or_centro or
                    'central' in ubicacion_or_centro or
                    'farmacia' in ubicacion_or_centro
                )
                if not es_almacen_central:
                    logger.info(
                        f"Fila {fila_num}: OMITIDA - lote {numero_lote} pertenece a "
                        f"'{ubicacion_excel or centro_excel}', no a Farmacia Central"
                    )
                    omitidos_centro += 1
                    resultado.agregar_omitido()
                    continue
            
            # ========== CANTIDAD INICIAL (requerido) ==========
            cant_raw = get_val('cantidad_inicial', '0')
            try:
                cantidad_inicial = int(float(cant_raw))
            except:
                resultado.agregar_error(fila_num, 'cantidad', f'Cantidad inválida: {cant_raw}')
                continue
            
            # ISS-IMPORT-TOLERANTE: Filas con cantidad 0 se OMITEN (no error)
            # Esto permite importar archivos donde algunas filas aún no tienen stock
            if cantidad_inicial <= 0:
                cant_contrato_preview = get_val('cantidad_contrato')
                if cant_contrato_preview and int(float(cant_contrato_preview)) > 0:
                    logger.info(
                        f"Fila {fila_num}: OMITIDA - Lote {numero_lote} con cantidad 0 "
                        f"(contrato pendiente: {cant_contrato_preview}). "
                        f"Importe cuando llegue la primera entrega."
                    )
                else:
                    logger.info(
                        f"Fila {fila_num}: OMITIDA - Lote {numero_lote} con cantidad 0. "
                        f"No se puede registrar un lote sin unidades recibidas."
                    )
                resultado.agregar_omitido()
                continue
            
            # ========== FECHA CADUCIDAD (requerido) ==========
            idx_cad = col_map['caducidad']
            fecha_cad_raw = fila[idx_cad].value if idx_cad < len(fila) else None
            
            try:
                fecha_caducidad = _parse_fecha_excel(fecha_cad_raw, 'Fecha Caducidad')
                if not fecha_caducidad:
                    raise ValueError('Fecha vacía')
                
                # VALIDACIÓN: Fechas de caducidad no pueden estar más de 8 años en el futuro
                # Usa relativedelta para considerar años bisiestos correctamente
                from dateutil.relativedelta import relativedelta
                fecha_actual = date.today()
                fecha_maxima = fecha_actual + relativedelta(years=8)
                
                if fecha_caducidad > fecha_maxima:
                    resultado.agregar_error(fila_num, 'caducidad', 
                        f'Fecha de caducidad muy lejana ({fecha_caducidad.strftime("%d/%m/%Y")}). '
                        f'Máximo permitido: 8 años desde hoy ({fecha_maxima.strftime("%d/%m/%Y")}). '
                        f'Verifique que el formato sea correcto (DD/MM/AAAA).')
                    continue
                
            except Exception as e:
                resultado.agregar_error(fila_num, 'caducidad', f'Fecha inválida: {e}')
                continue
            
            # ========== CAMPOS OPCIONALES ==========
            numero_contrato = get_val('contrato')
            marca = get_val('marca')
            
            # ISS-INV-002: cantidad_contrato es OPCIONAL.
            # Si la columna no existe o el valor es vacío, se guarda como NULL.
            cantidad_contrato = None
            if 'cantidad_contrato' in col_map:
                cant_contrato_raw = get_val('cantidad_contrato')
                if cant_contrato_raw:
                    try:
                        cantidad_contrato = int(float(cant_contrato_raw))
                        if cantidad_contrato < 0:
                            cantidad_contrato = None  # No negativos
                    except:
                        pass  # Si no se puede parsear, se deja como NULL

            # cantidad_contrato_global: total contratado para TODA la clave de producto
            cantidad_contrato_global = None
            if 'cantidad_contrato_global' in col_map:
                cant_global_raw = get_val('cantidad_contrato_global')
                if cant_global_raw:
                    try:
                        cantidad_contrato_global = int(float(cant_global_raw))
                        if cantidad_contrato_global < 0:
                            cantidad_contrato_global = None
                    except:
                        pass
            
            # Fecha de entrega (OPCIONAL - campo informativo para trazabilidad)
            # Nota: Se almacena en 'fecha_fabricacion' por compatibilidad histórica del modelo
            fecha_fabricacion = None
            if 'recepcion' in col_map:
                idx_fab = col_map['recepcion']
                fecha_fab_raw = fila[idx_fab].value if idx_fab < len(fila) else None
                if fecha_fab_raw:
                    try:
                        fecha_fabricacion = _parse_fecha_excel(fecha_fab_raw, 'Fecha Entrega')
                        if fila_num <= 15:  # Log primeras 15 filas para diagnóstico
                            logger.info(f"[FECHA] Fila {fila_num}: fecha_recepcion PARSEADA = {fecha_fabricacion} (raw={type(fecha_fab_raw).__name__}:{fecha_fab_raw})")
                    except Exception as e:
                        # WARNING visible para diagnóstico en producción
                        logger.warning(f"[FECHA-ERROR] Fila {fila_num}: No se pudo parsear fecha entrega '{fecha_fab_raw}' - {e}")
                else:
                    if fila_num <= 15:
                        logger.info(f"[FECHA] Fila {fila_num}: celda de recepción VACÍA (idx={idx_fab})")
            else:
                if fila_num <= 5:  # Log solo primeras filas
                    logger.warning(f"[FECHA] Fila {fila_num}: columna 'recepcion' NO MAPEADA en este Excel")
            
            # Precio
            precio_raw = get_val('precio', '0')
            try:
                precio_str = str(precio_raw).replace(',', '.').replace('$', '').replace(' ', '')
                precio_unitario = max(Decimal('0'), Decimal(precio_str))
            except:
                precio_unitario = Decimal('0')
            
            # Activo
            activo = True
            if 'activo' in col_map:
                activo_raw = get_val('activo', 'activo')
                activo = _parse_bool(activo_raw)
            
            # Acumular fila parseada para consolidación
            filas_parseadas.append({
                'fila_num': fila_num,
                'producto': producto,
                'producto_id': producto.pk,
                'clave_producto': clave_producto,
                'lote_base': numero_lote,       # Número de lote original del Excel
                'numero_lote': numero_lote,      # Puede cambiar tras consolidación con sufijo
                'cantidad_inicial': cantidad_inicial,
                'cantidad_contrato': cantidad_contrato,
                'cantidad_contrato_global': cantidad_contrato_global,
                'fecha_caducidad': fecha_caducidad,
                'fecha_fabricacion': fecha_fabricacion,
                'precio_unitario': precio_unitario,
                'numero_contrato': numero_contrato,
                'marca': marca,
                'ubicacion_excel': ubicacion_excel,
                'activo': activo,
            })
            
        except Exception as exc:
            logger.exception(f"Error parseando fila {fila_num}: {exc}")
            resultado.agregar_error(fila_num, 'general', str(exc))
    
    # ========================================================================
    # FASE 2: CONSOLIDACIÓN - Agrupar parcialidades y diferenciar caducidades
    # ========================================================================
    filas_consolidadas = _consolidar_filas_importacion(filas_parseadas)
    
    total_pre_consolidacion = len(filas_parseadas)
    total_post_consolidacion = len(filas_consolidadas)
    if total_pre_consolidacion != total_post_consolidacion:
        consolidados_archivo = total_pre_consolidacion - total_post_consolidacion
        logger.info(
            f"ISS-IMPORT-CONSOLIDATION: {total_pre_consolidacion} filas → "
            f"{total_post_consolidacion} lotes consolidados "
            f"({consolidados_archivo} filas combinadas por parcialidades)"
        )
    
    # Verificar sufijos contra BD existente para evitar colisiones
    for fila in filas_consolidadas:
        numero_lote = fila['numero_lote']
        producto = fila['producto']
        
        # Si el lote ya tiene sufijo asignado por consolidación, verificar que no colisione con BD
        if numero_lote != fila['lote_base']:
            existente_bd = Lote.objects.filter(
                producto=producto,
                numero_lote__iexact=numero_lote,
                activo=True
            )
            if centro:
                existente_bd = existente_bd.filter(centro=centro)
            else:
                existente_bd = existente_bd.filter(centro__isnull=True)
            
            if existente_bd.exists():
                # Colisión: buscar siguiente sufijo disponible
                sufijo = 2
                while True:
                    candidato = f"{fila['lote_base']}.{sufijo}"
                    existe = Lote.objects.filter(
                        producto=producto,
                        numero_lote__iexact=candidato,
                        activo=True
                    )
                    if centro:
                        existe = existe.filter(centro=centro)
                    else:
                        existe = existe.filter(centro__isnull=True)
                    
                    if not existe.exists():
                        fila['numero_lote'] = candidato
                        break
                    sufijo += 1
                    if sufijo > 100:  # Seguridad contra loops infinitos
                        break
    
    # ========================================================================
    # FASE 3: PERSISTENCIA - Crear/actualizar lotes en BD (atómico)
    # ========================================================================
    with transaction.atomic():
        # BLOQUEO DURO CCG: Deshabilitado - permitir importación sin validar CCG
        # El CCG se puede ajustar manualmente en el Excel según necesidades operativas
        # _validar_ccg_antes_de_importar(filas_consolidadas, centro)

        for fila in filas_consolidadas:
            try:
                producto = fila['producto']
                numero_lote = fila['numero_lote']
                cantidad_inicial = fila['cantidad_inicial']
                cantidad_contrato = fila['cantidad_contrato']
                fecha_caducidad = fila['fecha_caducidad']
                numero_contrato = fila.get('numero_contrato')
                marca = fila.get('marca')
                
                # ========== VERIFICAR DUPLICADO EN BD Y CONSOLIDAR ==========
                lote_existente_activo = Lote.objects.filter(
                    producto=producto, 
                    numero_lote__iexact=numero_lote,
                    activo=True
                )
                
                if centro:
                    lote_existente_activo = lote_existente_activo.filter(centro=centro)
                else:
                    lote_existente_activo = lote_existente_activo.filter(centro__isnull=True)
                
                if lote_existente_activo.exists():
                    lote = lote_existente_activo.first()
                    
                    contrato_igual = (
                        (not numero_contrato and not lote.numero_contrato) or
                        (numero_contrato and lote.numero_contrato and 
                         str(numero_contrato).strip().upper() == str(lote.numero_contrato).strip().upper())
                    )
                    marca_igual = (
                        (not marca and not lote.marca) or
                        (marca and lote.marca and 
                         str(marca).strip().upper() == str(lote.marca).strip().upper())
                    )
                    fecha_igual = (lote.fecha_caducidad == fecha_caducidad)
                    
                    if contrato_igual and marca_igual and fecha_igual:
                        # CONSOLIDAR: Sumar cantidades al lote existente (atómico con F())
                        cantidad_inicial_anterior = lote.cantidad_inicial
                        cantidad_contrato_lote = lote.cantidad_contrato
                        
                        # Obtener fecha de entrega del Excel para la parcialidad y el lote
                        nueva_fecha_fab = fila.get('fecha_fabricacion')
                        logger.info(f"[LOTE-CONSOLIDAR] {lote.numero_lote}: fecha del Excel={nueva_fecha_fab}, fecha actual en BD={lote.fecha_fabricacion}")
                        
                        # Preparar datos de actualización
                        update_data = {
                            'cantidad_actual': F('cantidad_actual') + cantidad_inicial,
                            'cantidad_inicial': F('cantidad_inicial') + cantidad_inicial,
                        }
                        
                        # SIEMPRE actualizar fecha_fabricacion si viene en el Excel
                        # O si el lote existente no tiene fecha y ahora sí viene
                        tiene_fecha_valida = nueva_fecha_fab is not None
                        if tiene_fecha_valida:
                            update_data['fecha_fabricacion'] = nueva_fecha_fab
                            logger.info(f"[LOTE-CONSOLIDAR] Actualizando fecha_fabricacion de {lote.numero_lote}: {lote.fecha_fabricacion} -> {nueva_fecha_fab}")
                        elif lote.fecha_fabricacion is None:
                            # Si no viene fecha del Excel pero el lote no tiene fecha,
                            # usar la fecha actual como fallback
                            from django.utils import timezone
                            update_data['fecha_fabricacion'] = timezone.now().date()
                            logger.warning(f"[LOTE-CONSOLIDAR] {lote.numero_lote}: Sin fecha en Excel, usando fecha actual como fallback")
                        
                        Lote.objects.filter(pk=lote.pk).update(**update_data)
                        lote.refresh_from_db()
                        
                        # P0-1: IDEMPOTENCIA - Verificar fingerprint antes de procesar parcialidad
                        from core.utils.parcialidad_merge import (
                            merge_or_create_parcialidad,
                            calcular_row_fingerprint,
                            verificar_fingerprint_existente,
                            registrar_fingerprint,
                        )
                        from django.utils import timezone
                        
                        # Calcular fingerprint único para esta fila
                        row_fingerprint = calcular_row_fingerprint(
                            file_checksum=file_checksum,
                            row_number=fila.get('fila_num'),
                            lote_id=lote.pk,
                            clave_producto=fila.get('clave_producto'),
                            proveedor=fila.get('proveedor'),
                            factura=fila.get('numero_factura'),
                            fecha_entrega_raw=str(fila.get('fecha_fabricacion_raw', fila.get('fecha_fabricacion', ''))),
                            cantidad=cantidad_inicial,
                        )
                        
                        # Verificar si esta fila ya fue procesada (reimportación)
                        fp_existe, fp_registro = verificar_fingerprint_existente(row_fingerprint)
                        if fp_existe:
                            logger.info(
                                f"[IMPORT-SKIP] Fila {fila.get('fila_num')} ya importada previamente "
                                f"(fingerprint: {row_fingerprint[:16]}...). Saltando."
                            )
                            # No contar como error ni éxito - es un skip válido
                            continue
                        
                        fecha_parcialidad = nueva_fecha_fab or timezone.now().date()
                        nota_parcialidad = 'Entrega via importación Excel (consolidación)'
                        
                        # P0-2: Si no hay fecha válida, NO permitir merge (evita falsos positivos masivos)
                        allow_merge = tiene_fecha_valida
                        if not allow_merge:
                            nota_parcialidad += f' | Fila {fila.get("fila_num")} sin fecha original'
                        
                        # Usar merge_or_create para evitar duplicados y consolidar entregas equivalentes
                        try:
                            resultado_merge = merge_or_create_parcialidad(
                                lote=lote,
                                fecha_entrega=fecha_parcialidad,
                                cantidad=cantidad_inicial,
                                usuario=usuario,
                                notas=nota_parcialidad,
                                archivo_nombre=archivo_nombre,
                                fila_num=fila.get('fila_num'),
                                allow_merge=allow_merge,  # P0-2: False si no hay fecha válida
                            )
                            parcialidad = resultado_merge['parcialidad']
                            fue_merge = resultado_merge['merged']
                            
                            # P0-1: Registrar fingerprint después de éxito
                            registrar_fingerprint(
                                fingerprint=row_fingerprint,
                                lote=lote,
                                parcialidad=parcialidad,
                                file_checksum=file_checksum,
                                row_number=fila.get('fila_num'),
                                archivo_nombre=archivo_nombre,
                                usuario=usuario,
                                action_taken='MERGED' if fue_merge else 'CREATED',
                                cantidad=cantidad_inicial,
                            )
                            
                            if fue_merge:
                                logger.info(
                                    f"[PARCIALIDAD-IMPORT] MERGE para lote {numero_lote}: "
                                    f"+{cantidad_inicial} uds, nivel={resultado_merge.get('nivel_match')}, "
                                    f"total={parcialidad.cantidad}"
                                )
                            else:
                                logger.info(
                                    f"[PARCIALIDAD-IMPORT] Nueva entrega para lote {numero_lote}: "
                                    f"+{cantidad_inicial} uds, fecha={fecha_parcialidad}, "
                                    f"allow_merge={allow_merge}"
                                )
                        except Exception as parcialidad_error:
                            logger.warning(
                                f"[PARCIALIDAD-IMPORT] Error creando parcialidad para lote {numero_lote}: "
                                f"{parcialidad_error}"
                            )
                        
                        Producto.objects.filter(pk=producto.pk).update(
                            stock_actual=F('stock_actual') + cantidad_inicial
                        )
                        
                        pendiente_str = ""
                        if cantidad_contrato_lote:
                            pendiente = cantidad_contrato_lote - lote.cantidad_inicial
                            pendiente_str = f", contrato: {cantidad_contrato_lote}, pendiente: {pendiente}"
                        
                        filas_origen = fila.get('filas_origen', [fila.get('fila_num', '?')])
                        logger.info(
                            f"Filas {filas_origen}: CONSOLIDADO CON BD lote {numero_lote} "
                            f"producto {fila['clave_producto']} - "
                            f"surtido anterior: {cantidad_inicial_anterior}, sumado: +{cantidad_inicial}, "
                            f"nuevo surtido: {lote.cantidad_inicial}, stock: {lote.cantidad_actual}{pendiente_str}"
                        )
                        
                        creados += 1
                        resultado.agregar_exito(es_actualizacion=True)  # Es actualización/consolidación
                        continue
                    else:
                        diferencias = []
                        if not fecha_igual:
                            diferencias.append(f"caducidad BD:{lote.fecha_caducidad} vs Excel:{fecha_caducidad}")
                        if not contrato_igual:
                            diferencias.append(f"contrato BD:'{lote.numero_contrato}' vs Excel:'{numero_contrato}'")
                        if not marca_igual:
                            diferencias.append(f"marca BD:'{lote.marca}' vs Excel:'{marca}'")
                        
                        resultado.agregar_error(fila.get('fila_num', 0), 'lote', 
                            f'Lote {numero_lote} ya existe para producto {fila["clave_producto"]} '
                            f'pero con datos diferentes: {", ".join(diferencias)}. '
                            f'Corrija el Excel o use otro número de lote.')
                        continue
                
                # Si existe inactivo, permitir crear nuevo
                lote_existente_inactivo = Lote.objects.filter(
                    producto=producto, 
                    numero_lote__iexact=numero_lote,
                    activo=False
                )
                if centro:
                    lote_existente_inactivo = lote_existente_inactivo.filter(centro=centro)
                
                # Obtener fecha de entrega para guardar en el lote y la parcialidad
                fecha_recepcion = fila.get('fecha_fabricacion')
                
                # Log de diagnóstico antes de crear el lote
                logger.info(f"[LOTE-CREAR] {numero_lote}: fecha_recepcion={fecha_recepcion} (type={type(fecha_recepcion).__name__})")
                
                # Si no hay fecha de entrega del Excel, usar fecha actual
                # Esto asegura que SIEMPRE haya fecha tanto en el lote como en la parcialidad
                from django.utils import timezone
                if fecha_recepcion is None:
                    fecha_recepcion = timezone.now().date()
                    logger.warning(f"[LOTE-CREAR] {numero_lote}: Sin fecha en Excel, usando fecha actual: {fecha_recepcion}")
                
                # Crear lote
                lote_nuevo = Lote.objects.create(
                    producto=producto,
                    centro=centro,
                    numero_lote=numero_lote,
                    cantidad_inicial=cantidad_inicial,
                    cantidad_actual=cantidad_inicial,
                    cantidad_contrato=cantidad_contrato,  # NULL si no se proporcionó
                    cantidad_contrato_global=fila.get('cantidad_contrato_global'),
                    fecha_caducidad=fecha_caducidad,
                    fecha_fabricacion=fecha_recepcion,  # Mantener por compatibilidad
                    precio_unitario=fila.get('precio_unitario', Decimal('0')),
                    numero_contrato=numero_contrato,
                    marca=marca,
                    ubicacion='Almacén Central',
                    activo=fila.get('activo', True),
                )

                # Registrar el usuario importador en created_by_id (columna existe en DB
                # pero no está declarada en el modelo Django managed=False)
                if usuario and getattr(usuario, 'is_authenticated', False) and getattr(usuario, 'pk', None):
                    try:
                        with connection.cursor() as cur:
                            cur.execute(
                                'UPDATE lotes SET created_by_id = %s WHERE id = %s AND created_by_id IS NULL',
                                [usuario.pk, lote_nuevo.pk],
                            )
                    except Exception:
                        pass  # No bloquear la importación si este UPDATE falla

                # Crear parcialidad inicial para el historial de entregas
                # P0-1: IDEMPOTENCIA - Verificar fingerprint antes de procesar
                from core.utils.parcialidad_merge import (
                    merge_or_create_parcialidad,
                    calcular_row_fingerprint,
                    verificar_fingerprint_existente,
                    registrar_fingerprint,
                )
                
                # Determinar si hay fecha válida para permitir merge
                tiene_fecha_valida_nuevo = fila.get('fecha_fabricacion') is not None
                
                # Calcular fingerprint único para esta fila (usando numero_lote porque el lote es nuevo)
                row_fingerprint_nuevo = calcular_row_fingerprint(
                    file_checksum=file_checksum,
                    row_number=fila.get('fila_num'),
                    numero_lote=numero_lote,  # Usar numero_lote porque aún no tenemos ID
                    clave_producto=fila.get('clave_producto'),
                    proveedor=fila.get('proveedor'),
                    factura=fila.get('numero_factura'),
                    fecha_entrega_raw=str(fila.get('fecha_fabricacion_raw', fila.get('fecha_fabricacion', ''))),
                    cantidad=cantidad_inicial,
                )
                
                # Verificar si esta fila ya fue procesada (reimportación)
                fp_existe_nuevo, fp_registro_nuevo = verificar_fingerprint_existente(row_fingerprint_nuevo)
                if fp_existe_nuevo:
                    logger.info(
                        f"[IMPORT-SKIP-NEW] Fila {fila.get('fila_num')} (lote {numero_lote}) ya importada "
                        f"previamente (fingerprint: {row_fingerprint_nuevo[:16]}...). Saltando creación."
                    )
                    # No crear el lote ni la parcialidad - ya fue procesado
                    continue
                
                # La fecha ya está garantizada (del Excel o fallback a fecha actual)
                nota_inicial = 'Carga inicial por importación Excel'
                
                # P0-2: Si no hay fecha válida, NO permitir merge
                allow_merge_nuevo = tiene_fecha_valida_nuevo
                if not allow_merge_nuevo:
                    nota_inicial += f' | Fila {fila.get("fila_num")} sin fecha original'
                
                # Usar merge_or_create para consistencia con la lógica de consolidación
                try:
                    resultado_merge_nuevo = merge_or_create_parcialidad(
                        lote=lote_nuevo,
                        fecha_entrega=fecha_recepcion,
                        cantidad=cantidad_inicial,
                        usuario=usuario,
                        notas=nota_inicial,
                        archivo_nombre=archivo_nombre,
                        fila_num=fila.get('fila_num'),
                        allow_merge=allow_merge_nuevo,
                    )
                    parcialidad = resultado_merge_nuevo['parcialidad']
                    
                    # Registrar fingerprint después de éxito
                    registrar_fingerprint(
                        fingerprint=row_fingerprint_nuevo,
                        lote=lote_nuevo,
                        parcialidad=parcialidad,
                        file_checksum=file_checksum,
                        row_number=fila.get('fila_num'),
                        archivo_nombre=archivo_nombre,
                        usuario=usuario,
                        action_taken='CREATED',
                        cantidad=cantidad_inicial,
                    )
                    
                    logger.info(
                        f"[PARCIALIDAD] Creada para lote nuevo {numero_lote}: "
                        f"{cantidad_inicial} uds, fecha={fecha_recepcion}"
                    )
                except Exception as parcialidad_error:
                    logger.warning(
                        f"[PARCIALIDAD] Error creando parcialidad para lote nuevo {numero_lote}: "
                        f"{parcialidad_error}"
                    )
                
                Producto.objects.filter(pk=producto.pk).update(
                    stock_actual=F('stock_actual') + cantidad_inicial
                )
                
                creados += 1
                resultado.agregar_exito()
                
            except Exception as exc:
                logger.exception(f"Error creando lote {fila.get('numero_lote', '?')}: {exc}")
                resultado.agregar_error(fila.get('fila_num', 0), 'general', str(exc))

        # Auto-propagar cantidad_contrato_global: para cada producto+contrato,
        # si algún lote tiene ccg definido, propagarlo a todos los lotes del mismo grupo.
        _propagar_contrato_global(filas_consolidadas, centro)

    result = resultado.get_dict()
    result['creados'] = creados
    # DEBUG: Información de columnas detectadas
    result['_debug_col_map'] = {k: v for k, v in col_map.items()}
    result['_debug_encabezados'] = encabezados[:15]  # Solo los primeros 15
    result['_debug_recepcion_detectada'] = 'recepcion' in col_map
    # ADVERTENCIA si no se detectó columna de fecha de entrega
    if 'recepcion' not in col_map:
        result['advertencia_fecha'] = (
            'ADVERTENCIA: No se detectó columna de "Fecha Entrega" en el Excel. '
            'Los lotes se crearán con la fecha actual. '
            'Encabezados detectados: ' + ', '.join(encabezados[:10])
        )
    if consolidados_archivo > 0:
        result['consolidados_archivo'] = consolidados_archivo
        result['nota_consolidacion'] = (
            f'{consolidados_archivo} fila(s) consolidada(s) en el archivo '
            f'(parcialidades con mismo lote+producto+caducidad se sumaron).'
        )
    # ISS-FIX: Informar sobre filas omitidas por pertenecer a otros centros
    if omitidos_centro > 0:
        result['omitidos_centro'] = omitidos_centro
        result['nota_centros'] = (
            f'{omitidos_centro} fila(s) omitida(s) por pertenecer a centros penitenciarios. '
            f'Los lotes de centros se gestionan mediante transferencias, no por importación directa.'
        )
    return result


def crear_log_importacion(usuario, tipo, archivo_nombre, resultado_dict):
    """
    Crea registro de ImportacionLog con los campos correctos de la BD.
    
    Campos de la tabla importacion_logs:
    - archivo, tipo_importacion, usuario_id, registros_totales,
    - registros_exitosos, registros_fallidos, errores (jsonb),
    - estado, fecha_inicio, fecha_fin
    """
    try:
        ImportacionLog.objects.create(
            usuario=usuario if usuario and getattr(usuario, 'is_authenticated', False) else None,
            archivo=archivo_nombre[:255],  # Max 255 chars
            tipo_importacion=tipo[:50],  # Max 50 chars
            registros_totales=resultado_dict.get('total_registros', 0),
            registros_exitosos=resultado_dict.get('registros_exitosos', 0),
            registros_fallidos=resultado_dict.get('registros_fallidos', 0),
            errores=resultado_dict.get('errores', []),  # Campo JSONField
            estado='procesando' if resultado_dict.get('exitosa') else ('parcial' if resultado_dict.get('registros_exitosos', 0) > 0 else 'fallida'),
            fecha_fin=timezone.now(),  # Marcar como completado
        )
    except Exception as exc:
        logger.error(f"Error creando ImportacionLog: {exc}")


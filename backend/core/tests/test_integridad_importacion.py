# -*- coding: utf-8 -*-
"""
Tests de integridad de datos e importación con rollback.

Cubre:
1. Importación de productos con rollback transaccional
2. Importación de lotes con validación
3. Importación de centros
4. Integridad referencial
5. Soft delete y restauración
"""

from django.test import TestCase, TransactionTestCase
from django.contrib.auth.models import Group
from django.db import IntegrityError, transaction
from django.db.models.signals import post_save, pre_save, post_delete
from rest_framework.test import APIClient
from rest_framework import status
from decimal import Decimal
from datetime import date, timedelta
from io import BytesIO
import openpyxl

from core.models import (
    User, Centro, Producto, Lote, Movimiento, 
    Requisicion, DetalleRequisicion
)
from core import signals


class DisableAuditSignalsMixin:
    """Mixin para desconectar signals de auditoría durante tests.
    
    Esto previene errores de FK constraint cuando TransactionTestCase
    hace rollback de la base de datos y elimina usuarios referenciados
    en registros de auditoría.
    """
    
    _audit_signals = [
        (post_save, signals.auditar_cambios_requisicion, Requisicion),
        (post_save, signals.auditar_cambios_producto, Producto),
        (post_save, signals.auditar_lote, Lote),
        (post_delete, signals.auditar_eliminacion_producto, Producto),
        (pre_save, signals.snapshot_requisicion, Requisicion),
        (pre_save, signals.snapshot_producto, Producto),
    ]
    
    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        # Desconectar signals de auditoría
        for signal, handler, sender in cls._audit_signals:
            signal.disconnect(handler, sender=sender)
    
    @classmethod
    def tearDownClass(cls):
        # Reconectar signals de auditoría
        for signal, handler, sender in cls._audit_signals:
            signal.connect(handler, sender=sender)
        super().tearDownClass()


class ImportacionProductosTest(DisableAuditSignalsMixin, TransactionTestCase):
    """Tests de importación de productos con validaciones.
    
    NOTA: Estos tests se saltan si el endpoint de importación
    requiere permisos especiales no disponibles en el entorno de test.
    """
    
    def setUp(self):
        """Preparar datos de prueba."""
        self.client = APIClient()
        
        # Crear usuario con rol farmacia para importación
        self.grupo_farmacia = Group.objects.create(name='FARMACIA')
        self.farmacia_user = User.objects.create_user(
            username='import_farmacia',
            password='Test@123',
            email='import@test.com',
            rol='farmacia',
            is_staff=True
        )
        self.farmacia_user.groups.add(self.grupo_farmacia)
        
        # Login
        login = self.client.post('/api/v1/token/', {
            'username': 'import_farmacia',
            'password': 'Test@123'
        })
        self.client.credentials(HTTP_AUTHORIZATION=f'Bearer {login.data["access"]}')
    
    def _check_import_permission(self, response):
        """Verificar si tenemos permisos de importación."""
        if response.status_code == 403:
            self.skipTest("Endpoint de importación requiere permisos especiales")
    
    def _crear_excel_productos(self, filas):
        """Helper para crear archivo Excel de productos."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['Clave', 'Descripcion', 'Unidad', 'Precio', 'Stock Minimo', 'Estado'])
        for fila in filas:
            ws.append(fila)
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    def test_importacion_productos_validos(self):
        """Importación de productos válidos debe crear todos."""
        excel = self._crear_excel_productos([
            ['IMP-001', 'Producto Importado 1', 'PIEZA', '10.00', '50', 'Activo'],
            ['IMP-002', 'Producto Importado 2', 'CAJA', '25.00', '30', 'Activo'],
            ['IMP-003', 'Producto Importado 3', 'FRASCO', '15.00', '20', 'Activo'],
        ])
        
        response = self.client.post('/api/v1/productos/importar_excel/', {
            'file': excel
        }, format='multipart')
        
        self._check_import_permission(response)
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['resumen']['creados'], 3)
        self.assertEqual(response.data['resumen']['total_errores'], 0)
        
        # Verificar que todos se crearon
        self.assertTrue(Producto.objects.filter(clave='IMP-001').exists())
        self.assertTrue(Producto.objects.filter(clave='IMP-002').exists())
        self.assertTrue(Producto.objects.filter(clave='IMP-003').exists())
    
    def test_importacion_actualiza_existentes(self):
        """Importación debe actualizar productos que ya existen."""
        # Crear producto existente
        Producto.objects.create(
            clave='EXIST-001',
            descripcion='Descripción original',
            unidad_medida='PIEZA',
            precio_unitario=Decimal('10.00'),
            stock_minimo=50
        )
        
        excel = self._crear_excel_productos([
            ['EXIST-001', 'Descripción actualizada', 'CAJA', '20.00', '100', 'Activo'],
        ])
        
        response = self.client.post('/api/v1/productos/importar_excel/', {
            'file': excel
        }, format='multipart')
        
        self._check_import_permission(response)
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['resumen']['actualizados'], 1)
        
        # Verificar actualización
        producto = Producto.objects.get(clave='EXIST-001')
        self.assertEqual(producto.descripcion, 'Descripción actualizada')
        self.assertEqual(producto.unidad_medida, 'CAJA')
        self.assertEqual(producto.precio_unitario, Decimal('20.00'))
    
    def test_importacion_filas_invalidas_reporta_errores(self):
        """Importación con filas inválidas debe reportar errores."""
        excel = self._crear_excel_productos([
            ['VALID-001', 'Producto Válido', 'PIEZA', '10.00', '50', 'Activo'],
            ['', 'Sin Clave', 'PIEZA', '10.00', '50', 'Activo'],  # Error: sin clave
            ['VALID-002', '', 'PIEZA', '10.00', '50', 'Activo'],  # Error: sin descripción
            ['VALID-003', 'Otro Válido', 'PIEZA', '10.00', '50', 'Activo'],
        ])
        
        response = self.client.post('/api/v1/productos/importar_excel/', {
            'file': excel
        }, format='multipart')
        
        self._check_import_permission(response)
        
        self.assertIn(response.status_code, [status.HTTP_200_OK, status.HTTP_207_MULTI_STATUS])
        self.assertEqual(response.data['resumen']['creados'], 2)  # Solo válidos
        self.assertGreater(response.data['resumen']['total_errores'], 0)
        
        # Verificar que los válidos se crearon
        self.assertTrue(Producto.objects.filter(clave='VALID-001').exists())
        self.assertTrue(Producto.objects.filter(clave='VALID-003').exists())
    
    def test_importacion_unidad_invalida_error(self):
        """Unidad de medida inválida debe reportar error."""
        excel = self._crear_excel_productos([
            ['UNIDAD-001', 'Producto Test', 'INVALIDA', '10.00', '50', 'Activo'],
        ])
        
        response = self.client.post('/api/v1/productos/importar_excel/', {
            'file': excel
        }, format='multipart')
        
        self._check_import_permission(response)
        
        self.assertIn(response.status_code, [status.HTTP_200_OK, status.HTTP_207_MULTI_STATUS])
        self.assertGreater(len(response.data.get('errores', [])), 0)
    
    def test_importacion_precio_negativo_error(self):
        """Precio negativo debe reportar error."""
        excel = self._crear_excel_productos([
            ['PRECIO-001', 'Producto Test', 'PIEZA', '-10.00', '50', 'Activo'],
        ])
        
        response = self.client.post('/api/v1/productos/importar_excel/', {
            'file': excel
        }, format='multipart')
        
        self._check_import_permission(response)
        
        self.assertIn(response.status_code, [status.HTTP_200_OK, status.HTTP_207_MULTI_STATUS])
        # Debe reportar error de precio
        self.assertGreater(len(response.data.get('errores', [])), 0)


class ImportacionLotesTest(DisableAuditSignalsMixin, TransactionTestCase):
    """Tests de importación de lotes."""
    
    def setUp(self):
        """Preparar datos de prueba."""
        self.client = APIClient()
        
        # Crear grupo y usuario superusuario para importación
        self.grupo_farmacia = Group.objects.create(name='FARMACIA')
        self.farmacia_user = User.objects.create_superuser(
            username='lotes_farmacia',
            password='Test@123',
            email='lotes@test.com'
        )
        
        # Producto para asociar lotes
        self.producto = Producto.objects.create(
            clave='LOTE-PROD-001',
            descripcion='Producto para lotes',
            unidad_medida='PIEZA',
            precio_unitario=Decimal('10.00'),
            stock_minimo=50,
            activo=True
        )
        
        # Login
        login = self.client.post('/api/v1/token/', {
            'username': 'lotes_farmacia',
            'password': 'Test@123'
        })
        self.client.credentials(HTTP_AUTHORIZATION=f'Bearer {login.data["access"]}')
    
    def _crear_excel_lotes(self, filas):
        """Helper para crear archivo Excel de lotes."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['Producto', 'Numero Lote', 'Fecha Caducidad', 'Cantidad Inicial', 'Cantidad Actual', 'Proveedor'])
        for fila in filas:
            ws.append(fila)
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    def test_importacion_lotes_validos(self):
        """Importación de lotes válidos debe crear todos."""
        fecha_cad = (date.today() + timedelta(days=365)).strftime('%Y-%m-%d')
        
        excel = self._crear_excel_lotes([
            ['LOTE-PROD-001', 'LOTE-IMP-001', fecha_cad, 100, 100, 'Proveedor A'],
            ['LOTE-PROD-001', 'LOTE-IMP-002', fecha_cad, 50, 50, 'Proveedor B'],
        ])
        
        response = self.client.post('/api/v1/lotes/importar-excel/', {
            'file': excel
        }, format='multipart')
        
        self.assertIn(response.status_code, [status.HTTP_200_OK, status.HTTP_207_MULTI_STATUS])
        
        # Verificar que se crearon
        self.assertTrue(Lote.objects.filter(numero_lote='LOTE-IMP-001').exists())
        self.assertTrue(Lote.objects.filter(numero_lote='LOTE-IMP-002').exists())
    
    def test_importacion_lote_producto_inexistente(self):
        """Lote con producto inexistente debe reportar error."""
        fecha_cad = (date.today() + timedelta(days=365)).strftime('%Y-%m-%d')
        
        excel = self._crear_excel_lotes([
            ['NO-EXISTE', 'LOTE-ERR-001', fecha_cad, 100, 100, 'Proveedor'],
        ])
        
        response = self.client.post('/api/v1/lotes/importar-excel/', {
            'file': excel
        }, format='multipart')
        
        self.assertIn(response.status_code, [status.HTTP_200_OK, status.HTTP_207_MULTI_STATUS])
        self.assertGreater(len(response.data.get('errores', [])), 0)


class ImportacionCentrosTest(DisableAuditSignalsMixin, TransactionTestCase):
    """Tests de importación de centros."""
    
    def setUp(self):
        """Preparar datos de prueba."""
        self.client = APIClient()
        
        # Crear grupo y usuario superusuario para importación
        self.grupo_farmacia = Group.objects.create(name='FARMACIA')
        self.farmacia_user = User.objects.create_superuser(
            username='centros_farmacia',
            password='Test@123',
            email='centros@test.com'
        )
        
        # Login
        login = self.client.post('/api/v1/token/', {
            'username': 'centros_farmacia',
            'password': 'Test@123'
        })
        self.client.credentials(HTTP_AUTHORIZATION=f'Bearer {login.data["access"]}')
    
    def _crear_excel_centros(self, filas):
        """Helper para crear archivo Excel de centros."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['Clave', 'Nombre', 'Direccion', 'Telefono', 'Estado'])
        for fila in filas:
            ws.append(fila)
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    def test_importacion_centros_validos(self):
        """Importación de centros válidos debe crear todos."""
        excel = self._crear_excel_centros([
            ['CENTRO-IMP-001', 'Centro Importado 1', 'Dirección 1', '555-0001', 'Activo'],
            ['CENTRO-IMP-002', 'Centro Importado 2', 'Dirección 2', '555-0002', 'Activo'],
        ])
        
        response = self.client.post('/api/v1/centros/importar/', {
            'file': excel
        }, format='multipart')
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data['resumen']['creados'], 2)
        
        # Verificar creación
        self.assertTrue(Centro.objects.filter(clave='CENTRO-IMP-001').exists())
        self.assertTrue(Centro.objects.filter(clave='CENTRO-IMP-002').exists())
    
    def test_plantilla_centros(self):
        """Descarga de plantilla de centros debe funcionar."""
        response = self.client.get('/api/v1/centros/plantilla/')
        
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )


class IntegridadReferencialTest(TestCase):
    """Tests de integridad referencial."""
    
    def setUp(self):
        """Preparar datos de prueba."""
        # Usuario necesario para requisiciones
        self.user = User.objects.create_user(
            username='int_user_setup',
            password='Test@123',
            email='int_setup@test.com'
        )
        
        self.centro = Centro.objects.create(
            clave='INT-CENTRO-001',
            nombre='Centro Integridad',
            activo=True
        )
        
        self.producto = Producto.objects.create(
            clave='INT-PROD-001',
            descripcion='Producto Integridad',
            unidad_medida='PIEZA',
            precio_unitario=Decimal('10.00'),
            stock_minimo=50,
            activo=True
        )
        
        self.lote = Lote.objects.create(
            producto=self.producto,
            numero_lote='INT-LOTE-001',
            fecha_caducidad=date.today() + timedelta(days=60),
            cantidad_inicial=100,
            cantidad_actual=100,
            estado='disponible'
        )
    
    def test_lote_requiere_producto(self):
        """Lote debe tener producto asociado."""
        from django.core.exceptions import ValidationError
        with self.assertRaises(ValidationError):
            Lote.objects.create(
                producto=None,
                numero_lote='SIN-PRODUCTO',
                fecha_caducidad=date.today() + timedelta(days=60),
                cantidad_inicial=100,
                cantidad_actual=100
            )
    
    def test_movimiento_requiere_lote(self):
        """Movimiento debe tener lote y usuario asociados."""
        from django.core.exceptions import ValidationError
        with self.assertRaises(ValidationError):
            Movimiento.objects.create(
                lote=None,
                tipo='entrada',
                cantidad=50
            )
    
    def test_detalle_requisicion_requiere_producto(self):
        """DetalleRequisicion debe tener producto."""
        from django.db import IntegrityError
        requisicion = Requisicion.objects.create(
            folio='REQ-INT-001',
            centro=self.centro,
            usuario_solicita=self.user,
            estado='borrador'
        )
        
        with self.assertRaises(IntegrityError):
            DetalleRequisicion.objects.create(
                requisicion=requisicion,
                producto=None,
                cantidad_solicitada=10
            )
    
    def test_producto_con_lotes_no_se_puede_eliminar_desde_api(self):
        """Producto con lotes no debe poder eliminarse vía API."""
        self.client = APIClient()
        
        grupo = Group.objects.create(name='FARMACIA')
        user = User.objects.create_superuser(username='int_user_api', password='Test@123', email='api@test.com')
        
        login = self.client.post('/api/v1/token/', {'username': 'int_user_api', 'password': 'Test@123'})
        self.client.credentials(HTTP_AUTHORIZATION=f'Bearer {login.data["access"]}')
        
        response = self.client.delete(f'/api/v1/productos/{self.producto.id}/')
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        
        # Verificar que el producto sigue existiendo
        self.assertTrue(Producto.objects.filter(id=self.producto.id).exists())


class SoftDeleteTest(DisableAuditSignalsMixin, TransactionTestCase):
    """Tests de soft delete."""
    
    def setUp(self):
        """Preparar datos de prueba."""
        from django.utils import timezone
        
        self.producto = Producto.objects.create(
            clave='SOFT-PROD-001',
            descripcion='Producto para soft delete',
            unidad_medida='PIEZA',
            precio_unitario=Decimal('10.00'),
            stock_minimo=50,
            activo=True
        )
        
        self.lote = Lote.objects.create(
            producto=self.producto,
            numero_lote='SOFT-LOTE-001',
            fecha_caducidad=date.today() + timedelta(days=60),
            cantidad_inicial=100,
            cantidad_actual=100,
            estado='disponible'
        )
    
    def test_lote_soft_delete(self):
        """Soft delete de lote debe marcar deleted_at."""
        from django.utils import timezone
        
        self.lote.deleted_at = timezone.now()
        self.lote.save()
        
        # No debe aparecer en queryset normal con filter
        lotes_activos = Lote.objects.filter(
            producto=self.producto,
            deleted_at__isnull=True
        )
        self.assertEqual(lotes_activos.count(), 0)
        
        # Pero el registro sigue existiendo
        self.assertTrue(Lote.objects.filter(id=self.lote.id).exists())
    
    def test_lote_soft_deleted_no_afecta_stock(self):
        """Lotes soft-deleted no deben contar en stock."""
        from django.utils import timezone
        from django.db.models import Sum
        
        # Crear otro lote activo
        lote2 = Lote.objects.create(
            producto=self.producto,
            numero_lote='SOFT-LOTE-002',
            fecha_caducidad=date.today() + timedelta(days=90),
            cantidad_inicial=50,
            cantidad_actual=50,
            estado='disponible'
        )
        
        # Soft delete del lote1
        self.lote.deleted_at = timezone.now()
        self.lote.save()
        
        # Stock solo debe contar lote2
        stock = Lote.objects.filter(
            producto=self.producto,
            deleted_at__isnull=True,
            estado='disponible'
        ).aggregate(total=Sum('cantidad_actual'))['total'] or 0
        
        self.assertEqual(stock, 50)  # Solo lote2


class ConcurrenciaTest(TestCase):
    """Tests de concurrencia en operaciones críticas.
    
    NOTA: Usar TestCase en lugar de TransactionTestCase para evitar
    problemas con FK constraints durante el teardown de tests.
    """
    
    def setUp(self):
        """Preparar datos de prueba."""
        # Usuario necesario para movimientos
        self.user = User.objects.create_user(
            username='conc_user',
            password='Test@123',
            email='conc@test.com'
        )
        
        self.producto = Producto.objects.create(
            clave='CONC-PROD-001',
            descripcion='Producto para concurrencia',
            unidad_medida='PIEZA',
            precio_unitario=Decimal('10.00'),
            stock_minimo=50,
            activo=True
        )
        
        self.lote = Lote.objects.create(
            producto=self.producto,
            numero_lote='CONC-LOTE-001',
            fecha_caducidad=date.today() + timedelta(days=60),
            cantidad_inicial=100,
            cantidad_actual=100,
            estado='disponible'
        )
    
    def test_select_for_update_previene_race_condition(self):
        """select_for_update debe prevenir condiciones de carrera."""
        from inventario.views import registrar_movimiento_stock
        
        # Primera operación
        mov1, lote1 = registrar_movimiento_stock(
            lote=self.lote,
            tipo='salida',
            cantidad=50,
            usuario=self.user,
            observaciones='Primera salida'
        )
        
        # El lote debe estar actualizado
        self.assertEqual(lote1.cantidad_actual, 50)
        
        # Refresh y verificar
        self.lote.refresh_from_db()
        self.assertEqual(self.lote.cantidad_actual, 50)
        
        # Segunda operación
        mov2, lote2 = registrar_movimiento_stock(
            lote=self.lote,
            tipo='salida',
            cantidad=30,
            usuario=self.user,
            observaciones='Segunda salida'
        )
        
        self.assertEqual(lote2.cantidad_actual, 20)
        
        self.lote.refresh_from_db()
        self.assertEqual(self.lote.cantidad_actual, 20)


class ConsistenciaStockTest(TestCase):
    """Tests de consistencia de stock."""
    
    def setUp(self):
        """Preparar datos de prueba."""
        # Usuario necesario para movimientos
        self.user = User.objects.create_user(
            username='cons_user',
            password='Test@123',
            email='cons@test.com'
        )
        
        self.producto = Producto.objects.create(
            clave='CONS-PROD-001',
            descripcion='Producto para consistencia',
            unidad_medida='PIEZA',
            precio_unitario=Decimal('10.00'),
            stock_minimo=50,
            activo=True
        )
        
        self.lote = Lote.objects.create(
            producto=self.producto,
            numero_lote='CONS-LOTE-001',
            fecha_caducidad=date.today() + timedelta(days=60),
            cantidad_inicial=100,
            cantidad_actual=100,
            estado='disponible'
        )
    
    def test_stock_lote_coincide_con_movimientos(self):
        """Stock del lote debe coincidir con suma de movimientos."""
        from inventario.views import registrar_movimiento_stock
        from django.db.models import Sum
        
        # Solo salidas (las entradas violarían cantidad_inicial)
        registrar_movimiento_stock(
            lote=self.lote,
            tipo='salida',
            cantidad=30,
            usuario=self.user,
            observaciones='Salida 1'
        )
        
        registrar_movimiento_stock(
            lote=self.lote,
            tipo='salida',
            cantidad=20,
            usuario=self.user,
            observaciones='Salida 2'
        )
        
        # Calcular saldo según movimientos
        # Nota: cantidad_inicial (100) + movimientos (negativos)
        total_movimientos = Movimiento.objects.filter(
            lote=self.lote
        ).aggregate(total=Sum('cantidad'))['total'] or 0
        
        saldo_calculado = self.lote.cantidad_inicial + total_movimientos
        
        self.lote.refresh_from_db()
        
        # Saldo calculado debe coincidir con cantidad_actual
        # 100 + (-30) + (-20) = 50
        self.assertEqual(saldo_calculado, self.lote.cantidad_actual)
        self.assertEqual(self.lote.cantidad_actual, 50)

"""
ViewSets para la API REST del sistema de farmacia penitenciaria
"""

from rest_framework import viewsets, status, filters, mixins
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.pagination import PageNumberPagination
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.views import APIView
from rest_framework.throttling import AnonRateThrottle, UserRateThrottle
from rest_framework_simplejwt.views import TokenObtainPairView
from rest_framework_simplejwt.tokens import RefreshToken
from django.utils import timezone
from django.http import HttpResponse
from django.db import transaction
from django.db.models import Q, Sum, F
from django.db.models.functions import Coalesce
from django.views.decorators.cache import cache_page
from django.utils.decorators import method_decorator
from datetime import date, timedelta, datetime
from io import BytesIO
import openpyxl
import logging
import requests

from core.models import (
    User, Centro, Producto, Lote, Requisicion, DetalleRequisicion,
    Movimiento, AuditoriaLog, ImportacionLog, Notificacion, UserProfile,
    ConfiguracionSistema,
    # Módulo Compras Caja Chica
    CompraCajaChica, DetalleCompraCajaChica, InventarioCajaChica,
    MovimientoCajaChica, HistorialCompraCajaChica
)
from core.utils.pdf_reports import generar_reporte_auditoria, generar_reporte_trazabilidad
from core.serializers import (
    UserSerializer, CentroSerializer, UserMeSerializer,
    ProductoSerializer, LoteSerializer, RequisicionSerializer,
    DetalleRequisicionSerializer, MovimientoSerializer,
    AuditoriaLogSerializer, ImportacionLogSerializer, NotificacionSerializer,
    ConfiguracionSistemaSerializer,
    # Módulo Compras Caja Chica
    CompraCajaChicaSerializer, CompraCajaChicaListSerializer,
    DetalleCompraCajaChicaSerializer, InventarioCajaChicaSerializer,
    MovimientoCajaChicaSerializer, HistorialCompraCajaChicaSerializer
)
from core.permissions import (
    IsFarmaciaRole, IsCentroUser, CanAuthorizeRequisicion, CanViewNotifications, 
    CanViewProfile, IsVistaRole, IsSuperuserOnly, CanManageDispensaciones,
    CanManageComprasCajaChica
)
# ISS-SEC: Mixin de confirmación obligatoria
from core.mixins import ConfirmationRequiredMixin
from django.conf import settings

logger = logging.getLogger(__name__)


class StandardResultsSetPagination(PageNumberPagination):
    page_size = 50
    page_size_query_param = 'page_size'
    max_page_size = 1000


# ============================================
# THROTTLING - CLASES DE RATE LIMITING
# ============================================

class LoginThrottle(AnonRateThrottle):
    """Rate limit específico para login: 5 intentos por minuto por IP."""
    scope = 'login'


class PasswordChangeThrottle(UserRateThrottle):
    """Rate limit para cambio de contraseña: 3 intentos por minuto."""
    scope = 'password_change'


# ============================================
# ISS-013: JERARQUÍA DE ROLES PARA VALIDACIÓN
# ============================================
# Menor número = mayor privilegio
# Esto previene escalamiento de privilegios
# FLUJO V2: Incluye roles jerárquicos del centro
ROLE_HIERARCHY = {
    # Nivel 0: Superusuario (acceso total)
    'superuser': 0,
    'superusuario': 0,
    
    # Nivel 1: Administradores del sistema
    'admin': 1,
    'admin_sistema': 1,
    
    # Nivel 2: Personal de Farmacia Central
    'farmacia': 2,
    'admin_farmacia': 2,
    
    # Nivel 3: Directivos del Centro (FLUJO V2)
    'director_centro': 3,
    'administrador_centro': 3,
    
    # Nivel 4: Personal operativo del Centro (FLUJO V2)
    'medico': 4,
    
    # Nivel 5: Usuarios de consulta
    'centro': 5,
    'usuario_centro': 5,
    'vista': 5,
    'usuario_vista': 5,
    'usuario_normal': 5,
}


def get_role_level(user, rol=None):
    """
    ISS-013: Obtiene el nivel jerárquico de un usuario o rol.
    Menor nivel = mayor privilegio.
    """
    if user and user.is_superuser:
        return 0
    
    target_rol = rol or (getattr(user, 'rol', '') or '').lower()
    return ROLE_HIERARCHY.get(target_rol, 99)


class UserViewSet(ConfirmationRequiredMixin, viewsets.ModelViewSet):
    """
    ViewSet para gestion de usuarios.
    
    ISS-SEC: Requiere confirmación para operaciones de eliminación
    """
    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = StandardResultsSetPagination
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['username', 'email', 'first_name', 'last_name', 'adscripcion']
    ordering_fields = ['username', 'date_joined']
    ordering = ['-date_joined']
    
    # ISS-SEC: Configuración de confirmación
    require_delete_confirmation = True
    
    def _apply_filters(self, queryset, params):
        """Aplica filtros comunes a queryset de usuarios"""
        # Filtro por rol
        rol = params.get('rol')
        if rol:
            queryset = queryset.filter(rol=rol)
        
        # Filtro por estado activo/inactivo
        is_active = params.get('is_active')
        if is_active is not None:
            if is_active in ['true', 'True', '1', True]:
                queryset = queryset.filter(is_active=True)
            elif is_active in ['false', 'False', '0', False]:
                queryset = queryset.filter(is_active=False)
        
        # Filtro por centro
        centro = params.get('centro')
        if centro:
            queryset = queryset.filter(centro_id=centro)
        
        return queryset

    def get_serializer_class(self):
        if self.action in ['me', 'me_change_password']:
            return UserMeSerializer
        return super().get_serializer_class()

    def get_permissions(self):
        """
        Permisos por acción:
        - me, me_change_password: Cualquier autenticado
        - list, retrieve: Autenticado (queryset filtra según rol)
        - create, update, destroy, cambiar_password: Solo FARMACIA/Admin
        """
        if self.action in ['me', 'me_change_password']:
            return [IsAuthenticated(), CanViewProfile()]
        if self.action in ['create', 'update', 'partial_update', 'destroy', 'cambiar_password', 
                           'exportar_excel', 'importar_excel']:
            return [IsAuthenticated(), IsFarmaciaRole()]
        return [IsAuthenticated()]

    def _is_farmacia_or_admin(self, user):
        """Verifica si el usuario es farmacia o admin"""
        if not user or not user.is_authenticated:
            return False
        if user.is_superuser:
            return True
        rol = (getattr(user, 'rol', '') or '').lower()
        if rol in ['admin_sistema', 'superusuario', 'farmacia', 'admin_farmacia']:
            return True
        group_names = set(g.name.upper() for g in user.groups.all())
        return bool({'FARMACIA_ADMIN', 'FARMACIA', 'ADMIN'} & group_names)

    def get_queryset(self):
        """
        ISS-005 FIX (audit7): Filtra usuarios por centro Y jerarquía de roles.
        
        Un usuario solo puede ver:
        - Superusuario/Admin: Todos los usuarios
        - Farmacia: Usuarios de menor privilegio
        - Usuario de centro: Solo usuarios de su centro con menor o igual privilegio
        """
        user = self.request.user
        if user.is_superuser:
            qs = User.objects.all()
        elif self._is_farmacia_or_admin(user):
            # ISS-005 FIX: Farmacia ve todos excepto superusuarios
            qs = User.objects.filter(is_superuser=False)
        else:
            # Usuario no admin solo ve usuarios de su centro
            if hasattr(user, 'centro') and user.centro:
                qs = User.objects.filter(centro=user.centro)
                
                # ISS-005 FIX (audit7): Filtrar por jerarquía de roles
                # Solo ver usuarios de menor o igual privilegio
                user_level = get_role_level(user)
                
                # Obtener roles que este usuario puede ver (igual o menor privilegio)
                roles_visibles = [rol for rol, nivel in ROLE_HIERARCHY.items() 
                                  if nivel >= user_level]
                
                # Filtrar por roles permitidos
                qs = qs.filter(rol__in=roles_visibles)
                
                logger.debug(
                    f"ISS-005: Usuario {user.username} (nivel {user_level}) "
                    f"puede ver roles: {roles_visibles}"
                )
            else:
                # Sin centro, solo ve a sí mismo
                qs = User.objects.filter(id=user.id)
        
        # Aplicar filtros server-side
        qs = self._apply_filters(qs, self.request.query_params)
        
        return qs.select_related('centro')
    
    def get_object(self):
        """
        HALLAZGO #7 FIX: Prevenir IDOR validando que el objeto solicitado
        esté en el queryset filtrado del usuario.
        """
        obj = super().get_object()
        
        # Verificar que el usuario tiene permiso para ver este objeto específico
        # (ya está validado por get_queryset, pero double-check por seguridad)
        if not self.request.user.is_superuser and not self._is_farmacia_or_admin(self.request.user):
            # Usuario de centro: verificar que sea del mismo centro
            if hasattr(self.request.user, 'centro') and self.request.user.centro:
                if hasattr(obj, 'centro') and obj.centro != self.request.user.centro:
                    from rest_framework.exceptions import PermissionDenied
                    raise PermissionDenied('No tiene permiso para acceder a este usuario.')
        
        return obj

    def _validate_role_hierarchy(self, requesting_user, target_role, target_user=None):
        """
        ISS-013 + ISS-004 FIX: Valida que el usuario que hace la petición tenga suficientes
        privilegios para asignar el rol objetivo.
        
        Reglas:
        - Superusuarios pueden hacer cualquier cosa
        - Un usuario no puede asignar un rol de mayor o igual privilegio que el suyo
        - Un usuario no puede modificar a otro de mayor o igual privilegio
        - ISS-004 FIX: Un usuario NO puede modificar a otro del MISMO rol (excepto superusuarios)
        """
        from rest_framework.exceptions import PermissionDenied
        
        requesting_level = get_role_level(requesting_user)
        target_level = get_role_level(None, target_role)
        
        # Si el rol objetivo tiene mayor privilegio (menor nivel) que el solicitante
        if target_level <= requesting_level and not requesting_user.is_superuser:
            raise PermissionDenied(
                f'No puede asignar el rol "{target_role}". '
                f'Solo usuarios con mayor privilegio pueden asignar este rol.'
            )
        
        # Si estamos modificando un usuario existente, verificar que tengamos privilegio sobre él
        if target_user:
            target_user_level = get_role_level(target_user)
            
            # ISS-004 FIX: No permitir modificar usuarios del mismo nivel o superior
            # Excepto para superusuarios que pueden modificar a cualquiera
            if target_user_level <= requesting_level and not requesting_user.is_superuser:
                raise PermissionDenied(
                    f'No puede modificar a este usuario. '
                    f'Solo usuarios con mayor privilegio pueden modificarlo.'
                )
            
            # ISS-004 FIX: Prevenir que un usuario se modifique a sí mismo con privilegios elevados
            # (excepto superusuarios)
            if target_user.pk == requesting_user.pk and not requesting_user.is_superuser:
                # Solo permitir modificar campos básicos, no el rol
                if target_role and target_role.lower() != requesting_user.rol.lower():
                    raise PermissionDenied(
                        'No puede cambiar su propio rol. Contacte a un administrador.'
                    )

    def perform_create(self, serializer):
        """
        ISS-013: Valida jerarquía de roles antes de crear usuario.
        """
        target_role = self.request.data.get('rol', '').lower()
        is_superuser = self.request.data.get('is_superuser', False)
        
        # Si intentan crear un superusuario, validar que quien lo hace sea superusuario
        if is_superuser and not self.request.user.is_superuser:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('Solo superusuarios pueden crear otros superusuarios.')
        
        # Validar jerarquía de roles
        if target_role:
            self._validate_role_hierarchy(self.request.user, target_role)
        
        # Guardar el usuario
        user = serializer.save()
        
        # Registrar en auditoría
        AuditoriaLog.objects.create(
            usuario=self.request.user,
            accion='CREATE',
            modelo='User',
            objeto_id=str(user.id),
            datos_nuevos={'username': user.username, 'rol': user.rol, 'email': user.email},
            detalles={'objeto_repr': user.username, 'creado_por': self.request.user.username},
            ip_address=self.request.META.get('REMOTE_ADDR')
        )
        logger.info(f"Usuario {user.username} creado por {self.request.user.username}")

    def perform_update(self, serializer):
        """
        ISS-013: Valida jerarquía de roles antes de actualizar usuario.
        """
        instance = serializer.instance
        target_role = self.request.data.get('rol', instance.rol).lower() if self.request.data.get('rol') else None
        is_superuser = self.request.data.get('is_superuser')
        
        # Si intentan hacer superusuario, validar que quien lo hace sea superusuario
        if is_superuser and not self.request.user.is_superuser:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('Solo superusuarios pueden otorgar privilegios de superusuario.')
        
        # Validar que tenemos privilegio sobre el usuario a modificar
        self._validate_role_hierarchy(self.request.user, target_role or instance.rol, instance)
        
        # Si cambia el rol, validar también el nuevo rol
        if target_role and target_role != instance.rol:
            self._validate_role_hierarchy(self.request.user, target_role)
        
        # Guardar datos anteriores para auditoría
        old_data = {
            'username': instance.username,
            'rol': instance.rol,
            'email': instance.email,
            'is_active': instance.is_active,
        }
        
        # Guardar el usuario
        user = serializer.save()
        
        # Detectar cambios
        new_data = {
            'username': user.username,
            'rol': user.rol,
            'email': user.email,
            'is_active': user.is_active,
        }
        cambios = {k: (old_data[k], new_data[k]) for k in new_data if old_data.get(k) != new_data.get(k)}
        
        if cambios:
            AuditoriaLog.objects.create(
                usuario=self.request.user,
                accion='UPDATE',
                modelo='User',
                objeto_id=str(user.id),
                datos_anteriores=old_data,
                datos_nuevos=new_data,
                detalles={'objeto_repr': user.username, 'modificado_por': self.request.user.username, 'cambios': list(cambios.keys())},
                ip_address=self.request.META.get('REMOTE_ADDR')
            )
        logger.info(f"Usuario {user.username} actualizado por {self.request.user.username}: {cambios}")

    def destroy(self, request, *args, **kwargs):
        """
        ISS-004 FIX: Valida jerarquía de roles antes de eliminar usuario.
        
        - Un usuario no puede eliminarse a sí mismo
        - Un usuario no puede eliminar a otro de igual o mayor privilegio
        """
        from rest_framework.exceptions import PermissionDenied
        
        instance = self.get_object()
        
        # Prevenir auto-eliminación
        if instance.pk == request.user.pk:
            return Response(
                {'error': 'No puede eliminarse a sí mismo'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Validar jerarquía (excepto superusuarios)
        if not request.user.is_superuser:
            requesting_level = get_role_level(request.user)
            target_level = get_role_level(instance)
            
            if target_level <= requesting_level:
                raise PermissionDenied(
                    'No puede eliminar a este usuario. '
                    'Solo usuarios con mayor privilegio pueden eliminarlo.'
                )
        
        # Registrar en auditoría antes de eliminar
        AuditoriaLog.objects.create(
            usuario=request.user,
            accion='DELETE',
            modelo='User',
            objeto_id=str(instance.id),
            datos_anteriores={
                'username': instance.username,
                'email': instance.email,
                'rol': instance.rol,
                'centro_id': instance.centro_id,
            },
            detalles={'objeto_repr': instance.username, 'eliminado_por': request.user.username},
            ip_address=request.META.get('REMOTE_ADDR')
        )
        
        logger.info(f"Usuario {instance.username} eliminado por {request.user.username}")
        return super().destroy(request, *args, **kwargs)

    @action(detail=False, methods=['get', 'patch'], url_path='me')
    def me(self, request):
        """GET/PATCH /api/usuarios/me/ - Perfil del usuario autenticado"""
        try:
            UserProfile.objects.get_or_create(usuario=request.user)
        except Exception as e:
            # Si la tabla user_profiles no existe, continuar sin profile
            logger.warning(f"No se pudo crear UserProfile para {request.user}: {e}")
        
        if request.method == 'PATCH':
            # Guardar datos anteriores para auditoría
            old_data = {
                'email': request.user.email,
                'first_name': request.user.first_name,
                'last_name': request.user.last_name,
            }
            profile = getattr(request.user, 'profile', None)
            if profile:
                old_data['telefono'] = profile.telefono
                old_data['rol'] = profile.rol
            
            serializer = UserMeSerializer(request.user, data=request.data, partial=True)
            serializer.is_valid(raise_exception=True)
            serializer.save()
            
            # Refrescar usuario y profile para capturar cambios
            request.user.refresh_from_db()
            
            # Detectar cambios para auditoría
            new_data = {
                'email': request.user.email,
                'first_name': request.user.first_name,
                'last_name': request.user.last_name,
            }
            # Refrescar profile para obtener telefono actualizado
            profile = getattr(request.user, 'profile', None)
            if profile:
                profile.refresh_from_db()
                new_data['telefono'] = profile.telefono
                new_data['rol'] = profile.rol
            
            cambios = {k: (old_data.get(k), new_data.get(k)) 
                      for k in new_data if old_data.get(k) != new_data.get(k)}
            
            if cambios:
                AuditoriaLog.objects.create(
                    usuario=request.user,
                    accion='UPDATE',
                    modelo='Usuario',
                    objeto_id=str(request.user.id),
                    datos_nuevos=cambios,
                    detalles={'objeto_repr': str(request.user)}
                )
            
            updated_user = User.objects.select_related('profile').get(pk=request.user.pk)
            updated = UserMeSerializer(updated_user)
            return Response(updated.data)
        else:
            serializer = UserMeSerializer(request.user)
            return Response(serializer.data)

    @action(detail=False, methods=['post'], url_path='me/change-password', throttle_classes=[PasswordChangeThrottle])
    def me_change_password(self, request):
        """
        POST /api/usuarios/me/change-password/
        
        Cambio de contraseña del usuario autenticado.
        Requiere: old_password, new_password, confirm_password
        
        Validaciones:
        - Contraseña actual correcta
        - Nueva contraseña ≥8 caracteres
        - Al menos una mayúscula
        - Al menos un número
        - Diferente a la anterior
        
        ISS-PASSWORD FIX: Incluye verificación post-guardado.
        """
        from django.db import transaction
        
        user = request.user
        old_password = request.data.get('old_password')
        new_password = request.data.get('new_password')
        confirm_password = request.data.get('confirm_password')

        if not all([old_password, new_password, confirm_password]):
            return Response({'error': 'Debe proporcionar old_password, new_password y confirm_password'}, status=status.HTTP_400_BAD_REQUEST)

        if new_password != confirm_password:
            return Response({'error': 'Las contraseñas nuevas no coinciden'}, status=status.HTTP_400_BAD_REQUEST)

        if not user.check_password(old_password):
            logger.warning("Intento de cambio de contraseña fallido para %s: contraseña actual incorrecta", user.username)
            # Registrar intento fallido en auditoría
            AuditoriaLog.objects.create(
                usuario=user,
                accion='UPDATE',
                modelo='Usuario',
                objeto_id=str(user.id),
                detalles={'objeto_repr': str(user), 'razon': 'Contraseña actual incorrecta', 'resultado': 'fallido'}
            )
            return Response({'error': 'Contraseña actual incorrecta'}, status=status.HTTP_400_BAD_REQUEST)

        # ISS-022: Usar validadores de Django en lugar de reglas duplicadas
        from django.contrib.auth.password_validation import validate_password
        from django.core.exceptions import ValidationError as DjangoValidationError
        
        try:
            validate_password(new_password, user)
        except DjangoValidationError as e:
            return Response({'error': '; '.join(e.messages)}, status=status.HTTP_400_BAD_REQUEST)

        if old_password == new_password:
            return Response({'error': 'La nueva contraseña debe ser diferente a la anterior'}, status=status.HTTP_400_BAD_REQUEST)

        # ISS-PASSWORD FIX: Usar transacción y verificar post-guardado
        try:
            with transaction.atomic():
                user.set_password(new_password)
                user.save()
                
                # Verificar que la contraseña se guardó correctamente
                user.refresh_from_db()
                if not user.check_password(new_password):
                    logger.error(f"me_change_password - PASSWORD VERIFICATION FAILED for {user.username}!")
                    # Reintentar
                    user.set_password(new_password)
                    user.save(update_fields=['password'])
                    user.refresh_from_db()
                    if not user.check_password(new_password):
                        raise Exception("No se pudo guardar la contraseña después de reintentar")
                    logger.info(f"me_change_password - Password fixed on retry for {user.username}")
        except Exception as e:
            logger.error(f"me_change_password - Error crítico al guardar contraseña: {e}")
            return Response({
                'error': f'Error crítico al guardar contraseña. Contacte al administrador.'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
        # HALLAZGO #3: Invalidar todos los tokens JWT existentes del usuario
        # Esto previene que tokens robados sigan funcionando después del cambio de contraseña
        try:
            from rest_framework_simplejwt.token_blacklist.models import OutstandingToken
            # Eliminar todos los tokens pendientes del usuario (los marca como inválidos)
            OutstandingToken.objects.filter(user=user).delete()
        except Exception as e:
            # Si la app token_blacklist no está habilitada, loguear advertencia
            logger.warning(f'No se pudieron invalidar tokens JWT para {user.username}: {e}')
        
        # Registrar cambio exitoso en auditoría
        AuditoriaLog.objects.create(
            usuario=user,
            accion='UPDATE',
            modelo='Usuario',
            objeto_id=str(user.id),
            detalles={'objeto_repr': str(user), 'resultado': 'Contraseña actualizada exitosamente', 'tokens_invalidados': True}
        )
        
        logger.info("Contraseña actualizada para usuario %s (verificada, tokens JWT invalidados)", user.username)
        return Response({
            'message': 'Contraseña actualizada exitosamente',
            'nota': 'Todas las sesiones activas han sido cerradas. Debe iniciar sesión nuevamente.'
        })

    @action(detail=True, methods=['post'], url_path='cambiar-password')
    def cambiar_password(self, request, pk=None):
        """POST /api/usuarios/{id}/cambiar-password/ - Admin cambia password de otro usuario
        
        ISS-PASSWORD FIX: Incluye verificación post-guardado para asegurar que la contraseña
        se guardó correctamente y es funcional.
        """
        from django.db import transaction
        
        # Solo superusuarios o farmacia pueden cambiar passwords de otros
        if not request.user.is_superuser and request.user.rol not in ['admin_sistema', 'farmacia', 'admin_farmacia']:
            return Response({'error': 'No tiene permisos para cambiar contraseñas de otros usuarios'}, status=status.HTTP_403_FORBIDDEN)
        
        try:
            usuario = User.objects.get(pk=pk)
        except User.DoesNotExist:
            return Response({'error': 'Usuario no encontrado'}, status=status.HTTP_404_NOT_FOUND)
        
        # No permitir cambiar password de superusuarios (solo otro superuser puede)
        if usuario.is_superuser and not request.user.is_superuser:
            return Response({'error': 'No puede cambiar la contraseña de un superusuario'}, status=status.HTTP_403_FORBIDDEN)
        
        new_password = request.data.get('new_password')
        
        if not new_password:
            return Response({'error': 'Debe proporcionar new_password'}, status=status.HTTP_400_BAD_REQUEST)
        
        # HALLAZGO #6: Usar validate_password de Django (consistencia con cambiar_mi_password)
        from django.contrib.auth.password_validation import validate_password
        from django.core.exceptions import ValidationError as DjangoValidationError
        
        try:
            validate_password(new_password, usuario)
        except DjangoValidationError as e:
            return Response({'error': '; '.join(e.messages)}, status=status.HTTP_400_BAD_REQUEST)
        
        # ISS-PASSWORD FIX: Usar transacción y verificar post-guardado
        try:
            with transaction.atomic():
                usuario.set_password(new_password)
                usuario.save()
                
                # Verificar que la contraseña se guardó correctamente
                usuario.refresh_from_db()
                if not usuario.check_password(new_password):
                    logger.error(f"cambiar_password - PASSWORD VERIFICATION FAILED for {usuario.username}!")
                    # Reintentar
                    usuario.set_password(new_password)
                    usuario.save(update_fields=['password'])
                    usuario.refresh_from_db()
                    if not usuario.check_password(new_password):
                        raise Exception("No se pudo guardar la contraseña después de reintentar")
                    logger.info(f"cambiar_password - Password fixed on retry for {usuario.username}")
        except Exception as e:
            logger.error(f"cambiar_password - Error crítico al guardar contraseña: {e}")
            return Response({
                'error': f'Error crítico al guardar contraseña: {str(e)}. Contacte al administrador.'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
        # HALLAZGO #3: Invalidar todos los tokens JWT existentes del usuario
        try:
            from rest_framework_simplejwt.token_blacklist.models import OutstandingToken
            OutstandingToken.objects.filter(user=usuario).delete()
        except Exception as e:
            logger.warning(f'No se pudieron invalidar tokens JWT para {usuario.username}: {e}')
        
        # Registrar en auditoría
        AuditoriaLog.objects.create(
            usuario=request.user,
            accion='UPDATE',
            modelo='User',
            objeto_id=str(usuario.id),
            detalles={'objeto_repr': usuario.username, 'cambiado_por': request.user.username, 'cambio': 'password'},
            ip_address=request.META.get('REMOTE_ADDR')
        )
        
        logger.info("Contraseña de usuario %s actualizada por %s (verificada)", usuario.username, request.user.username)
        return Response({'message': f'Contraseña de {usuario.username} actualizada exitosamente'})

    @action(detail=False, methods=['get'], url_path='exportar-excel')
    def exportar_excel(self, request):
        """GET /api/usuarios/exportar_excel/ - Exporta usuarios a Excel
        
        Acepta los mismos filtros que el listado:
        - search: búsqueda en username, email, first_name, last_name, adscripcion
        - rol: filtrar por rol específico
        - is_active: true/false para filtrar por estado
        - centro: ID del centro para filtrar
        """
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from django.http import HttpResponse
        from django.db.models import Q
        
        # Admin y Farmacia pueden exportar usuarios
        if not request.user.is_superuser and request.user.rol not in ['admin_sistema', 'farmacia', 'admin_farmacia']:
            return Response({'error': 'No tiene permisos para exportar usuarios'}, status=status.HTTP_403_FORBIDDEN)
        
        # Obtener queryset base según permisos del usuario
        user = request.user
        if user.is_superuser or self._is_farmacia_or_admin(user):
            usuarios = User.objects.all()
        else:
            # Usuario no admin solo puede exportar usuarios de su centro
            if hasattr(user, 'centro') and user.centro:
                usuarios = User.objects.filter(centro=user.centro)
            else:
                usuarios = User.objects.filter(id=user.id)
        
        # Aplicar filtros server-side (mismos que el listado)
        usuarios = self._apply_filters(usuarios, request.query_params)
        
        # Aplicar búsqueda de texto (igual que search_fields)
        search = request.query_params.get('search', '').strip()
        if search:
            usuarios = usuarios.filter(
                Q(username__icontains=search) |
                Q(email__icontains=search) |
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(adscripcion__icontains=search)
            )
        
        usuarios = usuarios.select_related('centro').order_by('username')
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Usuarios'
        
        headers = ['#', 'Usuario', 'Email', 'Nombre', 'Apellidos', 'Rol', 'Centro', 'Activo', 'Fecha Registro']
        ws.append(headers)
        
        header_fill = PatternFill(start_color='632842', end_color='632842', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        for idx, u in enumerate(usuarios, start=1):
            ws.append([
                idx,
                u.username,
                u.email,
                u.first_name,
                u.last_name,
                u.rol or 'Sin rol',
                u.centro.nombre if u.centro else '-',
                'Si' if u.is_active else 'No',
                u.date_joined.strftime('%Y-%m-%d %H:%M') if u.date_joined else '-'
            ])
        
        for col in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=usuarios_{timezone.now().strftime("%Y%m%d")}.xlsx'
        wb.save(response)
        return response

    @action(detail=False, methods=['post'], url_path='importar-excel')
    def importar_excel(self, request):
        """POST /api/usuarios/importar-excel/ - Importa usuarios desde Excel
        
        Columnas esperadas (deben coincidir con la plantilla):
        0. username (requerido, único, se normaliza a minúsculas)
        1. email (opcional, se genera si falta)
        2. first_name (opcional)
        3. last_name (opcional)
        4. password (opcional, mín 8 chars; default: temporal que requiere cambio)
        5. rol (opcional: admin_sistema, admin_farmacia, farmacia, centro, vista, usuario_normal, usuario_vista; default: centro)
        6. centro_clave (opcional: ID o nombre del centro a asignar)
        7. adscripcion (opcional)
        8. telefono (informativo, no se guarda)
        9. activo (Si/No, default: Si)
        
        SEGURIDAD:
        - Tamaño máximo de archivo: 5MB
        - Máximo de filas: 1000
        - Solo archivos .xlsx
        - Usernames normalizados a minúsculas
        - Contraseñas temporales NO se exponen en respuesta
        """
        import openpyxl
        import re
        from .models import Centro
        
        # === VALIDACIÓN DE PERMISOS ===
        if not request.user.is_superuser and request.user.rol not in ['admin_sistema', 'farmacia', 'admin_farmacia']:
            return Response({'error': 'No tiene permisos para importar usuarios'}, status=status.HTTP_403_FORBIDDEN)
        
        archivo = request.FILES.get('file')
        if not archivo:
            return Response({'error': 'No se recibió archivo'}, status=status.HTTP_400_BAD_REQUEST)
        
        # === VALIDACIÓN DE ARCHIVO (ISS-001) ===
        MAX_FILE_SIZE = 5 * 1024 * 1024  # 5MB
        MAX_ROWS = 1000
        ALLOWED_EXTENSIONS = ['.xlsx']
        
        # Validar tamaño
        if archivo.size > MAX_FILE_SIZE:
            return Response(
                {'error': f'Archivo demasiado grande. Máximo permitido: {MAX_FILE_SIZE // (1024*1024)}MB'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Validar extensión
        file_ext = '.' + archivo.name.split('.')[-1].lower() if '.' in archivo.name else ''
        if file_ext not in ALLOWED_EXTENSIONS:
            return Response(
                {'error': f'Tipo de archivo no permitido. Solo se aceptan: {ALLOWED_EXTENSIONS}'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Validar contenido (magic bytes para xlsx/zip)
        archivo.seek(0)
        header = archivo.read(4)
        archivo.seek(0)
        if header[:4] != b'PK\x03\x04':  # ZIP/XLSX magic bytes
            return Response(
                {'error': 'El archivo no parece ser un Excel válido (.xlsx)'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # === CONSTANTES DE VALIDACIÓN ===
        # Todos los roles válidos según ROLES_USUARIO en constants.py
        ROLES_VALIDOS = [
            # Roles actuales
            'admin', 'farmacia', 'vista',
            'medico', 'administrador_centro', 'director_centro', 'centro',
            # Legacy (compatibilidad)
            'admin_sistema', 'admin_farmacia', 'usuario_normal', 'usuario_vista', 'superusuario',
        ]
        USERNAME_PATTERN = re.compile(r'^[a-z0-9_.-]{3,50}$')
        EMAIL_PATTERN = re.compile(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$')
        
        try:
            wb = openpyxl.load_workbook(archivo, read_only=True, data_only=True)
            ws = wb.active
            
            # Contar filas antes de procesar
            row_count = sum(1 for _ in ws.iter_rows(min_row=2, max_row=MAX_ROWS + 2, values_only=True))
            if row_count > MAX_ROWS:
                return Response(
                    {'error': f'Demasiadas filas. Máximo permitido: {MAX_ROWS}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Reabrir para procesar (read_only no permite re-iterar)
            archivo.seek(0)
            wb = openpyxl.load_workbook(archivo)
            ws = wb.active
            
            creados = 0
            actualizados = 0
            errores = []
            usuarios_con_reset = 0  # Solo conteo, NO exponemos contraseñas (ISS-002)
            
            # ── OPTIMIZACIÓN BULK: precargar en memoria para evitar N round-trips a BD ──
            import secrets
            import string
            
            # 1 query para todos los centros activos
            centros_por_nombre = {c.nombre.upper(): c for c in Centro.objects.filter(activo=True)}
            centros_por_id = {str(c.id): c for c in centros_por_nombre.values()}
            
            # 1 query para todos los usuarios existentes (para detectar duplicados)
            existing_users = {u.username.lower(): u for u in User.objects.all().select_related('centro')}
            
            users_to_create = []    # User instances (in-memory, password ya hasheado)
            users_to_update = []    # User instances existentes modificados
            UPDATE_FIELDS = ['email', 'first_name', 'last_name', 'rol', 'centro', 'adscripcion', 'is_active']
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=MAX_ROWS + 1, values_only=True), start=2):
                valores = list(row) + [None] * 10
                
                # === EXTRACCIÓN Y NORMALIZACIÓN (ISS-003) ===
                # Columnas: Username(0), Email(1), Nombre(2), Apellidos(3),
                #           Password(4), Rol(5), Centro ID(6), Adscripción(7),
                #           Teléfono(8), Activo(9)
                username_raw = str(valores[0] or '').strip()
                username = username_raw.lower()  # Normalizar a minúsculas
                email = str(valores[1] or '').strip().lower()
                first_name = str(valores[2] or '').strip()[:100]  # Limitar longitud
                last_name = str(valores[3] or '').strip()[:100]
                password = str(valores[4] or '').strip()
                rol = str(valores[5] or 'centro').strip().lower()
                # Normalizar Centro ID: Excel devuelve números como float (1.0, 2.0)
                # str(1.0) = "1.0" no coincide con str(c.id) = "1" → corregir
                _centro_raw = str(valores[6] or '').strip()
                try:
                    centro_clave_id = str(int(float(_centro_raw))) if _centro_raw else ''
                except (ValueError, TypeError):
                    centro_clave_id = _centro_raw.upper()
                centro_clave = _centro_raw.upper()  # Para búsqueda por nombre
                adscripcion = str(valores[7] or '').strip()[:200]
                # valores[8] = Teléfono (informativo, no se guarda en User)
                activo_raw = str(valores[9] or 'si').strip().lower()
                is_active = activo_raw not in ('no', 'false', '0', 'inactivo')
                
                # Saltar filas vacías
                if not username:
                    continue
                
                # === VALIDACIONES ESTRICTAS ===
                
                # Validar formato de username
                if not USERNAME_PATTERN.match(username):
                    errores.append({
                        'fila': row_idx, 
                        'campo': 'username',
                        'error': f'Username "{username}" inválido. Solo letras minúsculas, números, guiones y puntos (3-50 chars)'
                    })
                    continue
                
                # Validar rol estrictamente (NO usar default silencioso)
                if rol not in ROLES_VALIDOS:
                    errores.append({
                        'fila': row_idx,
                        'campo': 'rol', 
                        'error': f'Rol "{rol}" no válido. Roles permitidos: {ROLES_VALIDOS}'
                    })
                    continue
                
                # Validar email si se proporciona
                if email and not EMAIL_PATTERN.match(email):
                    errores.append({
                        'fila': row_idx,
                        'campo': 'email',
                        'error': f'Email "{email}" no tiene formato válido'
                    })
                    continue
                
                # Buscar centro desde caché (sin DB call)
                # Intentar por ID numérico normalizado, luego por nombre
                centro = None
                if centro_clave or centro_clave_id:
                    centro = centros_por_id.get(centro_clave_id) or centros_por_nombre.get(centro_clave)
                    if not centro and centro_clave:
                        errores.append({
                            'fila': row_idx,
                            'campo': 'centro_clave',
                            'error': f'Centro "{centro_clave}" no encontrado o inactivo'
                        })
                        continue
                
                # Generar password seguro si no se proporciona o es débil
                requiere_reset = False
                if not password or len(password) < 8:
                    chars = string.ascii_letters + string.digits + '!@#$%&*'
                    password = ''.join(secrets.choice(chars) for _ in range(12))
                    requiere_reset = True
                
                # === ACUMULAR EN MEMORIA (sin DB calls) ===
                existing_user = existing_users.get(username)
                
                if existing_user is None:
                    # Nuevo usuario: construir en memoria con password hasheado
                    user = User(
                        username=username,
                        email=email or f'{username}@sistema.local',
                        first_name=first_name,
                        last_name=last_name,
                        rol=rol,
                        centro=centro,
                        adscripcion=adscripcion or '',
                        is_active=is_active,
                    )
                    user.set_password(password)  # Hashea sin tocar BD
                    users_to_create.append(user)
                    # Registrar en caché local para detectar duplicados dentro del mismo archivo
                    existing_users[username] = user
                    if requiere_reset:
                        usuarios_con_reset += 1
                else:
                    # Actualizar usuario existente en memoria
                    if email:
                        existing_user.email = email
                    if first_name:
                        existing_user.first_name = first_name
                    if last_name:
                        existing_user.last_name = last_name
                    existing_user.rol = rol
                    if centro:
                        existing_user.centro = centro
                    if adscripcion:
                        existing_user.adscripcion = adscripcion
                    existing_user.is_active = is_active
                    users_to_update.append(existing_user)
            
            # ── PERSISTIR EN BD: 2 llamadas bulk en lugar de N individuales ──
            with transaction.atomic():
                if users_to_create:
                    User.objects.bulk_create(users_to_create, ignore_conflicts=False)
                    creados = len(users_to_create)
                
                if users_to_update:
                    User.objects.bulk_update(users_to_update, UPDATE_FIELDS)
                    actualizados = len(users_to_update)
            
            # Registrar auditoría resumen
            logger.info(
                f"Importación masiva por {request.user.username}: "
                f"{creados} creados, {actualizados} actualizados, {len(errores)} errores"
            )
            
            # === RESPUESTA SIN CONTRASEÑAS (ISS-002) ===
            return Response({
                'mensaje': 'Importación completada exitosamente',
                'resumen': {
                    'creados': creados,
                    'actualizados': actualizados,
                    'errores': len(errores),
                    'usuarios_requieren_cambio_password': usuarios_con_reset
                },
                # NO incluimos contraseñas - deben gestionarse por canal seguro
                'errores': errores[:20],  # Limitar errores mostrados
                'nota': (
                    f'{usuarios_con_reset} usuarios fueron creados con contraseña temporal. '
                    'Las credenciales deben entregarse por un canal seguro (NO por esta API). '
                    'Los usuarios deberán cambiar su contraseña en el primer inicio de sesión.'
                ) if usuarios_con_reset > 0 else None,
                'advertencia': (
                    'Se encontraron errores en algunas filas. Revise el detalle y corrija el archivo.'
                ) if errores else None
            })
            
        except openpyxl.utils.exceptions.InvalidFileException:
            return Response(
                {'error': 'El archivo no es un Excel válido o está corrupto'},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            logger.error(f"Error al importar usuarios: {e}")
            return Response(
                {'error': 'Error inesperado al procesar archivo. Contacte al administrador.'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'], url_path='plantilla')
    def plantilla_usuarios(self, request):
        """
        Descarga plantilla Excel actualizada para importación de usuarios.
        
        Usa el generador estandarizado con el esquema real de la base de datos.
        Columnas: Username, Email, Nombre, Apellidos, Password, Rol, Centro ID, 
        Adscripción, Teléfono, Activo
        
        SEGURIDAD: Los usuarios creados sin contraseña requieren cambio en primer login.
        """
        from core.utils.excel_templates import generar_plantilla_usuarios
        return generar_plantilla_usuarios()



# NOTA: ProductoViewSet, LoteViewSet, RequisicionViewSet y CentroViewSet
# están en inventario/views.py para evitar duplicación.
# Importar desde allí si es necesario.


class DetalleRequisicionViewSet(viewsets.ModelViewSet):
    """
    ViewSet para detalles de requisiciones.
    
    SEGURIDAD: Filtra por centro del usuario para prevenir IDOR/BOLA.
    Admin/farmacia pueden ver todos; usuarios de centro solo los de su centro.
    """
    serializer_class = DetalleRequisicionSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = StandardResultsSetPagination  # HALLAZGO #9: Paginación forzada
    
    def get_queryset(self):
        """
        SEGURIDAD: Filtrar detalles por centro/rol del usuario.
        
        - Admin/farmacia: acceso a todos los detalles
        - Centro: solo detalles de requisiciones de su centro
        """
        user = self.request.user
        
        # Base queryset con select_related para prevenir N+1
        queryset = DetalleRequisicion.objects.select_related(
            'requisicion', 'requisicion__centro', 'requisicion__solicitante',
            'producto', 'lote', 'lote__producto', 'lote__centro'
        )
        
        # Admin y farmacia ven todo
        if user.is_superuser:
            return queryset
        
        rol = (getattr(user, 'rol', '') or '').lower()
        group_names = set(g.name.upper() for g in user.groups.all())
        
        es_admin_o_farmacia = (
            rol in {'admin', 'admin_sistema', 'superusuario', 'farmacia', 'admin_farmacia', 'farmaceutico'} or
            'FARMACIA_ADMIN' in group_names or
            'FARMACEUTICO' in group_names
        )
        
        if es_admin_o_farmacia:
            return queryset
        
        # Usuarios de centro solo ven detalles de requisiciones de su centro
        user_centro = getattr(user, 'centro', None)
        if user_centro:
            return queryset.filter(requisicion__centro=user_centro)
        
        # Sin centro asignado, no ve nada
        return queryset.none()


class AuditoriaLogViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet para registros de auditoria (solo lectura)"""
    queryset = AuditoriaLog.objects.all()
    serializer_class = AuditoriaLogSerializer
    # Solo Superuser/Admin puede leer el log de auditoría.
    permission_classes = [IsSuperuserOnly]
    pagination_class = StandardResultsSetPagination
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['modelo', 'usuario__username', 'accion']
    ordering_fields = ['timestamp']
    ordering = ['-timestamp']
    
    # ISS-037: Límite máximo de registros para evitar slow queries
    MAX_AUDIT_RECORDS = 10000
    
    def get_queryset(self):
        """Filtrado avanzado por parámetros de query.
        
        ISS-037: Se aplica un límite máximo para evitar escaneo de tabla completa.
        """
        queryset = AuditoriaLog.objects.select_related('usuario').order_by('-timestamp')
        
        # Filtro por acción
        accion = self.request.query_params.get('accion')
        if accion:
            queryset = queryset.filter(accion=accion)
        
        # Filtro por modelo
        modelo = self.request.query_params.get('modelo')
        if modelo:
            queryset = queryset.filter(modelo__icontains=modelo)
        
        # Filtro por usuario
        usuario = self.request.query_params.get('usuario')
        if usuario:
            queryset = queryset.filter(
                Q(usuario__username__icontains=usuario) |
                Q(usuario__first_name__icontains=usuario) |
                Q(usuario__last_name__icontains=usuario)
            )
        
        # Filtro por fecha inicio
        fecha_inicio = self.request.query_params.get('fecha_inicio')
        if fecha_inicio:
            queryset = queryset.filter(timestamp__date__gte=fecha_inicio)
        
        # Filtro por fecha fin
        fecha_fin = self.request.query_params.get('fecha_fin')
        if fecha_fin:
            queryset = queryset.filter(timestamp__date__lte=fecha_fin)
        
        # ISS-037: La paginación manejará el límite de registros
        return queryset

    @action(detail=False, methods=['get'])
    def exportar(self, request):
        """
        Exporta logs de auditoría a Excel.
        GET /api/auditoria/exportar/
        """
        try:
            logs = self.get_queryset()[:5000]  # Limitar a 5000 registros
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Auditoria'
            
            # Título
            ws.merge_cells('A1:G1')
            titulo = ws['A1']
            titulo.value = 'REPORTE DE AUDITORÍA'
            titulo.font = openpyxl.styles.Font(bold=True, size=14, color='632842')
            titulo.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            
            ws.merge_cells('A2:G2')
            fecha = ws['A2']
            fecha.value = f'Generado el {timezone.now().strftime("%d/%m/%Y %H:%M")}'
            fecha.font = openpyxl.styles.Font(size=10, italic=True)
            fecha.alignment = openpyxl.styles.Alignment(horizontal='center')
            
            ws.append([])
            
            # Encabezados
            headers = ['#', 'Fecha', 'Usuario', 'Acción', 'Modelo', 'Objeto', 'IP']
            ws.append(headers)
            
            header_fill = openpyxl.styles.PatternFill(start_color='632842', end_color='632842', fill_type='solid')
            header_font = openpyxl.styles.Font(bold=True, color='FFFFFF', size=11)
            for cell in ws[4]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            
            # Datos
            for idx, log in enumerate(logs, 1):
                objeto_repr = ''
                if log.detalles and isinstance(log.detalles, dict):
                    objeto_repr = log.detalles.get('objeto_repr', '')
                if not objeto_repr:
                    objeto_repr = f"{log.modelo} #{log.objeto_id}" if log.objeto_id else log.modelo
                
                ws.append([
                    idx,
                    log.timestamp.strftime('%d/%m/%Y %H:%M:%S') if log.timestamp else '',
                    log.usuario.username if log.usuario else 'Sistema',
                    log.accion,
                    log.modelo,
                    str(objeto_repr),  # SIN TRUNCAR
                    log.ip_address or ''
                ])
            
            # Ajustar anchos - Columna Descripción más ancha para texto completo
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 80  # Más ancho para descripciones completas
            ws.column_dimensions['G'].width = 15
            
            from django.http import HttpResponse
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename=Auditoria_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            wb.save(response)
            
            return response
            
        except Exception as e:
            logger.error(f"Error al exportar auditoría: {e}")
            return Response({
                'error': 'Error al exportar auditoría',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path='exportar-pdf')
    def exportar_pdf(self, request):
        """
        Exporta logs de auditoría a PDF con fondo institucional.
        GET /api/auditoria/exportar-pdf/
        
        Ideal para reportes oficiales de auditoría y cumplimiento normativo.
        """
        try:
            logs = self.get_queryset()[:2000]  # Limitar para PDFs
            
            # Preparar datos para el generador
            auditoria_data = []
            for log in logs:
                objeto_repr = ''
                if log.detalles and isinstance(log.detalles, dict):
                    objeto_repr = log.detalles.get('objeto_repr', '')
                if not objeto_repr:
                    objeto_repr = f"{log.modelo} #{log.objeto_id}" if log.objeto_id else log.modelo
                
                auditoria_data.append({
                    'fecha': log.timestamp,
                    'usuario': log.usuario.username if log.usuario else 'Sistema',
                    'accion': log.accion,
                    'modelo': log.modelo,
                    'objeto_repr': objeto_repr,
                    'ip_address': log.ip_address or ''
                })
            
            # Filtros aplicados
            filtros = {
                'fecha_inicio': request.query_params.get('fecha_inicio'),
                'fecha_fin': request.query_params.get('fecha_fin'),
                'usuario': request.query_params.get('usuario'),
                'accion': request.query_params.get('accion'),
                'modelo': request.query_params.get('modelo'),
            }
            filtros = {k: v for k, v in filtros.items() if v}
            
            # Generar PDF
            pdf_buffer = generar_reporte_auditoria(auditoria_data, filtros)
            
            response = HttpResponse(pdf_buffer, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename=Auditoria_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf'
            return response
            
        except Exception as e:
            logger.error(f"Error al exportar auditoría PDF: {e}")
            return Response({
                'error': 'Error al exportar auditoría',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'])
    def opciones_filtro(self, request):
        """
        Devuelve opciones para filtros de auditoría.
        GET /api/auditoria/opciones_filtro/
        
        Retorna:
        - acciones: Lista de acciones únicas registradas
        - modelos: Lista de modelos auditados
        """
        from .constants import ACCIONES_AUDITORIA, MODELOS_AUDITADOS
        
        # Acciones que realmente existen en la base de datos
        acciones_db = list(
            AuditoriaLog.objects.values_list('accion', flat=True)
            .distinct()
            .order_by('accion')
        )
        
        # Modelos que realmente existen en la base de datos
        modelos_db = list(
            AuditoriaLog.objects.values_list('modelo', flat=True)
            .distinct()
            .order_by('modelo')
        )
        
        return Response({
            'acciones': acciones_db,
            'modelos': modelos_db,
            'catalogo_acciones': dict(ACCIONES_AUDITORIA),
            'catalogo_modelos': MODELOS_AUDITADOS
        })


class ImportacionLogViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet para registros de importacion (solo lectura)"""
    queryset = ImportacionLog.objects.all()
    serializer_class = ImportacionLogSerializer
    permission_classes = [IsAuthenticated, IsFarmaciaRole]
    pagination_class = StandardResultsSetPagination
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['modelo', 'usuario__username', 'estado']
    ordering_fields = ['fecha_inicio']
    ordering = ['-fecha_inicio']


class NotificacionViewSet(
    mixins.ListModelMixin,
    mixins.RetrieveModelMixin,
    mixins.DestroyModelMixin,
    viewsets.GenericViewSet
):
    """
    ViewSet para notificaciones del usuario.
    
    Las notificaciones son generadas automáticamente por el sistema:
    - Cambios de estado en requisiciones
    - Alertas de stock crítico
    - Alertas de lotes por caducar
    
    Endpoints:
    - GET /api/notificaciones/ - Lista notificaciones del usuario
    - GET /api/notificaciones/{id}/ - Detalle de notificación
    - DELETE /api/notificaciones/{id}/ - Eliminar notificación propia
    - POST /api/notificaciones/{id}/marcar-leida/ - Marcar como leída
    - POST /api/notificaciones/marcar-todas-leidas/ - Marcar todas como leídas
    - GET /api/notificaciones/no-leidas-count/ - Contador de no leídas
    """
    http_method_names = ['get', 'post', 'delete', 'head', 'options']
    serializer_class = NotificacionSerializer
    permission_classes = [IsAuthenticated, CanViewNotifications]
    pagination_class = StandardResultsSetPagination
    # ISS-FIX: Desactivar OrderingFilter de DRF, manejamos ordering manualmente
    filter_backends = []
    ordering = ['-created_at']

    def get_queryset(self):
        try:
            queryset = Notificacion.objects.filter(usuario=self.request.user)
            
            # ISS-FIX: Manejar ordering manualmente con alias fecha_creacion -> created_at
            ordering = self.request.query_params.get('ordering', '-created_at')
            if 'fecha_creacion' in ordering:
                # Reemplazar fecha_creacion por created_at
                ordering = ordering.replace('fecha_creacion', 'created_at')
            
            # Aplicar ordering solo si es un campo válido
            valid_fields = ['created_at', '-created_at', 'leida', '-leida']
            if ordering in valid_fields:
                queryset = queryset.order_by(ordering)
            else:
                queryset = queryset.order_by('-created_at')
            
            tipo = self.request.query_params.get('tipo')
            if tipo:
                queryset = queryset.filter(tipo=tipo)

            leida = self.request.query_params.get('leida')
            if leida in ['true', 'false']:
                queryset = queryset.filter(leida=leida == 'true')

            fecha_desde = self.request.query_params.get('desde')
            fecha_hasta = self.request.query_params.get('hasta')
            try:
                if fecha_desde:
                    queryset = queryset.filter(created_at__date__gte=fecha_desde)
                if fecha_hasta:
                    queryset = queryset.filter(created_at__date__lte=fecha_hasta)
            except Exception:
                pass

            return queryset
        except Exception as e:
            # ISS-FIX: Si la tabla notificaciones no existe, devolver queryset vacío
            logger.warning(f"Error accediendo a notificaciones: {e}")
            return Notificacion.objects.none()
    
    def list(self, request, *args, **kwargs):
        """Override list para manejar errores de tabla inexistente."""
        try:
            return super().list(request, *args, **kwargs)
        except Exception as e:
            logger.warning(f"Error listando notificaciones: {e}")
            return Response({'count': 0, 'results': []})

    @action(detail=True, methods=['post'], url_path='marcar-leida')
    def marcar_leida(self, request, pk=None):
        """POST /api/notificaciones/{id}/marcar-leida/"""
        try:
            notificacion = self.get_object()
            notificacion.leida = True
            notificacion.save()
            return Response({'leida': True})
        except Exception as e:
            logger.warning(f"Error marcando notificación como leída: {e}")
            return Response({'leida': False, 'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'], url_path='marcar-todas-leidas')
    def marcar_todas_leidas(self, request):
        """
        POST /api/notificaciones/marcar-todas-leidas/
        
        Marca como leídas las notificaciones que coinciden con los filtros actuales.
        Respeta los query params: tipo, desde, hasta, leida.
        Solo afecta las notificaciones del usuario autenticado (get_queryset ya filtra).
        """
        try:
            # get_queryset() ya aplica filtros de tipo, desde, hasta, leida y usuario
            updated = self.get_queryset().filter(leida=False).update(leida=True)
            return Response({'marcadas': updated})
        except Exception as e:
            logger.warning(f"Error marcando notificaciones como leídas: {e}")
            return Response({'marcadas': 0})

    @action(detail=False, methods=['get'], url_path='no-leidas-count',
           permission_classes=[AllowAny])
    def no_leidas_count(self, request):
        """GET /api/notificaciones/no-leidas-count/
        
        ISS-FIX: Permite acceso sin autenticación para evitar errores 401
        durante el polling del frontend cuando el token expira.
        Devuelve 0 si el usuario no está autenticado.
        """
        if not request.user or not request.user.is_authenticated:
            return Response({'no_leidas': 0})
        try:
            count = self.get_queryset().filter(leida=False).count()
            return Response({'no_leidas': count})
        except Exception as e:
            # ISS-FIX: Si la tabla no existe o hay error, devolver 0 sin fallar
            logger.warning(f"Error contando notificaciones no leídas: {e}")
            return Response({'no_leidas': 0})
    
    @action(detail=False, methods=['get'], url_path='diagnostico')
    def diagnostico(self, request):
        """
        ISS-DEBUG: Endpoint de diagnóstico para notificaciones.
        
        ISS-SEC-FIX: Restringido a admin/farmacia solamente.
        Retorna estadísticas básicas sin exponer SQL ni datos sensibles.
        """
        # ISS-SEC-FIX: Solo admin/farmacia pueden acceder a diagnóstico
        user = request.user
        if not user.is_superuser and not (getattr(user, 'rol', '') or '').lower() in ['admin', 'admin_sistema', 'farmacia']:
            return Response({'error': 'No autorizado'}, status=status.HTTP_403_FORBIDDEN)
        
        try:
            # Obtener todas las notificaciones del usuario (sin filtros)
            todas = Notificacion.objects.filter(usuario=request.user)
            no_leidas = todas.filter(leida=False)
            
            # Obtener las últimas 10 notificaciones con campos básicos (sin datos sensibles)
            ultimas = todas.order_by('-created_at')[:10]
            
            return Response({
                'usuario': {
                    'id': request.user.pk,
                    'username': request.user.username,
                },
                'estadisticas': {
                    'total': todas.count(),
                    'no_leidas': no_leidas.count(),
                    'leidas': todas.filter(leida=True).count(),
                },
                'ultimas_notificaciones': [
                    {
                        'id': n.pk,
                        'titulo': n.titulo,
                        'tipo': n.tipo,
                        'leida': n.leida,
                        'created_at': n.created_at.isoformat() if n.created_at else None,
                    }
                    for n in ultimas
                ],
                # ISS-SEC-FIX: NO exponer consulta_sql ni datos completos
            })
        except Exception as e:
            logger.exception(f"Error en diagnóstico de notificaciones: {e}")
            return Response({
                'error': 'Error interno',
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def destroy(self, request, *args, **kwargs):
        """
        Eliminar notificación propia.
        
        Cualquier usuario autenticado puede eliminar sus propias notificaciones.
        El queryset ya filtra por usuario, así que solo pueden borrar las suyas.
        """
        try:
            instance = self.get_object()
            self.perform_destroy(instance)
            return Response(status=status.HTTP_204_NO_CONTENT)
        except Exception as e:
            return Response(
                {'error': f'No se pudo eliminar la notificación: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )


class ReportesViewSet(viewsets.ViewSet):
    """ViewSet para reportes del sistema"""
    permission_classes = [IsAuthenticated, IsFarmaciaRole]

    @action(detail=False, methods=['get'])
    def inventario(self, request):
        """
        GET /api/reportes/inventario/?formato=excel|pdf|json&centro=id|todos
        
        ISS-OPT: Optimizado con annotations para evitar N+1 queries.
        """
        from django.http import FileResponse
        from django.db.models import Sum, Count, Q, F, Case, When, Value, CharField
        from django.db.models.functions import Coalesce
        
        try:
            formato = request.query_params.get('formato', 'json')
            centro_filtro = request.query_params.get('centro', 'todos')
            hoy = timezone.now().date()
            
            # Filtro de lotes activos y vigentes (no vencidos)
            filtro_lotes_vigentes = Q(
                lotes__activo=True,
                lotes__cantidad_actual__gt=0
            ) & (Q(lotes__fecha_caducidad__gte=hoy) | Q(lotes__fecha_caducidad__isnull=True))
            
            # Filtro por centro si se especifica
            if centro_filtro and centro_filtro != 'todos':
                if centro_filtro == 'central':
                    filtro_lotes_vigentes &= Q(lotes__centro__isnull=True)
                else:
                    try:
                        centro_id = int(centro_filtro)
                        filtro_lotes_vigentes &= Q(lotes__centro_id=centro_id)
                    except (ValueError, TypeError):
                        pass
            
            # Query optimizada con annotations
            queryset = Producto.objects.filter(activo=True).annotate(
                stock_calculado=Coalesce(
                    Sum('lotes__cantidad_actual', filter=filtro_lotes_vigentes),
                    0
                ),
                lotes_activos_count=Count(
                    'lotes',
                    filter=filtro_lotes_vigentes
                )
            ).values(
                'id', 'clave', 'descripcion', 'unidad_medida',
                'stock_minimo', 'precio_unitario',
                'stock_calculado', 'lotes_activos_count'
            )
            
            datos = []
            for p in queryset:
                stock = p['stock_calculado'] or 0
                stock_minimo = p['stock_minimo'] or 0
                precio = float(p['precio_unitario'] or 0)
                
                # Calcular nivel de stock
                if stock == 0:
                    nivel = 'critico'
                elif stock_minimo > 0:
                    ratio = stock / stock_minimo
                    if ratio < 0.5:
                        nivel = 'critico'
                    elif ratio < 1:
                        nivel = 'bajo'
                    elif ratio > 2:
                        nivel = 'alto'
                    else:
                        nivel = 'normal'
                else:
                    nivel = 'normal'
                
                datos.append({
                    'id': p['id'],
                    'clave': p['clave'],
                    'descripcion': p['descripcion'],
                    'unidad_medida': p['unidad_medida'],
                    'stock_actual': stock,
                    'stock_minimo': stock_minimo,
                    'nivel_stock': nivel,
                    'nivel': nivel,
                    'precio_unitario': precio,
                    'valor_inventario': float(stock * precio),
                    'lotes_activos': p['lotes_activos_count'] or 0
                })

            # Calcular resumen
            productos_bajo_minimo = sum(
                1 for d in datos 
                if d['stock_actual'] < d['stock_minimo'] and d['stock_minimo'] > 0
            )
            
            resumen = {
                'total_productos': len(datos),
                'stock_total': sum(d['stock_actual'] for d in datos),
                'productos_sin_stock': sum(1 for d in datos if d['stock_actual'] == 0),
                'productos_bajo_minimo': productos_bajo_minimo,
                'productos_stock_critico': sum(1 for d in datos if d['nivel_stock'] == 'critico'),
                'valor_total_inventario': sum(d['valor_inventario'] for d in datos)
            }

            if formato == 'pdf':
                from core.utils.pdf_reports import generar_reporte_inventario
                pdf_buffer = generar_reporte_inventario(datos)
                response = FileResponse(
                    pdf_buffer,
                    content_type='application/pdf',
                    as_attachment=False,
                    filename=f'reporte_inventario_{timezone.now().strftime("%Y%m%d")}.pdf'
                )
                response['Content-Disposition'] = f'inline; filename="reporte_inventario_{timezone.now().strftime("%Y%m%d")}.pdf"'
                return response

            if formato == 'excel':
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.title = 'Inventario'

                headers = ['Clave', 'Descripcion', 'Stock Actual', 'Stock Minimo', 'Nivel', 'Precio Unitario', 'Valor Total']
                sheet.append(headers)

                for d in datos:
                    sheet.append([
                        d['clave'],
                        d['descripcion'],
                        d['stock_actual'],
                        d['stock_minimo'],
                        d['nivel_stock'],
                        d['precio_unitario'],
                        d['valor_inventario']
                    ])

                for row in sheet.iter_rows(min_row=1, max_row=len(datos) + 1):
                    for cell in row:
                        cell.alignment = openpyxl.styles.Alignment(horizontal='center')

                buffer = BytesIO()
                workbook.save(buffer)
                buffer.seek(0)

                return FileResponse(
                    buffer,
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True,
                    filename='reporte_inventario.xlsx'
                )

            return Response({
                'reporte': 'inventario',
                'fecha_generacion': timezone.now().isoformat(),
                'centro': centro_filtro,
                'datos': datos,
                'resumen': resumen
            })
        except Exception as e:
            logger.error(f"Error en reporte inventario: {e}")
            return Response(
                {'error': 'Error generando reporte de inventario', 'detail': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @method_decorator(cache_page(60 * 5))  # Cache 5 minutos
    @action(detail=False, methods=['get'])
    def caducidades(self, request):
        """GET /api/reportes/caducidades/?dias=30&formato=json|pdf"""
        from django.http import FileResponse
        from core.utils.pdf_reports import generar_reporte_caducidades
        
        dias = int(request.query_params.get('dias', 30))
        formato = request.query_params.get('formato', 'json')
        fecha_limite = date.today() + timedelta(days=dias)

        lotes = Lote.objects.filter(
            fecha_caducidad__lte=fecha_limite,
            activo=True
        ).select_related('producto').order_by('fecha_caducidad')

        datos = []
        for lote in lotes:
            dias_restantes = lote.dias_para_caducar()
            datos.append({
                'id': lote.id,
                'producto_clave': lote.producto.clave,
                'producto_descripcion': lote.producto.descripcion,
                'numero_lote': lote.numero_lote,
                'fecha_caducidad': lote.fecha_caducidad.isoformat(),
                'dias_restantes': dias_restantes,
                'alerta': lote.alerta_caducidad(),
                'cantidad_actual': lote.cantidad_actual,
                'marca': lote.marca
            })

        resumen = {
            'total_lotes_proximos': len(datos),
            'lotes_vencidos': sum(1 for d in datos if d['alerta'] == 'vencido'),
            'lotes_criticos': sum(1 for d in datos if d['alerta'] == 'critico'),
            'lotes_proximos': sum(1 for d in datos if d['alerta'] == 'proximo')
        }

        if formato == 'pdf':
            pdf_buffer = generar_reporte_caducidades(datos, dias=dias)
            response = FileResponse(
                pdf_buffer,
                content_type='application/pdf',
                as_attachment=False,
                filename=f'reporte_caducidades_{timezone.now().strftime("%Y%m%d")}.pdf'
            )
            response['Content-Disposition'] = f'inline; filename="reporte_caducidades_{timezone.now().strftime("%Y%m%d")}.pdf"'
            return response

        return Response({
            'reporte': 'caducidades',
            'fecha_generacion': timezone.now().isoformat(),
            'datos': datos,
            'resumen': resumen
        })

    @method_decorator(cache_page(60 * 10))  # Cache 10 minutos
    @action(detail=False, methods=['get'])
    def requisiciones(self, request):
        """GET /api/reportes/requisiciones/?formato=json|pdf|excel"""
        from django.http import FileResponse, HttpResponse
        from core.utils.pdf_reports import generar_reporte_requisiciones
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        formato = request.query_params.get('formato', 'json')
        estado = request.query_params.get('estado')
        fecha_inicio = request.query_params.get('fecha_inicio')
        fecha_fin = request.query_params.get('fecha_fin')

        queryset = Requisicion.objects.select_related(
            'centro_origen', 'centro_destino', 'solicitante', 'autorizador'
        ).prefetch_related('detalles', 'detalles__producto')

        filtros = {}
        if estado:
            queryset = queryset.filter(estado=estado)
            filtros['estado'] = estado

        if fecha_inicio:
            queryset = queryset.filter(fecha_solicitud__gte=fecha_inicio)
            filtros['fecha_inicio'] = fecha_inicio

        if fecha_fin:
            queryset = queryset.filter(fecha_solicitud__lte=fecha_fin)
            filtros['fecha_fin'] = fecha_fin

        datos = []
        for req in queryset:
            centro_nombre = ''
            if req.centro_destino:
                centro_nombre = req.centro_destino.nombre
            elif req.centro_origen:
                centro_nombre = req.centro_origen.nombre
            
            # Incluir productos/detalles para PDF y Excel
            productos = []
            for detalle in req.detalles.all():
                productos.append({
                    'clave': detalle.producto.clave if detalle.producto else 'N/A',
                    'nombre': detalle.producto.descripcion if detalle.producto else 'N/A',
                    'cantidad_solicitada': detalle.cantidad_solicitada,
                    'cantidad_autorizada': detalle.cantidad_autorizada or 0,
                    'cantidad_surtida': detalle.cantidad_surtida or 0,
                })
            
            datos.append({
                'id': req.id,
                'folio': req.folio,
                'centro': centro_nombre,
                'centro_nombre': centro_nombre,
                'estado': req.estado,
                'fecha_solicitud': req.fecha_solicitud.isoformat(),
                'total_items': len(productos),
                'usuario_solicita': req.solicitante.get_full_name() if req.solicitante else (req.solicitante.username if req.solicitante else '-'),
                'productos': productos,
            })

        if formato == 'pdf':
            pdf_buffer = generar_reporte_requisiciones(datos, filtros=filtros)
            response = FileResponse(
                pdf_buffer,
                content_type='application/pdf',
                as_attachment=False,
                filename=f'reporte_requisiciones_{timezone.now().strftime("%Y%m%d")}.pdf'
            )
            response['Content-Disposition'] = f'inline; filename="reporte_requisiciones_{timezone.now().strftime("%Y%m%d")}.pdf"'
            return response
        
        if formato == 'excel':
            # Generar Excel con detalles de productos
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Requisiciones'
            
            # Estilos
            header_fill = PatternFill(start_color='632842', end_color='632842', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=10)
            subheader_fill = PatternFill(start_color='9F2241', end_color='9F2241', fill_type='solid')
            subheader_font = Font(bold=True, color='FFFFFF', size=9)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Título
            ws.merge_cells('A1:H1')
            ws['A1'] = 'REPORTE DE REQUISICIONES CON DETALLE'
            ws['A1'].font = Font(bold=True, size=14, color='632842')
            ws['A1'].alignment = Alignment(horizontal='center')
            
            ws.merge_cells('A2:H2')
            ws['A2'] = f'Generado el {timezone.now().strftime("%d/%m/%Y %H:%M")}'
            ws['A2'].font = Font(italic=True, size=10)
            ws['A2'].alignment = Alignment(horizontal='center')
            
            # Encabezados principales
            row = 4
            headers = ['Folio', 'Centro', 'Estado', 'Fecha', 'Solicitante', 'Clave Producto', 'Producto', 'Cant. Solicitada', 'Cant. Autorizada', 'Cant. Surtida']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            row += 1
            
            # Datos con detalles
            for req in datos:
                productos = req.get('productos', [])
                if not productos:
                    # Requisición sin productos
                    ws.cell(row=row, column=1, value=req['folio']).border = border
                    ws.cell(row=row, column=2, value=req['centro_nombre']).border = border
                    ws.cell(row=row, column=3, value=req['estado'].upper()).border = border
                    ws.cell(row=row, column=4, value=req['fecha_solicitud'][:10]).border = border
                    ws.cell(row=row, column=5, value=req['usuario_solicita']).border = border
                    ws.cell(row=row, column=6, value='-').border = border
                    ws.cell(row=row, column=7, value='Sin productos').border = border
                    ws.cell(row=row, column=8, value=0).border = border
                    ws.cell(row=row, column=9, value=0).border = border
                    ws.cell(row=row, column=10, value=0).border = border
                    row += 1
                else:
                    # Primera fila con datos de requisición
                    first_row = True
                    for prod in productos:
                        ws.cell(row=row, column=1, value=req['folio'] if first_row else '').border = border
                        ws.cell(row=row, column=2, value=req['centro_nombre'] if first_row else '').border = border
                        ws.cell(row=row, column=3, value=req['estado'].upper() if first_row else '').border = border
                        ws.cell(row=row, column=4, value=req['fecha_solicitud'][:10] if first_row else '').border = border
                        ws.cell(row=row, column=5, value=req['usuario_solicita'] if first_row else '').border = border
                        ws.cell(row=row, column=6, value=prod['clave']).border = border
                        ws.cell(row=row, column=7, value=prod['nombre']).border = border
                        ws.cell(row=row, column=8, value=prod['cantidad_solicitada']).border = border
                        ws.cell(row=row, column=9, value=prod['cantidad_autorizada']).border = border
                        ws.cell(row=row, column=10, value=prod['cantidad_surtida']).border = border
                        first_row = False
                        row += 1
            
            # Ajustar anchos
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 35
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 12
            ws.column_dimensions['E'].width = 30
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 40
            ws.column_dimensions['H'].width = 15
            ws.column_dimensions['I'].width = 15
            ws.column_dimensions['J'].width = 15
            
            # Respuesta
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="requisiciones_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
            wb.save(response)
            return response

        return Response({
            'reporte': 'requisiciones',
            'fecha_generacion': timezone.now().isoformat(),
            'datos': datos
        })


# ============================================
# AUTENTICACION JWT
# ============================================

class CustomTokenObtainPairView(TokenObtainPairView):
    """
    View personalizado para login que retorna tokens + datos del usuario.
    Usa CustomTokenObtainPairSerializer de serializers_jwt.py
    ✅ Incluye rate limiting para prevenir fuerza bruta.
    """
    from core.serializers_jwt import CustomTokenObtainPairSerializer
    serializer_class = CustomTokenObtainPairSerializer
    throttle_classes = [LoginThrottle]

    def _verify_captcha(self, token, remote_ip=None):
        """
        Valida el token de reCAPTCHA contra el servicio de Google.
        Retorna True si es válido o si la validación está deshabilitada.
        """
        if not settings.RECAPTCHA_ENABLED:
            return True
        secret = settings.RECAPTCHA_SECRET_KEY
        if not secret:
            logger.warning("RECAPTCHA_ENABLED sin RECAPTCHA_SECRET_KEY configurado")
            return False
        try:
            resp = requests.post(
                'https://www.google.com/recaptcha/api/siteverify',
                data={'secret': secret, 'response': token, 'remoteip': remote_ip},
                timeout=5
            )
            data = resp.json()
            return bool(data.get('success'))
        except Exception as exc:
            logger.error(f"Error verificando reCAPTCHA: {exc}")
            return False

    def post(self, request, *args, **kwargs):
        if settings.RECAPTCHA_ENABLED:
            captcha_token = request.data.get('captcha_token') or request.data.get('captcha')
            if not captcha_token:
                return Response({'error': 'Captcha requerido'}, status=status.HTTP_400_BAD_REQUEST)
            if not self._verify_captcha(captcha_token, request.META.get('REMOTE_ADDR')):
                return Response({'error': 'Captcha inválido'}, status=status.HTTP_400_BAD_REQUEST)
        return super().post(request, *args, **kwargs)


class LogoutView(APIView):
    """Endpoint para logout (blacklist del refresh token)."""
    permission_classes = [IsAuthenticated]

    def post(self, request):
        try:
            refresh_token = request.data.get("refresh")
            if not refresh_token:
                # Si no se envía refresh, solo confirmamos logout
                # (el access token expirará naturalmente)
                return Response({"message": "Logout exitoso (access token invalidado)"}, status=status.HTTP_200_OK)
            
            token = RefreshToken(refresh_token)
            token.blacklist()
            return Response({"message": "Logout exitoso (refresh blacklisted)"}, status=status.HTTP_200_OK)
        except Exception as e:
            # Si el refresh ya expiró o es inválido, igual consideramos logout exitoso
            return Response({"message": "Logout completado (refresh inválido/expirado)", "detail": str(e)}, status=status.HTTP_200_OK)


class DevAutoLoginView(APIView):
    """
    SOLO DESARROLLO: Autologin sin credenciales.
    Automaticamente deshabilitado en produccion (DEBUG=False).
    """
    permission_classes = [AllowAny]

    def post(self, request):
        from django.conf import settings
        if not settings.DEBUG:
            return Response(
                {'error': 'Este endpoint solo esta disponible en modo desarrollo'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        username = request.data.get('username', 'admin')
        
        try:
            user = User.objects.get(username=username)
        except User.DoesNotExist:
            return Response(
                {'error': f'Usuario {username} no encontrado'},
                status=status.HTTP_404_NOT_FOUND
            )

        refresh = RefreshToken.for_user(user)
        return Response({
            'refresh': str(refresh),
            'access': str(refresh.access_token),
            'user': UserSerializer(user).data
        })

# NOTA: UserProfileView eliminada - usar /api/usuarios/me/ del UserViewSet
# class UserProfileView eliminada (código muerto, duplicaba funcionalidad)


class ConfiguracionSistemaViewSet(viewsets.ViewSet):
    """
    ViewSet para la configuración global del sistema (colores del tema).
    
    GET  /api/configuracion/tema/  - Obtiene la configuración actual (público)
    PUT  /api/configuracion/tema/  - Actualiza la configuración (solo superusuario)
    POST /api/configuracion/tema/aplicar-tema/  - Aplica un tema predefinido (solo superusuario)
    POST /api/configuracion/tema/restablecer/    - Restablece al tema por defecto (solo superusuario)
    """
    
    def get_permissions(self):
        """
        GET es público (para cargar el tema al inicio).
        PUT, POST requieren superusuario.
        """
        if self.action in ['retrieve', 'list']:
            return [AllowAny()]
        return [IsAuthenticated()]
    
    def list(self, request):
        """
        GET /api/configuracion/
        Retorna la configuración actual del sistema.
        Público para que el frontend pueda cargar configuraciones al iniciar.
        """
        # ConfiguracionSistema es un modelo clave-valor, retornar todas las configuraciones públicas
        configs = ConfiguracionSistema.objects.filter(es_publica=True)
        serializer = ConfiguracionSistemaSerializer(configs, many=True)
        return Response(serializer.data)
    
    def retrieve(self, request, pk=None):
        """
        GET /api/configuracion/{clave}/
        Retorna una configuración específica por clave.
        """
        try:
            config = ConfiguracionSistema.objects.get(clave=pk)
            if not config.es_publica and not request.user.is_superuser:
                return Response(
                    {'error': 'Configuración no accesible'},
                    status=status.HTTP_403_FORBIDDEN
                )
            serializer = ConfiguracionSistemaSerializer(config)
            return Response(serializer.data)
        except ConfiguracionSistema.DoesNotExist:
            return Response(
                {'error': 'Configuración no encontrada'},
                status=status.HTTP_404_NOT_FOUND
            )
    
    def update(self, request, pk=None):
        """
        PUT /api/configuracion/{clave}/
        Actualiza una configuración del sistema.
        Solo superusuarios pueden modificar.
        """
        if not request.user.is_superuser:
            return Response(
                {'error': 'Solo superusuarios pueden modificar la configuración del sistema'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        try:
            config = ConfiguracionSistema.objects.get(clave=pk)
        except ConfiguracionSistema.DoesNotExist:
            return Response(
                {'error': 'Configuración no encontrada'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        serializer = ConfiguracionSistemaSerializer(config, data=request.data, partial=True)
        
        if serializer.is_valid():
            serializer.save()
            logger.info(f"Configuración '{pk}' actualizada por {request.user.username}")
            return Response(serializer.data)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['post'], url_path='bulk-update')
    def bulk_update(self, request):
        """
        POST /api/configuracion/bulk-update/
        Actualiza múltiples configuraciones.
        Body: { "configuraciones": [{"clave": "...", "valor": "..."}, ...] }
        """
        if not request.user.is_superuser:
            return Response(
                {'error': 'Solo superusuarios pueden modificar la configuración'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        configuraciones = request.data.get('configuraciones', [])
        actualizadas = 0
        errores = []
        
        for cfg in configuraciones:
            clave = cfg.get('clave')
            valor = cfg.get('valor')
            if not clave:
                continue
            try:
                config, created = ConfiguracionSistema.objects.update_or_create(
                    clave=clave,
                    defaults={'valor': str(valor) if valor is not None else ''}
                )
                actualizadas += 1
            except Exception as e:
                errores.append({'clave': clave, 'error': str(e)})
        
        logger.info(f"{actualizadas} configuraciones actualizadas por {request.user.username}")
        return Response({
            'mensaje': f'{actualizadas} configuraciones actualizadas',
            'errores': errores
        })
    
    @action(detail=False, methods=['post'], url_path='aplicar-tema')
    def aplicar_tema(self, request):
        """
        POST /api/configuracion/tema/aplicar-tema/
        Aplica un tema predefinido al sistema.
        Solo superusuarios pueden modificar.
        """
        try:
            if not request.user.is_superuser:
                return Response(
                    {'error': 'Solo superusuarios pueden aplicar temas'},
                    status=status.HTTP_403_FORBIDDEN
                )
            
            tema_nombre = request.data.get('tema')
            if not tema_nombre:
                return Response(
                    {'error': 'Se requiere el nombre del tema'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Definición de temas predefinidos con sus colores
            TEMAS_PREDEFINIDOS = {
                'default': {
                    'nombre': 'Por Defecto (Institucional)',
                    'color_primario': '#9F2241',
                    'color_primario_hover': '#6B1839',
                    'color_secundario': '#424242',
                    'color_secundario_hover': '#2E2E2E',
                    'color_exito': '#4CAF50',
                    'color_exito_hover': '#3d8b40',
                    'color_alerta': '#FF9800',
                    'color_alerta_hover': '#e68900',
                    'color_error': '#F44336',
                    'color_error_hover': '#d32f2f',
                    'color_info': '#2196F3',
                    'color_info_hover': '#1976D2',
                    'color_fondo_principal': '#F5F5F5',
                    'color_fondo_sidebar': '#9F2241',
                    'color_fondo_header': '#9F2241',
                    'color_texto_principal': '#212121',
                    'color_texto_sidebar': '#FFFFFF',
                    'color_texto_header': '#FFFFFF',
                    'color_texto_links': '#9F2241',
                    'color_borde_inputs': '#d1d5db',
                    'color_borde_focus': '#9F2241',
                    'reporte_color_encabezado': '#9F2241',
                    'reporte_color_texto': '#1f2937',
                },
                'dark': {
                    'nombre': 'Oscuro',
                    'color_primario': '#1F2937',
                    'color_primario_hover': '#111827',
                    'color_secundario': '#374151',
                    'color_secundario_hover': '#1F2937',
                    'color_exito': '#10B981',
                    'color_exito_hover': '#059669',
                    'color_alerta': '#F59E0B',
                    'color_alerta_hover': '#D97706',
                    'color_error': '#EF4444',
                    'color_error_hover': '#DC2626',
                    'color_info': '#3B82F6',
                    'color_info_hover': '#2563EB',
                    'color_fondo_principal': '#111827',
                    'color_fondo_sidebar': '#1F2937',
                    'color_fondo_header': '#1F2937',
                    'color_texto_principal': '#F9FAFB',
                    'color_texto_sidebar': '#F9FAFB',
                    'color_texto_header': '#F9FAFB',
                    'color_texto_links': '#60A5FA',
                    'color_borde_inputs': '#4B5563',
                    'color_borde_focus': '#3B82F6',
                    'reporte_color_encabezado': '#1F2937',
                    'reporte_color_texto': '#1f2937',
                },
                'green': {
                    'nombre': 'Verde Institucional',
                    'color_primario': '#166534',
                    'color_primario_hover': '#14532D',
                    'color_secundario': '#15803D',
                    'color_secundario_hover': '#166534',
                    'color_exito': '#22C55E',
                    'color_exito_hover': '#16A34A',
                    'color_alerta': '#EAB308',
                    'color_alerta_hover': '#CA8A04',
                    'color_error': '#DC2626',
                    'color_error_hover': '#B91C1C',
                    'color_info': '#0EA5E9',
                    'color_info_hover': '#0284C7',
                    'color_fondo_principal': '#F0FDF4',
                    'color_fondo_sidebar': '#166534',
                    'color_fondo_header': '#166534',
                    'color_texto_principal': '#14532D',
                    'color_texto_sidebar': '#FFFFFF',
                    'color_texto_header': '#FFFFFF',
                    'color_texto_links': '#166534',
                    'color_borde_inputs': '#86EFAC',
                    'color_borde_focus': '#22C55E',
                    'reporte_color_encabezado': '#166534',
                    'reporte_color_texto': '#14532D',
                },
                'purple': {
                    'nombre': 'Púrpura',
                    'color_primario': '#7C3AED',
                    'color_primario_hover': '#6D28D9',
                    'color_secundario': '#8B5CF6',
                    'color_secundario_hover': '#7C3AED',
                    'color_exito': '#10B981',
                    'color_exito_hover': '#059669',
                    'color_alerta': '#F59E0B',
                    'color_alerta_hover': '#D97706',
                    'color_error': '#EF4444',
                    'color_error_hover': '#DC2626',
                    'color_info': '#3B82F6',
                    'color_info_hover': '#2563EB',
                    'color_fondo_principal': '#FAF5FF',
                    'color_fondo_sidebar': '#7C3AED',
                    'color_fondo_header': '#7C3AED',
                    'color_texto_principal': '#581C87',
                    'color_texto_sidebar': '#FFFFFF',
                    'color_texto_header': '#FFFFFF',
                    'color_texto_links': '#7C3AED',
                    'color_borde_inputs': '#C4B5FD',
                    'color_borde_focus': '#8B5CF6',
                    'reporte_color_encabezado': '#7C3AED',
                    'reporte_color_texto': '#581C87',
                },
            }
            
            if tema_nombre not in TEMAS_PREDEFINIDOS:
                return Response(
                    {'error': f'Tema "{tema_nombre}" no es válido. Opciones: {list(TEMAS_PREDEFINIDOS.keys())}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            tema_colores = TEMAS_PREDEFINIDOS[tema_nombre]
            
            # Intentar actualizar TemaGlobal si existe
            try:
                from core.models import TemaGlobal
                tema_global = TemaGlobal.objects.filter(es_activo=True).first()
                if not tema_global:
                    tema_global = TemaGlobal.objects.first()
                
                if tema_global:
                    # Actualizar los colores del tema global
                    for campo, valor in tema_colores.items():
                        if hasattr(tema_global, campo):
                            setattr(tema_global, campo, valor)
                    tema_global.save()
                    logger.info(f"TemaGlobal actualizado con tema '{tema_nombre}' por {request.user.username}")
            except Exception as e:
                logger.warning(f"No se pudo actualizar TemaGlobal: {e}")
            
            # También actualizar ConfiguracionSistema para compatibilidad
            configuraciones_actualizadas = 0
            for clave, valor in tema_colores.items():
                try:
                    config, created = ConfiguracionSistema.objects.update_or_create(
                        clave=clave,
                        defaults={'valor': str(valor), 'es_publica': True}
                    )
                    configuraciones_actualizadas += 1
                except Exception as e:
                    logger.warning(f"No se pudo actualizar ConfiguracionSistema '{clave}': {e}")
            
            # Guardar el tema activo
            ConfiguracionSistema.objects.update_or_create(
                clave='tema_activo',
                defaults={'valor': tema_nombre, 'es_publica': True}
            )
            
            logger.info(f"Tema '{tema_nombre}' aplicado por {request.user.username}")
            
            # Generar respuesta con CSS variables para el frontend
            css_variables = {
                '--color-primary': tema_colores.get('color_primario', '#9F2241'),
                '--color-primary-hover': tema_colores.get('color_primario_hover', '#6B1839'),
                '--color-primary-light': f"rgba({self._hex_to_rgb(tema_colores.get('color_primario', '#9F2241'))}, 0.2)",
                '--color-secondary': tema_colores.get('color_secundario', '#424242'),
                '--color-accent': '#BC955C',
                '--color-background': tema_colores.get('color_fondo_principal', '#F5F5F5'),
                '--color-sidebar-bg': tema_colores.get('color_fondo_sidebar', '#9F2241'),
                '--color-header-bg': tema_colores.get('color_fondo_header', '#9F2241'),
                '--color-card-bg': '#FFFFFF',
                '--color-text': tema_colores.get('color_texto_principal', '#212121'),
                '--color-text-secondary': '#757575',
                '--color-sidebar-text': tema_colores.get('color_texto_sidebar', '#FFFFFF'),
                '--color-header-text': tema_colores.get('color_texto_header', '#FFFFFF'),
                '--color-success': tema_colores.get('color_exito', '#4CAF50'),
                '--color-warning': tema_colores.get('color_alerta', '#FF9800'),
                '--color-error': tema_colores.get('color_error', '#F44336'),
                '--color-info': tema_colores.get('color_info', '#2196F3'),
            }
            
            return Response({
                'mensaje': f'Tema "{tema_nombre}" aplicado correctamente',
                'configuracion': {
                    'tema_activo': tema_nombre,
                    'nombre_sistema': 'Sistema de Farmacia Penitenciaria',
                    'css_variables': css_variables,
                    **tema_colores
                }
            })
            
        except Exception as e:
            logger.error(f"Error al aplicar tema: {str(e)}")
            return Response(
                {'error': f'Error interno al aplicar tema: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def _hex_to_rgb(self, hex_color):
        """Convierte color hexadecimal a formato RGB para CSS rgba()"""
        try:
            hex_color = hex_color.lstrip('#')
            return ', '.join(str(int(hex_color[i:i+2], 16)) for i in (0, 2, 4))
        except Exception:
            return '159, 34, 65'  # Default institucional
    
    @action(detail=False, methods=['post'], url_path='restablecer')
    def restablecer(self, request):
        """
        POST /api/configuracion/tema/restablecer/
        Restablece el tema a los valores institucionales por defecto.
        Solo superusuarios pueden modificar.
        """
        try:
            if not request.user.is_superuser:
                return Response(
                    {'error': 'Solo superusuarios pueden restablecer el tema'},
                    status=status.HTTP_403_FORBIDDEN
                )
            
            # Usar el mismo método aplicar_tema con 'default'
            request._request.method = 'POST'
            from django.http import QueryDict
            original_data = request.data
            request._full_data = {'tema': 'default'}
            
            result = self.aplicar_tema(request)
            
            # Restaurar datos originales
            request._full_data = original_data
            
            logger.info(f"Tema restablecido a institucional por {request.user.username}")
            return result
            
        except Exception as e:
            logger.error(f"Error al restablecer tema: {str(e)}")
            return Response(
                {'error': f'Error interno al restablecer tema: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    # End of ConfiguracionSistemaViewSet


# ============================================================================
# VIEWSET PARA TEMA GLOBAL
# ============================================================================

from core.models import TemaGlobal
from core.serializers import TemaGlobalSerializer, TemaGlobalPublicoSerializer


class TemaGlobalViewSet(viewsets.ViewSet):
    """Gestión del tema global acorde al esquema actual (campos simples)."""

    def get_permissions(self):
        if self.action in ['tema_activo']:
            return [AllowAny()]
        # Para acciones de modificación, usar IsFarmaciaRole (admin + farmacia)
        if self.action in ['update', 'restablecer_institucional', 'subir_logo', 'eliminar_logo']:
            return [IsAuthenticated(), IsFarmaciaRole()]
        return [IsAuthenticated()]

    def _get_tema_activo(self):
        tema = TemaGlobal.objects.filter(es_activo=True).first()
        if tema:
            return tema
        return TemaGlobal.objects.first()

    @action(detail=False, methods=['get'], url_path='activo')
    def tema_activo(self, request):
        tema = self._get_tema_activo()
        if not tema:
            return Response({'error': 'No hay tema configurado'}, status=status.HTTP_404_NOT_FOUND)
        serializer = TemaGlobalPublicoSerializer(tema, context={'request': request})
        return Response(serializer.data)

    def list(self, request):
        tema = self._get_tema_activo()
        if not tema:
            return Response({'error': 'No hay tema configurado'}, status=status.HTTP_404_NOT_FOUND)
        serializer = TemaGlobalSerializer(tema, context={'request': request})
        return Response(serializer.data)

    def update(self, request, pk=None):
        # Permiso controlado por get_permissions() -> IsFarmaciaRole
        tema = self._get_tema_activo()
        if not tema:
            tema = TemaGlobal.objects.create(nombre='Tema Sistema', es_activo=True)

        serializer = TemaGlobalSerializer(tema, data=request.data, partial=True, context={'request': request})
        if serializer.is_valid():
            tema_guardado = serializer.save()
            AuditoriaLog.objects.create(
                usuario=request.user,
                accion='UPDATE',
                modelo='TemaGlobal',
                objeto_id=str(tema_guardado.id),
                datos_nuevos=serializer.data,
                detalles={'objeto_repr': f'Tema global: {tema_guardado.nombre}'}
            )
            logger.info(f"Tema global actualizado por {request.user.username}")
            return Response({
                'mensaje': 'Tema actualizado correctamente',
                'tema': TemaGlobalSerializer(tema_guardado, context={'request': request}).data
            })

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'], url_path='restablecer')
    def restablecer_institucional(self, request):
        """
        POST /api/tema/restablecer/
        Restablece el tema global a valores institucionales por defecto.
        Permiso: admin o farmacia (IsFarmaciaRole)
        """
        # Permiso controlado por get_permissions() -> IsFarmaciaRole
        try:
            tema = self._get_tema_activo()
            if not tema:
                tema = TemaGlobal.objects.create(nombre='Tema Institucional', es_activo=True)

            # Valores institucionales por defecto (guinda)
            valores_institucionales = {
                'nombre': 'Tema Institucional',
                'color_primario': '#9F2241',
                'color_primario_hover': '#6B1839',
                'color_secundario': '#424242',
                'color_secundario_hover': '#212121',
                'color_exito': '#4CAF50',
                'color_exito_hover': '#388E3C',
                'color_alerta': '#FF9800',
                'color_alerta_hover': '#F57C00',
                'color_error': '#F44336',
                'color_error_hover': '#D32F2F',
                'color_info': '#2196F3',
                'color_info_hover': '#1976D2',
                'color_fondo_principal': '#F5F5F5',
                'color_fondo_sidebar': '#9F2241',
                'color_fondo_header': '#9F2241',
                'color_texto_principal': '#212121',
                'color_texto_sidebar': '#FFFFFF',
                'color_texto_header': '#FFFFFF',
                'color_texto_links': '#9F2241',
                'color_borde_inputs': '#E0E0E0',
                'color_borde_focus': '#9F2241',
                'reporte_color_encabezado': '#9F2241',
                'reporte_color_texto': '#FFFFFF',
            }

            for campo, valor in valores_institucionales.items():
                setattr(tema, campo, valor)
            tema.save()

            # Registrar en auditoría
            AuditoriaLog.objects.create(
                usuario=request.user,
                accion='UPDATE',
                modelo='TemaGlobal',
                objeto_id=str(tema.id),
                datos_nuevos=valores_institucionales,
                detalles={'objeto_repr': 'Tema restablecido a institucional'}
            )

            logger.info(f"Tema restablecido a institucional por {request.user.username}")

            return Response({
                'mensaje': 'Tema restablecido a valores institucionales',
                'tema': TemaGlobalSerializer(tema, context={'request': request}).data
            })

        except Exception as e:
            logger.error(f"Error al restablecer tema: {str(e)}")
            return Response(
                {'error': f'Error interno: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['post'], url_path='subir-logo/(?P<tipo>[^/.]+)')
    def subir_logo(self, request, tipo=None):
        """
        POST /api/tema/subir-logo/{tipo}/
        Sube un logo para el tema global.
        Tipos válidos: header, login, reportes, favicon
        Permiso: admin o farmacia (IsFarmaciaRole)
        """
        # Permiso controlado por get_permissions() -> IsFarmaciaRole

        tipos_validos = ['header', 'login', 'reportes', 'favicon']
        if tipo not in tipos_validos:
            return Response(
                {'error': f'Tipo de logo inválido. Opciones: {tipos_validos}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        archivo = request.FILES.get('archivo') or request.FILES.get('logo')
        if not archivo:
            return Response(
                {'error': 'No se proporcionó archivo. Enviar como "archivo" o "logo".'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Validar tipo de archivo
        tipos_permitidos = ['image/png', 'image/jpeg', 'image/jpg', 'image/webp', 'image/x-icon', 'image/vnd.microsoft.icon']
        if archivo.content_type not in tipos_permitidos:
            return Response(
                {'error': f'Tipo de archivo no permitido: {archivo.content_type}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Validar tamaño (2MB máximo)
        if archivo.size > 2 * 1024 * 1024:
            return Response(
                {'error': 'El archivo no puede superar 2MB'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            import os
            from django.conf import settings
            from django.core.files.storage import default_storage

            tema = self._get_tema_activo()
            if not tema:
                tema = TemaGlobal.objects.create(nombre='Tema Sistema', es_activo=True)

            # Guardar archivo
            extension = archivo.name.split('.')[-1]
            nombre_archivo = f'tema/logos/{tipo}_{tema.id}.{extension}'
            
            # Eliminar archivo anterior si existe
            campo_url = f'logo_url' if tipo == 'header' else f'favicon_url' if tipo == 'favicon' else None
            if campo_url and hasattr(tema, campo_url):
                url_anterior = getattr(tema, campo_url)
                if url_anterior:
                    try:
                        default_storage.delete(url_anterior.replace('/media/', ''))
                    except Exception:
                        pass

            # Guardar nuevo archivo
            ruta_guardada = default_storage.save(nombre_archivo, archivo)
            url_completa = f'{settings.MEDIA_URL}{ruta_guardada}'

            # Actualizar campo correspondiente en TemaGlobal
            # Nota: Solo logo_url y favicon_url existen en el modelo actual
            if tipo == 'header':
                tema.logo_url = url_completa
            elif tipo == 'favicon':
                tema.favicon_url = url_completa
            # Para otros tipos, podríamos necesitar campos adicionales en el modelo

            tema.save()

            logger.info(f"Logo {tipo} actualizado por {request.user.username}")

            return Response({
                'mensaje': f'Logo {tipo} actualizado correctamente',
                'url': url_completa,
                'tema': TemaGlobalSerializer(tema, context={'request': request}).data
            })

        except Exception as e:
            logger.error(f"Error al subir logo {tipo}: {str(e)}")
            return Response(
                {'error': f'Error interno: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['delete'], url_path='eliminar-logo/(?P<tipo>[^/.]+)')
    def eliminar_logo(self, request, tipo=None):
        """
        DELETE /api/tema/eliminar-logo/{tipo}/
        Elimina un logo del tema global.
        Tipos válidos: header, login, reportes, favicon
        Permiso: admin o farmacia (IsFarmaciaRole)
        """
        # Permiso controlado por get_permissions() -> IsFarmaciaRole

        tipos_validos = ['header', 'login', 'reportes', 'favicon']
        if tipo not in tipos_validos:
            return Response(
                {'error': f'Tipo de logo inválido. Opciones: {tipos_validos}'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            from django.core.files.storage import default_storage

            tema = self._get_tema_activo()
            if not tema:
                return Response(
                    {'error': 'No hay tema configurado'},
                    status=status.HTTP_404_NOT_FOUND
                )

            # Obtener y limpiar campo correspondiente
            campo_url = 'logo_url' if tipo == 'header' else 'favicon_url' if tipo == 'favicon' else None
            
            if campo_url and hasattr(tema, campo_url):
                url_anterior = getattr(tema, campo_url)
                if url_anterior:
                    try:
                        # Intentar eliminar archivo físico
                        ruta = url_anterior.replace('/media/', '')
                        if default_storage.exists(ruta):
                            default_storage.delete(ruta)
                    except Exception as e:
                        logger.warning(f"No se pudo eliminar archivo físico: {e}")
                    
                    # Limpiar campo en BD
                    setattr(tema, campo_url, None)
                    tema.save()

            logger.info(f"Logo {tipo} eliminado por {request.user.username}")

            return Response({
                'mensaje': f'Logo {tipo} eliminado correctamente',
                'tema': TemaGlobalSerializer(tema, context={'request': request}).data
            })

        except Exception as e:
            logger.error(f"Error al eliminar logo {tipo}: {str(e)}")
            return Response(
                {'error': f'Error interno: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


# =============================================================================
# PRODUCTO IMAGEN VIEWSET
# =============================================================================

class ProductoImagenViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestionar imagenes de productos.
    """
    permission_classes = [IsAuthenticated]
    pagination_class = StandardResultsSetPagination
    
    def get_queryset(self):
        from core.models import ProductoImagen
        queryset = ProductoImagen.objects.select_related('producto').all()
        
        # Filtrar por producto si se especifica
        producto_id = self.request.query_params.get('producto')
        if producto_id:
            queryset = queryset.filter(producto_id=producto_id)
        
        return queryset.order_by('orden', '-es_principal')
    
    def get_serializer_class(self):
        from core.serializers import ProductoImagenSerializer
        return ProductoImagenSerializer
    
    def perform_create(self, serializer):
        serializer.save()
        logger.info(f"Imagen de producto creada por {self.request.user.username}")
    
    @action(detail=True, methods=['post'], url_path='set-principal')
    def set_principal(self, request, pk=None):
        """Establecer una imagen como principal."""
        from core.models import ProductoImagen
        imagen = self.get_object()
        
        # Quitar es_principal de otras imagenes del mismo producto
        ProductoImagen.objects.filter(
            producto_id=imagen.producto_id
        ).update(es_principal=False)
        
        # Establecer esta como principal
        imagen.es_principal = True
        imagen.save()
        
        return Response({'mensaje': 'Imagen establecida como principal'})


# =============================================================================
# LOTE DOCUMENTO VIEWSET
# =============================================================================

class LoteDocumentoViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestionar documentos de lotes (facturas, contratos).
    """
    permission_classes = [IsAuthenticated]
    pagination_class = StandardResultsSetPagination
    
    def get_queryset(self):
        from core.models import LoteDocumento
        queryset = LoteDocumento.objects.select_related('lote', 'lote__producto', 'created_by').all()
        
        # Filtrar por lote si se especifica
        lote_id = self.request.query_params.get('lote')
        if lote_id:
            queryset = queryset.filter(lote_id=lote_id)
        
        # Filtrar por tipo de documento
        tipo = self.request.query_params.get('tipo_documento')
        if tipo:
            queryset = queryset.filter(tipo_documento=tipo)
        
        return queryset.order_by('-created_at')
    
    def get_serializer_class(self):
        from core.serializers import LoteDocumentoSerializer
        return LoteDocumentoSerializer
    
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)
        logger.info(f"Documento de lote creado por {self.request.user.username}")


# =============================================================================
# DONACION VIEWSET
# =============================================================================

class DonacionViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestionar donaciones de medicamentos.
    Solo ADMIN y FARMACIA pueden crear/editar/procesar.
    VISTA puede consultar en solo lectura SI tiene perm_donaciones.
    
    ISS-FIX SEGURIDAD:
    - Lectura requiere perm_donaciones (no solo autenticación)
    - Scoping por centro: usuarios de centro solo ven donaciones a SU centro
    """
    pagination_class = StandardResultsSetPagination
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['numero', 'donante_nombre', 'donante_rfc']
    ordering_fields = ['fecha_donacion', 'fecha_recepcion', 'created_at']
    ordering = ['-fecha_recepcion']
    
    def get_permissions(self):
        """
        ISS-FIX: Permisos según la acción:
        - list, retrieve: HasDonacionesPermission (verifica perm_donaciones)
        - create, update, destroy, acciones: IsFarmaciaRole (admin/farmacia)
        """
        from core.permissions import HasDonacionesPermission, HasDonacionesWritePermission
        
        if self.action in ['list', 'retrieve']:
            # ISS-FIX: Requiere perm_donaciones, no solo autenticación
            return [IsAuthenticated(), HasDonacionesPermission()]
        # Crear, editar, eliminar, acciones solo admin/farmacia
        return [IsAuthenticated(), HasDonacionesWritePermission()]
    
    def get_queryset(self):
        from core.models import Donacion
        from core.permissions import _has_role
        
        # ISS-FIX: Incluir producto_donacion del nuevo catálogo
        queryset = Donacion.objects.select_related(
            'centro_destino', 'recibido_por'
        ).prefetch_related(
            'detalles', 
            'detalles__producto_donacion',  # Nuevo catálogo de donaciones
            'detalles__producto'  # Legacy (compatibilidad)
        )
        
        # ISS-FIX SEGURIDAD: Scoping por centro
        # Admin/farmacia/vista ven todo; usuarios de centro solo su centro
        user = self.request.user
        if not _has_role(user, ['admin', 'farmacia', 'vista']):
            # Usuario de centro: solo donaciones DESTINADAS a su centro
            user_centro = getattr(user, 'centro', None)
            if user_centro:
                queryset = queryset.filter(centro_destino=user_centro)
                logger.debug(
                    f"ISS-FIX: Usuario {user.username} de centro {user_centro.nombre} "
                    f"accediendo a donaciones filtradas"
                )
            else:
                # Usuario sin centro no ve nada
                logger.warning(
                    f"ISS-FIX: Usuario {user.username} sin centro intentó acceder a donaciones"
                )
                return Donacion.objects.none()
        
        # Filtros adicionales por query params
        estado = self.request.query_params.get('estado')
        if estado:
            queryset = queryset.filter(estado=estado)
        
        centro = self.request.query_params.get('centro')
        if centro:
            queryset = queryset.filter(centro_destino_id=centro)
        
        donante_tipo = self.request.query_params.get('donante_tipo')
        if donante_tipo:
            queryset = queryset.filter(donante_tipo=donante_tipo)
        
        # Filtro por rango de fechas
        fecha_desde = self.request.query_params.get('fecha_desde')
        fecha_hasta = self.request.query_params.get('fecha_hasta')
        if fecha_desde:
            queryset = queryset.filter(fecha_donacion__gte=fecha_desde)
        if fecha_hasta:
            queryset = queryset.filter(fecha_donacion__lte=fecha_hasta)
        
        return queryset
    
    def get_serializer_class(self):
        from core.serializers import DonacionSerializer
        return DonacionSerializer
    
    def perform_create(self, serializer):
        donacion = serializer.save(recibido_por=self.request.user)
        logger.info(f"Donacion {donacion.numero} creada por {self.request.user.username}")
    
    def destroy(self, request, *args, **kwargs):
        """
        Eliminar una donación completa.
        Solo se pueden eliminar donaciones que:
        - NO estén procesadas (ya tienen stock activo)
        - NO tengan entregas/salidas registradas
        
        Se eliminan en cascada los detalles asociados.
        """
        from core.models import SalidaDonacion
        
        donacion = self.get_object()
        
        # Verificar que no esté procesada
        if donacion.estado == 'procesada':
            return Response(
                {'error': 'No se puede eliminar una donación procesada. El stock ya está activo en el almacén.'},
                status=status.HTTP_409_CONFLICT
            )
        
        # Verificar que no tenga salidas registradas
        salidas_count = SalidaDonacion.objects.filter(detalle_donacion__donacion=donacion).count()
        if salidas_count > 0:
            return Response(
                {'error': f'No se puede eliminar: la donación tiene {salidas_count} entrega(s) registrada(s).'},
                status=status.HTTP_409_CONFLICT
            )
        
        try:
            numero = donacion.numero
            # Los detalles se eliminan en cascada por la FK on_delete=CASCADE
            donacion.delete()
            logger.info(f"Donacion {numero} eliminada por {request.user.username}")
            return Response({'mensaje': f'Donación {numero} eliminada correctamente'}, status=status.HTTP_200_OK)
        except Exception as e:
            logger.error(f"Error eliminando donación {donacion.numero}: {e}")
            return Response(
                {'error': f'Error al eliminar: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['post'], url_path='recibir')
    def recibir(self, request, pk=None):
        """Marcar una donacion como recibida."""
        donacion = self.get_object()
        
        if donacion.estado != 'pendiente':
            return Response(
                {'error': 'Solo se pueden recibir donaciones pendientes'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Validar que tenga productos
        if not donacion.detalles.exists():
            return Response(
                {'error': 'No se puede recibir una donación sin productos. Agregue al menos un producto antes de continuar.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        donacion.estado = 'recibida'
        donacion.recibido_por = request.user
        donacion.fecha_recepcion = timezone.now()
        donacion.save()
        
        logger.info(f"Donacion {donacion.numero} recibida por {request.user.username}")
        
        from core.serializers import DonacionSerializer
        return Response(DonacionSerializer(donacion).data)
    
    @action(detail=True, methods=['post'], url_path='procesar')
    def procesar(self, request, pk=None):
        """
        Procesar una donacion - activar stock disponible en almacen de donaciones.
        Las donaciones funcionan como ALMACEN SEPARADO, no afectan inventario principal.
        Las salidas se registran mediante SalidaDonacion.
        """
        donacion = self.get_object()
        
        if donacion.estado not in ['pendiente', 'recibida']:
            return Response(
                {'error': 'Solo se pueden procesar donaciones pendientes o recibidas'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Validar que tenga productos
        if not donacion.detalles.exists():
            return Response(
                {'error': 'No se puede procesar una donación sin productos. Agregue al menos un producto antes de continuar.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            with transaction.atomic():
                # Actualizar cantidad_disponible de cada detalle
                for detalle in donacion.detalles.all():
                    # Asegurar que cantidad_disponible = cantidad recibida
                    detalle.cantidad_disponible = detalle.cantidad
                    detalle.save()
                
                # Cambiar estado a procesada
                donacion.estado = 'procesada'
                donacion.save()
                
                logger.info(f"Donacion {donacion.numero} procesada por {request.user.username}")
                
                from core.serializers import DonacionSerializer
                return Response({
                    'mensaje': 'Donacion procesada correctamente. Stock disponible en almacen de donaciones.',
                    'donacion': DonacionSerializer(donacion).data
                })
        
        except Exception as e:
            logger.error(f"Error procesando donacion {donacion.numero}: {str(e)}")
            return Response(
                {'error': f'Error procesando donacion: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['post'], url_path='procesar-todas')
    def procesar_todas(self, request):
        """
        Procesar TODAS las donaciones pendientes o recibidas de una vez.
        Útil después de importaciones masivas.
        """
        from core.models import Donacion
        
        donaciones_pendientes = Donacion.objects.filter(
            estado__in=['pendiente', 'recibida']
        )
        
        if not donaciones_pendientes.exists():
            return Response(
                {'mensaje': 'No hay donaciones pendientes para procesar'},
                status=status.HTTP_200_OK
            )
        
        procesadas = 0
        errores = []
        
        try:
            with transaction.atomic():
                for donacion in donaciones_pendientes:
                    try:
                        # Validar que tenga productos
                        if not donacion.detalles.exists():
                            errores.append({
                                'donacion': donacion.numero,
                                'error': 'La donación no tiene productos'
                            })
                            continue
                        
                        # Actualizar cantidad_disponible de cada detalle
                        for detalle in donacion.detalles.all():
                            detalle.cantidad_disponible = detalle.cantidad
                            detalle.save()
                        
                        # Cambiar estado a procesada
                        donacion.estado = 'procesada'
                        donacion.save()
                        procesadas += 1
                        
                    except Exception as e:
                        errores.append({
                            'donacion': donacion.numero,
                            'error': str(e)
                        })
                
                logger.info(f"Procesamiento masivo por {request.user.username}: {procesadas} donaciones procesadas")
                
                return Response({
                    'mensaje': f'{procesadas} donaciones procesadas correctamente',
                    'procesadas': procesadas,
                    'errores': errores
                })
                
        except Exception as e:
            logger.error(f"Error en procesamiento masivo: {str(e)}")
            return Response(
                {'error': f'Error procesando donaciones: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['post'], url_path='rechazar')
    def rechazar(self, request, pk=None):
        """Rechazar una donacion."""
        donacion = self.get_object()
        
        if donacion.estado == 'procesada':
            return Response(
                {'error': 'No se puede rechazar una donacion ya procesada'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        motivo = request.data.get('motivo', '')
        donacion.estado = 'rechazada'
        donacion.notas = f"{donacion.notas or ''}\n\nRechazada: {motivo}".strip()
        donacion.save()
        
        logger.info(f"Donacion {donacion.numero} rechazada por {request.user.username}")
        
        from core.serializers import DonacionSerializer
        return Response(DonacionSerializer(donacion).data)
    
    @action(detail=False, methods=['get'], url_path='siguiente-numero')
    def siguiente_numero(self, request):
        """
        Genera el siguiente número de donación disponible.
        Formato: DON-YYYY-NNNN
        """
        from core.models import Donacion
        import datetime
        
        year = datetime.datetime.now().year
        prefix = f'DON-{year}-'
        
        # Buscar el último número de este año
        ultima = Donacion.objects.filter(
            numero__startswith=prefix
        ).order_by('-numero').first()
        
        if ultima:
            try:
                # Extraer el número secuencial
                ultimo_num = int(ultima.numero.replace(prefix, ''))
                siguiente = ultimo_num + 1
            except (ValueError, AttributeError):
                siguiente = 1
        else:
            siguiente = 1
        
        numero = f"{prefix}{siguiente:04d}"
        return Response({'numero': numero})
    
    @action(detail=False, methods=['get'], url_path='diagnostico')
    def diagnostico(self, request):
        """
        ISS-DEBUG: Endpoint de diagnóstico para el módulo de donaciones.
        
        Retorna información sobre:
        - Estadísticas de donaciones
        - Posibles errores de datos
        - Estado de la tabla
        """
        from core.models import Donacion, DetalleDonacion
        from django.db import connection
        
        try:
            # Verificar si la tabla existe
            with connection.cursor() as cursor:
                cursor.execute("SELECT COUNT(*) FROM donaciones")
                total_raw = cursor.fetchone()[0]
            
            # Estadísticas via ORM
            todas = Donacion.objects.all()
            por_estado = {}
            for estado in ['pendiente', 'recibida', 'procesada', 'rechazada']:
                por_estado[estado] = todas.filter(estado=estado).count()
            
            # Últimas 5 donaciones
            ultimas = todas.order_by('-created_at')[:5]
            ultimas_data = [
                {
                    'id': d.pk,
                    'numero': d.numero,
                    'donante_nombre': d.donante_nombre,
                    'estado': d.estado,
                    'centro_destino': d.centro_destino.nombre if d.centro_destino else None,
                    'detalles_count': d.detalles.count(),
                    'created_at': d.created_at.isoformat() if d.created_at else None,
                }
                for d in ultimas
            ]
            
            # Verificar detalles
            total_detalles = DetalleDonacion.objects.count()
            detalles_sin_producto = DetalleDonacion.objects.filter(producto__isnull=True).count()
            
            return Response({
                'tabla_existe': True,
                'total_donaciones_raw': total_raw,
                'total_donaciones_orm': todas.count(),
                'por_estado': por_estado,
                'ultimas_donaciones': ultimas_data,
                'detalles': {
                    'total': total_detalles,
                    'sin_producto': detalles_sin_producto,
                },
                'usuario': {
                    'id': request.user.pk,
                    'username': request.user.username,
                    'rol': getattr(request.user, 'rol', 'N/A'),
                }
            })
        except Exception as e:
            logger.exception(f"Error en diagnóstico de donaciones: {e}")
            return Response({
                'error': str(e),
                'tipo_error': type(e).__name__,
                'tabla_existe': False,
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path='exportar-excel')
    def exportar_excel(self, request):
        """
        Exporta las donaciones a Excel con formato profesional.
        Respeta los filtros aplicados en la consulta.
        """
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from django.http import HttpResponse
        from django.utils import timezone
        
        try:
            donaciones = self.get_queryset()
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Donaciones'
            
            # Título
            ws.merge_cells('A1:J1')
            ws['A1'].value = 'REPORTE DE DONACIONES RECIBIDAS'
            ws['A1'].font = Font(bold=True, size=14, color='632842')
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            
            ws.merge_cells('A2:J2')
            ws['A2'].value = f'Generado el {timezone.now().strftime("%d/%m/%Y %H:%M")}'
            ws['A2'].font = Font(size=10, italic=True)
            ws['A2'].alignment = Alignment(horizontal='center')
            
            ws.append([])
            
            # Encabezados
            headers = [
                'Número', 'Donante', 'Tipo Donante', 'RFC', 
                'Fecha Donación', 'Fecha Recepción', 'Centro Destino',
                'Estado', 'Productos', 'Unidades Totales'
            ]
            ws.append(headers)
            
            header_fill = PatternFill(start_color='632842', end_color='632842', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            
            for cell in ws[4]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Datos
            for donacion in donaciones:
                total_productos = donacion.detalles.count() if hasattr(donacion, 'detalles') else 0
                total_unidades = sum(d.cantidad for d in donacion.detalles.all()) if hasattr(donacion, 'detalles') else 0
                
                ws.append([
                    donacion.numero,
                    donacion.donante_nombre,
                    donacion.donante_tipo,
                    donacion.donante_rfc or '',
                    donacion.fecha_donacion.strftime('%d/%m/%Y') if donacion.fecha_donacion else '',
                    donacion.fecha_recepcion.strftime('%d/%m/%Y') if donacion.fecha_recepcion else '',
                    donacion.centro_destino.nombre if donacion.centro_destino else 'Sin asignar',
                    donacion.estado,
                    total_productos,
                    total_unidades
                ])
            
            # Ajustar anchos
            column_widths = [15, 35, 15, 15, 15, 15, 30, 12, 12, 15]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
            
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = f'donaciones_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            wb.save(response)
            return response
            
        except Exception as e:
            logger.error(f"Error exportando donaciones: {e}")
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path='plantilla-excel')
    def plantilla_excel(self, request):
        """
        Genera plantilla Excel SIMPLIFICADA para importar donaciones.
        Solo requiere datos básicos de la donación y los productos.
        Los productos DEBEN existir en el catálogo INDEPENDIENTE de productos de donación.
        """
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from django.http import HttpResponse
        
        try:
            wb = openpyxl.Workbook()
            
            # Estilos
            header_fill = PatternFill(start_color='632842', end_color='632842', fill_type='solid')
            example_font = Font(italic=True, color='888888')
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            # Obtener productos del catálogo INDEPENDIENTE de donaciones para ejemplos
            from core.models import ProductoDonacion
            productos_ejemplo = list(ProductoDonacion.objects.filter(activo=True).order_by('clave')[:3])
            
            # ========== HOJA PRINCIPAL: DONACIONES ==========
            ws = wb.active
            ws.title = 'Donaciones'
            
            # Título
            ws.merge_cells('A1:J1')
            ws['A1'].value = 'PLANTILLA DE DONACIONES (SIMPLIFICADA)'
            ws['A1'].font = Font(bold=True, size=14, color='632842')
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # Instrucciones breves
            ws['A2'].value = '⚠️ Elimine las filas de ejemplo (grises). Los productos DEBEN existir en el Catálogo de Productos de Donación.'
            ws['A2'].font = Font(italic=True, size=10, color='CC0000')
            ws.merge_cells('A2:J2')
            ws.append([])
            
            # Headers - ahora incluye fecha_caducidad y estado_producto
            headers = [
                'numero *',
                'donante *', 
                'tipo_donante',
                'fecha *',
                'producto_clave *',
                'cantidad *',
                'lote',
                'fecha_caducidad',
                'estado_producto',
                'notas'
            ]
            ws.append(headers)
            
            for cell in ws[4]:
                cell.fill = header_fill
                cell.font = Font(bold=True, color='FFFFFF')
                cell.border = thin_border
            
            # Ejemplos simplificados: una fila por producto donado
            # Cada fila tiene: donación + producto
            from datetime import date, timedelta
            fecha_ejemplo = (date.today() + timedelta(days=365)).strftime('%Y-%m-%d')  # 1 año de caducidad
            
            ejemplos = []
            if productos_ejemplo:
                for i, prod in enumerate(productos_ejemplo):
                    ejemplos.append([
                        '[EJEMPLO] DON-001 - ELIMINAR' if i == 0 else '',  # numero (solo primera fila de la donación)
                        '[EJEMPLO] Empresa Donante SA - ELIMINAR' if i == 0 else '',  # donante
                        'empresa' if i == 0 else '',  # tipo
                        '2024-01-15' if i == 0 else '',  # fecha
                        prod.clave,  # clave del producto del catálogo
                        100 + i * 50,  # cantidad
                        f'LOTE-{i+1}',  # lote
                        fecha_ejemplo,  # fecha_caducidad
                        'bueno',  # estado_producto (bueno/regular/deteriorado)
                        '[EJEMPLO] - ELIMINAR' if i == 0 else ''  # notas
                    ])
            else:
                # Si no hay productos, poner ejemplo genérico
                ejemplos.append([
                    '[EJEMPLO] DON-001 - ELIMINAR',
                    '[EJEMPLO] Empresa Donante SA - ELIMINAR',
                    'empresa',
                    '2024-01-15',
                    '⚠️ AGREGAR PRODUCTOS AL CATÁLOGO DE DONACIONES PRIMERO',
                    100,
                    'LOTE-001',
                    fecha_ejemplo,
                    'bueno',
                    '[EJEMPLO] - ELIMINAR'
                ])
            
            for ejemplo in ejemplos:
                ws.append(ejemplo)
            
            # Aplicar formato gris itálica a filas de ejemplo
            for row_num in range(5, 5 + len(ejemplos)):
                for cell in ws[row_num]:
                    cell.font = example_font
                    cell.border = thin_border
            
            # Ajustar anchos (10 columnas ahora)
            column_widths = [30, 35, 15, 12, 15, 10, 15, 15, 15, 30]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
            
            # ========== HOJA CATÁLOGO DE PRODUCTOS DE DONACIÓN ==========
            ws2 = wb.create_sheet(title='Catálogo Productos Donación')
            ws2['A1'].value = 'CATÁLOGO INDEPENDIENTE DE PRODUCTOS DE DONACIÓN'
            ws2['A1'].font = Font(bold=True, size=12, color='632842')
            ws2['A2'].value = 'Use estas claves en la columna "producto_clave" de la hoja Donaciones'
            ws2['A2'].font = Font(italic=True, size=10, color='333333')
            ws2['A3'].value = '⚠️ Este catálogo es INDEPENDIENTE del catálogo principal de la farmacia'
            ws2['A3'].font = Font(italic=True, size=10, color='CC0000')
            ws2.append([])
            ws2.append(['Clave', 'Nombre', 'Unidad', 'Presentación'])
            
            for cell in ws2[5]:
                cell.fill = header_fill
                cell.font = Font(bold=True, color='FFFFFF')
            
            # Mostrar productos del catálogo INDEPENDIENTE de donaciones
            productos = ProductoDonacion.objects.filter(activo=True).order_by('nombre')[:500]
            
            if productos.count() == 0:
                ws2.append(['⚠️ NO HAY PRODUCTOS EN EL CATÁLOGO DE DONACIONES'])
                ws2.append(['Agregue productos primero en: Donaciones → Catálogo Productos'])
            else:
                for prod in productos:
                    ws2.append([prod.clave, prod.nombre, prod.unidad_medida or 'PIEZA', prod.presentacion or ''])
            
            ws2.column_dimensions['A'].width = 20
            ws2.column_dimensions['B'].width = 50
            ws2.column_dimensions['C'].width = 15
            ws2.column_dimensions['D'].width = 25
            
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="plantilla_donaciones.xlsx"'
            wb.save(response)
            return response
            
        except Exception as e:
            logger.error(f"Error generando plantilla: {e}")
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['post'], url_path='importar-excel')
    def importar_excel(self, request):
        """
        Importa donaciones desde archivo Excel SIMPLIFICADO.
        
        Formato simplificado (una sola hoja):
        - numero: Número de la donación
        - donante: Nombre del donante  
        - tipo_donante: empresa, gobierno, ong, particular, otro
        - fecha: Fecha de la donación
        - producto_clave: Clave del producto (del catálogo INDEPENDIENTE de donaciones)
        - cantidad: Cantidad donada
        - lote: Número de lote (opcional)
        - notas: Observaciones (opcional)
        
        Los productos DEBEN existir en el catálogo INDEPENDIENTE de productos de donación.
        """
        import openpyxl
        from django.db import transaction
        from core.models import Donacion, DetalleDonacion, ProductoDonacion, Centro
        
        archivo = request.FILES.get('archivo') or request.FILES.get('file')
        if not archivo:
            return Response({'error': 'No se proporcionó archivo'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            wb = openpyxl.load_workbook(archivo, data_only=True)
            ws = wb.active
            
            # Funciones auxiliares
            def normalizar_header(val):
                if not val:
                    return ''
                return str(val).lower().strip().replace('_', ' ').replace('-', ' ').replace('*', '')
            
            def parse_fecha(val):
                if not val:
                    return None
                if hasattr(val, 'date'):
                    return val.date() if hasattr(val, 'date') else val
                if hasattr(val, 'strftime'):
                    return val
                val_str = str(val).strip()
                for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d', '%d.%m.%Y']:
                    try:
                        return datetime.strptime(val_str, fmt).date()
                    except:
                        continue
                return None
            
            def es_fila_ejemplo(row):
                """Detecta si una fila es un ejemplo que debe ignorarse"""
                for cell in row:
                    if cell:
                        cell_str = str(cell).upper()
                        if '[EJEMPLO]' in cell_str or 'ELIMINAR' in cell_str:
                            return True
                return False
            
            # Detectar encabezados
            ALIASES = {
                'numero': ['numero', 'número', 'num', 'no', 'id', 'codigo', 'folio'],
                'donante': ['donante', 'nombre donante', 'donante nombre', 'nombre', 'razon social'],
                'tipo_donante': ['tipo donante', 'donante tipo', 'tipo'],
                'fecha': ['fecha', 'fecha donacion', 'fecha recepcion'],
                'producto_clave': ['producto clave', 'clave producto', 'producto', 'clave', 'medicamento'],
                'cantidad': ['cantidad', 'cant', 'unidades'],
                'lote': ['lote', 'numero lote', 'no lote'],
                'fecha_caducidad': ['fecha caducidad', 'caducidad', 'vencimiento', 'fecha vencimiento', 'expira'],
                'estado_producto': ['estado producto', 'estado', 'condicion'],
                'notas': ['notas', 'observaciones', 'comentarios'],
            }
            
            # Buscar fila de encabezados
            header_row = 1
            col_map = {}
            for row_num in range(1, min(11, ws.max_row + 1)):
                row_values = [cell.value for cell in ws[row_num]]
                matches = 0
                temp_map = {}
                
                for col_idx, val in enumerate(row_values):
                    header_norm = normalizar_header(val)
                    if not header_norm:
                        continue
                    for field, aliases in ALIASES.items():
                        if field not in temp_map:
                            if header_norm in aliases or any(alias in header_norm for alias in aliases):
                                temp_map[field] = col_idx
                                matches += 1
                                break
                
                if matches >= 3:  # Al menos numero, donante/producto, cantidad
                    header_row = row_num
                    col_map = temp_map
                    break
            
            if not col_map:
                # Usar posiciones por defecto
                col_map = {
                    'numero': 0, 'donante': 1, 'tipo_donante': 2,
                    'fecha': 3, 'producto_clave': 4, 'cantidad': 5,
                    'lote': 6, 'notas': 7
                }
            
            def get_val(row, field, default=None):
                if field not in col_map:
                    return default
                idx = col_map[field]
                if idx < len(row):
                    val = row[idx]
                    return val if val not in [None, '', 'None'] else default
                return default
            
            resultados = {
                'donaciones_creadas': 0,
                'detalles_creados': 0,
                'errores': [],
                'exitos': []
            }
            
            donaciones_map = {}  # numero -> instancia
            filas_procesadas = 0
            filas_vacias = 0
            filas_ejemplo = 0
            
            with transaction.atomic():
                for row_num, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
                    if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                        filas_vacias += 1
                        continue
                    
                    # Ignorar filas de ejemplo
                    if es_fila_ejemplo(row):
                        filas_ejemplo += 1
                        continue
                    
                    filas_procesadas += 1
                    
                    try:
                        numero = get_val(row, 'numero')
                        donante = get_val(row, 'donante')
                        producto_clave = get_val(row, 'producto_clave')
                        cantidad_raw = get_val(row, 'cantidad')
                        
                        # Si no hay número pero hay producto, es continuación de donación anterior
                        if not numero and producto_clave and donaciones_map:
                            # Usar última donación
                            numero = list(donaciones_map.keys())[-1]
                        
                        if not numero:
                            continue  # Fila vacía o incompleta
                        
                        numero = str(numero).strip()
                        
                        # Crear donación si no existe
                        if numero not in donaciones_map:
                            if not donante:
                                resultados['errores'].append({
                                    'fila': row_num,
                                    'error': f'Donación {numero}: nombre del donante es obligatorio'
                                })
                                continue
                            
                            # Verificar si ya existe en BD
                            donacion_existente = Donacion.objects.filter(numero=numero).first()
                            if donacion_existente:
                                donaciones_map[numero] = donacion_existente
                            else:
                                # Parsear fecha
                                fecha = parse_fecha(get_val(row, 'fecha')) or timezone.now().date()
                                
                                # Normalizar tipo
                                tipo = str(get_val(row, 'tipo_donante', 'empresa')).lower().strip()
                                if tipo not in ['empresa', 'gobierno', 'ong', 'particular', 'otro']:
                                    tipo = 'empresa'
                                
                                donacion = Donacion.objects.create(
                                    numero=numero,
                                    donante_nombre=str(donante).strip()[:200],
                                    donante_tipo=tipo,
                                    fecha_donacion=fecha,
                                    recibido_por=request.user,
                                    estado='pendiente'
                                )
                                donaciones_map[numero] = donacion
                                resultados['donaciones_creadas'] += 1
                        
                        # Agregar producto si existe
                        if producto_clave:
                            producto_clave = str(producto_clave).strip()
                            
                            # Buscar en catálogo INDEPENDIENTE de productos de donación
                            producto_donacion = ProductoDonacion.objects.filter(clave__iexact=producto_clave, activo=True).first()
                            if not producto_donacion:
                                producto_donacion = ProductoDonacion.objects.filter(nombre__icontains=producto_clave, activo=True).first()
                            
                            if not producto_donacion:
                                resultados['errores'].append({
                                    'fila': row_num,
                                    'error': f'Producto "{producto_clave}" no encontrado en el catálogo de donaciones. Agregue el producto en Donaciones → Catálogo Productos.'
                                })
                                continue
                            
                            # Parsear cantidad
                            try:
                                cantidad = int(float(cantidad_raw)) if cantidad_raw else 0
                            except:
                                cantidad = 0
                            
                            if cantidad <= 0:
                                resultados['errores'].append({
                                    'fila': row_num,
                                    'error': 'Cantidad debe ser mayor a 0'
                                })
                                continue
                            
                            donacion = donaciones_map[numero]
                            
                            # Parsear fecha de caducidad si existe
                            fecha_caducidad_raw = get_val(row, 'fecha_caducidad')
                            fecha_caducidad = parse_fecha(fecha_caducidad_raw) if fecha_caducidad_raw else None
                            
                            # VALIDACIÓN: Fechas de caducidad no pueden estar más de 8 años en el futuro
                            # Usa relativedelta para considerar años bisiestos correctamente
                            if fecha_caducidad:
                                from dateutil.relativedelta import relativedelta
                                fecha_actual = date.today()
                                fecha_maxima = fecha_actual + relativedelta(years=8)
                                
                                if fecha_caducidad > fecha_maxima:
                                    resultados['errores'].append({
                                        'fila': row_num,
                                        'error': f'Fecha de caducidad muy lejana ({fecha_caducidad.strftime("%d/%m/%Y")}). '
                                                f'Máximo permitido: 8 años desde hoy ({fecha_maxima.strftime("%d/%m/%Y")}). '
                                                f'Verifique que el formato sea correcto (DD/MM/AAAA).'
                                    })
                                    continue
                            
                            # Estado del producto (default: bueno)
                            estado_prod = str(get_val(row, 'estado_producto', 'bueno')).lower().strip()
                            if estado_prod not in ['bueno', 'regular', 'deteriorado']:
                                estado_prod = 'bueno'
                            
                            DetalleDonacion.objects.create(
                                donacion=donacion,
                                producto=None,  # No usa catálogo principal
                                producto_donacion=producto_donacion,  # Usa catálogo independiente
                                numero_lote=str(get_val(row, 'lote', '')).strip()[:50] or None,
                                cantidad=cantidad,
                                cantidad_disponible=0,  # Stock NO disponible hasta procesar
                                fecha_caducidad=fecha_caducidad,
                                estado_producto=estado_prod,
                                notas=str(get_val(row, 'notas', '')).strip() or None
                            )
                            resultados['detalles_creados'] += 1
                            
                    except Exception as e:
                        resultados['errores'].append({
                            'fila': row_num,
                            'error': str(e)
                        })
            
            logger.info(f"Importación de donaciones por {request.user.username}: "
                       f"{resultados['donaciones_creadas']} donaciones, {resultados['detalles_creados']} detalles, "
                       f"filas procesadas: {filas_procesadas}, vacias: {filas_vacias}, ejemplo: {filas_ejemplo}")
            
            status_code = status.HTTP_200_OK if len(resultados['errores']) == 0 else status.HTTP_207_MULTI_STATUS
            
            return Response({
                'mensaje': 'Importación completada',
                'resultados': {
                    'exitosos': resultados['donaciones_creadas'] + resultados['detalles_creados'],
                    'fallidos': len(resultados['errores']),
                    'donaciones_creadas': resultados['donaciones_creadas'],
                    'detalles_creados': resultados['detalles_creados'],
                    'errores': resultados['errores'],
                    'filas_procesadas': filas_procesadas,
                    'filas_vacias': filas_vacias,
                    'filas_ejemplo': filas_ejemplo
                },
                'exitos': resultados['exitos'],
                'errores': resultados['errores']
            }, status=status_code)
            
        except Exception as e:
            logger.error(f"Error importando donaciones: {e}")
            return Response({
                'error': 'Error al procesar archivo',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# =============================================================================
# CATALOGO DE PRODUCTOS DONACIONES - COMPLETAMENTE INDEPENDIENTE
# =============================================================================

class ProductoDonacionViewSet(viewsets.ModelViewSet):
    """
    ViewSet para el catálogo INDEPENDIENTE de productos de donaciones.
    Este catálogo es COMPLETAMENTE SEPARADO del catálogo principal de productos.
    Las donaciones pueden tener productos con claves y nombres diferentes.
    
    Solo ADMIN y FARMACIA pueden crear/editar/eliminar.
    Cualquier usuario autenticado con permiso de donaciones puede ver.
    """
    pagination_class = StandardResultsSetPagination
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['clave', 'nombre', 'descripcion']
    ordering_fields = ['clave', 'nombre', 'created_at']
    ordering = ['nombre']
    
    def get_permissions(self):
        """Permisos según la acción:
        - list, retrieve: IsAuthenticated
        - create, update, destroy: IsFarmaciaRole (admin/farmacia)
        """
        if self.action in ['list', 'retrieve']:
            return [IsAuthenticated()]
        return [IsAuthenticated(), IsFarmaciaRole()]
    
    def get_queryset(self):
        from core.models import ProductoDonacion
        queryset = ProductoDonacion.objects.all()
        
        # Filtrar por activo (por defecto solo activos)
        activo = self.request.query_params.get('activo', 'true')
        if activo.lower() == 'true':
            queryset = queryset.filter(activo=True)
        elif activo.lower() == 'false':
            queryset = queryset.filter(activo=False)
        # Si activo='all', no filtrar
        
        return queryset.order_by('nombre')
    
    def get_serializer_class(self):
        from core.serializers import ProductoDonacionSerializer
        return ProductoDonacionSerializer
    
    def perform_create(self, serializer):
        producto = serializer.save()
        logger.info(f"Producto de donación {producto.clave} creado por {self.request.user.username}")
    
    def perform_update(self, serializer):
        producto = serializer.save()
        logger.info(f"Producto de donación {producto.clave} actualizado por {self.request.user.username}")
    
    def perform_destroy(self, instance):
        clave = instance.clave
        instance.delete()
        logger.info(f"Producto de donación {clave} eliminado por {self.request.user.username}")
    
    @action(detail=False, methods=['get'], url_path='buscar')
    def buscar(self, request):
        """Búsqueda rápida de productos de donación por clave o nombre."""
        from core.models import ProductoDonacion
        from core.serializers import ProductoDonacionSerializer
        
        q = request.query_params.get('q', '').strip()
        if len(q) < 2:
            return Response([])
        
        productos = ProductoDonacion.objects.filter(
            Q(clave__icontains=q) | Q(nombre__icontains=q),
            activo=True
        )[:20]
        
        return Response(ProductoDonacionSerializer(productos, many=True).data)

    @action(detail=False, methods=['get'], url_path='plantilla-excel')
    def plantilla_excel(self, request):
        """Genera plantilla Excel para importar productos de donación."""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Productos Donación'
        
        # Estilos
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='9F2241', end_color='9F2241', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers - campos de productos_donacion
        headers = [
            'clave *',
            'nombre *',
            'descripcion',
            'unidad_medida',
            'presentacion',
            'activo',
            'notas'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        # Fila de ejemplo
        ejemplo = [
            '[EJEMPLO] DON-MED-001',
            'Paracetamol 500mg Donación',
            'Tabletas analgésicas',
            'CAJA',
            'Caja con 20 tabletas',
            'SI',
            'Producto de ejemplo - eliminar esta fila'
        ]
        for col, val in enumerate(ejemplo, 1):
            cell = ws.cell(row=2, column=col, value=val)
            cell.border = thin_border
            cell.font = Font(italic=True, color='888888')
        
        # Segunda fila de ejemplo
        ejemplo2 = [
            '[EJEMPLO] DON-INS-001',
            'Jeringa 5ml Donación',
            'Jeringa desechable',
            'PIEZA',
            'Unidad',
            'SI',
            ''
        ]
        for col, val in enumerate(ejemplo2, 1):
            cell = ws.cell(row=3, column=col, value=val)
            cell.border = thin_border
            cell.font = Font(italic=True, color='888888')
        
        # Hoja de instrucciones
        ws_inst = wb.create_sheet('Instrucciones')
        instrucciones = [
            ['INSTRUCCIONES PARA IMPORTAR PRODUCTOS DE DONACIÓN'],
            [''],
            ['Este catálogo es INDEPENDIENTE del catálogo principal de productos.'],
            ['Puede usar claves y nombres diferentes a los del inventario general.'],
            [''],
            ['CAMPOS:'],
            ['- clave *: Código único del producto (requerido)'],
            ['- nombre *: Nombre del producto (requerido)'],
            ['- descripcion: Descripción detallada (opcional)'],
            ['- unidad_medida: PIEZA, CAJA, FRASCO, etc. (default: PIEZA)'],
            ['- presentacion: Forma de presentación (opcional)'],
            ['- activo: SI o NO (default: SI)'],
            ['- notas: Observaciones adicionales (opcional)'],
            [''],
            ['NOTAS:'],
            ['- Las filas con [EJEMPLO] serán ignoradas'],
            ['- Los campos marcados con * son obligatorios'],
            ['- Las claves duplicadas actualizarán el producto existente'],
        ]
        for row_idx, row in enumerate(instrucciones, 1):
            ws_inst.cell(row=row_idx, column=1, value=row[0] if row else '')
        ws_inst.column_dimensions['A'].width = 60
        
        # Preparar respuesta
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=plantilla_productos_donacion.xlsx'
        wb.save(response)
        
        return response

    @action(detail=False, methods=['get'], url_path='exportar-excel')
    def exportar_excel(self, request):
        """
        Exporta todos los productos de donación a Excel CON información de inventario.
        Incluye stock disponible sumando cantidad_disponible de detalle_donaciones.
        """
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse
        from django.db.models import Sum, Count
        from core.models import ProductoDonacion, DetalleDonacion
        
        # Obtener productos con su stock calculado desde detalle_donaciones
        productos = ProductoDonacion.objects.all().order_by('clave')
        
        # Calcular stock por producto_donacion_id
        stock_por_producto = {}
        detalles_agrupados = DetalleDonacion.objects.filter(
            producto_donacion__isnull=False
        ).values('producto_donacion_id').annotate(
            stock_total=Sum('cantidad_disponible'),
            num_lotes=Count('id')
        )
        for item in detalles_agrupados:
            stock_por_producto[item['producto_donacion_id']] = {
                'stock': item['stock_total'] or 0,
                'lotes': item['num_lotes'] or 0
            }
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Productos Donación'
        
        # Estilos
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='9F2241', end_color='9F2241', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        stock_fill_ok = PatternFill(start_color='D5F5E3', end_color='D5F5E3', fill_type='solid')
        stock_fill_zero = PatternFill(start_color='FADBD8', end_color='FADBD8', fill_type='solid')
        
        # Headers CON inventario
        headers = ['clave', 'nombre', 'descripcion', 'unidad_medida', 'presentacion', 'activo', 'stock_disponible', 'num_lotes', 'notas']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
            # Ajustar anchos
            if header in ['nombre', 'descripcion', 'notas']:
                ws.column_dimensions[get_column_letter(col)].width = 30
            elif header in ['stock_disponible', 'num_lotes']:
                ws.column_dimensions[get_column_letter(col)].width = 15
            else:
                ws.column_dimensions[get_column_letter(col)].width = 18
        
        for row_idx, prod in enumerate(productos, 2):
            info_stock = stock_por_producto.get(prod.id, {'stock': 0, 'lotes': 0})
            stock = info_stock['stock']
            lotes = info_stock['lotes']
            
            ws.cell(row=row_idx, column=1, value=prod.clave).border = thin_border
            ws.cell(row=row_idx, column=2, value=prod.nombre).border = thin_border
            ws.cell(row=row_idx, column=3, value=prod.descripcion or '').border = thin_border
            ws.cell(row=row_idx, column=4, value=prod.unidad_medida or 'PIEZA').border = thin_border
            ws.cell(row=row_idx, column=5, value=prod.presentacion or '').border = thin_border
            ws.cell(row=row_idx, column=6, value='SI' if prod.activo else 'NO').border = thin_border
            
            # Stock con color según disponibilidad
            stock_cell = ws.cell(row=row_idx, column=7, value=stock)
            stock_cell.border = thin_border
            stock_cell.alignment = Alignment(horizontal='center')
            stock_cell.fill = stock_fill_ok if stock > 0 else stock_fill_zero
            
            lotes_cell = ws.cell(row=row_idx, column=8, value=lotes)
            lotes_cell.border = thin_border
            lotes_cell.alignment = Alignment(horizontal='center')
            
            ws.cell(row=row_idx, column=9, value=prod.notas or '').border = thin_border
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=catalogo_donaciones_con_inventario_{timezone.now().strftime("%Y%m%d")}.xlsx'
        wb.save(response)
        
        return response

    @action(detail=False, methods=['post'], url_path='importar-excel')
    def importar_excel(self, request):
        """Importa productos de donación desde Excel."""
        import openpyxl
        from core.models import ProductoDonacion
        
        archivo = request.FILES.get('archivo')
        if not archivo:
            return Response({'error': 'No se proporcionó archivo'}, status=400)
        
        try:
            wb = openpyxl.load_workbook(archivo)
            ws = wb.active
        except Exception as e:
            return Response({'error': f'Error al leer archivo: {str(e)}'}, status=400)
        
        # Buscar fila de headers
        headers = []
        header_row = 1
        for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
            if row and any('clave' in str(cell).lower() for cell in row if cell):
                headers = [str(cell).lower().replace(' *', '').replace('*', '').strip() if cell else '' for cell in row]
                break
            header_row += 1
        
        if not headers or 'clave' not in headers:
            return Response({'error': 'No se encontró encabezado con columna "clave"'}, status=400)
        
        col_map = {h: idx for idx, h in enumerate(headers) if h}
        
        creados = 0
        actualizados = 0
        errores = []
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1):
            if not row or not any(row):
                continue
            
            # Ignorar filas de ejemplo
            if any('[EJEMPLO]' in str(cell).upper() for cell in row if cell):
                continue
            
            try:
                clave = str(row[col_map.get('clave', 0)] or '').strip()
                if not clave:
                    continue
                
                nombre = str(row[col_map.get('nombre', 1)] or '').strip()
                if not nombre:
                    errores.append(f'Fila {row_idx}: nombre es requerido')
                    continue
                
                # ISS-FIX: Truncar campos largos para evitar errores de BD
                descripcion = str(row[col_map.get('descripcion', 2)] or '').strip() if col_map.get('descripcion') is not None else ''
                unidad = str(row[col_map.get('unidad_medida', 3)] or 'PIEZA').strip().upper()[:50] if col_map.get('unidad_medida') is not None else 'PIEZA'
                presentacion = str(row[col_map.get('presentacion', 4)] or '').strip()[:255] if col_map.get('presentacion') is not None else ''
                
                # Advertir si se truncó
                presentacion_original = str(row[col_map.get('presentacion', 4)] or '').strip() if col_map.get('presentacion') is not None else ''
                if len(presentacion_original) > 255:
                    errores.append(f'Fila {row_idx}: presentación truncada (excedía 255 caracteres)')
                
                activo_val = str(row[col_map.get('activo', 5)] or 'SI').strip().upper() if col_map.get('activo') is not None else 'SI'
                activo = activo_val in ['SI', 'S', 'YES', 'Y', '1', 'TRUE', 'ACTIVO']
                
                notas = str(row[col_map.get('notas', 6)] or '').strip() if col_map.get('notas') is not None else ''
                
                # Buscar o crear
                producto, created = ProductoDonacion.objects.update_or_create(
                    clave=clave,
                    defaults={
                        'nombre': nombre,
                        'descripcion': descripcion or None,
                        'unidad_medida': unidad,
                        'presentacion': presentacion or None,
                        'activo': activo,
                        'notas': notas or None
                    }
                )
                
                if created:
                    creados += 1
                else:
                    actualizados += 1
                    
            except Exception as e:
                errores.append(f'Fila {row_idx}: {str(e)}')
        
        return Response({
            'success': True,
            'creados': creados,
            'actualizados': actualizados,
            'total': creados + actualizados,
            'errores': errores[:20] if errores else []
        })


class DetalleDonacionViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestionar detalles de donaciones.
    Usa el catálogo independiente de ProductoDonacion.
    Solo ADMIN y FARMACIA pueden modificar.
    
    ISS-FIX SEGURIDAD:
    - Lectura requiere perm_donaciones (no solo autenticación)
    - Scoping por centro: usuarios de centro solo ven detalles de donaciones a SU centro
    """
    pagination_class = StandardResultsSetPagination
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = [
        'producto_donacion__nombre', 'producto_donacion__clave',  # Nuevo catálogo
        'producto__nombre', 'producto__clave',  # Legacy
        'numero_lote', 'donacion__numero'
    ]
    ordering_fields = ['created_at', 'fecha_caducidad', 'cantidad_disponible']
    ordering = ['-created_at']
    
    def get_permissions(self):
        """
        ISS-FIX: Permisos según la acción:
        - list, retrieve, exportaciones: HasDonacionesPermission (verifica perm_donaciones)
        - create, update, destroy: IsFarmaciaRole
        """
        from core.permissions import HasDonacionesPermission, HasDonacionesWritePermission
        
        if self.action in ['list', 'retrieve', 'exportar_excel', 'exportar_pdf']:
            # ISS-FIX: Requiere perm_donaciones, no solo autenticación
            return [IsAuthenticated(), HasDonacionesPermission()]
        return [IsAuthenticated(), HasDonacionesWritePermission()]
    
    def get_queryset(self):
        from core.models import DetalleDonacion
        from core.permissions import _has_role
        from datetime import date, timedelta
        
        queryset = DetalleDonacion.objects.select_related(
            'donacion', 'donacion__centro_destino', 'producto', 'producto_donacion'
        )
        
        # ISS-FIX SEGURIDAD: Scoping por centro
        # Admin/farmacia/vista ven todo; usuarios de centro solo de su centro
        user = self.request.user
        if not _has_role(user, ['admin', 'farmacia', 'vista']):
            # Usuario de centro: solo detalles de donaciones a su centro
            user_centro = getattr(user, 'centro', None)
            if user_centro:
                queryset = queryset.filter(donacion__centro_destino=user_centro)
                logger.debug(
                    f"ISS-FIX: Usuario {user.username} de centro {user_centro.nombre} "
                    f"accediendo a detalles de donaciones filtrados"
                )
            else:
                logger.warning(
                    f"ISS-FIX: Usuario {user.username} sin centro intentó acceder a detalles donaciones"
                )
                return DetalleDonacion.objects.none()
        
        # Filtrar por donacion si se especifica
        donacion_id = self.request.query_params.get('donacion')
        if donacion_id:
            queryset = queryset.filter(donacion_id=donacion_id)
        else:
            # Si no se especifica donación, solo mostrar detalles de donaciones PROCESADAS
            # El inventario solo debe mostrar productos con stock activado
            queryset = queryset.filter(donacion__estado='procesada')
        
        # Filtrar solo con stock disponible
        solo_disponible = self.request.query_params.get('disponible')
        if solo_disponible == 'true':
            queryset = queryset.filter(cantidad_disponible__gt=0)
        elif solo_disponible == 'agotado':
            queryset = queryset.filter(cantidad_disponible=0)
        
        # Filtrar por estado de producto
        estado = self.request.query_params.get('estado_producto')
        if estado:
            queryset = queryset.filter(estado_producto=estado)
        
        # Filtrar por caducidad
        caducidad = self.request.query_params.get('caducidad')
        if caducidad:
            hoy = date.today()
            if caducidad == 'vencido':
                queryset = queryset.filter(fecha_caducidad__lt=hoy)
            elif caducidad == 'critico':  # <= 30 días
                queryset = queryset.filter(
                    fecha_caducidad__gte=hoy,
                    fecha_caducidad__lte=hoy + timedelta(days=30)
                )
            elif caducidad == 'proximo':  # 31-90 días
                queryset = queryset.filter(
                    fecha_caducidad__gt=hoy + timedelta(days=30),
                    fecha_caducidad__lte=hoy + timedelta(days=90)
                )
            elif caducidad == 'normal':  # > 90 días
                queryset = queryset.filter(fecha_caducidad__gt=hoy + timedelta(days=90))
        
        # Búsqueda por texto
        search = self.request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                Q(producto_donacion__nombre__icontains=search) |
                Q(producto_donacion__clave__icontains=search) |
                Q(producto__nombre__icontains=search) |
                Q(producto__clave__icontains=search) |
                Q(numero_lote__icontains=search) |
                Q(donacion__numero__icontains=search)
            )
        
        return queryset.order_by('-created_at')
    
    def get_serializer_class(self):
        from core.serializers import DetalleDonacionSerializer
        return DetalleDonacionSerializer
    
    @action(detail=False, methods=['get'], url_path='exportar-excel')
    def exportar_excel(self, request):
        """
        Exporta el inventario de donaciones a Excel con formato de trazabilidad.
        Respeta todos los filtros aplicados (search, disponible, estado_producto, caducidad).
        
        Formato estilo licitación/trazabilidad con columnas:
        - No., Clave, Descripción, Unidad, Presentación
        - Lote, Caducidad, Estado, Cantidad Recibida, Cantidad Disponible
        - Donación, Fecha Donación, Donante
        """
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse
        from django.utils import timezone
        
        try:
            # Obtener datos filtrados
            inventario = self.get_queryset()
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Inventario Donaciones'
            
            # ===== ENCABEZADO INSTITUCIONAL =====
            ws.merge_cells('A1:M1')
            titulo = ws['A1']
            titulo.value = 'SISTEMA DE FARMACIA PENITENCIARIA'
            titulo.font = Font(bold=True, size=14, color='632842')
            titulo.alignment = Alignment(horizontal='center', vertical='center')
            
            ws.merge_cells('A2:M2')
            subtitulo = ws['A2']
            subtitulo.value = 'INVENTARIO DE DONACIONES - FORMATO DE TRAZABILIDAD'
            subtitulo.font = Font(bold=True, size=12, color='9F2241')
            subtitulo.alignment = Alignment(horizontal='center', vertical='center')
            
            ws.merge_cells('A3:M3')
            fecha_gen = ws['A3']
            fecha_gen.value = f'Generado el: {timezone.now().strftime("%d/%m/%Y %H:%M:%S")}'
            fecha_gen.font = Font(italic=True, size=10, color='666666')
            fecha_gen.alignment = Alignment(horizontal='center')
            
            # Mostrar filtros activos
            filtros_activos = []
            if request.query_params.get('search'):
                filtros_activos.append(f"Búsqueda: '{request.query_params.get('search')}'")
            if request.query_params.get('disponible'):
                disp = request.query_params.get('disponible')
                filtros_activos.append(f"Disponibilidad: {disp}")
            if request.query_params.get('estado_producto'):
                filtros_activos.append(f"Estado: {request.query_params.get('estado_producto')}")
            if request.query_params.get('caducidad'):
                filtros_activos.append(f"Caducidad: {request.query_params.get('caducidad')}")
            
            if filtros_activos:
                ws.merge_cells('A4:M4')
                filtros_cell = ws['A4']
                filtros_cell.value = f'Filtros: {" | ".join(filtros_activos)}'
                filtros_cell.font = Font(italic=True, size=9, color='666666')
                filtros_cell.alignment = Alignment(horizontal='center')
                header_row = 6
            else:
                header_row = 5
            
            ws.append([])  # Fila vacía
            
            # ===== ESTILOS =====
            header_fill = PatternFill(start_color='9F2241', end_color='9F2241', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=10)
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Colores para semáforo de caducidad
            fill_vencido = PatternFill(start_color='FADBD8', end_color='FADBD8', fill_type='solid')
            fill_critico = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
            fill_proximo = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
            fill_ok = PatternFill(start_color='D5F5E3', end_color='D5F5E3', fill_type='solid')
            fill_agotado = PatternFill(start_color='F5B7B1', end_color='F5B7B1', fill_type='solid')
            
            # ===== ENCABEZADOS DE TABLA =====
            headers = [
                'No.', 'Clave', 'Descripción del Producto', 'Unidad', 'Presentación',
                'Número de Lote', 'Fecha Caducidad', 'Estado', 'Cant. Recibida', 
                'Cant. Disponible', 'No. Donación', 'Fecha Donación', 'Donante'
            ]
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=col_num, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Anchos de columna
            col_widths = {
                'A': 6,   # No.
                'B': 12,  # Clave
                'C': 40,  # Descripción
                'D': 10,  # Unidad
                'E': 15,  # Presentación
                'F': 15,  # Lote
                'G': 14,  # Caducidad
                'H': 12,  # Estado
                'I': 12,  # Recibida
                'J': 12,  # Disponible
                'K': 15,  # No. Donación
                'L': 14,  # Fecha Donación
                'M': 25,  # Donante
            }
            for col, width in col_widths.items():
                ws.column_dimensions[col].width = width
            
            # ===== DATOS =====
            from datetime import date, timedelta
            hoy = date.today()
            
            for idx, item in enumerate(inventario, 1):
                row_num = header_row + idx
                
                # Obtener datos del producto (priorizar producto_donacion)
                if item.producto_donacion:
                    clave = item.producto_donacion.clave
                    nombre = item.producto_donacion.nombre
                    unidad = item.producto_donacion.unidad_medida or 'PIEZA'
                    presentacion = item.producto_donacion.presentacion or ''
                elif item.producto:
                    clave = item.producto.clave
                    nombre = item.producto.nombre
                    unidad = item.producto.unidad_medida or 'PIEZA'
                    presentacion = item.producto.presentacion or ''
                else:
                    clave = '-'
                    nombre = 'Producto no especificado'
                    unidad = '-'
                    presentacion = '-'
                
                # Fecha caducidad y semáforo
                fecha_cad = item.fecha_caducidad
                fecha_cad_str = fecha_cad.strftime('%d/%m/%Y') if fecha_cad else 'N/A'
                
                # Determinar semáforo
                if fecha_cad:
                    dias_restantes = (fecha_cad - hoy).days
                    if dias_restantes < 0:
                        semaforo_fill = fill_vencido
                    elif dias_restantes <= 30:
                        semaforo_fill = fill_critico
                    elif dias_restantes <= 90:
                        semaforo_fill = fill_proximo
                    else:
                        semaforo_fill = fill_ok
                else:
                    semaforo_fill = None
                
                # Estado del producto
                estado = item.estado_producto or 'bueno'
                
                # Donación info
                don_numero = item.donacion.numero if item.donacion else '-'
                don_fecha = item.donacion.fecha_donacion.strftime('%d/%m/%Y') if item.donacion and item.donacion.fecha_donacion else '-'
                don_donante = item.donacion.donante_nombre if item.donacion else '-'
                
                # Escribir fila
                row_data = [
                    idx,
                    clave,
                    nombre,
                    unidad,
                    presentacion,
                    item.numero_lote or '-',
                    fecha_cad_str,
                    estado.capitalize(),
                    item.cantidad,
                    item.cantidad_disponible,
                    don_numero,
                    don_fecha,
                    don_donante
                ]
                
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical='center')
                    
                    # Aplicar color a columna de caducidad
                    if col_num == 7 and semaforo_fill:  # Columna Caducidad
                        cell.fill = semaforo_fill
                    
                    # Aplicar color a cantidad disponible si es 0
                    if col_num == 10 and item.cantidad_disponible == 0:
                        cell.fill = fill_agotado
                        cell.font = Font(bold=True, color='922B21')
            
            # ===== TOTALES =====
            total_row = header_row + len(inventario) + 2
            ws.cell(row=total_row, column=8, value='TOTALES:').font = Font(bold=True)
            ws.cell(row=total_row, column=9, value=sum(i.cantidad for i in inventario)).font = Font(bold=True)
            ws.cell(row=total_row, column=10, value=sum(i.cantidad_disponible for i in inventario)).font = Font(bold=True)
            
            # ===== LEYENDA =====
            leyenda_row = total_row + 2
            ws.cell(row=leyenda_row, column=1, value='Leyenda de colores (Caducidad):').font = Font(bold=True, size=9)
            
            leyenda_items = [
                ('Verde', 'Vigente (+90 días)', fill_ok),
                ('Amarillo', 'Próximo (31-90 días)', fill_proximo),
                ('Naranja', 'Crítico (≤30 días)', fill_critico),
                ('Rojo', 'Vencido', fill_vencido),
            ]
            
            for i, (color, desc, fill) in enumerate(leyenda_items):
                cell = ws.cell(row=leyenda_row + 1 + i, column=1, value=f'  {color}: {desc}')
                cell.font = Font(size=8)
            
            # ===== RESPUESTA =====
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = f'inventario_donaciones_trazabilidad_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            wb.save(response)
            return response
            
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error exportando inventario donaciones Excel: {e}")
            return Response(
                {'error': f'Error al exportar: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'], url_path='exportar-pdf')
    def exportar_pdf(self, request):
        """
        Exporta el inventario de donaciones a PDF con formato profesional.
        Respeta todos los filtros aplicados.
        """
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch, cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        from django.http import HttpResponse
        from django.utils import timezone
        from io import BytesIO
        from datetime import date, timedelta
        
        try:
            # Obtener datos filtrados
            inventario = list(self.get_queryset())
            hoy = date.today()
            
            # Crear buffer para PDF
            buffer = BytesIO()
            
            # Configurar documento en landscape para más columnas
            doc = SimpleDocTemplate(
                buffer,
                pagesize=landscape(letter),
                rightMargin=0.5*inch,
                leftMargin=0.5*inch,
                topMargin=0.5*inch,
                bottomMargin=0.5*inch
            )
            
            elements = []
            styles = getSampleStyleSheet()
            
            # Estilos personalizados
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                textColor=colors.HexColor('#632842'),
                alignment=TA_CENTER,
                spaceAfter=6
            )
            
            subtitle_style = ParagraphStyle(
                'CustomSubtitle',
                parent=styles['Heading2'],
                fontSize=12,
                textColor=colors.HexColor('#9F2241'),
                alignment=TA_CENTER,
                spaceAfter=12
            )
            
            info_style = ParagraphStyle(
                'InfoStyle',
                parent=styles['Normal'],
                fontSize=9,
                textColor=colors.HexColor('#666666'),
                alignment=TA_CENTER,
                spaceAfter=6
            )
            
            # ===== ENCABEZADO =====
            elements.append(Paragraph('SISTEMA DE FARMACIA PENITENCIARIA', title_style))
            elements.append(Paragraph('INVENTARIO DE DONACIONES', subtitle_style))
            elements.append(Paragraph(f'Generado: {timezone.now().strftime("%d/%m/%Y %H:%M")}', info_style))
            
            # Mostrar filtros activos
            filtros_activos = []
            if request.query_params.get('search'):
                filtros_activos.append(f"Búsqueda: '{request.query_params.get('search')}'")
            if request.query_params.get('disponible'):
                filtros_activos.append(f"Disponibilidad: {request.query_params.get('disponible')}")
            if request.query_params.get('estado_producto'):
                filtros_activos.append(f"Estado: {request.query_params.get('estado_producto')}")
            if request.query_params.get('caducidad'):
                filtros_activos.append(f"Caducidad: {request.query_params.get('caducidad')}")
            
            if filtros_activos:
                elements.append(Paragraph(f'Filtros: {" | ".join(filtros_activos)}', info_style))
            
            elements.append(Spacer(1, 0.2*inch))
            
            # ===== TABLA DE DATOS =====
            # ISS-FIX: Usar Paragraph para ajustar texto en celdas y evitar desbordamiento
            from reportlab.platypus import Paragraph as TableParagraph
            from reportlab.lib.styles import ParagraphStyle
            
            # Estilo para celdas con texto que puede desbordar
            cell_style = ParagraphStyle(
                'CellStyle',
                fontName='Helvetica',
                fontSize=7,
                leading=8,
                wordWrap='CJK',  # Permite wrap agresivo
            )
            cell_style_header = ParagraphStyle(
                'CellStyleHeader',
                fontName='Helvetica-Bold',
                fontSize=8,
                leading=9,
                textColor=colors.whitesmoke,
                alignment=TA_CENTER,
            )
            
            # Encabezados más cortos para PDF
            headers = [
                TableParagraph('#', cell_style_header),
                TableParagraph('Clave', cell_style_header),
                TableParagraph('Producto', cell_style_header),
                TableParagraph('Lote', cell_style_header),
                TableParagraph('Caducidad', cell_style_header),
                TableParagraph('Estado', cell_style_header),
                TableParagraph('Recib.', cell_style_header),
                TableParagraph('Disp.', cell_style_header),
                TableParagraph('Donación', cell_style_header),
                TableParagraph('Donante', cell_style_header),
            ]
            
            data = [headers]
            
            for idx, item in enumerate(inventario, 1):
                # Obtener datos del producto - SIN TRUNCAR, usar Paragraph para wrap
                if item.producto_donacion:
                    clave = item.producto_donacion.clave if item.producto_donacion.clave else '-'
                    nombre = item.producto_donacion.nombre if item.producto_donacion.nombre else '-'
                elif item.producto:
                    clave = item.producto.clave if item.producto.clave else '-'
                    nombre = item.producto.nombre if item.producto.nombre else '-'
                else:
                    clave = '-'
                    nombre = '-'
                
                # Fecha caducidad
                fecha_cad = item.fecha_caducidad
                fecha_cad_str = fecha_cad.strftime('%d/%m/%y') if fecha_cad else 'N/A'
                
                # Estado
                estado = (item.estado_producto or 'bueno').capitalize()
                
                # Donación - SIN TRUNCAR
                don_numero = item.donacion.numero if item.donacion and item.donacion.numero else '-'
                don_donante = item.donacion.donante_nombre if item.donacion and item.donacion.donante_nombre else '-'
                
                # ISS-FIX: Usar Paragraph para campos que pueden desbordar
                data.append([
                    str(idx),
                    TableParagraph(clave, cell_style),
                    TableParagraph(nombre, cell_style),
                    TableParagraph((item.numero_lote or '-'), cell_style),
                    fecha_cad_str,
                    estado,
                    str(item.cantidad),
                    str(item.cantidad_disponible),
                    TableParagraph(don_numero, cell_style),
                    TableParagraph(don_donante, cell_style),
                ])
            
            # Crear tabla
            col_widths = [0.4*inch, 0.9*inch, 2.2*inch, 0.9*inch, 0.8*inch, 0.6*inch, 0.5*inch, 0.5*inch, 0.9*inch, 1.2*inch]
            
            table = Table(data, colWidths=col_widths, repeatRows=1)
            
            # Estilo de tabla
            table_style = TableStyle([
                # Encabezados
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#9F2241')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                
                # Datos
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 7),
                ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # #
                ('ALIGN', (6, 1), (7, -1), 'CENTER'),  # Cantidades
                
                # Bordes
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('LINEBELOW', (0, 0), (-1, 0), 1.5, colors.HexColor('#632842')),
                
                # Alternar colores de fila
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F8F8')]),
            ])
            
            # Colorear filas según caducidad y disponibilidad
            for row_idx, item in enumerate(inventario, 1):
                fecha_cad = item.fecha_caducidad
                
                # Color por caducidad
                if fecha_cad:
                    dias_restantes = (fecha_cad - hoy).days
                    if dias_restantes < 0:
                        table_style.add('BACKGROUND', (4, row_idx), (4, row_idx), colors.HexColor('#FADBD8'))
                    elif dias_restantes <= 30:
                        table_style.add('BACKGROUND', (4, row_idx), (4, row_idx), colors.HexColor('#FCE4D6'))
                    elif dias_restantes <= 90:
                        table_style.add('BACKGROUND', (4, row_idx), (4, row_idx), colors.HexColor('#FFF2CC'))
                    else:
                        table_style.add('BACKGROUND', (4, row_idx), (4, row_idx), colors.HexColor('#D5F5E3'))
                
                # Color por disponibilidad
                if item.cantidad_disponible == 0:
                    table_style.add('BACKGROUND', (7, row_idx), (7, row_idx), colors.HexColor('#F5B7B1'))
                    table_style.add('TEXTCOLOR', (7, row_idx), (7, row_idx), colors.HexColor('#922B21'))
                    table_style.add('FONTNAME', (7, row_idx), (7, row_idx), 'Helvetica-Bold')
            
            table.setStyle(table_style)
            elements.append(table)
            
            # ===== TOTALES =====
            elements.append(Spacer(1, 0.2*inch))
            
            total_recibido = sum(i.cantidad for i in inventario)
            total_disponible = sum(i.cantidad_disponible for i in inventario)
            
            totales_data = [
                ['', '', '', '', '', '', 'TOTALES:', str(total_recibido), str(total_disponible), '']
            ]
            totales_table = Table(totales_data, colWidths=col_widths)
            totales_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('ALIGN', (6, 0), (8, 0), 'CENTER'),
                ('TEXTCOLOR', (6, 0), (8, 0), colors.HexColor('#632842')),
            ]))
            elements.append(totales_table)
            
            # ===== LEYENDA =====
            elements.append(Spacer(1, 0.3*inch))
            
            leyenda_style = ParagraphStyle(
                'Leyenda',
                parent=styles['Normal'],
                fontSize=7,
                textColor=colors.HexColor('#666666')
            )
            
            leyenda_text = '''
            <b>Leyenda de colores (Caducidad):</b> 
            <font color="#27AE60">■</font> Vigente (+90 días) | 
            <font color="#F39C12">■</font> Próximo (31-90 días) | 
            <font color="#E67E22">■</font> Crítico (≤30 días) | 
            <font color="#E74C3C">■</font> Vencido |
            <font color="#C0392B">■</font> Agotado (Disp.=0)
            '''
            elements.append(Paragraph(leyenda_text, leyenda_style))
            
            # Construir PDF
            doc.build(elements)
            
            # Respuesta
            buffer.seek(0)
            response = HttpResponse(buffer, content_type='application/pdf')
            filename = f'inventario_donaciones_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf'
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            return response
            
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error exportando inventario donaciones PDF: {e}")
            return Response(
                {'error': f'Error al exportar PDF: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


# =============================================================================
# SALIDA DONACION VIEWSET (Control de entregas del almacen donaciones)
# =============================================================================

class SalidaDonacionViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestionar salidas/entregas del almacen de donaciones.
    Control interno sin afectar movimientos principales.
    Solo ADMIN y FARMACIA pueden registrar entregas.
    
    ISS-FIX SEGURIDAD:
    - Lectura requiere perm_donaciones (no solo autenticación)
    - Scoping por centro: usuarios de centro solo ven entregas a SU centro
    
    Endpoints adicionales:
    - GET /salidas-donaciones/exportar-excel/ - Exportar entregas a Excel
    - POST /salidas-donaciones/importar-excel/ - Importar entregas desde Excel
    - GET /salidas-donaciones/plantilla-excel/ - Descargar plantilla de importación
    - DELETE /salidas-donaciones/{id}/ - Eliminar entrega NO finalizada (devuelve stock)
    """
    pagination_class = StandardResultsSetPagination
    http_method_names = ['get', 'post', 'delete', 'head', 'options']  # Permite eliminar entregas pendientes
    
    def get_permissions(self):
        """
        ISS-FIX: Permisos según la acción:
        - list, retrieve, exportaciones: HasDonacionesPermission (verifica perm_donaciones)
        - create, delete, importar_excel: IsFarmaciaRole
        """
        from core.permissions import HasDonacionesPermission, HasDonacionesWritePermission
        
        if self.action in ['list', 'retrieve', 'exportar_excel', 'plantilla_excel']:
            # ISS-FIX: Requiere perm_donaciones, no solo autenticación
            return [IsAuthenticated(), HasDonacionesPermission()]
        return [IsAuthenticated(), HasDonacionesWritePermission()]
    
    def destroy(self, request, *args, **kwargs):
        """
        Eliminar una salida de donación NO finalizada.
        Al eliminar, se devuelve el stock al detalle de donación si corresponde.
        Solo se pueden eliminar entregas que NO han sido confirmadas.
        """
        from django.db import transaction
        
        instance = self.get_object()
        
        # Verificar que no esté finalizada
        if instance.finalizado:
            return Response(
                {'error': 'No se puede eliminar una entrega que ya fue confirmada/finalizada'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            with transaction.atomic():
                # Guardar datos antes de eliminar
                cantidad_salida = instance.cantidad
                detalle = instance.detalle_donacion
                
                # Eliminar la salida primero
                instance.delete()
                
                # Devolver el stock al detalle de donación SOLO si no excede el máximo
                # El constraint chk_cantidad_disponible impide que cantidad_disponible > cantidad
                if detalle:
                    cantidad_maxima = detalle.cantidad
                    nuevo_disponible = detalle.cantidad_disponible + cantidad_salida
                    
                    # Solo actualizar si no viola el constraint
                    if nuevo_disponible <= cantidad_maxima:
                        detalle.cantidad_disponible = nuevo_disponible
                        detalle.save(update_fields=['cantidad_disponible'])
                    # Si ya está al máximo, no hacemos nada (datos legacy inconsistentes)
            
            return Response(status=status.HTTP_204_NO_CONTENT)
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error eliminando salida donación: {e}")
            return Response(
                {'error': f'Error al eliminar: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    def get_queryset(self):
        from core.models import SalidaDonacion
        from core.permissions import _has_role
        
        # ISS-DB-ALIGN: Incluir centro_destino, finalizado_por y ambos productos en select_related
        queryset = SalidaDonacion.objects.select_related(
            'detalle_donacion', 
            'detalle_donacion__producto',  # FK legacy
            'detalle_donacion__producto_donacion',  # FK nuevo catálogo donaciones
            'detalle_donacion__donacion', 
            'detalle_donacion__donacion__centro_destino',  # ISS-FIX: Para scoping
            'entregado_por',
            'centro_destino', 
            'finalizado_por'
        )
        
        # ISS-FIX SEGURIDAD: Scoping por centro
        # Admin/farmacia/vista ven todo; usuarios de centro solo de su centro
        user = self.request.user
        if not _has_role(user, ['admin', 'farmacia', 'vista']):
            # Usuario de centro: solo salidas a su centro
            user_centro = getattr(user, 'centro', None)
            if user_centro:
                # Filtrar por centro destino de la salida O centro destino de la donación
                queryset = queryset.filter(
                    Q(centro_destino=user_centro) | 
                    Q(detalle_donacion__donacion__centro_destino=user_centro)
                )
                logger.debug(
                    f"ISS-FIX: Usuario {user.username} de centro {user_centro.nombre} "
                    f"accediendo a salidas de donaciones filtradas"
                )
            else:
                logger.warning(
                    f"ISS-FIX: Usuario {user.username} sin centro intentó acceder a salidas donaciones"
                )
                return SalidaDonacion.objects.none()
        
        # Filtrar por detalle de donacion
        detalle_id = self.request.query_params.get('detalle_donacion')
        if detalle_id:
            queryset = queryset.filter(detalle_donacion_id=detalle_id)
        
        # Filtrar por donacion
        donacion_id = self.request.query_params.get('donacion')
        if donacion_id:
            queryset = queryset.filter(detalle_donacion__donacion_id=donacion_id)
        
        # Filtrar por destinatario
        destinatario = self.request.query_params.get('destinatario')
        if destinatario:
            queryset = queryset.filter(destinatario__icontains=destinatario)
        
        # ISS-DB-ALIGN: Filtrar por centro destino
        centro_destino = self.request.query_params.get('centro_destino')
        if centro_destino:
            queryset = queryset.filter(centro_destino_id=centro_destino)
        
        # ISS-DB-ALIGN: Filtrar por estado de finalización
        finalizado = self.request.query_params.get('finalizado')
        if finalizado is not None:
            is_finalizado = finalizado.lower() in ('true', '1', 'si', 'yes')
            queryset = queryset.filter(finalizado=is_finalizado)
        
        # Filtrar por fecha
        fecha_desde = self.request.query_params.get('fecha_desde')
        if fecha_desde:
            queryset = queryset.filter(fecha_entrega__date__gte=fecha_desde)
        
        fecha_hasta = self.request.query_params.get('fecha_hasta')
        if fecha_hasta:
            queryset = queryset.filter(fecha_entrega__date__lte=fecha_hasta)
        
        # Búsqueda general - ISS-FIX: Incluir producto_donacion (nuevo catálogo)
        search = self.request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                Q(destinatario__icontains=search) |
                Q(motivo__icontains=search) |
                # Nuevo catálogo de donaciones
                Q(detalle_donacion__producto_donacion__nombre__icontains=search) |
                Q(detalle_donacion__producto_donacion__clave__icontains=search) |
                # Legacy (compatibilidad)
                Q(detalle_donacion__producto__nombre__icontains=search) |
                Q(detalle_donacion__producto__clave__icontains=search) |
                # Donación relacionada
                Q(detalle_donacion__donacion__numero__icontains=search)
            )
        
        return queryset.order_by('-fecha_entrega')
    
    def get_serializer_class(self):
        from core.serializers import SalidaDonacionSerializer
        return SalidaDonacionSerializer
    
    @action(detail=False, methods=['get'], url_path='exportar-excel')
    def exportar_excel(self, request):
        """
        Exporta las entregas de donaciones a Excel con formato profesional.
        Respeta los filtros aplicados en la consulta (centro_destino, destinatario, fechas).
        Incluye resumen y totales por centro.
        """
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse
        from django.utils import timezone
        from collections import defaultdict
        
        try:
            entregas = list(self.get_queryset())
            
            # Obtener el centro filtrado para el título (si aplica)
            centro_filtro = self.request.query_params.get('centro_destino')
            fecha_desde = self.request.query_params.get('fecha_desde', '')
            fecha_hasta = self.request.query_params.get('fecha_hasta', '')
            
            centro_nombre = "Todos los Centros"
            if centro_filtro:
                try:
                    centro = Centro.objects.get(pk=centro_filtro)
                    centro_nombre = centro.nombre
                except Centro.DoesNotExist:
                    pass
            
            # Crear libro de Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Entregas Donaciones'
            
            # Estilos
            header_fill = PatternFill(start_color='722F37', end_color='722F37', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            title_font = Font(bold=True, size=14, color='722F37')
            subtitle_font = Font(size=10, italic=True, color='666666')
            total_fill = PatternFill(start_color='E8E8E8', end_color='E8E8E8', fill_type='solid')
            total_font = Font(bold=True, size=11)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Título del reporte
            ws.merge_cells('A1:H1')
            titulo_cell = ws['A1']
            titulo_cell.value = 'REPORTE DE ENTREGAS DE DONACIONES A CENTROS'
            titulo_cell.font = title_font
            titulo_cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 25
            
            # Subtítulo con filtros
            periodo = ""
            if fecha_desde and fecha_hasta:
                periodo = f" | Periodo: {fecha_desde} a {fecha_hasta}"
            elif fecha_desde:
                periodo = f" | Desde: {fecha_desde}"
            elif fecha_hasta:
                periodo = f" | Hasta: {fecha_hasta}"
            
            ws.merge_cells('A2:H2')
            subtitulo_cell = ws['A2']
            subtitulo_cell.value = f'Centro: {centro_nombre}{periodo} | Generado: {timezone.now().strftime("%d/%m/%Y %H:%M")}'
            subtitulo_cell.font = subtitle_font
            subtitulo_cell.alignment = Alignment(horizontal='center')
            
            # Resumen rápido
            ws.merge_cells('A3:H3')
            total_entregas = len(entregas)
            total_cantidad = sum(e.cantidad or 0 for e in entregas)
            resumen_cell = ws['A3']
            resumen_cell.value = f'Total Entregas: {total_entregas} | Total Unidades: {total_cantidad}'
            resumen_cell.font = Font(bold=True, size=11, color='722F37')
            resumen_cell.alignment = Alignment(horizontal='center')
            
            # Espacio
            ws.append([])
            
            # Encabezados
            headers = [
                '#', 'Fecha', 'Centro Destino', 'Producto', 'Clave',
                'Cantidad', 'Estado', 'Entregado Por'
            ]
            ws.append(headers)
            header_row = 5
            
            for col_num, cell in enumerate(ws[header_row], 1):
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            
            # Agrupar por centro para subtotales
            entregas_por_centro = defaultdict(list)
            
            # Datos
            row_num = header_row + 1
            for idx, entrega in enumerate(entregas, start=1):
                producto_nombre = ''
                producto_clave = ''
                
                if entrega.detalle_donacion:
                    producto_nombre = entrega.detalle_donacion.nombre_producto or ''
                    producto_clave = entrega.detalle_donacion.clave_producto or ''
                
                entregado_por_nombre = ''
                if entrega.entregado_por:
                    entregado_por_nombre = f"{entrega.entregado_por.first_name} {entrega.entregado_por.last_name}".strip()
                    if not entregado_por_nombre:
                        entregado_por_nombre = entrega.entregado_por.username
                
                fecha_str = entrega.fecha_entrega.strftime('%d/%m/%Y %H:%M') if entrega.fecha_entrega else ''
                centro_destino_nombre = entrega.centro_destino.nombre if entrega.centro_destino else entrega.destinatario or '-'
                estado = 'Entregado' if entrega.finalizado else 'Pendiente'
                
                # Guardar para resumen
                entregas_por_centro[centro_destino_nombre].append(entrega.cantidad or 0)
                
                row_data = [
                    idx,
                    fecha_str,
                    centro_destino_nombre,
                    producto_nombre,
                    producto_clave,
                    entrega.cantidad,
                    estado,
                    entregado_por_nombre
                ]
                ws.append(row_data)
                
                # Aplicar estilos a la fila
                for col, cell in enumerate(ws[row_num], 1):
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical='center')
                    if col == 6:  # Cantidad
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                
                row_num += 1
            
            # Agregar resumen por centro al final
            ws.append([])
            ws.append([])
            row_num += 2
            
            ws.merge_cells(f'A{row_num}:H{row_num}')
            resumen_titulo = ws[f'A{row_num}']
            resumen_titulo.value = 'RESUMEN POR CENTRO DESTINO'
            resumen_titulo.font = title_font
            resumen_titulo.alignment = Alignment(horizontal='center')
            row_num += 1
            
            # Encabezados resumen
            ws.append(['', '', 'Centro', '', '', 'Entregas', 'Unidades', ''])
            for col, cell in enumerate(ws[row_num], 1):
                if col in [3, 6, 7]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = thin_border
            row_num += 1
            
            # Datos resumen
            for centro, cantidades in sorted(entregas_por_centro.items()):
                ws.append(['', '', centro, '', '', len(cantidades), sum(cantidades), ''])
                for col, cell in enumerate(ws[row_num], 1):
                    if col in [3, 6, 7]:
                        cell.border = thin_border
                        if col in [6, 7]:
                            cell.alignment = Alignment(horizontal='center')
                row_num += 1
            
            # Total general
            ws.append(['', '', 'TOTAL GENERAL', '', '', total_entregas, total_cantidad, ''])
            for col, cell in enumerate(ws[row_num], 1):
                if col in [3, 6, 7]:
                    cell.fill = total_fill
                    cell.font = total_font
                    cell.border = thin_border
                    if col in [6, 7]:
                        cell.alignment = Alignment(horizontal='center')
            
            # Ajustar anchos de columna
            anchos = [6, 16, 35, 35, 12, 10, 12, 20]
            for col, ancho in enumerate(anchos, 1):
                ws.column_dimensions[get_column_letter(col)].width = ancho
            
            # Congelar encabezados
            ws.freeze_panes = 'A6'
            
            # Preparar respuesta
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
            # Nombre del archivo
            if centro_filtro:
                centro_slug = centro_nombre[:20].replace(' ', '_').replace('/', '-')
                filename = f'entregas_{centro_slug}_{timezone.now().strftime("%Y%m%d")}.xlsx'
            else:
                filename = f'entregas_todos_centros_{timezone.now().strftime("%Y%m%d")}.xlsx'
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            wb.save(response)
            return response
            
        except Exception as e:
            return Response(
                {'error': f'Error al exportar: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'], url_path='plantilla-excel')
    def plantilla_excel(self, request):
        """
        Genera una plantilla Excel para importación de entregas.
        Incluye ejemplos y validaciones.
        """
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.worksheet.datavalidation import DataValidation
        from django.http import HttpResponse
        
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Plantilla Entregas'
            
            # Título
            ws.merge_cells('A1:F1')
            titulo_cell = ws['A1']
            titulo_cell.value = 'PLANTILLA PARA IMPORTAR ENTREGAS DE DONACIONES'
            titulo_cell.font = Font(bold=True, size=14, color='632842')
            titulo_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Instrucciones
            ws.merge_cells('A2:F2')
            ws['A2'].value = 'Complete los datos siguiendo el formato indicado. Las columnas marcadas con * son obligatorias.'
            ws['A2'].font = Font(size=10, italic=True)
            
            ws.append([])
            
            # Encabezados
            headers = [
                'detalle_donacion_id *',  # ID del detalle de donación
                'cantidad *',             # Cantidad a entregar
                'destinatario *',         # Nombre del destinatario
                'motivo',                 # Motivo de la entrega
                'notas'                   # Notas adicionales
            ]
            ws.append(headers)
            
            # Estilo de encabezados
            header_fill = PatternFill(start_color='632842', end_color='632842', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            
            for col_num, cell in enumerate(ws[4], 1):
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Fila de ejemplo
            ws.append([1, 10, 'Juan Pérez', 'Tratamiento médico', 'Entrega programada'])
            
            # Agregar nota sobre detalle_donacion_id
            ws.append([])
            ws.append(['NOTA: El detalle_donacion_id lo puede obtener desde el inventario de donaciones.'])
            ws['A7'].font = Font(italic=True, color='666666')
            
            # Ajustar anchos
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 30
            ws.column_dimensions['D'].width = 30
            ws.column_dimensions['E'].width = 30
            
            # Segunda hoja con lista de detalles disponibles
            ws2 = wb.create_sheet(title='Inventario Disponible')
            ws2.merge_cells('A1:F1')
            ws2['A1'].value = 'INVENTARIO DE DONACIONES CON STOCK DISPONIBLE'
            ws2['A1'].font = Font(bold=True, size=12, color='632842')
            ws2['A1'].alignment = Alignment(horizontal='center')
            
            ws2.append([])
            headers2 = ['ID Detalle', 'Producto', 'Clave', 'Lote', 'Disponible', 'Donación']
            ws2.append(headers2)
            
            for cell in ws2[3]:
                cell.fill = header_fill
                cell.font = header_font
            
            # Obtener detalles con stock disponible
            from core.models import DetalleDonacion
            detalles = DetalleDonacion.objects.filter(
                cantidad_disponible__gt=0,
                donacion__estado='procesada'
            ).select_related('producto', 'donacion').order_by('producto__nombre')
            
            for det in detalles:
                ws2.append([
                    det.id,
                    det.producto.nombre if det.producto else '',
                    det.producto.clave if det.producto else '',
                    det.numero_lote or '',
                    det.cantidad_disponible,
                    det.donacion.numero if det.donacion else ''
                ])
            
            ws2.column_dimensions['A'].width = 12
            ws2.column_dimensions['B'].width = 40
            ws2.column_dimensions['C'].width = 15
            ws2.column_dimensions['D'].width = 15
            ws2.column_dimensions['E'].width = 12
            ws2.column_dimensions['F'].width = 15
            
            # Respuesta
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="plantilla_entregas_donaciones.xlsx"'
            
            wb.save(response)
            return response
            
        except Exception as e:
            return Response(
                {'error': f'Error al generar plantilla: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['post'], url_path='importar-excel')
    def importar_excel(self, request):
        """
        Importa entregas de donaciones desde un archivo Excel.
        
        Formato esperado:
        - detalle_donacion_id: ID del detalle de donación (obligatorio)
        - cantidad: Cantidad a entregar (obligatorio)
        - destinatario: Nombre del destinatario (obligatorio)
        - motivo: Motivo de la entrega (opcional)
        - notas: Notas adicionales (opcional)
        """
        import openpyxl
        from django.db import transaction
        from core.models import DetalleDonacion, SalidaDonacion
        
        archivo = request.FILES.get('archivo')
        if not archivo:
            return Response(
                {'error': 'No se proporcionó archivo'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if not archivo.name.endswith(('.xlsx', '.xls')):
            return Response(
                {'error': 'El archivo debe ser Excel (.xlsx o .xls)'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            wb = openpyxl.load_workbook(archivo, data_only=True)
            ws = wb.active
            
            # Buscar fila de encabezados (puede estar en fila 4 si usa plantilla)
            header_row = None
            for row_num in range(1, 10):
                cell_value = ws.cell(row=row_num, column=1).value
                if cell_value and 'detalle_donacion' in str(cell_value).lower():
                    header_row = row_num
                    break
            
            if not header_row:
                # Asumir que empieza en fila 1
                header_row = 1
            
            # Mapear columnas
            headers = {}
            for col_num, cell in enumerate(ws[header_row], 1):
                if cell.value:
                    header_name = str(cell.value).lower().strip()
                    header_name = header_name.replace(' *', '').replace('*', '')
                    headers[header_name] = col_num
            
            required_cols = ['detalle_donacion_id', 'cantidad', 'destinatario']
            for col in required_cols:
                if col not in headers:
                    return Response(
                        {'error': f'Columna requerida no encontrada: {col}'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
            
            # Procesar filas
            resultados = {
                'exitosos': 0,
                'fallidos': 0,
                'errores': []
            }
            
            with transaction.atomic():
                for row_num in range(header_row + 1, ws.max_row + 1):
                    # Verificar si la fila está vacía
                    detalle_id = ws.cell(row=row_num, column=headers['detalle_donacion_id']).value
                    if not detalle_id:
                        continue
                    
                    try:
                        cantidad = int(ws.cell(row=row_num, column=headers['cantidad']).value or 0)
                        destinatario = ws.cell(row=row_num, column=headers['destinatario']).value
                        motivo = ws.cell(row=row_num, column=headers.get('motivo', 0)).value if headers.get('motivo') else None
                        notas = ws.cell(row=row_num, column=headers.get('notas', 0)).value if headers.get('notas') else None
                        
                        # Validaciones
                        if not destinatario:
                            raise ValueError('Destinatario es requerido')
                        if cantidad <= 0:
                            raise ValueError('La cantidad debe ser mayor a 0')
                        
                        # Obtener detalle de donación
                        try:
                            detalle = DetalleDonacion.objects.select_related('donacion').get(pk=detalle_id)
                        except DetalleDonacion.DoesNotExist:
                            raise ValueError(f'Detalle de donación {detalle_id} no existe')
                        
                        # Verificar que la donación esté procesada
                        if detalle.donacion.estado != 'procesada':
                            raise ValueError(f'La donación {detalle.donacion.numero} no está procesada')
                        
                        # Verificar stock disponible
                        if cantidad > detalle.cantidad_disponible:
                            raise ValueError(
                                f'Stock insuficiente. Disponible: {detalle.cantidad_disponible}, Solicitado: {cantidad}'
                            )
                        
                        # Crear salida
                        salida = SalidaDonacion(
                            detalle_donacion=detalle,
                            cantidad=cantidad,
                            destinatario=str(destinatario).strip(),
                            motivo=str(motivo).strip() if motivo else None,
                            notas=str(notas).strip() if notas else None,
                            entregado_por=request.user
                        )
                        salida.save()
                        
                        resultados['exitosos'] += 1
                        
                    except Exception as e:
                        resultados['fallidos'] += 1
                        resultados['errores'].append({
                            'fila': row_num,
                            'error': str(e)
                        })
            
            return Response({
                'mensaje': f'Importación completada. Exitosos: {resultados["exitosos"]}, Fallidos: {resultados["fallidos"]}',
                'resultados': resultados
            })
            
        except Exception as e:
            return Response(
                {'error': f'Error al procesar archivo: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['post'], url_path='finalizar')
    def finalizar(self, request, pk=None):
        """
        Marca una entrega de donación como finalizada/entregada.
        
        ISS-DB-ALIGN: Ahora usa los campos reales de la BD:
        - finalizado: Boolean que indica si fue entregada
        - fecha_finalizado: Timestamp de la finalización
        - finalizado_por: Usuario que confirmó la entrega
        
        IMPORTANTE: El stock solo se descuenta al finalizar, no al crear la salida.
        Esto permite generar la hoja de entrega antes de confirmar.
        """
        from django.utils import timezone
        
        try:
            salida = self.get_object()
            
            # Verificar si ya está finalizada
            if salida.finalizado:
                return Response(
                    {'error': 'Esta entrega ya fue finalizada'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Usar el método del modelo que descuenta del inventario
            salida.finalizar(usuario=request.user)
            
            # Agregar nota de trazabilidad
            nota_finalizacion = f"\n[ENTREGADO] Confirmado por {request.user.username} el {timezone.now().strftime('%d/%m/%Y %H:%M')}"
            salida.notas = (salida.notas or '') + nota_finalizacion
            salida.save()
            
            logger.info(f"Salida de donación {salida.id} finalizada por {request.user.username} - Stock descontado")
            
            from core.serializers import SalidaDonacionSerializer
            return Response({
                'mensaje': 'Entrega finalizada y stock descontado correctamente',
                'salida': SalidaDonacionSerializer(salida).data
            })
        except ValueError as ve:
            return Response(
                {'error': str(ve)},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            logger.error(f"Error finalizando salida de donación: {str(e)}")
            return Response(
                {'error': f'Error al finalizar entrega: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['get'], url_path='recibo-pdf')
    def recibo_pdf(self, request, pk=None):
        """
        Genera un PDF de recibo para una salida de donación.
        Usa el formato oficial con "DONACIÓN" en el título.
        
        Parámetros:
        - finalizado: si es 'true', muestra sello de ENTREGADO en lugar de campos de firma
        - ids: lista de IDs separados por coma para incluir múltiples items (entregas agrupadas)
        
        ISS-DB-ALIGN: Ahora usa el campo 'finalizado' del modelo
        ISS-FIX: Soporta agrupación de entregas para mostrar todos los productos
        """
        from django.http import HttpResponse
        from django.utils import timezone
        from core.utils.pdf_generator import generar_hoja_entrega
        from datetime import timedelta
        
        try:
            salida = self.get_object()
            
            # ISS-FIX: Obtener IDs adicionales del query param o buscar entregas relacionadas
            ids_param = request.query_params.get('ids', '')
            salidas_ids = [int(i.strip()) for i in ids_param.split(',') if i.strip().isdigit()] if ids_param else []
            
            # Si no se pasaron IDs, buscar entregas del mismo grupo (mismo destinatario y minuto)
            if not salidas_ids:
                # Obtener el rango de tiempo (mismo minuto)
                fecha_base = salida.fecha_entrega
                if fecha_base:
                    # Crear rango de 1 minuto centrado en la fecha de entrega
                    fecha_inicio = fecha_base.replace(second=0, microsecond=0)
                    fecha_fin = fecha_inicio + timedelta(minutes=1)
                    
                    # Buscar otras salidas del mismo grupo
                    from core.models import SalidaDonacion
                    salidas_relacionadas = SalidaDonacion.objects.filter(
                        destinatario=salida.destinatario,
                        fecha_entrega__gte=fecha_inicio,
                        fecha_entrega__lt=fecha_fin
                    ).order_by('id')
                    salidas_ids = list(salidas_relacionadas.values_list('id', flat=True))
            
            # Asegurar que la salida actual esté en la lista
            if salida.id not in salidas_ids:
                salidas_ids = [salida.id] + salidas_ids
            
            # Obtener todas las salidas del grupo
            from core.models import SalidaDonacion
            salidas = SalidaDonacion.objects.filter(id__in=salidas_ids).select_related(
                'detalle_donacion__producto_donacion',
                'detalle_donacion__producto',
                'centro_destino'
            ).order_by('id')
            
            if not salidas.exists():
                salidas = [salida]  # Fallback a solo la salida actual
            
            # ISS-DB-ALIGN: Priorizar el campo real del modelo
            finalizado = salida.finalizado or request.query_params.get('finalizado', 'false').lower() == 'true'
            # Fallback: detectar por la nota (compatibilidad con datos antiguos)
            if not finalizado and salida.notas and '[ENTREGADO]' in salida.notas:
                finalizado = True
            
            # Construir datos para el PDF
            # ISS-DB-ALIGN: Usar centro_destino del modelo si existe
            centro_destino_nombre = salida.destinatario or 'Destinatario no especificado'
            if salida.centro_destino:
                centro_destino_nombre = salida.centro_destino.nombre
            
            # ISS-FIX: Construir lista de items con TODAS las salidas del grupo
            items = []
            for s in salidas:
                detalle = s.detalle_donacion
                producto_nombre = 'N/A'
                producto_clave = 'N/A'
                numero_lote = 'N/A'
                presentacion = 'N/A'
                
                if detalle:
                    # Usar las propiedades del modelo que manejan ambos casos
                    nombre = detalle.nombre_producto
                    if nombre and nombre != 'Sin producto':
                        producto_nombre = nombre
                    
                    # Obtener clave del producto
                    if detalle.producto_donacion:
                        producto_clave = detalle.producto_donacion.clave or 'N/A'
                    elif detalle.producto:
                        producto_clave = detalle.producto.clave or 'N/A'
                    
                    # numero_lote es un campo directo del modelo
                    if detalle.numero_lote:
                        numero_lote = detalle.numero_lote
                    
                    # Obtener presentación
                    if detalle.producto_donacion and detalle.producto_donacion.presentacion:
                        presentacion = detalle.producto_donacion.presentacion
                    elif detalle.producto_donacion and detalle.producto_donacion.unidad_medida:
                        presentacion = detalle.producto_donacion.unidad_medida
                    elif detalle.producto and detalle.producto.presentacion:
                        presentacion = detalle.producto.presentacion
                    elif detalle.producto and detalle.producto.unidad_medida:
                        presentacion = detalle.producto.unidad_medida
                
                items.append({
                    'clave': producto_clave,
                    'nombre': producto_nombre,
                    'descripcion': producto_nombre,
                    'presentacion': presentacion,
                    'lote': numero_lote,
                    'cantidad': s.cantidad,
                })
            
            logger.debug(f"PDF Donación - Generando con {len(items)} items para grupo de salidas: {salidas_ids}")
            
            # Preparar datos para generar_hoja_entrega con formato oficial
            fecha_obj = salida.fecha_entrega or timezone.now()
            if isinstance(fecha_obj, str):
                from datetime import datetime
                fecha_obj = datetime.fromisoformat(fecha_obj.replace('Z', '+00:00'))
            
            # Usar rango de IDs para el folio si hay múltiples
            if len(salidas_ids) > 1:
                folio = f"DON-{min(salidas_ids)}-{max(salidas_ids)}"
            else:
                folio = f"DON-{salida.id}"
            
            datos_entrega = {
                'grupo_salida': folio,
                'fecha': fecha_obj,
                'centro_destino': centro_destino_nombre,
                'items': items,
                # Datos de firma para el PDF (nombres en blanco para firma manual)
                'nombre_encargada': '',
                'cargo_encargada': 'ENCARGADA DEL ALMACÉN MÉDICO',
                'nombre_titular': '',
                'cargo_titular': 'TITULAR DE LA DIRECCIÓN DE REINSERCIÓN SOCIAL',
            }
            
            # Si está finalizado, usar fecha de finalización
            if finalizado:
                fecha_fin = salida.fecha_finalizado or salida.fecha_entrega
                if fecha_fin:
                    if isinstance(fecha_fin, str):
                        from datetime import datetime
                        fecha_fin = datetime.fromisoformat(fecha_fin.replace('Z', '+00:00'))
                    datos_entrega['fecha'] = fecha_fin
            
            # Generar PDF con formato oficial y "DONACIÓN" en título
            pdf_buffer = generar_hoja_entrega(
                datos_entrega,
                finalizado=finalizado,
                es_donacion=True  # Esto agrega "- DONACIÓN" al título
            )
            
            response = HttpResponse(
                pdf_buffer.getvalue(),
                content_type='application/pdf'
            )
            estado = 'Entregado' if finalizado else 'Pendiente'
            response['Content-Disposition'] = f'attachment; filename="Recibo_Donacion_{folio}_{estado}_{timezone.now().strftime("%Y%m%d")}.pdf"'
            
            logger.info(f"Recibo de donación generado para {len(items)} items (IDs: {salidas_ids}) por {request.user.username}")
            return response
            
        except Exception as e:
            logger.error(f"Error generando recibo de donación: {str(e)}")
            return Response(
                {'error': f'Error al generar recibo: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


# =============================================================================
# ISS-002 FIX: ENDPOINT DE CATÁLOGOS - Sincronizar enums frontend/backend
# =============================================================================

class CatalogosView(APIView):
    """
    Vista para obtener todos los catálogos del sistema.
    
    Expone las constantes definidas en backend para sincronizar con frontend.
    Esto evita discrepancias entre los valores válidos en front y backend.
    
    Endpoints:
    - GET /api/catalogos/ - Todos los catálogos
    - GET /api/catalogos/unidades-medida/ - Solo unidades de medida
    - GET /api/catalogos/categorias/ - Solo categorías
    - GET /api/catalogos/vias-administracion/ - Solo vías de administración
    - GET /api/catalogos/estados-requisicion/ - Solo estados de requisición
    - GET /api/catalogos/tipos-movimiento/ - Solo tipos de movimiento
    - GET /api/catalogos/roles/ - Solo roles de usuario
    """
    permission_classes = [AllowAny]  # Catálogos son públicos para formularios de login/registro
    
    def get(self, request, catalogo=None):
        from core.constants import (
            UNIDADES_MEDIDA,
            CATEGORIAS_PRODUCTO,
            ESTADOS_REQUISICION,
            TIPOS_MOVIMIENTO,
            ROLES_USUARIO,
        )
        
        # Vías de administración (no están en constants, definimos aquí)
        VIAS_ADMINISTRACION = [
            ('ORAL', 'Oral'),
            ('INTRAVENOSA', 'Intravenosa'),
            ('INTRAMUSCULAR', 'Intramuscular'),
            ('SUBCUTANEA', 'Subcutánea'),
            ('TOPICA', 'Tópica'),
            ('INHALATORIA', 'Inhalatoria'),
            ('RECTAL', 'Rectal'),
            ('OFTALMICA', 'Oftálmica'),
            ('OTICA', 'Ótica'),
            ('NASAL', 'Nasal'),
        ]
        
        # Construir respuesta según el catálogo solicitado
        catalogos = {
            'unidades_medida': [{'value': u[0], 'label': u[1]} for u in UNIDADES_MEDIDA],
            'categorias': [{'value': c[0], 'label': c[1]} for c in CATEGORIAS_PRODUCTO],
            'vias_administracion': [{'value': v[0], 'label': v[1]} for v in VIAS_ADMINISTRACION],
            'estados_requisicion': [{'value': e[0], 'label': e[1]} for e in ESTADOS_REQUISICION],
            'tipos_movimiento': [{'value': t[0], 'label': t[1]} for t in TIPOS_MOVIMIENTO],
            'roles': [{'value': r[0], 'label': r[1]} for r in ROLES_USUARIO],
        }
        
        # Si se solicita un catálogo específico
        if catalogo:
            catalogo_key = catalogo.replace('-', '_')
            if catalogo_key in catalogos:
                return Response({
                    catalogo_key: catalogos[catalogo_key]
                })
            return Response(
                {'error': f'Catálogo no encontrado: {catalogo}'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Retornar todos los catálogos
        return Response(catalogos)


# =============================================================================
# ADMIN: LIMPIEZA DE DATOS PARA REINICIO COMPLETO
# =============================================================================

class AdminLimpiarDatosView(APIView):
    """
    Vista EXCLUSIVA para SUPERUSUARIOS para limpiar datos operativos del sistema.
    
    Permite dejar el sistema "en blanco" para que farmacia y centros puedan
    empezar a usar el sistema desde cero después de capacitación.
    
    SOPORTA ELIMINACIÓN SELECTIVA:
    - productos: Elimina productos, imágenes, lotes, parcialidades, documentos, movimientos
    - lotes: Elimina lotes, documentos, parcialidades, hojas recolección (no productos)
    - requisiciones: Elimina requisiciones, detalles, historial, ajustes
    - movimientos: Elimina solo movimientos
    - donaciones: Elimina donaciones, detalles y salidas de donaciones
    - notificaciones: Elimina notificaciones de usuarios
    - dispensaciones: Elimina dispensaciones Formato C
    - pacientes: Elimina pacientes/internos
    - caja_chica: Elimina módulo de caja chica
    - todos: Elimina todo lo anterior
    
    NO ELIMINA (configuración del sistema):
    - Usuarios y sus perfiles
    - Centros
    - Configuración del sistema
    - Tema global (estilos)
    - Logs de auditoría (para mantener trazabilidad)
    - Permisos de Django
    - Grupos de Django
    
    Endpoints:
    - GET /api/admin/limpiar-datos/ - Obtener estadísticas detalladas
    - POST /api/admin/limpiar-datos/ - Ejecutar limpieza (requiere confirmación y categoría)
    
    Versión: 2.1 - Incluye parcialidades de lotes, dispensaciones, pacientes y caja chica
    """
    permission_classes = [IsAuthenticated, IsSuperuserOnly]
    
    def get(self, request):
        """
        Retorna estadísticas detalladas de lo que se eliminaría por categoría.
        """
        from core.models import (
            Producto, Lote, Movimiento, Requisicion, DetalleRequisicion,
            HojaRecoleccion, DetalleHojaRecoleccion, LoteDocumento,
            ProductoImagen, ImportacionLog, DetalleDonacion, Donacion, SalidaDonacion,
            Notificacion, ProductoDonacion,
            # Nuevos módulos
            Paciente, Dispensacion, DetalleDispensacion, HistorialDispensacion,
            CompraCajaChica, DetalleCompraCajaChica, InventarioCajaChica,
            MovimientoCajaChica, HistorialCompraCajaChica
        )
        from django.db import connection
        
        # Solo superusuarios
        if not request.user.is_superuser:
            return Response(
                {'error': 'Solo SUPERUSUARIOS pueden acceder a esta función'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        # Conteos
        productos_count = Producto.objects.count()
        lotes_count = Lote.objects.count()
        lotes_centros_count = Lote.objects.filter(centro__isnull=False).count()  # Lotes duplicados en centros
        movimientos_count = Movimiento.objects.count()
        movimientos_salida_count = Movimiento.objects.filter(tipo='salida').count()  # Solo salidas (para requisiciones)
        requisiciones_count = Requisicion.objects.count()
        detalles_req_count = DetalleRequisicion.objects.count()
        hojas_recoleccion_count = HojaRecoleccion.objects.count()
        
        # Conteos de donaciones
        donaciones_count = Donacion.objects.count()
        detalles_donacion_count = DetalleDonacion.objects.count()
        salidas_donacion_count = SalidaDonacion.objects.count()
        productos_donacion_count = ProductoDonacion.objects.count()
        
        # Conteo de notificaciones
        notificaciones_count = Notificacion.objects.count()
        
        # Conteos de dispensaciones (Formato C)
        pacientes_count = Paciente.objects.count()
        dispensaciones_count = Dispensacion.objects.count()
        detalle_dispensaciones_count = DetalleDispensacion.objects.count()
        historial_dispensaciones_count = HistorialDispensacion.objects.count()
        
        # Conteos de caja chica
        compras_caja_chica_count = CompraCajaChica.objects.count()
        detalle_compras_caja_chica_count = DetalleCompraCajaChica.objects.count()
        inventario_caja_chica_count = InventarioCajaChica.objects.count()
        movimientos_caja_chica_count = MovimientoCajaChica.objects.count()
        historial_compras_caja_chica_count = HistorialCompraCajaChica.objects.count()
        
        # Conteos con raw SQL para tablas sin modelo
        with connection.cursor() as cursor:
            cursor.execute("SELECT COUNT(*) FROM requisicion_ajustes_cantidad")
            ajustes_count = cursor.fetchone()[0]
            cursor.execute("SELECT COUNT(*) FROM requisicion_historial_estados")
            historial_count = cursor.fetchone()[0]
            cursor.execute("SELECT COUNT(*) FROM lote_parcialidades")
            lote_parcialidades_count = cursor.fetchone()[0]
            # Fingerprints de importación (tabla puede no existir)
            try:
                cursor.execute("SELECT COUNT(*) FROM parcialidad_import_fingerprints")
                fingerprints_count = cursor.fetchone()[0]
            except Exception:
                fingerprints_count = 0
        
        # Verificar si hay donaciones con productos
        productos_con_donaciones = DetalleDonacion.objects.values('producto_id').distinct().count()
        
        # Estadísticas organizadas por categoría
        # NOTA: 'total' debe reflejar TODOS los registros que se eliminarán realmente
        stats = {
            'categorias': {
                'productos': {
                    'nombre': 'Productos e Inventario',
                    'descripcion': 'Elimina productos, lotes, movimientos y requisiciones',
                    'total': (
                        # Productos
                        productos_count + ProductoImagen.objects.count() +
                        # Lotes
                        lotes_count + LoteDocumento.objects.count() + lote_parcialidades_count + fingerprints_count +
                        # Movimientos
                        movimientos_count +
                        # Requisiciones
                        requisiciones_count + detalles_req_count + historial_count + ajustes_count +
                        # Hojas
                        hojas_recoleccion_count + DetalleHojaRecoleccion.objects.count() +
                        # Dispensaciones (se eliminan por FK a productos)
                        dispensaciones_count + detalle_dispensaciones_count + historial_dispensaciones_count +
                        # Donaciones (se eliminan por FK a productos)
                        donaciones_count + detalles_donacion_count + salidas_donacion_count +
                        # Caja chica (se elimina por FK a productos)
                        compras_caja_chica_count + detalle_compras_caja_chica_count + inventario_caja_chica_count + movimientos_caja_chica_count + historial_compras_caja_chica_count
                        # NO incluye: pacientes, notificaciones (esos solo con 'todos')
                    ),
                    'detalle': {
                        'productos': productos_count,
                        'lotes': lotes_count,
                        'movimientos': movimientos_count,
                        'requisiciones': requisiciones_count,
                        'dispensaciones': dispensaciones_count,
                        'donaciones': donaciones_count,
                        'caja_chica': compras_caja_chica_count,
                    },
                    'dependencias': ['Elimina TODO el inventario'],
                },
                'lotes': {
                    'nombre': 'Solo Lotes',
                    'descripcion': 'Elimina lotes y movimientos asociados, mantiene productos',
                    'total': (
                        lotes_count + LoteDocumento.objects.count() + lote_parcialidades_count + fingerprints_count +
                        hojas_recoleccion_count + DetalleHojaRecoleccion.objects.count() +
                        movimientos_count  # TODOS los movimientos
                    ),
                    'detalle': {
                        'lotes': lotes_count,
                        'lote_documentos': LoteDocumento.objects.count(),
                        'lote_parcialidades': lote_parcialidades_count,
                        'fingerprints_importacion': fingerprints_count,
                        'movimientos': movimientos_count,
                        'hojas_recoleccion': hojas_recoleccion_count,
                    },
                    'dependencias': ['Los productos quedarán con stock 0'],
                },
                'requisiciones': {
                    'nombre': 'Requisiciones',
                    'descripcion': 'Elimina requisiciones, detalles, historial, movimientos de salida y lotes de centros',
                    'total': (
                        requisiciones_count + detalles_req_count + historial_count + ajustes_count +
                        hojas_recoleccion_count + DetalleHojaRecoleccion.objects.count() +
                        movimientos_salida_count + lotes_centros_count
                    ),
                    'detalle': {
                        'requisiciones': requisiciones_count,
                        'detalles_requisicion': detalles_req_count,
                        'requisicion_historial_estados': historial_count,
                        'requisicion_ajustes_cantidad': ajustes_count,
                        'hojas_recoleccion': hojas_recoleccion_count,
                        'movimientos_salida': movimientos_salida_count,
                        'lotes_centros': lotes_centros_count,
                    },
                    'dependencias': ['Los productos quedarán con stock 0'],
                },
                'movimientos': {
                    'nombre': 'Movimientos',
                    'descripcion': 'Elimina historial de movimientos y lotes duplicados en centros',
                    'total': movimientos_count + lotes_centros_count,
                    'detalle': {
                        'movimientos': movimientos_count,
                        'lotes_en_centros': lotes_centros_count,
                    },
                    'dependencias': ['También eliminará: lotes duplicados en centros (creados por salidas)'],
                },
                'donaciones': {
                    'nombre': 'Donaciones',
                    'descripcion': 'Elimina donaciones, detalles, salidas y catálogo de productos',
                    'total': donaciones_count + detalles_donacion_count + salidas_donacion_count + productos_donacion_count,
                    'detalle': {
                        'donaciones': donaciones_count,
                        'detalles_donacion': detalles_donacion_count,
                        'salidas_donacion': salidas_donacion_count,
                        'productos_donacion': productos_donacion_count,
                    },
                    'dependencias': ['Incluye el catálogo de productos de donación'],
                },
                'notificaciones': {
                    'nombre': 'Notificaciones',
                    'descripcion': 'Elimina todas las notificaciones de usuarios',
                    'total': notificaciones_count,
                    'detalle': {
                        'notificaciones': notificaciones_count,
                    },
                    'dependencias': [],
                },
                'dispensaciones': {
                    'nombre': 'Dispensaciones (Formato C)',
                    'descripcion': 'Elimina dispensaciones a pacientes, sus detalles e historial',
                    'total': dispensaciones_count + detalle_dispensaciones_count + historial_dispensaciones_count,
                    'detalle': {
                        'dispensaciones': dispensaciones_count,
                        'detalle_dispensaciones': detalle_dispensaciones_count,
                        'historial_dispensaciones': historial_dispensaciones_count,
                    },
                    'dependencias': ['NO elimina pacientes (deben eliminarse por separado)'],
                },
                'pacientes': {
                    'nombre': 'Pacientes/Internos',
                    'descripcion': 'Elimina el catálogo de pacientes/internos registrados',
                    'total': pacientes_count,
                    'detalle': {
                        'pacientes': pacientes_count,
                    },
                    'dependencias': ['También eliminará: todas las dispensaciones vinculadas a pacientes'],
                },
                'caja_chica': {
                    'nombre': 'Caja Chica Farmacia',
                    'descripcion': 'Elimina compras, detalles, inventario, movimientos e historial de caja chica',
                    'total': compras_caja_chica_count + detalle_compras_caja_chica_count + inventario_caja_chica_count + movimientos_caja_chica_count + historial_compras_caja_chica_count,
                    'detalle': {
                        'compras_caja_chica': compras_caja_chica_count,
                        'detalle_compras_caja_chica': detalle_compras_caja_chica_count,
                        'inventario_caja_chica': inventario_caja_chica_count,
                        'movimientos_caja_chica': movimientos_caja_chica_count,
                        'historial_compras_caja_chica': historial_compras_caja_chica_count,
                    },
                    'dependencias': [],
                },
                'todos': {
                    'nombre': 'Todo el Sistema',
                    'descripcion': 'Limpieza completa: productos, lotes, requisiciones, movimientos, donaciones, notificaciones, dispensaciones, pacientes y caja chica',
                    'total': (
                        # Productos
                        productos_count + ProductoImagen.objects.count() +
                        # Lotes
                        lotes_count + LoteDocumento.objects.count() + lote_parcialidades_count + fingerprints_count +
                        # Requisiciones
                        requisiciones_count + detalles_req_count + historial_count + ajustes_count +
                        # Hojas recolección
                        hojas_recoleccion_count + DetalleHojaRecoleccion.objects.count() +
                        # Movimientos
                        movimientos_count +
                        # Donaciones
                        donaciones_count + detalles_donacion_count + salidas_donacion_count + productos_donacion_count +
                        # Notificaciones
                        notificaciones_count +
                        # Dispensaciones
                        dispensaciones_count + detalle_dispensaciones_count + historial_dispensaciones_count +
                        # Pacientes
                        pacientes_count +
                        # Caja chica
                        compras_caja_chica_count + detalle_compras_caja_chica_count + inventario_caja_chica_count + 
                        movimientos_caja_chica_count + historial_compras_caja_chica_count
                    ),
                    'detalle': {
                        'productos': productos_count,
                        'lotes': lotes_count,
                        'requisiciones': requisiciones_count,
                        'movimientos': movimientos_count,
                        'donaciones': donaciones_count,
                        'notificaciones': notificaciones_count,
                        'dispensaciones': dispensaciones_count,
                        'pacientes': pacientes_count,
                        'caja_chica': compras_caja_chica_count + detalle_compras_caja_chica_count + inventario_caja_chica_count + movimientos_caja_chica_count + historial_compras_caja_chica_count,
                    },
                    'dependencias': ['ELIMINA TODO (incluyendo donaciones y notificaciones)'],
                },
            },
            'resumen': {
                'productos': productos_count,
                'lotes': lotes_count,
                'movimientos': movimientos_count,
                'requisiciones': requisiciones_count,
                'donaciones': donaciones_count,
                'notificaciones': notificaciones_count,
                'dispensaciones': dispensaciones_count,
                'pacientes': pacientes_count,
                'caja_chica': compras_caja_chica_count,
            },
            'no_se_eliminara': [
                'Usuarios y perfiles',
                'Centros',
                'Configuración del sistema',
                'Tema global (estilos)',
                'Logs de auditoría',
                'Permisos y grupos',
            ],
            'advertencias': [],
        }
        
        # Advertir si hay donaciones vinculadas a productos
        if productos_con_donaciones > 0:
            stats['advertencias'].append(
                f'Hay {productos_con_donaciones} productos vinculados a donaciones. '
                'El sistema de donaciones seguirá funcionando pero los productos '
                'aparecerán como "producto eliminado" en el historial de donaciones.'
            )
        
        return Response(stats)
    
    def post(self, request):
        """
        Ejecuta la limpieza de datos operativos según la categoría seleccionada.
        Requiere: {"confirmar": true, "categoria": "productos|lotes|requisiciones|movimientos|dispensaciones|pacientes|caja_chica|todos"}
        """
        from core.models import (
            Producto, Lote, Movimiento, Requisicion, DetalleRequisicion,
            HojaRecoleccion, DetalleHojaRecoleccion, LoteDocumento,
            ProductoImagen, ImportacionLog, AuditoriaLog,
            Donacion, DetalleDonacion, SalidaDonacion, Notificacion,
            # Nuevos módulos
            Paciente, Dispensacion, DetalleDispensacion, HistorialDispensacion,
            CompraCajaChica, DetalleCompraCajaChica, InventarioCajaChica,
            MovimientoCajaChica, HistorialCompraCajaChica
        )
        from django.db import connection
        
        # Solo superusuarios
        user = request.user
        if not user.is_superuser:
            return Response(
                {'error': 'Solo SUPERUSUARIOS pueden ejecutar esta función'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        # Verificar confirmación
        confirmar = request.data.get('confirmar', False)
        if not confirmar:
            return Response(
                {'error': 'Debe enviar {"confirmar": true} para ejecutar la limpieza'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Obtener categoría
        categoria = request.data.get('categoria', 'todos').lower()
        categorias_validas = ['productos', 'lotes', 'requisiciones', 'movimientos', 'donaciones', 'notificaciones', 'dispensaciones', 'pacientes', 'caja_chica', 'todos']
        
        if categoria not in categorias_validas:
            return Response(
                {'error': f'Categoría inválida. Use una de: {", ".join(categorias_validas)}'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            with transaction.atomic():
                eliminados = {}
                
                # ============================================================
                # PRE-FIX: Corregir lotes corruptos antes de cualquier operación
                # Lotes con cantidad_inicial=0 o NULL violan CHECK constraint
                # ============================================================
                with connection.cursor() as cursor:
                    cursor.execute("""
                        UPDATE lotes 
                        SET cantidad_inicial = GREATEST(1, cantidad_actual) 
                        WHERE cantidad_inicial IS NULL OR cantidad_inicial < 1
                    """)
                    lotes_corregidos = cursor.rowcount
                    if lotes_corregidos > 0:
                        eliminados['lotes_corregidos_pre'] = lotes_corregidos
                
                # ============================================================
                # ELIMINACIÓN SELECTIVA - Usar SQL directo para evitar 
                # problemas con modelos managed=False (Supabase)
                # ============================================================
                
                if categoria == 'movimientos':
                    # Movimientos y lotes duplicados en centros (creados por salidas)
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM movimientos")
                        eliminados['movimientos'] = cursor.rowcount
                    
                    # ISS-FIX: Eliminar lotes de centros (son duplicados creados por movimientos)
                    # Los lotes originales de Farmacia Central (centro_id IS NULL) se mantienen
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM lotes WHERE centro_id IS NOT NULL")
                        eliminados['lotes_centros'] = cursor.rowcount
                    
                    # ISS-FIX: Recalcular cantidad_inicial de lotes restantes
                    # Sin movimientos, cantidad_inicial debe ser igual a cantidad_actual
                    # NOTA: Usar GREATEST(1, cantidad_actual) para respetar CHECK constraint cantidad_inicial > 0
                    with connection.cursor() as cursor:
                        cursor.execute("UPDATE lotes SET cantidad_inicial = GREATEST(1, cantidad_actual) WHERE centro_id IS NULL")
                        eliminados['lotes_recalculados'] = cursor.rowcount
                
                elif categoria == 'donaciones':
                    # Eliminar donaciones en orden de dependencias FK
                    # 1. Salidas de donaciones
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM salidas_donaciones")
                        eliminados['salidas_donacion'] = cursor.rowcount
                    # 2. Detalles de donaciones
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM detalle_donaciones")
                        eliminados['detalles_donacion'] = cursor.rowcount
                    # 3. Donaciones
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM donaciones")
                        eliminados['donaciones'] = cursor.rowcount
                    # 4. Catálogo de productos de donación
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM productos_donacion")
                        eliminados['productos_donacion'] = cursor.rowcount
                
                elif categoria == 'notificaciones':
                    # Solo notificaciones
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM notificaciones")
                        eliminados['notificaciones'] = cursor.rowcount
                
                elif categoria == 'dispensaciones':
                    # Eliminar dispensaciones en orden de dependencias FK
                    # 1. Historial de dispensaciones
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM historial_dispensaciones")
                        eliminados['historial_dispensaciones'] = cursor.rowcount
                    # 2. Detalles de dispensaciones
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM detalle_dispensaciones")
                        eliminados['detalle_dispensaciones'] = cursor.rowcount
                    # 3. Dispensaciones
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM dispensaciones")
                        eliminados['dispensaciones'] = cursor.rowcount
                
                elif categoria == 'pacientes':
                    # Eliminar pacientes (y sus dispensaciones por cascada FK)
                    # 1. Primero historial de dispensaciones
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM historial_dispensaciones")
                        eliminados['historial_dispensaciones'] = cursor.rowcount
                    # 2. Detalles de dispensaciones
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM detalle_dispensaciones")
                        eliminados['detalle_dispensaciones'] = cursor.rowcount
                    # 3. Dispensaciones
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM dispensaciones")
                        eliminados['dispensaciones'] = cursor.rowcount
                    # 4. Pacientes
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM pacientes")
                        eliminados['pacientes'] = cursor.rowcount
                
                elif categoria == 'caja_chica':
                    # Eliminar caja chica en orden de dependencias FK
                    # 1. Historial de compras
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM historial_compras_caja_chica")
                        eliminados['historial_compras_caja_chica'] = cursor.rowcount
                    # 2. Movimientos de caja chica
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM movimientos_caja_chica")
                        eliminados['movimientos_caja_chica'] = cursor.rowcount
                    # 3. Detalles de compras
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM detalle_compras_caja_chica")
                        eliminados['detalle_compras_caja_chica'] = cursor.rowcount
                    # 4. Compras
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM compras_caja_chica")
                        eliminados['compras_caja_chica'] = cursor.rowcount
                    # 5. Inventario de caja chica
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM inventario_caja_chica")
                        eliminados['inventario_caja_chica'] = cursor.rowcount
                
                elif categoria == 'requisiciones':
                    # 1. Detalles de hojas de recolección (pueden tener FK a requisiciones)
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM detalle_hojas_recoleccion")
                        eliminados['detalles_hojas_recoleccion'] = cursor.rowcount
                    
                    # 2. Hojas de recolección
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM hojas_recoleccion")
                        eliminados['hojas_recoleccion'] = cursor.rowcount
                    
                    # 3. Ajustes de cantidad
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM requisicion_ajustes_cantidad")
                        eliminados['requisicion_ajustes_cantidad'] = cursor.rowcount
                    
                    # 4. Detalles de requisición
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM detalles_requisicion")
                        eliminados['detalles_requisicion'] = cursor.rowcount
                    
                    # 5. Historial de estados
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM requisicion_historial_estados")
                        eliminados['requisicion_historial_estados'] = cursor.rowcount
                    
                    # 6. Movimientos de SALIDA (tipo='salida') - son los de requisiciones
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM movimientos WHERE tipo = 'salida'")
                        eliminados['movimientos_salida'] = cursor.rowcount
                    
                    # ISS-FIX: Eliminar lotes de centros (creados por requisiciones)
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM lotes WHERE centro_id IS NOT NULL")
                        eliminados['lotes_centros'] = cursor.rowcount
                    
                    # 7. Requisiciones
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM requisiciones")
                        eliminados['requisiciones'] = cursor.rowcount
                
                elif categoria == 'lotes':
                    # 1. Movimientos - TODOS porque sin lotes no tienen sentido
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM movimientos")
                        eliminados['movimientos'] = cursor.rowcount
                    
                    # 2. Detalles de hojas de recolección
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM detalle_hojas_recoleccion")
                        eliminados['detalles_hojas_recoleccion'] = cursor.rowcount
                    
                    # 3. Hojas de recolección
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM hojas_recoleccion")
                        eliminados['hojas_recoleccion'] = cursor.rowcount
                    
                    # 4. Documentos de lotes
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM lote_documentos")
                        eliminados['lote_documentos'] = cursor.rowcount
                    
                    # 4.1 Parcialidades de lotes (entregas parciales)
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM lote_parcialidades")
                        eliminados['lote_parcialidades'] = cursor.rowcount
                    
                    # 4.2 Fingerprints de importación de parcialidades
                    with connection.cursor() as cursor:
                        try:
                            cursor.execute("DELETE FROM parcialidad_import_fingerprints")
                            eliminados['parcialidad_import_fingerprints'] = cursor.rowcount
                        except Exception:
                            pass  # Tabla puede no existir en algunos entornos
                    
                    # 5. Limpiar referencias a lotes en detalles_requisicion
                    with connection.cursor() as cursor:
                        cursor.execute("UPDATE detalles_requisicion SET lote_id = NULL WHERE lote_id IS NOT NULL")
                        eliminados['detalles_requisicion_limpios'] = cursor.rowcount
                    
                    # 5.1 Limpiar referencias a lotes en detalle_dispensaciones
                    with connection.cursor() as cursor:
                        cursor.execute("UPDATE detalle_dispensaciones SET lote_id = NULL WHERE lote_id IS NOT NULL")
                        eliminados['detalle_dispensaciones_limpios'] = cursor.rowcount
                    
                    # 6. Lotes - DELETE con CASCADE en mente
                    with connection.cursor() as cursor:
                        # Primero contar cuántos hay
                        cursor.execute("SELECT COUNT(*) FROM lotes")
                        lotes_antes = cursor.fetchone()[0]
                        
                        # Ejecutar DELETE
                        cursor.execute("DELETE FROM lotes")
                        eliminados['lotes'] = cursor.rowcount
                        
                        # Verificar que realmente se eliminaron
                        cursor.execute("SELECT COUNT(*) FROM lotes")
                        lotes_despues = cursor.fetchone()[0]
                        
                        if lotes_despues > 0:
                            logger.warning(f"⚠️ LOTES NO ELIMINADOS COMPLETAMENTE: {lotes_despues} permanecen en BD")
                            eliminados['lotes_restantes_advertencia'] = lotes_despues
                        else:
                            logger.info(f"✅ Todos los lotes eliminados correctamente ({lotes_antes} -> {lotes_despues})")
                    
                    # 7. Actualizar stock de productos a 0
                    with connection.cursor() as cursor:
                        cursor.execute("UPDATE productos SET stock_actual = 0")
                
                elif categoria == 'productos':
                    # Elimina productos Y todo lo que depende de ellos
                    # ORDEN FK-SAFE: primero eliminar tablas hijas que referencian productos
                    
                    # ====== CAJA CHICA (detalle_compras_caja_chica.producto_id → productos) ======
                    # 0. Historial de compras caja chica
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM historial_compras_caja_chica")
                        eliminados['historial_compras_caja_chica'] = cursor.rowcount
                    # 0.1 Movimientos de caja chica
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM movimientos_caja_chica")
                        eliminados['movimientos_caja_chica'] = cursor.rowcount
                    # 0.2 Detalles de compras caja chica (referencia producto_id)
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM detalle_compras_caja_chica")
                        eliminados['detalle_compras_caja_chica'] = cursor.rowcount
                    # 0.3 Compras caja chica
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM compras_caja_chica")
                        eliminados['compras_caja_chica'] = cursor.rowcount
                    # 0.4 Inventario caja chica (referencia producto_id)
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM inventario_caja_chica")
                        eliminados['inventario_caja_chica'] = cursor.rowcount
                    
                    # ====== DISPENSACIONES (detalle_dispensaciones.producto_id → ON DELETE RESTRICT) ======
                    # 1. Historial de dispensaciones
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM historial_dispensaciones")
                        eliminados['historial_dispensaciones'] = cursor.rowcount
                    # 2. Detalles de dispensaciones (referencia producto_id)
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM detalle_dispensaciones")
                        eliminados['detalle_dispensaciones'] = cursor.rowcount
                    # 3. Dispensaciones
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM dispensaciones")
                        eliminados['dispensaciones'] = cursor.rowcount
                    
                    # ====== DONACIONES (detalle_donaciones.producto_id → ON DELETE RESTRICT) ======
                    # 4. Salidas de donación
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM salidas_donaciones")
                        eliminados['salidas_donacion'] = cursor.rowcount
                    # 5. Detalles de donación (referencia producto_id)
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM detalle_donaciones")
                        eliminados['detalles_donacion'] = cursor.rowcount
                    # 6. Donaciones
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM donaciones")
                        eliminados['donaciones'] = cursor.rowcount
                    
                    # ====== HOJAS DE RECOLECCIÓN ======
                    # 7. Detalles de hojas de recolección
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM detalle_hojas_recoleccion")
                        eliminados['detalles_hojas_recoleccion'] = cursor.rowcount
                    # 8. Hojas de recolección
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM hojas_recoleccion")
                        eliminados['hojas_recoleccion'] = cursor.rowcount
                    
                    # ====== REQUISICIONES ======
                    # 9. Ajustes de cantidad
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM requisicion_ajustes_cantidad")
                        eliminados['requisicion_ajustes_cantidad'] = cursor.rowcount
                    # 10. Detalles de requisición
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM detalles_requisicion")
                        eliminados['detalles_requisicion'] = cursor.rowcount
                    # 11. Historial de estados
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM requisicion_historial_estados")
                        eliminados['requisicion_historial_estados'] = cursor.rowcount
                    # 12. Movimientos
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM movimientos")
                        eliminados['movimientos'] = cursor.rowcount
                    # 13. Requisiciones
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM requisiciones")
                        eliminados['requisiciones'] = cursor.rowcount
                    
                    # ====== LOTES ======
                    # 14. Documentos de lotes
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM lote_documentos")
                        eliminados['lote_documentos'] = cursor.rowcount
                    # 14.1 Parcialidades de lotes (entregas parciales)
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM lote_parcialidades")
                        eliminados['lote_parcialidades'] = cursor.rowcount
                    
                    # 14.2 Fingerprints de importación de parcialidades
                    with connection.cursor() as cursor:
                        try:
                            cursor.execute("DELETE FROM parcialidad_import_fingerprints")
                            eliminados['parcialidad_import_fingerprints'] = cursor.rowcount
                        except Exception:
                            pass  # Tabla puede no existir en algunos entornos
                    
                    # 15. Lotes
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM lotes")
                        eliminados['lotes'] = cursor.rowcount
                    
                    # ====== PRODUCTOS (ahora seguro - sin hijos que lo referencien) ======
                    # 16. Imágenes de productos
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM producto_imagenes")
                        eliminados['producto_imagenes'] = cursor.rowcount
                    # 17. Productos
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM productos")
                        eliminados['productos'] = cursor.rowcount
                    
                    # 18. Logs de importación
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM importacion_logs")
                        eliminados['importacion_logs'] = cursor.rowcount
                
                else:  # categoria == 'todos'
                    # LIMPIEZA COMPLETA - Orden FK-safe: tablas hijas ANTES que padres
                    # CRÍTICO: detalle_dispensaciones, detalle_donaciones y detalle_compras_caja_chica tienen
                    # producto_id con ON DELETE RESTRICT → deben eliminarse ANTES de productos
                    
                    # 0. Limpiar FKs que referencian lotes antes de eliminarlos
                    with connection.cursor() as cursor:
                        cursor.execute("UPDATE detalle_dispensaciones SET lote_id = NULL WHERE lote_id IS NOT NULL")
                        cursor.execute("UPDATE detalles_requisicion SET lote_id = NULL WHERE lote_id IS NOT NULL")
                    
                    # ==================== CAJA CHICA FARMACIA ====================
                    # Debe eliminarse ANTES de productos (detalle_compras_caja_chica.producto_id, inventario_caja_chica.producto_id)
                    # 0a. Historial de compras caja chica
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM historial_compras_caja_chica")
                        eliminados['historial_compras_caja_chica'] = cursor.rowcount
                    
                    # 0b. Movimientos de caja chica
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM movimientos_caja_chica")
                        eliminados['movimientos_caja_chica'] = cursor.rowcount
                    
                    # 0c. Detalles de compras caja chica (referencia producto_id)
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM detalle_compras_caja_chica")
                        eliminados['detalle_compras_caja_chica'] = cursor.rowcount
                    
                    # 0d. Compras caja chica
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM compras_caja_chica")
                        eliminados['compras_caja_chica'] = cursor.rowcount
                    
                    # 0e. Inventario caja chica (referencia producto_id)
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM inventario_caja_chica")
                        eliminados['inventario_caja_chica'] = cursor.rowcount
                    
                    # ==================== DISPENSACIONES (FORMATO C) ====================
                    # Deben eliminarse ANTES de productos (detalle_dispensaciones.producto_id)
                    # 1. Historial de dispensaciones
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM historial_dispensaciones")
                        eliminados['historial_dispensaciones'] = cursor.rowcount
                    
                    # 2. Detalles de dispensaciones
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM detalle_dispensaciones")
                        eliminados['detalle_dispensaciones'] = cursor.rowcount
                    
                    # 3. Dispensaciones
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM dispensaciones")
                        eliminados['dispensaciones'] = cursor.rowcount
                    
                    # 4. Pacientes
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM pacientes")
                        eliminados['pacientes'] = cursor.rowcount
                    
                    # ==================== DONACIONES ====================
                    # Deben eliminarse ANTES de productos (detalle_donaciones.producto_id)
                    # 5. Salidas de donación
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM salidas_donaciones")
                        eliminados['salidas_donacion'] = cursor.rowcount
                    
                    # 6. Detalles de donación
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM detalle_donaciones")
                        eliminados['detalles_donacion'] = cursor.rowcount
                    
                    # 7. Donaciones
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM donaciones")
                        eliminados['donaciones'] = cursor.rowcount
                    
                    # 8. Catálogo de productos de donación
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM productos_donacion")
                        eliminados['productos_donacion'] = cursor.rowcount
                    
                    # ==================== NOTIFICACIONES ====================
                    # 9. Notificaciones (sin FKs críticas)
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM notificaciones")
                        eliminados['notificaciones'] = cursor.rowcount
                    
                    # ==================== REQUISICIONES Y HOJAS ====================
                    # 10. Detalles de hojas de recolección
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM detalle_hojas_recoleccion")
                        eliminados['detalles_hojas_recoleccion'] = cursor.rowcount
                    
                    # 11. Hojas de recolección
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM hojas_recoleccion")
                        eliminados['hojas_recoleccion'] = cursor.rowcount
                    
                    # 12. Ajustes de cantidad
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM requisicion_ajustes_cantidad")
                        eliminados['requisicion_ajustes_cantidad'] = cursor.rowcount
                    
                    # 13. Detalles de requisición
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM detalles_requisicion")
                        eliminados['detalles_requisicion'] = cursor.rowcount
                    
                    # 14. Historial de estados
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM requisicion_historial_estados")
                        eliminados['requisicion_historial_estados'] = cursor.rowcount
                    
                    # 15. Movimientos
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM movimientos")
                        eliminados['movimientos'] = cursor.rowcount
                    
                    # 16. Requisiciones
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM requisiciones")
                        eliminados['requisiciones'] = cursor.rowcount
                    
                    # ==================== LOTES ====================
                    # 17. Documentos de lotes
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM lote_documentos")
                        eliminados['lote_documentos'] = cursor.rowcount
                    
                    # 17.1 Parcialidades de lotes (entregas parciales)
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM lote_parcialidades")
                        eliminados['lote_parcialidades'] = cursor.rowcount
                    
                    # 17.2 Fingerprints de importación de parcialidades
                    with connection.cursor() as cursor:
                        try:
                            cursor.execute("DELETE FROM parcialidad_import_fingerprints")
                            eliminados['parcialidad_import_fingerprints'] = cursor.rowcount
                        except Exception:
                            pass  # Tabla puede no existir en algunos entornos
                    
                    # 18. Lotes - con verificación
                    with connection.cursor() as cursor:
                        cursor.execute("SELECT COUNT(*) FROM lotes")
                        lotes_antes = cursor.fetchone()[0]
                        
                        cursor.execute("DELETE FROM lotes")
                        eliminados['lotes'] = cursor.rowcount
                        
                        cursor.execute("SELECT COUNT(*) FROM lotes")
                        lotes_despues = cursor.fetchone()[0]
                        
                        if lotes_despues > 0:
                            logger.error(f"❌ {lotes_despues} LOTES NO ELIMINADOS")
                            eliminados['lotes_restantes_advertencia'] = lotes_despues
                        else:
                            logger.info(f"✅ Lotes eliminados: {lotes_antes}")
                    
                    # ==================== PRODUCTOS ====================
                    # AHORA SEGURO: todas las tablas hijas ya fueron eliminadas
                    # 19. Imágenes de productos
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM producto_imagenes")
                        eliminados['producto_imagenes'] = cursor.rowcount
                    
                    # 20. Productos
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM productos")
                        eliminados['productos'] = cursor.rowcount
                    
                    # 21. Logs de importación
                    with connection.cursor() as cursor:
                        cursor.execute("DELETE FROM importacion_logs")
                        eliminados['importacion_logs'] = cursor.rowcount
                
                # Calcular totales
                total_eliminados = sum(eliminados.values())
                
                # ISS-FIX: INVALIDAR CACHÉ DEL DASHBOARD después de limpiar datos
                # Esto asegura que el dashboard se actualice inmediatamente
                from django.core.cache import cache
                from core.models import Centro
                
                # Invalidar caché global
                cache.delete('dashboard_resumen_global')
                cache.delete('dashboard_graficas_global')
                
                # Invalidar caché de cada centro
                for centro in Centro.objects.all():
                    cache.delete(f'dashboard_resumen_{centro.id}')
                    cache.delete(f'dashboard_graficas_{centro.id}')
                
                logger.info("Caché del dashboard invalidado después de limpieza de datos")
                
                # Nombres de categorías para el log
                nombres_categorias = {
                    'productos': 'PRODUCTOS E INVENTARIO',
                    'lotes': 'SOLO LOTES',
                    'requisiciones': 'REQUISICIONES',
                    'movimientos': 'MOVIMIENTOS',
                    'donaciones': 'DONACIONES',
                    'notificaciones': 'NOTIFICACIONES',
                    'dispensaciones': 'DISPENSACIONES (FORMATO C)',
                    'pacientes': 'PACIENTES/INTERNOS',
                    'caja_chica': 'CAJA CHICA FARMACIA',
                    'todos': 'TODO EL SISTEMA (INCLUYE DISPENSACIONES, CAJA CHICA Y DONACIONES)',
                }
                
                # Registrar en auditoría (NO se elimina)
                AuditoriaLog.objects.create(
                    usuario=user,
                    accion='LIMPIEZA_DATOS',
                    modelo='SISTEMA',
                    objeto_id=None,
                    datos_anteriores=None,
                    datos_nuevos=eliminados,
                    detalles={
                        'tipo': f'LIMPIEZA_SELECTIVA_{categoria.upper()}',
                        'categoria': categoria,
                        'categoria_nombre': nombres_categorias.get(categoria, categoria),
                        'ejecutado_por': user.username,
                        'email': user.email,
                        'fecha': timezone.now().isoformat(),
                        'registros_eliminados': total_eliminados,
                    },
                    ip_address=request.META.get('REMOTE_ADDR', ''),
                    user_agent=request.META.get('HTTP_USER_AGENT', '')[:500] if request.META.get('HTTP_USER_AGENT') else None,
                )
                
                logger.warning(
                    f"🗑️ LIMPIEZA DE DATOS [{categoria.upper()}] ejecutada por {user.username} ({user.email}): "
                    f"{total_eliminados} registros eliminados. Detalle: {eliminados}"
                )
                
                return Response({
                    'success': True,
                    'mensaje': f'✅ Limpieza de {nombres_categorias.get(categoria, categoria)} completada exitosamente.',
                    'categoria': categoria,
                    'eliminados': eliminados,
                    'total_registros_eliminados': total_eliminados,
                    'no_eliminado': [
                        'Usuarios y perfiles',
                        'Centros', 
                        'Configuración del sistema',
                        'Tema global',
                        'Auditoría',
                    ],
                    'ejecutado_por': user.username,
                    'fecha': timezone.now().isoformat(),
                })
                
        except Exception as e:
            logger.error(f"❌ Error en limpieza de datos [{categoria}]: {e}", exc_info=True)
            return Response(
                {'error': f'Error al limpiar datos: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


# =============================================================================
# MÓDULO DE DISPENSACIÓN A PACIENTES (FORMATO C)
# =============================================================================

from core.models import Paciente, Dispensacion, DetalleDispensacion, HistorialDispensacion
from core.serializers import (
    PacienteSerializer, PacienteSimpleSerializer,
    DispensacionSerializer, DispensacionListSerializer,
    DetalleDispensacionSerializer, HistorialDispensacionSerializer
)


class PacienteViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestión de Pacientes/Internos.
    
    ISS-DISP: Este módulo es operado por MÉDICO y CENTRO.
    FARMACIA solo puede auditar (ver) pero NO crear/editar.
    
    Proporciona operaciones CRUD completas con filtrado por centro,
    búsqueda por expediente/nombre y exportación.
    """
    queryset = Paciente.objects.all()
    serializer_class = PacienteSerializer
    permission_classes = [IsAuthenticated, CanManageDispensaciones]
    pagination_class = StandardResultsSetPagination
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['numero_expediente', 'nombre', 'apellido_paterno', 'apellido_materno', 'curp']
    ordering_fields = ['numero_expediente', 'nombre', 'apellido_paterno', 'created_at']
    ordering = ['-created_at']
    
    def get_queryset(self):
        """Filtra pacientes según centro del usuario y parámetros de búsqueda"""
        queryset = super().get_queryset().select_related('centro', 'created_by')
        user = self.request.user
        
        # SEGURIDAD: Filtrar por centro del usuario
        # - Superusuario y farmacia ven todo
        # - Centro solo ve lo suyo
        # - Sin centro = sin datos
        if user.is_superuser or user.rol in ['admin', 'admin_sistema', 'farmacia']:
            # Admin y farmacia pueden ver todo (para auditoría)
            pass
        elif user.centro:
            # Usuario de centro: SOLO ve pacientes de su centro
            queryset = queryset.filter(centro=user.centro)
        else:
            # Sin centro asignado = no ve nada (seguridad)
            return queryset.none()
        
        # Filtros opcionales (solo si tiene permiso de ver)
        centro_id = self.request.query_params.get('centro')
        if centro_id:
            # Si es usuario de centro, ignorar este filtro (ya está filtrado)
            if user.is_superuser or user.rol in ['admin', 'admin_sistema', 'farmacia']:
                queryset = queryset.filter(centro_id=centro_id)
        
        activo = self.request.query_params.get('activo')
        if activo is not None:
            queryset = queryset.filter(activo=activo.lower() == 'true')
        
        sexo = self.request.query_params.get('sexo')
        if sexo:
            queryset = queryset.filter(sexo=sexo.upper())
        
        dormitorio = self.request.query_params.get('dormitorio')
        if dormitorio:
            queryset = queryset.filter(dormitorio__icontains=dormitorio)
        
        return queryset
    
    def get_serializer_class(self):
        if self.action == 'list' and self.request.query_params.get('simple') == 'true':
            return PacienteSimpleSerializer
        return PacienteSerializer
    
    def create(self, request, *args, **kwargs):
        """
        Crea un paciente con validación de centro.
        
        ISS-SEC-FIX: Usuarios de centro SOLO pueden crear pacientes en su propio centro.
        Admin/farmacia pueden especificar cualquier centro.
        """
        user = request.user
        data = request.data.copy()
        
        # ISS-SEC-FIX: Validación estricta de centro
        if user.is_superuser or user.rol in ['admin', 'admin_sistema', 'farmacia']:
            # Admin/farmacia pueden especificar centro libremente
            pass
        else:
            # Usuario de centro: FORZAR su centro (ignorar payload)
            if user.centro:
                data['centro'] = user.centro.id
            else:
                return Response(
                    {'detail': 'Usuario sin centro asignado no puede crear pacientes'},
                    status=status.HTTP_403_FORBIDDEN
                )
        
        serializer = self.get_serializer(data=data)
        serializer.is_valid(raise_exception=True)
        serializer.save(created_by=user)
        
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    
    def update(self, request, *args, **kwargs):
        """
        Actualiza un paciente con validación de centro.
        
        ISS-SEC-FIX: Usuarios de centro SOLO pueden editar pacientes de su propio centro.
        """
        user = request.user
        instance = self.get_object()
        data = request.data.copy()
        
        # ISS-SEC-FIX: Validar permisos sobre el paciente
        if not (user.is_superuser or user.rol in ['admin', 'admin_sistema', 'farmacia']):
            # Verificar que el paciente sea del centro del usuario
            if instance.centro_id != user.centro_id:
                return Response(
                    {'detail': 'No tiene permisos para editar este paciente'},
                    status=status.HTTP_403_FORBIDDEN
                )
            # No permitir cambiar el centro
            if 'centro' in data:
                data['centro'] = user.centro.id  # Forzar su centro
        
        serializer = self.get_serializer(instance, data=data, partial=kwargs.get('partial', False))
        serializer.is_valid(raise_exception=True)
        serializer.save()
        
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def autocomplete(self, request):
        """Endpoint para autocompletado de pacientes"""
        query = request.query_params.get('q', '')
        centro_id = request.query_params.get('centro')
        limit = int(request.query_params.get('limit', 10))
        
        queryset = self.get_queryset().filter(activo=True)
        
        if query:
            queryset = queryset.filter(
                Q(numero_expediente__icontains=query) |
                Q(nombre__icontains=query) |
                Q(apellido_paterno__icontains=query) |
                Q(apellido_materno__icontains=query)
            )
        
        if centro_id:
            queryset = queryset.filter(centro_id=centro_id)
        
        queryset = queryset[:limit]
        serializer = PacienteSimpleSerializer(queryset, many=True)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def historial_dispensaciones(self, request, pk=None):
        """Obtiene el historial de dispensaciones de un paciente"""
        paciente = self.get_object()
        dispensaciones = paciente.dispensaciones.all().order_by('-fecha_dispensacion')[:50]
        serializer = DispensacionListSerializer(dispensaciones, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def exportar_excel(self, request):
        """Exporta la lista de pacientes a Excel"""
        queryset = self.get_queryset()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pacientes"
        
        # Encabezados
        headers = [
            'No. Expediente', 'Nombre', 'Apellido Paterno', 'Apellido Materno',
            'CURP', 'Fecha Nacimiento', 'Sexo', 'Centro', 'Dormitorio', 'Celda',
            'Tipo Sangre', 'Activo', 'Fecha Ingreso'
        ]
        ws.append(headers)
        
        # Datos
        for paciente in queryset:
            ws.append([
                paciente.numero_expediente,
                paciente.nombre,
                paciente.apellido_paterno,
                paciente.apellido_materno or '',
                paciente.curp or '',
                str(paciente.fecha_nacimiento) if paciente.fecha_nacimiento else '',
                paciente.get_sexo_display() if paciente.sexo else '',
                paciente.centro.nombre if paciente.centro else '',
                paciente.dormitorio or '',
                paciente.celda or '',
                paciente.tipo_sangre or '',
                'Sí' if paciente.activo else 'No',
                str(paciente.fecha_ingreso) if paciente.fecha_ingreso else '',
            ])
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in column)
            ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=pacientes_{timezone.now().strftime("%Y%m%d")}.xlsx'
        return response

    @action(detail=False, methods=['get'])
    def plantilla_importacion(self, request):
        """Descarga plantilla Excel para importación de pacientes/PPL"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "PPL"
        
        # Encabezados - TODOS obligatorios excepto información médica
        headers = [
            'numero_expediente*', 'curp*', 'nombre*', 'apellido_paterno*', 'apellido_materno*',
            'fecha_nacimiento*', 'sexo*', 'centro*', 'dormitorio*', 'celda*',
            'tipo_sangre', 'alergias', 'enfermedades_cronicas', 'observaciones_medicas',
            'fecha_ingreso'
        ]
        ws.append(headers)
        
        # Obtener los centros para mostrar ejemplo real
        centros = Centro.objects.filter(activo=True).order_by('nombre')[:1]
        centro_ejemplo = centros.first().nombre if centros.exists() else 'CENTRO PENITENCIARIO EJEMPLO'
        
        # Fila de ejemplo
        ws.append([
            'EXP-001234', 'PELJ900101HDFRRS09', 'Juan', 'Pérez', 'López',
            '1990-01-15', 'M', centro_ejemplo, 'Dorm. 5', 'Celda 12',
            'O+', 'Penicilina', 'Diabetes', 'Requiere dieta especial',
            '2024-06-01'
        ])
        
        # Hoja de centros disponibles
        ws_centros = wb.create_sheet("Centros Disponibles")
        ws_centros.append(['ID', 'Nombre del Centro'])
        for centro in Centro.objects.filter(activo=True).order_by('nombre'):
            ws_centros.append([centro.id, centro.nombre])
        ws_centros.column_dimensions['A'].width = 10
        ws_centros.column_dimensions['B'].width = 50
        
        # Estilo para encabezados de centros
        from openpyxl.styles import Font, PatternFill
        header_fill_centros = PatternFill(start_color="235B4E", end_color="235B4E", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws_centros[1]:
            cell.fill = header_fill_centros
            cell.font = header_font
        
        # Instrucciones en otra hoja
        ws_inst = wb.create_sheet("Instrucciones")
        instrucciones = [
            ["INSTRUCCIONES PARA IMPORTACIÓN DE PPL"],
            ["(Personas Privadas de la Libertad)"],
            [""],
            ["Campos OBLIGATORIOS (marcados con *):"],
            ["- numero_expediente*: Identificador único del PPL (ej: EXP-001234)"],
            ["- curp*: CURP a 18 caracteres exactos"],
            ["- nombre*: Nombre(s) del PPL"],
            ["- apellido_paterno*: Primer apellido"],
            ["- apellido_materno*: Segundo apellido"],
            ["- fecha_nacimiento*: Formato YYYY-MM-DD (ej: 1990-01-15)"],
            ["- sexo*: M (Masculino) o F (Femenino)"],
            ["- centro*: Nombre EXACTO del centro (ver hoja 'Centros Disponibles')"],
            ["- dormitorio*: Identificador del dormitorio (ej: Dorm. 5)"],
            ["- celda*: Número o identificador de celda (ej: Celda 12)"],
            [""],
            ["Campos opcionales (información médica):"],
            ["- tipo_sangre: O+, O-, A+, A-, B+, B-, AB+, AB-"],
            ["- alergias: Texto libre con alergias conocidas"],
            ["- enfermedades_cronicas: Texto libre"],
            ["- observaciones_medicas: Notas adicionales"],
            ["- fecha_ingreso: Fecha de ingreso al centro (YYYY-MM-DD)"],
            [""],
            ["NOTAS IMPORTANTES:"],
            ["- Todos los campos marcados con * son OBLIGATORIOS"],
            ["- El nombre del centro debe coincidir EXACTAMENTE (ver hoja 'Centros Disponibles')"],
            ["- Si el expediente ya existe, se ACTUALIZAN los datos"],
            ["- El CURP debe tener exactamente 18 caracteres"],
            ["- Máximo 1000 registros por archivo"],
            ["- Solo archivos .xlsx"],
        ]
        for row in instrucciones:
            ws_inst.append(row)
        ws_inst.column_dimensions['A'].width = 70
        
        # Estilo de encabezados
        header_fill = PatternFill(start_color="9F2241", end_color="9F2241", fill_type="solid")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        for column in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in column)
            ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 30)
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=plantilla_pacientes_ppl.xlsx'
        return response

    @action(detail=False, methods=['post'], url_path='importar-excel')
    def importar_excel(self, request):
        """
        Importa pacientes/PPL desde archivo Excel.
        
        Columnas esperadas:
        - numero_expediente* (único)
        - nombre*
        - apellido_paterno*
        - apellido_materno
        - curp
        - fecha_nacimiento (YYYY-MM-DD)
        - sexo (M/F)
        - centro_clave* (clave del centro)
        - dormitorio
        - celda
        - tipo_sangre
        - alergias
        - enfermedades_cronicas
        - observaciones_medicas
        - fecha_ingreso (YYYY-MM-DD)
        
        Si el expediente ya existe, se actualizan los datos.
        """
        import re
        from datetime import datetime
        
        archivo = request.FILES.get('file')
        if not archivo:
            return Response({'error': 'No se recibió archivo'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Validaciones de archivo
        MAX_FILE_SIZE = 5 * 1024 * 1024  # 5MB
        MAX_ROWS = 1000
        
        if archivo.size > MAX_FILE_SIZE:
            return Response(
                {'error': f'Archivo demasiado grande. Máximo: {MAX_FILE_SIZE // (1024*1024)}MB'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        file_ext = '.' + archivo.name.split('.')[-1].lower() if '.' in archivo.name else ''
        if file_ext != '.xlsx':
            return Response(
                {'error': 'Solo se aceptan archivos .xlsx'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Validar magic bytes
        archivo.seek(0)
        header = archivo.read(4)
        archivo.seek(0)
        if header[:4] != b'PK\x03\x04':
            return Response(
                {'error': 'El archivo no es un Excel válido'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Patrones de validación
        CURP_PATTERN = re.compile(r'^[A-Z]{4}\d{6}[HM][A-Z]{5}[A-Z0-9]\d$')
        TIPOS_SANGRE = ['O+', 'O-', 'A+', 'A-', 'B+', 'B-', 'AB+', 'AB-']
        
        try:
            wb = openpyxl.load_workbook(archivo)
            ws = wb.active
            
            creados = 0
            actualizados = 0
            errores = []
            user = request.user
            
            # Determinar centro del usuario si no es admin/farmacia
            centro_usuario = None
            if not user.is_superuser and user.rol not in ['admin', 'admin_sistema', 'farmacia']:
                centro_usuario = user.centro
            
            with transaction.atomic():
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=MAX_ROWS + 1, values_only=True), start=2):
                    if row_idx > MAX_ROWS + 1:
                        errores.append(f"Se alcanzó el límite de {MAX_ROWS} registros")
                        break
                    
                    valores = list(row) + [None] * 20
                    
                    # Extraer valores - nuevo orden con CURP después de expediente
                    numero_expediente = str(valores[0] or '').strip().upper()
                    curp = str(valores[1] or '').strip().upper() if valores[1] else None
                    nombre = str(valores[2] or '').strip().title()
                    apellido_paterno = str(valores[3] or '').strip().title()
                    apellido_materno = str(valores[4] or '').strip().title() if valores[4] else None
                    fecha_nacimiento_str = str(valores[5] or '').strip() if valores[5] else None
                    sexo = str(valores[6] or '').strip().upper() if valores[6] else None
                    centro_nombre = str(valores[7] or '').strip() if valores[7] else None
                    dormitorio = str(valores[8] or '').strip() if valores[8] else None
                    celda = str(valores[9] or '').strip() if valores[9] else None
                    tipo_sangre = str(valores[10] or '').strip().upper() if valores[10] else None
                    alergias = str(valores[11] or '').strip() if valores[11] else None
                    enfermedades_cronicas = str(valores[12] or '').strip() if valores[12] else None
                    observaciones_medicas = str(valores[13] or '').strip() if valores[13] else None
                    fecha_ingreso_str = str(valores[14] or '').strip() if valores[14] else None
                    
                    # Saltar filas vacías
                    if not numero_expediente:
                        continue
                    
                    # Validar campos obligatorios (todos son obligatorios ahora)
                    if not curp or len(curp) != 18:
                        errores.append(f"Fila {row_idx}: CURP inválido o faltante (debe tener 18 caracteres)")
                        continue
                    if not nombre:
                        errores.append(f"Fila {row_idx}: Falta nombre")
                        continue
                    if not apellido_paterno:
                        errores.append(f"Fila {row_idx}: Falta apellido paterno")
                        continue
                    if not apellido_materno:
                        errores.append(f"Fila {row_idx}: Falta apellido materno")
                        continue
                    if not fecha_nacimiento_str:
                        errores.append(f"Fila {row_idx}: Falta fecha de nacimiento")
                        continue
                    if not sexo or sexo not in ['M', 'F']:
                        errores.append(f"Fila {row_idx}: Sexo inválido (debe ser M o F)")
                        continue
                    if not dormitorio:
                        errores.append(f"Fila {row_idx}: Falta dormitorio")
                        continue
                    if not celda:
                        errores.append(f"Fila {row_idx}: Falta celda")
                        continue
                    
                    # Validar/obtener centro
                    centro = None
                    if centro_usuario:
                        # Usuario de centro solo puede importar a su centro
                        centro = centro_usuario
                    elif centro_nombre:
                        try:
                            # Buscar centro por nombre (búsqueda exacta insensible a mayúsculas)
                            centro = Centro.objects.get(nombre__iexact=centro_nombre)
                        except Centro.DoesNotExist:
                            # Intentar búsqueda por ID si es numérico
                            try:
                                centro = Centro.objects.get(id=int(centro_nombre))
                            except (Centro.DoesNotExist, ValueError):
                                errores.append(f"Fila {row_idx}: Centro '{centro_nombre}' no encontrado (ver hoja 'Centros Disponibles' en la plantilla)")
                                continue
                    else:
                        errores.append(f"Fila {row_idx}: Falta el nombre del centro")
                        continue
                    
                    # Validar CURP si se proporciona
                    if curp and not CURP_PATTERN.match(curp):
                        errores.append(f"Fila {row_idx}: CURP inválido '{curp}'")
                        curp = None  # Continuar sin CURP
                    
                    # Parsear fecha de nacimiento
                    fecha_nacimiento = None
                    if fecha_nacimiento_str:
                        try:
                            # Intentar varios formatos
                            for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y']:
                                try:
                                    fecha_nacimiento = datetime.strptime(fecha_nacimiento_str, fmt).date()
                                    break
                                except ValueError:
                                    continue
                        except:
                            pass
                    
                    # Parsear fecha de ingreso
                    fecha_ingreso = None
                    if fecha_ingreso_str:
                        try:
                            for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y']:
                                try:
                                    fecha_ingreso = datetime.strptime(fecha_ingreso_str, fmt).date()
                                    break
                                except ValueError:
                                    continue
                        except:
                            pass
                    
                    # Validar sexo
                    if sexo and sexo not in ['M', 'F']:
                        sexo = None
                    
                    # Validar tipo de sangre
                    if tipo_sangre and tipo_sangre not in TIPOS_SANGRE:
                        tipo_sangre = None
                    
                    # Crear o actualizar paciente
                    try:
                        # Verificar si ya existe para determinar created_by
                        existing_paciente = Paciente.objects.filter(numero_expediente=numero_expediente).first()
                        
                        paciente, created = Paciente.objects.update_or_create(
                            numero_expediente=numero_expediente,
                            defaults={
                                'nombre': nombre,
                                'apellido_paterno': apellido_paterno,
                                'apellido_materno': apellido_materno,
                                'curp': curp,
                                'fecha_nacimiento': fecha_nacimiento,
                                'sexo': sexo,
                                'centro': centro,
                                'dormitorio': dormitorio,
                                'celda': celda,
                                'tipo_sangre': tipo_sangre,
                                'alergias': alergias,
                                'enfermedades_cronicas': enfermedades_cronicas,
                                'observaciones_medicas': observaciones_medicas,
                                'fecha_ingreso': fecha_ingreso,
                                'activo': True,
                                'created_by': existing_paciente.created_by if existing_paciente else user,
                            }
                        )
                        
                        if created:
                            creados += 1
                        else:
                            actualizados += 1
                            
                    except Exception as e:
                        errores.append(f"Fila {row_idx}: Error al guardar - {str(e)}")
                        continue
            
            return Response({
                'mensaje': f'Importación completada: {creados} creados, {actualizados} actualizados',
                'creados': creados,
                'actualizados': actualizados,
                'errores': errores[:50],  # Limitar errores mostrados
                'total_errores': len(errores)
            })
            
        except Exception as e:
            logger.error(f"Error en importación de pacientes: {e}", exc_info=True)
            return Response(
                {'error': f'Error al procesar archivo: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class DispensacionViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestión de Dispensaciones de medicamentos.
    
    ISS-DISP: Este módulo es operado por MÉDICO y CENTRO.
    FARMACIA solo puede auditar (ver) pero NO crear/editar/dispensar/cancelar.
    
    Incluye operaciones CRUD, dispensar, cancelar y generación de PDF Formato C.
    """
    queryset = Dispensacion.objects.all()
    serializer_class = DispensacionSerializer
    permission_classes = [IsAuthenticated, CanManageDispensaciones]
    pagination_class = StandardResultsSetPagination
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['folio', 'paciente__numero_expediente', 'paciente__nombre', 'medico_prescriptor']
    ordering_fields = ['folio', 'fecha_dispensacion', 'created_at']
    ordering = ['-fecha_dispensacion']
    
    def get_queryset(self):
        """
        Filtra dispensaciones según rol y centro del usuario.
        
        ISS-SEC-FIX: Aislamiento de datos entre médicos del mismo centro
        - Superusuario/Admin/Farmacia: Ven TODO (modo auditoría global)
        - Director/Admin de Centro: Ven todas las de su centro
        - Médico: SOLO ve las que él mismo creó (aislamiento por usuario)
        - Centro genérico: Ve las de su centro (lectura)
        - Sin centro: No ve nada
        """
        queryset = super().get_queryset().select_related(
            'paciente', 'centro', 'dispensado_por', 'autorizado_por', 'created_by'
        ).prefetch_related('detalles', 'detalles__producto', 'detalles__lote')
        
        user = self.request.user
        
        # SEGURIDAD: Filtrar por centro y rol del usuario
        if user.is_superuser or user.rol in ['admin', 'admin_sistema', 'farmacia']:
            # Admin y farmacia pueden ver todo (modo auditoría global)
            pass
        elif user.centro:
            # ISS-FIX: Todos los usuarios del centro pueden VER todas las dispensaciones
            # de su centro (médicos, directores, etc.) para coordinación clínica.
            # La restricción de EDICIÓN se hace en update/destroy, no en visibilidad.
            queryset = queryset.filter(centro=user.centro)
        else:
            # Sin centro asignado = no ve nada (seguridad)
            return queryset.none()
        
        # Filtros opcionales (solo válidos para admin/farmacia)
        centro_id = self.request.query_params.get('centro')
        if centro_id:
            if user.is_superuser or user.rol in ['admin', 'admin_sistema', 'farmacia']:
                queryset = queryset.filter(centro_id=centro_id)
        
        paciente_id = self.request.query_params.get('paciente')
        if paciente_id:
            queryset = queryset.filter(paciente_id=paciente_id)
        
        estado = self.request.query_params.get('estado')
        if estado:
            queryset = queryset.filter(estado=estado)
        
        tipo = self.request.query_params.get('tipo')
        if tipo:
            queryset = queryset.filter(tipo_dispensacion=tipo)
        
        # Filtro por fecha
        fecha_desde = self.request.query_params.get('fecha_desde')
        fecha_hasta = self.request.query_params.get('fecha_hasta')
        if fecha_desde:
            queryset = queryset.filter(fecha_dispensacion__date__gte=fecha_desde)
        if fecha_hasta:
            queryset = queryset.filter(fecha_dispensacion__date__lte=fecha_hasta)
        
        return queryset
    
    def get_serializer_class(self):
        if self.action == 'list':
            return DispensacionListSerializer
        return DispensacionSerializer
    
    def update(self, request, *args, **kwargs):
        """
        Actualiza una dispensación.
        ISS-FIX: Solo el creador puede editar (o admin/farmacia).
        """
        dispensacion = self.get_object()
        user = request.user
        
        # Solo el creador, admin o farmacia pueden editar
        if not (user.is_superuser or user.rol in ['admin', 'admin_sistema', 'farmacia']):
            if dispensacion.created_by_id != user.id:
                return Response(
                    {'error': 'Solo el usuario que creó esta dispensación puede editarla'},
                    status=status.HTTP_403_FORBIDDEN
                )
        
        return super().update(request, *args, **kwargs)
    
    def create(self, request, *args, **kwargs):
        """
        Crea una dispensación con sus detalles en una sola operación.
        
        Recibe:
        - paciente: ID del paciente
        - centro: ID del centro (opcional, se toma del usuario)
        - tipo_dispensacion: 'normal', 'urgente', 'cronica'
        - medico_prescriptor: Nombre del médico
        - diagnostico: Diagnóstico del paciente
        - indicaciones / indicaciones_medicas: Indicaciones médicas
        - observaciones: Observaciones adicionales
        - detalles: Lista de productos a dispensar
        """
        try:
            with transaction.atomic():
                user = request.user
                data = request.data.copy()
                
                # Normalizar campo indicaciones
                if 'indicaciones_medicas' in data and 'indicaciones' not in data:
                    data['indicaciones'] = data.pop('indicaciones_medicas')
                
                # Extraer detalles para procesarlos después
                detalles_data = data.pop('detalles', [])
                
                # ISS-SEC-FIX: Validación estricta de centro
                # Admin/farmacia pueden especificar centro, otros usuarios DEBEN usar su propio centro
                if user.is_superuser or user.rol in ['admin', 'admin_sistema', 'farmacia']:
                    # Admin/farmacia pueden especificar centro libremente
                    if not data.get('centro') and user.centro:
                        data['centro'] = user.centro.id
                else:
                    # Usuario de centro: FORZAR su centro (ignorar payload)
                    if user.centro:
                        data['centro'] = user.centro.id
                    else:
                        return Response(
                            {'detail': 'Usuario sin centro asignado no puede crear dispensaciones'},
                            status=status.HTTP_403_FORBIDDEN
                        )
                
                # ISS-SEC-FIX: Validar que el paciente pertenezca al centro del usuario
                paciente_id = data.get('paciente')
                if paciente_id:
                    try:
                        paciente = Paciente.objects.get(pk=paciente_id)
                        # Si no es admin/farmacia, verificar que el paciente sea de su centro
                        if not (user.is_superuser or user.rol in ['admin', 'admin_sistema', 'farmacia']):
                            if paciente.centro_id != user.centro_id:
                                return Response(
                                    {'detail': 'No puede crear dispensaciones para pacientes de otro centro'},
                                    status=status.HTTP_403_FORBIDDEN
                                )
                    except Paciente.DoesNotExist:
                        return Response(
                            {'detail': 'Paciente no encontrado'},
                            status=status.HTTP_404_NOT_FOUND
                        )
                
                # Crear la dispensación
                serializer = self.get_serializer(data=data)
                serializer.is_valid(raise_exception=True)
                dispensacion = serializer.save(created_by=user)
                
                # Crear los detalles
                detalles_creados = []
                for detalle_data in detalles_data:
                    # Normalizar campo indicaciones del detalle
                    if 'indicaciones' in detalle_data and 'notas' not in detalle_data:
                        detalle_data['notas'] = detalle_data.pop('indicaciones')
                    
                    detalle = DetalleDispensacion.objects.create(
                        dispensacion=dispensacion,
                        producto_id=detalle_data.get('producto'),
                        lote_id=detalle_data.get('lote'),
                        cantidad_prescrita=detalle_data.get('cantidad_prescrita', 0),
                        dosis=detalle_data.get('dosis'),
                        frecuencia=detalle_data.get('frecuencia'),
                        duracion_tratamiento=detalle_data.get('duracion_tratamiento'),
                        notas=detalle_data.get('notas'),
                    )
                    detalles_creados.append(detalle.id)
                
                logger.info(f"Dispensación {dispensacion.id} creada con {len(detalles_creados)} detalles: {detalles_creados}")
                
                # Registrar en historial (no crítico, no debe fallar la creación)
                try:
                    HistorialDispensacion.objects.create(
                        dispensacion=dispensacion,
                        accion='crear',
                        estado_nuevo='pendiente',
                        usuario=user,
                        detalles={'total_items': len(detalles_data)},
                        ip_address=request.META.get('REMOTE_ADDR', '')
                    )
                except Exception as hist_error:
                    logger.warning(f"No se pudo crear historial de dispensación: {hist_error}")
                
                # Recargar la dispensación con prefetch para incluir los detalles recién creados
                dispensacion = Dispensacion.objects.prefetch_related(
                    'detalles', 'detalles__producto', 'detalles__lote'
                ).get(pk=dispensacion.pk)
                
                # Retornar la dispensación completa con detalles
                result_serializer = self.get_serializer(dispensacion)
                return Response(result_serializer.data, status=status.HTTP_201_CREATED)
                
        except Exception as e:
            logger.error(f"Error al crear dispensación: {e}", exc_info=True)
            return Response(
                {'detail': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )

    def destroy(self, request, *args, **kwargs):
        """
        Elimina una dispensación.
        Solo se pueden eliminar dispensaciones en estado 'pendiente'.
        ISS-FIX: Solo el creador puede eliminar (o admin/farmacia).
        """
        dispensacion = self.get_object()
        user = request.user
        
        # Solo el creador, admin o farmacia pueden eliminar
        if not (user.is_superuser or user.rol in ['admin', 'admin_sistema', 'farmacia']):
            if dispensacion.created_by_id != user.id:
                return Response(
                    {'error': 'Solo el usuario que creó esta dispensación puede eliminarla'},
                    status=status.HTTP_403_FORBIDDEN
                )
        
        if dispensacion.estado == 'dispensada':
            return Response(
                {'error': 'No se puede eliminar una dispensación ya completada. Use la opción de cancelar si es necesario.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if dispensacion.estado == 'cancelada':
            return Response(
                {'error': 'No se puede eliminar una dispensación cancelada.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Solo permitir eliminar dispensaciones pendientes
        if dispensacion.estado != 'pendiente':
            return Response(
                {'error': f'No se puede eliminar una dispensación en estado "{dispensacion.estado}".'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            folio = dispensacion.folio
            paciente_expediente = dispensacion.paciente.numero_expediente if dispensacion.paciente else 'N/A'
            
            # Registrar en historial antes de eliminar (no crítico)
            try:
                HistorialDispensacion.objects.create(
                    dispensacion=dispensacion,
                    accion='eliminar',
                    estado_anterior=dispensacion.estado,
                    estado_nuevo='eliminada',
                    usuario=request.user,
                    detalles={'folio': folio, 'paciente': paciente_expediente},
                    ip_address=request.META.get('REMOTE_ADDR', '')
                )
            except Exception as hist_error:
                logger.warning(f"No se pudo crear historial al eliminar dispensación: {hist_error}")
            
            logger.info(f"Usuario {request.user} eliminó dispensación {folio} (paciente: {paciente_expediente})")
            
            # Eliminar la dispensación (los detalles se eliminarán en cascada)
            dispensacion.delete()
            
            return Response(
                {'message': f'Dispensación {folio} eliminada correctamente'},
                status=status.HTTP_200_OK
            )
            
        except Exception as e:
            logger.error(f"Error al eliminar dispensación: {e}", exc_info=True)
            return Response(
                {'error': f'Error al eliminar la dispensación: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['post'])
    def dispensar(self, request, pk=None):
        """
        Procesa la dispensación de medicamentos.
        Actualiza el estado y descuenta del inventario.
        
        Si se envían detalles específicos, usa esas cantidades.
        Si no se envían detalles, dispensa las cantidades prescritas automáticamente.
        
        Parámetro opcional 'forzar_parcial': Si es True, dispensa lo disponible aunque sea menos de lo prescrito.
        """
        dispensacion = self.get_object()
        
        if dispensacion.estado in ['dispensada', 'cancelada']:
            return Response(
                {'error': f'La dispensación ya está {dispensacion.estado}'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Verificar que tenga detalles
        detalles_existentes = dispensacion.detalles.all()
        if not detalles_existentes.exists():
            return Response(
                {'error': 'La dispensación no tiene medicamentos registrados'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        forzar_parcial = request.data.get('forzar_parcial', False)
        
        try:
            # ISS-SEC FIX (audit6): TODO dentro de transaction.atomic con select_for_update
            # para evitar race conditions en detalles y lotes
            with transaction.atomic():
                # PASO 1: Bloquear dispensación y detalles para evitar concurrencia
                dispensacion = Dispensacion.objects.select_for_update().get(pk=dispensacion.pk)
                
                # Re-verificar estado después del bloqueo
                if dispensacion.estado in ['dispensada', 'cancelada']:
                    return Response(
                        {'error': f'La dispensación ya está {dispensacion.estado}'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                
                # Bloquear todos los detalles con select_for_update
                # ISS-FIX: No usar select_related('lote') con select_for_update()
                # porque lote es nullable (LEFT OUTER JOIN) y PostgreSQL no permite
                # FOR UPDATE en el lado nullable de un outer join.
                # Los lotes se bloquean individualmente más adelante con select_for_update().
                detalles_bloqueados = list(
                    dispensacion.detalles.select_for_update().select_related('producto')
                )
                
                if not detalles_bloqueados:
                    return Response(
                        {'error': 'La dispensación no tiene medicamentos registrados'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                
                errores_stock = []
                detalles_a_procesar = []
                
                for detalle in detalles_bloqueados:
                    if not detalle.lote_id:
                        errores_stock.append(f'{detalle.producto.nombre}: No tiene lote asignado')
                        continue
                    
                    if detalle.cantidad_prescrita <= 0:
                        continue
                    
                    # ISS-SEC FIX (audit6): Calcular REMANENTE = prescrita - ya_dispensada
                    # Esto evita sobre-dispensación en reintentos/parciales
                    cantidad_ya_dispensada = detalle.cantidad_dispensada or 0
                    remanente = detalle.cantidad_prescrita - cantidad_ya_dispensada
                    
                    if remanente <= 0:
                        # Ya se dispensó todo lo prescrito, saltear
                        continue
                    
                    # Obtener y bloquear el lote
                    try:
                        lote = Lote.objects.select_for_update().get(id=detalle.lote_id)
                    except Lote.DoesNotExist:
                        errores_stock.append(f'{detalle.producto.nombre}: El lote ya no existe')
                        continue
                    
                    # ISS-SEC FIX: VALIDACIÓN CRÍTICA de centro del lote
                    # Caso 1: Dispensación para un centro específico
                    if dispensacion.centro_id:
                        # Si el lote NO tiene centro (es de Farmacia Central), BLOQUEAR
                        if not lote.centro_id:
                            errores_stock.append(
                                f'{detalle.producto.nombre} (Lote: {lote.numero_lote}): '
                                f'El lote pertenece a Farmacia Central y no puede usarse para '
                                f'dispensaciones en {dispensacion.centro.nombre}. '
                                f'Seleccione un lote del centro {dispensacion.centro.nombre}.'
                            )
                            continue
                        # Si el lote tiene centro diferente al de la dispensación, BLOQUEAR
                        if lote.centro_id != dispensacion.centro_id:
                            errores_stock.append(
                                f'{detalle.producto.nombre} (Lote: {lote.numero_lote}): '
                                f'El lote no pertenece al centro de esta dispensación. '
                                f'Seleccione un lote del centro {dispensacion.centro.nombre}.'
                            )
                            continue
                    # Caso 2: Dispensación desde Farmacia Central (centro_id NULL)
                    else:
                        if lote.centro_id:
                            errores_stock.append(
                                f'{detalle.producto.nombre} (Lote: {lote.numero_lote}): '
                                f'El lote pertenece a un centro y no puede usarse para '
                                f'dispensaciones de Farmacia Central. '
                                f'Seleccione un lote de Farmacia Central.'
                            )
                            continue
                    
                    stock_disponible = lote.cantidad_actual
                    # ISS-SEC FIX (audit6): Usar remanente, NO cantidad_prescrita completa
                    cantidad_solicitada = remanente
                    
                    if stock_disponible <= 0:
                        errores_stock.append(
                            f'{detalle.producto.nombre} (Lote: {lote.numero_lote}): Sin stock disponible'
                        )
                    elif stock_disponible < cantidad_solicitada:
                        if not forzar_parcial:
                            errores_stock.append(
                                f'{detalle.producto.nombre} (Lote: {lote.numero_lote}): '
                                f'Stock insuficiente. Disponible: {stock_disponible}, Remanente a dispensar: {cantidad_solicitada}'
                            )
                        else:
                            # Si forzar_parcial, agregar con la cantidad disponible
                            detalles_a_procesar.append({
                                'detalle': detalle,
                                'lote': lote,
                                'cantidad': stock_disponible,
                                'parcial': True
                            })
                    else:
                        # Stock suficiente para el remanente
                        detalles_a_procesar.append({
                            'detalle': detalle,
                            'lote': lote,
                            'cantidad': cantidad_solicitada,
                            'parcial': False
                        })
                
                # Si hay errores de stock y no se fuerza parcial, devolver error 400
                if errores_stock and not forzar_parcial:
                    return Response(
                        {
                            'error': 'Stock insuficiente para completar la dispensación',
                            'detalles_error': errores_stock,
                            'sugerencia': 'Puede editar las cantidades o usar la opción de dispensación parcial'
                        },
                        status=status.HTTP_400_BAD_REQUEST
                    )
                
                # Si no hay nada que procesar
                if not detalles_a_procesar:
                    # ISS-SEC FIX (audit6): Verificar si ya está todo dispensado
                    total_ya_dispensado = sum((d.cantidad_dispensada or 0) for d in detalles_bloqueados)
                    total_prescrito = sum(d.cantidad_prescrita for d in detalles_bloqueados)
                    if total_ya_dispensado >= total_prescrito:
                        return Response(
                            {'error': 'Esta dispensación ya fue completada. No hay remanente por dispensar.'},
                            status=status.HTTP_400_BAD_REQUEST
                        )
                    return Response(
                        {'error': 'No hay medicamentos con stock disponible para dispensar'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                
                # PASO 2: Procesar la dispensación (ya dentro de transaction.atomic)
                total_dispensado = 0
                total_prescrito = 0
                
                for item in detalles_a_procesar:
                    detalle = item['detalle']
                    # ISS-SEC FIX (audit6): El lote ya está bloqueado desde PASO 1
                    lote = item['lote']
                    cantidad = item['cantidad']
                    
                    # Verificar de nuevo el stock (verificación redundante por seguridad)
                    if lote.cantidad_actual < cantidad:
                        cantidad = lote.cantidad_actual
                    
                    if cantidad <= 0:
                        continue
                    
                    # Guardar stock antes de descontar para la validación del movimiento
                    stock_antes = lote.cantidad_actual
                    
                    # ISS-SEC-006: Descontar del lote con F() atómico (defensa en profundidad)
                    Lote.objects.filter(pk=lote.pk).update(
                        cantidad_actual=F('cantidad_actual') - cantidad
                    )
                    lote.refresh_from_db()
                    
                    # Crear movimiento de salida
                    # ISS-FIX: Establecer _stock_pre_movimiento para que el validador
                    # del modelo use el stock correcto (antes del descuento)
                    movimiento = Movimiento(
                        tipo='salida',
                        subtipo_salida='dispensacion',
                        producto=detalle.producto,
                        lote=lote,
                        cantidad=cantidad,
                        centro_origen=dispensacion.centro,
                        usuario=request.user,
                        motivo=f'Dispensación {dispensacion.folio} - Paciente: {dispensacion.paciente.numero_expediente}',
                        referencia=dispensacion.folio,
                        numero_expediente=dispensacion.paciente.numero_expediente,
                    )
                    movimiento._stock_pre_movimiento = stock_antes
                    movimiento.save(skip_stock_update=True)
                    
                    # ISS-SEC FIX: Actualizar detalle ACUMULANDO cantidad_dispensada
                    # En dispensaciones parciales/reintentos, no sobrescribir
                    cantidad_previa = detalle.cantidad_dispensada or 0
                    detalle.cantidad_dispensada = cantidad_previa + cantidad
                    # Determinar estado basado en TOTAL acumulado vs prescrito
                    detalle.estado = 'dispensado' if detalle.cantidad_dispensada >= detalle.cantidad_prescrita else 'parcial'
                    detalle.save()
                    
                    total_dispensado += cantidad
                    total_prescrito += detalle.cantidad_prescrita
                
                # ISS-SEC FIX: Calcular estado con el TOTAL HISTÓRICO de todos los detalles
                # No solo los procesados en esta transacción
                total_dispensado_historico = sum(
                    (d.cantidad_dispensada or 0) for d in dispensacion.detalles.all()
                )
                total_prescrito_global = sum(
                    d.cantidad_prescrita for d in dispensacion.detalles.all()
                )
                
                if total_dispensado_historico >= total_prescrito_global:
                    dispensacion.estado = 'dispensada'
                elif total_dispensado_historico > 0:
                    dispensacion.estado = 'parcial'
                
                dispensacion.dispensado_por = request.user
                dispensacion.fecha_dispensacion = timezone.now()
                dispensacion.save()
                
                # Registrar en historial
                try:
                    HistorialDispensacion.objects.create(
                        dispensacion=dispensacion,
                        accion='dispensar',
                        estado_anterior='pendiente',
                        estado_nuevo=dispensacion.estado,
                        usuario=request.user,
                        detalles={
                            'total_dispensado': total_dispensado,
                            'total_prescrito': total_prescrito,
                            'parcial': dispensacion.estado == 'parcial'
                        },
                        ip_address=request.META.get('REMOTE_ADDR', '')
                    )
                except Exception as hist_error:
                    logger.warning(f"No se pudo crear historial al dispensar: {hist_error}")
                
                # ISS-FIX: Invalidar caché del dashboard tras dispensar
                # para que el inventario por centro se actualice inmediatamente
                try:
                    from django.core.cache import cache as dashboard_cache
                    dashboard_cache.delete('dashboard_resumen_global')
                    dashboard_cache.delete('dashboard_graficas_global')
                    if dispensacion.centro_id:
                        dashboard_cache.delete(f'dashboard_resumen_{dispensacion.centro_id}')
                        dashboard_cache.delete(f'dashboard_graficas_{dispensacion.centro_id}')
                except Exception as cache_err:
                    logger.warning(f'Error al invalidar caché del dashboard tras dispensar: {cache_err}')
                
                serializer = self.get_serializer(dispensacion)
                return Response({
                    'dispensacion': serializer.data,
                    'mensaje': f'Dispensación {"completada" if dispensacion.estado == "dispensada" else "parcial"} exitosamente',
                    'total_dispensado': total_dispensado,
                    'total_prescrito': total_prescrito
                })
                
        except Exception as e:
            logger.error(f"Error al dispensar: {e}", exc_info=True)
            return Response(
                {'error': f'Error al procesar la dispensación: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['post'])
    def cancelar(self, request, pk=None):
        """Cancela una dispensación pendiente"""
        dispensacion = self.get_object()
        
        if dispensacion.estado == 'dispensada':
            return Response(
                {'error': 'No se puede cancelar una dispensación ya completada'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if dispensacion.estado == 'cancelada':
            return Response(
                {'error': 'La dispensación ya está cancelada'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        motivo = request.data.get('motivo', '')
        if not motivo:
            return Response(
                {'error': 'Debe proporcionar un motivo de cancelación'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        estado_anterior = dispensacion.estado
        dispensacion.estado = 'cancelada'
        dispensacion.motivo_cancelacion = motivo
        dispensacion.save()
        
        # Registrar en historial (no crítico)
        try:
            HistorialDispensacion.objects.create(
                dispensacion=dispensacion,
                accion='cancelar',
                estado_anterior=estado_anterior,
                estado_nuevo='cancelada',
                usuario=request.user,
                detalles={'motivo': motivo},
                ip_address=request.META.get('REMOTE_ADDR', '')
            )
        except Exception as hist_error:
            logger.warning(f"No se pudo crear historial al cancelar: {hist_error}")
        
        serializer = self.get_serializer(dispensacion)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def exportar_pdf(self, request, pk=None):
        """Genera el PDF Formato C de dispensación"""
        dispensacion = self.get_object()
        
        try:
            from core.utils.pdf_reports import generar_formato_c_dispensacion
            
            logger.info(f"Generando PDF para dispensación {dispensacion.folio}")
            pdf_buffer = generar_formato_c_dispensacion(dispensacion)
            logger.info(f"PDF generado exitosamente para {dispensacion.folio}")
            
            response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="formato_c_{dispensacion.folio}.pdf"'
            return response
            
        except ImportError as e:
            logger.error(f"Error de importación generando PDF: {e}", exc_info=True)
            return Response(
                {'error': f'Error de importación: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        except Exception as e:
            logger.error(f"Error generando PDF Formato C: {e}", exc_info=True)
            return Response(
                {'error': 'Error al generar PDF del Formato C'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['get'])
    def historial(self, request, pk=None):
        """Obtiene el historial de cambios de una dispensación"""
        dispensacion = self.get_object()
        try:
            historial = dispensacion.historial.all().order_by('-created_at')
            serializer = HistorialDispensacionSerializer(historial, many=True)
            return Response(serializer.data)
        except Exception as e:
            logger.warning(f"Error al obtener historial de dispensación {pk}: {e}")
            # Retornar lista vacía en caso de error (tabla no existe o está vacía)
            return Response([])
    
    @action(detail=True, methods=['post'])
    def agregar_detalle(self, request, pk=None):
        """Agrega un item a la dispensación"""
        dispensacion = self.get_object()
        
        if dispensacion.estado in ['dispensada', 'cancelada']:
            return Response(
                {'error': 'No se pueden agregar items a esta dispensación'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        serializer = DetalleDispensacionSerializer(data={
            **request.data,
            'dispensacion': dispensacion.id
        })
        
        if serializer.is_valid():
            serializer.save()
            
            # Registrar en historial (no crítico)
            try:
                HistorialDispensacion.objects.create(
                    dispensacion=dispensacion,
                    accion='agregar_item',
                    usuario=request.user,
                    detalles={'producto_id': request.data.get('producto')},
                    ip_address=request.META.get('REMOTE_ADDR', '')
                )
            except Exception as hist_error:
                logger.warning(f"No se pudo crear historial al agregar detalle: {hist_error}")
            
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get'], url_path='control-mensual-cprs')
    def control_mensual_cprs(self, request):
        """
        Genera reporte PDF "Control mensual de almacén (A)" para el centro CPRS.
        
        Parámetros:
        - mes: Número del mes (1-12). Default: mes actual
        - anio: Año (YYYY). Default: año actual
        - centro: ID del centro (automático si usuario tiene centro asignado)
        
        Permisos:
        - Usuario de centro: solo puede generar reporte de su centro
        - Admin/Farmacia: puede generar de cualquier centro
        """
        from datetime import datetime
        from dateutil.relativedelta import relativedelta
        from django.db.models import Sum, Q
        from django.db.models.functions import Coalesce
        from core.utils.pdf_reports import generar_control_mensual_cprs
        from core.models import Lote, Movimiento
        
        user = request.user
        
        try:
            # Obtener parámetros
            mes = int(request.query_params.get('mes', timezone.now().month))
            anio = int(request.query_params.get('anio', timezone.now().year))
            centro_param = request.query_params.get('centro', None)
            
            # Validar mes
            if mes < 1 or mes > 12:
                return Response({'error': 'Mes debe estar entre 1 y 12'}, status=status.HTTP_400_BAD_REQUEST)
            
            # Determinar centro
            centro = None
            centro_nombre = None
            
            # Si el usuario tiene centro asignado, usarlo por defecto
            if user.centro:
                centro = user.centro
                centro_nombre = centro.nombre
            
            # Si se especifica centro_param y el usuario tiene permisos globales
            if centro_param and centro_param not in ['', 'null', 'undefined']:
                if user.is_superuser or user.rol in ['admin', 'admin_sistema', 'farmacia']:
                    try:
                        centro = Centro.objects.get(pk=int(centro_param))
                        centro_nombre = centro.nombre
                    except Centro.DoesNotExist:
                        return Response({'error': 'Centro no encontrado'}, status=status.HTTP_404_NOT_FOUND)
                elif user.centro and str(user.centro.id) != centro_param:
                    return Response({
                        'error': 'No tiene permiso para generar reporte de otro centro'
                    }, status=status.HTTP_403_FORBIDDEN)
            
            if not centro:
                return Response({
                    'error': 'Debe especificar un centro o tener uno asignado'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Calcular fechas del periodo
            fecha_inicio = datetime(anio, mes, 1)
            if mes == 12:
                fecha_fin = datetime(anio + 1, 1, 1) - relativedelta(days=1)
            else:
                fecha_fin = datetime(anio, mes + 1, 1) - relativedelta(days=1)
            
            # Preparar datos del periodo
            periodo_data = {
                'mes': mes,
                'anio': anio,
                'fecha_inicio': fecha_inicio,
                'fecha_fin': fecha_fin,
                'fecha_elaboracion': timezone.now(),
            }
            
            # Obtener lotes del centro con movimientos en el periodo
            lotes_base = Lote.objects.filter(centro=centro, activo=True)
            lotes_base = lotes_base.select_related('producto').order_by('producto__clave', 'numero_lote')
            
            productos_data = []
            for lote in lotes_base:
                # ISS-FIX: Cálculo correcto de existencia anterior
                # 1. Sumar TODOS los movimientos ANTES del periodo para obtener existencia_anterior
                # 2. Sumar movimientos DEL periodo para entradas/salidas
                # 3. Existencia final = existencia_anterior + entradas - salidas
                
                # Movimientos ANTES del periodo (para calcular existencia anterior)
                movimientos_antes = Movimiento.objects.filter(
                    lote=lote,
                    fecha__lt=fecha_inicio
                )
                
                existencia_anterior = 0
                for mov in movimientos_antes:
                    tipo = (mov.tipo or '').lower()
                    if tipo in ['entrada', 'ajuste_positivo', 'devolucion']:
                        existencia_anterior += abs(mov.cantidad)
                    elif tipo in ['salida', 'ajuste', 'ajuste_negativo', 'merma', 'caducidad', 'transferencia']:
                        existencia_anterior -= abs(mov.cantidad)
                
                # Si no hay movimientos antes, usar cantidad_inicial del lote como base
                # Solo si el lote fue creado antes del periodo
                if not movimientos_antes.exists():
                    # Verificar si el lote existía antes del periodo
                    if lote.created_at and lote.created_at.date() < fecha_inicio.date():
                        existencia_anterior = lote.cantidad_inicial or 0
                    else:
                        existencia_anterior = 0
                
                # Obtener movimientos del periodo actual
                movimientos_periodo = Movimiento.objects.filter(
                    lote=lote,
                    fecha__gte=fecha_inicio,
                    fecha__lte=fecha_fin
                ).select_related('requisicion')
                
                # Calcular entradas y salidas del periodo
                entradas = 0
                salidas = 0
                doc_entrada = ''
                
                for mov in movimientos_periodo:
                    tipo = (mov.tipo or '').lower()
                    if tipo in ['entrada', 'ajuste_positivo', 'devolucion']:
                        entradas += abs(mov.cantidad)
                        # Obtener número de requisición directamente
                        if not doc_entrada and mov.requisicion:
                            req_num = mov.requisicion.numero
                            if req_num.startswith('REQ-'):
                                doc_entrada = req_num[4:]
                            else:
                                doc_entrada = req_num
                        elif not doc_entrada:
                            ref = mov.referencia or mov.motivo or ''
                            import re
                            match = re.search(r'REQ-(\d{8}-\d+)', ref)
                            if match:
                                doc_entrada = match.group(1)
                    elif tipo in ['salida', 'ajuste', 'ajuste_negativo', 'merma', 'caducidad', 'transferencia']:
                        salidas += abs(mov.cantidad)
                
                # Existencia final calculada correctamente
                existencia_final = existencia_anterior + entradas - salidas
                
                # Validar que no sea negativo (indica error en datos)
                if existencia_final < 0:
                    logger.warning(
                        f"Existencia negativa detectada para lote {lote.numero_lote}: "
                        f"anterior={existencia_anterior}, entradas={entradas}, salidas={salidas}, final={existencia_final}"
                    )
                    # Ajustar a 0 mínimo para el reporte (el stock real del lote)
                    existencia_final = max(0, lote.cantidad_actual or 0)
                
                # Solo incluir si hay movimientos o stock
                if existencia_anterior > 0 or entradas > 0 or salidas > 0 or existencia_final > 0:
                    productos_data.append({
                        'producto_clave': lote.producto.clave if lote.producto else 'N/A',
                        'producto_nombre': lote.producto.nombre if lote.producto else 'N/A',
                        'numero_lote': lote.numero_lote or '',
                        'presentacion': lote.producto.presentacion if lote.producto else 'N/A',
                        'fecha_caducidad': lote.fecha_caducidad,
                        'existencia_anterior': max(0, existencia_anterior),  # Nunca mostrar negativos
                        'documento_entrada': doc_entrada,
                        'entradas': entradas,
                        'salidas': salidas,
                        'existencia_final': existencia_final,
                    })
            
            # Si no hay datos, generar reporte vacío con mensaje
            if not productos_data:
                productos_data = [{
                    'producto_clave': '-',
                    'producto_nombre': 'No hay movimientos registrados en este periodo',
                    'presentacion': '-',
                    'fecha_caducidad': None,
                    'existencia_anterior': 0,
                    'documento_entrada': '',
                    'entradas': 0,
                    'salidas': 0,
                    'existencia_final': 0,
                }]
            
            # Generar PDF
            pdf_buffer = generar_control_mensual_cprs(
                periodo_data=periodo_data,
                productos_data=productos_data,
                centro_nombre=centro_nombre
            )
            
            # Nombre del archivo
            meses_nombres = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                           'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
            mes_nombre = meses_nombres[mes]
            filename = f"Control_Mensual_{centro_nombre.replace(' ', '_')}_{mes_nombre}_{anio}.pdf"
            
            response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
            
        except Exception as e:
            logger.error(f"Error generando Control Mensual CPRS: {e}", exc_info=True)
            return Response(
                {'error': f'Error al generar reporte: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=False, methods=['get'], url_path='formato-c-consolidado')
    def formato_c_consolidado(self, request):
        """
        Genera PDF consolidado "Tarjeta para la distribución de Insumos médicos (C)".
        
        Una página por producto con todas las dispensaciones del periodo.
        
        Parámetros:
        - fecha_inicio: YYYY-MM-DD (requerido)
        - fecha_fin: YYYY-MM-DD (requerido)
        - hora_inicio: HH:MM (opcional, default 00:00)
        - hora_fin: HH:MM (opcional, default 23:59)
        - centro: ID del centro (automático si usuario tiene centro asignado)
        """
        from datetime import datetime as dt_class
        from core.utils.pdf_reports import generar_formato_c_consolidado
        
        user = request.user
        
        try:
            # ---- Parámetros de fecha ----
            fecha_inicio_str = request.query_params.get('fecha_inicio')
            fecha_fin_str = request.query_params.get('fecha_fin')
            
            if not fecha_inicio_str or not fecha_fin_str:
                return Response(
                    {'error': 'Los parámetros fecha_inicio y fecha_fin son requeridos (formato YYYY-MM-DD)'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            
            try:
                fecha_inicio = dt_class.strptime(fecha_inicio_str, '%Y-%m-%d').date()
                fecha_fin = dt_class.strptime(fecha_fin_str, '%Y-%m-%d').date()
            except ValueError:
                return Response(
                    {'error': 'Formato de fecha inválido. Use YYYY-MM-DD'},
                    status=status.HTTP_400_BAD_REQUEST,
                )
            
            # Horas opcionales
            hora_inicio_str = request.query_params.get('hora_inicio', '00:00')
            hora_fin_str = request.query_params.get('hora_fin', '23:59')
            
            try:
                hora_inicio = dt_class.strptime(hora_inicio_str, '%H:%M').time()
                hora_fin = dt_class.strptime(hora_fin_str, '%H:%M').time()
            except ValueError:
                hora_inicio = dt_class.strptime('00:00', '%H:%M').time()
                hora_fin = dt_class.strptime('23:59', '%H:%M').time()
            
            datetime_inicio = timezone.make_aware(dt_class.combine(fecha_inicio, hora_inicio))
            datetime_fin = timezone.make_aware(dt_class.combine(fecha_fin, hora_fin))
            
            # ---- Determinar centro ----
            centro = None
            centro_nombre = 'Todos los Centros'
            centro_param = request.query_params.get('centro')
            
            if user.centro:
                centro = user.centro
                centro_nombre = centro.nombre
            
            if centro_param and centro_param not in ('', 'null', 'undefined'):
                if user.is_superuser or user.rol in ('admin', 'admin_sistema', 'farmacia'):
                    try:
                        centro = Centro.objects.get(pk=int(centro_param))
                        centro_nombre = centro.nombre
                    except Centro.DoesNotExist:
                        return Response({'error': 'Centro no encontrado'}, status=status.HTTP_404_NOT_FOUND)
                elif user.centro and str(user.centro.id) != centro_param:
                    return Response(
                        {'error': 'No tiene permiso para generar reporte de otro centro'},
                        status=status.HTTP_403_FORBIDDEN,
                    )
            
            if not centro:
                if not (user.is_superuser or user.rol in ('admin', 'admin_sistema', 'farmacia')):
                    return Response(
                        {'error': 'No se encontró centro asignado al usuario'},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
            
            # ---- Filtrar dispensaciones ----
            dispensaciones = Dispensacion.objects.filter(
                estado='dispensada',
                fecha_dispensacion__gte=datetime_inicio,
                fecha_dispensacion__lte=datetime_fin,
            ).select_related('paciente', 'centro', 'dispensado_por')
            
            if centro:
                dispensaciones = dispensaciones.filter(centro=centro)
            
            dispensaciones = dispensaciones.order_by('fecha_dispensacion')
            
            # ---- Generar PDF ----
            pdf_buffer = generar_formato_c_consolidado(
                dispensaciones_qs=dispensaciones,
                centro_nombre=centro_nombre,
                fecha_inicio=fecha_inicio,
                fecha_fin=fecha_fin,
            )
            
            filename = (
                f"Formato_C_Consolidado_{centro_nombre.replace(' ', '_')}_"
                f"{fecha_inicio.strftime('%Y%m%d')}-{fecha_fin.strftime('%Y%m%d')}.pdf"
            )
            
            response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response
            
        except Exception as e:
            logger.error(f"Error generando Formato C Consolidado: {e}", exc_info=True)
            return Response(
                {'error': f'Error al generar reporte: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class DetalleDispensacionViewSet(viewsets.ModelViewSet):
    """
    ViewSet para detalles de dispensación.
    
    ISS-DISP: Este módulo es operado por MÉDICO y CENTRO.
    FARMACIA solo puede auditar (ver) pero NO crear/editar.
    """
    queryset = DetalleDispensacion.objects.all()
    serializer_class = DetalleDispensacionSerializer
    permission_classes = [IsAuthenticated, CanManageDispensaciones]
    pagination_class = StandardResultsSetPagination
    
    def get_queryset(self):
        queryset = super().get_queryset().select_related('producto', 'lote', 'dispensacion')
        
        dispensacion_id = self.request.query_params.get('dispensacion')
        if dispensacion_id:
            queryset = queryset.filter(dispensacion_id=dispensacion_id)
        
        return queryset


# =====================================================
# MÓDULO: COMPRAS DE CAJA CHICA DEL CENTRO
# =====================================================

class CompraCajaChicaViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestión de Compras de Caja Chica.
    
    FLUJO CON VERIFICACIÓN DE FARMACIA:
    1. Centro crea solicitud con productos (pendiente)
    2. Centro envía a Farmacia para verificar stock (enviada_farmacia)
    3. Farmacia confirma NO disponibilidad (sin_stock_farmacia) o rechaza (rechazada_farmacia)
    4. Centro envía a Admin (enviada_admin)
    5. Admin revisa y autoriza (autorizada_admin)
    6. Admin envía a Director (enviada_director)
    7. Director autoriza (autorizada) - Lista para comprar
    8. Se realiza la compra (comprada)
    9. Se reciben productos (recibida)
    
    PERMISOS:
    - Centro: Crear, editar, enviar a farmacia
    - Farmacia: Verificar disponibilidad, confirmar o rechazar
    - Admin Centro: Autorizar, enviar a director
    - Director Centro: Autorización final
    - Cualquier rol centro: Registrar compra y recepción
    """
    queryset = CompraCajaChica.objects.all()
    serializer_class = CompraCajaChicaSerializer
    permission_classes = [IsAuthenticated, CanManageComprasCajaChica]
    pagination_class = StandardResultsSetPagination
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['folio', 'proveedor_nombre', 'motivo_compra']
    ordering_fields = ['fecha_solicitud', 'fecha_compra', 'total', 'estado']
    ordering = ['-fecha_solicitud']
    
    def get_queryset(self):
        """Filtra compras según centro y rol del usuario"""
        queryset = super().get_queryset().select_related(
            'centro', 'solicitante', 'autorizado_por', 'recibido_por', 'requisicion_origen',
            'administrador_centro', 'director_centro', 'rechazado_por', 'verificado_por_farmacia'
        ).prefetch_related('detalles')
        user = self.request.user
        
        # SEGURIDAD: Filtrar por centro del usuario
        if user.is_superuser or user.rol in ['admin', 'admin_sistema', 'farmacia', 'admin_farmacia']:
            # Admin y farmacia pueden ver todo (para auditoría/verificación)
            pass
        elif user.centro:
            # Usuario de centro: SOLO ve compras de su centro
            queryset = queryset.filter(centro=user.centro)
        else:
            # Sin centro asignado = no ve nada (seguridad)
            return queryset.none()
        
        # Filtros opcionales
        centro_id = self.request.query_params.get('centro')
        if centro_id:
            if user.is_superuser or user.rol in ['admin', 'admin_sistema', 'farmacia', 'admin_farmacia']:
                queryset = queryset.filter(centro_id=centro_id)
        
        estado = self.request.query_params.get('estado')
        if estado:
            queryset = queryset.filter(estado=estado)
        
        fecha_desde = self.request.query_params.get('fecha_desde')
        if fecha_desde:
            queryset = queryset.filter(fecha_solicitud__date__gte=fecha_desde)
        
        fecha_hasta = self.request.query_params.get('fecha_hasta')
        if fecha_hasta:
            queryset = queryset.filter(fecha_solicitud__date__lte=fecha_hasta)
        
        return queryset
    
    def get_serializer_class(self):
        if self.action == 'list':
            return CompraCajaChicaListSerializer
        return CompraCajaChicaSerializer
    
    # ========== FLUJO FARMACIA: VERIFICACIÓN DE DISPONIBILIDAD ==========
    
    @action(detail=True, methods=['post'], url_path='enviar-farmacia')
    def enviar_farmacia(self, request, pk=None):
        """Centro envía solicitud a Farmacia para verificar disponibilidad"""
        compra = self.get_object()
        
        if compra.estado != 'pendiente':
            return Response(
                {'error': 'Solo se pueden enviar compras pendientes a farmacia'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Validar que tenga al menos un producto
        if compra.detalles.count() == 0:
            return Response(
                {'error': 'Debe agregar al menos un producto antes de enviar'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        compra.estado = 'enviada_farmacia'
        compra.fecha_envio_farmacia = timezone.now()
        compra.save()
        
        # Registrar en historial
        HistorialCompraCajaChica.objects.create(
            compra=compra,
            estado_anterior='pendiente',
            estado_nuevo='enviada_farmacia',
            usuario=request.user,
            accion='enviar_farmacia',
            observaciones='Enviada a Farmacia para verificar disponibilidad de productos',
            ip_address=request.META.get('REMOTE_ADDR', '')
        )
        
        return Response(CompraCajaChicaSerializer(compra, context={'request': request}).data)
    
    @action(detail=True, methods=['post'], url_path='confirmar-sin-stock')
    def confirmar_sin_stock(self, request, pk=None):
        """Farmacia confirma que NO tiene stock - libera la compra para el centro"""
        compra = self.get_object()
        
        if compra.estado != 'enviada_farmacia':
            return Response(
                {'error': 'Esta compra no está pendiente de verificación por farmacia'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Validar rol de farmacia
        if not (request.user.is_superuser or request.user.rol in ['farmacia', 'admin_farmacia', 'admin', 'admin_sistema']):
            return Response(
                {'error': 'Solo usuarios de farmacia pueden verificar disponibilidad'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        # Obtener stock verificado si se proporciona
        stock_verificado = request.data.get('stock_verificado', 0)
        respuesta = request.data.get('respuesta', 'Verificado: No hay stock disponible en farmacia central')
        
        compra.estado = 'sin_stock_farmacia'
        compra.fecha_respuesta_farmacia = timezone.now()
        compra.verificado_por_farmacia = request.user
        compra.respuesta_farmacia = respuesta
        compra.stock_farmacia_verificado = stock_verificado
        compra.save()
        
        HistorialCompraCajaChica.objects.create(
            compra=compra,
            estado_anterior='enviada_farmacia',
            estado_nuevo='sin_stock_farmacia',
            usuario=request.user,
            accion='confirmar_sin_stock',
            observaciones=f'{respuesta}. Stock verificado: {stock_verificado}',
            ip_address=request.META.get('REMOTE_ADDR', '')
        )
        
        return Response(CompraCajaChicaSerializer(compra, context={'request': request}).data)
    
    @action(detail=True, methods=['post'], url_path='rechazar-tiene-stock')
    def rechazar_tiene_stock(self, request, pk=None):
        """Farmacia rechaza indicando que SÍ tiene stock disponible"""
        compra = self.get_object()
        
        if compra.estado != 'enviada_farmacia':
            return Response(
                {'error': 'Esta compra no está pendiente de verificación por farmacia'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Validar rol de farmacia
        if not (request.user.is_superuser or request.user.rol in ['farmacia', 'admin_farmacia', 'admin', 'admin_sistema']):
            return Response(
                {'error': 'Solo usuarios de farmacia pueden verificar disponibilidad'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        stock_disponible = request.data.get('stock_disponible', 0)
        respuesta = request.data.get('respuesta', '')
        
        if not respuesta:
            respuesta = f'Hay stock disponible en farmacia central ({stock_disponible} unidades). Se recomienda hacer requisición normal.'
        
        compra.estado = 'rechazada_farmacia'
        compra.fecha_respuesta_farmacia = timezone.now()
        compra.verificado_por_farmacia = request.user
        compra.respuesta_farmacia = respuesta
        compra.stock_farmacia_verificado = stock_disponible
        compra.save()
        
        HistorialCompraCajaChica.objects.create(
            compra=compra,
            estado_anterior='enviada_farmacia',
            estado_nuevo='rechazada_farmacia',
            usuario=request.user,
            accion='rechazar_tiene_stock',
            observaciones=respuesta,
            ip_address=request.META.get('REMOTE_ADDR', '')
        )
        
        return Response(CompraCajaChicaSerializer(compra, context={'request': request}).data)
    
    # ========== FLUJO MULTINIVEL: ACCIONES ==========
    
    @action(detail=True, methods=['post'], url_path='enviar-admin')
    def enviar_admin(self, request, pk=None):
        """Centro envía solicitud al Administrador después de confirmación de farmacia"""
        compra = self.get_object()
        
        if compra.estado != 'sin_stock_farmacia':
            return Response(
                {'error': 'La compra debe ser confirmada por farmacia (sin stock) antes de enviar a admin'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        compra.estado = 'enviada_admin'
        compra.fecha_envio_admin = timezone.now()
        compra.save()
        
        # Registrar en historial
        HistorialCompraCajaChica.objects.create(
            compra=compra,
            estado_anterior='pendiente',
            estado_nuevo='enviada_admin',
            usuario=request.user,
            accion='enviar_admin',
            observaciones='Enviada para autorización del administrador',
            ip_address=request.META.get('REMOTE_ADDR', '')
        )
        
        return Response(CompraCajaChicaSerializer(compra, context={'request': request}).data)
    
    @action(detail=True, methods=['post'], url_path='autorizar-admin')
    def autorizar_admin(self, request, pk=None):
        """Administrador del Centro autoriza la solicitud"""
        compra = self.get_object()
        
        if compra.estado != 'enviada_admin':
            return Response(
                {'error': 'Esta compra no está en estado de autorización por admin'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Validar rol del usuario
        if request.user.rol not in ['administrador_centro', 'admin', 'admin_sistema']:
            return Response(
                {'error': 'No tiene permisos para autorizar como administrador'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        compra.estado = 'autorizada_admin'
        compra.fecha_autorizacion_admin = timezone.now()
        compra.administrador_centro = request.user
        compra.save()
        
        HistorialCompraCajaChica.objects.create(
            compra=compra,
            estado_anterior='enviada_admin',
            estado_nuevo='autorizada_admin',
            usuario=request.user,
            accion='autorizar_admin',
            observaciones=request.data.get('observaciones', 'Autorizada por administrador'),
            ip_address=request.META.get('REMOTE_ADDR', '')
        )
        
        return Response(CompraCajaChicaSerializer(compra, context={'request': request}).data)
    
    @action(detail=True, methods=['post'], url_path='enviar-director')
    def enviar_director(self, request, pk=None):
        """Administrador envía al Director para autorización final"""
        compra = self.get_object()
        
        if compra.estado != 'autorizada_admin':
            return Response(
                {'error': 'La compra debe estar autorizada por admin primero'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        compra.estado = 'enviada_director'
        compra.fecha_envio_director = timezone.now()
        compra.save()
        
        HistorialCompraCajaChica.objects.create(
            compra=compra,
            estado_anterior='autorizada_admin',
            estado_nuevo='enviada_director',
            usuario=request.user,
            accion='enviar_director',
            observaciones='Enviada para autorización del director',
            ip_address=request.META.get('REMOTE_ADDR', '')
        )
        
        return Response(CompraCajaChicaSerializer(compra, context={'request': request}).data)
    
    @action(detail=True, methods=['post'], url_path='autorizar-director')
    def autorizar_director(self, request, pk=None):
        """Director del Centro da autorización final"""
        compra = self.get_object()
        
        if compra.estado != 'enviada_director':
            return Response(
                {'error': 'Esta compra no está en estado de autorización por director'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Validar rol del usuario
        if request.user.rol not in ['director_centro', 'director', 'admin', 'admin_sistema']:
            return Response(
                {'error': 'No tiene permisos para autorizar como director'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        compra.estado = 'autorizada'
        compra.fecha_autorizacion_director = timezone.now()
        compra.director_centro = request.user
        compra.autorizado_por = request.user  # Para compatibilidad
        compra.save()
        
        HistorialCompraCajaChica.objects.create(
            compra=compra,
            estado_anterior='enviada_director',
            estado_nuevo='autorizada',
            usuario=request.user,
            accion='autorizar_director',
            observaciones=request.data.get('observaciones', 'Autorizada por director - Lista para compra'),
            ip_address=request.META.get('REMOTE_ADDR', '')
        )
        
        return Response(CompraCajaChicaSerializer(compra, context={'request': request}).data)
    
    @action(detail=True, methods=['post'])
    def rechazar(self, request, pk=None):
        """Rechaza una solicitud (Admin o Director)"""
        compra = self.get_object()
        
        if compra.estado not in ['enviada_admin', 'enviada_director']:
            return Response(
                {'error': 'Esta compra no puede ser rechazada en su estado actual'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        motivo = request.data.get('motivo')
        if not motivo:
            return Response(
                {'error': 'Debe proporcionar un motivo de rechazo'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        estado_anterior = compra.estado
        compra.estado = 'rechazada'
        compra.motivo_rechazo = motivo
        compra.rechazado_por = request.user
        compra.save()
        
        HistorialCompraCajaChica.objects.create(
            compra=compra,
            estado_anterior=estado_anterior,
            estado_nuevo='rechazada',
            usuario=request.user,
            accion='rechazar',
            observaciones=motivo,
            ip_address=request.META.get('REMOTE_ADDR', '')
        )
        
        return Response(CompraCajaChicaSerializer(compra, context={'request': request}).data)
    
    @action(detail=True, methods=['post'])
    def devolver(self, request, pk=None):
        """Devuelve la solicitud al estado anterior para corrección"""
        compra = self.get_object()
        
        estado_anterior = compra.estado
        
        if compra.estado == 'enviada_farmacia':
            compra.estado = 'pendiente'
        elif compra.estado == 'enviada_admin':
            compra.estado = 'sin_stock_farmacia'
        elif compra.estado == 'enviada_director':
            compra.estado = 'autorizada_admin'
        else:
            return Response(
                {'error': 'Esta compra no puede ser devuelta en su estado actual'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        observaciones = request.data.get('observaciones', 'Devuelta para corrección')
        compra.observaciones = observaciones
        compra.save()
        
        HistorialCompraCajaChica.objects.create(
            compra=compra,
            estado_anterior=estado_anterior,
            estado_nuevo=compra.estado,
            usuario=request.user,
            accion='devolver',
            observaciones=observaciones,
            ip_address=request.META.get('REMOTE_ADDR', '')
        )
        
        return Response(CompraCajaChicaSerializer(compra, context={'request': request}).data)
    
    def destroy(self, request, *args, **kwargs):
        """
        Elimina una compra de caja chica.
        - Compras pendientes: puede eliminar el centro que la creó
        - Compras en cualquier otro estado (excepto cancelada/rechazada): solo admin de farmacia
        - Compras canceladas o rechazadas: no se pueden eliminar
        """
        compra = self.get_object()
        user = request.user
        
        # Estados que no se pueden eliminar
        if compra.estado in ['cancelada', 'rechazada', 'rechazada_farmacia']:
            return Response(
                {'error': 'No se puede eliminar una compra cancelada o rechazada'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Estados confirmados (ya pasaron del pendiente)
        estados_confirmados = [
            'enviada_farmacia', 'sin_stock_farmacia', 'enviada_admin',
            'autorizada_admin', 'enviada_director', 'autorizada',
            'comprada', 'recibida'
        ]
        
        # Si está en estado confirmado, solo admin de farmacia o superuser puede eliminar
        if compra.estado in estados_confirmados:
            es_admin_farmacia = user.is_superuser or user.rol in ['admin', 'admin_sistema', 'farmacia', 'admin_farmacia']
            if not es_admin_farmacia:
                return Response(
                    {'error': 'Solo el administrador de farmacia puede eliminar compras confirmadas'},
                    status=status.HTTP_403_FORBIDDEN
                )
        
        # Si está pendiente, verificar que sea del mismo centro
        if compra.estado == 'pendiente':
            if not user.is_superuser and user.rol not in ['admin', 'admin_sistema', 'farmacia', 'admin_farmacia']:
                if user.centro and user.centro != compra.centro:
                    return Response(
                        {'error': 'Solo puede eliminar compras de su propio centro'},
                        status=status.HTTP_403_FORBIDDEN
                    )
        
        try:
            folio = compra.folio
            centro_nombre = compra.centro.nombre if compra.centro else 'N/A'
            
            # Registrar en historial antes de eliminar
            try:
                HistorialCompraCajaChica.objects.create(
                    compra=compra,
                    estado_anterior=compra.estado,
                    estado_nuevo='eliminada',
                    usuario=user,
                    accion='eliminar',
                    observaciones=f'Compra eliminada por {user.username}',
                    ip_address=request.META.get('REMOTE_ADDR', '')
                )
            except Exception as hist_error:
                logger.warning(f"No se pudo crear historial al eliminar compra: {hist_error}")
            
            logger.info(f"Usuario {user} eliminó compra caja chica {folio} (centro: {centro_nombre})")
            
            # Eliminar la compra (los detalles se eliminarán en cascada)
            compra.delete()
            
            return Response(
                {'message': f'Compra {folio} eliminada correctamente'},
                status=status.HTTP_200_OK
            )
            
        except Exception as e:
            logger.error(f"Error al eliminar compra caja chica: {e}", exc_info=True)
            return Response(
                {'error': f'Error al eliminar la compra: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    # ========== ACCIONES LEGACY (mantener compatibilidad) ==========
    
    @action(detail=True, methods=['post'])
    def autorizar(self, request, pk=None):
        """Autoriza una solicitud de compra de caja chica (legacy - autorización directa)"""
        compra = self.get_object()
        
        if compra.estado != 'pendiente':
            return Response(
                {'error': 'Solo se pueden autorizar compras pendientes'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        compra.estado = 'autorizada'
        compra.autorizado_por = request.user
        compra.save()
        
        # Registrar en historial
        HistorialCompraCajaChica.objects.create(
            compra=compra,
            estado_anterior='pendiente',
            estado_nuevo='autorizada',
            usuario=request.user,
            accion='autorizar',
            observaciones=request.data.get('observaciones', ''),
            ip_address=request.META.get('REMOTE_ADDR', '')
        )
        
        return Response(CompraCajaChicaSerializer(compra).data)
    
    @action(detail=True, methods=['post'])
    def registrar_compra(self, request, pk=None):
        """Registra que la compra fue realizada"""
        compra = self.get_object()
        
        if compra.estado not in ['pendiente', 'autorizada']:
            return Response(
                {'error': 'Esta compra no puede ser registrada'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        fecha_compra = request.data.get('fecha_compra')
        numero_factura = request.data.get('numero_factura')
        
        if not fecha_compra:
            return Response(
                {'error': 'Debe proporcionar la fecha de compra'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        compra.estado = 'comprada'
        compra.fecha_compra = fecha_compra
        compra.numero_factura = numero_factura
        
        # Actualizar datos del proveedor si se envían
        if request.data.get('proveedor_nombre'):
            compra.proveedor_nombre = request.data.get('proveedor_nombre')
        if request.data.get('proveedor_contacto'):
            compra.proveedor_contacto = request.data.get('proveedor_contacto')
        
        compra.save()
        
        # Recalcular total basado en los detalles actualizados
        total_compra = 0
        
        # Actualizar cantidades compradas en detalles
        detalles = request.data.get('detalles', [])
        for detalle_data in detalles:
            try:
                detalle = DetalleCompraCajaChica.objects.get(
                    id=detalle_data.get('id'),
                    compra=compra
                )
                detalle.cantidad_comprada = detalle_data.get('cantidad_comprada', detalle.cantidad_solicitada)
                detalle.precio_unitario = detalle_data.get('precio_unitario', detalle.precio_unitario)
                detalle.numero_lote = detalle_data.get('numero_lote')
                detalle.fecha_caducidad = detalle_data.get('fecha_caducidad')
                detalle.importe = detalle.precio_unitario * detalle.cantidad_comprada
                detalle.save()
                total_compra += detalle.importe
            except DetalleCompraCajaChica.DoesNotExist:
                pass
        
        # Actualizar total de la compra
        if total_compra > 0:
            compra.subtotal = total_compra
            compra.iva = total_compra * 0.16
            compra.total = compra.subtotal + compra.iva
            compra.save(update_fields=['subtotal', 'iva', 'total'])
        
        # Registrar en historial
        HistorialCompraCajaChica.objects.create(
            compra=compra,
            estado_anterior='autorizada',
            estado_nuevo='comprada',
            usuario=request.user,
            accion='registrar_compra',
            observaciones=f'Factura: {numero_factura or "Sin factura"} - Total: ${compra.total}',
            ip_address=request.META.get('REMOTE_ADDR', '')
        )
        
        return Response(CompraCajaChicaSerializer(compra).data)
    
    @action(detail=True, methods=['post'])
    def recibir(self, request, pk=None):
        """
        Recibe los productos de la compra y los ingresa al inventario de caja chica.
        Este inventario es SEPARADO del inventario principal de farmacia.
        """
        compra = self.get_object()
        
        if compra.estado != 'comprada':
            return Response(
                {'error': 'Solo se pueden recibir compras en estado "comprada"'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        with transaction.atomic():
            # Actualizar compra
            compra.estado = 'recibida'
            compra.fecha_recepcion = timezone.now()
            compra.recibido_por = request.user
            compra.save()
            
            # Procesar cada detalle y crear inventario
            detalles_recibidos = request.data.get('detalles', [])
            for detalle_data in detalles_recibidos:
                try:
                    detalle = DetalleCompraCajaChica.objects.get(
                        id=detalle_data.get('id'),
                        compra=compra
                    )
                    cantidad_recibida = detalle_data.get('cantidad_recibida', detalle.cantidad_comprada)
                    detalle.cantidad_recibida = cantidad_recibida
                    detalle.save()
                    
                    # Crear o actualizar inventario de caja chica
                    if cantidad_recibida > 0:
                        inventario, created = InventarioCajaChica.objects.get_or_create(
                            centro=compra.centro,
                            producto=detalle.producto,
                            numero_lote=detalle.numero_lote or f"CC-{compra.folio}",
                            defaults={
                                'descripcion_producto': detalle.descripcion_producto,
                                'fecha_caducidad': detalle.fecha_caducidad,
                                'cantidad_inicial': cantidad_recibida,
                                'cantidad_actual': cantidad_recibida,
                                'compra': compra,
                                'detalle_compra': detalle,
                                'precio_unitario': detalle.precio_unitario,
                            }
                        )
                        
                        if not created:
                            # Actualizar inventario existente
                            cantidad_anterior = inventario.cantidad_actual
                            inventario.cantidad_actual += cantidad_recibida
                            inventario.cantidad_inicial += cantidad_recibida
                            inventario.save()
                        else:
                            cantidad_anterior = 0
                        
                        # Registrar movimiento de entrada
                        MovimientoCajaChica.objects.create(
                            inventario=inventario,
                            tipo='entrada',
                            cantidad=cantidad_recibida,
                            cantidad_anterior=cantidad_anterior,
                            cantidad_nueva=inventario.cantidad_actual,
                            referencia=compra.folio,
                            motivo=f"Recepción de compra {compra.folio}",
                            usuario=request.user
                        )
                        
                except DetalleCompraCajaChica.DoesNotExist:
                    continue
            
            # Registrar en historial
            HistorialCompraCajaChica.objects.create(
                compra=compra,
                estado_anterior='comprada',
                estado_nuevo='recibida',
                usuario=request.user,
                accion='recibir',
                ip_address=request.META.get('REMOTE_ADDR', '')
            )
        
        return Response(CompraCajaChicaSerializer(compra).data)
    
    @action(detail=True, methods=['post'])
    def cancelar(self, request, pk=None):
        """Cancela una solicitud de compra"""
        compra = self.get_object()
        
        if compra.estado in ['recibida', 'cancelada']:
            return Response(
                {'error': 'Esta compra no puede ser cancelada'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        motivo = request.data.get('motivo')
        if not motivo:
            return Response(
                {'error': 'Debe proporcionar un motivo de cancelación'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        estado_anterior = compra.estado
        compra.estado = 'cancelada'
        compra.motivo_cancelacion = motivo
        compra.save()
        
        # Registrar en historial
        HistorialCompraCajaChica.objects.create(
            compra=compra,
            estado_anterior=estado_anterior,
            estado_nuevo='cancelada',
            usuario=request.user,
            accion='cancelar',
            observaciones=motivo,
            ip_address=request.META.get('REMOTE_ADDR', '')
        )
        
        return Response(CompraCajaChicaSerializer(compra).data)
    
    @action(detail=True, methods=['post'], url_path='agregar-detalle')
    def agregar_detalle(self, request, pk=None):
        """Agrega un producto a la compra"""
        compra = self.get_object()
        
        if compra.estado not in ['pendiente', 'autorizada']:
            return Response(
                {'error': 'No se pueden agregar productos a esta compra'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        serializer = DetalleCompraCajaChicaSerializer(data={
            **request.data,
            'compra': compra.id
        })
        
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['get'])
    def resumen(self, request):
        """Resumen de compras de caja chica por centro"""
        queryset = self.get_queryset()
        
        resumen = {
            'total_compras': queryset.count(),
            'pendientes': queryset.filter(estado='pendiente').count(),
            'enviadas_admin': queryset.filter(estado='enviada_admin').count(),
            'autorizadas_admin': queryset.filter(estado='autorizada_admin').count(),
            'enviadas_director': queryset.filter(estado='enviada_director').count(),
            'autorizadas': queryset.filter(estado='autorizada').count(),
            'compradas': queryset.filter(estado='comprada').count(),
            'recibidas': queryset.filter(estado='recibida').count(),
            'canceladas': queryset.filter(estado='cancelada').count(),
            'rechazadas': queryset.filter(estado='rechazada').count(),
            'monto_total': queryset.exclude(estado__in=['cancelada', 'rechazada']).aggregate(
                total=Sum('total')
            )['total'] or 0,
        }
        
        return Response(resumen)


class DetalleCompraCajaChicaViewSet(viewsets.ModelViewSet):
    """ViewSet para detalles de compra de caja chica"""
    queryset = DetalleCompraCajaChica.objects.all()
    serializer_class = DetalleCompraCajaChicaSerializer
    permission_classes = [IsAuthenticated, CanManageComprasCajaChica]
    pagination_class = StandardResultsSetPagination
    
    def get_queryset(self):
        queryset = super().get_queryset().select_related('producto', 'compra', 'compra__centro')
        user = self.request.user
        
        # ISS-SEC-FIX: Filtrar por centro del usuario (IDOR/Tenancy)
        # - Admin/farmacia pueden ver todo (auditoría)
        # - Usuarios de centro solo ven detalles de compras de su centro
        if user.is_superuser or user.rol in ['admin', 'admin_sistema', 'farmacia']:
            pass  # Admin y farmacia ven todo
        elif user.centro:
            # Usuario de centro: SOLO ve detalles de compras de su centro
            queryset = queryset.filter(compra__centro=user.centro)
        else:
            # Sin centro asignado = no ve nada
            return queryset.none()
        
        compra_id = self.request.query_params.get('compra')
        if compra_id:
            # ISS-SEC-FIX: Validar que la compra pertenezca al centro del usuario
            if not (user.is_superuser or user.rol in ['admin', 'admin_sistema', 'farmacia']):
                if user.centro:
                    # Verificar que la compra sea del centro del usuario
                    queryset = queryset.filter(compra_id=compra_id, compra__centro=user.centro)
                else:
                    return queryset.none()
            else:
                queryset = queryset.filter(compra_id=compra_id)
        
        return queryset


class InventarioCajaChicaViewSet(viewsets.ModelViewSet):
    """
    ViewSet para inventario de caja chica del centro.
    
    Este inventario es SEPARADO del inventario principal de farmacia.
    Contiene los productos comprados por el centro con recursos propios.
    
    FARMACIA puede VER todo el inventario para auditoría y prevenir malas prácticas.
    """
    queryset = InventarioCajaChica.objects.all()
    serializer_class = InventarioCajaChicaSerializer
    permission_classes = [IsAuthenticated, CanManageComprasCajaChica]
    pagination_class = StandardResultsSetPagination
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['descripcion_producto', 'numero_lote']
    ordering_fields = ['created_at', 'cantidad_actual', 'fecha_caducidad']
    ordering = ['-created_at']
    
    def get_queryset(self):
        """Filtra inventario según centro del usuario"""
        queryset = super().get_queryset().select_related('centro', 'producto', 'compra')
        user = self.request.user
        
        # SEGURIDAD: Filtrar por centro del usuario
        # - Superusuario y farmacia ven todo (auditoría)
        # - Centro solo ve lo suyo
        # - Sin centro = sin datos
        if user.is_superuser or user.rol in ['admin', 'admin_sistema', 'farmacia']:
            # Admin y farmacia pueden ver todo (para auditoría)
            pass
        elif user.centro:
            # Usuario de centro: SOLO ve inventario de su centro
            queryset = queryset.filter(centro=user.centro)
        else:
            # Sin centro asignado = no ve nada (seguridad)
            return queryset.none()
        
        # Filtros opcionales (solo válidos para admin/farmacia)
        centro_id = self.request.query_params.get('centro')
        if centro_id:
            if user.is_superuser or user.rol in ['admin', 'admin_sistema', 'farmacia']:
                queryset = queryset.filter(centro_id=centro_id)
        
        activo = self.request.query_params.get('activo')
        if activo is not None:
            queryset = queryset.filter(activo=activo.lower() == 'true')
        
        con_stock = self.request.query_params.get('con_stock')
        if con_stock and con_stock.lower() == 'true':
            queryset = queryset.filter(cantidad_actual__gt=0)
        
        return queryset
    
    @action(detail=True, methods=['post'])
    def registrar_salida(self, request, pk=None):
        """Registra una salida del inventario de caja chica"""
        inventario = self.get_object()
        
        cantidad = request.data.get('cantidad')
        if not cantidad or cantidad <= 0:
            return Response(
                {'error': 'Debe proporcionar una cantidad válida'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if cantidad > inventario.cantidad_actual:
            return Response(
                {'error': f'Stock insuficiente. Disponible: {inventario.cantidad_actual}'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        referencia = request.data.get('referencia', '')  # Expediente del paciente, etc.
        motivo = request.data.get('motivo', 'Uso/Dispensación')
        
        with transaction.atomic():
            cantidad_anterior = inventario.cantidad_actual
            inventario.cantidad_actual -= cantidad
            inventario.save()
            
            # Registrar movimiento
            MovimientoCajaChica.objects.create(
                inventario=inventario,
                tipo='salida',
                cantidad=cantidad,
                cantidad_anterior=cantidad_anterior,
                cantidad_nueva=inventario.cantidad_actual,
                referencia=referencia,
                motivo=motivo,
                usuario=request.user
            )
        
        return Response(InventarioCajaChicaSerializer(inventario).data)
    
    @action(detail=True, methods=['post'])
    def ajustar(self, request, pk=None):
        """Realiza un ajuste de inventario"""
        inventario = self.get_object()
        
        nueva_cantidad = request.data.get('cantidad')
        motivo = request.data.get('motivo')
        
        if nueva_cantidad is None or nueva_cantidad < 0:
            return Response(
                {'error': 'Debe proporcionar una cantidad válida'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if not motivo:
            return Response(
                {'error': 'Debe proporcionar un motivo para el ajuste'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        with transaction.atomic():
            cantidad_anterior = inventario.cantidad_actual
            diferencia = nueva_cantidad - cantidad_anterior
            
            inventario.cantidad_actual = nueva_cantidad
            inventario.save()
            
            # Determinar tipo de ajuste
            if diferencia > 0:
                tipo = 'ajuste_positivo'
            elif diferencia < 0:
                tipo = 'ajuste_negativo'
            else:
                return Response(
                    {'error': 'La cantidad nueva es igual a la actual'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Registrar movimiento
            MovimientoCajaChica.objects.create(
                inventario=inventario,
                tipo=tipo,
                cantidad=abs(diferencia),
                cantidad_anterior=cantidad_anterior,
                cantidad_nueva=nueva_cantidad,
                motivo=motivo,
                usuario=request.user
            )
        
        return Response(InventarioCajaChicaSerializer(inventario).data)
    
    @action(detail=False, methods=['get'])
    def resumen(self, request):
        """Resumen del inventario de caja chica"""
        queryset = self.get_queryset()
        
        resumen = {
            'total_items': queryset.count(),
            'items_con_stock': queryset.filter(cantidad_actual__gt=0).count(),
            'items_agotados': queryset.filter(cantidad_actual=0).count(),
            'items_por_caducar': queryset.filter(
                fecha_caducidad__lte=timezone.now().date() + timedelta(days=90),
                fecha_caducidad__gt=timezone.now().date()
            ).count(),
            'items_caducados': queryset.filter(
                fecha_caducidad__lte=timezone.now().date()
            ).count(),
            'valor_total': sum(
                item.precio_unitario * item.cantidad_actual 
                for item in queryset.filter(cantidad_actual__gt=0)
            ),
        }
        
        return Response(resumen)
    
    @action(detail=False, methods=['get'])
    def exportar(self, request):
        """Exporta el inventario de caja chica a Excel para análisis y comparaciones"""
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from io import BytesIO
        
        queryset = self.get_queryset().select_related('centro', 'producto', 'compra')
        
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventario Caja Chica"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="722F37", end_color="722F37", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Encabezados
        headers = [
            "Centro", "Producto", "Descripción", "Lote", "Caducidad",
            "Stock Inicial", "Stock Actual", "Precio Unit.", "Valor Total",
            "Folio Compra", "Fecha Compra", "Estado"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Datos
        for row, item in enumerate(queryset, 2):
            # Determinar estado
            if item.cantidad_actual == 0:
                estado = "Agotado"
            elif item.fecha_caducidad and item.fecha_caducidad <= timezone.now().date():
                estado = "Caducado"
            elif item.fecha_caducidad and item.fecha_caducidad <= timezone.now().date() + timedelta(days=90):
                estado = "Por caducar"
            else:
                estado = "Activo"
            
            valores = [
                item.centro.nombre if item.centro else 'N/A',
                item.producto.nombre if item.producto else item.descripcion_producto or 'N/A',
                item.descripcion_producto or (item.producto.descripcion if item.producto else 'N/A'),
                item.numero_lote or 'N/A',
                item.fecha_caducidad.strftime('%d/%m/%Y') if item.fecha_caducidad else 'N/A',
                item.cantidad_inicial or 0,
                item.cantidad_actual or 0,
                float(item.precio_unitario) if item.precio_unitario else 0,
                float(item.precio_unitario or 0) * (item.cantidad_actual or 0),
                item.compra.folio if item.compra else 'N/A',
                item.compra.fecha_compra.strftime('%d/%m/%Y') if item.compra and item.compra.fecha_compra else 'N/A',
                estado,
            ]
            
            for col, value in enumerate(valores, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = thin_border
                if col in [6, 7]:  # Stock
                    cell.alignment = Alignment(horizontal="center")
                elif col in [8, 9]:  # Precios
                    cell.alignment = Alignment(horizontal="right")
                    cell.number_format = '#,##0.00'
        
        # Ajustar anchos de columna
        anchos = [30, 30, 40, 15, 12, 12, 12, 12, 15, 20, 12, 12]
        for col, ancho in enumerate(anchos, 1):
            ws.column_dimensions[get_column_letter(col)].width = ancho
        
        # Congelar primera fila
        ws.freeze_panes = 'A2'
        
        # Guardar en buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Generar nombre de archivo
        centro_nombre = request.query_params.get('centro', 'todos')
        if centro_nombre != 'todos':
            try:
                centro = Centro.objects.get(pk=centro_nombre)
                centro_nombre = centro.nombre[:20].replace(' ', '_')
            except:
                centro_nombre = 'centro'
        
        fecha = timezone.now().strftime('%Y%m%d')
        filename = f"inventario_caja_chica_{centro_nombre}_{fecha}.xlsx"
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class MovimientoCajaChicaViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet de solo lectura para movimientos de inventario de caja chica"""
    queryset = MovimientoCajaChica.objects.all()
    serializer_class = MovimientoCajaChicaSerializer
    permission_classes = [IsAuthenticated, CanManageComprasCajaChica]
    pagination_class = StandardResultsSetPagination
    ordering = ['-created_at']
    
    def get_queryset(self):
        queryset = super().get_queryset().select_related('inventario', 'inventario__centro', 'usuario')
        user = self.request.user
        
        # ISS-SEC-FIX: Filtrar por centro del usuario (IDOR/Tenancy)
        # - Admin/farmacia pueden ver todo (auditoría)
        # - Usuarios de centro solo ven movimientos de su centro
        if user.is_superuser or user.rol in ['admin', 'admin_sistema', 'farmacia']:
            pass  # Admin y farmacia ven todo
        elif user.centro:
            # Usuario de centro: SOLO ve movimientos de inventario de su centro
            queryset = queryset.filter(inventario__centro=user.centro)
        else:
            # Sin centro asignado = no ve nada
            return queryset.none()
        
        inventario_id = self.request.query_params.get('inventario')
        if inventario_id:
            # ISS-SEC-FIX: Validar que el inventario pertenezca al centro del usuario
            if not (user.is_superuser or user.rol in ['admin', 'admin_sistema', 'farmacia']):
                if user.centro:
                    queryset = queryset.filter(inventario_id=inventario_id, inventario__centro=user.centro)
                else:
                    return queryset.none()
            else:
                queryset = queryset.filter(inventario_id=inventario_id)
        
        tipo = self.request.query_params.get('tipo')
        if tipo:
            queryset = queryset.filter(tipo=tipo)
        
        return queryset

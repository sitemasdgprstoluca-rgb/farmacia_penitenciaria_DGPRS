# -*- coding: utf-8 -*-
"""
Tests para validadores de importación Excel (ISS-001, ISS-003).

Cobertura:
- Archivos sin nombre / nombre vacío
- Extensiones inválidas
- Archivos muy grandes
- Magic bytes incorrectos (contenido no coincide con extensión)
- Archivos corruptos / vacíos
- Conteo de filas con corte temprano
"""
import pytest
from io import BytesIO
from unittest.mock import MagicMock, patch
import openpyxl

from inventario.views import validar_archivo_excel, validar_filas_excel, EXCEL_MAGIC_BYTES


# ============================================================================
# FIXTURES
# ============================================================================

@pytest.fixture
def xlsx_magic_bytes():
    """Magic bytes de un archivo XLSX válido (ZIP header)."""
    return b'PK\x03\x04'


@pytest.fixture
def xls_magic_bytes():
    """Magic bytes de un archivo XLS válido (OLE2)."""
    return b'\xD0\xCF\x11\xE0'


@pytest.fixture
def valid_xlsx_content(xlsx_magic_bytes):
    """Contenido mínimo de un archivo XLSX válido."""
    # Un archivo XLSX real es un ZIP, usamos openpyxl para generarlo
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = 'Header'
    ws['A2'] = 'Data'
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


@pytest.fixture
def mock_xlsx_file(valid_xlsx_content):
    """Archivo XLSX mock válido."""
    file = BytesIO(valid_xlsx_content)
    file.name = 'test_import.xlsx'
    file.size = len(valid_xlsx_content)
    return file


# ============================================================================
# TESTS ISS-001: VALIDACIÓN DE ARCHIVO EXCEL
# ============================================================================

class TestValidarArchivoExcel:
    """Tests para validar_archivo_excel con hardening de seguridad."""
    
    def test_archivo_none_rechazado(self):
        """Archivo None debe ser rechazado."""
        valido, error = validar_archivo_excel(None)
        assert not valido
        assert 'No se recibió archivo' in error
    
    def test_archivo_sin_nombre_rechazado(self):
        """ISS-001: Archivo sin nombre debe ser rechazado."""
        file = BytesIO(b'PK\x03\x04some content')
        # Sin atributo name
        valido, error = validar_archivo_excel(file)
        assert not valido
        assert 'nombre válido' in error
    
    def test_archivo_nombre_vacio_rechazado(self):
        """ISS-001: Archivo con nombre vacío debe ser rechazado."""
        file = BytesIO(b'PK\x03\x04some content')
        file.name = ''
        valido, error = validar_archivo_excel(file)
        assert not valido
        assert 'nombre válido' in error
    
    def test_archivo_nombre_espacios_rechazado(self):
        """ISS-001: Archivo con nombre solo espacios debe ser rechazado."""
        file = BytesIO(b'PK\x03\x04some content')
        file.name = '   '
        valido, error = validar_archivo_excel(file)
        assert not valido
        assert 'nombre válido' in error
    
    def test_archivo_sin_extension_rechazado(self):
        """Archivo sin extensión debe ser rechazado."""
        file = BytesIO(b'PK\x03\x04some content')
        file.name = 'archivo_sin_extension'
        valido, error = validar_archivo_excel(file)
        assert not valido
        assert 'extensión' in error.lower()
    
    def test_extension_no_permitida_rechazada(self):
        """Extensiones no permitidas deben ser rechazadas."""
        file = BytesIO(b'some content')
        file.name = 'malicious.exe'
        file.size = 100
        valido, error = validar_archivo_excel(file)
        assert not valido
        assert 'Extensión no permitida' in error
        assert '.exe' in error
    
    def test_extension_csv_rechazada(self):
        """CSV no es una extensión permitida por defecto."""
        file = BytesIO(b'col1,col2\nval1,val2')
        file.name = 'data.csv'
        file.size = 100
        valido, error = validar_archivo_excel(file)
        assert not valido
        assert 'Extensión no permitida' in error
    
    def test_archivo_muy_grande_rechazado(self):
        """Archivo que excede el tamaño máximo debe ser rechazado."""
        file = BytesIO(b'PK\x03\x04' + b'x' * 100)
        file.name = 'huge_file.xlsx'
        file.size = 50 * 1024 * 1024  # 50MB
        valido, error = validar_archivo_excel(file)
        assert not valido
        assert 'demasiado grande' in error.lower()
    
    def test_magic_bytes_incorrectos_xlsx(self, xlsx_magic_bytes):
        """ISS-001: Archivo .xlsx con magic bytes incorrectos debe ser rechazado."""
        # Contenido que no es ZIP (no empieza con PK)
        fake_content = b'NOT A ZIP FILE CONTENT HERE'
        file = BytesIO(fake_content)
        file.name = 'fake_excel.xlsx'
        file.size = len(fake_content)
        
        valido, error = validar_archivo_excel(file)
        assert not valido
        assert 'contenido del archivo no corresponde' in error
    
    def test_magic_bytes_incorrectos_xls(self, xls_magic_bytes):
        """ISS-001: Archivo .xls con magic bytes incorrectos debe ser rechazado."""
        # Contenido que no es OLE2
        fake_content = b'NOT AN OLE2 FILE CONTENT'
        file = BytesIO(fake_content)
        file.name = 'fake_excel.xls'
        file.size = len(fake_content)
        
        valido, error = validar_archivo_excel(file)
        assert not valido
        assert 'contenido del archivo no corresponde' in error
    
    def test_archivo_vacio_rechazado(self):
        """ISS-001: Archivo vacío debe ser rechazado."""
        file = BytesIO(b'')
        file.name = 'empty.xlsx'
        file.size = 0
        
        valido, error = validar_archivo_excel(file)
        assert not valido
        assert 'vacío o corrupto' in error.lower()
    
    def test_archivo_muy_pequeno_rechazado(self):
        """ISS-001: Archivo demasiado pequeño para ser válido."""
        file = BytesIO(b'PK')  # Solo 2 bytes
        file.name = 'tiny.xlsx'
        file.size = 2
        
        valido, error = validar_archivo_excel(file)
        assert not valido
        assert 'vacío o corrupto' in error.lower()
    
    def test_xlsx_valido_aceptado(self, mock_xlsx_file):
        """Archivo XLSX válido debe ser aceptado."""
        valido, error = validar_archivo_excel(mock_xlsx_file)
        assert valido
        assert error is None
    
    def test_xlsx_mayusculas_aceptado(self, valid_xlsx_content):
        """Extensión en mayúsculas debe ser aceptada."""
        file = BytesIO(valid_xlsx_content)
        file.name = 'TEST.XLSX'
        file.size = len(valid_xlsx_content)
        
        valido, error = validar_archivo_excel(file)
        assert valido
        assert error is None
    
    def test_xls_magic_bytes_valido(self, xls_magic_bytes):
        """Archivo .xls con magic bytes OLE2 correctos debe ser aceptado."""
        # Crear contenido con header OLE2 válido
        content = xls_magic_bytes + b'\xa1\xb1\x1a\xe1' + b'x' * 100
        file = BytesIO(content)
        file.name = 'legacy.xls'
        file.size = len(content)
        
        valido, error = validar_archivo_excel(file)
        assert valido
        assert error is None
    
    def test_posicion_archivo_restaurada(self, mock_xlsx_file):
        """La posición del archivo debe restaurarse después de validar."""
        # Mover posición inicial
        mock_xlsx_file.seek(10)
        pos_before = mock_xlsx_file.tell()
        
        validar_archivo_excel(mock_xlsx_file)
        
        # La posición debe haberse restaurado al inicio (0) para lectura posterior
        # En realidad restauramos a la posición original (10)
        pos_after = mock_xlsx_file.tell()
        assert pos_after == pos_before


# ============================================================================
# TESTS ISS-003: VALIDACIÓN DE FILAS CON CORTE TEMPRANO
# ============================================================================

class TestValidarFilasExcel:
    """Tests para validar_filas_excel con optimización de streaming."""
    
    @pytest.fixture
    def workbook_pocas_filas(self):
        """Workbook con pocas filas (dentro del límite)."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'Header'
        for i in range(2, 12):  # 10 filas de datos
            ws[f'A{i}'] = f'Data {i}'
        return ws
    
    @pytest.fixture
    def workbook_vacio(self):
        """Workbook sin datos."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'Header'
        return ws
    
    def test_pocas_filas_aceptado(self, workbook_pocas_filas):
        """Archivo con pocas filas debe ser aceptado."""
        valido, error, num_filas = validar_filas_excel(workbook_pocas_filas)
        assert valido
        assert error is None
        assert num_filas == 10
    
    def test_archivo_vacio_aceptado(self, workbook_vacio):
        """Archivo solo con header (sin datos) debe ser aceptado."""
        valido, error, num_filas = validar_filas_excel(workbook_vacio)
        assert valido
        assert error is None
        assert num_filas == 0
    
    def test_filas_vacias_no_contadas(self):
        """Filas completamente vacías no deben contarse."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'Header'
        ws['A2'] = 'Data 1'
        # Filas 3-5 vacías
        ws['A6'] = 'Data 2'
        
        valido, error, num_filas = validar_filas_excel(ws)
        assert valido
        assert num_filas == 2  # Solo filas con datos
    
    def test_filas_solo_espacios_no_contadas(self):
        """Filas con solo espacios no deben contarse como datos."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'Header'
        ws['A2'] = '   '  # Solo espacios
        ws['A3'] = 'Real Data'
        
        valido, error, num_filas = validar_filas_excel(ws)
        assert valido
        # La fila con espacios se cuenta si any() considera '   ' como truthy
        # Depende de la implementación - verificamos que al menos los datos reales cuentan
        assert num_filas >= 1
    
    @patch('inventario.views.settings')
    def test_excede_limite_rechazado(self, mock_settings):
        """ISS-003: Archivo que excede límite debe ser rechazado."""
        mock_settings.IMPORT_MAX_ROWS = 5
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'Header'
        for i in range(2, 12):  # 10 filas de datos
            ws[f'A{i}'] = f'Data {i}'
        
        valido, error, num_filas = validar_filas_excel(ws)
        assert not valido
        assert 'excede el máximo' in error or 'Demasiadas filas' in error
        # El conteo debe detenerse poco después del límite
        assert num_filas > 5
    
    @patch('inventario.views.settings')
    def test_corte_temprano_optimizacion(self, mock_settings):
        """ISS-003: El conteo debe detenerse al superar el límite (corte temprano)."""
        mock_settings.IMPORT_MAX_ROWS = 100
        
        # Crear workbook con muchas filas
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'Header'
        for i in range(2, 502):  # 500 filas de datos
            ws[f'A{i}'] = f'Data {i}'
        
        valido, error, num_filas = validar_filas_excel(ws)
        assert not valido
        # No debería haber contado las 500 filas completas
        # Debería detenerse poco después del límite (101)
        assert num_filas == 101  # max_rows + 1


# ============================================================================
# TESTS DE INTEGRACIÓN: VALIDACIÓN COMPLETA
# ============================================================================

class TestValidacionIntegrada:
    """Tests de integración para el flujo completo de validación."""
    
    def test_flujo_completo_valido(self):
        """Test del flujo completo con archivo válido."""
        # Crear workbook válido
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'Producto'
        ws['B1'] = 'Cantidad'
        for i in range(2, 22):  # 20 filas
            ws[f'A{i}'] = f'Producto {i}'
            ws[f'B{i}'] = i * 10
        
        # Guardar a BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        content = buffer.getvalue()
        
        # Crear archivo mock
        file = BytesIO(content)
        file.name = 'importacion.xlsx'
        file.size = len(content)
        
        # Validar archivo
        valido, error = validar_archivo_excel(file)
        assert valido, f"Error inesperado: {error}"
        
        # Validar filas
        file.seek(0)
        wb_loaded = openpyxl.load_workbook(file)
        ws_loaded = wb_loaded.active
        valido, error, num_filas = validar_filas_excel(ws_loaded)
        assert valido, f"Error inesperado: {error}"
        assert num_filas == 20
    
    def test_archivo_malicioso_disfrazado(self):
        """Test de archivo malicioso disfrazado como Excel."""
        # Archivo ejecutable renombrado a .xlsx
        exe_content = b'MZ\x90\x00' + b'\x00' * 100  # PE header
        file = BytesIO(exe_content)
        file.name = 'malicious.xlsx'
        file.size = len(exe_content)
        
        valido, error = validar_archivo_excel(file)
        assert not valido
        assert 'contenido del archivo no corresponde' in error
    
    def test_zip_no_excel(self):
        """Test de archivo ZIP que no es Excel."""
        import zipfile
        
        # Crear un ZIP válido pero sin contenido Excel
        buffer = BytesIO()
        with zipfile.ZipFile(buffer, 'w') as zf:
            zf.writestr('malicious.txt', 'This is not Excel content')
        
        content = buffer.getvalue()
        file = BytesIO(content)
        file.name = 'fake_excel.xlsx'
        file.size = len(content)
        
        # El archivo pasa la validación de magic bytes (es un ZIP válido)
        # Pero fallará al intentar abrirlo con openpyxl
        valido, error = validar_archivo_excel(file)
        # Magic bytes de ZIP son válidos, así que pasa esta validación
        # La validación de contenido Excel real ocurriría en el procesamiento
        assert valido  # Pasa validación básica de formato

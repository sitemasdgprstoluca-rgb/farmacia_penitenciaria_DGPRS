"""
ISS-035: Exportaciones streaming.

Sistema de exportación eficiente para grandes volúmenes de datos
usando streaming para evitar problemas de memoria.
"""
import csv
import io
import json
import logging
from datetime import datetime, date
from typing import Generator, Iterable, Dict, List, Any, Callable, Optional
from decimal import Decimal

from django.http import StreamingHttpResponse, HttpResponse
from django.db.models import QuerySet

logger = logging.getLogger(__name__)


class Echo:
    """Pseudo-buffer que retorna lo que se escribe."""
    def write(self, value):
        return value


class StreamingExporter:
    """
    ISS-035: Exportador streaming para grandes datasets.
    
    Permite exportar millones de registros sin agotar memoria,
    escribiendo directamente al response HTTP.
    """
    
    CHUNK_SIZE = 1000  # Registros por chunk
    
    @classmethod
    def export_csv_streaming(
        cls,
        queryset: QuerySet,
        columns: List[Dict[str, Any]],
        filename: str = "export.csv",
        chunk_size: int = None
    ) -> StreamingHttpResponse:
        """
        ISS-035: Exporta queryset a CSV usando streaming.
        
        Args:
            queryset: QuerySet a exportar (debe tener iterator())
            columns: Lista de columnas [{
                'field': 'campo_modelo',
                'header': 'Encabezado CSV',
                'formatter': callable opcional
            }]
            filename: Nombre del archivo
            chunk_size: Tamaño de cada chunk
            
        Returns:
            StreamingHttpResponse con CSV
        """
        chunk_size = chunk_size or cls.CHUNK_SIZE
        
        def generate_csv():
            buffer = Echo()
            writer = csv.writer(buffer)
            
            # Escribir encabezados
            headers = [col['header'] for col in columns]
            yield writer.writerow(headers)
            
            # Iterar en chunks usando iterator()
            count = 0
            for obj in queryset.iterator(chunk_size=chunk_size):
                row = []
                for col in columns:
                    value = cls._get_value(obj, col['field'])
                    if 'formatter' in col:
                        value = col['formatter'](value)
                    row.append(cls._format_csv_value(value))
                
                yield writer.writerow(row)
                count += 1
                
                if count % 10000 == 0:
                    logger.info(f"ISS-035: Exportados {count} registros...")
            
            logger.info(f"ISS-035: Exportación CSV completada: {count} registros")
        
        response = StreamingHttpResponse(
            generate_csv(),
            content_type='text/csv; charset=utf-8'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response['X-Total-Count'] = str(queryset.count())
        
        return response
    
    @classmethod
    def export_json_streaming(
        cls,
        queryset: QuerySet,
        serializer_func: Callable,
        filename: str = "export.json",
        chunk_size: int = None
    ) -> StreamingHttpResponse:
        """
        ISS-035: Exporta queryset a JSON usando streaming.
        
        Args:
            queryset: QuerySet a exportar
            serializer_func: Función que serializa cada objeto
            filename: Nombre del archivo
            chunk_size: Tamaño de cada chunk
            
        Returns:
            StreamingHttpResponse con JSON
        """
        chunk_size = chunk_size or cls.CHUNK_SIZE
        
        def generate_json():
            yield '{"data": ['
            
            first = True
            count = 0
            
            for obj in queryset.iterator(chunk_size=chunk_size):
                if not first:
                    yield ','
                first = False
                
                data = serializer_func(obj)
                yield json.dumps(data, default=cls._json_serializer)
                count += 1
                
                if count % 10000 == 0:
                    logger.info(f"ISS-035: Exportados {count} registros JSON...")
            
            yield f'], "total": {count}, "exported_at": "{datetime.now().isoformat()}"' + '}'
            logger.info(f"ISS-035: Exportación JSON completada: {count} registros")
        
        response = StreamingHttpResponse(
            generate_json(),
            content_type='application/json; charset=utf-8'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
    
    @classmethod
    def export_excel_streaming(
        cls,
        queryset: QuerySet,
        columns: List[Dict[str, Any]],
        filename: str = "export.xlsx",
        sheet_name: str = "Datos"
    ) -> HttpResponse:
        """
        ISS-035: Exporta a Excel con optimización de memoria.
        
        Nota: openpyxl write_only mode requiere cerrar el workbook,
        por lo que usa buffer en memoria pero optimizado.
        """
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.error("openpyxl no está instalado")
            raise ImportError("Instalar openpyxl para exportación Excel")
        
        # Usar write_only para optimizar memoria
        wb = Workbook(write_only=True)
        ws = wb.create_sheet(sheet_name)
        
        # Encabezados
        headers = [col['header'] for col in columns]
        ws.append(headers)
        
        # Datos
        count = 0
        for obj in queryset.iterator(chunk_size=cls.CHUNK_SIZE):
            row = []
            for col in columns:
                value = cls._get_value(obj, col['field'])
                if 'formatter' in col:
                    value = col['formatter'](value)
                row.append(cls._format_excel_value(value))
            
            ws.append(row)
            count += 1
            
            if count % 10000 == 0:
                logger.info(f"ISS-035: Procesados {count} registros Excel...")
        
        # Guardar a buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        logger.info(f"ISS-035: Exportación Excel completada: {count} registros")
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response['X-Total-Count'] = str(count)
        
        return response
    
    @classmethod
    def _get_value(cls, obj: Any, field_path: str) -> Any:
        """Obtiene valor de un objeto siguiendo path con puntos."""
        value = obj
        for part in field_path.split('.'):
            if value is None:
                return None
            if hasattr(value, part):
                value = getattr(value, part)
                if callable(value):
                    value = value()
            elif isinstance(value, dict):
                value = value.get(part)
            else:
                return None
        return value
    
    @classmethod
    def _format_csv_value(cls, value: Any) -> str:
        """Formatea valor para CSV."""
        if value is None:
            return ""
        if isinstance(value, bool):
            return "Sí" if value else "No"
        if isinstance(value, datetime):
            return value.strftime('%Y-%m-%d %H:%M:%S')
        if isinstance(value, date):
            return value.strftime('%Y-%m-%d')
        if isinstance(value, Decimal):
            return str(value)
        return str(value)
    
    @classmethod
    def _format_excel_value(cls, value: Any) -> Any:
        """Formatea valor para Excel."""
        if value is None:
            return ""
        if isinstance(value, Decimal):
            return float(value)
        if isinstance(value, bool):
            return "Sí" if value else "No"
        return value
    
    @classmethod
    def _json_serializer(cls, obj: Any) -> Any:
        """Serializador JSON personalizado."""
        if isinstance(obj, (datetime, date)):
            return obj.isoformat()
        if isinstance(obj, Decimal):
            return float(obj)
        if hasattr(obj, '__dict__'):
            return str(obj)
        raise TypeError(f"Object of type {type(obj)} is not JSON serializable")


class ReportExporter:
    """
    ISS-035: Exportador específico para reportes del sistema.
    """
    
    @staticmethod
    def export_productos(queryset, formato: str = 'csv') -> HttpResponse:
        """Exporta lista de productos."""
        columns = [
            {'field': 'clave', 'header': 'Clave'},
            {'field': 'descripcion', 'header': 'Descripción'},
            {'field': 'unidad_medida', 'header': 'Unidad'},
            {'field': 'precio_unitario', 'header': 'Precio'},
            {'field': 'stock_minimo', 'header': 'Stock Mínimo'},
            {'field': 'activo', 'header': 'Activo'},
            {'field': 'created_at', 'header': 'Fecha Creación'},
        ]
        
        filename = f"productos_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        if formato == 'csv':
            return StreamingExporter.export_csv_streaming(
                queryset, columns, f"{filename}.csv"
            )
        elif formato == 'excel':
            return StreamingExporter.export_excel_streaming(
                queryset, columns, f"{filename}.xlsx"
            )
        else:
            raise ValueError(f"Formato no soportado: {formato}")
    
    @staticmethod
    def export_lotes(queryset, formato: str = 'csv') -> HttpResponse:
        """Exporta lista de lotes."""
        columns = [
            {'field': 'numero_lote', 'header': 'Número Lote'},
            {'field': 'producto.clave', 'header': 'Producto'},
            {'field': 'producto.descripcion', 'header': 'Descripción'},
            {'field': 'cantidad_inicial', 'header': 'Cantidad Inicial'},
            {'field': 'cantidad_actual', 'header': 'Cantidad Actual'},
            {'field': 'fecha_caducidad', 'header': 'Fecha Caducidad'},
            {'field': 'estado', 'header': 'Estado'},
            {'field': 'centro.nombre', 'header': 'Centro'},
            {'field': 'numero_contrato', 'header': 'Contrato'},
            {'field': 'fecha_entrada', 'header': 'Fecha Entrada'},
        ]
        
        filename = f"lotes_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        if formato == 'csv':
            return StreamingExporter.export_csv_streaming(
                queryset.select_related('producto', 'centro'),
                columns, f"{filename}.csv"
            )
        elif formato == 'excel':
            return StreamingExporter.export_excel_streaming(
                queryset.select_related('producto', 'centro'),
                columns, f"{filename}.xlsx"
            )
        else:
            raise ValueError(f"Formato no soportado: {formato}")
    
    @staticmethod
    def export_movimientos(queryset, formato: str = 'csv') -> HttpResponse:
        """Exporta movimientos de inventario."""
        columns = [
            {'field': 'id', 'header': 'ID'},
            {'field': 'tipo', 'header': 'Tipo'},
            {'field': 'lote.numero_lote', 'header': 'Lote'},
            {'field': 'lote.producto.clave', 'header': 'Producto'},
            {'field': 'cantidad', 'header': 'Cantidad'},
            {'field': 'usuario.username', 'header': 'Usuario'},
            {'field': 'centro.nombre', 'header': 'Centro'},
            {'field': 'fecha', 'header': 'Fecha'},
            {'field': 'observaciones', 'header': 'Observaciones'},
        ]
        
        filename = f"movimientos_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        if formato == 'csv':
            return StreamingExporter.export_csv_streaming(
                queryset.select_related('lote__producto', 'usuario', 'centro'),
                columns, f"{filename}.csv"
            )
        elif formato == 'excel':
            return StreamingExporter.export_excel_streaming(
                queryset.select_related('lote__producto', 'usuario', 'centro'),
                columns, f"{filename}.xlsx"
            )
        else:
            raise ValueError(f"Formato no soportado: {formato}")
    
    @staticmethod
    def export_requisiciones(queryset, formato: str = 'csv') -> HttpResponse:
        """Exporta requisiciones."""
        columns = [
            {'field': 'folio', 'header': 'Folio'},
            {'field': 'centro.nombre', 'header': 'Centro'},
            {'field': 'estado', 'header': 'Estado'},
            {'field': 'usuario_solicita.username', 'header': 'Solicitante'},
            {'field': 'fecha_solicitud', 'header': 'Fecha Solicitud'},
            {'field': 'fecha_autorizacion', 'header': 'Fecha Autorización'},
            {'field': 'fecha_surtido', 'header': 'Fecha Surtido'},
            {'field': 'observaciones', 'header': 'Observaciones'},
        ]
        
        filename = f"requisiciones_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        if formato == 'csv':
            return StreamingExporter.export_csv_streaming(
                queryset.select_related('centro', 'usuario_solicita'),
                columns, f"{filename}.csv"
            )
        elif formato == 'excel':
            return StreamingExporter.export_excel_streaming(
                queryset.select_related('centro', 'usuario_solicita'),
                columns, f"{filename}.xlsx"
            )
        else:
            raise ValueError(f"Formato no soportado: {formato}")


def stream_large_export(
    queryset: QuerySet,
    columns: List[Dict],
    formato: str = 'csv',
    filename_prefix: str = 'export'
) -> HttpResponse:
    """
    ISS-035: Helper para exportación rápida.
    
    Uso:
        return stream_large_export(
            Producto.objects.all(),
            [{'field': 'clave', 'header': 'Clave'}, ...],
            formato='csv'
        )
    """
    filename = f"{filename_prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    
    if formato == 'csv':
        return StreamingExporter.export_csv_streaming(
            queryset, columns, f"{filename}.csv"
        )
    elif formato in ('excel', 'xlsx'):
        return StreamingExporter.export_excel_streaming(
            queryset, columns, f"{filename}.xlsx"
        )
    elif formato == 'json':
        def serializer(obj):
            return {col['field']: StreamingExporter._get_value(obj, col['field']) 
                    for col in columns}
        return StreamingExporter.export_json_streaming(
            queryset, serializer, f"{filename}.json"
        )
    else:
        raise ValueError(f"Formato no soportado: {formato}")

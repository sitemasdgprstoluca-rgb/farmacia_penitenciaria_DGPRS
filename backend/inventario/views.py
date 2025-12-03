from rest_framework import viewsets, status, serializers, permissions, mixins
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.pagination import PageNumberPagination
from django.core.paginator import InvalidPage
from django.core.cache import cache
from django.http import HttpResponse
from django.db import transaction
from django.db.models import Q, Sum, Count, F, IntegerField
from django.db.models.functions import Coalesce
from django.utils import timezone
from django.conf import settings
from django.contrib.auth.models import Group
from datetime import datetime, timedelta, date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import random
import os
import logging
from io import BytesIO

logger = logging.getLogger(__name__)


# -----------------------------------------------------------
# VALIDADORES DE IMPORTACIÓN EXCEL
# -----------------------------------------------------------

# Magic bytes para formatos Excel válidos
# XLSX/XLSM son archivos ZIP (PK\x03\x04)
# XLS antiguo usa formato OLE2 (Compound File Binary Format)
EXCEL_MAGIC_BYTES = {
    '.xlsx': b'PK\x03\x04',  # ZIP archive (Office Open XML)
    '.xlsm': b'PK\x03\x04',  # ZIP archive with macros
    '.xls': b'\xD0\xCF\x11\xE0',  # OLE2 Compound Document
}


def leer_archivo_con_limite(file, max_bytes):
    """
    ISS-001: Lee un archivo con límite estricto de bytes.
    
    Previene DoS por archivos enormes que no declaran tamaño correcto.
    Lee en chunks y aborta si se excede el límite.
    
    Args:
        file: Archivo a leer (UploadedFile o similar)
        max_bytes: Máximo de bytes permitidos
    
    Returns:
        tuple: (BytesIO con contenido, bytes_leidos) o (None, error_message)
    """
    CHUNK_SIZE = 64 * 1024  # 64KB chunks
    buffer = BytesIO()
    bytes_leidos = 0
    
    try:
        # Asegurar que estamos al inicio del archivo
        if hasattr(file, 'seek'):
            file.seek(0)
        
        while True:
            chunk = file.read(CHUNK_SIZE)
            if not chunk:
                break
            
            bytes_leidos += len(chunk)
            
            # ISS-001: Corte estricto si excede el límite
            if bytes_leidos > max_bytes:
                logger.warning(
                    f"Archivo rechazado: tamaño real ({bytes_leidos} bytes) "
                    f"excede límite ({max_bytes} bytes)"
                )
                return None, f'El archivo excede el tamaño máximo permitido ({max_bytes / 1024 / 1024:.1f}MB)'
            
            buffer.write(chunk)
        
        buffer.seek(0)
        return buffer, bytes_leidos
        
    except Exception as e:
        logger.error(f"Error leyendo archivo: {e}")
        return None, f'Error al leer el archivo: {str(e)}'


def validar_archivo_excel(file):
    """
    Valida archivo Excel antes de procesarlo con múltiples capas de seguridad.
    
    Validaciones:
    1. Archivo presente y con nombre válido
    2. Extensión permitida (.xlsx, .xls)
    3. Tamaño dentro de límites
    4. Magic bytes correctos (contenido real coincide con extensión)
    
    Retorna: (es_valido, mensaje_error)
    """
    # 1. Validar que hay archivo
    if not file:
        return False, 'No se recibió archivo'
    
    # 2. ISS-001: Validar nombre obligatorio - rechazar archivos sin nombre
    nombre = getattr(file, 'name', None)
    if not nombre or not nombre.strip():
        return False, 'El archivo debe tener un nombre válido'
    
    nombre_lower = nombre.lower().strip()
    ext = os.path.splitext(nombre_lower)[1]
    
    # 3. Validar extensión
    if not ext:
        return False, 'El archivo debe tener una extensión (.xlsx o .xls)'
    
    extensiones_permitidas = getattr(settings, 'IMPORT_ALLOWED_EXTENSIONS', ['.xlsx', '.xls'])
    if ext not in extensiones_permitidas:
        return False, f'Extensión no permitida: {ext}. Use: {", ".join(extensiones_permitidas)}'
    
    # 4. Validar tamaño declarado ANTES de leer contenido (primera línea de defensa)
    max_size_mb = getattr(settings, 'IMPORT_MAX_FILE_SIZE_MB', 10)
    max_size_bytes = max_size_mb * 1024 * 1024
    
    file_size = getattr(file, 'size', None)
    if file_size is not None and file_size > max_size_bytes:
        return False, f'Archivo demasiado grande: {file_size / 1024 / 1024:.1f}MB. Máximo: {max_size_mb}MB'
    
    # 5. ISS-001: Validar magic bytes - contenido real del archivo
    try:
        # Guardar posición actual y leer primeros bytes
        pos = file.tell() if hasattr(file, 'tell') else 0
        header = file.read(8)  # Leer suficientes bytes para identificar formato
        
        # Restaurar posición para que el archivo pueda ser procesado después
        if hasattr(file, 'seek'):
            file.seek(pos)
        
        if not header or len(header) < 4:
            return False, 'Archivo vacío o corrupto'
        
        # Verificar magic bytes según extensión
        expected_magic = EXCEL_MAGIC_BYTES.get(ext)
        if expected_magic:
            if not header.startswith(expected_magic):
                # El contenido no coincide con la extensión declarada
                logger.warning(
                    f"Archivo rechazado: extensión {ext} pero magic bytes incorrectos. "
                    f"Header: {header[:8].hex()}"
                )
                return False, f'El contenido del archivo no corresponde a un archivo {ext} válido'
    except Exception as e:
        logger.error(f"Error validando magic bytes: {e}")
        return False, 'Error al validar el contenido del archivo'
    
    return True, None


def cargar_workbook_seguro(file):
    """
    ISS-001: Carga un workbook Excel con límites de seguridad estrictos.
    
    1. Lee el archivo con límite real de bytes (no confía en file.size)
    2. Usa read_only=True para streaming y menor uso de memoria
    3. Usa data_only=True para ignorar fórmulas (previene ataques por fórmulas)
    
    Returns:
        tuple: (workbook, error_message) - workbook es None si hay error
    """
    max_size_mb = getattr(settings, 'IMPORT_MAX_FILE_SIZE_MB', 10)
    max_size_bytes = max_size_mb * 1024 * 1024
    
    # Leer archivo con límite estricto de bytes
    buffer, resultado = leer_archivo_con_limite(file, max_size_bytes)
    
    if buffer is None:
        # resultado contiene el mensaje de error
        return None, resultado
    
    try:
        # ISS-001: Usar read_only=True para streaming y data_only=True para ignorar fórmulas
        # Esto reduce significativamente el uso de memoria en archivos grandes
        wb = openpyxl.load_workbook(
            buffer, 
            read_only=True,  # Streaming mode - no carga todo en memoria
            data_only=True   # Ignora fórmulas - solo valores (previene ataques)
        )
        return wb, None
    except Exception as e:
        logger.error(f"Error cargando workbook: {e}")
        return None, f'Error al procesar el archivo Excel: {str(e)}'


def validar_filas_excel(ws):
    """
    Valida número de filas en worksheet con streaming y corte temprano.
    
    ISS-003: Optimizado para no recorrer archivos enormes innecesariamente.
    Se detiene al superar el límite máximo permitido.
    
    Retorna: (es_valido, mensaje_error, num_filas)
    """
    max_rows = getattr(settings, 'IMPORT_MAX_ROWS', 5000)
    # Margen adicional para detener el conteo (evitar contar millones de filas)
    cutoff = max_rows + 1
    
    num_filas = 0
    # ISS-003: Streaming con corte temprano - detenerse apenas superamos el límite
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Solo contar filas que tienen al menos un valor no vacío
        if any(cell is not None and str(cell).strip() for cell in row):
            num_filas += 1
            if num_filas > max_rows:
                # Corte temprano: ya sabemos que excede el límite
                return False, f'El archivo excede el máximo de {max_rows} filas permitidas', num_filas
    
    return True, None, num_filas
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

from core.models import Producto, Lote, Movimiento, Centro, Requisicion, DetalleRequisicion, HojaRecoleccion
from core.serializers import (
    ProductoSerializer, LoteSerializer, MovimientoSerializer, 
    CentroSerializer, RequisicionSerializer, DetalleRequisicionSerializer,
    HojaRecoleccionSerializer
)

from django.contrib.auth import get_user_model
from core.permissions import (
    IsAdminRole, IsFarmaciaRole, IsCentroRole, IsVistaRole,
    IsFarmaciaAdminOrReadOnly, CanAuthorizeRequisicion
)
from core.constants import (
    ESTADOS_REQUISICION,
    PAGINATION_DEFAULT_PAGE_SIZE,
    PAGINATION_MAX_PAGE_SIZE,
    UNIDADES_MEDIDA,
    REQUISICION_GRUPOS_ESTADO,
)

User = get_user_model()


# ============================================================================
# HELPERS DE SEGURIDAD - Filtrado por centro para roles no-admin
# ============================================================================

def is_farmacia_or_admin(user):
    """
    Verifica si el usuario es farmacia, admin o vista (puede ver datos globales).
    
    Roles privilegiados con acceso global:
    - admin: admin_sistema, superusuario, administrador
    - farmacia: farmacia, admin_farmacia, farmaceutico, usuario_farmacia
    - vista: vista, usuario_vista (solo lectura global)
    """
    if not user or not user.is_authenticated:
        return False
    if user.is_superuser:
        return True
    
    rol = (getattr(user, 'rol', '') or '').lower()
    
    # Roles con acceso global (pueden ver todo)
    ROLES_GLOBALES = {
        # Admin
        'admin', 'admin_sistema', 'superusuario', 'administrador',
        # Farmacia
        'farmacia', 'admin_farmacia', 'farmaceutico', 'usuario_farmacia',
        # Vista (solo lectura pero global)
        'vista', 'usuario_vista',
    }
    
    if rol in ROLES_GLOBALES:
        return True
    
    # Verificar grupos (por si el rol no est� en campo directo)
    group_names = {g.name.upper() for g in user.groups.all()}
    GRUPOS_GLOBALES = {'FARMACIA_ADMIN', 'FARMACEUTICO', 'VISTA_USER'}
    
    return bool(group_names & GRUPOS_GLOBALES)


def get_user_centro(user):
    """Obtiene el centro asignado al usuario."""
    return getattr(user, 'centro', None) or getattr(getattr(user, 'profile', None), 'centro', None)


def registrar_movimiento_stock(*, lote, tipo, cantidad, usuario=None, centro=None, requisicion=None, observaciones='', skip_centro_check=False):
    """
    Helper central para registrar un movimiento y actualizar cantidad_actual del lote.
    
    ISS-002: Incluye validación de pertenencia de centro para prevenir manipulación
    de inventario entre centros.
    
    Args:
        lote: Instancia del Lote a modificar
        tipo: 'entrada', 'salida' o 'ajuste'
        cantidad: Cantidad a mover (positivo)
        usuario: Usuario que realiza la operación (opcional)
        centro: Centro donde se registra el movimiento (opcional)
        requisicion: Requisición asociada (opcional)
        observaciones: Texto descriptivo (opcional)
        skip_centro_check: Si True, omite validación de pertenencia (solo para operaciones
                          de sistema/admin que ya validaron permisos)
    
    Raises:
        ValidationError: Si los datos son inválidos o hay problemas de autorización
    
    Returns:
        tuple: (movimiento_creado, lote_actualizado)
    """
    tipo_normalizado = (tipo or '').lower()
    if tipo_normalizado not in ('entrada', 'salida', 'ajuste'):
        raise serializers.ValidationError({'tipo': 'Tipo de movimiento no valido'})
    if cantidad is None:
        raise serializers.ValidationError({'cantidad': 'Cantidad requerida'})
    try:
        cantidad_int = int(cantidad)
    except (TypeError, ValueError):
        raise serializers.ValidationError({'cantidad': 'La cantidad debe ser un numero entero'})

    # =========================================================================
    # ISS-002: Validación de pertenencia de centro
    # =========================================================================
    # Si se pasa un centro explícito, verificar que el lote pertenezca a ese centro
    # Esto previene que usuarios de un centro manipulen stock de otros centros
    if not skip_centro_check and centro is not None:
        lote_centro = getattr(lote, 'centro', None)
        if lote_centro is not None and lote_centro.pk != centro.pk:
            logger.warning(
                f"Intento de operación de stock rechazado: "
                f"usuario={getattr(usuario, 'username', 'N/A')}, "
                f"lote_centro={lote_centro.pk}, centro_solicitado={centro.pk}"
            )
            raise serializers.ValidationError({
                'centro': 'El lote no pertenece al centro especificado. Operación no autorizada.'
            })
    
    # Si el usuario no es admin/farmacia, verificar que tenga acceso al centro del lote
    if not skip_centro_check and usuario is not None and hasattr(usuario, 'is_authenticated') and usuario.is_authenticated:
        if not is_farmacia_or_admin(usuario):
            user_centro = get_user_centro(usuario)
            lote_centro = getattr(lote, 'centro', None)
            if lote_centro is not None and user_centro is not None:
                if user_centro.pk != lote_centro.pk:
                    logger.warning(
                        f"Acceso no autorizado a lote de otro centro: "
                        f"usuario={usuario.username}, user_centro={user_centro.pk}, "
                        f"lote_centro={lote_centro.pk}"
                    )
                    raise serializers.ValidationError({
                        'lote': 'No tiene permiso para operar sobre lotes de otros centros.'
                    })

    delta = cantidad_int
    if tipo_normalizado == 'salida' and delta > 0:
        delta = -delta
    if tipo_normalizado == 'entrada' and delta < 0:
        delta = abs(delta)

    with transaction.atomic():
        lote_ref = Lote.objects.select_for_update().get(pk=lote.pk)
        stock_disponible = lote_ref.cantidad_actual

        if delta < 0 and abs(delta) > stock_disponible:
            raise serializers.ValidationError({
                'cantidad': f'Stock insuficiente en el lote (disponible {stock_disponible}).'
            })

        nuevo_stock = stock_disponible + delta
        if nuevo_stock < 0:
            raise serializers.ValidationError({'cantidad': 'La operacion dejaria el lote con stock negativo'})

        # Actualizar stock y estado de disponibilidad
        lote_ref.cantidad_actual = nuevo_stock
        
        # Para entradas, tambi�n actualizar cantidad_inicial si es necesario
        # Esto permite recibir stock adicional en lotes existentes
        update_fields = ['cantidad_actual', 'estado', 'updated_at']
        if tipo_normalizado == 'entrada' and nuevo_stock > lote_ref.cantidad_inicial:
            lote_ref.cantidad_inicial = nuevo_stock
            update_fields.append('cantidad_inicial')
        
        if nuevo_stock == 0:
            lote_ref.estado = 'agotado'
        elif lote_ref.estado == 'agotado':
            lote_ref.estado = 'disponible'
        lote_ref.save(update_fields=update_fields)

        movimiento = Movimiento(
            tipo=tipo_normalizado,
            lote=lote_ref,
            centro=centro,
            requisicion=requisicion,
            usuario=usuario if usuario and getattr(usuario, 'is_authenticated', False) else None,
            cantidad=delta,
            observaciones=observaciones or ''
        )
        # Guardar stock previo para evitar fallos de validaci?n al crear el movimiento
        movimiento._stock_pre_movimiento = stock_disponible
        movimiento.save()

    return movimiento, lote_ref


class CustomPagination(PageNumberPagination):
    """Paginacin unificada para todos los listados del API."""
    page_size = PAGINATION_DEFAULT_PAGE_SIZE
    page_size_query_param = 'page_size'
    max_page_size = PAGINATION_MAX_PAGE_SIZE

    def paginate_queryset(self, queryset, request, view=None):
        try:
            return super().paginate_queryset(queryset, request, view=view)
        except InvalidPage:
            # Si la pgina solicitada no existe, devolver la ltima disponible sin 404
            self.page = self.paginator.page(self.paginator.num_pages or 1)
            self.request = request
            return list(self.page)

# ISS-002: UserSerializer eliminado - usar el de core.serializers
# from core.serializers import UserSerializer (si se necesita)

# NOTA: UserViewSet está en core/views.py - importar desde allí

class ProductoViewSet(viewsets.ModelViewSet):
    queryset = Producto.objects.select_related('created_by').all()
    serializer_class = ProductoSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = CustomPagination

    def get_permissions(self):
        """
        Permisos personalizados por accion:
        - list, retrieve: IsAuthenticated
        - create, update, destroy, toggle_activo, importar_excel, exportar_excel, auditoria: IsFarmaciaRole
        """
        acciones_farmacia = [
            'create', 'update', 'partial_update', 'destroy', 
            'toggle_activo', 'importar_excel', 'exportar_excel', 'auditoria'
        ]
        if self.action in acciones_farmacia:
            from core.permissions import IsFarmaciaRole
            return [IsAuthenticated(), IsFarmaciaRole()]
        return [IsAuthenticated()]

    def get_queryset(self):
        queryset = Producto.objects.select_related('created_by').all()
        
        activo = self.request.query_params.get('activo')
        if activo == 'true':
            queryset = queryset.filter(activo=True)
        elif activo == 'false':
            queryset = queryset.filter(activo=False)
        
        unidad = self.request.query_params.get('unidad_medida')
        if unidad and unidad != '':
            queryset = queryset.filter(unidad_medida=unidad)
        
        search = self.request.query_params.get('search')
        if search and search.strip():
            queryset = queryset.filter(
                Q(clave__icontains=search) | 
                Q(descripcion__icontains=search)
            )

        stock_status = self.request.query_params.get('stock_status')
        if stock_status:
            queryset = queryset.annotate(
                stock_total_calc=Coalesce(
                    Sum(
                        'lotes__cantidad_actual',
                        filter=Q(
                            lotes__deleted_at__isnull=True,
                            lotes__estado='disponible',
                            lotes__cantidad_actual__gt=0
                        )
                    ),
                    0
                )
            )
            status_val = stock_status.lower()
            if status_val == 'sin_stock':
                queryset = queryset.filter(stock_total_calc__lte=0)
            elif status_val == 'critico':
                queryset = queryset.filter(
                    stock_total_calc__gt=0,
                    stock_minimo__gt=0,
                    stock_total_calc__lt=F('stock_minimo') * 0.5
                )
            elif status_val == 'bajo':
                queryset = queryset.filter(
                    Q(
                        stock_minimo__gt=0,
                        stock_total_calc__gte=F('stock_minimo') * 0.5,
                        stock_total_calc__lt=F('stock_minimo')
                    ) | Q(
                        stock_minimo__lte=0,
                        stock_total_calc__gt=0,
                        stock_total_calc__lt=25
                    )
                )
            elif status_val == 'normal':
                queryset = queryset.filter(
                    Q(
                        stock_minimo__gt=0,
                        stock_total_calc__gte=F('stock_minimo'),
                        stock_total_calc__lte=F('stock_minimo') * 2
                    ) | Q(
                        stock_minimo__lte=0,
                        stock_total_calc__gte=25,
                        stock_total_calc__lt=100
                    )
                )
            elif status_val == 'alto':
                queryset = queryset.filter(
                    Q(
                        stock_minimo__gt=0,
                        stock_total_calc__gt=F('stock_minimo') * 2
                    ) | Q(
                        stock_minimo__lte=0,
                        stock_total_calc__gte=100
                    )
                )
        
        return queryset.order_by('-created_at')
    
    def create(self, request, *args, **kwargs):
        """
        Crea un nuevo producto.
        
        Validaciones:
        - Clave unica
        - Campos requeridos
        - Formato correcto
        """
        try:
            serializer = self.get_serializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            self.perform_create(serializer)
            
            headers = self.get_success_headers(serializer.data)
            return Response(
                serializer.data, 
                status=status.HTTP_201_CREATED, 
                headers=headers
            )
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            logger.error(f"Error al crear producto: {str(e)}", exc_info=True)
            return Response({'error': 'Error al crear producto. Verifique los datos ingresados.'}, status=status.HTTP_400_BAD_REQUEST)
    
    def update(self, request, *args, **kwargs):
        """
        Actualiza un producto existente.
        
        Validaciones:
        - Clave unica (si se modifica)
        - Datos validos
        """
        try:
            partial = kwargs.pop('partial', False)
            instance = self.get_object()
            serializer = self.get_serializer(instance, data=request.data, partial=partial)
            serializer.is_valid(raise_exception=True)
            self.perform_update(serializer)
            
            return Response(serializer.data)
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            logger.error(f"Error al actualizar producto: {str(e)}", exc_info=True)
            return Response({'error': 'Error al actualizar producto. Verifique los datos ingresados.'}, status=status.HTTP_400_BAD_REQUEST)
    
    def destroy(self, request, *args, **kwargs):
        """
        Elimina un producto.
        
        Validaciones:
        - No puede eliminarse si tiene lotes asociados
        - Confirmacion de eliminacion
        """
        instance = self.get_object()
        
        try:
            if instance.lotes.exists():
                return Response({
                    'error': 'No se puede eliminar el producto',
                    'razon': 'Tiene lotes asociados',
                    'sugerencia': 'Elimine primero los lotes o marque el producto como inactivo'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            instance.delete()
            return Response({'mensaje': 'Producto eliminado exitosamente'}, status=status.HTTP_204_NO_CONTENT)
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            logger.error(f"Error al eliminar producto: {str(e)}", exc_info=True)
            return Response({'error': 'Error al eliminar producto'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=True, methods=['post'], url_path='toggle-activo')
    def toggle_activo(self, request, pk=None):
        """
        Activa o desactiva un producto.
        POST /api/productos/{id}/toggle-activo/
        
        Reglas:
        - No se puede desactivar un producto con stock disponible > 0
        - Usa update() directo para evitar validaci�n de otros campos
        """
        try:
            producto = self.get_object()
            nuevo_estado = not producto.activo
            
            # Si se va a desactivar, verificar que no tenga stock disponible
            if not nuevo_estado:  # Desactivando
                stock_disponible = producto.lotes.filter(
                    estado='disponible',
                    cantidad_actual__gt=0
                ).aggregate(total=Sum('cantidad_actual'))['total'] or 0
                
                if stock_disponible > 0:
                    return Response({
                        'error': 'No se puede desactivar el producto',
                        'razon': f'Tiene {stock_disponible} unidades en stock disponible',
                        'sugerencia': 'Transfiera o agote el inventario antes de desactivar'
                    }, status=status.HTTP_400_BAD_REQUEST)
            
            # Usar update() directo para evitar validaci�n de otros campos
            Producto.objects.filter(pk=producto.pk).update(activo=nuevo_estado)
            
            estado = 'activado' if nuevo_estado else 'desactivado'
            return Response({
                'mensaje': f'Producto {estado} exitosamente',
                'activo': nuevo_estado,
                'id': producto.id
            }, status=status.HTTP_200_OK)
        except Producto.DoesNotExist:
            return Response({'error': 'Producto no encontrado'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"Error en toggle_activo: {str(e)}", exc_info=True)
            return Response({'error': f'Error interno: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=True, methods=['get'], url_path='auditoria')
    def auditoria(self, request, pk=None):
        """
        Obtiene el historial de cambios de un producto.
        GET /api/productos/{id}/auditoria/
        """
        try:
            from core.models import AuditoriaLog
            producto = self.get_object()
            
            # Buscar logs de auditor�a relacionados con este producto
            logs = AuditoriaLog.objects.filter(
                Q(tabla='producto') | Q(tabla='core_producto'),
                registro_id=str(producto.id)
            ).order_by('-fecha')[:50]
            
            historial = []
            for log in logs:
                historial.append({
                    'id': log.id,
                    'fecha': log.fecha.isoformat() if log.fecha else None,
                    'usuario': log.usuario.get_full_name() or log.usuario.username if log.usuario else 'Sistema',
                    'accion': log.accion,
                    'cambios': log.cambios if hasattr(log, 'cambios') else None,
                    'ip': log.ip if hasattr(log, 'ip') else None,
                })
            
            return Response({
                'producto': {
                    'id': producto.id,
                    'clave': producto.clave,
                    'descripcion': producto.descripcion,
                },
                'historial': historial,
                'total': len(historial)
            })
        except Producto.DoesNotExist:
            return Response({'error': 'Producto no encontrado'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"Error en auditoria: {str(e)}", exc_info=True)
            return Response({'error': 'Error al obtener auditoría'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['get'], url_path='exportar-excel')
    def exportar_excel(self, request):
        """
        Exporta todos los productos a un archivo Excel.
        
        Columnas:
        - #, Clave, Descripcion, Unidad, Precio, Stock Minimo, Stock Actual, Lotes, Estado
        """
        try:
            productos = self.get_queryset()
            
            # Crear libro de Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Productos'
            
            # Encabezados
            headers = ['#', 'Clave', 'Descripcion', 'Unidad Medida', 'Precio Unitario', 'Stock Minimo', 'Stock Actual', 'Lotes Activos', 'Estado']
            ws.append(headers)
            
            # Estilo de encabezados
            header_fill = PatternFill(start_color='632842', end_color='632842', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=12)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Datos de productos
            for idx, producto in enumerate(productos, start=1):
                stock_actual = producto.lotes.filter(estado='disponible').aggregate(total=Sum('cantidad_actual'))['total'] or 0
                lotes_activos = producto.lotes.filter(estado='disponible', cantidad_actual__gt=0).count()
                
                ws.append([
                    idx,
                    producto.clave,
                    producto.descripcion,
                    producto.unidad_medida,
                    float(producto.precio_unitario) if producto.precio_unitario else 0,
                    producto.stock_minimo,
                    stock_actual,
                    lotes_activos,
                    'Activo' if producto.activo else 'Inactivo'
                ])
                
                # Colorear fila si el stock esta por debajo del minimo
                if stock_actual < producto.stock_minimo:
                    for col in range(1, 10):
                        ws.cell(row=idx+1, column=col).fill = PatternFill(
                            start_color='FFF4E6', 
                            end_color='FFF4E6', 
                            fill_type='solid'
                        )
            
            # Ajustar anchos de columna
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 50
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 15
            ws.column_dimensions['H'].width = 15
            ws.column_dimensions['I'].width = 12
            
            # Generar respuesta
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename=Productos_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            wb.save(response)
            
            return response
            
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            return Response({
                'error': 'Error al exportar productos',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['post'], url_path='importar-excel')
    def importar_excel(self, request):
        """
        Importa productos desde un archivo Excel.
        
        Formato esperado:
        Fila 1: Encabezados (se ignora)
        Columnas: Clave | Descripcion | Unidad | Precio | Stock Minimo | Estado
        
        Limites de seguridad:
        - Tamano maximo: configurado en IMPORT_MAX_FILE_SIZE_MB (default 10MB)
        - Filas maximas: configurado en IMPORT_MAX_ROWS (default 5000)
        - Extensiones: .xlsx, .xls
        """
        file = request.FILES.get('file')
        
        # Validar archivo
        es_valido, error_msg = validar_archivo_excel(file)
        if not es_valido:
            return Response({
                'error': 'Archivo invalido',
                'mensaje': error_msg
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            # ISS-001: Usar carga segura con limite real de bytes
            wb, error_carga = cargar_workbook_seguro(file)
            if wb is None:
                return Response({
                    'error': 'Error al procesar archivo',
                    'mensaje': error_carga
                }, status=status.HTTP_400_BAD_REQUEST)
            
            ws = wb.active
            
            # Validar numero de filas
            filas_validas, error_filas, num_filas = validar_filas_excel(ws)
            if not filas_validas:
                wb.close()  # Liberar recursos en modo read_only
                return Response({
                    'error': 'Archivo demasiado grande',
                    'mensaje': error_filas
                }, status=status.HTTP_400_BAD_REQUEST)
            
            creados = 0
            actualizados = 0
            errores = []
            exitos = []
            
            # Procesar cada fila (empezando desde la fila 2)
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Extraer valores (asegurarse de tener al menos 6 columnas)
                    valores = list(row) + [None] * 6
                    clave = valores[0]
                    descripcion = valores[1]
                    unidad_medida = valores[2]
                    precio_unitario = valores[3]
                    stock_minimo = valores[4]
                    estado = valores[5]
                    
                    # Validar campos requeridos
                    if not clave or not descripcion:
                        errores.append({'fila': row_idx, 'error': 'Clave y descripcion son obligatorios'})
                        continue
                    
                    # Validar unidad
                    unidad_limpia = str(unidad_medida).strip().upper() if unidad_medida else 'PIEZA'
                    if unidad_limpia not in dict(UNIDADES_MEDIDA):
                        errores.append({'fila': row_idx, 'error': f'Unidad no valida: {unidad_limpia}'})
                        continue

                    # Validar precio y stock
                    try:
                        precio_val = float(precio_unitario) if precio_unitario not in [None, ''] else 0.0
                        if precio_val < 0:
                            raise ValueError
                    except Exception:
                        errores.append({'fila': row_idx, 'error': 'Precio invalido'})
                        continue

                    try:
                        stock_min = int(stock_minimo) if stock_minimo not in [None, ''] else 10
                        if stock_min < 0:
                            raise ValueError
                    except Exception:
                        errores.append({'fila': row_idx, 'error': 'Stock minimo invalido'})
                        continue

                    # Limpiar y preparar datos
                    datos = {
                        'descripcion': str(descripcion).strip(),
                        'unidad_medida': unidad_limpia,
                        'precio_unitario': precio_val,
                        'stock_minimo': stock_min,
                        'activo': str(estado).lower() in ['activo', 'sI', 'si', 'si', 'true', '1', 'yes'] if estado else True
                    }
                    
                    # Crear o actualizar producto
                    producto, created = Producto.objects.update_or_create(
                        clave=str(clave).upper().strip(),
                        defaults=datos
                    )
                    
                    if created:
                        creados += 1
                    else:
                        actualizados += 1
                    exitos.append({'fila': row_idx, 'producto_id': producto.id, 'clave': producto.clave})
                        
                except Exception as e:
                    errores.append({'fila': row_idx, 'error': str(e)})
            
            return Response({
                'mensaje': 'Importacion completada',
                'resumen': {
                    'creados': creados,
                    'actualizados': actualizados,
                    'total_procesados': creados + actualizados,
                    'total_errores': len(errores)
                },
                'exitos': exitos,
                'errores': errores,
                'exito': len(errores) == 0
            }, status=status.HTTP_200_OK if len(errores) == 0 else status.HTTP_207_MULTI_STATUS)
            
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            return Response({
                'error': 'Error al procesar archivo',
                'mensaje': str(e),
                'sugerencia': 'Verifique que el archivo tenga el formato correcto: Clave, Descripcion, Unidad, Precio, Stock Minimo, Estado'
            }, status=status.HTTP_400_BAD_REQUEST)


class CentroViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestionar centros penitenciarios.
    
    Funcionalidades:
    - CRUD completo
    - Busqueda por clave, nombre y direccion
    - Filtrado por estado activo/inactivo
    - Exportar a Excel con formato profesional
    - Importar desde Excel con validaciones
    - Obtener requisiciones por centro
    """
    queryset = Centro.objects.all()
    serializer_class = CentroSerializer
    permission_classes = [IsFarmaciaAdminOrReadOnly]
    pagination_class = CustomPagination

    def _user_centro(self, user):
        return getattr(user, 'centro', None) or getattr(getattr(user, 'profile', None), 'centro', None)

    def get_queryset(self):
        """Filtra centros segun parametros"""
        queryset = Centro.objects.all()
        user = getattr(self.request, 'user', None)
        
        # Admin, Farmacia y Superusuarios pueden ver todos los centros
        # Otros usuarios solo ven su propio centro
        if user and not user.is_superuser and not is_farmacia_or_admin(user):
            user_centro = self._user_centro(user)
            if user_centro:
                queryset = queryset.filter(id=user_centro.id)
            else:
                return Centro.objects.none()
        
        # Filtro por busqueda
        search = self.request.query_params.get('search')
        if search and search.strip():
            queryset = queryset.filter(
                Q(clave__icontains=search) | 
                Q(nombre__icontains=search) | 
                Q(direccion__icontains=search)
            )
        
        # Filtro por estado activo
        activo = self.request.query_params.get('activo')
        if activo == 'true':
            queryset = queryset.filter(activo=True)
        elif activo == 'false':
            queryset = queryset.filter(activo=False)
        
        return queryset.order_by('-created_at')
    
    def create(self, request, *args, **kwargs):
        """Crea un nuevo centro"""
        try:
            logger.debug(f"CREAR CENTRO - Body: {request.data}")
            
            serializer = self.get_serializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            centro = serializer.save()
            
            logger.info(f"Centro creado: {centro.clave} - {centro.nombre}")
            
            return Response({
                'mensaje': 'Centro creado exitosamente',
                'centro': CentroSerializer(centro).data
            }, status=status.HTTP_201_CREATED)
            
        except serializers.ValidationError as e:
            logger.warning(f"Error de validación al crear centro: {e.detail}")
            return Response({
                'error': 'Error de validacion',
                'detalles': e.detail
            }, status=status.HTTP_400_BAD_REQUEST)
            
        except Exception as e:
            logger.exception(f"Error inesperado al crear centro: {e}")
            # traceback removido por seguridad (ISS-008)
            return Response({
                'error': 'Error al crear centro',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def update(self, request, *args, **kwargs):
        """Actualiza un centro existente"""
        try:
            partial = kwargs.pop('partial', False)
            instance = self.get_object()
            serializer = self.get_serializer(instance, data=request.data, partial=partial)
            serializer.is_valid(raise_exception=True)
            self.perform_update(serializer)
            
            return Response({
                'mensaje': 'Centro actualizado exitosamente',
                'centro': serializer.data
            })
        except serializers.ValidationError as e:
            return Response(
                {'error': 'Error de validacion', 'detalles': e.detail},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            return Response(
                {'error': 'Error al actualizar centro', 'mensaje': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def destroy(self, request, *args, **kwargs):
        """
        Elimina un centro.
        
        Validaciones:
        - No puede eliminarse si tiene requisiciones asociadas
        - No puede eliminarse si tiene usuarios asignados
        """
        instance = self.get_object()
        
        try:
            # Verificar requisiciones
            if hasattr(instance, 'requisiciones') and instance.requisiciones.exists():
                total_requisiciones = instance.requisiciones.count()
                requisiciones_activas = instance.requisiciones.exclude(
                    estado__in=['CANCELADA', 'SURTIDA']
                ).count()
                
                return Response({
                    'error': 'No se puede eliminar el centro',
                    'razon': 'Tiene requisiciones asociadas',
                    'total_requisiciones': total_requisiciones,
                    'requisiciones_activas': requisiciones_activas,
                    'sugerencia': 'Marque el centro como inactivo en lugar de eliminarlo'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Verificar usuarios asignados
            if hasattr(instance, 'usuarios') and instance.usuarios.exists():
                total_usuarios = instance.usuarios.count()
                
                return Response({
                    'error': 'No se puede eliminar el centro',
                    'razon': 'Tiene usuarios asignados',
                    'total_usuarios': total_usuarios,
                    'sugerencia': 'Reasigne los usuarios a otro centro o marque el centro como inactivo'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Si no tiene relaciones, se puede eliminar
            clave_eliminada = instance.clave
            nombre_eliminado = instance.nombre
            instance.delete()
            
            return Response({
                'mensaje': 'Centro eliminado exitosamente',
                'centro_eliminado': f"{clave_eliminada} - {nombre_eliminado}"
            }, status=status.HTTP_204_NO_CONTENT)
            
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            return Response({
                'error': 'Error al eliminar centro',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['post'], url_path='toggle-activo')
    def toggle_activo(self, request, pk=None):
        """
        Activa o desactiva un centro.
        POST /api/centros/{id}/toggle-activo/
        
        Usa update() directo para evitar validaci�n de otros campos.
        """
        try:
            centro = self.get_object()
            nuevo_estado = not centro.activo
            
            # Usar update() directo para evitar validaci�n de otros campos
            Centro.objects.filter(pk=centro.pk).update(activo=nuevo_estado)
            
            estado = 'activado' if nuevo_estado else 'desactivado'
            return Response({
                'mensaje': f'Centro {estado} exitosamente',
                'activo': nuevo_estado,
                'id': centro.id,
                'clave': centro.clave,
                'nombre': centro.nombre
            }, status=status.HTTP_200_OK)
        except Centro.DoesNotExist:
            return Response({'error': 'Centro no encontrado'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"Error en toggle_activo centro: {str(e)}", exc_info=True)
            return Response({'error': f'Error interno: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['get'])
    def inventario(self, request, pk=None):
        """Devuelve inventario resumido del centro a partir de lotes asociados a movimientos del centro."""
        centro = self.get_object()
        user_centro = self._user_centro(request.user)
        if not request.user.is_superuser:
            if not user_centro or user_centro.id != centro.id:
                return Response({'error': 'Solo puedes ver inventario de tu centro'}, status=status.HTTP_403_FORBIDDEN)

        # Lotes que han tenido movimientos en este centro
        lote_ids = Movimiento.objects.filter(centro=centro).values_list('lote_id', flat=True)
        lotes = Lote.objects.filter(
            Q(id__in=lote_ids) | Q(centro=centro),
            estado='disponible',
            deleted_at__isnull=True,
            cantidad_actual__gt=0
        ).select_related('producto')

        inventario_dict = {}
        for lote in lotes:
            prod = lote.producto
            item = inventario_dict.setdefault(prod.id, {
                'producto_id': prod.id,
                'clave': prod.clave,
                'producto': prod.descripcion,
                'cantidad_disponible': 0,
                'lote_proximo_caducar': None,
                'fecha_caducidad': None,
            })
            item['cantidad_disponible'] += lote.cantidad_actual
            if lote.fecha_caducidad:
                fecha_actual = item['fecha_caducidad']
                if fecha_actual is None or lote.fecha_caducidad < fecha_actual:
                    item['lote_proximo_caducar'] = lote.numero_lote
                    item['fecha_caducidad'] = lote.fecha_caducidad

        # Si no hay lotes asociados, caer al agregado por movimientos para no dejar vac�o
        inventario = list(inventario_dict.values())
        if not inventario:
            movimientos = Movimiento.objects.filter(centro=centro)
            agregados = movimientos.values('lote__producto').annotate(cantidad=Coalesce(Sum('cantidad'), 0))
            for item in agregados:
                producto = Producto.objects.filter(id=item['lote__producto']).first()
                if not producto:
                    continue
                inventario.append({
                    'producto_id': producto.id,
                    'clave': producto.clave,
                    'producto': producto.descripcion,
                    'cantidad_disponible': max(0, item['cantidad']),
                    'lote_proximo_caducar': None,
                    'fecha_caducidad': None,
                })

        return Response({
            'centro': centro.nombre,
            'centro_id': centro.id,
            'total_productos': len(inventario),
            'inventario': inventario
        })
    
    @action(detail=False, methods=['get'], url_path='exportar-excel')
    def exportar_excel(self, request):
        """
        Exporta todos los centros a Excel con formato profesional.
        
        Columnas:
        - #, Clave, Nombre, Direccion, Telefono, Total Requisiciones, Estado
        """
        try:
            centros = self.get_queryset()
            
            # Crear libro de Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Centros Penitenciarios'
            
            # Titulo del reporte
            ws.merge_cells('A1:G1')
            titulo_cell = ws['A1']
            titulo_cell.value = 'REPORTE DE CENTROS PENITENCIARIOS'
            titulo_cell.font = Font(bold=True, size=14, color='632842')
            titulo_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Fecha de generacion
            ws.merge_cells('A2:G2')
            fecha_cell = ws['A2']
            fecha_cell.value = f'Generado el {timezone.now().strftime("%d/%m/%Y %H:%M")}'

            fecha_cell.font = Font(size=10, italic=True)
            fecha_cell.alignment = Alignment(horizontal='center')
            
            # Espacio
            ws.append([])
            
            # Encabezados
            headers = ['#', 'Clave', 'Nombre', 'Direccion', 'Telefono', 'Total Requisiciones', 'Estado']
            ws.append(headers)
            
            # Estilo de encabezados
            header_fill = PatternFill(start_color='632842', end_color='632842', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            for col_num, cell in enumerate(ws[4], 1):
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # Datos de centros
            for idx, centro in enumerate(centros, start=1):
                # Calcular total de requisiciones si existe la relacion
                total_requisiciones = 0
                if hasattr(centro, 'requisiciones'):
                    total_requisiciones = centro.requisiciones.count()
                
                ws.append([
                    idx,
                    centro.clave,
                    centro.nombre,
                    centro.direccion or 'Sin direccion',
                    centro.telefono or 'Sin telefono',
                    total_requisiciones,
                    'Activo' if centro.activo else 'Inactivo'
                ])
                
                # Estilo para filas
                row_num = idx + 4
                for cell in ws[row_num]:
                    cell.alignment = Alignment(vertical='center')
                    
                # Colorear estado
                estado_cell = ws.cell(row=row_num, column=7)
                if centro.activo:
                    estado_cell.fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
                    estado_cell.font = Font(color='155724', bold=True)
                else:
                    estado_cell.fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')
                    estado_cell.font = Font(color='721C24', bold=True)
            
            # Ajustar anchos de columna
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 50
            ws.column_dimensions['D'].width = 40
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 20
            ws.column_dimensions['G'].width = 12
            
            # Agregar bordes
            from openpyxl.styles import Border, Side
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=7):
                for cell in row:
                    cell.border = thin_border
            
            # Resumen al final
            ws.append([])
            resumen_row = ws.max_row + 1
            ws.merge_cells(f'A{resumen_row}:C{resumen_row}')
            resumen_cell = ws[f'A{resumen_row}']
            resumen_cell.value = f'TOTAL DE CENTROS: {centros.count()}'
            resumen_cell.font = Font(bold=True, size=11)
            resumen_cell.alignment = Alignment(horizontal='left')
            
            # Generar respuesta
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename=Centros_Penitenciarios_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            wb.save(response)
            
            return response
            
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            return Response({
                'error': 'Error al exportar centros',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['post'], url_path='importar')
    def importar_excel(self, request):
        """
        Importa centros desde Excel.
        
        Formato esperado:
        - Clave (requerido)
        - Nombre (requerido)
        - Direccion (opcional)
        - Telefono (opcional)
        - Estado (Activo/Inactivo)
        
        Limites de seguridad:
        - Tamano maximo: configurado en IMPORT_MAX_FILE_SIZE_MB (default 10MB)
        - Filas maximas: configurado en IMPORT_MAX_ROWS (default 5000)
        - Extensiones: .xlsx, .xls
        """
        file = request.FILES.get('file')
        
        # Validar archivo
        es_valido, error_msg = validar_archivo_excel(file)
        if not es_valido:
            return Response({
                'error': 'Archivo invalido',
                'mensaje': error_msg
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            # ISS-001: Usar carga segura con limite real de bytes
            wb, error_carga = cargar_workbook_seguro(file)
            if wb is None:
                return Response({
                    'error': 'Error al procesar archivo',
                    'mensaje': error_carga
                }, status=status.HTTP_400_BAD_REQUEST)
            
            ws = wb.active
            
            # Validar numero de filas
            filas_validas, error_filas, num_filas = validar_filas_excel(ws)
            if not filas_validas:
                wb.close()  # Liberar recursos en modo read_only
                return Response({
                    'error': 'Archivo demasiado grande',
                    'mensaje': error_filas
                }, status=status.HTTP_400_BAD_REQUEST)
            
            creados = 0
            actualizados = 0
            errores = []
            
            # Procesar filas
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Extraer datos
                    clave, nombre, direccion, telefono, estado = row[:5]
                    
                    # Validar requeridos
                    if not clave or not nombre:
                        errores.append(f'Fila {row_idx}: Clave y nombre son requeridos')
                        continue
                    
                    # Preparar datos
                    datos = {
                        'nombre': str(nombre).strip(),
                        'direccion': str(direccion).strip() if direccion else '',
                        'telefono': str(telefono).strip() if telefono else '',
                        'activo': str(estado).lower() in ['activo', 'si', 'si', 'true', '1'] if estado else True
                    }
                    
                    # Crear o actualizar
                    centro, created = Centro.objects.update_or_create(
                        clave=str(clave).upper().strip(),
                        defaults=datos
                    )
                    
                    if created:
                        creados += 1
                    else:
                        actualizados += 1
                        
                except Exception as e:
                    errores.append(f'Fila {row_idx}: {str(e)}')
            
            return Response({
                'mensaje': 'Importacion completada',
                'resumen': {
                    'creados': creados,
                    'actualizados': actualizados,
                    'total_procesados': creados + actualizados,
                    'errores_encontrados': len(errores)
                },
                'errores': errores[:10] if errores else [],  # Maximo 10 errores
                'tiene_mas_errores': len(errores) > 10,
                'exito': len(errores) == 0
            }, status=status.HTTP_200_OK if len(errores) == 0 else status.HTTP_207_MULTI_STATUS)
            
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            return Response({
                'error': 'Error al procesar el archivo',
                'mensaje': str(e),
                'sugerencia': 'Verifique que el archivo tenga el formato correcto (Clave, Nombre, Direccion, Telefono, Estado)'
            }, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['get'])
    def requisiciones(self, request, pk=None):
        """Obtiene todas las requisiciones de un centro"""
        try:
            centro = self.get_object()
            
            if not hasattr(centro, 'requisiciones'):
                return Response({
                    'centro': {
                        'id': centro.id,
                        'clave': centro.clave,
                        'nombre': centro.nombre
                    },
                    'requisiciones': [],
                    'total': 0,
                    'mensaje': 'No hay requisiciones disponibles'
                })
            
            requisiciones = centro.requisiciones.all().order_by('-fecha_solicitud')
            
            # Agrupar por estado
            por_estado = {}
            for req in requisiciones:
                estado = req.estado
                if estado not in por_estado:
                    por_estado[estado] = 0
                por_estado[estado] += 1
            
            requisiciones_data = []
            for req in requisiciones:
                requisiciones_data.append({
                    'id': req.id,
                    'folio': req.folio,
                    'estado': req.estado,
                    'fecha_solicitud': req.fecha_solicitud,
                    'total_items': req.items.count() if hasattr(req, 'items') else 0
                })
            
            return Response({
                'centro': {
                    'id': centro.id,
                    'clave': centro.clave,
                    'nombre': centro.nombre
                },
                'estadisticas': {
                    'total': requisiciones.count(),
                    'por_estado': por_estado
                },
                'requisiciones': requisiciones_data
            })
            
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            return Response({
                'error': 'Error al obtener requisiciones',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['get'], url_path='plantilla')
    def plantilla_centros(self, request):
        """Descarga plantilla de Excel para importaci?n de centros."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Centros'
        headers = ['Clave', 'Nombre', 'Direccion', 'Telefono', 'Estado']
        ws.append(headers)
        ws.append(['CENTRO-001', 'Centro Ejemplo', 'Direccion de ejemplo', '555-0000', 'Activo'])
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=Plantilla_Centros.xlsx'
        wb.save(response)
        return response


class LoteViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestionar lotes.
    
    Funcionalidades:
    - CRUD completo
    - Filtrado por producto
    - Filtrado por estado de caducidad
    - Busqueda por numero de lote
    - Validaciones de integridad
    """
    queryset = Lote.objects.select_related('producto').all()
    serializer_class = LoteSerializer
    permission_classes = [IsFarmaciaAdminOrReadOnly]
    pagination_class = CustomPagination

    def get_queryset(self):
        """
        Filtra lotes segun parametros.
        
        Parametros:
        - producto: ID del producto
        - activo: true/false
        - caducidad: vencido/critico/proximo/normal
        - search: busqueda por numero de lote
        - centro: ID del centro o 'central' para farmacia (solo admin/farmacia/vista)
        
        Seguridad: Usuarios de centro solo ven lotes de su centro.
        Admin/farmacia/vista ven todo por defecto, pueden filtrar con ?centro=.
        """
        queryset = Lote.objects.select_related('producto', 'centro', 'lote_origen').filter(deleted_at__isnull=True).exclude(estado='retirado')
        
        # SEGURIDAD: Filtrar por centro segun rol
        user = self.request.user
        if not is_farmacia_or_admin(user):
            # Usuario de centro: forzado a su centro
            user_centro = get_user_centro(user)
            if user_centro:
                queryset = queryset.filter(centro=user_centro)
            else:
                return Lote.objects.none()
        else:
            # Admin/farmacia/vista: pueden filtrar por centro especifico
            centro_param = self.request.query_params.get('centro')
            if centro_param:
                if centro_param == 'central':
                    # Filtrar solo farmacia central (centro=NULL)
                    queryset = queryset.filter(centro__isnull=True)
                else:
                    queryset = queryset.filter(centro_id=centro_param)
        
        # Filtrar por producto
        producto = self.request.query_params.get('producto')
        if producto:
            queryset = queryset.filter(producto_id=producto)
        
        # Filtrar por estado activo
        activo = self.request.query_params.get('activo')
        if activo == 'true':
            queryset = queryset.filter(deleted_at__isnull=True)
        elif activo == 'false':
            queryset = queryset.filter(deleted_at__isnull=False)
        
        # Busqueda por numero de lote, clave o descripcion producto (ISS-003)
        search = self.request.query_params.get('search')
        if search and search.strip():
            search_term = search.strip()
            queryset = queryset.filter(
                Q(numero_lote__icontains=search_term) |
                Q(producto__clave__icontains=search_term) |
                Q(producto__descripcion__icontains=search_term)
            )
        
        # Filtrar por estado de caducidad seg�n especificaci�n SIFP:
        # ?? Normal: > 6 meses (180 d�as)
        # ?? Pr�ximo: 3-6 meses (90-180 d�as)
        # ?? Cr�tico: < 3 meses (90 d�as)
        # ?? Vencido: < 0 d�as
        caducidad = self.request.query_params.get('caducidad')
        if caducidad:
            from datetime import date, timedelta
            hoy = date.today()
            
            if caducidad == 'vencido':
                queryset = queryset.filter(fecha_caducidad__lt=hoy)
            elif caducidad == 'critico':
                # Menos de 3 meses (< 90 d�as) pero no vencido
                queryset = queryset.filter(
                    fecha_caducidad__gte=hoy,
                    fecha_caducidad__lt=hoy + timedelta(days=90)
                )
            elif caducidad == 'proximo':
                # Entre 3 y 6 meses (90-180 d�as)
                queryset = queryset.filter(
                    fecha_caducidad__gte=hoy + timedelta(days=90),
                    fecha_caducidad__lt=hoy + timedelta(days=180)
                )
            elif caducidad == 'normal':
                # M�s de 6 meses (> 180 d�as)
                queryset = queryset.filter(fecha_caducidad__gte=hoy + timedelta(days=180))
        
        # Filtrar por stock m�nimo (para cat�logo de requisiciones)
        stock_min = self.request.query_params.get('stock_min')
        if stock_min:
            try:
                queryset = queryset.filter(cantidad_actual__gte=int(stock_min))
            except (ValueError, TypeError):
                pass
        
        # Filtrar por existencia de stock (con_stock/sin_stock)
        con_stock = self.request.query_params.get('con_stock')
        if con_stock == 'con_stock':
            queryset = queryset.filter(cantidad_actual__gt=0)
        elif con_stock == 'sin_stock':
            queryset = queryset.filter(cantidad_actual=0)
        
        # Filtrar solo lotes disponibles (no vencidos) para el cat�logo
        solo_disponibles = self.request.query_params.get('solo_disponibles')
        if solo_disponibles == 'true':
            from datetime import date
            queryset = queryset.filter(
                estado='disponible',
                fecha_caducidad__gt=date.today()
            )
        
        return queryset.order_by('-created_at')
    
    def create(self, request, *args, **kwargs):
        """Crea un nuevo lote con validaciones"""
        try:
            serializer = self.get_serializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            self.perform_create(serializer)
            
            headers = self.get_success_headers(serializer.data)
            return Response(
                serializer.data, 
                status=status.HTTP_201_CREATED, 
                headers=headers
            )
        except serializers.ValidationError as e:
            return Response(
                {'error': 'Error de validacion', 'detalles': e.detail}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            return Response(
                {'error': 'Error al crear lote', 'mensaje': str(e)}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def update(self, request, *args, **kwargs):
        """Actualiza un lote existente"""
        try:
            partial = kwargs.pop('partial', False)
            instance = self.get_object()
            serializer = self.get_serializer(instance, data=request.data, partial=partial)
            serializer.is_valid(raise_exception=True)
            self.perform_update(serializer)
            
            return Response(serializer.data)
        except serializers.ValidationError as e:
            return Response(
                {'error': 'Error de validacion', 'detalles': e.detail}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            return Response(
                {'error': 'Error al actualizar lote', 'mensaje': str(e)}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def destroy(self, request, *args, **kwargs):
        """
        Elimina un lote.
        
        Validaciones:
        - No puede eliminarse si tiene movimientos asociados
        """
        instance = self.get_object()
        
        try:
            # Verificar si tiene movimientos
            if Movimiento.objects.filter(lote=instance).exists():
                total_movimientos = Movimiento.objects.filter(lote=instance).count()
                
                return Response({
                    'error': 'No se puede eliminar el lote',
                    'razon': 'Tiene movimientos asociados',
                    'total_movimientos': total_movimientos,
                    'sugerencia': 'Marque el lote como inactivo en lugar de eliminarlo'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Si no tiene movimientos, se puede eliminar
            numero_lote = instance.numero_lote
            producto_clave = instance.producto.clave
            instance.delete()
            
            return Response({
                'mensaje': 'Lote eliminado exitosamente',
                'lote_eliminado': numero_lote,
                'producto': producto_clave
            }, status=status.HTTP_204_NO_CONTENT)
            
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            return Response({
                'error': 'Error al eliminar lote',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'])
    def exportar_excel(self, request):
        """Exporta lotes aplicando los mismos filtros de listado."""
        try:
            lotes = self.get_queryset()
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Lotes'

            ws.merge_cells('A1:H1')
            ws['A1'] = 'REPORTE DE LOTES'
            ws['A1'].font = Font(bold=True, size=14, color='632842')
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

            ws.append([])
            headers = ['#', 'Producto', 'Numero de lote', 'Caducidad', 'Cantidad inicial', 'Cantidad actual', 'Estado', 'Proveedor']
            ws.append(headers)
            header_fill = PatternFill(start_color='632842', end_color='632842', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            for cell in ws[3]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            for idx, lote in enumerate(lotes, 1):
                ws.append([
                    idx,
                    getattr(lote.producto, 'clave', ''),
                    lote.numero_lote,
                    lote.fecha_caducidad.strftime('%Y-%m-%d') if lote.fecha_caducidad else '',
                    lote.cantidad_inicial,
                    lote.cantidad_actual,
                    lote.estado,
                    lote.proveedor or ''
                ])

            for col, width in zip(['A','B','C','D','E','F','G','H'], [6,12,18,14,14,14,12,16]):
                ws.column_dimensions[col].width = width

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename=Lotes_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            wb.save(response)
            return response
        except Exception as exc:
            # traceback removido por seguridad (ISS-008)
            return Response({'error': 'Error al exportar lotes', 'mensaje': str(exc)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path='exportar-excel')
    def exportar_excel_lotes(self, request):
        """Alias para exportar_excel con url_path correcto."""
        return self.exportar_excel(request)

    @action(detail=False, methods=['post'], url_path='importar-excel')
    def importar_excel(self, request):
        """
        Importa lotes desde Excel con validaciones.
        
        Limites de seguridad:
        - Tamano maximo: configurado en IMPORT_MAX_FILE_SIZE_MB (default 10MB)
        - Filas maximas: configurado en IMPORT_MAX_ROWS (default 5000)
        - Extensiones: .xlsx, .xls
        """
        file = request.FILES.get('file')
        
        # Validar archivo
        es_valido, error_msg = validar_archivo_excel(file)
        if not es_valido:
            return Response({
                'error': 'Archivo invalido',
                'mensaje': error_msg
            }, status=status.HTTP_400_BAD_REQUEST)

        try:
            # ISS-001: Usar carga segura con limite real de bytes
            wb, error_carga = cargar_workbook_seguro(file)
            if wb is None:
                return Response({
                    'error': 'Error al procesar archivo',
                    'mensaje': error_carga
                }, status=status.HTTP_400_BAD_REQUEST)
            
            ws = wb.active
            
            # Validar numero de filas
            filas_validas, error_filas, num_filas = validar_filas_excel(ws)
            if not filas_validas:
                wb.close()  # Liberar recursos en modo read_only
                return Response({
                    'error': 'Archivo demasiado grande',
                    'mensaje': error_filas
                }, status=status.HTTP_400_BAD_REQUEST)
            
            exitos = []
            errores = []

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    producto_clave, numero_lote, fecha_cad, cantidad_inicial, cantidad_actual, proveedor = (row + (None,)*6)[:6]

                    if not producto_clave or not numero_lote:
                        errores.append({'fila': row_idx, 'error': 'Producto y numero de lote son obligatorios'})
                        continue

                    producto = Producto.objects.filter(clave__iexact=str(producto_clave).strip()).first()
                    if not producto:
                        errores.append({'fila': row_idx, 'error': f'Producto no encontrado: {producto_clave}'})
                        continue

                    try:
                        fecha_val = None
                        if fecha_cad:
                            if isinstance(fecha_cad, (datetime, date)):
                                fecha_val = fecha_cad.date() if hasattr(fecha_cad, 'date') else fecha_cad
                            else:
                                fecha_val = datetime.strptime(str(fecha_cad), '%Y-%m-%d').date()
                    except Exception:
                        errores.append({'fila': row_idx, 'error': 'Fecha de caducidad invalida'})
                        continue

                    try:
                        cant_ini = int(cantidad_inicial) if cantidad_inicial not in [None, ''] else 0
                        cant_act = int(cantidad_actual) if cantidad_actual not in [None, ''] else cant_ini
                        if cant_ini <= 0 or cant_act < 0:
                            raise ValueError
                    except Exception:
                        errores.append({'fila': row_idx, 'error': 'Cantidades invalidas'})
                        continue

                    lote, created = Lote.objects.update_or_create(
                        producto=producto,
                        numero_lote=str(numero_lote).strip().upper(),
                        defaults={
                            'fecha_caducidad': fecha_val or date.today(),
                            'cantidad_inicial': cant_ini,
                            'cantidad_actual': cant_act,
                            'proveedor': proveedor or ''
                        }
                    )
                    exitos.append({'fila': row_idx, 'lote_id': lote.id, 'numero_lote': lote.numero_lote, 'created': created})
                except Exception as exc:
                    errores.append({'fila': row_idx, 'error': str(exc)})

            status_code = status.HTTP_200_OK if not errores else status.HTTP_207_MULTI_STATUS
            return Response({
                'mensaje': 'Importacion de lotes completada',
                'resumen': {
                    'exitos': len(exitos),
                    'errores': len(errores),
                    'total': len(exitos) + len(errores)
                },
                'exitos': exitos,
                'errores': errores
            }, status=status_code)
        except Exception as exc:
            # traceback removido por seguridad (ISS-008)
            return Response({'error': 'Error al procesar archivo', 'mensaje': str(exc)}, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=False, methods=['get'])
    def por_vencer(self, request):
        """
        Obtiene lotes proximos a vencer.
        
        Parametros:
        - dias: numero de dias (default: 30)
        """
        try:
            from datetime import date, timedelta
            
            dias = int(request.query_params.get('dias', 30))
            fecha_limite = date.today() + timedelta(days=dias)
            
            lotes = Lote.objects.select_related('producto').filter(
                deleted_at__isnull=True,
                estado__in=['disponible', 'agotado', 'bloqueado'],
                cantidad_actual__gt=0,
                fecha_caducidad__lte=fecha_limite
            ).order_by('fecha_caducidad')
            
            serializer = self.get_serializer(lotes, many=True)
            
            return Response({
                'total': lotes.count(),
                'dias_configurados': dias,
                'fecha_limite': fecha_limite,
                'lotes': serializer.data
            })
            
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            return Response({
                'error': 'Error al obtener lotes por vencer',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path='por_caducar')
    def por_caducar(self, request):
        """Alias compatible para el frontend: lotes proximos a vencer."""
        try:
            from datetime import date, timedelta

            dias = int(request.query_params.get('dias', 90))
            hoy = date.today()
            fecha_limite = hoy + timedelta(days=dias)
            lotes = Lote.objects.select_related('producto').filter(
                deleted_at__isnull=True,
                cantidad_actual__gt=0,
                fecha_caducidad__gt=hoy,
                fecha_caducidad__lte=fecha_limite
            ).order_by('fecha_caducidad')

            page = self.paginate_queryset(lotes)
            if page is not None:
                serializer = self.get_serializer(page, many=True)
                return self.get_paginated_response(serializer.data)

            serializer = self.get_serializer(lotes, many=True)
            return Response(serializer.data)
        except Exception as exc:
            # traceback removido por seguridad (ISS-008)
            return Response({'error': 'Error al obtener lotes por caducar', 'mensaje': str(exc)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'])
    def vencidos(self, request):
        """Lotes con caducidad vencida y stock disponible."""
        try:
            from datetime import date

            hoy = date.today()
            lotes = Lote.objects.select_related('producto').filter(
                deleted_at__isnull=True,
                cantidad_actual__gt=0,
                fecha_caducidad__lt=hoy
            ).order_by('fecha_caducidad')

            page = self.paginate_queryset(lotes)
            if page is not None:
                serializer = self.get_serializer(page, many=True)
                return self.get_paginated_response(serializer.data)

            serializer = self.get_serializer(lotes, many=True)
            return Response(serializer.data)
        except Exception as exc:
            # traceback removido por seguridad (ISS-008)
            return Response({'error': 'Error al obtener lotes vencidos', 'mensaje': str(exc)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=True, methods=['get'])
    def historial(self, request, pk=None):
        """Obtiene el historial de movimientos de un lote"""
        try:
            lote = self.get_object()
            
            movimientos = Movimiento.objects.filter(lote=lote).select_related(
                'lote__producto'
            ).order_by('-fecha')
            
            from django.db.models import Sum
            
            total_entradas = movimientos.filter(tipo='entrada').aggregate(
                total=Sum('cantidad')
            )['total'] or 0
            
            total_salidas = movimientos.filter(tipo='salida').aggregate(
                total=Sum('cantidad')
            )['total'] or 0
            
            movimientos_data = []
            for mov in movimientos:
                movimientos_data.append({
                    'id': mov.id,
                    'tipo': mov.tipo,
                    'cantidad': mov.cantidad,
                    'fecha': mov.fecha,
                    'observaciones': mov.observaciones or ''
                })
            
            return Response({
                'lote': {
                    'id': lote.id,
                    'numero_lote': lote.numero_lote,
                    'producto': lote.producto.clave,
                    'cantidad_actual': lote.cantidad_actual
                },
                'estadisticas': {
                    'total_entradas': total_entradas,
                    'total_salidas': total_salidas,
                    'diferencia': total_entradas - total_salidas
                },
                'movimientos': movimientos_data,
                'total_movimientos': len(movimientos_data)
            })
            
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            return Response({
                'error': 'Error al obtener historial',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['post'])
    def ajustar_stock(self, request, pk=None):
        """
        Ajusta stock del lote y crea movimiento asociado (entrada/salida/ajuste).
        """
        lote = self.get_object()
        tipo = request.data.get('tipo', 'ajuste')
        cantidad = request.data.get('cantidad')
        observaciones = request.data.get('observaciones', '')

        try:
            movimiento, lote_actualizado = registrar_movimiento_stock(
                lote=lote,
                tipo=tipo,
                cantidad=cantidad,
                usuario=request.user if request.user.is_authenticated else None,
                centro=None,
                requisicion=None,
                observaciones=observaciones
            )
            return Response({
                'mensaje': 'Stock ajustado correctamente',
                'lote': self.get_serializer(lote_actualizado).data,
                'movimiento_id': movimiento.id
            })
        except serializers.ValidationError as exc:
            return Response({'error': 'Error de validacion', 'detalles': exc.detail}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as exc:
            # traceback removido por seguridad (ISS-008)
            return Response({'error': 'Error al ajustar stock', 'mensaje': str(exc)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['get'], url_path='lotes-derivados')
    def lotes_derivados(self, request, pk=None):
        """
        Obtiene los lotes derivados de un lote de farmacia (vinculados a centros).
        
        Solo aplica para lotes de farmacia central (centro=NULL).
        Muestra todos los centros que tienen stock de este lote.
        """
        try:
            lote = self.get_object()
            
            # Verificar que es un lote de farmacia
            if lote.centro is not None:
                return Response({
                    'error': 'Solo los lotes de farmacia central tienen lotes derivados',
                    'lote_id': lote.id,
                    'centro': lote.centro.nombre if lote.centro else None
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Obtener lotes derivados
            derivados = Lote.objects.filter(
                lote_origen=lote,
                deleted_at__isnull=True
            ).select_related('centro', 'producto')
            
            # Calcular totales
            from django.db.models import Sum
            total_derivados = derivados.count()
            stock_total_centros = derivados.aggregate(total=Sum('cantidad_actual'))['total'] or 0
            
            derivados_data = []
            for d in derivados:
                derivados_data.append({
                    'id': d.id,
                    'centro_id': d.centro.id if d.centro else None,
                    'centro_nombre': d.centro.nombre if d.centro else None,
                    'centro_clave': d.centro.clave if d.centro else None,
                    'numero_lote': d.numero_lote,
                    'cantidad_actual': d.cantidad_actual,
                    'cantidad_inicial': d.cantidad_inicial,
                    'fecha_caducidad': d.fecha_caducidad,
                    'estado': d.estado,
                    'created_at': d.created_at
                })
            
            return Response({
                'lote_farmacia': {
                    'id': lote.id,
                    'numero_lote': lote.numero_lote,
                    'producto_clave': lote.producto.clave,
                    'producto_descripcion': lote.producto.descripcion,
                    'cantidad_actual': lote.cantidad_actual,
                    'fecha_caducidad': lote.fecha_caducidad
                },
                'resumen': {
                    'total_centros_con_stock': total_derivados,
                    'stock_total_en_centros': stock_total_centros,
                    'stock_farmacia': lote.cantidad_actual,
                    'stock_total_sistema': lote.cantidad_actual + stock_total_centros
                },
                'lotes_derivados': derivados_data
            })
            
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            return Response({
                'error': 'Error al obtener lotes derivados',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['get'], url_path='trazabilidad')
    def trazabilidad_lote(self, request, pk=None):
        """
        Obtiene la trazabilidad completa de un lote:
        - Si es lote de farmacia: muestra derivados en centros
        - Si es lote de centro: muestra origen en farmacia
        """
        try:
            lote = self.get_object()
            
            result = {
                'lote': {
                    'id': lote.id,
                    'numero_lote': lote.numero_lote,
                    'producto_clave': lote.producto.clave,
                    'producto_descripcion': lote.producto.descripcion,
                    'cantidad_actual': lote.cantidad_actual,
                    'fecha_caducidad': lote.fecha_caducidad,
                    'es_lote_farmacia': lote.centro is None,
                    'ubicacion': lote.centro.nombre if lote.centro else 'Farmacia Central'
                },
                'origen': None,
                'derivados': []
            }
            
            # Si es lote de centro, mostrar origen
            if lote.lote_origen:
                result['origen'] = {
                    'id': lote.lote_origen.id,
                    'numero_lote': lote.lote_origen.numero_lote,
                    'cantidad_actual': lote.lote_origen.cantidad_actual,
                    'ubicacion': 'Farmacia Central'
                }
            
            # Si es lote de farmacia, mostrar derivados
            if lote.centro is None:
                derivados = Lote.objects.filter(
                    lote_origen=lote,
                    deleted_at__isnull=True
                ).select_related('centro')
                
                for d in derivados:
                    result['derivados'].append({
                        'id': d.id,
                        'centro_nombre': d.centro.nombre if d.centro else None,
                        'centro_clave': d.centro.clave if d.centro else None,
                        'cantidad_actual': d.cantidad_actual
                    })
            
            # Movimientos relacionados
            movimientos = Movimiento.objects.filter(
                lote=lote
            ).select_related('requisicion', 'usuario', 'centro').order_by('-fecha')[:20]
            
            result['movimientos'] = [{
                'id': m.id,
                'tipo': m.tipo,
                'cantidad': m.cantidad,
                'fecha': m.fecha,
                'requisicion_folio': m.requisicion.folio if m.requisicion else None,
                'observaciones': m.observaciones
            } for m in movimientos]
            
            return Response(result)
            
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            return Response({
                'error': 'Error al obtener trazabilidad',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['post'], url_path='subir-documento')
    def subir_documento(self, request, pk=None):
        """
        Sube un documento PDF asociado al lote.
        
        El documento puede ser factura, remisi�n, contrato, etc.
        
        PERMISOS:
        - Solo usuarios de farmacia/admin pueden subir documentos
        
        DATOS REQUERIDOS (multipart/form-data):
        - documento: Archivo PDF
        - nombre (opcional): Nombre descriptivo del documento
        """
        from django.core.files.uploadhandler import MemoryFileUploadHandler
        
        lote = self.get_object()
        
        # PERMISOS: Solo farmacia/admin pueden subir documentos
        user = request.user
        if not user.is_superuser and not is_farmacia_or_admin(user):
            return Response({
                'error': 'Solo el personal de farmacia puede subir documentos'
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Validar que se envi� un archivo
        documento = request.FILES.get('documento')
        if not documento:
            return Response({
                'error': 'No se envi� ning�n documento'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Validar tipo de archivo (solo PDF)
        if not documento.name.lower().endswith('.pdf'):
            return Response({
                'error': 'Solo se permiten archivos PDF'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Validar tama�o (m�ximo 10MB)
        if documento.size > 10 * 1024 * 1024:
            return Response({
                'error': 'El archivo no puede superar los 10MB'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Obtener nombre descriptivo
        nombre_documento = request.data.get('nombre', documento.name)
        
        # Guardar documento
        lote.documento_pdf = documento
        lote.documento_nombre = nombre_documento
        lote.save(update_fields=['documento_pdf', 'documento_nombre'])
        
        return Response({
            'mensaje': 'Documento subido correctamente',
            'documento_nombre': lote.documento_nombre,
            'documento_url': request.build_absolute_uri(lote.documento_pdf.url) if lote.documento_pdf else None
        })

    @action(detail=True, methods=['delete'], url_path='eliminar-documento')
    def eliminar_documento(self, request, pk=None):
        """
        Elimina el documento PDF asociado al lote.
        
        PERMISOS:
        - Solo usuarios de farmacia/admin pueden eliminar documentos
        """
        lote = self.get_object()
        
        # PERMISOS: Solo farmacia/admin pueden eliminar documentos
        user = request.user
        if not user.is_superuser and not is_farmacia_or_admin(user):
            return Response({
                'error': 'Solo el personal de farmacia puede eliminar documentos'
            }, status=status.HTTP_403_FORBIDDEN)
        
        if not lote.documento_pdf:
            return Response({
                'error': 'El lote no tiene documento asociado'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Eliminar archivo f�sico
        lote.documento_pdf.delete(save=False)
        lote.documento_nombre = ''
        lote.save(update_fields=['documento_pdf', 'documento_nombre'])
        
        return Response({
            'mensaje': 'Documento eliminado correctamente'
        })


class MovimientoViewSet(
    mixins.ListModelMixin,
    mixins.RetrieveModelMixin,
    mixins.CreateModelMixin,
    viewsets.GenericViewSet
):
    """
    ViewSet para gestionar movimientos de inventario.
    
    PERMISOS:
    - Admin/Farmacia: acceso completo a todos los movimientos
    - Centro: puede VER y CREAR movimientos en sus propios lotes (salidas/ajustes)
    - Vista: solo lectura
    
    FILTROS (alineados con exportaci�n):
    - tipo: entrada/salida/ajuste
    - centro: ID del centro
    - producto: ID del producto
    - lote: ID del lote
    - fecha_inicio: YYYY-MM-DD
    - fecha_fin: YYYY-MM-DD
    - search: b�squeda en observaciones, n�mero de lote, producto
    
    Esto permite auditor�a completa de consumos en cada centro.
    """
    queryset = Movimiento.objects.select_related('lote__producto', 'centro', 'usuario').all()
    serializer_class = MovimientoSerializer
    permission_classes = [IsCentroRole]  # Centro puede operar en sus lotes
    pagination_class = CustomPagination
    http_method_names = ['get', 'post', 'head', 'options']

    def get_queryset(self):
        """
        Filtra movimientos segun parametros.
        
        Parametros (alineados con exportaci�n):
        - tipo: entrada/salida/ajuste
        - centro: ID del centro (solo admin/farmacia/vista)
        - producto: ID del producto
        - lote: ID del lote
        - fecha_inicio: fecha m�nima (YYYY-MM-DD)
        - fecha_fin: fecha m�xima (YYYY-MM-DD)
        - search: b�squeda en observaciones, lote, producto
        
        Seguridad: Usuarios de centro solo ven movimientos de su centro.
        Admin/farmacia/vista ven todo por defecto, pueden filtrar con ?centro=.
        """
        queryset = Movimiento.objects.select_related('lote__producto', 'centro', 'usuario')
        
        # SEGURIDAD: Filtrar por centro segun rol
        user = self.request.user
        if not is_farmacia_or_admin(user):
            # Usuario de centro: forzado a su centro
            user_centro = get_user_centro(user)
            if user_centro:
                queryset = queryset.filter(lote__centro=user_centro)
            else:
                return Movimiento.objects.none()
        else:
            # Admin/farmacia/vista: pueden filtrar por centro especifico
            centro_param = self.request.query_params.get('centro')
            if centro_param:
                queryset = queryset.filter(lote__centro_id=centro_param)
        
        # Filtro por tipo
        tipo = self.request.query_params.get('tipo')
        if tipo:
            queryset = queryset.filter(tipo=tipo.lower())
        
        # Filtro por producto
        producto = self.request.query_params.get('producto')
        if producto:
            queryset = queryset.filter(lote__producto_id=producto)
        
        # Filtro por lote
        lote = self.request.query_params.get('lote')
        if lote:
            queryset = queryset.filter(lote_id=lote)
        
        # Filtro por rango de fechas
        fecha_inicio = self.request.query_params.get('fecha_inicio')
        if fecha_inicio:
            queryset = queryset.filter(fecha__date__gte=fecha_inicio)
        
        fecha_fin = self.request.query_params.get('fecha_fin')
        if fecha_fin:
            queryset = queryset.filter(fecha__date__lte=fecha_fin)
        
        # B�squeda en observaciones, lote y producto
        search = self.request.query_params.get('search')
        if search and search.strip():
            search_term = search.strip()
            queryset = queryset.filter(
                Q(observaciones__icontains=search_term) |
                Q(lote__numero_lote__icontains=search_term) |
                Q(lote__producto__clave__icontains=search_term) |
                Q(lote__producto__descripcion__icontains=search_term)
            )
        
        return queryset.order_by('-fecha')

    def perform_create(self, serializer):
        """
        Crea un movimiento validando permisos por centro.
        
        SEGURIDAD:
        - Admin/farmacia: pueden crear cualquier movimiento en cualquier lote
        - Usuario de centro: solo pueden crear movimientos en lotes de su centro
          y solo ciertos tipos: 'salida' (consumo), 'ajuste' (inventario f�sico)
        - Usuario de centro NO puede crear 'entrada' (solo v�a surtido de requisici�n)
        """
        user = self.request.user
        lote = serializer.validated_data.get('lote')
        tipo = serializer.validated_data.get('tipo', '').lower()
        
        # Validar que usuario de centro solo opere con sus lotes
        if not is_farmacia_or_admin(user):
            user_centro = get_user_centro(user)
            
            # Validar que el lote pertenece al centro del usuario
            if lote and lote.centro != user_centro:
                raise serializers.ValidationError({
                    'lote': 'Solo puedes registrar movimientos en lotes de tu centro'
                })
            
            # Validar tipos de movimiento permitidos para centros
            # Centros pueden: salida (consumo), ajuste (inventario f�sico)
            # Centros NO pueden: entrada (solo v�a surtido autom�tico)
            tipos_permitidos_centro = ['salida', 'ajuste']
            if tipo not in tipos_permitidos_centro:
                raise serializers.ValidationError({
                    'tipo': f'Los centros solo pueden registrar: {", ".join(tipos_permitidos_centro)}. Las entradas se generan autom�ticamente al surtir requisiciones.'
                })
        
        movimiento, _ = registrar_movimiento_stock(
            lote=lote,
            tipo=serializer.validated_data.get('tipo'),
            cantidad=serializer.validated_data.get('cantidad'),
            usuario=user,
            centro=serializer.validated_data.get('centro') or (lote.centro if lote else None),
            requisicion=serializer.validated_data.get('requisicion'),
            observaciones=serializer.validated_data.get('observaciones')
        )
        # Dejar instancia lista para serializer.data
        serializer.instance = movimiento

    @action(detail=False, methods=['get'], url_path='trazabilidad-pdf')
    def trazabilidad_pdf(self, request):
        """
        Genera PDF de trazabilidad de un producto.
        Par�metros: ?producto_clave=XXX
        """
        from core.utils.pdf_reports import generar_reporte_trazabilidad
        
        clave = request.query_params.get('producto_clave')
        if not clave:
            return Response({'error': 'Se requiere producto_clave'}, status=status.HTTP_400_BAD_REQUEST)
        
        producto = Producto.objects.filter(clave__iexact=clave).first()
        if not producto:
            return Response({'error': 'Producto no encontrado'}, status=status.HTTP_404_NOT_FOUND)
        
        try:
            # Obtener movimientos del producto
            movimientos = Movimiento.objects.filter(
                lote__producto=producto
            ).select_related('lote', 'centro', 'usuario').order_by('-fecha')[:100]
            
            trazabilidad_data = []
            for mov in movimientos:
                trazabilidad_data.append({
                    'fecha': mov.fecha.strftime('%d/%m/%Y %H:%M') if mov.fecha else 'N/A',
                    'tipo': mov.tipo.upper(),
                    'lote': mov.lote.numero_lote if mov.lote else 'N/A',
                    'cantidad': mov.cantidad,
                    'centro': mov.centro.nombre if mov.centro else 'Farmacia Central',
                    'usuario': mov.usuario.get_full_name() if mov.usuario else 'Sistema',
                    'observaciones': mov.observaciones or ''
                })
            
            producto_info = {
                'clave': producto.clave,
                'descripcion': producto.descripcion,
                'unidad_medida': producto.unidad_medida,
                'stock_actual': producto.get_stock_actual() if hasattr(producto, 'get_stock_actual') else 0,
                'stock_minimo': producto.stock_minimo,
                'precio_unitario': float(producto.precio_unitario) if producto.precio_unitario else 0,
            }
            
            pdf_buffer = generar_reporte_trazabilidad(trazabilidad_data, producto_info=producto_info)
            
            response = HttpResponse(
                pdf_buffer.getvalue(),
                content_type='application/pdf'
            )
            response['Content-Disposition'] = f'attachment; filename="Trazabilidad_{clave}_{timezone.now().strftime("%Y%m%d")}.pdf"'
            
            return response
            
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            return Response({
                'error': 'Error al generar PDF de trazabilidad',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path='trazabilidad-lote-pdf')
    def trazabilidad_lote_pdf(self, request):
        """
        Genera PDF de trazabilidad de un lote espec�fico.
        Par�metros: ?numero_lote=XXX
        """
        from core.utils.pdf_reports import generar_reporte_trazabilidad
        
        numero_lote = request.query_params.get('numero_lote')
        if not numero_lote:
            return Response({'error': 'Se requiere numero_lote'}, status=status.HTTP_400_BAD_REQUEST)
        
        lote = Lote.objects.filter(numero_lote__iexact=numero_lote).select_related('producto', 'centro').first()
        if not lote:
            return Response({'error': 'Lote no encontrado'}, status=status.HTTP_404_NOT_FOUND)
        
        try:
            # Obtener movimientos del lote
            movimientos = Movimiento.objects.filter(
                lote=lote
            ).select_related('lote', 'centro', 'usuario').order_by('-fecha')[:100]
            
            trazabilidad_data = []
            for mov in movimientos:
                trazabilidad_data.append({
                    'fecha': mov.fecha.strftime('%d/%m/%Y %H:%M') if mov.fecha else 'N/A',
                    'tipo': mov.tipo.upper(),
                    'lote': mov.lote.numero_lote if mov.lote else 'N/A',
                    'cantidad': mov.cantidad,
                    'centro': mov.centro.nombre if mov.centro else 'Farmacia Central',
                    'usuario': mov.usuario.get_full_name() if mov.usuario else 'Sistema',
                    'observaciones': mov.observaciones or ''
                })
            
            producto_info = {
                'clave': lote.producto.clave if lote.producto else 'N/A',
                'descripcion': lote.producto.descripcion if lote.producto else 'N/A',
                'unidad_medida': lote.producto.unidad_medida if lote.producto else 'N/A',
                'stock_actual': lote.cantidad_actual,
                'stock_minimo': lote.producto.stock_minimo if lote.producto else 0,
                'numero_lote': lote.numero_lote,
                'fecha_caducidad': lote.fecha_caducidad.strftime('%d/%m/%Y') if lote.fecha_caducidad else 'N/A',
                'proveedor': lote.proveedor or 'No especificado',
            }
            
            pdf_buffer = generar_reporte_trazabilidad(trazabilidad_data, producto_info=producto_info)
            
            response = HttpResponse(
                pdf_buffer.getvalue(),
                content_type='application/pdf'
            )
            response['Content-Disposition'] = f'attachment; filename="Trazabilidad_Lote_{numero_lote}_{timezone.now().strftime("%Y%m%d")}.pdf"'
            
            return response
            
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            return Response({
                'error': 'Error al generar PDF de trazabilidad del lote',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path='exportar-pdf')
    def exportar_pdf(self, request):
        """
        Genera PDF de movimientos con filtros opcionales.
        """
        from core.utils.pdf_reports import generar_reporte_movimientos
        
        try:
            # Aplicar filtros
            queryset = self.get_queryset()
            
            tipo = request.query_params.get('tipo')
            if tipo:
                queryset = queryset.filter(tipo=tipo.lower())
            
            fecha_inicio = request.query_params.get('fecha_inicio')
            if fecha_inicio:
                queryset = queryset.filter(fecha__gte=fecha_inicio)
            
            fecha_fin = request.query_params.get('fecha_fin')
            if fecha_fin:
                queryset = queryset.filter(fecha__lte=fecha_fin)
            
            movimientos = queryset[:200]  # Limitar para PDF
            
            movimientos_data = []
            for mov in movimientos:
                movimientos_data.append({
                    'fecha': mov.fecha.strftime('%d/%m/%Y %H:%M') if mov.fecha else 'N/A',
                    'tipo': mov.tipo.upper(),
                    'producto': mov.lote.producto.clave if mov.lote and mov.lote.producto else 'N/A',
                    'lote': mov.lote.numero_lote if mov.lote else 'N/A',
                    'cantidad': mov.cantidad,
                    'centro': mov.centro.nombre if mov.centro else 'Farmacia Central',
                    'usuario': mov.usuario.get_full_name() if mov.usuario else 'Sistema',
                })
            
            pdf_buffer = generar_reporte_movimientos(movimientos_data)
            
            response = HttpResponse(
                pdf_buffer.getvalue(),
                content_type='application/pdf'
            )
            response['Content-Disposition'] = f'attachment; filename="Movimientos_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
            
            return response
            
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            return Response({
                'error': 'Error al generar PDF de movimientos',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path='exportar-excel')
    def exportar_excel(self, request):
        """
        Genera Excel de movimientos con filtros opcionales.
        """
        try:
            # Aplicar filtros
            queryset = self.get_queryset()
            
            tipo = request.query_params.get('tipo')
            if tipo:
                queryset = queryset.filter(tipo=tipo.lower())
            
            fecha_inicio = request.query_params.get('fecha_inicio')
            if fecha_inicio:
                queryset = queryset.filter(fecha__gte=fecha_inicio)
            
            fecha_fin = request.query_params.get('fecha_fin')
            if fecha_fin:
                queryset = queryset.filter(fecha__lte=fecha_fin)
            
            producto = request.query_params.get('producto')
            if producto:
                queryset = queryset.filter(lote__producto_id=producto)
            
            centro = request.query_params.get('centro')
            if centro:
                queryset = queryset.filter(Q(centro_id=centro) | Q(lote__centro_id=centro))
            
            search = request.query_params.get('search')
            if search:
                queryset = queryset.filter(
                    Q(lote__numero_lote__icontains=search) |
                    Q(lote__producto__descripcion__icontains=search) |
                    Q(observaciones__icontains=search)
                )
            
            movimientos = queryset[:1000]  # Limitar para Excel
            
            # Crear libro de Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Movimientos'
            
            # T�tulo
            ws.merge_cells('A1:H1')
            ws['A1'] = 'REPORTE DE MOVIMIENTOS'
            ws['A1'].font = Font(bold=True, size=14, color='632842')
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # Fecha
            ws.merge_cells('A2:H2')
            ws['A2'] = f'Generado el {timezone.now().strftime("%d/%m/%Y %H:%M")}'
            ws['A2'].alignment = Alignment(horizontal='center')
            
            # Encabezados
            headers = ['#', 'Fecha', 'Tipo', 'Producto', 'Lote', 'Cantidad', 'Centro', 'Usuario']
            ws.append([])
            ws.append(headers)
            
            header_fill = PatternFill(start_color='632842', end_color='632842', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            for cell in ws[4]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Datos
            for idx, mov in enumerate(movimientos, 1):
                ws.append([
                    idx,
                    mov.fecha.strftime('%d/%m/%Y %H:%M') if mov.fecha else 'N/A',
                    mov.tipo.upper(),
                    mov.lote.producto.descripcion[:50] if mov.lote and mov.lote.producto else 'N/A',
                    mov.lote.numero_lote if mov.lote else 'N/A',
                    mov.cantidad,
                    mov.centro.nombre if mov.centro else (mov.lote.centro.nombre if mov.lote and mov.lote.centro else 'Farmacia Central'),
                    mov.usuario.get_full_name() or mov.usuario.username if mov.usuario else 'Sistema',
                ])
            
            # Ajustar anchos
            column_widths = [8, 18, 12, 50, 15, 12, 25, 20]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
            
            # Respuesta
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="Movimientos_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
            wb.save(response)
            return response
            
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            return Response({
                'error': 'Error al generar Excel de movimientos',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class RequisicionViewSet(viewsets.ModelViewSet):
    """
    CRUD y flujo de requisiciones con estados en minusculas
    (borrador -> enviada -> autorizada/parcial -> surtida o rechazada/cancelada).
    Las validaciones se alinean con los campos reales del modelo y la UI.
    """
    queryset = Requisicion.objects.select_related('centro', 'usuario_solicita', 'usuario_autoriza').prefetch_related('detalles__producto').all()
    serializer_class = RequisicionSerializer
    permission_classes = [IsCentroRole]
    pagination_class = CustomPagination

    def _user_centro(self, user):
        return getattr(user, 'centro', None) or getattr(getattr(user, 'profile', None), 'centro', None)

    def _validar_stock_items(self, items, centro=None):
        """
        Valida stock disponible. Si se proporciona centro, se usa stock por centro
        calculado a partir de movimientos (entradas - salidas); de lo contrario usa stock global.
        """
        errores = []
        for item_data in items:
            producto_id = item_data.get('producto')
            if not producto_id:
                continue
            producto = Producto.objects.filter(id=producto_id).first()
            if not producto:
                continue
            try:
                cantidad = int(item_data.get('cantidad_autorizada') or item_data.get('cantidad_solicitada') or 0)
            except (TypeError, ValueError):
                continue

            disponible = producto.get_stock_actual()
            if centro:
                # Intentar calcular stock por centro desde lotes asociados al centro o con movimientos del centro
                lote_ids = Movimiento.objects.filter(centro=centro, lote__producto=producto).values_list('lote_id', flat=True)
                disponible_lotes = Lote.objects.filter(
                    Q(id__in=lote_ids) | Q(centro=centro),
                    estado='disponible',
                    deleted_at__isnull=True,
                    cantidad_actual__gt=0
                ).aggregate(total=Coalesce(Sum('cantidad_actual'), 0))['total'] or 0
                # Si no hay lotes asociados, caer al agregado de movimientos
                if disponible_lotes > 0:
                    disponible = disponible_lotes
                else:
                    disponible = Movimiento.objects.filter(
                        centro=centro,
                        lote__producto=producto
                    ).aggregate(total=Coalesce(Sum('cantidad'), 0))['total'] or 0

            if cantidad > disponible:
                errores.append({
                    'producto': producto.clave,
                    'disponible': disponible,
                    'solicitado': cantidad
                })
        return errores

    def get_queryset(self):
        queryset = Requisicion.objects.select_related('centro', 'usuario_solicita', 'usuario_autoriza').prefetch_related('detalles__producto')
        user = getattr(self.request, 'user', None)
        
        # SEGURIDAD: Filtrar por centro segun rol
        if user and not is_farmacia_or_admin(user):
            # Usuario de centro: forzado a su centro
            user_centro = self._user_centro(user)
            if user_centro:
                queryset = queryset.filter(centro=user_centro)
            else:
                return Requisicion.objects.none()
        else:
            # Admin/farmacia/vista: pueden filtrar por centro especifico
            centro_param = self.request.query_params.get('centro')
            if centro_param:
                queryset = queryset.filter(centro_id=centro_param)

        estado = self.request.query_params.get('estado')
        if estado:
            queryset = queryset.filter(estado=estado.lower())

        grupo = self.request.query_params.get('grupo_estado')
        if grupo and grupo in REQUISICION_GRUPOS_ESTADO:
            queryset = queryset.filter(estado__in=REQUISICION_GRUPOS_ESTADO[grupo])

        search = self.request.query_params.get('search')
        if search and search.strip():
            queryset = queryset.filter(folio__icontains=search.strip())

        # Filtros de fecha
        fecha_desde = self.request.query_params.get('fecha_desde')
        if fecha_desde:
            queryset = queryset.filter(fecha_solicitud__date__gte=fecha_desde)
        
        fecha_hasta = self.request.query_params.get('fecha_hasta')
        if fecha_hasta:
            queryset = queryset.filter(fecha_solicitud__date__lte=fecha_hasta)

        return queryset.order_by('-fecha_solicitud')
    
    def create(self, request, *args, **kwargs):
        """Crea requisicion en estado borrador/enviada segun data."""
        try:
            data = request.data.copy()
            fecha = timezone.now()
            folio = f"REQ-{fecha.strftime('%Y%m%d')}-{random.randint(1000, 9999)}"
            while Requisicion.objects.filter(folio=folio).exists():
                folio = f"REQ-{fecha.strftime('%Y%m%d')}-{random.randint(1000, 9999)}"

            data.setdefault('estado', 'borrador')
            data['estado'] = str(data.get('estado')).lower()
            data['folio'] = folio

            solicitante = request.user if request.user.is_authenticated else None
            es_privilegiado = False
            if solicitante:
                data['usuario_solicita'] = getattr(solicitante, 'id', None)
                centro_user = self._user_centro(solicitante)
                es_privilegiado = is_farmacia_or_admin(solicitante)
                if centro_user and not es_privilegiado:
                    if data.get('centro') and int(data.get('centro')) != centro_user.id and not solicitante.is_superuser:
                        return Response({'error': 'No puedes crear requisiciones para otro centro'}, status=status.HTTP_403_FORBIDDEN)
                    data['centro'] = centro_user.id
                elif not centro_user and not es_privilegiado and not solicitante.is_superuser:
                    return Response({'error': 'El usuario no tiene centro asignado'}, status=status.HTTP_403_FORBIDDEN)
                elif not data.get('centro') and centro_user:
                    data['centro'] = centro_user.id

            items_data = request.data.get('items', []) or request.data.get('detalles', []) or []
            centro = Centro.objects.filter(id=data.get('centro')).first() if data.get('centro') else None
            if not centro and solicitante and not solicitante.is_superuser:
                return Response({'error': 'No se encontro el centro para validar stock'}, status=status.HTTP_400_BAD_REQUEST)
            # Para requisiciones, validar stock global (no del centro solicitante)
            # El centro solicitante pide stock, no lo provee
            errores_stock = []
            if es_privilegiado:
                errores_stock = self._validar_stock_items(items_data, centro=None)
                if errores_stock:
                    return Response({'error': 'Stock insuficiente', 'detalles': errores_stock}, status=status.HTTP_400_BAD_REQUEST)
            else:
                product_ids = [item.get('producto') for item in items_data if item.get('producto')]
                if product_ids and Lote.objects.filter(producto_id__in=product_ids, deleted_at__isnull=True).exists():
                    errores_stock = self._validar_stock_items(items_data, centro=centro)
                    if errores_stock:
                        return Response({'error': 'Stock insuficiente', 'detalles': errores_stock}, status=status.HTTP_400_BAD_REQUEST)

            # Agregar detalles a data para que el serializer los procese
            data['detalles'] = items_data
            
            serializer = self.get_serializer(data=data)
            serializer.is_valid(raise_exception=True)
            requisicion = serializer.save()

            return Response({
                'mensaje': 'Requisicion creada exitosamente',
                'requisicion': RequisicionSerializer(requisicion).data
            }, status=status.HTTP_201_CREATED)
        except serializers.ValidationError as exc:
            return Response({'error': 'Error de validacion', 'detalles': exc.detail}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as exc:
            # traceback removido por seguridad (ISS-008)
            return Response({'error': 'Error al crear requisicion', 'mensaje': str(exc)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def update(self, request, *args, **kwargs):
        """Solo permite editar si sigue en borrador."""
        requisicion = self.get_object()
        if (requisicion.estado or '').lower() != 'borrador':
            return Response({'error': 'Solo se pueden editar requisiciones en estado BORRADOR', 'estado_actual': requisicion.estado}, status=status.HTTP_400_BAD_REQUEST)

        centro_user = self._user_centro(request.user)
        if not request.user.is_superuser and centro_user and requisicion.centro_id != centro_user.id:
            return Response({'error': 'No puedes editar requisiciones de otro centro'}, status=status.HTTP_403_FORBIDDEN)

        partial = kwargs.pop('partial', False)
        serializer = self.get_serializer(requisicion, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        requisicion = serializer.save()

        items_data = request.data.get('items') or request.data.get('detalles') or []
        if items_data:
            errores_stock = self._validar_stock_items(items_data, centro=requisicion.centro)
            if errores_stock:
                return Response({'error': 'Stock insuficiente', 'detalles': errores_stock}, status=status.HTTP_400_BAD_REQUEST)
            requisicion.detalles.all().delete()
            for item_data in items_data:
                producto_id = item_data.get('producto')
                cant = item_data.get('cantidad_solicitada')
                if not producto_id or cant in [None, '']:
                    continue
                DetalleRequisicion.objects.create(
                    requisicion=requisicion,
                    producto_id=producto_id,
                    cantidad_solicitada=int(cant),
                    cantidad_autorizada=int(item_data.get('cantidad_autorizada') or 0),
                    observaciones=item_data.get('observaciones', '')
                )

        return Response({'mensaje': 'Requisicion actualizada exitosamente', 'requisicion': RequisicionSerializer(requisicion).data})
    
    def destroy(self, request, *args, **kwargs):
        requisicion = self.get_object()
        if (requisicion.estado or '').lower() != 'borrador':
            return Response({'error': 'Solo se pueden eliminar requisiciones en estado BORRADOR', 'estado_actual': requisicion.estado}, status=status.HTTP_400_BAD_REQUEST)
        centro_user = self._user_centro(request.user)
        if not request.user.is_superuser:
            if not centro_user or requisicion.centro_id != centro_user.id:
                return Response({'error': 'No puedes eliminar requisiciones de otro centro'}, status=status.HTTP_403_FORBIDDEN)
        folio = requisicion.folio
        requisicion.delete()
        return Response({'mensaje': 'Requisicion eliminada', 'folio_eliminado': folio}, status=status.HTTP_204_NO_CONTENT)

    @action(detail=True, methods=['post'])
    def enviar(self, request, pk=None):
        requisicion = self.get_object()
        if (requisicion.estado or '').lower() != 'borrador':
            return Response({'error': 'Solo se pueden enviar requisiciones en estado BORRADOR', 'estado_actual': requisicion.estado}, status=status.HTTP_400_BAD_REQUEST)
        centro_user = self._user_centro(request.user)
        es_privilegiado = is_farmacia_or_admin(request.user)
        if not request.user.is_superuser and not es_privilegiado:
            if not centro_user or requisicion.centro_id != centro_user.id:
                return Response({'error': 'No puedes enviar requisiciones de otro centro'}, status=status.HTTP_403_FORBIDDEN)
        if not requisicion.detalles.exists():
            return Response({'error': 'La requisicion debe tener al menos un producto'}, status=status.HTTP_400_BAD_REQUEST)
        requisicion.estado = 'enviada'
        requisicion.save(update_fields=['estado'])
        return Response({'mensaje': 'Requisicion enviada', 'requisicion': RequisicionSerializer(requisicion).data})

    @action(detail=True, methods=['post'])
    def autorizar(self, request, pk=None):
        requisicion = self.get_object()
        if (requisicion.estado or '').lower() != 'enviada':
            return Response({'error': 'Solo se pueden autorizar requisiciones en estado ENVIADA', 'estado_actual': requisicion.estado}, status=status.HTTP_400_BAD_REQUEST)

        centro_user = self._user_centro(request.user)
        es_privilegiado = is_farmacia_or_admin(request.user)
        if not request.user.is_superuser and not es_privilegiado:
            if not centro_user:
                return Response({'error': 'El usuario no tiene centro asignado'}, status=status.HTTP_403_FORBIDDEN)
            if requisicion.centro_id != centro_user.id:
                return Response({'error': 'No puedes autorizar requisiciones de otro centro'}, status=status.HTTP_403_FORBIDDEN)
        rol = (getattr(request.user, 'rol', '') or '').lower()
        if not request.user.is_superuser and rol not in ['admin_sistema', 'farmacia', 'admin_farmacia', 'superusuario'] and not request.user.is_staff:
            return Response({'error': 'No tienes permiso para autorizar requisiciones'}, status=status.HTTP_403_FORBIDDEN)

        items_data = request.data.get('items') or request.data.get('detalles') or []
        for item_data in items_data:
            item_id = item_data.get('id')
            cant = item_data.get('cantidad_autorizada')
            if item_id is None or cant is None:
                continue
            try:
                item = requisicion.detalles.get(id=item_id)
            except DetalleRequisicion.DoesNotExist:
                continue
            item.cantidad_autorizada = max(0, int(cant))
            item.save()

        requisicion.estado = 'autorizada'
        requisicion.fecha_autorizacion = timezone.now()
        if request.user and request.user.is_authenticated:
            requisicion.usuario_autoriza = request.user
        requisicion.save(update_fields=['estado', 'fecha_autorizacion', 'usuario_autoriza'])

        return Response({'mensaje': 'Requisicion autorizada', 'requisicion': RequisicionSerializer(requisicion).data})

    @action(detail=True, methods=['post'])
    def rechazar(self, request, pk=None):
        requisicion = self.get_object()
        if (requisicion.estado or '').lower() != 'enviada':
            return Response({'error': 'Solo se pueden rechazar requisiciones en estado ENVIADA', 'estado_actual': requisicion.estado}, status=status.HTTP_400_BAD_REQUEST)
        centro_user = self._user_centro(request.user)
        es_privilegiado = is_farmacia_or_admin(request.user)
        if not request.user.is_superuser and not es_privilegiado:
            if not centro_user or requisicion.centro_id != centro_user.id:
                return Response({'error': 'No puedes rechazar requisiciones de otro centro'}, status=status.HTTP_403_FORBIDDEN)
        motivo = request.data.get('observaciones') or request.data.get('comentario') or ''
        if not motivo.strip():
            return Response({'error': 'Debe proporcionar un motivo de rechazo'}, status=status.HTTP_400_BAD_REQUEST)
        requisicion.estado = 'rechazada'
        requisicion.motivo_rechazo = motivo
        requisicion.observaciones = motivo
        requisicion.save(update_fields=['estado', 'motivo_rechazo', 'observaciones'])
        return Response({'mensaje': 'Requisicion rechazada', 'requisicion': RequisicionSerializer(requisicion).data})

    @action(detail=True, methods=['post'])
    def cancelar(self, request, pk=None):
        requisicion = self.get_object()
        estado_actual = (requisicion.estado or '').lower()
        if estado_actual in ['surtida', 'cancelada', 'rechazada']:
            return Response({'error': f'No se puede cancelar una requisicion en estado {requisicion.estado}'}, status=status.HTTP_400_BAD_REQUEST)
        centro_user = self._user_centro(request.user)
        if not request.user.is_superuser:
            if not centro_user or requisicion.centro_id != centro_user.id:
                return Response({'error': 'No puedes cancelar requisiciones de otro centro'}, status=status.HTTP_403_FORBIDDEN)
        requisicion.estado = 'cancelada'
        motivo = request.data.get('observaciones') or request.data.get('comentario')
        if motivo:
            requisicion.observaciones = motivo
        requisicion.save(update_fields=['estado', 'observaciones'])
        return Response({'mensaje': 'Requisicion cancelada', 'requisicion': RequisicionSerializer(requisicion).data})

    @action(detail=True, methods=['post'])
    def surtir(self, request, pk=None):
        """
        Surte una requisici�n autorizada.
        
        PERMISOS:
        - Superuser: puede surtir cualquier requisici�n
        - Farmacia (admin_farmacia, farmacia): puede surtir cualquier requisici�n
        - Centro: solo puede surtir requisiciones de su propio centro
        
        L�GICA DE STOCK:
        - Primero usa lotes de farmacia central (centro=NULL) ? crea entrada en centro destino
        - Si no hay en farmacia central, usa lotes del centro solicitante (salida interna)
        """
        requisicion = self.get_object()
        estado_actual = (requisicion.estado or '').lower()
        if estado_actual not in ['autorizada', 'parcial']:
            return Response({'error': 'Solo se pueden surtir requisiciones autorizadas'}, status=status.HTTP_400_BAD_REQUEST)
        
        # PERMISOS: Farmacia/admin puede surtir cualquier requisici�n
        user = request.user
        if not user.is_superuser and not is_farmacia_or_admin(user):
            # Usuario de centro: solo puede surtir su propio centro
            centro_user = self._user_centro(user)
            if not centro_user or requisicion.centro_id != centro_user.id:
                return Response({'error': 'No puedes surtir requisiciones de otro centro'}, status=status.HTTP_403_FORBIDDEN)

        # SEGURIDAD: Solo usar lotes de farmacia central (centro=None) o del centro solicitante
        # Esto evita descuentos de stock entre centros
        centro_requisicion = requisicion.centro
        
        errores_stock = []
        for detalle in requisicion.detalles.select_related('producto'):
            requerido = (detalle.cantidad_autorizada or detalle.cantidad_solicitada) - (detalle.cantidad_surtida or 0)
            if requerido <= 0:
                continue
            # Lotes de farmacia central (centro=None) o del centro de destino
            disponible = Lote.objects.filter(
                Q(centro__isnull=True) | Q(centro=centro_requisicion),
                producto=detalle.producto,
                estado='disponible',
                deleted_at__isnull=True,
                cantidad_actual__gt=0
            ).aggregate(total=Sum('cantidad_actual'))['total'] or 0
            if disponible < requerido:
                errores_stock.append({
                    'producto': detalle.producto.clave,
                    'requerido': requerido,
                    'disponible': disponible
                })
        if errores_stock:
            return Response({'error': 'Stock insuficiente para surtir', 'detalles': errores_stock}, status=status.HTTP_400_BAD_REQUEST)

        with transaction.atomic():
            for detalle in requisicion.detalles.select_related('producto'):
                pendiente = (detalle.cantidad_autorizada or detalle.cantidad_solicitada) - (detalle.cantidad_surtida or 0)
                if pendiente <= 0:
                    continue
                # FEFO: Lotes de farmacia central o del centro, ordenados por caducidad
                lotes = Lote.objects.filter(
                    Q(centro__isnull=True) | Q(centro=centro_requisicion),
                    producto=detalle.producto,
                    estado='disponible',
                    deleted_at__isnull=True,
                    cantidad_actual__gt=0
                ).order_by('fecha_caducidad', 'id')

                for lote in lotes:
                    if pendiente <= 0:
                        break
                    usar = min(pendiente, lote.cantidad_actual)
                    
                    # PASO 1: Registrar SALIDA del lote origen (farmacia central o mismo centro)
                    movimiento_salida, lote_actualizado = registrar_movimiento_stock(
                        lote=lote,
                        tipo='salida',
                        cantidad=usar,
                        centro=lote.centro,  # Centro del lote origen
                        usuario=request.user if request.user.is_authenticated else None,
                        requisicion=requisicion,
                        observaciones=f'SALIDA_POR_REQUISICION {requisicion.folio}'
                    )
                    
                    # PASO 2: Si el lote era de farmacia central (centro=None), crear ENTRADA en centro destino
                    # REGLA DE ORO: El lote en el centro es una COPIA FIEL del lote de farmacia
                    # - Mismo producto
                    # - Mismo n�mero de lote (SIN modificar)
                    # - Misma fecha de caducidad
                    # - Vinculado a lote_origen (trazabilidad obligatoria)
                    if lote.centro is None and centro_requisicion:
                        # Buscar lote existente en el centro destino con MISMO n�mero de lote
                        lote_destino = Lote.objects.filter(
                            producto=detalle.producto,
                            numero_lote=lote.numero_lote,  # MISMO n�mero de lote
                            centro=centro_requisicion,
                            deleted_at__isnull=True
                        ).first()
                        
                        if lote_destino:
                            # Si existe y est� agotado, reactivarlo
                            if lote_destino.estado == 'agotado':
                                lote_destino.estado = 'disponible'
                                lote_destino.save(update_fields=['estado'])
                        else:
                            # Crear nuevo lote: COPIA FIEL del lote de farmacia
                            # El n�mero de lote es ID�NTICO al de farmacia
                            lote_destino = Lote(
                                producto=detalle.producto,
                                numero_lote=lote.numero_lote,  # MISMO n�mero de lote
                                centro=centro_requisicion,
                                fecha_caducidad=lote.fecha_caducidad,  # MISMA caducidad
                                cantidad_inicial=usar,
                                cantidad_actual=0,  # Se actualiza con el movimiento
                                estado='disponible',
                                precio_compra=lote.precio_compra,
                                proveedor=lote.proveedor,
                                lote_origen=lote,  # VINCULACI�N OBLIGATORIA al lote de farmacia
                                observaciones=f'Transferido de farmacia central via requisicion {requisicion.folio}'
                            )
                            # Guardar sin validaci�n full_clean para permitir la creaci�n
                            lote_destino.save()
                        
                        # Registrar ENTRADA en el centro destino
                        # IMPORTANTE: Asociar la requisici�n al movimiento de entrada para trazabilidad completa
                        movimiento_entrada, _ = registrar_movimiento_stock(
                            lote=lote_destino,
                            tipo='entrada',
                            cantidad=usar,
                            centro=centro_requisicion,
                            usuario=request.user if request.user.is_authenticated else None,
                            requisicion=None,  # La requisici�n solo se registra en la salida para no duplicar conteos
                            observaciones=f'ENTRADA_POR_REQUISICION {requisicion.folio}'
                        )
                    
                    detalle.cantidad_surtida = (detalle.cantidad_surtida or 0) + usar
                    detalle.save(update_fields=['cantidad_surtida'])
                    pendiente -= usar

            detalles_actualizados = Requisicion.objects.get(pk=requisicion.pk).detalles.all()
            completada = all(
                (d.cantidad_autorizada or d.cantidad_solicitada) <= (d.cantidad_surtida or 0)
                for d in detalles_actualizados
            )
            requisicion.estado = 'surtida' if completada else 'parcial'
            requisicion.save(update_fields=['estado'])

        return Response({'mensaje': 'Requisicion surtida', 'requisicion': RequisicionSerializer(requisicion).data})

    @action(detail=True, methods=['post'], url_path='marcar-recibida')
    def marcar_recibida(self, request, pk=None):
        """
        Marca una requisici�n surtida como recibida por el centro.
        
        PERMISOS:
        - Solo usuarios del centro receptor pueden marcar como recibida
        - Solo requisiciones en estado 'surtida' pueden marcarse
        
        DATOS REQUERIDOS:
        - lugar_entrega: Lugar donde se recibi�
        - observaciones_recepcion: Observaciones de la recepci�n (opcional)
        """
        from django.utils import timezone
        
        requisicion = self.get_object()
        estado_actual = (requisicion.estado or '').lower()
        
        if estado_actual != 'surtida':
            return Response({
                'error': 'Solo se pueden marcar como recibidas las requisiciones surtidas',
                'estado_actual': requisicion.estado
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # PERMISOS: Solo usuarios del centro receptor pueden confirmar recepci�n
        user = request.user
        if not user.is_superuser:
            centro_user = self._user_centro(user)
            if not centro_user or requisicion.centro_id != centro_user.id:
                return Response({
                    'error': 'Solo el centro receptor puede confirmar la recepci�n'
                }, status=status.HTTP_403_FORBIDDEN)
        
        # Obtener datos del request
        lugar_entrega = request.data.get('lugar_entrega', '')
        observaciones_recepcion = request.data.get('observaciones_recepcion', '')
        
        if not lugar_entrega:
            return Response({
                'error': 'El lugar de entrega es requerido'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Actualizar la requisici�n
        requisicion.estado = 'recibida'
        requisicion.fecha_recibido = timezone.now()
        requisicion.usuario_recibe = user
        requisicion.lugar_entrega = lugar_entrega
        requisicion.observaciones_recepcion = observaciones_recepcion
        requisicion.save(update_fields=[
            'estado', 'fecha_recibido', 'usuario_recibe', 
            'lugar_entrega', 'observaciones_recepcion'
        ])
        
        return Response({
            'mensaje': 'Requisici�n marcada como recibida',
            'requisicion': RequisicionSerializer(requisicion).data
        })

    @action(detail=True, methods=['get'], url_path='hoja-recoleccion')
    def hoja_recoleccion(self, request, pk=None):
        """
        Genera y descarga el PDF de la hoja de recolecci�n para una requisici�n.
        Solo disponible para requisiciones autorizadas, parciales o surtidas.
        """
        from core.utils.pdf_generator import generar_hoja_recoleccion
        
        requisicion = self.get_object()
        estado = (requisicion.estado or '').lower()
        
        # Validar que la requisici�n puede generar hoja
        if estado not in ['autorizada', 'parcial', 'surtida']:
            return Response({
                'error': 'Solo se pueden generar hojas para requisiciones autorizadas, parciales o surtidas',
                'estado_actual': requisicion.estado
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            # Generar el PDF
            pdf_buffer = generar_hoja_recoleccion(requisicion)
            
            # Crear respuesta HTTP con el PDF
            response = HttpResponse(
                pdf_buffer.getvalue(),
                content_type='application/pdf'
            )
            folio_safe = (requisicion.folio or f'REQ-{requisicion.id}').replace('/', '-')
            response['Content-Disposition'] = f'attachment; filename="Hoja_Recoleccion_{folio_safe}.pdf"'
            
            return response
            
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            return Response({
                'error': 'Error al generar la hoja de recolecci�n',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['get'], url_path='pdf-rechazo')
    def pdf_rechazo(self, request, pk=None):
        """
        Genera y descarga el PDF de rechazo para una requisici�n rechazada.
        """
        from core.utils.pdf_generator import generar_pdf_rechazo
        
        requisicion = self.get_object()
        estado = (requisicion.estado or '').lower()
        
        if estado != 'rechazada':
            return Response({
                'error': 'Solo se pueden generar PDFs de rechazo para requisiciones rechazadas',
                'estado_actual': requisicion.estado
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            pdf_buffer = generar_pdf_rechazo(requisicion)
            
            response = HttpResponse(
                pdf_buffer.getvalue(),
                content_type='application/pdf'
            )
            folio_safe = (requisicion.folio or f'REQ-{requisicion.id}').replace('/', '-')
            response['Content-Disposition'] = f'attachment; filename="Requisicion_Rechazada_{folio_safe}.pdf"'
            
            return response
            
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            return Response({
                'error': 'Error al generar el PDF de rechazo',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'])
    def estadisticas(self, request):
        """Estad�sticas de requisiciones filtradas por centro si aplica."""
        try:
            # SEGURIDAD: Filtrar por centro del usuario si no es admin/farmacia
            user = request.user
            base_queryset = Requisicion.objects.all()
            
            if not is_farmacia_or_admin(user):
                user_centro = get_user_centro(user)
                if user_centro:
                    base_queryset = base_queryset.filter(centro=user_centro)
                else:
                    base_queryset = Requisicion.objects.none()
            
            total = base_queryset.count()
            por_estado = {estado: base_queryset.filter(estado=estado).count() for estado, _ in ESTADOS_REQUISICION}
            
            # Top centros solo para admin/farmacia
            por_centro = []
            if is_farmacia_or_admin(user):
                centros = Centro.objects.annotate(total_requisiciones=Count('requisiciones')).filter(total_requisiciones__gt=0).order_by('-total_requisiciones')[:10]
                for centro in centros:
                    por_centro.append({'centro': centro.nombre, 'total': centro.total_requisiciones})
            
            return Response({'total': total, 'por_estado': por_estado, 'top_centros': por_centro})
        except Exception as exc:
            # traceback removido por seguridad (ISS-008)
            return Response({'error': 'Error al obtener estadisticas', 'mensaje': str(exc)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path='resumen_estados')
    def resumen_estados(self, request):
        """Resumen de conteos por estado y por grupo logico. Filtrado por centro si aplica."""
        try:
            # SEGURIDAD: Filtrar por centro del usuario si no es admin/farmacia
            user = request.user
            base_queryset = Requisicion.objects.all()
            
            if not is_farmacia_or_admin(user):
                user_centro = get_user_centro(user)
                if user_centro:
                    base_queryset = base_queryset.filter(centro=user_centro)
                else:
                    base_queryset = Requisicion.objects.none()
            
            por_estado = {estado.upper(): base_queryset.filter(estado=estado).count() for estado, _ in ESTADOS_REQUISICION}
            por_grupo = {}
            for nombre, estados in REQUISICION_GRUPOS_ESTADO.items():
                por_grupo[nombre] = base_queryset.filter(estado__in=estados).count()
            return Response({'por_estado': por_estado, 'por_grupo': por_grupo})
        except Exception as exc:
            # traceback removido por seguridad (ISS-008)
            return Response({'error': 'Error al obtener resumen de estados', 'mensaje': str(exc)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def dashboard_resumen(request):
    """
    Resumen del dashboard con KPIs y �ltimos movimientos.
    
    SEGURIDAD:
    - Usuarios de centro: solo ven datos de su centro
    - Admin/farmacia/vista: ven datos globales por defecto
    - Admin/farmacia/vista pueden usar ?centro=ID para filtrar por centro espec�fico
    """
    try:
        # SEGURIDAD: Filtrar por centro si el usuario no es admin/farmacia/vista
        user = request.user
        filtrar_por_centro = not is_farmacia_or_admin(user)
        user_centro = get_user_centro(user) if filtrar_por_centro else None
        
        # Admin/farmacia/vista puede filtrar por centro espec�fico
        centro_param = request.query_params.get('centro')
        if centro_param and centro_param not in ['', 'null', 'undefined', 'todos']:
            if is_farmacia_or_admin(user):
                try:
                    user_centro = Centro.objects.get(pk=int(centro_param))
                    filtrar_por_centro = True
                except (Centro.DoesNotExist, ValueError, TypeError):
                    pass
        
        # === PRODUCTOS ===
        total_productos = Producto.objects.filter(activo=True).count()
        
        # === LOTES ===
        lotes_query = Lote.objects.filter(
            estado='disponible', 
            deleted_at__isnull=True,
            cantidad_actual__gt=0
        )
        
        if filtrar_por_centro and user_centro:
            lotes_query = lotes_query.filter(centro=user_centro)
        
        stock_total = lotes_query.aggregate(
            total=Coalesce(Sum('cantidad_actual'), 0, output_field=IntegerField())
        )['total']
        lotes_activos = lotes_query.count()

        # === MOVIMIENTOS DEL MES ===
        inicio_mes = timezone.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        
        movimientos_base = Movimiento.objects.all()
        if filtrar_por_centro and user_centro:
            movimientos_base = movimientos_base.filter(
                Q(centro=user_centro) | Q(lote__centro=user_centro)
            )
        
        # Contar movimientos del mes actual
        movimientos_mes = movimientos_base.filter(fecha__gte=inicio_mes).count()
        
        # Si no hay movimientos este mes, mostrar total general como referencia
        total_movimientos = movimientos_base.count() if movimientos_mes == 0 else movimientos_mes

        # === �LTIMOS MOVIMIENTOS ===
        ultimos_movimientos = movimientos_base.select_related(
            'lote__producto', 'lote__centro', 'centro', 'requisicion', 'usuario'
        ).order_by('-fecha')[:10]
        
        movimientos_data = []
        for mov in ultimos_movimientos:
            lote = mov.lote
            producto_desc = 'N/A'
            producto_clave = 'N/A'
            lote_numero = 'N/A'
            
            if lote:
                lote_numero = lote.numero_lote or 'N/A'
                if lote.producto:
                    producto_desc = lote.producto.descripcion or 'N/A'
                    producto_clave = lote.producto.clave or 'N/A'
            
            # Determinar origen/destino
            lote_centro = lote.centro if lote else None
            mov_centro = mov.centro
            
            if mov.tipo == 'entrada':
                origen = mov.documento_referencia or 'Proveedor'
                destino = lote_centro.nombre if lote_centro else 'Farmacia Central'
            else:
                origen = lote_centro.nombre if lote_centro else 'Farmacia Central'
                destino = mov_centro.nombre if mov_centro else 'Consumo/Ajuste'
            
            movimientos_data.append({
                'id': mov.id,
                'tipo_movimiento': mov.tipo.upper(),
                'producto__descripcion': producto_desc,
                'producto__clave': producto_clave,
                'lote__codigo_lote': lote_numero,
                'cantidad': abs(mov.cantidad),
                'fecha_movimiento': mov.fecha.isoformat() if mov.fecha else None,
                'observaciones': mov.observaciones or '',
                'origen': origen,
                'destino': destino,
                'requisicion_folio': mov.requisicion.folio if mov.requisicion else None,
                'documento_referencia': mov.documento_referencia or None,
                'usuario': (mov.usuario.get_full_name() or mov.usuario.username) if mov.usuario else 'Sistema',
            })

        return Response({
            'kpi': {
                'total_productos': total_productos,
                'stock_total': max(0, stock_total),
                'lotes_activos': lotes_activos,
                'movimientos_mes': total_movimientos  # Usa total si no hay del mes actual
            },
            'ultimos_movimientos': movimientos_data
        })
        
    except Exception as exc:
        # traceback removido por seguridad (ISS-008)
        logger.exception('Error en dashboard_resumen')
        return Response({
            'kpi': {'total_productos': 0, 'stock_total': 0, 'lotes_activos': 0, 'movimientos_mes': 0},
            'ultimos_movimientos': [],
            'error': 'Error interno del servidor'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def dashboard_graficas(request):
    """
    Datos para graficas del dashboard.
    Retorna consumo_mensual, stock_por_centro y requisiciones_por_estado.
    
    SEGURIDAD: Filtra por centro del usuario si no es admin/farmacia.
    Admin/farmacia puede usar ?centro=ID para filtrar por centro espec�fico.
    """
    try:
        from dateutil.relativedelta import relativedelta
        
        # SEGURIDAD: Determinar filtro de centro
        user = request.user
        filtrar_por_centro = not is_farmacia_or_admin(user)
        user_centro = get_user_centro(user) if filtrar_por_centro else None
        
        # Admin/farmacia puede filtrar por centro espec�fico
        centro_param = request.query_params.get('centro')
        if centro_param and centro_param not in ['', 'null', 'undefined', 'todos']:
            if is_farmacia_or_admin(user):
                try:
                    user_centro = Centro.objects.get(pk=int(centro_param))
                    filtrar_por_centro = True
                except (Centro.DoesNotExist, ValueError, TypeError):
                    pass
        
        hoy = timezone.now().date()
        
        # =========================================
        # 1. CONSUMO MENSUAL (�LTIMOS 6 MESES)
        # =========================================
        consumo_mensual = []
        
        for i in range(6):
            # Calcular el mes (hace 5, 4, 3, 2, 1, 0 meses)
            fecha_mes = hoy - relativedelta(months=5-i)
            mes_inicio = fecha_mes.replace(day=1)
            
            # Calcular fin del mes
            if fecha_mes.month == 12:
                mes_fin = fecha_mes.replace(year=fecha_mes.year + 1, month=1, day=1)
            else:
                mes_fin = fecha_mes.replace(month=fecha_mes.month + 1, day=1)
            
            # Base query
            mov_mes = Movimiento.objects.filter(
                fecha__date__gte=mes_inicio,
                fecha__date__lt=mes_fin
            )
            
            if filtrar_por_centro and user_centro:
                mov_mes = mov_mes.filter(
                    Q(centro=user_centro) | Q(lote__centro=user_centro)
                )
            
            # Calcular entradas y salidas
            entradas = mov_mes.filter(tipo='entrada').aggregate(
                total=Coalesce(Sum('cantidad'), 0, output_field=IntegerField())
            )['total']
            
            salidas = mov_mes.filter(tipo='salida').aggregate(
                total=Coalesce(Sum('cantidad'), 0, output_field=IntegerField())
            )['total']
            
            consumo_mensual.append({
                'mes': mes_inicio.strftime('%b'),
                'entradas': max(0, abs(entradas)),
                'salidas': max(0, abs(salidas)),
            })
        
        # =========================================
        # 2. STOCK POR CENTRO (TODOS LOS CENTROS CON STOCK > 0)
        # =========================================
        stock_por_centro = []
        
        if not filtrar_por_centro:
            # Farmacia Central (lotes sin centro asignado)
            stock_farmacia = Lote.objects.filter(
                centro__isnull=True,
                estado='disponible',
                deleted_at__isnull=True,
                cantidad_actual__gt=0
            ).aggregate(
                total=Coalesce(Sum('cantidad_actual'), 0, output_field=IntegerField())
            )['total']
            
            # Solo agregar Farmacia Central si tiene stock
            if stock_farmacia > 0:
                stock_por_centro.append({
                    'centro': 'Farmacia Central',
                    'stock': stock_farmacia
                })
            
            # Todos los centros activos CON STOCK
            for centro in Centro.objects.filter(activo=True).order_by('nombre'):
                stock = Lote.objects.filter(
                    centro=centro,
                    estado='disponible',
                    deleted_at__isnull=True,
                    cantidad_actual__gt=0
                ).aggregate(
                    total=Coalesce(Sum('cantidad_actual'), 0, output_field=IntegerField())
                )['total']
                
                # Solo agregar centros con stock > 0
                if stock > 0:
                    # Truncar nombre largo
                    nombre = centro.nombre
                    if len(nombre) > 20:
                        nombre = nombre[:17] + '...'
                    
                    stock_por_centro.append({
                        'centro': nombre,
                        'stock': stock
                    })
        else:
            # Usuario de centro: solo su stock
            if user_centro:
                stock = Lote.objects.filter(
                    centro=user_centro,
                    estado='disponible',
                    deleted_at__isnull=True,
                    cantidad_actual__gt=0
                ).aggregate(
                    total=Coalesce(Sum('cantidad_actual'), 0, output_field=IntegerField())
                )['total']
                
                stock_por_centro.append({
                    'centro': user_centro.nombre,
                    'stock': max(0, stock)
                })
        
        # =========================================
        # 3. REQUISICIONES POR ESTADO
        # =========================================
        requisiciones_qs = Requisicion.objects.all()
        if filtrar_por_centro and user_centro:
            requisiciones_qs = requisiciones_qs.filter(centro=user_centro)
        
        estados_agg = requisiciones_qs.values('estado').annotate(
            cantidad=Count('id')
        ).order_by('estado')
        
        requisiciones_por_estado = []
        for item in estados_agg:
            if item['cantidad'] > 0:
                requisiciones_por_estado.append({
                    'estado': item['estado'].upper(),
                    'cantidad': item['cantidad']
                })
        
        return Response({
            'consumo_mensual': consumo_mensual,
            'stock_por_centro': stock_por_centro,
            'requisiciones_por_estado': requisiciones_por_estado
        })
        
    except Exception as exc:
        # traceback removido por seguridad (ISS-008)
        logger.exception('Error en dashboard_graficas')
        return Response({
            'consumo_mensual': [],
            'stock_por_centro': [],
            'requisiciones_por_estado': [],
            'error': 'Error interno del servidor'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def trazabilidad_producto(request, clave):
    """
    Trazabilidad de un producto identificado por clave (case-insensitive).
    Retorna lotes, movimientos y alertas de stock/caducidad.
    
    SEGURIDAD: Filtra por centro del usuario si no es admin/farmacia.
    """
    try:
        # SEGURIDAD: Determinar filtro de centro
        user = request.user
        if not user or not user.is_authenticated:
            return Response({'error': 'Autenticacion requerida'}, status=status.HTTP_403_FORBIDDEN)
        rol_usuario = (getattr(user, 'rol', '') or '').lower()
        if rol_usuario == 'vista':
            return Response({'error': 'No tienes permiso para trazabilidad'}, status=status.HTTP_403_FORBIDDEN)
        filtrar_por_centro = not is_farmacia_or_admin(user)
        user_centro = get_user_centro(user) if filtrar_por_centro else None
        
        centro_param = request.query_params.get('centro')
        if centro_param and is_farmacia_or_admin(user):
            try:
                user_centro = Centro.objects.get(pk=centro_param)
                filtrar_por_centro = True
            except Centro.DoesNotExist:
                pass
        
        producto = Producto.objects.filter(clave__iexact=clave).first()
        if not producto:
            return Response({'error': 'Producto no encontrado', 'clave_buscada': clave}, status=status.HTTP_404_NOT_FOUND)

        lotes = Lote.objects.filter(producto=producto, deleted_at__isnull=True)
        
        # Aplicar filtro de centro
        if filtrar_por_centro and user_centro:
            lotes = lotes.filter(centro=user_centro)
        
        lotes = lotes.order_by('-created_at')
        lotes_data = []
        from datetime import date, timedelta

        for lote in lotes:
            dias_caducidad = (lote.fecha_caducidad - date.today()).days if lote.fecha_caducidad else None
            if dias_caducidad is None:
                estado_caducidad = 'DESCONOCIDO'
            elif dias_caducidad < 0:
                estado_caducidad = 'VENCIDO'
            elif dias_caducidad <= 7:
                estado_caducidad = 'CRITICO'
            elif dias_caducidad <= 30:
                estado_caducidad = 'PROXIMO'
            else:
                estado_caducidad = 'NORMAL'

            movimientos_lote = Movimiento.objects.filter(lote=lote)
            total_entradas = movimientos_lote.filter(tipo='entrada').aggregate(total=Sum('cantidad'))['total'] or 0
            total_salidas = movimientos_lote.filter(tipo='salida').aggregate(total=Sum('cantidad'))['total'] or 0

            lotes_data.append({
                'id': lote.id,
                'numero_lote': lote.numero_lote,
                'fecha_caducidad': lote.fecha_caducidad.isoformat() if lote.fecha_caducidad else None,
                'dias_para_caducar': dias_caducidad,
                'estado_caducidad': estado_caducidad,
                'cantidad_actual': lote.cantidad_actual,
                'cantidad_inicial': lote.cantidad_inicial,
                'total_entradas': total_entradas,
                'total_salidas': total_salidas,
                'proveedor': lote.proveedor or 'N/A',
                'precio_compra': str(lote.precio_compra) if lote.precio_compra else None,
                # Campos de trazabilidad de contratos (solo para ADMIN/FARMACIA)
                'numero_contrato': (lote.numero_contrato or '') if is_farmacia_or_admin(user) else None,
                'marca': (lote.marca or '') if is_farmacia_or_admin(user) else None,
                'activo': getattr(lote, 'activo', True),
                'created_at': lote.created_at.isoformat()
            })

        movimientos = Movimiento.objects.filter(lote__producto=producto).select_related('lote')
        
        # Aplicar filtro de centro a movimientos
        if filtrar_por_centro and user_centro:
            movimientos = movimientos.filter(lote__centro=user_centro)
        
        movimientos = movimientos.order_by('-fecha')[:100]
        movimientos_data = []
        for mov in movimientos:
            movimientos_data.append({
                'id': mov.id,
                'tipo_movimiento': mov.tipo.upper(),
                'tipo': mov.tipo.upper(),
                'lote': mov.lote.numero_lote if mov.lote else 'N/A',
                'cantidad': mov.cantidad,
                'fecha_movimiento': mov.fecha.isoformat(),
                'observaciones': mov.observaciones or ''
            })

        stock_total = lotes.filter(estado='disponible').aggregate(total=Sum('cantidad_actual'))['total'] or 0
        lotes_activos = lotes.filter(estado='disponible', cantidad_actual__gt=0).count()
        total_lotes = lotes.count()
        
        # Totales de entradas/salidas filtrados por centro si aplica
        mov_query = Movimiento.objects.filter(lote__producto=producto)
        if filtrar_por_centro and user_centro:
            mov_query = mov_query.filter(lote__centro=user_centro)
        total_entradas_prod = mov_query.filter(tipo='entrada').aggregate(total=Sum('cantidad'))['total'] or 0
        total_salidas_prod = mov_query.filter(tipo='salida').aggregate(total=Sum('cantidad'))['total'] or 0

        fecha_limite = date.today() + timedelta(days=30)
        lotes_proximos_vencer = lotes.filter(cantidad_actual__gt=0, fecha_caducidad__lte=fecha_limite, fecha_caducidad__gte=date.today()).count()
        lotes_vencidos = lotes.filter(cantidad_actual__gt=0, fecha_caducidad__lt=date.today()).count()

        alertas = []
        if stock_total < producto.stock_minimo:
            alertas.append({'tipo': 'STOCK_BAJO', 'mensaje': f'Stock actual ({stock_total}) por debajo del minimo ({producto.stock_minimo})', 'nivel': 'CRITICO'})
        if lotes_vencidos > 0:
            alertas.append({'tipo': 'LOTES_VENCIDOS', 'mensaje': f'{lotes_vencidos} lote(s) vencido(s) con stock', 'nivel': 'CRITICO'})
        if lotes_proximos_vencer > 0:
            alertas.append({'tipo': 'PROXIMOS_VENCER', 'mensaje': f'{lotes_proximos_vencer} lote(s) proximo(s) a vencer (30 dias)', 'nivel': 'ADVERTENCIA'})

        return Response({
            'codigo': producto.clave,
            'producto': {
                'id': producto.id,
                'clave': producto.clave,
                'descripcion': producto.descripcion,
                'unidad_medida': producto.unidad_medida,
                'stock_minimo': producto.stock_minimo,
                'precio_unitario': str(producto.precio_unitario) if producto.precio_unitario else None,
                'activo': producto.activo
            },
            'estadisticas': {
                'stock_total': stock_total,
                'total_lotes': total_lotes,
                'lotes_activos': lotes_activos,
                'total_entradas': total_entradas_prod,
                'total_salidas': total_salidas_prod,
                'diferencia': total_entradas_prod - total_salidas_prod,
                'lotes_proximos_vencer': lotes_proximos_vencer,
                'lotes_vencidos': lotes_vencidos,
                'bajo_minimo': stock_total < producto.stock_minimo
            },
            'lotes': lotes_data,
            'movimientos': movimientos_data,
            'total_movimientos': Movimiento.objects.filter(lote__producto=producto).count(),
            'alertas': alertas
        })
    except Exception as exc:
        # traceback removido por seguridad (ISS-008)
        return Response({'error': 'Error al obtener trazabilidad del producto', 'mensaje': str(exc)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
def trazabilidad_lote(request, codigo):
    """
    Trazabilidad completa de un lote por su numero.
    
    SEGURIDAD: Filtra por centro del usuario si no es admin/farmacia.
    """
    try:
        if not request.user or not request.user.is_authenticated or not is_farmacia_or_admin(request.user):
            return Response({'error': 'Solo usuarios de farmacia o administradores pueden acceder a reportes'}, status=status.HTTP_403_FORBIDDEN)

        # SEGURIDAD: Determinar filtro de centro
        user = request.user
        filtrar_por_centro = not is_farmacia_or_admin(user)
        user_centro = get_user_centro(user) if filtrar_por_centro else None
        
        lote_query = Lote.objects.select_related('producto').filter(numero_lote__iexact=codigo)
        
        # Aplicar filtro de centro
        if filtrar_por_centro and user_centro:
            lote_query = lote_query.filter(centro=user_centro)
        
        lote = lote_query.first()
        if not lote:
            return Response({'error': 'Lote no encontrado', 'codigo_buscado': codigo}, status=status.HTTP_404_NOT_FOUND)

        movimientos = Movimiento.objects.filter(lote=lote).order_by('fecha')
        historial = []
        saldo = 0
        for mov in movimientos:
            saldo += mov.cantidad
            historial.append({
                'id': mov.id,
                'fecha': mov.fecha.isoformat(),
                'tipo': mov.tipo.upper(),
                'cantidad': mov.cantidad,
                'saldo': saldo,
                'observaciones': mov.observaciones or ''
            })

        total_entradas = movimientos.filter(tipo='entrada').aggregate(total=Sum('cantidad'))['total'] or 0
        total_salidas = movimientos.filter(tipo='salida').aggregate(total=Sum('cantidad'))['total'] or 0

        from datetime import date
        dias_caducidad = (lote.fecha_caducidad - date.today()).days if lote.fecha_caducidad else None
        if dias_caducidad is None:
            estado_caducidad = 'DESCONOCIDO'
        elif dias_caducidad < 0:
            estado_caducidad = 'VENCIDO'
        elif dias_caducidad <= 7:
            estado_caducidad = 'CRITICO'
        elif dias_caducidad <= 30:
            estado_caducidad = 'PROXIMO'
        else:
            estado_caducidad = 'NORMAL'

        alertas = []
        if dias_caducidad is not None:
            if dias_caducidad < 0:
                alertas.append({'tipo': 'VENCIDO', 'mensaje': f'Lote vencido hace {abs(dias_caducidad)} dias', 'nivel': 'CRITICO'})
            elif dias_caducidad <= 7:
                alertas.append({'tipo': 'CRITICO', 'mensaje': f'Caduca en {dias_caducidad} dias', 'nivel': 'CRITICO'})
            elif dias_caducidad <= 30:
                alertas.append({'tipo': 'PROXIMO', 'mensaje': f'Caduca en {dias_caducidad} dias', 'nivel': 'ADVERTENCIA'})

        return Response({
            'id': lote.id,
            'numero_lote': lote.numero_lote,
            'producto': lote.producto.clave,
            'lote': {
                'id': lote.id,
                'numero_lote': lote.numero_lote,
                'producto': lote.producto.clave,
                'producto_descripcion': lote.producto.descripcion,
                'cantidad_actual': lote.cantidad_actual,
                'cantidad_inicial': lote.cantidad_inicial,
                'estado': lote.estado,
                'fecha_caducidad': lote.fecha_caducidad.isoformat() if lote.fecha_caducidad else None,
                'dias_para_caducar': dias_caducidad,
                'estado_caducidad': estado_caducidad,
                'proveedor': lote.proveedor,
                # Campos de trazabilidad de contratos (solo para ADMIN/FARMACIA)
                'numero_contrato': lote.numero_contrato if is_farmacia_or_admin(user) else None,
                'marca': lote.marca if is_farmacia_or_admin(user) else None,
            },
            'estadisticas': {
                'total_entradas': total_entradas,
                'total_salidas': total_salidas,
                'diferencia': total_entradas - total_salidas,
                'cantidad_actual': lote.cantidad_actual,
                'saldo_calculado': saldo,
                'diferencia_stock': saldo - lote.cantidad_actual,
                'consistente': saldo == lote.cantidad_actual
            },
            'movimientos': historial,
            'historial': historial,  # compatibilidad hacia atr?s
            'total_movimientos': movimientos.count(),
            'alertas': alertas
        })
    except Exception as exc:
        # traceback removido por seguridad (ISS-008)
        return Response({'error': 'Error al obtener trazabilidad del lote', 'mensaje': str(exc)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
def reporte_inventario(request):
    """
    Genera reporte de inventario actual.
    Por defecto devuelve JSON. Con ?formato=excel devuelve Excel.
    
    SEGURIDAD: Filtra por centro del usuario si no es admin/farmacia.
    Admin/farmacia puede usar ?centro=ID para filtrar.
    
    Par�metros:
    - centro: ID del centro o 'central' para farmacia central
    - nivel_stock: alto, bajo, normal, sin_stock
    - formato: json (default), excel, pdf
    """
    try:
        if not request.user or not request.user.is_authenticated or not is_farmacia_or_admin(request.user):
            return Response({'error': 'Solo usuarios de farmacia o administradores pueden acceder a reportes'}, status=status.HTTP_403_FORBIDDEN)
        # SEGURIDAD: Determinar filtro de centro
        user = request.user
        filtrar_por_centro = not is_farmacia_or_admin(user)
        user_centro = get_user_centro(user) if filtrar_por_centro else None
        
        # Admin/farmacia puede filtrar por centro espec�fico
        centro_param = request.query_params.get('centro')
        if centro_param and is_farmacia_or_admin(user):
            if centro_param == 'central':
                # Filtrar solo farmacia central
                filtrar_por_centro = True
                user_centro = None  # centro=NULL significa farmacia central
            else:
                try:
                    user_centro = Centro.objects.get(pk=centro_param)
                    filtrar_por_centro = True
                except Centro.DoesNotExist:
                    pass
        
        formato = request.query_params.get('formato', 'json')
        nivel_stock_filtro = request.query_params.get('nivel_stock', '').lower().strip()
        productos = Producto.objects.filter(activo=True).order_by('clave')
        
        # Construir datos
        datos = []
        total_productos = 0
        total_stock = 0
        productos_bajo_minimo = 0
        productos_sin_stock = 0
        
        for idx, producto in enumerate(productos, 1):
            # Aplicar filtro de centro si corresponde
            lotes_query = producto.lotes.filter(
                deleted_at__isnull=True,
                estado='disponible'
            )
            if filtrar_por_centro:
                if user_centro:
                    lotes_query = lotes_query.filter(centro=user_centro)
                else:
                    # centro=NULL significa farmacia central
                    lotes_query = lotes_query.filter(centro__isnull=True)
            
            stock_total = lotes_query.aggregate(total=Sum('cantidad_actual'))['total'] or 0
            lotes_activos = lotes_query.filter(cantidad_actual__gt=0).count()

            nivel = 'alto'
            if stock_total == 0:
                nivel = 'sin_stock'
                productos_sin_stock += 1
            elif stock_total < producto.stock_minimo:
                nivel = 'bajo'
                productos_bajo_minimo += 1
            elif stock_total < producto.stock_minimo * 1.5:
                nivel = 'normal'
            
            # Filtrar por nivel_stock si se especific�
            if nivel_stock_filtro and nivel != nivel_stock_filtro:
                continue

            # Se incluye 'nivel_stock' para compatibilidad con el frontend
            datos.append({
                '#': idx,
                'clave': producto.clave,
                'descripcion': producto.descripcion,
                'unidad': producto.unidad_medida,
                'unidad_medida': producto.unidad_medida,  # alias esperado por frontend
                'stock_minimo': producto.stock_minimo,
                'stock_actual': stock_total,
                'lotes_activos': lotes_activos,
                'nivel': nivel,
                'nivel_stock': nivel,
                'precio_unitario': float(producto.precio_unitario) if producto.precio_unitario else 0.0,
            })
            total_productos += 1
            total_stock += stock_total
        
        resumen = {
            'total_productos': total_productos,
            'total_stock': total_stock,
            'stock_total': total_stock,  # alias para compatibilidad frontend
            'productos_bajo_minimo': productos_bajo_minimo,
            'productos_sin_stock': productos_sin_stock,
        }
        
        # Si formato es JSON, devolver datos
        if formato == 'json':
            return Response({
                'datos': datos,
                'resumen': resumen
            })
        
        # Formato PDF
        if formato == 'pdf':
            from core.utils.pdf_reports import generar_reporte_inventario
            
            # Preparar datos para el generador PDF
            productos_data = []
            for item in datos:
                productos_data.append({
                    'clave': item['clave'],
                    'descripcion': item['descripcion'],
                    'unidad_medida': item['unidad'],
                    'stock_minimo': item['stock_minimo'],
                    'stock_actual': item['stock_actual'],
                    'lotes_activos': item['lotes_activos'],
                    'nivel': item['nivel'],
                    'precio_unitario': item['precio_unitario']
                })
            
            filtros = {
                'fecha_generacion': timezone.now().strftime('%d/%m/%Y %H:%M')
            }
            if filtrar_por_centro and user_centro:
                filtros['centro'] = user_centro.nombre
            
            pdf_buffer = generar_reporte_inventario(productos_data, formato='pdf', filtros=filtros)
            
            response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = f"attachment; filename=Inventario_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            return response
        
        # Formato Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Inventario'

        ws.merge_cells('A1:I1')
        titulo = ws['A1']
        titulo.value = 'REPORTE DE INVENTARIO ACTUAL'
        titulo.font = Font(bold=True, size=14, color='632842')
        titulo.alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells('A2:I2')
        subtitulo = ws['A2']
        subtitulo.value = f"Generado el {timezone.now().strftime('%d/%m/%Y %H:%M:%S')}"
        subtitulo.alignment = Alignment(horizontal='center')
        subtitulo.font = Font(size=10, italic=True)

        ws.append([])
        headers = ['#', 'Clave', 'Descripci�n', 'Unidad', 'Stock M�n.', 'Stock Actual', 'Lotes', 'Nivel', 'Precio']
        ws.append(headers)

        header_fill = PatternFill(start_color='632842', end_color='632842', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        for cell in ws[4]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for item in datos:
            ws.append([
                item['#'],
                item['clave'],
                item['descripcion'][:70],
                item['unidad'],
                item['stock_minimo'],
                item['stock_actual'],
                item['lotes_activos'],
                item['nivel'].upper(),
                item['precio_unitario']
            ])

        ws.append([])
        resumen_row = ws.max_row + 1
        ws[f'B{resumen_row}'] = 'Total de Productos'
        ws[f'C{resumen_row}'] = resumen['total_productos']
        ws[f'B{resumen_row + 1}'] = 'Stock Total'
        ws[f'C{resumen_row + 1}'] = resumen['total_stock']
        ws[f'B{resumen_row + 2}'] = 'Productos bajo m�nimo'
        ws[f'C{resumen_row + 2}'] = resumen['productos_bajo_minimo']

        for col, width in zip(['A','B','C','D','E','F','G','H','I'], [8,14,45,10,14,14,10,12,12]):
            ws.column_dimensions[col].width = width

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f"attachment; filename=Inventario_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(response)
        return response

    except Exception as e:
        # traceback removido por seguridad (ISS-008)
        return Response({'error': 'Error al generar reporte', 'mensaje': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
@api_view(['GET'])
def reporte_movimientos(request):
    """
    Genera reporte de movimientos con filtros.
    
    SEGURIDAD: Filtra por centro del usuario si no es admin/farmacia.
    Admin/farmacia puede usar ?centro=ID para filtrar.
    
    Parametros:
    - fecha_inicio: Fecha inicial (YYYY-MM-DD)
    - fecha_fin: Fecha final (YYYY-MM-DD)
    - tipo: ENTRADA o SALIDA
    - centro: ID del centro (solo admin/farmacia)
    - formato: excel o pdf
    """
    try:
        logger.info(f"Generando reporte de movimientos. Params: {dict(request.query_params)}")
        
        # SEGURIDAD: Determinar filtro de centro
        user = request.user
        filtrar_por_centro = not is_farmacia_or_admin(user)
        user_centro = get_user_centro(user) if filtrar_por_centro else None
        
        # Admin/farmacia puede filtrar por centro específico
        centro_param = request.query_params.get('centro')
        if centro_param and is_farmacia_or_admin(user):
            try:
                user_centro = Centro.objects.get(pk=centro_param)
                filtrar_por_centro = True
            except Centro.DoesNotExist:
                pass
        
        # Obtener parametros
        fecha_inicio = request.query_params.get('fecha_inicio')
        fecha_fin = request.query_params.get('fecha_fin')
        tipo = request.query_params.get('tipo')
        formato = request.query_params.get('formato', 'excel')
        
        # Filtrar movimientos
        movimientos = Movimiento.objects.select_related('lote__producto').all()
        
        # Aplicar filtro de centro
        if filtrar_por_centro and user_centro:
            movimientos = movimientos.filter(lote__centro=user_centro)
        
        if fecha_inicio:
            movimientos = movimientos.filter(fecha__gte=fecha_inicio)
        if fecha_fin:
            movimientos = movimientos.filter(fecha__lte=fecha_fin)
        if tipo:
            movimientos = movimientos.filter(tipo=tipo.lower())
        
        movimientos = movimientos.order_by('-fecha')
        
        # Construir datos
        datos = []
        total_entradas = 0
        total_salidas = 0
        
        for mov in movimientos:
            amount = abs(mov.cantidad) if mov.tipo == 'salida' else mov.cantidad
            datos.append({
                'fecha': mov.fecha.strftime('%d/%m/%Y %H:%M'),
                'tipo': mov.tipo.upper(),
                'producto': f"{getattr(mov.lote.producto, 'clave', 'N/A')} - {getattr(mov.lote.producto, 'descripcion', '')[:40]}",
                'lote': getattr(mov.lote, 'numero_lote', 'N/A') if mov.lote else 'N/A',
                'cantidad': amount,
                'observaciones': mov.observaciones or ''
            })
            if mov.tipo == 'entrada':
                total_entradas += amount
            else:
                total_salidas += amount
        
        resumen = {
            'total_movimientos': len(datos),
            'total_entradas': total_entradas,
            'total_salidas': total_salidas,
            'diferencia': total_entradas - total_salidas
        }
        
        # Formato JSON
        if formato == 'json':
            return Response({
                'datos': datos,
                'resumen': resumen
            })
        
        # Formato PDF
        if formato == 'pdf':
            from core.utils.pdf_reports import generar_reporte_movimientos
            
            # Preparar datos para el generador PDF
            movimientos_data = []
            for item in datos:
                movimientos_data.append({
                    'fecha': item['fecha'],
                    'tipo': item['tipo'],
                    'producto': item['producto'],
                    'lote': item['lote'],
                    'cantidad': item['cantidad'],
                    'observaciones': item['observaciones']
                })
            
            filtros = {
                'fecha_generacion': timezone.now().strftime('%d/%m/%Y %H:%M')
            }
            if fecha_inicio:
                filtros['fecha_inicio'] = fecha_inicio
            if fecha_fin:
                filtros['fecha_fin'] = fecha_fin
            if tipo:
                filtros['tipo'] = tipo
            if filtrar_por_centro and user_centro:
                filtros['centro'] = user_centro.nombre
            
            pdf_buffer = generar_reporte_movimientos(movimientos_data, filtros=filtros)
            
            response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = f"attachment; filename=Movimientos_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            return response
        
        if formato == 'excel':
            # Generar Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Movimientos'
            
            # Titulo
            ws.merge_cells('A1:G1')
            titulo_cell = ws['A1']
            titulo_cell.value = 'REPORTE DE MOVIMIENTOS'
            titulo_cell.font = Font(bold=True, size=14, color='632842')
            titulo_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Filtros aplicados
            filtros_text = []
            if fecha_inicio:
                filtros_text.append(f'Desde: {fecha_inicio}')
            if fecha_fin:
                filtros_text.append(f'Hasta: {fecha_fin}')
            if tipo:
                filtros_text.append(f'Tipo: {tipo}')
            
            ws.merge_cells('A2:G2')
            filtros_cell = ws['A2']
            filtros_cell.value = ' | '.join(filtros_text) if filtros_text else 'Sin filtros'
            filtros_cell.font = Font(size=10, italic=True)
            filtros_cell.alignment = Alignment(horizontal='center')
            
            ws.append([])  # Linea en blanco
            
            # Encabezados
            headers = ['#', 'Fecha', 'Tipo', 'Producto', 'Lote', 'Cantidad', 'Observaciones']
            ws.append(headers)
            
            # Estilo encabezados
            header_fill = PatternFill(start_color='632842', end_color='632842', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            
            for cell in ws[4]:  # Fila 4 tiene los encabezados
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Datos
            total_entradas = 0
            total_salidas = 0
            
            for idx, mov in enumerate(movimientos, 1):
                amount = abs(mov.cantidad) if mov.tipo == 'salida' else mov.cantidad
                ws.append([
                    idx,
                    mov.fecha.strftime('%d/%m/%Y %H:%M'),
                    mov.tipo.upper(),
                    f"{getattr(mov.lote.producto, 'clave', 'N/A')} - {getattr(mov.lote.producto, 'descripcion', '')[:40]}",
                    getattr(mov.lote, 'numero_lote', 'N/A') if mov.lote else 'N/A',
                    amount,
                    mov.observaciones or ''
                ])
                
                if mov.tipo == 'entrada':
                    total_entradas += amount
                else:
                    total_salidas += amount
                
                # Colorear por tipo
                row_num = idx + 4
                tipo_cell = ws.cell(row=row_num, column=3)
                if mov.tipo == 'entrada':
                    tipo_cell.fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
                    tipo_cell.font = Font(color='155724', bold=True)
                else:
                    tipo_cell.fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')
                    tipo_cell.font = Font(color='721C24', bold=True)
            
            # Resumen
            ws.append([])
            resumen_row = ws.max_row + 1
            ws[f'D{resumen_row}'] = 'TOTAL ENTRADAS:'
            ws[f'D{resumen_row}'].font = Font(bold=True)
            ws[f'F{resumen_row}'] = total_entradas
            ws[f'F{resumen_row}'].font = Font(bold=True, color='155724')
            
            ws[f'D{resumen_row + 1}'] = 'TOTAL SALIDAS:'
            ws[f'D{resumen_row + 1}'].font = Font(bold=True)
            ws[f'F{resumen_row + 1}'] = total_salidas
            ws[f'F{resumen_row + 1}'].font = Font(bold=True, color='721C24');
            
            ws[f'D{resumen_row + 2}'] = 'DIFERENCIA:'
            ws[f'D{resumen_row + 2}'].font = Font(bold=True)
            ws[f'F{resumen_row + 2}'] = total_entradas - total_salidas
            ws[f'F{resumen_row + 2}'].font = Font(bold=True)
            
            # Ajustar anchos
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 18
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 50
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 12
            ws.column_dimensions['G'].width = 30
            
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename=Movimientos_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            wb.save(response)
            
            print(f" Reporte Excel generado: {movimientos.count()} movimientos")
            
            return response
            
        # Formato no soportado
        return Response({
            'error': 'Formato no soportado',
            'formatos_disponibles': ['json', 'pdf', 'excel']
        }, status=status.HTTP_400_BAD_REQUEST)
            
    except Exception as e:
        print(f" Error generando reporte: {str(e)}")
        # traceback removido por seguridad (ISS-008)
        return Response({
            'error': 'Error al generar reporte de movimientos',
            'mensaje': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
def reporte_caducidades(request):
    """
    Genera reporte de lotes proximos a caducar.
    Por defecto devuelve JSON. Con ?formato=excel devuelve Excel.
    
    SEGURIDAD: Filtra por centro del usuario si no es admin/farmacia.
    Admin/farmacia puede usar ?centro=ID para filtrar.
    
    Par�metros:
    - dias: N�mero de d�as de anticipaci�n (default: 30)
    - centro: ID del centro o 'central' para farmacia central
    - estado: vencido, critico, proximo (filtra por estado espec�fico)
    - formato: json (default), excel, pdf
    """
    try:
        # SEGURIDAD: Determinar filtro de centro
        user = request.user
        filtrar_por_centro = not is_farmacia_or_admin(user)
        user_centro = get_user_centro(user) if filtrar_por_centro else None
        
        # Admin/farmacia puede filtrar por centro específico
        centro_param = request.query_params.get('centro')
        if centro_param and is_farmacia_or_admin(user):
            if centro_param == 'central':
                filtrar_por_centro = True
                user_centro = None
            else:
                try:
                    user_centro = Centro.objects.get(pk=centro_param)
                    filtrar_por_centro = True
                except Centro.DoesNotExist:
                    pass
        
        dias = int(request.query_params.get('dias', 30))
        formato = request.query_params.get('formato', 'json')
        estado_filtro = request.query_params.get('estado', '').lower().strip()
        
        fecha_limite = date.today() + timedelta(days=dias)
        
        # Obtener lotes proximos a vencer
        lotes = Lote.objects.filter(
            deleted_at__isnull=True,
            cantidad_actual__gt=0,
            fecha_caducidad__lte=fecha_limite
        ).select_related('producto')
        
        # Aplicar filtro de centro
        if filtrar_por_centro:
            if user_centro:
                lotes = lotes.filter(centro=user_centro)
            else:
                lotes = lotes.filter(centro__isnull=True)
        
        lotes = lotes.order_by('fecha_caducidad')
        
        # Construir datos
        datos = []
        vencidos = 0
        criticos = 0
        proximos = 0
        
        for lote in lotes:
            dias_restantes = (lote.fecha_caducidad - date.today()).days
            
            if dias_restantes < 0:
                estado = 'vencido'
                vencidos += 1
            elif dias_restantes <= 7:
                estado = 'critico'
                criticos += 1
            else:
                estado = 'proximo'
                proximos += 1
            
            # Filtrar por estado si se especific�
            if estado_filtro and estado != estado_filtro:
                continue
            
            datos.append({
                'producto': f"{lote.producto.clave} - {lote.producto.descripcion}",
                'lote': lote.numero_lote,
                'caducidad': lote.fecha_caducidad.isoformat(),
                'dias_restantes': dias_restantes,
                'stock': lote.cantidad_actual,
                'estado': estado,
            })
        
        resumen = {
            'total': len(datos),
            'vencidos': vencidos,
            'criticos': criticos,
            'proximos': proximos,
            'dias_filtro': dias,
        }
        
        # Si formato es JSON, devolver datos
        if formato == 'json':
            return Response({
                'datos': datos,
                'resumen': resumen
            })
        
        # Formato PDF
        if formato == 'pdf':
            from core.utils.pdf_reports import generar_reporte_caducidades
            
            # Preparar datos para el generador PDF
            lotes_data = []
            for item in datos:
                lotes_data.append({
                    'producto': item['producto'],
                    'numero_lote': item['lote'],
                    'fecha_caducidad': item['caducidad'],
                    'dias_restantes': item['dias_restantes'],
                    'cantidad_actual': item['stock'],
                    'estado': item['estado']
                })
            
            filtros = {
                'dias_anticipacion': dias,
                'fecha_generacion': timezone.now().strftime('%d/%m/%Y %H:%M')
            }
            if filtrar_por_centro and user_centro:
                filtros['centro'] = user_centro.nombre
            
            pdf_buffer = generar_reporte_caducidades(lotes_data, dias=dias, filtros=filtros)
            
            response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = f"attachment; filename=Caducidades_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            return response
        
        # Generar Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Caducidades'
        
        ws.merge_cells('A1:G1')
        titulo_cell = ws['A1']
        titulo_cell.value = f'REPORTE DE LOTES PROXIMOS A CADUCAR ({dias} DIAS)'
        titulo_cell.font = Font(bold=True, size=14, color='632842')
        titulo_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        ws.merge_cells('A2:G2')
        fecha_cell = ws['A2']
        fecha_cell.value = f'Generado el {timezone.now().strftime("%d/%m/%Y %H:%M")}'
        fecha_cell.font = Font(size=10, italic=True)
        fecha_cell.alignment = Alignment(horizontal='center')
        
        ws.append([])
        
        headers = ['#', 'Producto', 'Lote', 'Caducidad', 'D�as Restantes', 'Stock', 'Estado']
        ws.append(headers)
        
        header_fill = PatternFill(start_color='632842', end_color='632842', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        
        for cell in ws[4]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for idx, item in enumerate(datos, 1):
            ws.append([
                idx,
                item['producto'][:50],
                item['lote'],
                item['caducidad'],
                item['dias_restantes'],
                item['stock'],
                item['estado'].upper()
            ])
        
        ws.append([])
        resumen_row = ws.max_row + 1
        ws[f'B{resumen_row}'] = 'Total:'
        ws[f'C{resumen_row}'] = resumen['total']
        ws[f'B{resumen_row + 1}'] = 'Vencidos:'
        ws[f'C{resumen_row + 1}'] = resumen['vencidos']
        ws[f'B{resumen_row + 2}'] = 'Cr�ticos:'
        ws[f'C{resumen_row + 2}'] = resumen['criticos']
        ws[f'B{resumen_row + 3}'] = 'Pr�ximos:'
        ws[f'C{resumen_row + 3}'] = resumen['proximos']
        
        for col, width in zip(['A','B','C','D','E','F','G'], [8,50,20,15,15,12,12]):
            ws.column_dimensions[col].width = width
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=Caducidades_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        wb.save(response)
        
        return response
        
    except Exception as e:
        # traceback removido por seguridad (ISS-008)
        return Response({
            'error': 'Error al generar reporte de caducidades',
            'mensaje': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def reporte_requisiciones(request):
    """
    Genera reporte de requisiciones con filtros.
    Por defecto devuelve JSON. Con ?formato=excel devuelve Excel.
    
    SEGURIDAD: Filtra por centro del usuario si no es admin/farmacia.
    Admin/farmacia puede usar ?centro=ID para filtrar.
    
    Parametros:
    - fecha_inicio: Fecha inicial (YYYY-MM-DD)
    - fecha_fin: Fecha final (YYYY-MM-DD)
    - estado: Estado de la requisicion
    - centro: ID del centro (solo admin/farmacia)
    - formato: json/excel
    """
    try:
        # SEGURIDAD: Determinar filtro de centro
        user = request.user
        filtrar_por_centro = not is_farmacia_or_admin(user)
        user_centro = get_user_centro(user) if filtrar_por_centro else None
        
        fecha_inicio = request.query_params.get('fecha_inicio')
        fecha_fin = request.query_params.get('fecha_fin')
        estado = request.query_params.get('estado')
        centro_param = request.query_params.get('centro') or request.query_params.get('centro_id')
        formato = request.query_params.get('formato', 'json')
        
        requisiciones = Requisicion.objects.select_related('centro', 'usuario_solicita').all()
        
        # Aplicar filtro de centro obligatorio para usuarios de centro
        if filtrar_por_centro and user_centro:
            requisiciones = requisiciones.filter(centro=user_centro)
        elif centro_param and is_farmacia_or_admin(user):
            requisiciones = requisiciones.filter(centro_id=centro_param)
        
        if fecha_inicio:
            requisiciones = requisiciones.filter(fecha_solicitud__gte=fecha_inicio)
        if fecha_fin:
            requisiciones = requisiciones.filter(fecha_solicitud__lte=fecha_fin)
        if estado:
            requisiciones = requisiciones.filter(estado__iexact=estado)
        
        requisiciones = requisiciones.order_by('-fecha_solicitud')
        
        # Construir datos
        datos = []
        estados_count = {}
        
        for req in requisiciones:
            estado_req = req.estado.upper()
            estados_count[estado_req] = estados_count.get(estado_req, 0) + 1
            
            datos.append({
                'id': req.id,
                'folio': req.folio or f'REQ-{req.id}',
                'centro': req.centro.nombre if req.centro else 'N/A',
                'estado': estado_req,
                'fecha_solicitud': req.fecha_solicitud.isoformat() if req.fecha_solicitud else None,
                'total_productos': req.detalles.count(),
                'solicitante': req.usuario_solicita.get_full_name() if req.usuario_solicita else 'N/A',
            })
        
        resumen = {
            'total': len(datos),
            'por_estado': estados_count,
        }
        
        # Si formato es JSON, devolver datos
        if formato == 'json':
            return Response({
                'datos': datos,
                'resumen': resumen
            })
        
        # Formato PDF
        if formato == 'pdf':
            from core.utils.pdf_reports import generar_reporte_requisiciones
            
            # Preparar datos para el generador PDF
            requisiciones_data = []
            for item in datos:
                requisiciones_data.append({
                    'folio': item['folio'],
                    'centro': item['centro'],
                    'estado': item['estado'],
                    'fecha_solicitud': item['fecha_solicitud'],
                    'total_productos': item['total_productos'],
                    'solicitante': item['solicitante']
                })
            
            filtros = {
                'fecha_generacion': timezone.now().strftime('%d/%m/%Y %H:%M')
            }
            if fecha_inicio:
                filtros['fecha_inicio'] = fecha_inicio
            if fecha_fin:
                filtros['fecha_fin'] = fecha_fin
            if estado:
                filtros['estado'] = estado
            if filtrar_por_centro and user_centro:
                filtros['centro'] = user_centro.nombre
            
            pdf_buffer = generar_reporte_requisiciones(requisiciones_data, filtros=filtros)
            
            response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = f"attachment; filename=Requisiciones_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            return response
        
        # Generar Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Requisiciones'
        
        ws.merge_cells('A1:G1')
        titulo_cell = ws['A1']
        titulo_cell.value = 'REPORTE DE REQUISICIONES'
        titulo_cell.font = Font(bold=True, size=14, color='632842')
        titulo_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        ws.merge_cells('A2:G2')
        fecha_cell = ws['A2']
        fecha_cell.value = f'Generado el {timezone.now().strftime("%d/%m/%Y %H:%M")}'
        fecha_cell.font = Font(size=10, italic=True)
        fecha_cell.alignment = Alignment(horizontal='center')
        
        ws.append([])
        
        headers = ['#', 'Folio', 'Centro', 'Fecha', 'Estado', 'Solicitante', 'Productos']
        ws.append(headers)
        
        header_fill = PatternFill(start_color='632842', end_color='632842', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        
        for cell in ws[4]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for idx, item in enumerate(datos, 1):
            ws.append([
                idx,
                item['folio'],
                item['centro'],
                item['fecha_solicitud'][:10] if item['fecha_solicitud'] else 'N/A',
                item['estado'],
                item['solicitante'],
                item['total_productos']
            ])
        
        for col, width in zip(['A','B','C','D','E','F','G'], [6,18,25,12,15,25,10]):
            ws.column_dimensions[col].width = width
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=Requisiciones_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        wb.save(response)
        
        return response
        
    except Exception as e:
        # traceback removido por seguridad (ISS-008)
        return Response({
            'error': 'Error al generar reporte de requisiciones',
            'mensaje': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def reporte_medicamentos_por_caducar(request):
    """Resumen JSON de productos con lotes proximos a caducar.
    
    SEGURIDAD: Filtra por centro del usuario si no es admin/farmacia.
    """
    try:
        # SEGURIDAD: Determinar filtro de centro
        user = request.user
        filtrar_por_centro = not is_farmacia_or_admin(user)
        user_centro = get_user_centro(user) if filtrar_por_centro else None
        
        centro_param = request.query_params.get('centro')
        if centro_param and is_farmacia_or_admin(user):
            try:
                user_centro = Centro.objects.get(pk=centro_param)
                filtrar_por_centro = True
            except Centro.DoesNotExist:
                pass
        
        dias = int(request.query_params.get('dias', 30))
        hoy = date.today()
        limite = hoy + timedelta(days=dias)
        lotes = Lote.objects.filter(
            deleted_at__isnull=True,
            cantidad_actual__gt=0,
            fecha_caducidad__gt=hoy,
            fecha_caducidad__lte=limite
        ).select_related('producto')
        
        # Aplicar filtro de centro
        if filtrar_por_centro and user_centro:
            lotes = lotes.filter(centro=user_centro)

        agregados = {}
        for lote in lotes:
            prod = lote.producto
            key = prod.id
            entry = agregados.setdefault(key, {
                'producto_id': prod.id,
                'clave': prod.clave,
                'descripcion': prod.descripcion,
                'stock_total': 0,
                'lotes': 0,
                'primer_vencimiento': None,
            })
            entry['lotes'] += 1
            entry['stock_total'] += lote.cantidad_actual
            fecha = lote.fecha_caducidad
            if entry['primer_vencimiento'] is None or fecha < entry['primer_vencimiento']:
                entry['primer_vencimiento'] = fecha

        resultados = sorted(
            agregados.values(),
            key=lambda x: x.get('primer_vencimiento') or limite
        )
        for res in resultados:
            fv = res['primer_vencimiento']
            res['primer_vencimiento'] = fv.isoformat() if fv else None

        return Response({
            'total_productos': len(resultados),
            'total_lotes': lotes.count(),
            'dias_configurados': dias,
            'resultados': resultados
        })
    except Exception as exc:
        # traceback removido por seguridad (ISS-008)
        return Response({'error': 'Error al obtener medicamentos por caducar', 'mensaje': str(exc)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def reporte_bajo_stock(request):
    """Productos con stock por debajo del minimo.
    
    SEGURIDAD: Filtra por centro del usuario si no es admin/farmacia.
    """
    try:
        # SEGURIDAD: Determinar filtro de centro
        user = request.user
        filtrar_por_centro = not is_farmacia_or_admin(user)
        user_centro = get_user_centro(user) if filtrar_por_centro else None
        
        centro_param = request.query_params.get('centro')
        if centro_param and is_farmacia_or_admin(user):
            try:
                user_centro = Centro.objects.get(pk=centro_param)
                filtrar_por_centro = True
            except Centro.DoesNotExist:
                pass
        
        productos = Producto.objects.filter(activo=True)
        resultados = []
        for prod in productos:
            lotes_query = prod.lotes.filter(
                deleted_at__isnull=True,
                estado='disponible'
            )
            # Aplicar filtro de centro
            if filtrar_por_centro and user_centro:
                lotes_query = lotes_query.filter(centro=user_centro)
            
            stock = lotes_query.aggregate(total=Sum('cantidad_actual'))['total'] or 0
            if stock < prod.stock_minimo:
                resultados.append({
                    'producto_id': prod.id,
                    'clave': prod.clave,
                    'descripcion': prod.descripcion,
                    'stock_actual': stock,
                    'stock_minimo': prod.stock_minimo,
                    'diferencia': prod.stock_minimo - stock
                })
        resultados = sorted(resultados, key=lambda x: x['diferencia'], reverse=True)
        return Response({'total': len(resultados), 'resultados': resultados})
    except Exception as exc:
        # traceback removido por seguridad (ISS-008)
        return Response({'error': 'Error al obtener productos en bajo stock', 'mensaje': str(exc)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def reporte_consumo(request):
    """
    Consumo (salidas) por producto en un rango de fechas.
    
    SEGURIDAD: Filtra por centro del usuario si no es admin/farmacia.
    """
    try:
        # SEGURIDAD: Determinar filtro de centro
        user = request.user
        filtrar_por_centro = not is_farmacia_or_admin(user)
        user_centro = get_user_centro(user) if filtrar_por_centro else None
        
        centro_param = request.query_params.get('centro')
        if centro_param and is_farmacia_or_admin(user):
            try:
                user_centro = Centro.objects.get(pk=centro_param)
                filtrar_por_centro = True
            except Centro.DoesNotExist:
                pass
        
        fecha_inicio = request.query_params.get('fecha_inicio')
        fecha_fin = request.query_params.get('fecha_fin')
        movimientos = Movimiento.objects.select_related('lote__producto').filter(tipo='salida')
        
        # Aplicar filtro de centro
        if filtrar_por_centro and user_centro:
            movimientos = movimientos.filter(lote__centro=user_centro)
        
        if fecha_inicio:
            movimientos = movimientos.filter(fecha__gte=fecha_inicio)
        if fecha_fin:
            movimientos = movimientos.filter(fecha__lte=fecha_fin)

        agregados = {}
        for mov in movimientos:
            prod = getattr(mov.lote, 'producto', None)
            if not prod:
                continue
            key = prod.id
            entry = agregados.setdefault(key, {
                'producto_id': prod.id,
                'clave': prod.clave,
                'descripcion': prod.descripcion,
                'total_salidas': 0
            })
            entry['total_salidas'] += abs(mov.cantidad or 0)

        resultados = sorted(agregados.values(), key=lambda x: x['total_salidas'], reverse=True)
        return Response({'total_productos': len(resultados), 'resultados': resultados})
    except Exception as exc:
        # traceback removido por seguridad (ISS-008)
        return Response({'error': 'Error al obtener consumo', 'mensaje': str(exc)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
def reportes_precarga(request):
    """
    Obtiene datos para precargar formularios de reportes.
    
    SEGURIDAD: Filtra centros y lotes seg�n el rol del usuario.
    - Admin/farmacia: ven todos los centros
    - Usuario de centro: solo ve su centro
    
    Retorna:
    - Lista de productos activos
    - Lista de centros activos (filtrada seg�n rol)
    - Tipos de movimiento disponibles
    """
    try:
        # SEGURIDAD: Determinar filtro de centro
        user = request.user
        es_admin_farmacia = is_farmacia_or_admin(user)
        user_centro = get_user_centro(user) if not es_admin_farmacia else None
        
        productos = list(Producto.objects.filter(activo=True).values('id', 'clave', 'descripcion').order_by('clave'))
        
        # Filtrar centros seg�n rol
        if es_admin_farmacia:
            centros = list(Centro.objects.filter(activo=True).values('id', 'clave', 'nombre').order_by('clave'))
        elif user_centro:
            centros = [{'id': user_centro.id, 'clave': user_centro.clave, 'nombre': user_centro.nombre}]
        else:
            centros = []
        
        # Filtrar lotes seg�n rol
        lotes_query = Lote.objects.filter(deleted_at__isnull=True)
        if not es_admin_farmacia and user_centro:
            lotes_query = lotes_query.filter(centro=user_centro)
        lotes = list(lotes_query.values('id', 'numero_lote', 'producto_id'))
        
        return Response({
            'productos': productos,
            'centros': centros,
            'lotes': lotes,
            'tipos_movimiento': ['ENTRADA', 'SALIDA']
        })
        
    except Exception as e:
        # traceback removido por seguridad (ISS-008)
        return Response({
            'error': 'Error al obtener datos de precarga',
            'mensaje': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class HojaRecoleccionViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet de solo lectura para Hojas de Recoleccion.
    Las hojas se generan automaticamente al autorizar requisiciones.
    """
    queryset = HojaRecoleccion.objects.select_related(
        'requisicion', 'requisicion__centro', 'generado_por'
    ).prefetch_related('detalles')
    serializer_class = HojaRecoleccionSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = CustomPagination

    def get_queryset(self):
        """Filtra por centro si el usuario no es admin/farmacia."""
        queryset = super().get_queryset()
        user = self.request.user
        if not user or not user.is_authenticated:
            return HojaRecoleccion.objects.none()
        if user.is_superuser:
            return queryset
        rol = getattr(user, 'rol', '').lower()
        if rol in ('admin', 'farmacia', 'administrador', 'usuario_farmacia'):
            return queryset
        # Filtrar por centro del usuario
        user_centro = getattr(user, 'centro', None)
        if user_centro:
            return queryset.filter(requisicion__centro=user_centro)
        return HojaRecoleccion.objects.none()

    @action(detail=True, methods=['get'])
    def verificar_integridad(self, request, pk=None):
        """Verifica que el hash de la hoja coincida con su contenido."""
        import hashlib
        import json
        hoja = self.get_object()
        contenido = json.dumps(hoja.contenido_json, sort_keys=True, ensure_ascii=False)
        hash_calculado = hashlib.sha256(contenido.encode('utf-8')).hexdigest()
        return Response({
            'folio': hoja.folio_hoja,
            'hash_almacenado': hoja.hash_contenido,
            'hash_calculado': hash_calculado,
            'integridad_ok': hash_calculado == hoja.hash_contenido
        })

    @action(detail=True, methods=['get'])
    def pdf(self, request, pk=None):
        """Genera y descarga el PDF de la hoja de recolecci�n."""
        from core.utils.pdf_generator import generar_hoja_recoleccion
        
        hoja = self.get_object()
        requisicion = hoja.requisicion
        
        try:
            pdf_buffer = generar_hoja_recoleccion(requisicion)
            
            response = HttpResponse(
                pdf_buffer.getvalue(),
                content_type='application/pdf'
            )
            folio_safe = (hoja.folio_hoja or f'HOJA-{hoja.id}').replace('/', '-')
            response['Content-Disposition'] = f'attachment; filename="Hoja_Recoleccion_{folio_safe}.pdf"'
            
            return response
            
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            return Response({
                'error': 'Error al generar PDF',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)








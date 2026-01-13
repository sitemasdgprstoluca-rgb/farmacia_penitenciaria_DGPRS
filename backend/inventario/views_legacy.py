from rest_framework import viewsets, status, serializers, mixins
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.pagination import PageNumberPagination
from django.core.paginator import InvalidPage
from django.core.cache import cache
from django.core.exceptions import ValidationError
from django.http import HttpResponse
from django.db import transaction
from django.db.models import Q, Sum, Count, F, IntegerField, Subquery, OuterRef
from django.db.models.functions import Coalesce
from django.utils import timezone
from django.conf import settings
from datetime import datetime, timedelta, date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os
import logging
import random  # ISS-FIX: Necesario para generar números de requisición
from io import BytesIO

# ISS-004 FIX (audit9): Pillow para validación profunda de imágenes
try:
    from PIL import Image
    PILLOW_AVAILABLE = True
except ImportError:
    PILLOW_AVAILABLE = False
    Image = None

# ISS-011, ISS-021, ISS-030: Import de servicios transaccionales
from inventario.services import CentroPermissionMixin

logger = logging.getLogger(__name__)


# -----------------------------------------------------------
# VALIDADORES DE IMPORTACIÓN EXCEL
# -----------------------------------------------------------

# Magic bytes para formatos Excel válidos
# XLSX/XLSM son archivos ZIP (PK\x03\x04)
# XLS antiguo usa formato OLE2 (Compound File Binary Format)
# ISS-008 FIX (audit7): XLSM eliminado - archivos con macros NO permitidos
EXCEL_MAGIC_BYTES = {
    '.xlsx': b'PK\x03\x04',  # ZIP archive (Office Open XML)
    # '.xlsm': REMOVIDO - ISS-008 FIX: No permitir archivos con macros
    '.xls': b'\xD0\xCF\x11\xE0',  # OLE2 Compound Document
}

# ISS-008 FIX (audit7): Extensiones explícitamente prohibidas por seguridad
EXCEL_EXTENSIONS_BLOCKED = {'.xlsm', '.xlsb', '.xltm', '.xla', '.xlam'}  # Formatos con macros

# ISS-006: Magic bytes y MIME types para imágenes permitidas en firmas
IMAGE_MAGIC_BYTES = {
    '.jpg': [b'\xff\xd8\xff'],  # JPEG
    '.jpeg': [b'\xff\xd8\xff'],  # JPEG
    '.png': [b'\x89PNG\r\n\x1a\n'],  # PNG
    '.gif': [b'GIF87a', b'GIF89a'],  # GIF
    '.webp': [b'RIFF'],  # WebP (RIFF container)
}

ALLOWED_IMAGE_EXTENSIONS = {'.jpg', '.jpeg', '.png', '.gif', '.webp'}
ALLOWED_IMAGE_MIMES = {
    'image/jpeg', 'image/png', 'image/gif', 'image/webp'
}

# ISS-005 FIX (audit7): Magic bytes para validación de PDF
PDF_MAGIC_BYTES = b'%PDF-'  # Todos los PDFs válidos empiezan con este header
PDF_MAX_SIZE_MB = 10  # Tamaño máximo en MB
PDF_MAX_SIZE_BYTES = PDF_MAX_SIZE_MB * 1024 * 1024

# ISS-002 FIX (audit9): Content-Types válidos para PDF (validación estricta)
PDF_VALID_CONTENT_TYPES = {'application/pdf', 'application/x-pdf'}

# ISS-001 FIX (audit9): Estado inicial SOLO borrador para TODOS los usuarios
# El flujo jerárquico OBLIGA: borrador → pendiente_admin → pendiente_director → enviada
# NUNCA se puede crear una requisición directamente en 'enviada'
ESTADO_INICIAL_UNICO = 'borrador'
# Mantener por compatibilidad pero DEPRECADO - usar ESTADO_INICIAL_UNICO
ESTADOS_INICIALES_VALIDOS = {'borrador'}  # ISS-001 FIX: Eliminado 'enviada'
ESTADO_INICIAL_CENTRO = 'borrador'
ESTADO_INICIAL_FARMACIA = 'borrador'


def validar_archivo_pdf(file, max_size_mb=PDF_MAX_SIZE_MB):
    """
    ISS-005 FIX (audit7): Valida archivo PDF antes de guardarlo.
    
    Validaciones:
    1. Archivo presente y con nombre válido
    2. Extensión .pdf
    3. Tamaño dentro de límites
    4. Magic bytes correctos (%PDF-)
    5. Content-Type MIME válido (si disponible)
    
    Retorna: (es_valido, mensaje_error)
    """
    # 1. Validar que hay archivo
    if not file:
        return False, 'No se recibió archivo PDF'
    
    # 2. Validar nombre obligatorio
    nombre = getattr(file, 'name', None)
    if not nombre or not nombre.strip():
        return False, 'El archivo debe tener un nombre válido'
    
    # 3. Validar extensión
    extension = ('.' + nombre.split('.')[-1].lower()) if '.' in nombre else ''
    if extension != '.pdf':
        return False, f'Solo se permiten archivos PDF (.pdf). Extensión recibida: {extension}'
    
    # 4. Validar tamaño
    max_size_bytes = max_size_mb * 1024 * 1024
    if hasattr(file, 'size') and file.size > max_size_bytes:
        return False, f'El archivo excede el tamaño máximo de {max_size_mb}MB ({file.size / 1024 / 1024:.2f}MB)'
    
    # 5. Validar magic bytes - leer primeros bytes con límite
    try:
        buffer_result, bytes_leidos = leer_archivo_con_limite(file, max_size_bytes + 1024)
        
        if buffer_result is None:
            return False, bytes_leidos  # bytes_leidos contiene el mensaje de error
        
        # Validar tamaño real
        if bytes_leidos > max_size_bytes:
            return False, f'El archivo excede el tamaño máximo de {max_size_mb}MB (tamaño real: {bytes_leidos / 1024 / 1024:.2f}MB)'
        
        # Verificar magic bytes de PDF
        header = buffer_result.read(8)
        buffer_result.seek(0)  # Restaurar posición
        
        if not header.startswith(PDF_MAGIC_BYTES):
            logger.warning(
                f"ISS-005: Archivo {nombre} rechazado - magic bytes incorrectos: {header[:8]!r}"
            )
            return False, 'El archivo no es un PDF válido (magic bytes incorrectos)'
        
    except Exception as e:
        logger.error(f"ISS-005: Error validando PDF {nombre}: {e}")
        return False, f'Error al validar el archivo PDF: {str(e)}'
    
    # 6. ISS-002 FIX (audit9): Validar Content-Type ESTRICTAMENTE
    # No solo advertir, BLOQUEAR archivos con Content-Type sospechoso
    content_type = getattr(file, 'content_type', None)
    if content_type:
        if content_type not in PDF_VALID_CONTENT_TYPES:
            logger.error(
                f"ISS-002: Archivo {nombre} BLOQUEADO - content-type inválido: {content_type}. "
                f"Solo se permiten: {PDF_VALID_CONTENT_TYPES}"
            )
            return False, (
                f'Content-Type inválido: {content_type}. '
                f'Solo se permiten archivos PDF (application/pdf)'
            )
    
    return True, None


def leer_archivo_con_limite(file, max_bytes):
    """
    ISS-001: Lee un archivo con límite estricto de bytes.
    
    Previene DoS por archivos enormes que no declaran tamaño correcto.
    Lee en chunks y aborta si se excede el límite.
    
    Args:
        file: Archivo a leer (UploadedFile o similar)
        max_bytes: Máximo de bytes permitidos
    
    Returns:
        tuple: (BytesIO con contenido, bytes_leidos) o (None, error_message)
    """
    CHUNK_SIZE = 64 * 1024  # 64KB chunks
    buffer = BytesIO()
    bytes_leidos = 0
    
    try:
        # Asegurar que estamos al inicio del archivo
        if hasattr(file, 'seek'):
            file.seek(0)
        
        while True:
            chunk = file.read(CHUNK_SIZE)
            if not chunk:
                break
            
            bytes_leidos += len(chunk)
            
            # ISS-001: Corte estricto si excede el límite
            if bytes_leidos > max_bytes:
                logger.warning(
                    f"Archivo rechazado: tamaño real ({bytes_leidos} bytes) "
                    f"excede límite ({max_bytes} bytes)"
                )
                return None, f'El archivo excede el tamaño máximo permitido ({max_bytes / 1024 / 1024:.1f}MB)'
            
            buffer.write(chunk)
        
        buffer.seek(0)
        return buffer, bytes_leidos
        
    except Exception as e:
        logger.error(f"Error leyendo archivo: {e}")
        return None, f'Error al leer el archivo: {str(e)}'


def validar_archivo_imagen(file, max_size_mb=2):
    """
    ISS-006: Valida archivo de imagen antes de guardarlo.
    
    Validaciones:
    1. Archivo presente y con nombre válido
    2. Extensión permitida (.jpg, .jpeg, .png, .gif, .webp)
    3. Tamaño dentro de límites (default 2MB)
    4. Magic bytes correctos (contenido real corresponde a imagen)
    5. Content-Type MIME válido (si disponible)
    
    Retorna: (es_valido, mensaje_error)
    """
    # 1. Validar que hay archivo
    if not file:
        return False, 'No se recibió archivo de imagen'
    
    # 2. Validar nombre obligatorio
    nombre = getattr(file, 'name', None)
    if not nombre or not nombre.strip():
        return False, 'El archivo debe tener un nombre válido'
    
    nombre_lower = nombre.lower().strip()
    ext = os.path.splitext(nombre_lower)[1]
    
    # 3. Validar extensión
    if not ext:
        return False, 'El archivo debe tener una extensión (.jpg, .png, etc.)'
    
    if ext not in ALLOWED_IMAGE_EXTENSIONS:
        return False, f'Extensión no permitida: {ext}. Use: {", ".join(ALLOWED_IMAGE_EXTENSIONS)}'
    
    # 4. Validar tamaño
    max_size_bytes = max_size_mb * 1024 * 1024
    file_size = getattr(file, 'size', None)
    if file_size is not None and file_size > max_size_bytes:
        return False, f'Imagen demasiado grande: {file_size / 1024 / 1024:.1f}MB. Máximo: {max_size_mb}MB'
    
    # 5. Validar Content-Type MIME si está disponible
    content_type = getattr(file, 'content_type', None)
    if content_type and content_type.lower() not in ALLOWED_IMAGE_MIMES:
        logger.warning(
            f"Archivo rechazado: content_type {content_type} no permitido. "
            f"Nombre: {nombre}"
        )
        return False, f'Tipo de archivo no permitido: {content_type}. Use imágenes JPG, PNG, GIF o WebP'
    
    # 6. Validar magic bytes - contenido real del archivo
    try:
        pos = file.tell() if hasattr(file, 'tell') else 0
        header = file.read(16)  # Leer suficientes bytes
        
        if hasattr(file, 'seek'):
            file.seek(pos)
        
        if not header or len(header) < 4:
            return False, 'Archivo vacío o corrupto'
        
        # Verificar magic bytes según extensión
        expected_magics = IMAGE_MAGIC_BYTES.get(ext, [])
        if expected_magics:
            match = False
            for magic in expected_magics:
                if header.startswith(magic):
                    match = True
                    break
            
            if not match:
                logger.warning(
                    f"Archivo imagen rechazado: extensión {ext} pero magic bytes incorrectos. "
                    f"Header: {header[:8].hex()}, Nombre: {nombre}"
                )
                return False, f'El contenido del archivo no corresponde a una imagen {ext} válida'
    
    except Exception as e:
        logger.error(f"Error validando magic bytes de imagen: {e}")
        return False, 'Error al validar el contenido del archivo de imagen'
    
    # 7. ISS-004 FIX (audit9): Validación profunda con Pillow
    # Intenta abrir la imagen para detectar archivos corruptos o maliciosos
    if PILLOW_AVAILABLE and Image is not None:
        try:
            pos = file.tell() if hasattr(file, 'tell') else 0
            
            # Leer contenido completo para Pillow
            if hasattr(file, 'seek'):
                file.seek(0)
            content = file.read()
            
            # Restaurar posición
            if hasattr(file, 'seek'):
                file.seek(pos)
            
            # Intentar abrir con Pillow
            img_buffer = BytesIO(content)
            with Image.open(img_buffer) as img:
                # Verificar que realmente se puede decodificar
                img.verify()
                
                # Reabrir para validar dimensiones (verify() invalida el handle)
                img_buffer.seek(0)
                with Image.open(img_buffer) as img_check:
                    width, height = img_check.size
                    
                    # Límites de dimensiones para evitar DoS
                    MAX_IMAGE_DIMENSION = 4096  # 4K máximo
                    if width > MAX_IMAGE_DIMENSION or height > MAX_IMAGE_DIMENSION:
                        logger.warning(
                            f"Imagen rechazada por dimensiones excesivas: {width}x{height}, "
                            f"máximo: {MAX_IMAGE_DIMENSION}x{MAX_IMAGE_DIMENSION}"
                        )
                        return False, f'Imagen demasiado grande: {width}x{height}. Máximo: {MAX_IMAGE_DIMENSION}x{MAX_IMAGE_DIMENSION} píxeles'
                    
                    # Verificar formato reportado por Pillow
                    pillow_format = img_check.format
                    expected_formats = {
                        '.jpg': ['JPEG'],
                        '.jpeg': ['JPEG'],
                        '.png': ['PNG'],
                        '.gif': ['GIF'],
                        '.webp': ['WEBP'],
                    }
                    
                    if ext in expected_formats and pillow_format not in expected_formats[ext]:
                        logger.warning(
                            f"Imagen rechazada: extensión {ext} pero Pillow detectó {pillow_format}. "
                            f"Nombre: {nombre}"
                        )
                        return False, f'El contenido del archivo ({pillow_format}) no corresponde a la extensión {ext}'
            
            # ISS-006 FIX (audit11): Asegurar que el stream quede en posición original
            if hasattr(file, 'seek'):
                file.seek(0)
                    
        except Exception as e:
            logger.warning(
                f"Imagen rechazada: Pillow no pudo abrir/verificar el archivo. "
                f"Error: {e}, Nombre: {nombre}"
            )
            return False, 'Imagen corrupta o formato no válido - no se puede procesar'
    else:
        # ISS-006 FIX (audit11): Rechazar imágenes si Pillow no está disponible
        # La validación solo con magic bytes no es suficiente para seguridad
        logger.error(
            "Pillow NO disponible - rechazando imagen por seguridad. "
            "Instale Pillow con: pip install Pillow"
        )
        return False, 'Validación de imágenes no disponible. Contacte al administrador.'
    
    # ISS-006 FIX (audit11): Asegurar stream en posición 0 para el caller
    if hasattr(file, 'seek'):
        file.seek(0)
    
    return True, None


def validar_archivo_excel(file):
    """
    ISS-008 FIX (audit7): Valida archivo Excel antes de procesarlo con múltiples capas de seguridad.
    
    Validaciones:
    1. Archivo presente y con nombre válido
    2. Extensión permitida (.xlsx, .xls) - NO xlsm ni otros con macros
    3. Tamaño dentro de límites
    4. Magic bytes correctos (contenido real coincide con extensión)
    
    ISS-008 FIX (audit7): Archivos con macros (.xlsm, .xlsb, etc) están BLOQUEADOS
    por seguridad. Solo se permiten formatos sin código ejecutable.
    
    Retorna: (es_valido, mensaje_error)
    """
    # 1. Validar que hay archivo
    if not file:
        return False, 'No se recibió archivo'
    
    # 2. ISS-001: Validar nombre obligatorio - rechazar archivos sin nombre
    nombre = getattr(file, 'name', None)
    if not nombre or not nombre.strip():
        return False, 'El archivo debe tener un nombre válido'
    
    nombre_lower = nombre.lower().strip()
    ext = os.path.splitext(nombre_lower)[1]
    
    # 3. Validar extensión
    if not ext:
        return False, 'El archivo debe tener una extensión (.xlsx o .xls)'
    
    # ISS-008 FIX (audit7): Bloquear extensiones con macros ANTES de cualquier otra validación
    if ext in EXCEL_EXTENSIONS_BLOCKED:
        logger.warning(
            f"ISS-008: Archivo rechazado por extensión con macros: {nombre}. "
            f"Extensiones bloqueadas: {EXCEL_EXTENSIONS_BLOCKED}"
        )
        return False, (
            f'Extensión {ext} no permitida por seguridad. '
            f'Los archivos con macros (.xlsm, .xlsb, etc.) están bloqueados. '
            f'Por favor convierta a .xlsx (sin macros) antes de importar.'
        )
    
    extensiones_permitidas = getattr(settings, 'IMPORT_ALLOWED_EXTENSIONS', ['.xlsx', '.xls'])
    if ext not in extensiones_permitidas:
        return False, f'Extensión no permitida: {ext}. Use: {", ".join(extensiones_permitidas)}'
    
    # 4. Validar tamaño declarado ANTES de leer contenido (primera línea de defensa)
    max_size_mb = getattr(settings, 'IMPORT_MAX_FILE_SIZE_MB', 10)
    max_size_bytes = max_size_mb * 1024 * 1024
    
    file_size = getattr(file, 'size', None)
    if file_size is not None and file_size > max_size_bytes:
        return False, f'Archivo demasiado grande: {file_size / 1024 / 1024:.1f}MB. Máximo: {max_size_mb}MB'
    
    # 5. ISS-001: Validar magic bytes - contenido real del archivo
    try:
        # Guardar posición actual y leer primeros bytes
        pos = file.tell() if hasattr(file, 'tell') else 0
        header = file.read(8)  # Leer suficientes bytes para identificar formato
        
        # Restaurar posición para que el archivo pueda ser procesado después
        if hasattr(file, 'seek'):
            file.seek(pos)
        
        if not header or len(header) < 4:
            return False, 'Archivo vacío o corrupto'
        
        # Verificar magic bytes según extensión
        expected_magic = EXCEL_MAGIC_BYTES.get(ext)
        if expected_magic:
            if not header.startswith(expected_magic):
                # El contenido no coincide con la extensión declarada
                logger.warning(
                    f"Archivo rechazado: extensión {ext} pero magic bytes incorrectos. "
                    f"Header: {header[:8].hex()}"
                )
                return False, f'El contenido del archivo no corresponde a un archivo {ext} válido'
    except Exception as e:
        logger.error(f"Error validando magic bytes: {e}")
        return False, 'Error al validar el contenido del archivo'
    
    return True, None


def cargar_workbook_seguro(file):
    """
    ISS-001: Carga un workbook Excel con límites de seguridad estrictos.
    
    1. Lee el archivo con límite real de bytes (no confía en file.size)
    2. Usa read_only=True para streaming y menor uso de memoria
    3. Usa data_only=True para ignorar fórmulas (previene ataques por fórmulas)
    
    Returns:
        tuple: (workbook, error_message) - workbook es None si hay error
    """
    max_size_mb = getattr(settings, 'IMPORT_MAX_FILE_SIZE_MB', 10)
    max_size_bytes = max_size_mb * 1024 * 1024
    
    # Leer archivo con límite estricto de bytes
    buffer, resultado = leer_archivo_con_limite(file, max_size_bytes)
    
    if buffer is None:
        # resultado contiene el mensaje de error
        return None, resultado
    
    try:
        # ISS-001: Usar read_only=True para streaming y data_only=True para ignorar fórmulas
        # Esto reduce significativamente el uso de memoria en archivos grandes
        wb = openpyxl.load_workbook(
            buffer, 
            read_only=True,  # Streaming mode - no carga todo en memoria
            data_only=True   # Ignora fórmulas - solo valores (previene ataques)
        )
        return wb, None
    except Exception as e:
        logger.error(f"Error cargando workbook: {e}")
        return None, f'Error al procesar el archivo Excel: {str(e)}'


def validar_filas_excel(ws):
    """
    Valida número de filas en worksheet con streaming y corte temprano.
    
    ISS-003: Optimizado para no recorrer archivos enormes innecesariamente.
    Se detiene al superar el límite máximo permitido.
    
    Retorna: (es_valido, mensaje_error, num_filas)
    """
    max_rows = getattr(settings, 'IMPORT_MAX_ROWS', 5000)
    # Margen adicional para detener el conteo (evitar contar millones de filas)
    cutoff = max_rows + 1
    
    num_filas = 0
    # ISS-003: Streaming con corte temprano - detenerse apenas superamos el límite
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Solo contar filas que tienen al menos un valor no vacío
        if any(cell is not None and str(cell).strip() for cell in row):
            num_filas += 1
            if num_filas > max_rows:
                # Corte temprano: ya sabemos que excede el límite
                return False, f'El archivo excede el máximo de {max_rows} filas permitidas', num_filas
    
    return True, None, num_filas

from core.models import Producto, Lote, Movimiento, Centro, Requisicion, DetalleRequisicion, HojaRecoleccion, LoteDocumento
from core.serializers import (
    ProductoSerializer, LoteSerializer, MovimientoSerializer, 
    CentroSerializer, RequisicionSerializer, RequisicionListSerializer, HojaRecoleccionSerializer, LoteDocumentoSerializer
)

from django.contrib.auth import get_user_model
from core.permissions import (
    IsCentroRole, IsCentroCanManageInventory, RoleHelper,  # ISS-MEDICO FIX
    IsFarmaciaAdminOrVistaReadOnly, IsCentroOwnResourcesOnly  # ISS-MEDICO FIX: permisos restrictivos
)
from core.constants import (
    ESTADOS_REQUISICION,
    PAGINATION_DEFAULT_PAGE_SIZE,
    PAGINATION_MAX_PAGE_SIZE,
    REQUISICION_GRUPOS_ESTADO,
    TRANSICIONES_REQUISICION,
    PERMISOS_FLUJO_REQUISICION,
)

User = get_user_model()


# ============================================================================
# HELPERS DE SEGURIDAD - Filtrado por centro para roles no-admin
# ISS-AUDIT FIX: Separación de funciones para claridad y seguridad
# ============================================================================

def is_farmacia_or_admin(user):
    """
    ISS-AUDIT FIX: Verifica si el usuario es farmacia o admin (roles de escritura).
    
    IMPORTANTE: Esta función NO incluye 'vista' - usar has_global_read_access()
    para operaciones de solo lectura que incluyan vista.
    
    Roles con acceso de escritura:
    - admin: admin_sistema, superusuario, administrador
    - farmacia: farmacia, admin_farmacia, farmaceutico, usuario_farmacia
    """
    if not user or not user.is_authenticated:
        return False
    if user.is_superuser:
        return True
    
    rol = (getattr(user, 'rol', '') or '').lower()
    
    # ISS-AUDIT FIX: Solo roles admin y farmacia (NO vista)
    ROLES_ESCRITURA = {
        # Admin
        'admin', 'admin_sistema', 'superusuario', 'administrador',
        # Farmacia
        'farmacia', 'admin_farmacia', 'farmaceutico', 'usuario_farmacia',
    }
    
    if rol in ROLES_ESCRITURA:
        return True
    
    # Verificar grupos (por si el rol no esta en campo directo)
    group_names = {g.name.upper() for g in user.groups.all()}
    GRUPOS_ESCRITURA = {'FARMACIA_ADMIN', 'FARMACEUTICO'}
    
    return bool(group_names & GRUPOS_ESCRITURA)


def has_global_read_access(user):
    """
    ISS-AUDIT FIX: Verifica si el usuario puede leer datos globales (incluye vista).
    
    Usar esta función para operaciones de SOLO LECTURA donde el rol vista
    debe tener acceso.
    
    Roles con lectura global:
    - admin, farmacia (heredan de is_farmacia_or_admin)
    - vista: vista, usuario_vista (solo lectura)
    """
    # Admin y farmacia siempre tienen acceso de lectura
    if is_farmacia_or_admin(user):
        return True
    
    if not user or not user.is_authenticated:
        return False
    
    rol = (getattr(user, 'rol', '') or '').lower()
    
    # Roles de solo lectura global
    ROLES_VISTA = {'vista', 'usuario_vista'}
    
    if rol in ROLES_VISTA:
        return True
    
    # Verificar grupos
    group_names = {g.name.upper() for g in user.groups.all()}
    return 'VISTA_USER' in group_names


def get_user_centro(user):
    """Obtiene el centro asignado al usuario."""
    return getattr(user, 'centro', None) or getattr(getattr(user, 'profile', None), 'centro', None)


def _get_rol_efectivo(user):
    """
    ISS-DIRECTOR FIX: Obtiene el rol efectivo del usuario para validaciones de flujo.
    
    Esta función debe usarse en TODAS las validaciones de permisos del flujo
    para garantizar consistencia con lo que se envía al frontend.
    
    La lógica es idéntica a _resolve_rol en serializers.py pero devuelve
    el rol en minúsculas para usar con PERMISOS_FLUJO_REQUISICION.
    
    IMPORTANTE: Ahora infiere roles específicos (director_centro, administrador_centro, medico)
    basándose en los permisos personalizados del usuario cuando el campo rol está vacío.
    
    Returns:
        str: Rol normalizado en minúsculas (medico, administrador_centro, director_centro, etc.)
    """
    if not user:
        return 'sin_rol'
    if user.is_superuser:
        return 'admin'
    
    # Obtener rol del campo usuario
    rol_campo = (getattr(user, 'rol', '') or '').lower().strip()
    
    # Si hay rol en el campo, usarlo directamente
    if rol_campo and rol_campo not in ['null', 'none', '']:
        return rol_campo
    
    # ISS-DIRECTOR FIX: Inferir rol si el campo está vacío
    # Esto evita que usuarios con rol vacío pierdan sus permisos específicos
    
    # Si es staff pero no superuser, es farmacia
    if getattr(user, 'is_staff', False):
        return 'farmacia'
    
    # ISS-DIRECTOR FIX: Inferir rol específico por permisos personalizados
    # Verificar permisos del flujo para roles de centro específicos
    if getattr(user, 'perm_autorizar_director', None) is True:
        return 'director_centro'
    
    if getattr(user, 'perm_autorizar_admin', None) is True:
        return 'administrador_centro'
    
    if getattr(user, 'perm_crear_requisicion', None) is True:
        # Usuario con permiso de crear requisiciones y centro = médico
        centro = getattr(user, 'centro', None) or getattr(user, 'centro_id', None)
        if centro:
            return 'medico'
    
    # Si tiene centro asignado sin permisos específicos, es usuario de centro genérico
    centro = getattr(user, 'centro', None) or getattr(user, 'centro_id', None)
    if centro:
        return 'centro'
    
    # Default: vista (más restrictivo)
    return 'vista'


def invalidar_cache_dashboard(centro_id=None):
    """
    ISS-005: Invalida el caché del dashboard cuando hay cambios en datos.
    
    Se llama automáticamente al registrar movimientos, crear requisiciones, etc.
    
    Args:
        centro_id: ID del centro afectado (opcional). Si es None, invalida solo el global.
                   Si se pasa, invalida tanto el del centro como el global.
    """
    try:
        # Siempre invalidar el global
        cache.delete('dashboard_resumen_global')
        cache.delete('dashboard_graficas_global')
        
        # Si hay un centro específico, invalidar también ese
        if centro_id:
            cache.delete(f'dashboard_resumen_{centro_id}')
            cache.delete(f'dashboard_graficas_{centro_id}')
    except Exception as e:
        # No fallar si hay problemas con el caché
        logger.warning(f'Error al invalidar caché del dashboard: {e}')


def registrar_movimiento_stock(*, lote, tipo, cantidad, usuario=None, centro=None, requisicion=None, observaciones='', skip_centro_check=False, subtipo_salida=None, numero_expediente=None):
    """
    Helper central para registrar un movimiento y actualizar cantidad_actual del lote.
    
    ISS-002: Incluye validación de pertenencia de centro para prevenir manipulación
    de inventario entre centros.
    
    ISS-003 FIX: Los ajustes negativos requieren observaciones obligatorias con
    longitud mínima para auditoría y prevención de robo hormiga.
    
    MEJORA FLUJO 5: Soporte para subtipo_salida y numero_expediente para
    trazabilidad de pacientes en salidas por receta médica.
    
    Args:
        lote: Instancia del Lote a modificar
        tipo: 'entrada', 'salida' o 'ajuste'
        cantidad: Cantidad a mover (positivo)
        usuario: Usuario que realiza la operación (opcional)
        centro: Centro donde se registra el movimiento (opcional)
        requisicion: Requisición asociada (opcional)
        observaciones: Texto descriptivo (obligatorio para ajustes negativos)
        skip_centro_check: Si True, omite validación de pertenencia (solo para operaciones
                          de sistema/admin que ya validaron permisos)
        subtipo_salida: Tipo de salida (receta, consumo_interno, merma, etc.)
        numero_expediente: Número de expediente del paciente (para recetas)
    
    Raises:
        ValidationError: Si los datos son inválidos o hay problemas de autorización
    
    Returns:
        tuple: (movimiento_creado, lote_actualizado)
    """
    tipo_normalizado = (tipo or '').lower()
    if tipo_normalizado not in ('entrada', 'salida', 'ajuste'):
        raise serializers.ValidationError({'tipo': 'Tipo de movimiento no valido'})
    if cantidad is None:
        raise serializers.ValidationError({'cantidad': 'Cantidad requerida'})
    try:
        cantidad_int = int(cantidad)
    except (TypeError, ValueError):
        raise serializers.ValidationError({'cantidad': 'La cantidad debe ser un numero entero'})

    # =========================================================================
    # ISS-003 FIX: Validación de observaciones para ajustes negativos
    # =========================================================================
    # Los ajustes que reducen stock (negativos) requieren justificación obligatoria
    # para prevenir robo hormiga y asegurar trazabilidad de pérdidas
    LONGITUD_MINIMA_OBSERVACION = 10
    es_ajuste_negativo = tipo_normalizado == 'ajuste' and cantidad_int < 0
    
    if es_ajuste_negativo:
        if not observaciones or len(observaciones.strip()) < LONGITUD_MINIMA_OBSERVACION:
            raise serializers.ValidationError({
                'observaciones': f'Los ajustes negativos requieren una justificación de al menos {LONGITUD_MINIMA_OBSERVACION} caracteres. '
                                 f'Explique el motivo del ajuste (merma, caducidad, rotura, etc.)'
            })

    # =========================================================================
    # ISS-002: Validación de pertenencia de centro
    # =========================================================================
    # Si se pasa un centro explícito, verificar que el lote pertenezca a ese centro
    # Esto previene que usuarios de un centro manipulen stock de otros centros
    if not skip_centro_check and centro is not None:
        lote_centro = getattr(lote, 'centro', None)
        if lote_centro is not None and lote_centro.pk != centro.pk:
            logger.warning(
                f"Intento de operación de stock rechazado: "
                f"usuario={getattr(usuario, 'username', 'N/A')}, "
                f"lote_centro={lote_centro.pk}, centro_solicitado={centro.pk}"
            )
            raise serializers.ValidationError({
                'centro': 'El lote no pertenece al centro especificado. Operación no autorizada.'
            })
    
    # Si el usuario no es admin/farmacia, verificar que tenga acceso al centro del lote
    if not skip_centro_check and usuario is not None and hasattr(usuario, 'is_authenticated') and usuario.is_authenticated:
        if not is_farmacia_or_admin(usuario):
            user_centro = get_user_centro(usuario)
            lote_centro = getattr(lote, 'centro', None)
            if lote_centro is not None and user_centro is not None:
                if user_centro.pk != lote_centro.pk:
                    logger.warning(
                        f"Acceso no autorizado a lote de otro centro: "
                        f"usuario={usuario.username}, user_centro={user_centro.pk}, "
                        f"lote_centro={lote_centro.pk}"
                    )
                    raise serializers.ValidationError({
                        'lote': 'No tiene permiso para operar sobre lotes de otros centros.'
                    })

    delta = cantidad_int
    if tipo_normalizado == 'salida' and delta > 0:
        delta = -delta
    if tipo_normalizado == 'entrada' and delta < 0:
        delta = abs(delta)

    with transaction.atomic():
        lote_ref = Lote.objects.select_for_update().get(pk=lote.pk)
        stock_disponible = lote_ref.cantidad_actual

        if delta < 0 and abs(delta) > stock_disponible:
            raise serializers.ValidationError({
                'cantidad': f'Stock insuficiente en el lote (disponible {stock_disponible}).'
            })

        nuevo_stock = stock_disponible + delta
        if nuevo_stock < 0:
            raise serializers.ValidationError({'cantidad': 'La operacion dejaria el lote con stock negativo'})

        # ISS-014 FIX: Usar F() para actualización atómica del stock
        # Aunque tenemos select_for_update, F() proporciona doble seguridad contra race conditions
        update_dict = {
            'cantidad_actual': F('cantidad_actual') + delta,
            'activo': nuevo_stock > 0,
            'updated_at': timezone.now()
        }
        
        # ISS-FIX: Para entradas (reabastecimiento), SUMAR a cantidad_inicial
        # Esto refleja que se recibió más mercancía del mismo contrato/lote
        # Ej: cantidad_inicial=84, entrada=50 → cantidad_inicial=134
        # FIX: Usar abs(cantidad_int) para asegurar valor entero positivo
        if tipo_normalizado == 'entrada':
            update_dict['cantidad_inicial'] = F('cantidad_inicial') + abs(cantidad_int)
        
        Lote.objects.filter(pk=lote_ref.pk).update(**update_dict)
        lote_ref.refresh_from_db()

        # Crear movimiento con campos correctos de la BD
        # ISS-MEDICO FIX v2: La cantidad siempre se guarda como positiva en BD
        # (hay un CHECK constraint chk_movimiento_cantidad_positiva)
        # El tipo (entrada/salida/ajuste) indica si suma o resta
        movimiento = Movimiento(
            tipo=tipo_normalizado,
            producto=lote_ref.producto,
            lote=lote_ref,
            centro_destino=centro if tipo_normalizado == 'entrada' else None,
            centro_origen=centro if tipo_normalizado != 'entrada' else None,
            requisicion=requisicion,
            usuario=usuario if usuario and getattr(usuario, 'is_authenticated', False) else None,
            cantidad=abs(cantidad_int),  # Siempre positivo, el tipo indica la operación
            motivo=observaciones or '',
            # MEJORA FLUJO 5: Campos de trazabilidad para pacientes
            subtipo_salida=subtipo_salida if tipo_normalizado == 'salida' else None,
            numero_expediente=numero_expediente if tipo_normalizado == 'salida' and subtipo_salida == 'receta' else None
        )
        # Guardar stock previo para evitar fallos de validacion al crear el movimiento
        movimiento._stock_pre_movimiento = stock_disponible
        movimiento.save()
        
        # ISS-005: Invalidar caché del dashboard al registrar movimientos
        centro_afectado = centro.id if centro else (lote_ref.centro.id if lote_ref.centro else None)
        invalidar_cache_dashboard(centro_afectado)

    return movimiento, lote_ref


class CustomPagination(PageNumberPagination):
    """Paginacin unificada para todos los listados del API."""
    page_size = PAGINATION_DEFAULT_PAGE_SIZE
    page_size_query_param = 'page_size'
    max_page_size = PAGINATION_MAX_PAGE_SIZE

    def paginate_queryset(self, queryset, request, view=None):
        try:
            return super().paginate_queryset(queryset, request, view=view)
        except InvalidPage:
            # Si la pgina solicitada no existe, devolver la ltima disponible sin 404
            self.page = self.paginator.page(self.paginator.num_pages or 1)
            self.request = request
            return list(self.page)

# ISS-002: UserSerializer eliminado - usar el de core.serializers
# from core.serializers import UserSerializer (si se necesita)

# NOTA: UserViewSet está en core/views.py - importar desde allí

class ProductoViewSet(viewsets.ModelViewSet):
    queryset = Producto.objects.prefetch_related('lotes').all()  # HALLAZGO #8 FIX
    serializer_class = ProductoSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = CustomPagination

    def get_permissions(self):
        """
        Permisos personalizados por accion:
        - list, retrieve: IsAuthenticated
        - create, update, destroy, toggle_activo, importar_excel, exportar_excel, auditoria: IsFarmaciaRole
        """
        acciones_farmacia = [
            'create', 'update', 'partial_update', 'destroy', 
            'toggle_activo', 'importar_excel', 'exportar_excel', 'auditoria'
        ]
        if self.action in acciones_farmacia:
            from core.permissions import IsFarmaciaRole
            return [IsAuthenticated(), IsFarmaciaRole()]
        return [IsAuthenticated()]
    
    def get_object(self):
        """
        HALLAZGO #7 FIX: Validar acceso por centro en operaciones de escritura.
        Usuarios de centro solo pueden modificar productos si tienen permisos explícitos.
        """
        obj = super().get_object()
        
        # Solo validar en operaciones de escritura (create/update/delete)
        if self.action in ['update', 'partial_update', 'destroy']:
            user = self.request.user
            if not is_farmacia_or_admin(user) and not user.is_superuser:
                from rest_framework.exceptions import PermissionDenied
                raise PermissionDenied(
                    'Solo usuarios con rol farmacia/admin pueden modificar productos.'
                )
        
        return obj

    def get_queryset(self):
        queryset = Producto.objects.all()
        user = self.request.user
        
        # ISS-FIX: Determinar el centro para filtrar stock
        # Usuarios CENTRO solo ven productos con stock en SU centro (lo que farmacia les ha surtido)
        # Admin/Farmacia/Vista ven stock de farmacia central por defecto
        centro_param = self.request.query_params.get('centro')
        
        # ISS-FIX: Flag para determinar si el usuario es de centro
        es_usuario_centro = not is_farmacia_or_admin(user) and not user.is_superuser
        
        if es_usuario_centro:
            # Usuario de centro - filtrar stock solo por su centro
            user_centro = get_user_centro(user)
            if user_centro:
                # Anotar stock_calculado basado SOLO en lotes de su centro
                queryset = queryset.annotate(
                    stock_calculado=Coalesce(
                        Sum(
                            'lotes__cantidad_actual',
                            filter=Q(
                                lotes__activo=True,
                                lotes__cantidad_actual__gt=0,
                                lotes__centro=user_centro
                            )
                        ),
                        0
                    ),
                    # ISS-FIX: Contar lotes del centro específico del usuario
                    lotes_centro_count=Count(
                        'lotes',
                        filter=Q(
                            lotes__activo=True,
                            lotes__cantidad_actual__gt=0,
                            lotes__centro=user_centro
                        )
                    )
                )
                # ISS-FIX: Usuarios de centro SOLO ven productos con stock > 0 en su centro
                # Ya no ven todos los 76 productos con stock = 0
                queryset = queryset.filter(stock_calculado__gt=0)
            else:
                # Usuario sin centro asignado - no ve ningún producto
                queryset = queryset.annotate(
                    stock_calculado=Coalesce(Sum('lotes__cantidad_actual', filter=Q(pk__isnull=True)), 0),
                    lotes_centro_count=Count('lotes', filter=Q(pk__isnull=True))
                ).filter(stock_calculado__gt=0)  # Filtro imposible = 0 resultados
        else:
            # Admin/Farmacia/Vista - pueden ver stock global o por centro específico
            if centro_param:
                if centro_param == 'central':
                    # Solo stock de farmacia central (centro=NULL)
                    queryset = queryset.annotate(
                        stock_calculado=Coalesce(
                            Sum(
                                'lotes__cantidad_actual',
                                filter=Q(
                                    lotes__activo=True,
                                    lotes__cantidad_actual__gt=0,
                                    lotes__centro__isnull=True
                                )
                            ),
                            0
                        ),
                        lotes_centro_count=Count(
                            'lotes',
                            filter=Q(
                                lotes__activo=True,
                                lotes__cantidad_actual__gt=0,
                                lotes__centro__isnull=True
                            )
                        )
                    )
                else:
                    # Stock de un centro específico
                    queryset = queryset.annotate(
                        stock_calculado=Coalesce(
                            Sum(
                                'lotes__cantidad_actual',
                                filter=Q(
                                    lotes__activo=True,
                                    lotes__cantidad_actual__gt=0,
                                    lotes__centro_id=centro_param
                                )
                            ),
                            0
                        ),
                        lotes_centro_count=Count(
                            'lotes',
                            filter=Q(
                                lotes__activo=True,
                                lotes__cantidad_actual__gt=0,
                                lotes__centro_id=centro_param
                            )
                        )
                    )
            else:
                # Por defecto: stock de farmacia central (donde está el inventario principal)
                queryset = queryset.annotate(
                    stock_calculado=Coalesce(
                        Sum(
                            'lotes__cantidad_actual',
                            filter=Q(
                                lotes__activo=True,
                                lotes__cantidad_actual__gt=0,
                                lotes__centro__isnull=True
                            )
                        ),
                        0
                    ),
                    lotes_centro_count=Count(
                        'lotes',
                        filter=Q(
                            lotes__activo=True,
                            lotes__cantidad_actual__gt=0,
                            lotes__centro__isnull=True
                        )
                    )
                )
        
        activo = self.request.query_params.get('activo')
        if activo == 'true':
            queryset = queryset.filter(activo=True)
        elif activo == 'false':
            queryset = queryset.filter(activo=False)
        
        unidad = self.request.query_params.get('unidad_medida')
        if unidad and unidad != '':
            queryset = queryset.filter(unidad_medida=unidad)
        
        search = self.request.query_params.get('search')
        if search and search.strip():
            queryset = queryset.filter(
                Q(clave__icontains=search) | 
                Q(nombre__icontains=search)
            )

        stock_status = self.request.query_params.get('stock_status')
        if stock_status:
            # ISS-FIX: Usar stock_calculado ya anotado (respeta el centro del usuario)
            # Ya no es necesario crear stock_total_calc separado
            status_val = stock_status.lower()
            if status_val == 'sin_stock':
                queryset = queryset.filter(stock_calculado__lte=0)
            elif status_val == 'critico':
                queryset = queryset.filter(
                    stock_calculado__gt=0,
                    stock_minimo__gt=0,
                    stock_calculado__lt=F('stock_minimo') * 0.5
                )
            elif status_val == 'bajo':
                queryset = queryset.filter(
                    Q(
                        stock_minimo__gt=0,
                        stock_calculado__gte=F('stock_minimo') * 0.5,
                        stock_calculado__lt=F('stock_minimo')
                    ) | Q(
                        stock_minimo__lte=0,
                        stock_calculado__gt=0,
                        stock_calculado__lt=25
                    )
                )
            elif status_val == 'normal':
                queryset = queryset.filter(
                    Q(
                        stock_minimo__gt=0,
                        stock_calculado__gte=F('stock_minimo'),
                        stock_calculado__lte=F('stock_minimo') * 2
                    ) | Q(
                        stock_minimo__lte=0,
                        stock_calculado__gte=25,
                        stock_calculado__lt=100
                    )
                )
            elif status_val == 'alto':
                queryset = queryset.filter(
                    Q(
                        stock_minimo__gt=0,
                        stock_calculado__gt=F('stock_minimo') * 2
                    ) | Q(
                        stock_minimo__lte=0,
                        stock_calculado__gte=100
                    )
                )
        
        # PERFORMANCE: Agregar anotación de marca del lote principal
        # Subquery para obtener la marca del lote con mayor cantidad_actual
        from core.models import Lote
        lote_principal = Lote.objects.filter(
            producto=OuterRef('pk'),
            activo=True,
            cantidad_actual__gt=0
        ).order_by('-cantidad_actual').values('marca')[:1]
        
        queryset = queryset.annotate(
            marca_principal=Subquery(lote_principal)
        )
        
        # PERFORMANCE: Prefetch lotes activos para evitar N+1 en serializer
        from django.db.models import Prefetch
        queryset = queryset.prefetch_related(
            Prefetch(
                'lotes',
                queryset=Lote.objects.filter(activo=True, cantidad_actual__gt=0).select_related('centro'),
                to_attr='lotes_activos_prefetch'
            )
        )
        
        return queryset.order_by('-created_at')
    
    def create(self, request, *args, **kwargs):
        """
        Crea un nuevo producto.
        
        Validaciones:
        - Clave unica
        - Campos requeridos
        - Formato correcto
        """
        try:
            serializer = self.get_serializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            self.perform_create(serializer)
            
            headers = self.get_success_headers(serializer.data)
            return Response(
                serializer.data, 
                status=status.HTTP_201_CREATED, 
                headers=headers
            )
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            logger.error(f"Error al crear producto: {str(e)}", exc_info=True)
            return Response({'error': 'Error al crear producto. Verifique los datos ingresados.'}, status=status.HTTP_400_BAD_REQUEST)
    
    def update(self, request, *args, **kwargs):
        """
        Actualiza un producto existente.
        
        Validaciones:
        - Clave unica (si se modifica)
        - Datos validos
        """
        try:
            partial = kwargs.pop('partial', False)
            instance = self.get_object()
            serializer = self.get_serializer(instance, data=request.data, partial=partial)
            serializer.is_valid(raise_exception=True)
            self.perform_update(serializer)
            
            return Response(serializer.data)
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            logger.error(f"Error al actualizar producto: {str(e)}", exc_info=True)
            return Response({'error': 'Error al actualizar producto. Verifique los datos ingresados.'}, status=status.HTTP_400_BAD_REQUEST)
    
    def destroy(self, request, *args, **kwargs):
        """
        Elimina un producto.
        
        Validaciones:
        - No puede eliminarse si tiene lotes asociados
        - Confirmacion de eliminacion
        """
        instance = self.get_object()
        
        try:
            if instance.lotes.exists():
                return Response({
                    'error': 'No se puede eliminar el producto',
                    'razon': 'Tiene lotes asociados',
                    'sugerencia': 'Elimine primero los lotes o marque el producto como inactivo'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            instance.delete()
            return Response({'mensaje': 'Producto eliminado exitosamente'}, status=status.HTTP_204_NO_CONTENT)
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            logger.error(f"Error al eliminar producto: {str(e)}", exc_info=True)
            return Response({'error': 'Error al eliminar producto'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=True, methods=['post'], url_path='toggle-activo')
    def toggle_activo(self, request, pk=None):
        """
        Activa o desactiva un producto.
        POST /api/productos/{id}/toggle-activo/
        
        Reglas:
        - No se puede desactivar un producto con stock disponible > 0
        - Usa update() directo para evitar validacion de otros campos
        """
        try:
            producto = self.get_object()
            nuevo_estado = not producto.activo
            
            # Si se va a desactivar, verificar que no tenga stock disponible
            if not nuevo_estado:  # Desactivando
                stock_disponible = producto.lotes.filter(
                    activo=True,
                    cantidad_actual__gt=0
                ).aggregate(total=Sum('cantidad_actual'))['total'] or 0
                
                if stock_disponible > 0:
                    return Response({
                        'error': 'No se puede desactivar el producto',
                        'razon': f'Tiene {stock_disponible} unidades en stock disponible',
                        'sugerencia': 'Transfiera o agote el inventario antes de desactivar'
                    }, status=status.HTTP_400_BAD_REQUEST)
            
            # Usar update() directo para evitar validacion de otros campos
            Producto.objects.filter(pk=producto.pk).update(activo=nuevo_estado)
            
            estado = 'activado' if nuevo_estado else 'desactivado'
            return Response({
                'mensaje': f'Producto {estado} exitosamente',
                'activo': nuevo_estado,
                'id': producto.id
            }, status=status.HTTP_200_OK)
        except Producto.DoesNotExist:
            return Response({'error': 'Producto no encontrado'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"Error en toggle_activo: {str(e)}", exc_info=True)
            return Response({'error': f'Error interno: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=True, methods=['get'], url_path='auditoria')
    def auditoria(self, request, pk=None):
        """
        Obtiene el historial de cambios de un producto.
        GET /api/productos/{id}/auditoria/
        """
        try:
            from core.models import AuditoriaLogs
            producto = self.get_object()
            
            # Buscar logs de auditoria relacionados con este producto
            logs = AuditoriaLogs.objects.filter(
                Q(modelo='producto') | Q(modelo='core_producto') | Q(modelo='Producto'),
                objeto_id=str(producto.id)
            ).order_by('-timestamp')[:50]
            
            historial = []
            for log in logs:
                historial.append({
                    'id': log.id,
                    'fecha': log.timestamp.isoformat() if log.timestamp else None,
                    'usuario': log.usuario.get_full_name() or log.usuario.username if log.usuario else 'Sistema',
                    'accion': log.accion,
                    'cambios': log.datos_nuevos if log.datos_nuevos else None,
                    'ip': log.ip_address if log.ip_address else None,
                })
            
            return Response({
                'producto': {
                    'id': producto.id,
                    'clave': producto.clave,
                    'nombre': producto.nombre,
                },
                'historial': historial,
                'total': len(historial)
            })
        except Producto.DoesNotExist:
            return Response({'error': 'Producto no encontrado'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"Error en auditoria: {str(e)}", exc_info=True)
            return Response({'error': 'Error al obtener auditoría'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=True, methods=['get'], url_path='lotes')
    def lotes(self, request, pk=None):
        """
        Obtiene los lotes de un producto específico con información de caducidad.
        GET /api/productos/{id}/lotes/
        
        ISS-FIX (lotes-centro): Mejorado para diagnóstico de problemas de lotes por centro.
        
        Retorna:
        - Lista de lotes con cantidad, número de lote, fecha de caducidad y semáforo de alerta
        """
        from datetime import date, timedelta
        try:
            producto = self.get_object()
            user = request.user
            
            # ISS-FIX: Logging para diagnóstico
            logger.info(
                f"ISS-LOTES-PROD: Consultando lotes de producto {producto.clave} (ID: {producto.pk}). "
                f"Usuario: {user.username}, Rol: {getattr(user, 'rol', 'N/A')}"
            )
            
            # Obtener lotes del producto
            lotes_queryset = producto.lotes.filter(
                activo=True,
                cantidad_actual__gt=0
            )
            
            # ISS-FIX: Si es usuario de centro, filtrar solo lotes de su centro
            if not is_farmacia_or_admin(user) and not user.is_superuser:
                user_centro = get_user_centro(user)
                if user_centro:
                    # ISS-FIX: Log del filtro aplicado
                    lotes_antes = lotes_queryset.count()
                    lotes_queryset = lotes_queryset.filter(centro=user_centro)
                    lotes_despues = lotes_queryset.count()
                    logger.info(
                        f"ISS-LOTES-PROD: Filtrado por centro {user_centro.nombre} (ID: {user_centro.pk}). "
                        f"Lotes antes: {lotes_antes}, después: {lotes_despues}"
                    )
                else:
                    logger.warning(
                        f"ISS-LOTES-PROD: Usuario {user.username} sin centro asignado. "
                        f"Retornando 0 lotes."
                    )
                    lotes_queryset = lotes_queryset.none()
            else:
                # ISS-FIX: Admin/farmacia ven todos los lotes
                logger.info(
                    f"ISS-LOTES-PROD: Usuario admin/farmacia, mostrando todos los lotes. "
                    f"Total: {lotes_queryset.count()}"
                )
            
            # Ordenar por fecha de caducidad (más próximos a vencer primero)
            lotes_queryset = lotes_queryset.order_by('fecha_caducidad')
            
            hoy = date.today()
            lotes_data = []
            
            for lote in lotes_queryset:
                dias_para_caducar = (lote.fecha_caducidad - hoy).days if lote.fecha_caducidad else None
                
                # Determinar semáforo de caducidad
                if dias_para_caducar is None:
                    alerta_caducidad = 'sin_fecha'
                elif dias_para_caducar < 0:
                    alerta_caducidad = 'vencido'
                elif dias_para_caducar <= 30:
                    alerta_caducidad = 'critico'
                elif dias_para_caducar <= 90:
                    alerta_caducidad = 'proximo'
                else:
                    alerta_caducidad = 'normal'
                
                lotes_data.append({
                    'id': lote.id,
                    'numero_lote': lote.numero_lote,
                    'cantidad_actual': lote.cantidad_actual,
                    'fecha_caducidad': lote.fecha_caducidad.isoformat() if lote.fecha_caducidad else None,
                    'dias_para_caducar': dias_para_caducar,
                    'alerta_caducidad': alerta_caducidad,
                    'centro_id': lote.centro_id,
                    'centro_nombre': lote.centro.nombre if lote.centro else 'Almacén Central',
                })
                
                # ISS-FIX: Log detallado de cada lote
                logger.debug(
                    f"ISS-LOTES-PROD: Lote ID={lote.id}, NumLote={lote.numero_lote}, "
                    f"Cantidad={lote.cantidad_actual}, Centro={lote.centro_id}"
                )
            
            # ISS-FIX: Log del resultado final
            logger.info(
                f"ISS-LOTES-PROD: Retornando {len(lotes_data)} lotes para producto {producto.clave}. "
                f"Stock total: {sum(l['cantidad_actual'] for l in lotes_data)}"
            )
            
            return Response({
                'producto': {
                    'id': producto.id,
                    'clave': producto.clave,
                    'nombre': producto.nombre,
                    'unidad_medida': producto.unidad_medida,
                },
                'lotes': lotes_data,
                'total_lotes': len(lotes_data),
                'total_stock': sum(l['cantidad_actual'] for l in lotes_data),
            })
        except Producto.DoesNotExist:
            return Response({'error': 'Producto no encontrado'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"Error en lotes: {str(e)}", exc_info=True)
            return Response({'error': 'Error al obtener lotes'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=True, methods=['get'], url_path='lotes-diagnostico')
    def lotes_diagnostico(self, request, pk=None):
        """
        ISS-FIX (lotes-centro): Endpoint de diagnóstico para problemas de lotes.
        GET /api/productos/{id}/lotes-diagnostico/
        
        Solo accesible para admin/farmacia. Muestra TODOS los lotes del producto
        sin filtros de centro para diagnóstico.
        """
        from datetime import date
        
        # Solo admin/farmacia puede usar este endpoint
        if not is_farmacia_or_admin(request.user) and not request.user.is_superuser:
            return Response({
                'error': 'Este endpoint solo está disponible para administradores'
            }, status=status.HTTP_403_FORBIDDEN)
        
        try:
            producto = self.get_object()
            
            # Obtener TODOS los lotes del producto (incluyendo inactivos y sin stock)
            todos_lotes = producto.lotes.all().order_by('centro_id', 'numero_lote')
            
            hoy = date.today()
            lotes_data = []
            
            for lote in todos_lotes:
                lotes_data.append({
                    'id': lote.id,
                    'numero_lote': lote.numero_lote,
                    'cantidad_inicial': lote.cantidad_inicial,
                    'cantidad_actual': lote.cantidad_actual,
                    'fecha_caducidad': lote.fecha_caducidad.isoformat() if lote.fecha_caducidad else None,
                    'activo': lote.activo,
                    'centro_id': lote.centro_id,
                    'centro_nombre': lote.centro.nombre if lote.centro else 'Almacén Central (NULL)',
                    'created_at': lote.created_at.isoformat() if lote.created_at else None,
                    'updated_at': lote.updated_at.isoformat() if lote.updated_at else None,
                })
            
            # Agrupar por centro
            lotes_por_centro = {}
            for lote in lotes_data:
                centro_key = lote['centro_nombre']
                if centro_key not in lotes_por_centro:
                    lotes_por_centro[centro_key] = {
                        'centro_id': lote['centro_id'],
                        'lotes': [],
                        'total_stock': 0,
                        'total_lotes': 0,
                    }
                lotes_por_centro[centro_key]['lotes'].append(lote)
                lotes_por_centro[centro_key]['total_stock'] += lote['cantidad_actual']
                lotes_por_centro[centro_key]['total_lotes'] += 1
            
            return Response({
                'producto': {
                    'id': producto.id,
                    'clave': producto.clave,
                    'nombre': producto.nombre,
                },
                'diagnostico': {
                    'total_lotes_bd': len(lotes_data),
                    'total_stock_global': sum(l['cantidad_actual'] for l in lotes_data),
                    'lotes_activos': sum(1 for l in lotes_data if l['activo']),
                    'lotes_con_stock': sum(1 for l in lotes_data if l['cantidad_actual'] > 0),
                },
                'lotes_por_centro': lotes_por_centro,
                'todos_los_lotes': lotes_data,
            })
            
        except Producto.DoesNotExist:
            return Response({'error': 'Producto no encontrado'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"Error en lotes_diagnostico: {str(e)}", exc_info=True)
            return Response({'error': f'Error: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path='exportar-excel')
    def exportar_excel(self, request):
        """
        Exporta todos los productos a un archivo Excel.
        
        Columnas alineadas con schema real de productos:
        - #, Clave, Nombre, Nombre Comercial, Categoria, Unidad Medida, Stock Minimo, Stock Actual,
          Sustancia Activa, Presentacion, Concentracion, Via Admin, Requiere Receta, Controlado, 
          Lotes Activos, Estado
        """
        try:
            productos = self.get_queryset()
            
            # Crear libro de Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Productos'
            
            # Encabezados alineados con schema de Supabase (incluye nombre_comercial)
            headers = ['#', 'Clave', 'Nombre', 'Nombre Comercial', 'Categoria', 'Unidad Medida', 
                       'Stock Minimo', 'Stock Actual', 'Sustancia Activa', 'Presentacion',
                       'Concentracion', 'Via Admin', 'Requiere Receta', 'Controlado', 
                       'Lotes Activos', 'Estado']
            ws.append(headers)
            
            # Estilo de encabezados
            header_fill = PatternFill(start_color='632842', end_color='632842', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=12)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Datos de productos
            for idx, producto in enumerate(productos, start=1):
                stock_actual = producto.lotes.filter(activo=True).aggregate(total=Sum('cantidad_actual'))['total'] or 0
                lotes_activos = producto.lotes.filter(activo=True, cantidad_actual__gt=0).count()
                
                ws.append([
                    idx,
                    producto.clave or '',
                    producto.nombre,
                    producto.nombre_comercial or '',
                    producto.categoria or '',
                    producto.unidad_medida,
                    producto.stock_minimo,
                    stock_actual,
                    producto.sustancia_activa or '',
                    producto.presentacion or '',
                    producto.concentracion or '',
                    producto.via_administracion or '',
                    'Sí' if producto.requiere_receta else 'No',
                    'Sí' if producto.es_controlado else 'No',
                    lotes_activos,
                    'Activo' if producto.activo else 'Inactivo'
                ])
                
                # Colorear fila si el stock esta por debajo del minimo
                if stock_actual < producto.stock_minimo:
                    for col in range(1, 17):
                        ws.cell(row=idx+1, column=col).fill = PatternFill(
                            start_color='FFF4E6', 
                            end_color='FFF4E6', 
                            fill_type='solid'
                        )
            
            # Ajustar anchos de columna
            ws.column_dimensions['A'].width = 6    # #
            ws.column_dimensions['B'].width = 18   # Clave
            ws.column_dimensions['C'].width = 40   # Nombre
            ws.column_dimensions['D'].width = 25   # Nombre Comercial
            ws.column_dimensions['E'].width = 18   # Categoria
            ws.column_dimensions['F'].width = 18   # Unidad Medida
            ws.column_dimensions['G'].width = 12   # Stock Minimo
            ws.column_dimensions['H'].width = 12   # Stock Actual
            ws.column_dimensions['I'].width = 20   # Sustancia Activa
            ws.column_dimensions['J'].width = 18   # Presentacion
            ws.column_dimensions['K'].width = 14   # Concentracion
            ws.column_dimensions['L'].width = 14   # Via Admin
            ws.column_dimensions['M'].width = 14   # Requiere Receta
            ws.column_dimensions['N'].width = 12   # Controlado
            ws.column_dimensions['O'].width = 12   # Lotes Activos
            ws.column_dimensions['P'].width = 10   # Estado
            
            # Generar respuesta
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename=Productos_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            wb.save(response)
            
            return response
            
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            return Response({
                'error': 'Error al exportar productos',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['post'], url_path='importar-excel')
    def importar_excel(self, request):
        """
        Importa productos desde archivo Excel con detección automática de columnas.
        
        Columnas reconocidas (flexibles, no importa el orden):
        - Clave (requerido): código único del producto
        - Nombre (requerido): nombre del producto
        - Unidad (requerido): unidad de medida (CAJA, PIEZA, etc.)
        - Stock Minimo (opcional): cantidad mínima en inventario
        - Categoria (opcional): medicamento, material_curacion, insumo
        - Sustancia Activa, Presentacion, Concentracion, Via Admin (opcionales)
        - Requiere Receta, Controlado (opcionales): Si/No
        - Estado/Activo (opcional): Activo/Inactivo
        
        Limites de seguridad:
        - Tamano maximo: 10MB
        - Filas maximas: 5000
        - Extensiones: .xlsx, .xls
        """
        file = request.FILES.get('file')
        
        # HALLAZGO #13 FIX: Validar archivo CON magic bytes
        es_valido, error_msg = validar_archivo_excel(file)
        if not es_valido:
            logger.warning(f"HALLAZGO #13: Archivo Excel rechazado en importar_excel: {error_msg}")
            return Response({
                'error': 'Archivo invalido',
                'mensaje': error_msg
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            # ISS-001: Usar carga segura con limite real de bytes
            wb, error_carga = cargar_workbook_seguro(file)
            if wb is None:
                return Response({
                    'error': 'Error al procesar archivo',
                    'mensaje': error_carga
                }, status=status.HTTP_400_BAD_REQUEST)
            
            ws = wb.active
            
            # Validar numero de filas
            filas_validas, error_filas, num_filas = validar_filas_excel(ws)
            if not filas_validas:
                wb.close()
                return Response({
                    'error': 'Archivo demasiado grande',
                    'mensaje': error_filas
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # ISS-FIX: Detectar fila de encabezados automáticamente
            COLUMN_ALIASES = {
                'clave': ['clave', 'codigo', 'código', 'clave producto', 'cod', 'id producto'],
                'nombre': ['nombre', 'nombre producto', 'descripcion', 'descripción', 'producto'],
                'nombre_comercial': ['nombre comercial', 'comercial', 'marca', 'brand'],
                'unidad_medida': ['unidad', 'unidad medida', 'um', 'unidad de medida', 'medida'],
                'stock_minimo': ['stock minimo', 'stock mínimo', 'stock min', 'minimo', 'mínimo'],
                'categoria': ['categoria', 'categoría', 'tipo', 'clasificacion'],
                'sustancia_activa': ['sustancia activa', 'sustancia', 'principio activo', 'activo'],
                'presentacion': ['presentacion', 'presentación', 'forma', 'forma farmaceutica'],
                'concentracion': ['concentracion', 'concentración', 'dosis'],
                'via_administracion': ['via admin', 'via administracion', 'vía admin', 'administracion', 'via'],
                'requiere_receta': ['requiere receta', 'receta', 'req receta'],
                'es_controlado': ['controlado', 'es controlado', 'control'],
                'estado': ['estado', 'activo', 'status'],
            }
            
            def normalizar_header(val):
                if not val:
                    return ''
                return str(val).lower().strip().replace('_', ' ').replace('-', ' ')
            
            # Buscar fila con encabezados (primeras 5 filas)
            header_row_idx = 1
            col_map = {}
            
            for row_num in range(1, min(6, ws.max_row + 1)):
                row_values = [cell.value for cell in ws[row_num]]
                headers_encontrados = 0
                temp_map = {}
                
                for col_idx, val in enumerate(row_values):
                    header_norm = normalizar_header(val)
                    if not header_norm:
                        continue
                    
                    for field, aliases in COLUMN_ALIASES.items():
                        if header_norm in aliases or any(alias in header_norm for alias in aliases):
                            if field not in temp_map:
                                temp_map[field] = col_idx
                                headers_encontrados += 1
                            break
                
                # Si encontramos al menos 2 columnas clave (clave, nombre), es la fila de headers
                if headers_encontrados >= 2 and ('clave' in temp_map or 'nombre' in temp_map):
                    header_row_idx = row_num
                    col_map = temp_map
                    break
            
            # Si no hay mapa, usar orden por defecto (con nombre_comercial)
            if not col_map:
                col_map = {
                    'clave': 0, 'nombre': 1, 'nombre_comercial': 2, 'unidad_medida': 3, 
                    'stock_minimo': 4, 'categoria': 5, 'sustancia_activa': 6, 'presentacion': 7,
                    'concentracion': 8, 'via_administracion': 9, 'requiere_receta': 10,
                    'es_controlado': 11, 'estado': 12
                }
            
            def get_val(row, field, default=None):
                if field not in col_map:
                    return default
                idx = col_map[field]
                if idx < len(row):
                    val = row[idx]
                    return val if val not in [None, '', 'None'] else default
                return default
            
            creados = 0
            actualizados = 0
            errores = []
            exitos = []
            
            # ISS-019: Envolver en transacción atómica
            with transaction.atomic():
                for row_idx, row in enumerate(ws.iter_rows(min_row=header_row_idx + 1, values_only=True), start=header_row_idx + 1):
                    try:
                        # Saltar filas vacías
                        if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                            continue
                        
                        # Extraer valores usando el mapa de columnas
                        clave = get_val(row, 'clave')
                        nombre = get_val(row, 'nombre')
                        nombre_comercial = get_val(row, 'nombre_comercial')
                        unidad_medida = get_val(row, 'unidad_medida')
                        stock_minimo = get_val(row, 'stock_minimo')
                        categoria = get_val(row, 'categoria')
                        sustancia_activa = get_val(row, 'sustancia_activa')
                        presentacion = get_val(row, 'presentacion')
                        concentracion = get_val(row, 'concentracion')
                        via_administracion = get_val(row, 'via_administracion')
                        requiere_receta = get_val(row, 'requiere_receta')
                        es_controlado = get_val(row, 'es_controlado')
                        estado = get_val(row, 'estado')
                        
                        # Validar campos requeridos
                        if not clave:
                            errores.append({'fila': row_idx, 'error': 'Clave es obligatoria'})
                            continue
                        if not nombre:
                            errores.append({'fila': row_idx, 'error': 'Nombre es obligatorio'})
                            continue
                        
                        # ISS-FIX: Permitir texto libre en unidad de medida para mejor manejo de farmacia
                        # Ejemplos: "CAJA CON 7 OVULOS", "GOTERO CON 15 MILILITROS"
                        unidad_limpia = str(unidad_medida).strip().upper() if unidad_medida else 'PIEZA'

                        try:
                            stock_min = int(float(stock_minimo)) if stock_minimo not in [None, ''] else 10
                            if stock_min < 0:
                                stock_min = 0
                        except Exception:
                            stock_min = 10  # Default

                        # Parsear campos booleanos
                        def parse_bool(val):
                            if val is None:
                                return False
                            return str(val).lower() in ['sí', 'si', 'true', '1', 'yes', 's', 'x', 'activo']

                        # Validar y normalizar categoría
                        from core.constants import CATEGORIAS_VALIDAS
                        categoria_limpia = str(categoria).strip().lower().replace(' ', '_') if categoria else 'medicamento'
                        if categoria_limpia not in CATEGORIAS_VALIDAS:
                            categoria_limpia = 'medicamento'

                        # Preparar datos
                        datos = {
                            'nombre': str(nombre).strip()[:500],
                            'nombre_comercial': str(nombre_comercial).strip()[:200] if nombre_comercial else '',
                            'unidad_medida': unidad_limpia,
                            'stock_minimo': stock_min,
                            'categoria': categoria_limpia,
                            'sustancia_activa': str(sustancia_activa).strip()[:200] if sustancia_activa else '',
                            'presentacion': str(presentacion).strip()[:200] if presentacion else '',
                            'concentracion': str(concentracion).strip()[:100] if concentracion else '',
                            'via_administracion': str(via_administracion).strip()[:50] if via_administracion else '',
                            'requiere_receta': parse_bool(requiere_receta),
                            'es_controlado': parse_bool(es_controlado),
                            'activo': str(estado).lower() in ['activo', 'sí', 'si', 'true', '1', 'yes', 's'] if estado else True
                        }
                        
                        # Ya no es necesario guardar en presentación porque unidad_medida acepta texto completo
                        
                        # Crear o actualizar producto
                        clave_limpia = str(clave).strip()[:50].upper()
                        
                        producto, created = Producto.objects.update_or_create(
                            clave=clave_limpia,
                            defaults=datos
                        )
                        
                        if created:
                            creados += 1
                        else:
                            actualizados += 1
                        exitos.append({'fila': row_idx, 'producto_id': producto.id, 'clave': producto.clave})
                            
                    except Exception as e:
                        errores.append({'fila': row_idx, 'error': str(e)})
            
            return Response({
                'mensaje': 'Importacion completada',
                'resumen': {
                    'creados': creados,
                    'actualizados': actualizados,
                    'total_procesados': creados + actualizados,
                    'total_errores': len(errores)
                },
                'exitos': exitos,
                'errores': errores,
                'exito': len(errores) == 0
            }, status=status.HTTP_200_OK if len(errores) == 0 else status.HTTP_207_MULTI_STATUS)
            
        except Exception as e:
            return Response({
                'error': 'Error al procesar archivo',
                'mensaje': str(e),
                'sugerencia': 'Verifique que el archivo tenga formato Excel válido con columnas: Clave, Nombre, Unidad, etc.'
            }, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get'], url_path='plantilla')
    def plantilla_productos(self, request):
        """
        Descarga plantilla Excel para importación de productos.
        
        Columnas:
        - Clave (REQUERIDO, único) - Código identificador del producto
        - Nombre (REQUERIDO) - Nombre del medicamento o insumo
        - Nombre Comercial (opcional) - Nombre comercial o marca
        - Unidad Medida (opcional) - Unidad de medida (PIEZA, CAJA, FRASCO, SOBRE, etc.)
        - Stock Minimo (opcional, default: 10) - Cantidad mínima de alerta
        - Categoria (opcional) - medicamento, material_curacion, insumo, equipo, otro
        - Sustancia Activa (opcional) - Principio activo
        - Presentacion (opcional) - Forma farmacéutica
        - Concentracion (opcional) - Dosis del principio activo
        - Via Admin (opcional) - Vía de administración
        - Requiere Receta (opcional) - Sí/No
        - Controlado (opcional) - Sí/No
        - Estado (opcional, default: Activo) - Activo/Inactivo
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Productos'
        
        # Headers que coinciden con importar_excel (con nombre_comercial)
        headers = [
            'Clave', 'Nombre', 'Nombre Comercial', 'Unidad Medida', 'Stock Minimo', 'Categoria',
            'Sustancia Activa', 'Presentacion', 'Concentracion', 
            'Via Admin', 'Requiere Receta', 'Controlado', 'Estado'
        ]
        ws.append(headers)
        
        # ============================================================
        # FILAS DE EJEMPLO - ELIMINAR ANTES DE USAR CON DATOS REALES
        # Estas filas son solo para mostrar el formato correcto.
        # ============================================================
        ws.append([
            'PRUEBA001', '[EJEMPLO] Paracetamol 500mg - ELIMINAR', 'Tempra', 'CAJA', 50, 'medicamento',
            'Paracetamol', 'Tableta', '500 mg',
            'oral', 'No', 'No', 'Activo'
        ])
        ws.append([
            'PRUEBA002', '[EJEMPLO] Ibuprofeno 400mg - ELIMINAR', 'Advil', 'FRASCO', 30, 'medicamento',
            'Ibuprofeno', 'Cápsula', '400 mg',
            'oral', 'No', 'No', 'Activo'
        ])
        ws.append([
            'PRUEBA003', '[EJEMPLO] Jeringa 10ml - ELIMINAR', '', 'PIEZA', 100, 'material_curacion',
            '', '', '',
            '', 'No', 'No', 'Activo'
        ])
        
        # Aplicar formato a headers
        from openpyxl.styles import Font, PatternFill, Alignment
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='9F2241', end_color='9F2241', fill_type='solid')
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
        
        # Estilo para filas de ejemplo (gris, itálica - sin fondo de color)
        example_font = Font(italic=True, color='888888')
        for row_num in range(2, 5):  # Filas 2, 3, 4 (ejemplos)
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col)
                cell.font = example_font
        
        # Ajustar ancho de columnas
        column_widths = {
            'A': 15,  # Clave
            'B': 45,  # Nombre (más ancho para ver el texto de ejemplo)
            'C': 20,  # Nombre Comercial
            'D': 14,  # Unidad Medida
            'E': 14,  # Stock Minimo
            'F': 18,  # Categoria
            'G': 20,  # Sustancia Activa
            'H': 15,  # Presentacion
            'I': 15,  # Concentracion
            'J': 12,  # Via Admin
            'K': 15,  # Requiere Receta
            'L': 12,  # Controlado
            'M': 10,  # Estado
        }
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # ============================================================
        # HOJA DE INSTRUCCIONES
        # ============================================================
        ws_instrucciones = wb.create_sheet(title='INSTRUCCIONES')
        
        instrucciones = [
            ['╔════════════════════════════════════════════════════════════════════╗'],
            ['║    INSTRUCCIONES PARA IMPORTACIÓN DE PRODUCTOS                    ║'],
            ['╚════════════════════════════════════════════════════════════════════╝'],
            [''],
            ['⚠️  IMPORTANTE: Las filas amarillas en la hoja "Productos" son EJEMPLOS.'],
            ['    ELIMÍNELAS antes de cargar sus datos reales.'],
            [''],
            ['────────────────────────────────────────────────────────────────────────'],
            ['COLUMNAS REQUERIDAS (obligatorias):'],
            ['────────────────────────────────────────────────────────────────────────'],
            ['• Clave      - Código único del producto (ej: 001, MED001, ABC123)'],
            ['• Nombre     - Nombre completo del producto'],
            [''],
            ['────────────────────────────────────────────────────────────────────────'],
            ['COLUMNAS OPCIONALES:'],
            ['────────────────────────────────────────────────────────────────────────'],
            ['• Nombre Comercial - Nombre comercial o marca (ej: Tempra, Advil)'],
            ['• Unidad Medida    - CAJA, PIEZA, FRASCO, SOBRE, TABLETA, etc. (default: PIEZA)'],
            ['• Stock Minimo     - Cantidad mínima para alertas (default: 10)'],
            ['• Categoria        - medicamento, material_curacion, insumo (default: medicamento)'],
            ['• Sustancia Activa - Principio activo del medicamento'],
            ['• Presentacion     - Forma farmacéutica (tableta, cápsula, jarabe, etc.)'],
            ['• Concentracion    - Dosis (ej: 500 mg, 10 ml)'],
            ['• Via Admin        - oral, intravenosa, tópica, etc.'],
            ['• Requiere Receta  - Sí / No (default: No)'],
            ['• Controlado       - Sí / No (default: No)'],
            ['• Estado           - Activo / Inactivo (default: Activo)'],
            [''],
            ['────────────────────────────────────────────────────────────────────────'],
            ['NOTAS:'],
            ['────────────────────────────────────────────────────────────────────────'],
            ['• Si la CLAVE ya existe, el producto se ACTUALIZARÁ con los nuevos datos.'],
            ['• Si la CLAVE no existe, se CREARÁ un nuevo producto.'],
            ['• Máximo 5000 productos por archivo.'],
            ['• Tamaño máximo de archivo: 10 MB.'],
            ['• Formatos aceptados: .xlsx, .xls'],
            [''],
        ]
        
        for row in instrucciones:
            ws_instrucciones.append(row)
        
        # Formato de instrucciones
        ws_instrucciones.column_dimensions['A'].width = 80
        for row in ws_instrucciones.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=False)
        
        # Poner hoja de Productos como activa
        wb.active = ws
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=Plantilla_Productos.xlsx'
        wb.save(response)
        return response


class CentroViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestionar centros penitenciarios.
    
    Funcionalidades:
    - CRUD completo
    - Busqueda por clave, nombre y direccion
    - Filtrado por estado activo/inactivo
    - Exportar a Excel con formato profesional
    - Importar desde Excel con validaciones
    - Obtener requisiciones por centro
    
    ISS-MEDICO FIX: Roles de centro (medico, etc.) reciben 403 en GET.
    Solo admin/farmacia/vista pueden ver el catálogo de centros.
    """
    queryset = Centro.objects.all()
    serializer_class = CentroSerializer
    # ISS-MEDICO FIX: Bloquear acceso a roles de centro (incluyendo médico)
    permission_classes = [IsFarmaciaAdminOrVistaReadOnly]
    pagination_class = CustomPagination

    def _user_centro(self, user):
        return getattr(user, 'centro', None) or getattr(getattr(user, 'profile', None), 'centro', None)

    def get_queryset(self):
        """Filtra centros segun parametros"""
        queryset = Centro.objects.all()
        user = getattr(self.request, 'user', None)
        
        # Admin, Farmacia y Superusuarios pueden ver todos los centros
        # Otros usuarios solo ven su propio centro
        if user and not user.is_superuser and not is_farmacia_or_admin(user):
            user_centro = self._user_centro(user)
            if user_centro:
                queryset = queryset.filter(id=user_centro.id)
            else:
                return Centro.objects.none()
        
        # Filtro por busqueda
        search = self.request.query_params.get('search')
        if search and search.strip():
            queryset = queryset.filter(
                Q(nombre__icontains=search) | 
                Q(direccion__icontains=search) |
                Q(email__icontains=search) |
                Q(telefono__icontains=search)
            )
        
        # Filtro por estado activo
        activo = self.request.query_params.get('activo')
        if activo == 'true':
            queryset = queryset.filter(activo=True)
        elif activo == 'false':
            queryset = queryset.filter(activo=False)
        
        # Ordenamiento (respeta parámetro del frontend)
        ordering = self.request.query_params.get('ordering', '-created_at')
        valid_orderings = ['nombre', '-nombre', 'created_at', '-created_at', 'activo', '-activo']
        if ordering in valid_orderings:
            queryset = queryset.order_by(ordering)
        else:
            queryset = queryset.order_by('-created_at')
        
        # PERFORMANCE: Agregar anotaciones para evitar N+1 queries en serializer
        from core.models import Requisicion
        from django.contrib.auth import get_user_model
        User = get_user_model()
        
        # Contar requisiciones donde el centro es origen O destino
        # Nota: Los related_name correctos son 'requisiciones_origen' y 'requisiciones_destino'
        queryset = queryset.annotate(
            requisiciones_count=Count(
                'requisiciones_origen',
                distinct=True
            ) + Count(
                'requisiciones_destino',
                distinct=True
            ),
            usuarios_count=Count(
                'usuarios',
                filter=Q(usuarios__is_active=True),
                distinct=True
            )
        )
        
        return queryset
    
    def create(self, request, *args, **kwargs):
        """Crea un nuevo centro"""
        try:
            logger.debug(f"CREAR CENTRO - Body: {request.data}")
            
            serializer = self.get_serializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            centro = serializer.save()
            
            logger.info(f"Centro creado: {centro.clave} - {centro.nombre}")
            
            return Response({
                'mensaje': 'Centro creado exitosamente',
                'centro': CentroSerializer(centro).data
            }, status=status.HTTP_201_CREATED)
            
        except serializers.ValidationError as e:
            logger.warning(f"Error de validación al crear centro: {e.detail}")
            return Response({
                'error': 'Error de validacion',
                'detalles': e.detail
            }, status=status.HTTP_400_BAD_REQUEST)
            
        except Exception as e:
            logger.exception(f"Error inesperado al crear centro: {e}")
            # traceback removido por seguridad (ISS-008)
            return Response({
                'error': 'Error al crear centro',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def update(self, request, *args, **kwargs):
        """Actualiza un centro existente"""
        try:
            partial = kwargs.pop('partial', False)
            instance = self.get_object()
            serializer = self.get_serializer(instance, data=request.data, partial=partial)
            serializer.is_valid(raise_exception=True)
            self.perform_update(serializer)
            
            return Response({
                'mensaje': 'Centro actualizado exitosamente',
                'centro': serializer.data
            })
        except serializers.ValidationError as e:
            return Response(
                {'error': 'Error de validacion', 'detalles': e.detail},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            return Response(
                {'error': 'Error al actualizar centro', 'mensaje': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def destroy(self, request, *args, **kwargs):
        """
        Elimina un centro.
        
        Validaciones:
        - No puede eliminarse si tiene requisiciones asociadas (como origen o destino)
        - No puede eliminarse si tiene usuarios asignados
        """
        instance = self.get_object()
        
        try:
            # Verificar requisiciones (como origen O destino)
            has_req_origen = hasattr(instance, 'requisiciones_origen') and instance.requisiciones_origen.exists()
            has_req_destino = hasattr(instance, 'requisiciones_destino') and instance.requisiciones_destino.exists()
            
            if has_req_origen or has_req_destino:
                total_origen = instance.requisiciones_origen.count() if has_req_origen else 0
                total_destino = instance.requisiciones_destino.count() if has_req_destino else 0
                total_requisiciones = total_origen + total_destino
                
                # Contar requisiciones activas
                activas_origen = instance.requisiciones_origen.exclude(
                    estado__in=['CANCELADA', 'SURTIDA']
                ).count() if has_req_origen else 0
                activas_destino = instance.requisiciones_destino.exclude(
                    estado__in=['CANCELADA', 'SURTIDA']
                ).count() if has_req_destino else 0
                requisiciones_activas = activas_origen + activas_destino
                
                return Response({
                    'error': 'No se puede eliminar el centro',
                    'razon': 'Tiene requisiciones asociadas',
                    'total_requisiciones': total_requisiciones,
                    'como_origen': total_origen,
                    'como_destino': total_destino,
                    'requisiciones_activas': requisiciones_activas,
                    'sugerencia': 'Marque el centro como inactivo en lugar de eliminarlo'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Verificar usuarios asignados (solo activos, consistente con serializer)
            if hasattr(instance, 'usuarios'):
                usuarios_activos = instance.usuarios.filter(is_active=True)
                if usuarios_activos.exists():
                    total_usuarios = usuarios_activos.count()
                    
                    return Response({
                        'error': 'No se puede eliminar el centro',
                        'razon': 'Tiene usuarios activos asignados',
                        'total_usuarios': total_usuarios,
                        'sugerencia': 'Reasigne los usuarios a otro centro o marque el centro como inactivo'
                    }, status=status.HTTP_400_BAD_REQUEST)
            
            # Verificar lotes con stock activo
            if hasattr(instance, 'lotes'):
                lotes_con_stock = instance.lotes.filter(activo=True, cantidad_actual__gt=0).count()
                if lotes_con_stock > 0:
                    return Response({
                        'error': 'No se puede eliminar el centro',
                        'razon': 'Tiene lotes con stock disponible',
                        'lotes_con_stock': lotes_con_stock,
                        'sugerencia': 'Transfiera el inventario a otro centro o marque el centro como inactivo'
                    }, status=status.HTTP_400_BAD_REQUEST)
            
            # Si no tiene relaciones, se puede eliminar
            clave_eliminada = instance.clave
            nombre_eliminado = instance.nombre
            instance.delete()
            
            return Response({
                'mensaje': 'Centro eliminado exitosamente',
                'centro_eliminado': f"{clave_eliminada} - {nombre_eliminado}"
            }, status=status.HTTP_204_NO_CONTENT)
            
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            return Response({
                'error': 'Error al eliminar centro',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['post'], url_path='toggle-activo')
    def toggle_activo(self, request, pk=None):
        """
        Activa o desactiva un centro.
        POST /api/centros/{id}/toggle-activo/
        
        Usa update() directo para evitar validacion de otros campos.
        """
        try:
            centro = self.get_object()
            nuevo_estado = not centro.activo
            
            # Usar update() directo para evitar validacion de otros campos
            Centro.objects.filter(pk=centro.pk).update(activo=nuevo_estado)
            
            estado = 'activado' if nuevo_estado else 'desactivado'
            return Response({
                'mensaje': f'Centro {estado} exitosamente',
                'activo': nuevo_estado,
                'id': centro.id,
                'clave': centro.clave,
                'nombre': centro.nombre
            }, status=status.HTTP_200_OK)
        except Centro.DoesNotExist:
            return Response({'error': 'Centro no encontrado'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"Error en toggle_activo centro: {str(e)}", exc_info=True)
            return Response({'error': f'Error interno: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['get'])
    def inventario(self, request, pk=None):
        """Devuelve inventario resumido del centro a partir de lotes asociados a movimientos del centro."""
        centro = self.get_object()
        user_centro = self._user_centro(request.user)
        if not request.user.is_superuser:
            if not user_centro or user_centro.id != centro.id:
                return Response({'error': 'Solo puedes ver inventario de tu centro'}, status=status.HTTP_403_FORBIDDEN)

        # Lotes que han tenido movimientos en este centro
        lote_ids = Movimiento.objects.filter(
            Q(centro_origen=centro) | Q(centro_destino=centro)
        ).values_list('lote_id', flat=True)
        lotes = Lote.objects.filter(
            Q(id__in=lote_ids) | Q(centro=centro),
            activo=True,
            cantidad_actual__gt=0
        ).select_related('producto')

        inventario_dict = {}
        for lote in lotes:
            prod = lote.producto
            item = inventario_dict.setdefault(prod.id, {
                'producto_id': prod.id,
                'clave': prod.clave,
                'producto': prod.nombre,
                'cantidad_disponible': 0,
                'lote_proximo_caducar': None,
                'fecha_caducidad': None,
            })
            item['cantidad_disponible'] += lote.cantidad_actual
            if lote.fecha_caducidad:
                fecha_actual = item['fecha_caducidad']
                if fecha_actual is None or lote.fecha_caducidad < fecha_actual:
                    item['lote_proximo_caducar'] = lote.numero_lote
                    item['fecha_caducidad'] = lote.fecha_caducidad

        # Si no hay lotes asociados, caer al agregado por movimientos para no dejar vacio
        inventario = list(inventario_dict.values())
        if not inventario:
            movimientos = Movimiento.objects.filter(
                Q(centro_origen=centro) | Q(centro_destino=centro)
            )
            agregados = movimientos.values('lote__producto').annotate(cantidad=Coalesce(Sum('cantidad'), 0))
            for item in agregados:
                producto = Producto.objects.filter(id=item['lote__producto']).first()
                if not producto:
                    continue
                inventario.append({
                    'producto_id': producto.id,
                    'clave': producto.clave,
                    'producto': producto.nombre,
                    'cantidad_disponible': max(0, item['cantidad']),
                    'lote_proximo_caducar': None,
                    'fecha_caducidad': None,
                })

        return Response({
            'centro': centro.nombre,
            'centro_id': centro.id,
            'total_productos': len(inventario),
            'inventario': inventario
        })
    
    @action(detail=False, methods=['get'], url_path='exportar-excel')
    def exportar_excel(self, request):
        """
        Exporta todos los centros a Excel con formato profesional.
        
        Columnas (alineadas con BD):
        - #, Nombre, Direccion, Telefono, Email, Total Requisiciones, Estado
        """
        try:
            centros = self.get_queryset()
            
            # Crear libro de Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Centros Penitenciarios'
            
            # Titulo del reporte
            ws.merge_cells('A1:G1')
            titulo_cell = ws['A1']
            titulo_cell.value = 'REPORTE DE CENTROS PENITENCIARIOS'
            titulo_cell.font = Font(bold=True, size=14, color='632842')
            titulo_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Fecha de generacion
            ws.merge_cells('A2:G2')
            fecha_cell = ws['A2']
            fecha_cell.value = f'Generado el {timezone.now().strftime("%d/%m/%Y %H:%M")}'

            fecha_cell.font = Font(size=10, italic=True)
            fecha_cell.alignment = Alignment(horizontal='center')
            
            # Espacio
            ws.append([])
            
            # Encabezados (sin Clave - campo no existe en BD)
            headers = ['#', 'Nombre', 'Direccion', 'Telefono', 'Email', 'Total Requisiciones', 'Estado']
            ws.append(headers)
            
            # Estilo de encabezados
            header_fill = PatternFill(start_color='632842', end_color='632842', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            for col_num, cell in enumerate(ws[4], 1):
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # Datos de centros
            for idx, centro in enumerate(centros, start=1):
                # Calcular total de requisiciones (origen + destino, consistente con serializer)
                total_requisiciones = 0
                if hasattr(centro, 'requisiciones_origen'):
                    total_requisiciones += centro.requisiciones_origen.count()
                if hasattr(centro, 'requisiciones_destino'):
                    total_requisiciones += centro.requisiciones_destino.count()
                
                ws.append([
                    idx,
                    centro.nombre,
                    centro.direccion or '',
                    centro.telefono or '',
                    getattr(centro, 'email', '') or '',
                    total_requisiciones,
                    'Activo' if centro.activo else 'Inactivo'
                ])
                
                # Estilo para filas
                row_num = idx + 4
                for cell in ws[row_num]:
                    cell.alignment = Alignment(vertical='center')
                    
                # Colorear estado
                estado_cell = ws.cell(row=row_num, column=7)
                if centro.activo:
                    estado_cell.fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
                    estado_cell.font = Font(color='155724', bold=True)
                else:
                    estado_cell.fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')
                    estado_cell.font = Font(color='721C24', bold=True)
            
            # Ajustar anchos de columna
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 50
            ws.column_dimensions['C'].width = 40
            ws.column_dimensions['D'].width = 18
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 20
            ws.column_dimensions['G'].width = 12
            
            # Agregar bordes
            from openpyxl.styles import Border, Side
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=7):
                for cell in row:
                    cell.border = thin_border
            
            # Resumen al final
            ws.append([])
            resumen_row = ws.max_row + 1
            ws.merge_cells(f'A{resumen_row}:C{resumen_row}')
            resumen_cell = ws[f'A{resumen_row}']
            resumen_cell.value = f'TOTAL DE CENTROS: {centros.count()}'
            resumen_cell.font = Font(bold=True, size=11)
            resumen_cell.alignment = Alignment(horizontal='left')
            
            # Generar respuesta
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename=Centros_Penitenciarios_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            wb.save(response)
            
            return response
            
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            return Response({
                'error': 'Error al exportar centros',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['post'], url_path='importar')
    def importar_excel(self, request):
        """
        Importa centros desde Excel.
        
        Formato esperado (columnas en orden):
        1. Nombre (REQUERIDO, único) - Nombre del centro penitenciario
        2. Direccion (opcional) - Dirección física
        3. Telefono (opcional) - Número de teléfono
        4. Email (opcional) - Correo electrónico
        5. Estado (opcional, default: Activo) - 'Activo' o 'Inactivo'
        
        Limites de seguridad:
        - Tamano maximo: configurado en IMPORT_MAX_FILE_SIZE_MB (default 10MB)
        - Filas maximas: configurado en IMPORT_MAX_ROWS (default 5000)
        - Extensiones: .xlsx, .xls
        
        Nota: Si el nombre ya existe, se actualizan los demás campos.
        """
        file = request.FILES.get('file')
        
        # HALLAZGO #13 FIX: Validar archivo CON magic bytes
        es_valido, error_msg = validar_archivo_excel(file)
        if not es_valido:
            logger.warning(f"HALLAZGO #13: Archivo Excel rechazado en importar_excel (CentroViewSet): {error_msg}")
            return Response({
                'error': 'Archivo invalido',
                'mensaje': error_msg
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            # ISS-001: Usar carga segura con limite real de bytes
            wb, error_carga = cargar_workbook_seguro(file)
            if wb is None:
                return Response({
                    'error': 'Error al procesar archivo',
                    'mensaje': error_carga
                }, status=status.HTTP_400_BAD_REQUEST)
            
            ws = wb.active
            
            # Validar numero de filas
            filas_validas, error_filas, num_filas = validar_filas_excel(ws)
            if not filas_validas:
                wb.close()  # Liberar recursos en modo read_only
                return Response({
                    'error': 'Archivo demasiado grande',
                    'mensaje': error_filas
                }, status=status.HTTP_400_BAD_REQUEST)
            
            creados = 0
            actualizados = 0
            errores = []
            
            # Procesar filas
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Saltar filas vacías
                    if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                        continue
                    
                    # Extraer datos - Formato: Nombre, Direccion, Telefono, Email, Estado
                    # (mínimo 1 columna: nombre)
                    nombre = row[0] if len(row) > 0 else None
                    direccion = row[1] if len(row) > 1 else None
                    telefono = row[2] if len(row) > 2 else None
                    email = row[3] if len(row) > 3 else None
                    estado = row[4] if len(row) > 4 else 'Activo'
                    
                    # Validar requeridos
                    if not nombre or str(nombre).strip() == '':
                        errores.append({'fila': row_idx, 'error': 'Nombre es requerido'})
                        continue
                    
                    nombre_limpio = str(nombre).strip()
                    
                    # Preparar datos
                    datos = {
                        'direccion': str(direccion).strip() if direccion else '',
                        'telefono': str(telefono).strip() if telefono else '',
                        'email': str(email).strip() if email else '',
                        'activo': str(estado).lower() in ['activo', 'si', 'sí', 'true', '1', 'yes'] if estado else True
                    }
                    
                    # Crear o actualizar usando nombre como identificador único
                    centro, created = Centro.objects.update_or_create(
                        nombre=nombre_limpio,
                        defaults=datos
                    )
                    
                    if created:
                        creados += 1
                    else:
                        actualizados += 1
                        
                except Exception as e:
                    errores.append(f'Fila {row_idx}: {str(e)}')
            
            return Response({
                'mensaje': 'Importacion completada',
                'resumen': {
                    'creados': creados,
                    'actualizados': actualizados,
                    'total_procesados': creados + actualizados,
                    'errores_encontrados': len(errores)
                },
                'errores': errores[:10] if errores else [],  # Maximo 10 errores
                'tiene_mas_errores': len(errores) > 10,
                'exito': len(errores) == 0
            }, status=status.HTTP_200_OK if len(errores) == 0 else status.HTTP_207_MULTI_STATUS)
            
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            logger.exception(f'Error en importacion de centros: {str(e)}')
            return Response({
                'error': 'Error al procesar el archivo',
                'mensaje': str(e),
                'sugerencia': 'Verifique que el archivo tenga el formato correcto: Nombre, Direccion, Telefono, Email, Estado'
            }, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['get'])
    def requisiciones(self, request, pk=None):
        """Obtiene todas las requisiciones de un centro"""
        try:
            centro = self.get_object()
            
            if not hasattr(centro, 'requisiciones'):
                return Response({
                    'centro': {
                        'id': centro.id,
                        'clave': centro.clave,
                        'nombre': centro.nombre
                    },
                    'requisiciones': [],
                    'total': 0,
                    'mensaje': 'No hay requisiciones disponibles'
                })
            
            requisiciones = centro.requisiciones_destino.all().order_by('-fecha_solicitud')
            
            # Agrupar por estado
            por_estado = {}
            for req in requisiciones:
                estado = req.estado
                if estado not in por_estado:
                    por_estado[estado] = 0
                por_estado[estado] += 1
            
            requisiciones_data = []
            for req in requisiciones:
                requisiciones_data.append({
                    'id': req.id,
                    'folio': req.folio,
                    'estado': req.estado,
                    'fecha_solicitud': req.fecha_solicitud,
                    'total_items': req.items.count() if hasattr(req, 'items') else 0
                })
            
            return Response({
                'centro': {
                    'id': centro.id,
                    'clave': centro.clave,
                    'nombre': centro.nombre
                },
                'estadisticas': {
                    'total': requisiciones.count(),
                    'por_estado': por_estado
                },
                'requisiciones': requisiciones_data
            })
            
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            return Response({
                'error': 'Error al obtener requisiciones',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['get'], url_path='plantilla')
    def plantilla_centros(self, request):
        """Descarga plantilla de Excel para importación de centros."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Centros'
        
        # Headers que coinciden con el modelo Centro
        headers = ['Nombre', 'Direccion', 'Telefono', 'Email', 'Estado']
        ws.append(headers)
        
        # Fila de ejemplo
        ws.append(['CENTRO PENITENCIARIO EJEMPLO', 'Av. Principal 123, Ciudad', '(555) 123-4567', 'centro@ejemplo.gob.mx', 'Activo'])
        
        # Aplicar formato a headers
        from openpyxl.styles import Font, PatternFill
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='9F2241', end_color='9F2241', fill_type='solid')
        for col in range(1, 6):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 45  # Nombre
        ws.column_dimensions['B'].width = 40  # Direccion
        ws.column_dimensions['C'].width = 18  # Telefono
        ws.column_dimensions['D'].width = 30  # Email
        ws.column_dimensions['E'].width = 12  # Estado
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=Plantilla_Centros.xlsx'
        wb.save(response)
        return response


class LoteViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestionar lotes.
    
    Funcionalidades:
    - CRUD completo
    - Filtrado por producto
    - Filtrado por estado de caducidad
    - Busqueda por numero de lote
    - Validaciones de integridad
    
    ISS-MEDICO FIX: Médicos pueden ver lotes (para crear requisiciones)
    pero NO pueden crear/editar/eliminar lotes.
    """
    queryset = Lote.objects.select_related('producto').all()
    serializer_class = LoteSerializer
    # ISS-MEDICO FIX: Médico puede leer pero NO escribir
    permission_classes = [IsCentroOwnResourcesOnly]
    pagination_class = CustomPagination

    def get_queryset(self):
        """
        Filtra lotes segun parametros.
        
        Parametros:
        - producto: ID del producto
        - activo: true/false
        - caducidad: vencido/critico/proximo/normal
        - search: busqueda por numero de lote
        - centro: ID del centro o 'central' para farmacia (solo admin/farmacia/vista)
        
        Seguridad: Usuarios de centro solo ven lotes de su centro.
        Admin/farmacia por defecto ven solo Farmacia Central, pueden filtrar con ?centro=.
        """
        queryset = Lote.objects.select_related('producto', 'centro').all()
        
        # SEGURIDAD: Filtrar por centro segun rol
        user = self.request.user
        
        # ISS-FIX: Usuarios de centro pueden ver lotes de farmacia central
        # cuando están creando requisiciones (para_requisicion=true)
        para_requisicion = self.request.query_params.get('para_requisicion', '').lower() == 'true'
        
        if not is_farmacia_or_admin(user):
            # Usuario de centro
            user_centro = get_user_centro(user)
            if not user_centro:
                return Lote.objects.none()
            
            if para_requisicion:
                # ISS-FIX: Para crear requisiciones, mostrar lotes de FARMACIA CENTRAL
                # porque las requisiciones se surten desde farmacia central
                queryset = queryset.filter(centro__isnull=True)
            else:
                # Por defecto: solo lotes de SU centro
                queryset = queryset.filter(centro=user_centro)
        else:
            # Admin/farmacia/vista: por defecto solo Farmacia Central
            # ISS-FIX: Evitar mostrar lotes duplicados de centros en vista de Farmacia
            centro_param = self.request.query_params.get('centro')
            if centro_param:
                if centro_param == 'central':
                    # Filtrar solo farmacia central (centro=NULL)
                    queryset = queryset.filter(centro__isnull=True)
                elif centro_param == 'todos':
                    # Mostrar todos los lotes (solo si se pide explícitamente)
                    pass  # No filtrar
                else:
                    # Filtrar por centro específico
                    queryset = queryset.filter(centro_id=centro_param)
            else:
                # Por defecto: solo lotes de Farmacia Central (centro=NULL)
                queryset = queryset.filter(centro__isnull=True)
        
        # Filtrar por producto
        producto = self.request.query_params.get('producto')
        if producto:
            queryset = queryset.filter(producto_id=producto)
        
        # ISS-FIX: Parámetro para incluir lotes inactivos (para reabastecimiento)
        incluir_inactivos = self.request.query_params.get('incluir_inactivos', '').lower() == 'true'
        
        # Filtrar por activo (el campo real en la BD)
        activo = self.request.query_params.get('activo')
        if activo is not None:
            if activo.lower() in ['true', '1', 'si']:
                queryset = queryset.filter(activo=True)
            elif activo.lower() in ['false', '0', 'no']:
                queryset = queryset.filter(activo=False)
        elif not incluir_inactivos:
            # Por defecto: solo lotes activos (a menos que se pida incluir inactivos)
            queryset = queryset.filter(activo=True)
        # Si incluir_inactivos=true y no hay filtro activo, mostrar todos
        
        # Busqueda por numero de lote, clave o nombre producto (ISS-003)
        search = self.request.query_params.get('search')
        if search and search.strip():
            search_term = search.strip()
            queryset = queryset.filter(
                Q(numero_lote__icontains=search_term) |
                Q(producto__clave__icontains=search_term) |
                Q(producto__nombre__icontains=search_term)
            )
        
        # Filtrar por estado de caducidad segun especificacion SIFP:
        # Normal: > 6 meses (180 dias)
        # Proximo: 3-6 meses (90-180 dias)
        # Critico: < 3 meses (90 dias)
        # Vencido: < 0 dias
        caducidad = self.request.query_params.get('caducidad')
        if caducidad:
            from datetime import date, timedelta
            hoy = date.today()
            
            if caducidad == 'vencido':
                queryset = queryset.filter(fecha_caducidad__lt=hoy)
            elif caducidad == 'critico':
                # Menos de 3 meses (< 90 dias) pero no vencido
                queryset = queryset.filter(
                    fecha_caducidad__gte=hoy,
                    fecha_caducidad__lt=hoy + timedelta(days=90)
                )
            elif caducidad == 'proximo':
                # Entre 3 y 6 meses (90-180 dias)
                queryset = queryset.filter(
                    fecha_caducidad__gte=hoy + timedelta(days=90),
                    fecha_caducidad__lt=hoy + timedelta(days=180)
                )
            elif caducidad == 'normal':
                # Mas de 6 meses (> 180 dias)
                queryset = queryset.filter(fecha_caducidad__gte=hoy + timedelta(days=180))
        
        # Filtrar por stock minimo (para catalogo de requisiciones)
        stock_min = self.request.query_params.get('stock_min')
        if stock_min:
            try:
                queryset = queryset.filter(cantidad_actual__gte=int(stock_min))
            except (ValueError, TypeError):
                pass
        
        # Filtrar por existencia de stock (con_stock/sin_stock)
        con_stock = self.request.query_params.get('con_stock')
        if con_stock == 'con_stock':
            queryset = queryset.filter(cantidad_actual__gt=0)
        elif con_stock == 'sin_stock':
            queryset = queryset.filter(cantidad_actual=0)
        
        # Filtrar solo lotes disponibles (no vencidos) para el catalogo
        solo_disponibles = self.request.query_params.get('solo_disponibles')
        if solo_disponibles == 'true':
            from datetime import date
            queryset = queryset.filter(
                activo=True,
                fecha_caducidad__gt=date.today()
            )
        
        # PERFORMANCE: Prefetch documentos para evitar N+1 queries
        return queryset.order_by('-created_at').prefetch_related('documentos')
    
    @transaction.atomic  # HALLAZGO #10 FIX: Garantizar atomicidad
    def create(self, request, *args, **kwargs):
        """Crea un nuevo lote con validaciones"""
        try:
            serializer = self.get_serializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            self.perform_create(serializer)
            
            headers = self.get_success_headers(serializer.data)
            return Response(
                serializer.data, 
                status=status.HTTP_201_CREATED, 
                headers=headers
            )
        except serializers.ValidationError as e:
            return Response(
                {'error': 'Error de validacion', 'detalles': e.detail}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            # HALLAZGO #12 FIX: No exponer stack trace completo
            logger.error(f"Error al crear lote: {str(e)}", exc_info=True)
            return Response(
                {'error': 'Error al crear lote. Contacte al administrador.'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def update(self, request, *args, **kwargs):
        """Actualiza un lote existente.
        
        ISS-010 FIX: Validación explícita de permisos para evitar IDOR.
        Solo admin/farmacia pueden modificar lotes de farmacia central.
        Usuarios de centro solo pueden modificar lotes de SU centro.
        """
        try:
            partial = kwargs.pop('partial', False)
            instance = self.get_object()
            
            # ISS-010: Validar permisos de escritura sobre este lote específico
            user = request.user
            if not is_farmacia_or_admin(user):
                user_centro = get_user_centro(user)
                lote_centro = instance.centro
                
                # Si el lote es de farmacia central o de otro centro, denegar
                if lote_centro is None or (user_centro and lote_centro.pk != user_centro.pk):
                    logger.warning(
                        f"ISS-010: Intento de modificación no autorizada de lote. "
                        f"Usuario={user.username}, lote={instance.numero_lote}, "
                        f"lote_centro={lote_centro}, user_centro={user_centro}"
                    )
                    return Response(
                        {'error': 'No tiene permisos para modificar este lote'},
                        status=status.HTTP_403_FORBIDDEN
                    )
            
            serializer = self.get_serializer(instance, data=request.data, partial=partial)
            serializer.is_valid(raise_exception=True)
            self.perform_update(serializer)
            
            return Response(serializer.data)
        except serializers.ValidationError as e:
            return Response(
                {'error': 'Error de validacion', 'detalles': e.detail}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            return Response(
                {'error': 'Error al actualizar lote', 'mensaje': str(e)}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def destroy(self, request, *args, **kwargs):
        """
        Elimina un lote.
        
        ISS-010 FIX: Validación explícita de permisos para evitar IDOR.
        
        Validaciones:
        - Permisos de escritura sobre el lote
        - No puede eliminarse si tiene movimientos asociados
        """
        instance = self.get_object()
        
        # ISS-010: Validar permisos de escritura sobre este lote específico
        user = request.user
        if not is_farmacia_or_admin(user):
            user_centro = get_user_centro(user)
            lote_centro = instance.centro
            
            # Si el lote es de farmacia central o de otro centro, denegar
            if lote_centro is None or (user_centro and lote_centro.pk != user_centro.pk):
                logger.warning(
                    f"ISS-010: Intento de eliminación no autorizada de lote. "
                    f"Usuario={user.username}, lote={instance.numero_lote}, "
                    f"lote_centro={lote_centro}, user_centro={user_centro}"
                )
                return Response(
                    {'error': 'No tiene permisos para eliminar este lote'},
                    status=status.HTTP_403_FORBIDDEN
                )
        
        try:
            # Verificar si tiene movimientos
            if Movimiento.objects.filter(lote=instance).exists():
                total_movimientos = Movimiento.objects.filter(lote=instance).count()
                
                return Response({
                    'error': 'No se puede eliminar el lote',
                    'razon': 'Tiene movimientos asociados',
                    'total_movimientos': total_movimientos,
                    'sugerencia': 'Marque el lote como inactivo en lugar de eliminarlo'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Si no tiene movimientos, se puede eliminar
            numero_lote = instance.numero_lote
            producto_clave = instance.producto.clave
            instance.delete()
            
            return Response({
                'mensaje': 'Lote eliminado exitosamente',
                'lote_eliminado': numero_lote,
                'producto': producto_clave
            }, status=status.HTTP_204_NO_CONTENT)
            
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            return Response({
                'error': 'Error al eliminar lote',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path='exportar-pdf')
    def exportar_pdf(self, request):
        """
        Genera PDF de inventario de lotes con filtros opcionales.
        
        Usa los mismos filtros que get_queryset para consistencia:
        - producto: ID del producto
        - activo: true/false
        - search: búsqueda en número de lote o producto
        - caducidad: vencido/critico/proximo/normal
        - con_stock: con_stock/sin_stock
        - centro: ID del centro o 'central'
        
        Respeta permisos de usuario:
        - Usuarios de centro solo ven lotes de su centro
        - Admin/Farmacia/Vista ven todo
        """
        from core.utils.pdf_reports import generar_reporte_lotes
        
        try:
            # Usar get_queryset que ya aplica filtros y permisos
            queryset = self.get_queryset()
            
            # Limitar a 500 lotes para PDF
            lotes = queryset[:500]
            
            # Preparar datos para el PDF
            lotes_data = []
            for lote in lotes:
                lotes_data.append({
                    'producto_clave': getattr(lote.producto, 'clave', '') if lote.producto else '',
                    'producto_nombre': getattr(lote.producto, 'nombre', '') if lote.producto else '',
                    'numero_lote': lote.numero_lote or '',
                    'fecha_fabricacion': lote.fecha_fabricacion.strftime('%Y-%m-%d') if lote.fecha_fabricacion else '',
                    'fecha_caducidad': lote.fecha_caducidad.strftime('%Y-%m-%d') if lote.fecha_caducidad else '',
                    'fecha_caducidad_raw': lote.fecha_caducidad,
                    'cantidad_inicial': lote.cantidad_inicial,
                    'cantidad_actual': lote.cantidad_actual,
                    'centro_nombre': getattr(lote.centro, 'nombre', 'Almacén Central') if lote.centro else 'Almacén Central',
                    'activo': lote.activo,
                })
            
            # Preparar filtros para mostrar en el PDF
            filtros = {}
            if request.query_params.get('producto'):
                try:
                    producto = Producto.objects.get(pk=request.query_params.get('producto'))
                    filtros['producto'] = f"{producto.clave} - {producto.nombre}"
                except Producto.DoesNotExist:
                    pass
            if request.query_params.get('centro'):
                centro_param = request.query_params.get('centro')
                if centro_param == 'central':
                    filtros['centro'] = 'Almacén Central'
                else:
                    try:
                        centro = Centro.objects.get(pk=centro_param)
                        filtros['centro'] = centro.nombre
                    except Centro.DoesNotExist:
                        pass
            if request.query_params.get('caducidad'):
                filtros['caducidad'] = request.query_params.get('caducidad')
            if request.query_params.get('con_stock'):
                filtros['con_stock'] = request.query_params.get('con_stock')
            if request.query_params.get('activo'):
                filtros['activo'] = request.query_params.get('activo')
            if request.query_params.get('search'):
                filtros['busqueda'] = request.query_params.get('search')
            
            pdf_buffer = generar_reporte_lotes(lotes_data, filtros)
            
            response = HttpResponse(
                pdf_buffer.getvalue(),
                content_type='application/pdf'
            )
            response['Content-Disposition'] = f'attachment; filename="Inventario_Lotes_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
            
            return response
            
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            return Response({
                'error': 'Error al generar PDF de lotes',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path='exportar-excel')
    def exportar_excel(self, request):
        """
        Exporta lotes aplicando los mismos filtros de listado.
        
        ISS-DB: Incluye campos principales de la tabla lotes:
        - clave (de producto), nombre_comercial (de producto)
        - numero_lote, fecha_fabricacion, fecha_caducidad
        - cantidad_inicial, cantidad_actual
        - precio_unitario, numero_contrato, marca
        - centro (nombre), activo
        """
        try:
            lotes = self.get_queryset()
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Lotes'

            ws.merge_cells('A1:N1')
            ws['A1'] = 'REPORTE DE LOTES - SISTEMA DE INVENTARIO FARMACEUTICO PENITENCIARIO'
            ws['A1'].font = Font(bold=True, size=14, color='632842')
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

            ws.append([])
            # Headers sin ubicación
            headers = [
                '#', 'Clave', 'Nombre Producto', 'Nombre Comercial', 'Número Lote',
                'Fecha Fabricación', 'Fecha Caducidad',
                'Cantidad Inicial', 'Cantidad Actual',
                'Precio Unitario', 'Número Contrato', 'Marca',
                'Centro', 'Activo'
            ]
            ws.append(headers)
            header_fill = PatternFill(start_color='632842', end_color='632842', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            for cell in ws[3]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            for idx, lote in enumerate(lotes, 1):
                ws.append([
                    idx,
                    getattr(lote.producto, 'clave', '') or '',
                    getattr(lote.producto, 'nombre', '') or '',
                    getattr(lote.producto, 'nombre_comercial', '') or '',
                    lote.numero_lote or '',
                    lote.fecha_fabricacion.strftime('%Y-%m-%d') if lote.fecha_fabricacion else '',
                    lote.fecha_caducidad.strftime('%Y-%m-%d') if lote.fecha_caducidad else '',
                    lote.cantidad_inicial,
                    lote.cantidad_actual,
                    float(lote.precio_unitario) if lote.precio_unitario else 0.00,
                    lote.numero_contrato or '',
                    lote.marca or '',
                    getattr(lote.centro, 'nombre', 'Almacén Central') if lote.centro else 'Almacén Central',
                    'Sí' if lote.activo else 'No'
                ])

            # Ajustar anchos de columna (sin ubicación)
            column_widths = [6, 15, 30, 20, 15, 14, 14, 12, 12, 12, 18, 15, 18, 8]
            for col_idx, width in enumerate(column_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename=Lotes_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            wb.save(response)
            return response
        except Exception as exc:
            # traceback removido por seguridad (ISS-008)
            return Response({'error': 'Error al exportar lotes', 'mensaje': str(exc)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['post'], url_path='importar-excel')
    def importar_excel(self, request):
        """
        Importa lotes desde Excel con detección automática de columnas.
        
        COLUMNAS OBLIGATORIAS:
        - Clave Producto* (REQUERIDO): clave única del producto
        - Nombre Producto* (REQUERIDO): nombre del producto (debe coincidir con clave)
        - Numero Lote* (REQUERIDO): identificador del lote
        - Fecha Caducidad* (REQUERIDO): fecha de vencimiento
        - Cantidad Inicial* (REQUERIDO): cantidad recibida
        
        COLUMNAS OPCIONALES:
        - Cantidad Actual: default = cantidad inicial
        - Fecha Fabricacion
        - Precio Unitario: default = 0
        - Numero Contrato
        - Marca
        - Centro/Centro ID: ID o nombre del centro
        - Activo: estado del lote
        
        IMPORTANTE: El sistema verifica que CLAVE y NOMBRE coincidan con el producto
        en la base de datos. Si hay discrepancia, se reporta error.
        
        Limites de seguridad:
        - Tamano maximo: 10MB
        - Filas maximas: 5000
        - Extensiones: .xlsx, .xls
        """
        file = request.FILES.get('file')
        
        # HALLAZGO #13 FIX: Validar archivo CON magic bytes
        es_valido, error_msg = validar_archivo_excel(file)
        if not es_valido:
            logger.warning(f"HALLAZGO #13: Archivo Excel rechazado en importar_excel (LoteViewSet): {error_msg}")
            return Response({
                'error': 'Archivo invalido',
                'mensaje': error_msg
            }, status=status.HTTP_400_BAD_REQUEST)

        try:
            # ISS-001: Usar carga segura con limite real de bytes
            wb, error_carga = cargar_workbook_seguro(file)
            if wb is None:
                return Response({
                    'error': 'Error al procesar archivo',
                    'mensaje': error_carga
                }, status=status.HTTP_400_BAD_REQUEST)
            
            ws = wb.active
            
            # Validar numero de filas
            filas_validas, error_filas, num_filas = validar_filas_excel(ws)
            if not filas_validas:
                wb.close()
                return Response({
                    'error': 'Archivo demasiado grande',
                    'mensaje': error_filas
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # ISS-FIX: Detectar fila de encabezados automáticamente
            # Buscar fila que contenga palabras clave como "lote", "producto", "cantidad"
            header_row_idx = 1
            col_map = {}
            
            # Mapeo de nombres de columna a campos internos
            # IMPORTANTE: El orden importa - los más específicos primero
            COLUMN_ALIASES = {
                'nombre_producto': ['nombre producto', 'nombre del producto', 'producto nombre', 'descripcion', 'descripción'],
                'nombre_comercial': ['nombre comercial', 'comercial', 'marca comercial'],  # Solo referencia, no se usa
                'producto': ['clave producto', 'clave', 'codigo producto', 'codigo', 'código', 'sku'],
                'numero_lote': ['numero lote', 'número lote', 'lote', 'no. lote', 'no lote', 'num lote'],
                'fecha_caducidad': ['fecha caducidad', 'caducidad', 'vencimiento', 'fecha vencimiento', 'expira', 'fecha expiracion'],
                'cantidad_inicial': ['cantidad inicial', 'cantidad', 'cant inicial', 'cant', 'qty'],
                'cantidad_actual': ['cantidad actual', 'cant actual', 'stock', 'existencia'],
                'fecha_fabricacion': ['fecha fabricacion', 'fecha fabricación', 'fabricacion', 'fabricación', 'manufactura'],
                'precio_unitario': ['precio unitario', 'precio', 'costo', 'valor', 'precio unit'],
                'numero_contrato': ['numero contrato', 'número contrato', 'contrato', 'no. contrato', 'no contrato'],
                'marca': ['marca', 'laboratorio', 'fabricante'],
                'centro': ['centro', 'centro id', 'centro_id', 'destino'],
                'activo': ['activo', 'estado', 'status'],
            }
            
            def normalizar_header(val):
                if not val:
                    return ''
                return str(val).lower().strip().replace('_', ' ').replace('-', ' ')
            
            # Buscar fila con encabezados (primeras 5 filas)
            for row_num in range(1, min(6, ws.max_row + 1)):
                row_values = [cell.value for cell in ws[row_num]]
                headers_encontrados = 0
                temp_map = {}
                
                for col_idx, val in enumerate(row_values):
                    header_norm = normalizar_header(val)
                    if not header_norm:
                        continue
                    
                    # Buscar coincidencia EXACTA primero (prioridad)
                    for field, aliases in COLUMN_ALIASES.items():
                        if field not in temp_map and header_norm in aliases:
                            temp_map[field] = col_idx
                            headers_encontrados += 1
                            break
                
                # Si encontramos al menos 3 columnas relevantes, es la fila de headers
                if headers_encontrados >= 3:
                    header_row_idx = row_num
                    col_map = temp_map
                    break
            
            # Si no hay mapa, usar orden por defecto (sin ubicacion)
            if not col_map:
                col_map = {
                    'producto': 0, 'numero_lote': 1, 'fecha_caducidad': 2,
                    'cantidad_inicial': 3, 'cantidad_actual': 4, 'fecha_fabricacion': 5,
                    'precio_unitario': 6, 'numero_contrato': 7, 'marca': 8,
                    'centro': 9
                }
            
            def get_val(row, field, default=None):
                if field not in col_map:
                    return default
                idx = col_map[field]
                if idx < len(row):
                    val = row[idx]
                    return val if val not in [None, '', 'None'] else default
                return default
            
            exitos = []
            errores = []

            for row_idx, row in enumerate(ws.iter_rows(min_row=header_row_idx + 1, values_only=True), start=header_row_idx + 1):
                try:
                    # Saltar filas vacías
                    if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                        continue
                    
                    # Extraer valores usando el mapa de columnas
                    # FIX: AMBOS son OBLIGATORIOS: Clave Y Nombre deben coincidir con producto en BD
                    clave_producto = get_val(row, 'producto')  # Columna "Clave Producto" - OBLIGATORIA
                    nombre_producto = get_val(row, 'nombre_producto')  # Columna "Nombre Producto" - OBLIGATORIA
                    numero_lote = get_val(row, 'numero_lote')
                    fecha_cad = get_val(row, 'fecha_caducidad')
                    cantidad_inicial = get_val(row, 'cantidad_inicial')
                    cantidad_actual = get_val(row, 'cantidad_actual')
                    fecha_fab = get_val(row, 'fecha_fabricacion')
                    precio_unitario = get_val(row, 'precio_unitario')
                    numero_contrato = get_val(row, 'numero_contrato')
                    marca = get_val(row, 'marca')
                    centro_ref = get_val(row, 'centro')

                    # FIX: Validar que AMBOS campos estén presentes: Clave Y Nombre
                    if not clave_producto:
                        errores.append({
                            'fila': row_idx, 
                            'error': f'Clave de producto es OBLIGATORIA. Nombre proporcionado: {nombre_producto or "N/A"}'
                        })
                        continue
                    
                    if not nombre_producto:
                        errores.append({
                            'fila': row_idx, 
                            'error': f'Nombre de producto es OBLIGATORIO. Clave proporcionada: {clave_producto}'
                        })
                        continue
                    
                    if not numero_lote:
                        errores.append({'fila': row_idx, 'error': 'Número de lote es obligatorio'})
                        continue

                    # FIX: Buscar producto por CLAVE y verificar que NOMBRE coincida
                    clave_busqueda = str(clave_producto).strip()
                    nombre_busqueda = str(nombre_producto).strip()
                    
                    try:
                        producto = Producto.objects.get(clave__iexact=clave_busqueda)
                    except Producto.DoesNotExist:
                        errores.append({
                            'fila': row_idx, 
                            'error': f'Clave "{clave_busqueda}" no encontrada en catálogo. '
                                     f'Nombre: "{nombre_busqueda}". Verifique el catálogo de productos.'
                        })
                        continue
                    
                    # VERIFICACIÓN CRÍTICA: El nombre en Excel debe coincidir con el nombre en BD
                    # Comparación flexible: ignorar mayúsculas/minúsculas y espacios extra
                    nombre_bd_normalizado = producto.nombre.strip().lower()
                    nombre_excel_normalizado = nombre_busqueda.lower()
                    
                    # Verificar si el nombre coincide (completo o parcialmente al inicio)
                    if not (nombre_bd_normalizado == nombre_excel_normalizado or 
                            nombre_bd_normalizado.startswith(nombre_excel_normalizado) or
                            nombre_excel_normalizado.startswith(nombre_bd_normalizado)):
                        errores.append({
                            'fila': row_idx, 
                            'error': f'DISCREPANCIA: Clave "{clave_busqueda}" corresponde a '
                                     f'"{producto.nombre}" en BD, pero Excel dice "{nombre_busqueda}". '
                                     f'Verifique que clave y nombre sean correctos.'
                        })
                        continue

                    # Parsear fecha de caducidad (varios formatos)
                    try:
                        fecha_cad_val = None
                        if fecha_cad:
                            if isinstance(fecha_cad, (datetime, date)):
                                fecha_cad_val = fecha_cad.date() if hasattr(fecha_cad, 'date') else fecha_cad
                            else:
                                fecha_str = str(fecha_cad).strip()
                                # Intentar varios formatos
                                for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d']:
                                    try:
                                        fecha_cad_val = datetime.strptime(fecha_str.split()[0], fmt).date()
                                        break
                                    except:
                                        continue
                        if not fecha_cad_val:
                            raise ValueError("No se pudo parsear fecha")
                    except Exception:
                        errores.append({'fila': row_idx, 'error': f'Fecha de caducidad invalida: {fecha_cad}'})
                        continue

                    # Parsear fecha de fabricacion (opcional)
                    fecha_fab_val = None
                    if fecha_fab:
                        try:
                            if isinstance(fecha_fab, (datetime, date)):
                                fecha_fab_val = fecha_fab.date() if hasattr(fecha_fab, 'date') else fecha_fab
                            else:
                                fecha_str = str(fecha_fab).strip()
                                for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%Y/%m/%d']:
                                    try:
                                        fecha_fab_val = datetime.strptime(fecha_str.split()[0], fmt).date()
                                        break
                                    except:
                                        continue
                        except Exception:
                            pass  # Ignorar fecha fabricacion invalida

                    # Parsear cantidades (más flexible)
                    try:
                        cant_ini = int(float(cantidad_inicial)) if cantidad_inicial not in [None, ''] else 0
                        cant_act = int(float(cantidad_actual)) if cantidad_actual not in [None, ''] else cant_ini
                        if cant_ini <= 0:
                            cant_ini = 1  # Minimo 1
                        if cant_act < 0:
                            cant_act = 0
                    except Exception:
                        errores.append({'fila': row_idx, 'error': f'Cantidades invalidas: inicial={cantidad_inicial}, actual={cantidad_actual}'})
                        continue

                    # Parsear precio unitario (opcional)
                    precio_val = 0
                    if precio_unitario not in [None, '']:
                        try:
                            precio_val = float(str(precio_unitario).replace(',', '.'))
                            if precio_val < 0:
                                precio_val = 0
                        except Exception:
                            pass  # Usar 0 si no es valido

                    # Preparar defaults para update_or_create
                    defaults = {
                        'fecha_caducidad': fecha_cad_val or date.today(),
                        'cantidad_inicial': cant_ini,
                        'cantidad_actual': cant_act,
                        'precio_unitario': precio_val,
                        'numero_contrato': str(numero_contrato).strip()[:100] if numero_contrato else '',
                        'marca': str(marca).strip()[:100] if marca else '',
                    }
                    
                    if fecha_fab_val:
                        defaults['fecha_fabricacion'] = fecha_fab_val
                    
                    # ISS-FIX: Asignar centro si se proporciona (por ID o nombre)
                    if centro_ref:
                        try:
                            centro_str = str(centro_ref).strip()
                            # Intentar por ID numerico primero
                            if centro_str.isdigit():
                                centro = Centro.objects.get(pk=int(centro_str))
                            else:
                                # Buscar por nombre
                                centro = Centro.objects.filter(nombre__icontains=centro_str).first()
                            if centro:
                                defaults['centro'] = centro
                        except (Centro.DoesNotExist, ValueError):
                            pass  # Ignorar centro invalido

                    lote, created = Lote.objects.update_or_create(
                        producto=producto,
                        numero_lote=str(numero_lote).strip().upper(),
                        defaults=defaults
                    )
                    exitos.append({'fila': row_idx, 'lote_id': lote.id, 'numero_lote': lote.numero_lote, 'created': created})
                except Exception as exc:
                    errores.append({'fila': row_idx, 'error': str(exc)})

            status_code = status.HTTP_200_OK if not errores else status.HTTP_207_MULTI_STATUS
            return Response({
                'mensaje': 'Importacion de lotes completada',
                'resumen': {
                    'exitos': len(exitos),
                    'errores': len(errores),
                    'total': len(exitos) + len(errores)
                },
                'exitos': exitos,
                'errores': errores
            }, status=status_code)
        except Exception as exc:
            # traceback removido por seguridad (ISS-008)
            return Response({'error': 'Error al procesar archivo', 'mensaje': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get'], url_path='plantilla')
    def plantilla_lotes(self, request):
        """
        Descarga plantilla Excel para importación de lotes.
        
        COLUMNAS OBLIGATORIAS (en orden):
        1. Clave Producto* (REQUERIDO) - Clave única del producto
        2. Nombre Producto* (REQUERIDO) - Debe coincidir con la clave
        3. Nombre Comercial (referencia) - Solo informativo
        4. Numero Lote* (REQUERIDO) - Identificador único del lote
        5. Fecha Caducidad* (REQUERIDO, YYYY-MM-DD)
        6. Cantidad Inicial* (REQUERIDO) - Cantidad recibida
        
        COLUMNAS OPCIONALES:
        7. Fecha Fabricacion (YYYY-MM-DD)
        8. Precio Unitario (default = 0)
        9. Numero Contrato
        10. Marca
        11. Activo (default = Activo)
        
        IMPORTANTE: El sistema verifica que CLAVE y NOMBRE coincidan con el producto
        en la base de datos. Si hay discrepancia (clave correcta pero nombre incorrecto),
        se reportará un error para evitar confusiones.
        
        NOTA: La ubicación se asigna automáticamente como "Almacén Central"
        y el centro queda NULL (representa Farmacia Central).
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Lotes'
        
        # Headers con Nombre Comercial como referencia visual
        headers = [
            'Clave Producto', 'Nombre Producto', 'Nombre Comercial', 'Numero Lote', 
            'Fecha Caducidad', 'Cantidad Inicial',
            'Fecha Fabricacion', 'Precio Unitario',
            'Numero Contrato', 'Marca', 'Activo'
        ]
        ws.append(headers)
        
        # ============================================================
        # FILAS DE EJEMPLO - ELIMINAR ANTES DE USAR CON DATOS REALES
        # Estas filas son solo para mostrar el formato correcto.
        # ============================================================
        from datetime import date, timedelta
        fecha_cad_ejemplo = (date.today() + timedelta(days=365)).strftime('%Y-%m-%d')
        fecha_fab_ejemplo = date.today().strftime('%Y-%m-%d')
        
        ws.append([
            'PRUEBA001', '[EJEMPLO] Paracetamol - ELIMINAR', 'Tempra', 'LOTE-PRUEBA-001', 
            fecha_cad_ejemplo, 100,
            fecha_fab_ejemplo, 25.50,
            'CONT-PRUEBA-001', '[EJEMPLO] Laboratorio - ELIMINAR', 'Activo'
        ])
        ws.append([
            'PRUEBA002', '[EJEMPLO] Ibuprofeno - ELIMINAR', 'Advil', 'LOTE-PRUEBA-002', 
            fecha_cad_ejemplo, 50,
            fecha_fab_ejemplo, 18.75,
            'CONT-PRUEBA-002', '[EJEMPLO] Farmacéutica - ELIMINAR', 'Activo'
        ])
        ws.append([
            'PRUEBA003', '[EJEMPLO] Jeringa - ELIMINAR', '', 'LOTE-PRUEBA-003', 
            fecha_cad_ejemplo, 200,
            '', 5.00,
            '', '[EJEMPLO] Material - ELIMINAR', 'Activo'
        ])
        
        # Aplicar formato a headers
        from openpyxl.styles import Font, PatternFill, Alignment
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='9F2241', end_color='9F2241', fill_type='solid')
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
        
        # Estilo para filas de ejemplo (gris, itálica - sin fondo de color)
        example_font = Font(italic=True, color='888888')
        for row_num in range(2, 5):  # Filas 2, 3, 4 (ejemplos)
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col)
                cell.font = example_font
        
        # Ajustar ancho de columnas
        column_widths = {
            'A': 15,  # Clave Producto
            'B': 40,  # Nombre Producto (referencia)
            'C': 18,  # Nombre Comercial
            'D': 20,  # Numero Lote
            'E': 16,  # Fecha Caducidad
            'F': 16,  # Cantidad Inicial
            'G': 18,  # Fecha Fabricacion
            'H': 15,  # Precio Unitario
            'I': 18,  # Numero Contrato
            'J': 35,  # Marca (más ancho para ver el texto de ejemplo)
            'K': 12,  # Activo
        }
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # ============================================================
        # HOJA DE INSTRUCCIONES
        # ============================================================
        ws_instrucciones = wb.create_sheet(title='INSTRUCCIONES')
        
        instrucciones = [
            ['╔════════════════════════════════════════════════════════════════════╗'],
            ['║    INSTRUCCIONES PARA IMPORTACIÓN DE LOTES                        ║'],
            ['╚════════════════════════════════════════════════════════════════════╝'],
            [''],
            ['⚠️  IMPORTANTE: Las filas grises en la hoja "Lotes" son EJEMPLOS.'],
            ['    ELIMÍNELAS antes de cargar sus datos reales.'],
            [''],
            ['════════════════════════════════════════════════════════════════════════'],
            ['⚠️  VERIFICACIÓN DE DOBLE CAMPO: CLAVE + NOMBRE'],
            ['════════════════════════════════════════════════════════════════════════'],
            ['El sistema verifica que AMBOS campos (Clave y Nombre) coincidan con'],
            ['el producto en la base de datos. Si hay discrepancia, se reportará error.'],
            ['Esto evita errores al sumar cantidades a productos incorrectos.'],
            [''],
            ['────────────────────────────────────────────────────────────────────────'],
            ['COLUMNAS REQUERIDAS (obligatorias):'],
            ['────────────────────────────────────────────────────────────────────────'],
            ['• Clave Producto* - OBLIGATORIA: Clave única del producto en el sistema'],
            ['• Nombre Producto* - OBLIGATORIO: Debe coincidir con la clave'],
            ['• Nombre Comercial - Solo referencia visual (ej: Tempra, Advil)'],
            ['• Numero Lote*    - Identificador único del lote'],
            ['• Fecha Caducidad* - Formato: YYYY-MM-DD (ej: 2026-12-31)'],
            ['• Cantidad Inicial* - Cantidad de unidades recibidas'],
            [''],
            ['────────────────────────────────────────────────────────────────────────'],
            ['COLUMNAS OPCIONALES:'],
            ['────────────────────────────────────────────────────────────────────────'],
            ['• Fecha Fabricacion - Formato: YYYY-MM-DD'],
            ['• Precio Unitario  - Precio por unidad (default: 0)'],
            ['• Numero Contrato  - Referencia del contrato de adquisición'],
            ['• Marca            - Laboratorio o fabricante'],
            ['• Activo           - Estado del lote (default: Activo)'],
            [''],
            ['────────────────────────────────────────────────────────────────────────'],
            ['NOTAS:'],
            ['────────────────────────────────────────────────────────────────────────'],
            ['• Los lotes se asignan automáticamente al Almacén Central (FARMACIA).'],
            ['• El PRODUCTO debe existir antes de importar lotes.'],
            ['• Verifique la CLAVE y NOMBRE del producto en el catálogo.'],
            ['• Si el lote ya existe (mismo producto + número de lote), se ACTUALIZARÁ.'],
            ['• La cantidad_actual se inicializa igual a cantidad_inicial.'],
            ['• El stock del producto se actualiza automáticamente.'],
            ['• Fechas aceptadas: YYYY-MM-DD, DD/MM/YYYY, DD-MM-YYYY'],
            ['• Máximo 5000 lotes por archivo.'],
            ['• Tamaño máximo de archivo: 10 MB.'],
            [''],
        ]
        
        for row in instrucciones:
            ws_instrucciones.append(row)
        
        # Formato de instrucciones
        ws_instrucciones.column_dimensions['A'].width = 80
        for row in ws_instrucciones.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=False)
        
        # Poner hoja de Lotes como activa
        wb.active = ws
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=Plantilla_Lotes.xlsx'
        wb.save(response)
        return response
    
    @action(detail=False, methods=['get'])
    def por_vencer(self, request):
        """
        Obtiene lotes proximos a vencer.
        
        Parametros:
        - dias: numero de dias (default: 30)
        """
        try:
            from datetime import date, timedelta
            
            dias = int(request.query_params.get('dias', 30))
            fecha_limite = date.today() + timedelta(days=dias)
            
            lotes = Lote.objects.select_related('producto').filter(
                activo=True,
                cantidad_actual__gt=0,
                fecha_caducidad__lte=fecha_limite
            ).order_by('fecha_caducidad')
            
            serializer = self.get_serializer(lotes, many=True)
            
            return Response({
                'total': lotes.count(),
                'dias_configurados': dias,
                'fecha_limite': fecha_limite,
                'lotes': serializer.data
            })
            
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            return Response({
                'error': 'Error al obtener lotes por vencer',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path='por-caducar')
    def por_caducar(self, request):
        """Alias compatible para el frontend: lotes proximos a vencer."""
        try:
            from datetime import date, timedelta

            dias = int(request.query_params.get('dias', 90))
            hoy = date.today()
            fecha_limite = hoy + timedelta(days=dias)
            lotes = Lote.objects.select_related('producto').filter(
                cantidad_actual__gt=0,
                fecha_caducidad__gt=hoy,
                fecha_caducidad__lte=fecha_limite
            ).order_by('fecha_caducidad')

            page = self.paginate_queryset(lotes)
            if page is not None:
                serializer = self.get_serializer(page, many=True)
                return self.get_paginated_response(serializer.data)

            serializer = self.get_serializer(lotes, many=True)
            return Response(serializer.data)
        except Exception as exc:
            # traceback removido por seguridad (ISS-008)
            return Response({'error': 'Error al obtener lotes por caducar', 'mensaje': str(exc)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'])
    def vencidos(self, request):
        """Lotes con caducidad vencida y stock disponible."""
        try:
            from datetime import date

            hoy = date.today()
            lotes = Lote.objects.select_related('producto').filter(
                cantidad_actual__gt=0,
                fecha_caducidad__lt=hoy
            ).order_by('fecha_caducidad')

            page = self.paginate_queryset(lotes)
            if page is not None:
                serializer = self.get_serializer(page, many=True)
                return self.get_paginated_response(serializer.data)

            serializer = self.get_serializer(lotes, many=True)
            return Response(serializer.data)
        except Exception as exc:
            # traceback removido por seguridad (ISS-008)
            return Response({'error': 'Error al obtener lotes vencidos', 'mensaje': str(exc)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=True, methods=['get'])
    def historial(self, request, pk=None):
        """Obtiene el historial de movimientos de un lote"""
        try:
            lote = self.get_object()
            
            movimientos = Movimiento.objects.filter(lote=lote).select_related(
                'lote__producto'
            ).order_by('-fecha')
            
            from django.db.models import Sum
            
            total_entradas = movimientos.filter(tipo='entrada').aggregate(
                total=Sum('cantidad')
            )['total'] or 0
            
            total_salidas = movimientos.filter(tipo='salida').aggregate(
                total=Sum('cantidad')
            )['total'] or 0
            
            movimientos_data = []
            for mov in movimientos:
                movimientos_data.append({
                    'id': mov.id,
                    'tipo': mov.tipo,
                    'cantidad': mov.cantidad,
                    'fecha': mov.fecha,
                    'observaciones': mov.observaciones or ''
                })
            
            return Response({
                'lote': {
                    'id': lote.id,
                    'numero_lote': lote.numero_lote,
                    'producto': lote.producto.clave,
                    'cantidad_actual': lote.cantidad_actual
                },
                'estadisticas': {
                    'total_entradas': total_entradas,
                    'total_salidas': total_salidas,
                    'diferencia': total_entradas - total_salidas
                },
                'movimientos': movimientos_data,
                'total_movimientos': len(movimientos_data)
            })
            
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            return Response({
                'error': 'Error al obtener historial',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['post'])
    def ajustar_stock(self, request, pk=None):
        """
        Ajusta stock del lote y crea movimiento asociado (entrada/salida/ajuste).
        """
        lote = self.get_object()
        tipo = request.data.get('tipo', 'ajuste')
        cantidad = request.data.get('cantidad')
        observaciones = request.data.get('observaciones', '')

        try:
            movimiento, lote_actualizado = registrar_movimiento_stock(
                lote=lote,
                tipo=tipo,
                cantidad=cantidad,
                usuario=request.user if request.user.is_authenticated else None,
                centro=None,
                requisicion=None,
                observaciones=observaciones
            )
            return Response({
                'mensaje': 'Stock ajustado correctamente',
                'lote': self.get_serializer(lote_actualizado).data,
                'movimiento_id': movimiento.id
            })
        except serializers.ValidationError as exc:
            return Response({'error': 'Error de validacion', 'detalles': exc.detail}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as exc:
            # traceback removido por seguridad (ISS-008)
            return Response({'error': 'Error al ajustar stock', 'mensaje': str(exc)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['get'], url_path='lotes-derivados')
    def lotes_derivados(self, request, pk=None):
        """
        Obtiene los lotes derivados de un lote de farmacia (vinculados a centros).
        
        Solo aplica para lotes de farmacia central (centro=NULL).
        Muestra todos los centros que tienen stock de este lote.
        """
        try:
            lote = self.get_object()
            
            # Verificar que es un lote de farmacia
            if lote.centro is not None:
                return Response({
                    'error': 'Solo los lotes de farmacia central tienen lotes derivados',
                    'lote_id': lote.id,
                    'centro': lote.centro.nombre if lote.centro else None
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Nota: En Supabase no hay lote_origen, se simplifica
            # Los lotes derivados se manejan diferente
            derivados = Lote.objects.none()
            
            # Calcular totales
            from django.db.models import Sum
            total_derivados = 0
            stock_total_centros = 0
            
            derivados_data = []
            # Código original removido - lote_origen no existe en Supabase
            
            return Response({
                'lote_farmacia': {
                    'id': lote.id,
                    'numero_lote': lote.numero_lote,
                    'producto_clave': lote.producto.clave,
                    'producto_nombre': lote.producto.nombre,
                    'cantidad_actual': lote.cantidad_actual,
                    'fecha_caducidad': lote.fecha_caducidad
                },
                'resumen': {
                    'total_centros_con_stock': total_derivados,
                    'stock_total_en_centros': stock_total_centros,
                    'stock_farmacia': lote.cantidad_actual,
                    'stock_total_sistema': lote.cantidad_actual + stock_total_centros
                },
                'lotes_derivados': derivados_data
            })
            
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            return Response({
                'error': 'Error al obtener lotes derivados',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['get'], url_path='trazabilidad')
    def trazabilidad_lote(self, request, pk=None):
        """
        Obtiene la trazabilidad completa de un lote:
        - Si es lote de farmacia: muestra derivados en centros
        - Si es lote de centro: muestra origen en farmacia
        """
        try:
            lote = self.get_object()
            
            result = {
                'lote': {
                    'id': lote.id,
                    'numero_lote': lote.numero_lote,
                    'producto_clave': lote.producto.clave,
                    'producto_descripcion': lote.producto.descripcion,
                    'cantidad_actual': lote.cantidad_actual,
                    'fecha_caducidad': lote.fecha_caducidad,
                    'es_lote_farmacia': lote.centro is None,
                    'ubicacion': lote.centro.nombre if lote.centro else 'Almacén Central'
                },
                'origen': None,
                'derivados': []
            }
            
            # En Supabase no hay lote_origen - trazabilidad simplificada
            # Los lotes de cada centro son independientes
            
            # Movimientos relacionados
            movimientos = Movimiento.objects.filter(
                lote=lote
            ).select_related('requisicion', 'usuario', 'centro_origen', 'centro_destino').order_by('-fecha')[:20]
            
            result['movimientos'] = [{
                'id': m.id,
                'tipo': m.tipo,
                'cantidad': m.cantidad,
                'fecha': m.fecha,
                'requisicion_folio': m.requisicion.folio if m.requisicion else None,
                'observaciones': m.observaciones
            } for m in movimientos]
            
            return Response(result)
            
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            return Response({
                'error': 'Error al obtener trazabilidad',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    # =========================================================================
    # ACCIONES DE DOCUMENTOS (facturas, contratos, remisiones)
    # =========================================================================
    
    @action(detail=True, methods=['get'], url_path='documentos')
    def listar_documentos(self, request, pk=None):
        """
        Lista todos los documentos asociados a un lote.
        """
        try:
            lote = self.get_object()
            documentos = LoteDocumento.objects.filter(lote=lote).order_by('-created_at')
            serializer = LoteDocumentoSerializer(documentos, many=True)
            return Response({
                'lote_id': lote.id,
                'numero_lote': lote.numero_lote,
                'total_documentos': documentos.count(),
                'documentos': serializer.data
            })
        except Exception as e:
            return Response({
                'error': 'Error al obtener documentos',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['post'], url_path='subir-documento')
    def subir_documento(self, request, pk=None):
        """
        Sube un documento (PDF) asociado al lote.
        
        Campos requeridos:
        - documento: archivo PDF (multipart)
        - tipo_documento: factura/contrato/remision/otro
        
        Campos opcionales:
        - numero_documento: número del documento
        - fecha_documento: fecha del documento (YYYY-MM-DD)
        - notas: notas adicionales
        """
        try:
            lote = self.get_object()
            
            # Validar permisos de escritura
            user = request.user
            if not is_farmacia_or_admin(user):
                user_centro = get_user_centro(user)
                lote_centro = lote.centro
                if lote_centro is None or (user_centro and lote_centro.pk != user_centro.pk):
                    return Response(
                        {'error': 'No tiene permisos para subir documentos a este lote'},
                        status=status.HTTP_403_FORBIDDEN
                    )
            
            # ISS-005 FIX (audit7): Validar archivo PDF con función centralizada
            archivo = request.FILES.get('documento')
            if not archivo:
                return Response(
                    {'error': 'Debe proporcionar un archivo'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Usar validador centralizado que verifica extensión, tamaño Y magic bytes
            es_valido, error_msg = validar_archivo_pdf(archivo)
            if not es_valido:
                return Response(
                    {'error': error_msg},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Validar tipo de documento
            tipo_documento = request.data.get('tipo_documento', 'otro')
            tipos_validos = ['factura', 'contrato', 'remision', 'otro']
            if tipo_documento not in tipos_validos:
                return Response(
                    {'error': f'Tipo de documento inválido. Valores permitidos: {tipos_validos}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Generar path único para el archivo
            import uuid
            timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
            unique_name = f"{tipo_documento}_{timestamp}_{uuid.uuid4().hex[:8]}.pdf"
            archivo_path = f"lotes/documentos/{lote.id}/{unique_name}"
            
            # ISS-001 FIX: Subir archivo al almacenamiento ANTES de crear registro
            from inventario.services.storage_service import get_storage_service, StorageError
            
            storage = get_storage_service()
            upload_result = storage.upload_file(
                file_content=archivo,
                file_path=archivo_path,
                content_type='application/pdf',
                metadata={
                    'lote_id': lote.id,
                    'tipo_documento': tipo_documento,
                    'uploaded_by': user.username if user.is_authenticated else 'anonymous'
                }
            )
            
            # ISS-001 FIX: Si falla la subida, NO crear registro (rollback)
            if not upload_result.get('success'):
                logger.error(
                    f"ISS-001: Fallo subida documento lote {lote.id}: {upload_result.get('error')}"
                )
                return Response({
                    'error': 'Error al guardar archivo en almacenamiento',
                    'detalle': upload_result.get('error', 'Error desconocido')
                }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
            # Parsear fecha si viene
            fecha_documento = None
            if request.data.get('fecha_documento'):
                try:
                    fecha_documento = datetime.strptime(
                        request.data.get('fecha_documento'), '%Y-%m-%d'
                    ).date()
                except ValueError:
                    pass
            
            # ISS-001 FIX: Solo crear registro si la subida fue exitosa
            try:
                documento = LoteDocumento.objects.create(
                    lote=lote,
                    tipo_documento=tipo_documento,
                    numero_documento=request.data.get('numero_documento', ''),
                    archivo=archivo_path,
                    nombre_archivo=nombre_archivo,
                    fecha_documento=fecha_documento,
                    notas=request.data.get('notas', ''),
                    created_by=user if user.is_authenticated else None
                )
            except Exception as db_error:
                # ISS-001 FIX: Si falla crear registro, eliminar archivo subido (rollback)
                logger.error(f"ISS-001: Error BD, revirtiendo subida: {db_error}")
                storage.delete_file(archivo_path)
                raise
            
            logger.info(
                f"ISS-001: Documento subido exitosamente - Lote: {lote.id}, "
                f"Path: {archivo_path}, Storage: {upload_result.get('storage')}"
            )
            
            serializer = LoteDocumentoSerializer(documento)
            return Response({
                'mensaje': 'Documento subido correctamente',
                'documento': serializer.data,
                'storage_info': {
                    'path': archivo_path,
                    'url': upload_result.get('url'),
                    'storage_type': upload_result.get('storage')
                }
            }, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            return Response({
                'error': 'Error al subir documento',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['delete'], url_path='eliminar-documento/(?P<doc_id>[0-9]+)')
    def eliminar_documento(self, request, pk=None, doc_id=None):
        """
        Elimina un documento específico del lote.
        """
        try:
            lote = self.get_object()
            
            # Validar permisos de escritura
            user = request.user
            if not is_farmacia_or_admin(user):
                user_centro = get_user_centro(user)
                lote_centro = lote.centro
                if lote_centro is None or (user_centro and lote_centro.pk != user_centro.pk):
                    return Response(
                        {'error': 'No tiene permisos para eliminar documentos de este lote'},
                        status=status.HTTP_403_FORBIDDEN
                    )
            
            # Buscar documento
            try:
                documento = LoteDocumento.objects.get(id=doc_id, lote=lote)
            except LoteDocumento.DoesNotExist:
                return Response(
                    {'error': 'Documento no encontrado'},
                    status=status.HTTP_404_NOT_FOUND
                )
            
            # ISS-002 FIX: Eliminar archivo del almacenamiento ANTES de borrar registro
            from inventario.services.storage_service import get_storage_service
            
            archivo_path = documento.archivo
            nombre = documento.nombre_archivo
            
            storage = get_storage_service()
            delete_result = storage.delete_file(archivo_path)
            
            # ISS-002 FIX: Registrar resultado de eliminación de storage
            if not delete_result.get('success'):
                logger.warning(
                    f"ISS-002: No se pudo eliminar archivo '{archivo_path}' del storage: "
                    f"{delete_result.get('error')}. Se eliminará el registro de BD igualmente."
                )
            else:
                logger.info(
                    f"ISS-002: Archivo eliminado de storage: {archivo_path}"
                )
            
            # Eliminar registro de BD
            documento.delete()
            
            logger.info(
                f"ISS-002: Documento eliminado - Lote: {lote.id}, "
                f"Archivo: {nombre}, Path: {archivo_path}"
            )
            
            return Response({
                'mensaje': 'Documento eliminado correctamente',
                'documento_eliminado': nombre,
                'storage_cleanup': delete_result.get('success', False)
            })
            
        except Exception as e:
            return Response({
                'error': 'Error al eliminar documento',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class MovimientoViewSet(
    mixins.ListModelMixin,
    mixins.RetrieveModelMixin,
    mixins.CreateModelMixin,
    viewsets.GenericViewSet
):
    """
    ViewSet para gestionar movimientos de inventario.
    
    PERMISOS:
    - Admin/Farmacia: acceso completo a todos los movimientos
    - Centro (administrador_centro, director_centro): puede VER y CREAR movimientos
    - Medico: NO puede crear movimientos (ISS-MEDICO FIX)
    - Vista: solo lectura
    
    FILTROS (alineados con exportacin):
    - tipo: entrada/salida/ajuste
    - centro: ID del centro
    - producto: ID del producto
    - lote: ID del lote
    - fecha_inicio: YYYY-MM-DD
    - fecha_fin: YYYY-MM-DD
    - search: bsqueda en observaciones, nmero de lote, producto
    
    Esto permite auditora completa de consumos en cada centro.
    """
    queryset = Movimiento.objects.select_related('lote__producto', 'centro_origen', 'centro_destino', 'usuario').all()
    serializer_class = MovimientoSerializer
    # ISS-MEDICO FIX: Usar permiso que excluye médico de operaciones de escritura
    permission_classes = [IsCentroCanManageInventory]
    pagination_class = CustomPagination
    http_method_names = ['get', 'post', 'head', 'options']

    def get_queryset(self):
        """
        Filtra movimientos segun parametros.
        
        Parametros (alineados con exportacin):
        - tipo: entrada/salida/ajuste
        - centro: ID del centro (solo admin/farmacia/vista)
        - producto: ID del producto
        - lote: ID del lote
        - fecha_inicio: fecha mnima (YYYY-MM-DD)
        - fecha_fin: fecha mxima (YYYY-MM-DD)
        - search: bsqueda en observaciones, lote, producto
        
        Seguridad: Usuarios de centro solo ven movimientos de su centro.
        Admin/farmacia/vista ven todo por defecto, pueden filtrar con ?centro=.
        """
        queryset = Movimiento.objects.select_related('lote__producto', 'centro_origen', 'centro_destino', 'usuario')
        
        # Obtener parámetros de filtro
        tipo = self.request.query_params.get('tipo')
        tipo_lower = (tipo or '').lower()
        
        # ISS-019 FIX: Filtrar por centro según rol usando has_global_read_access
        # que incluye admin, farmacia Y vista para lectura global
        user = self.request.user
        if not has_global_read_access(user):
            # Usuario de centro: forzado a su centro
            # Incluir movimientos donde el centro es origen, destino O el lote pertenece al centro
            user_centro = get_user_centro(user)
            if user_centro:
                queryset = queryset.filter(
                    Q(lote__centro=user_centro) | 
                    Q(centro_origen=user_centro) | 
                    Q(centro_destino=user_centro)
                )
            else:
                return Movimiento.objects.none()
        else:
            # Admin/farmacia/vista: pueden filtrar por centro especifico
            centro_param = self.request.query_params.get('centro')
            if centro_param and centro_param.lower() != 'todos':
                # ISS-FIX: Filtrar de forma ESTRICTA por centro según tipo de movimiento
                # Evitar usar lote__centro para que no se cuelen movimientos de otros contextos
                if tipo_lower == 'salida':
                    # Para SALIDAS: solo donde el centro es ORIGEN (salidas DESDE ese centro)
                    queryset = queryset.filter(centro_origen_id=centro_param)
                elif tipo_lower == 'entrada':
                    # Para ENTRADAS: solo donde el centro es DESTINO (entradas HACIA ese centro)
                    queryset = queryset.filter(centro_destino_id=centro_param)
                else:
                    # Sin tipo: movimientos donde el centro es origen O destino
                    queryset = queryset.filter(
                        Q(centro_origen_id=centro_param) | 
                        Q(centro_destino_id=centro_param)
                    )
        
        # Filtro por tipo (ya lo tenemos arriba, aplicar si existe)
        if tipo:
            queryset = queryset.filter(tipo=tipo.lower())
        
        # Filtro por producto
        producto = self.request.query_params.get('producto')
        if producto:
            queryset = queryset.filter(lote__producto_id=producto)
        
        # Filtro por lote (acepta ID numérico o número de lote como texto)
        lote = self.request.query_params.get('lote')
        if lote:
            if lote.isdigit():
                # Si es un número, buscar por ID
                queryset = queryset.filter(lote_id=lote)
            else:
                # Si es texto, buscar por número de lote (coincidencia parcial)
                queryset = queryset.filter(lote__numero_lote__icontains=lote)
        
        # Filtro por subtipo de salida (receta, consumo_interno, merma, etc.)
        subtipo_salida = self.request.query_params.get('subtipo_salida')
        if subtipo_salida:
            queryset = queryset.filter(subtipo_salida__iexact=subtipo_salida)
        
        # Filtro por rango de fechas
        fecha_inicio = self.request.query_params.get('fecha_inicio')
        if fecha_inicio:
            queryset = queryset.filter(fecha__date__gte=fecha_inicio)
        
        fecha_fin = self.request.query_params.get('fecha_fin')
        if fecha_fin:
            queryset = queryset.filter(fecha__date__lte=fecha_fin)
        
        # Busqueda en motivo, lote y producto
        search = self.request.query_params.get('search')
        if search and search.strip():
            search_term = search.strip()
            queryset = queryset.filter(
                Q(motivo__icontains=search_term) |
                Q(lote__numero_lote__icontains=search_term) |
                Q(lote__producto__clave__icontains=search_term) |
                Q(lote__producto__descripcion__icontains=search_term) |
                Q(numero_expediente__icontains=search_term)
            )
        
        return queryset.order_by('-fecha')

    def perform_create(self, serializer):
        """
        Crea un movimiento validando permisos por centro.
        
        SEGURIDAD:
        - Admin/farmacia: pueden crear cualquier movimiento en cualquier lote
        - Usuario de centro (administrador/director): pueden crear movimientos en lotes de su centro
          y solo ciertos tipos: 'salida' (consumo), 'ajuste' (inventario físico)
        - Médico: Solo puede crear movimientos de SALIDA (para dispensación a pacientes)
        - Usuario de centro NO puede crear 'entrada' (solo vía surtido de requisición)
        """
        user = self.request.user
        lote = serializer.validated_data.get('lote')
        tipo = serializer.validated_data.get('tipo', '').lower()
        motivo = serializer.validated_data.get('motivo', '').strip()
        
        # ISS-MEDICO FIX v2: Médicos SOLO pueden crear SALIDAS (para dispensación a pacientes)
        if RoleHelper.is_medico(user):
            if tipo != 'salida':
                raise serializers.ValidationError({
                    'tipo': 'Los médicos solo pueden registrar movimientos de SALIDA para dispensación a pacientes.'
                })
            # Observaciones/motivo son OBLIGATORIAS para médicos
            if not motivo:
                raise serializers.ValidationError({
                    'motivo': 'Las observaciones son obligatorias. Indique el motivo de la salida (ej: dispensación a paciente, nombre del paciente, etc.).'
                })
        
        # Validar que usuario de centro solo opere con sus lotes
        if not is_farmacia_or_admin(user):
            user_centro = get_user_centro(user)
            
            # Validar que el lote pertenece al centro del usuario
            if lote and lote.centro != user_centro:
                raise serializers.ValidationError({
                    'lote': 'Solo puedes registrar movimientos en lotes de tu centro'
                })
            
            # Validar tipos de movimiento permitidos para centros
            # Centros pueden: salida (consumo), ajuste (inventario fsico)
            # Centros NO pueden: entrada (solo va surtido automtico)
            tipos_permitidos_centro = ['salida', 'ajuste']
            if tipo not in tipos_permitidos_centro:
                raise serializers.ValidationError({
                    'tipo': f'Los centros solo pueden registrar: {", ".join(tipos_permitidos_centro)}. Las entradas se generan automticamente al surtir requisiciones.'
                })
        
        # MEJORA FLUJO 5: Extraer campos de trazabilidad
        subtipo_salida = serializer.validated_data.get('subtipo_salida')
        numero_expediente = serializer.validated_data.get('numero_expediente')
        
        # ISS-FIX-500: Convertir centro_id a objeto Centro si se pasa un ID
        centro_destino_raw = serializer.validated_data.get('centro')
        centro_destino = None
        if centro_destino_raw:
            if isinstance(centro_destino_raw, Centro):
                centro_destino = centro_destino_raw
            else:
                try:
                    centro_destino = Centro.objects.get(pk=int(centro_destino_raw))
                except (Centro.DoesNotExist, ValueError, TypeError):
                    raise serializers.ValidationError({
                        'centro': f'Centro con ID {centro_destino_raw} no encontrado'
                    })
        
        # ISS-FIX: Para transferencias desde Almacén Central a Centro,
        # el lote es del Almacén Central (centro=None) pero el destino es un Centro específico.
        # Debemos permitir esto para admin/farmacia usando skip_centro_check=True
        es_transferencia_almacen = is_farmacia_or_admin(user) and centro_destino and lote and lote.centro is None
        
        movimiento, _ = registrar_movimiento_stock(
            lote=lote,
            tipo=serializer.validated_data.get('tipo'),
            cantidad=serializer.validated_data.get('cantidad'),
            usuario=user,
            centro=centro_destino or (lote.centro if lote else None),
            requisicion=serializer.validated_data.get('requisicion'),
            # FIX: El serializer mapea 'observaciones' del frontend a 'motivo' via to_internal_value
            observaciones=serializer.validated_data.get('motivo', ''),
            subtipo_salida=subtipo_salida,
            numero_expediente=numero_expediente,
            # ISS-FIX: Saltear validación de centro para transferencias del Almacén Central
            skip_centro_check=es_transferencia_almacen
        )
        # Dejar instancia lista para serializer.data
        serializer.instance = movimiento

    @action(detail=False, methods=['get'], url_path='trazabilidad-pdf')
    def trazabilidad_pdf(self, request):
        """
        Genera PDF de trazabilidad de un producto.
        Parmetros: ?producto_clave=XXX
        
        SEGURIDAD: Filtra por centro del usuario si no es admin/farmacia.
        """
        from core.utils.pdf_reports import generar_reporte_trazabilidad
        
        # SEGURIDAD: Verificar permisos y determinar filtro de centro
        user = request.user
        filtrar_por_centro = not is_farmacia_or_admin(user)
        user_centro = get_user_centro(user) if filtrar_por_centro else None
        
        clave = request.query_params.get('producto_clave')
        if not clave:
            return Response({'error': 'Se requiere producto_clave'}, status=status.HTTP_400_BAD_REQUEST)
        
        producto = Producto.objects.filter(
            Q(clave__iexact=clave) | Q(descripcion__iexact=clave)
        ).first()
        if not producto:
            return Response({'error': 'Producto no encontrado'}, status=status.HTTP_404_NOT_FOUND)
        
        try:
            # Obtener movimientos del producto
            movimientos = Movimiento.objects.filter(
                lote__producto=producto
            ).select_related('lote', 'centro_origen', 'centro_destino', 'usuario')
            
            # ISS-FIX: Aplicar filtro de centro de forma ESTRICTA
            # Solo donde el centro es origen O destino (no por lote__centro)
            if filtrar_por_centro and user_centro:
                movimientos = movimientos.filter(
                    Q(centro_origen=user_centro) | Q(centro_destino=user_centro)
                )
            
            movimientos = movimientos.order_by('-fecha')[:100]
            
            trazabilidad_data = []
            for mov in movimientos:
                # ISS-FIX: Lógica clara para centro según tipo de movimiento
                tipo_upper = mov.tipo.upper()
                if tipo_upper == 'SALIDA':
                    centro_display = mov.centro_destino.nombre if mov.centro_destino else 'Farmacia Central'
                elif tipo_upper == 'ENTRADA':
                    centro_display = mov.centro_origen.nombre if mov.centro_origen else 'Farmacia Central'
                else:
                    centro_display = mov.centro_destino.nombre if mov.centro_destino else (mov.centro_origen.nombre if mov.centro_origen else 'Farmacia Central')
                
                # ISS-FIX: Mostrar username si no hay nombre completo
                if mov.usuario:
                    usuario_display = mov.usuario.get_full_name()
                    if not usuario_display or usuario_display.strip() == '':
                        usuario_display = mov.usuario.username
                else:
                    usuario_display = 'Sistema'
                
                trazabilidad_data.append({
                    'fecha': mov.fecha.strftime('%d/%m/%Y %H:%M') if mov.fecha else 'N/A',
                    'tipo': tipo_upper,
                    'lote': mov.lote.numero_lote if mov.lote else 'N/A',
                    'cantidad': mov.cantidad,
                    'centro': centro_display,
                    'usuario': usuario_display,
                    'observaciones': mov.motivo or ''
                })
            
            producto_info = {
                'clave': producto.clave,
                'descripcion': producto.nombre,  # Usar nombre como descripción principal
                'unidad_medida': producto.unidad_medida,
                'stock_actual': producto.get_stock_actual() if hasattr(producto, 'get_stock_actual') else 0,
                'stock_minimo': producto.stock_minimo,
                'precio_unitario': 0,  # precio_unitario está en Lote, no en Producto
            }
            
            pdf_buffer = generar_reporte_trazabilidad(trazabilidad_data, producto_info=producto_info)
            
            response = HttpResponse(
                pdf_buffer.getvalue(),
                content_type='application/pdf'
            )
            response['Content-Disposition'] = f'attachment; filename="Trazabilidad_{clave}_{timezone.now().strftime("%Y%m%d")}.pdf"'
            
            return response
            
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            return Response({
                'error': 'Error al generar PDF de trazabilidad',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path='trazabilidad-lote-pdf')
    def trazabilidad_lote_pdf(self, request):
        """
        Genera PDF de trazabilidad de un lote especfico.
        Parmetros: ?numero_lote=XXX
        
        SEGURIDAD: Solo admin/farmacia pueden acceder.
        """
        from core.utils.pdf_reports import generar_reporte_trazabilidad
        
        # SEGURIDAD: Solo admin/farmacia pueden exportar trazabilidad de lotes
        if not is_farmacia_or_admin(request.user):
            return Response({'error': 'Solo administradores y farmacia pueden exportar trazabilidad de lotes'}, status=status.HTTP_403_FORBIDDEN)
        
        numero_lote = request.query_params.get('numero_lote')
        if not numero_lote:
            return Response({'error': 'Se requiere numero_lote'}, status=status.HTTP_400_BAD_REQUEST)
        
        lote = Lote.objects.filter(numero_lote__iexact=numero_lote).select_related('producto', 'centro').first()
        if not lote:
            return Response({'error': 'Lote no encontrado'}, status=status.HTTP_404_NOT_FOUND)
        
        try:
            # Obtener movimientos del lote
            movimientos = Movimiento.objects.filter(
                lote=lote
            ).select_related('lote', 'centro_origen', 'centro_destino', 'usuario').order_by('-fecha')[:100]
            
            trazabilidad_data = []
            for mov in movimientos:
                # ISS-FIX: Lógica clara para centro según tipo de movimiento
                tipo_upper = mov.tipo.upper()
                if tipo_upper == 'SALIDA':
                    centro_display = mov.centro_destino.nombre if mov.centro_destino else 'Farmacia Central'
                elif tipo_upper == 'ENTRADA':
                    centro_display = mov.centro_origen.nombre if mov.centro_origen else 'Farmacia Central'
                else:
                    centro_display = mov.centro_destino.nombre if mov.centro_destino else (mov.centro_origen.nombre if mov.centro_origen else 'Farmacia Central')
                
                # ISS-FIX: Mostrar username si no hay nombre completo
                if mov.usuario:
                    usuario_display = mov.usuario.get_full_name()
                    if not usuario_display or usuario_display.strip() == '':
                        usuario_display = mov.usuario.username
                else:
                    usuario_display = 'Sistema'
                
                trazabilidad_data.append({
                    'fecha': mov.fecha.strftime('%d/%m/%Y %H:%M') if mov.fecha else 'N/A',
                    'tipo': tipo_upper,
                    'lote': mov.lote.numero_lote if mov.lote else 'N/A',
                    'cantidad': mov.cantidad,
                    'centro': centro_display,
                    'usuario': usuario_display,
                    'observaciones': mov.motivo or ''
                })
            
            # ISS-FIX: Usar nombre como fallback para descripcion, incluir numero_contrato
            descripcion_producto = 'N/A'
            if lote.producto:
                descripcion_producto = lote.producto.nombre or lote.producto.descripcion or 'N/A'
            
            producto_info = {
                'clave': lote.producto.clave if lote.producto else 'N/A',
                'descripcion': descripcion_producto,
                'unidad_medida': lote.producto.unidad_medida if lote.producto else 'N/A',
                'stock_actual': lote.cantidad_actual,
                'stock_minimo': lote.producto.stock_minimo if lote.producto else 0,
                'numero_lote': lote.numero_lote,
                'fecha_caducidad': lote.fecha_caducidad.strftime('%d/%m/%Y') if lote.fecha_caducidad else 'N/A',
                'proveedor': lote.marca or 'No especificado',
                'numero_contrato': lote.numero_contrato if lote.numero_contrato else 'N/A',
                'precio_unitario': float(lote.precio_unitario) if lote.precio_unitario else 0,
            }
            
            pdf_buffer = generar_reporte_trazabilidad(trazabilidad_data, producto_info=producto_info)
            
            response = HttpResponse(
                pdf_buffer.getvalue(),
                content_type='application/pdf'
            )
            response['Content-Disposition'] = f'attachment; filename="Trazabilidad_Lote_{numero_lote}_{timezone.now().strftime("%Y%m%d")}.pdf"'
            
            return response
            
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            return Response({
                'error': 'Error al generar PDF de trazabilidad del lote',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path='exportar-pdf')
    def exportar_pdf(self, request):
        """
        Genera PDF de movimientos con filtros opcionales.
        Filtros soportados: tipo, fecha_inicio, fecha_fin, producto, centro, lote, subtipo_salida, search
        """
        from core.utils.pdf_reports import generar_reporte_movimientos
        
        try:
            # Aplicar filtros (get_queryset ya aplica filtros base, aquí se duplican por consistencia explícita)
            queryset = self.get_queryset()
            
            tipo = request.query_params.get('tipo')
            if tipo:
                queryset = queryset.filter(tipo=tipo.lower())
            
            fecha_inicio = request.query_params.get('fecha_inicio')
            if fecha_inicio:
                queryset = queryset.filter(fecha__gte=fecha_inicio)
            
            fecha_fin = request.query_params.get('fecha_fin')
            if fecha_fin:
                queryset = queryset.filter(fecha__lte=fecha_fin)
            
            # FIX: Agregar filtros faltantes para consistencia total
            producto = request.query_params.get('producto')
            if producto:
                queryset = queryset.filter(lote__producto_id=producto)
            
            centro = request.query_params.get('centro')
            if centro:
                queryset = queryset.filter(Q(centro_origen_id=centro) | Q(centro_destino_id=centro) | Q(lote__centro_id=centro))
            
            lote = request.query_params.get('lote')
            if lote:
                if lote.isdigit():
                    queryset = queryset.filter(lote_id=lote)
                else:
                    queryset = queryset.filter(lote__numero_lote__icontains=lote)
            
            subtipo_salida = request.query_params.get('subtipo_salida')
            if subtipo_salida:
                queryset = queryset.filter(subtipo_salida__iexact=subtipo_salida)
            
            search = request.query_params.get('search')
            if search:
                queryset = queryset.filter(
                    Q(lote__numero_lote__icontains=search) |
                    Q(lote__producto__nombre__icontains=search) |
                    Q(lote__producto__descripcion__icontains=search) |
                    Q(motivo__icontains=search) |
                    Q(numero_expediente__icontains=search)
                )
            
            movimientos = queryset[:200]  # Limitar para PDF
            
            # Agrupar movimientos por referencia/transacción para PDF
            transacciones = {}
            total_entradas = 0
            total_salidas = 0
            
            for mov in movimientos:
                amount = abs(mov.cantidad) if mov.tipo == 'salida' else mov.cantidad
                ref = mov.referencia or f"MOV-{mov.id}"
                
                if ref not in transacciones:
                    transacciones[ref] = {
                        'referencia': ref,
                        'fecha': mov.fecha.strftime('%d/%m/%Y %H:%M') if mov.fecha else 'N/A',
                        'tipo': mov.tipo.upper(),
                        'centro_origen': mov.centro_origen.nombre if mov.centro_origen else 'Almacén Central',
                        'centro_destino': mov.centro_destino.nombre if mov.centro_destino else 'Almacén Central',
                        'total_productos': 0,
                        'total_cantidad': 0,
                        'detalles': []
                    }
                
                transacciones[ref]['detalles'].append({
                    'producto': f"{mov.lote.producto.clave if mov.lote and mov.lote.producto else 'N/A'} - {getattr(mov.lote.producto, 'descripcion', '')[:40] if mov.lote and mov.lote.producto else ''}",
                    'lote': mov.lote.numero_lote if mov.lote else 'N/A',
                    'cantidad': amount
                })
                transacciones[ref]['total_productos'] += 1
                transacciones[ref]['total_cantidad'] += amount
                
                if mov.tipo == 'entrada':
                    total_entradas += amount
                else:
                    total_salidas += amount
            
            movimientos_data = list(transacciones.values())
            resumen_data = {
                'total_transacciones': len(movimientos_data),
                'total_movimientos': sum(t['total_productos'] for t in movimientos_data),
                'total_entradas': total_entradas,
                'total_salidas': total_salidas,
                'diferencia': total_entradas - total_salidas
            }
            
            pdf_buffer = generar_reporte_movimientos(movimientos_data, resumen=resumen_data)
            
            response = HttpResponse(
                pdf_buffer.getvalue(),
                content_type='application/pdf'
            )
            response['Content-Disposition'] = f'attachment; filename="Movimientos_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
            
            return response
            
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            return Response({
                'error': 'Error al generar PDF de movimientos',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path='exportar-excel')
    def exportar_excel(self, request):
        """
        Genera Excel de movimientos con filtros opcionales.
        Filtros soportados: tipo, fecha_inicio, fecha_fin, producto, centro, lote, subtipo_salida, search
        """
        try:
            # Aplicar filtros (get_queryset ya aplica filtros base, aquí se duplican por consistencia explícita)
            queryset = self.get_queryset()
            
            tipo = request.query_params.get('tipo')
            if tipo:
                queryset = queryset.filter(tipo=tipo.lower())
            
            fecha_inicio = request.query_params.get('fecha_inicio')
            if fecha_inicio:
                queryset = queryset.filter(fecha__gte=fecha_inicio)
            
            fecha_fin = request.query_params.get('fecha_fin')
            if fecha_fin:
                queryset = queryset.filter(fecha__lte=fecha_fin)
            
            producto = request.query_params.get('producto')
            if producto:
                queryset = queryset.filter(lote__producto_id=producto)
            
            centro = request.query_params.get('centro')
            if centro:
                queryset = queryset.filter(Q(centro_origen_id=centro) | Q(centro_destino_id=centro) | Q(lote__centro_id=centro))
            
            # FIX: Agregar filtros faltantes para consistencia total
            lote = request.query_params.get('lote')
            if lote:
                if lote.isdigit():
                    queryset = queryset.filter(lote_id=lote)
                else:
                    queryset = queryset.filter(lote__numero_lote__icontains=lote)
            
            subtipo_salida = request.query_params.get('subtipo_salida')
            if subtipo_salida:
                queryset = queryset.filter(subtipo_salida__iexact=subtipo_salida)
            
            search = request.query_params.get('search')
            if search:
                queryset = queryset.filter(
                    Q(lote__numero_lote__icontains=search) |
                    Q(lote__producto__nombre__icontains=search) |
                    Q(lote__producto__descripcion__icontains=search) |
                    Q(motivo__icontains=search) |
                    Q(numero_expediente__icontains=search)
                )
            
            movimientos = queryset[:1000]  # Limitar para Excel
            
            # Crear libro de Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Movimientos'
            
            # Título - MEJORA FLUJO 5: Extender a 11 columnas (K)
            ws.merge_cells('A1:K1')
            ws['A1'] = 'REPORTE DE MOVIMIENTOS'
            ws['A1'].font = Font(bold=True, size=14, color='632842')
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # Fecha
            ws.merge_cells('A2:K2')
            ws['A2'] = f'Generado el {timezone.now().strftime("%d/%m/%Y %H:%M")}'
            ws['A2'].alignment = Alignment(horizontal='center')
            
            # Encabezados - MEJORA FLUJO 5: Incluir subtipo y expediente
            headers = ['#', 'Fecha', 'Tipo', 'Subtipo', 'Producto', 'Lote', 'Cantidad', 'Centro', 'Usuario', 'No. Expediente', 'Observaciones']
            ws.append([])
            ws.append(headers)
            
            header_fill = PatternFill(start_color='632842', end_color='632842', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            for cell in ws[4]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Datos - MEJORA FLUJO 5: Incluir campos de trazabilidad
            for idx, mov in enumerate(movimientos, 1):
                # ISS-FIX: Lógica clara para centro según tipo de movimiento
                tipo_upper = mov.tipo.upper()
                if tipo_upper == 'SALIDA':
                    centro_display = mov.centro_destino.nombre if mov.centro_destino else 'Farmacia Central'
                elif tipo_upper == 'ENTRADA':
                    centro_display = mov.centro_origen.nombre if mov.centro_origen else 'Farmacia Central'
                else:
                    centro_display = mov.centro_destino.nombre if mov.centro_destino else (mov.centro_origen.nombre if mov.centro_origen else 'Farmacia Central')
                
                # ISS-FIX: Mostrar username si no hay nombre completo
                if mov.usuario:
                    usuario_display = mov.usuario.get_full_name()
                    if not usuario_display or usuario_display.strip() == '':
                        usuario_display = mov.usuario.username
                else:
                    usuario_display = 'Sistema'
                
                ws.append([
                    idx,
                    mov.fecha.strftime('%d/%m/%Y %H:%M') if mov.fecha else 'N/A',
                    tipo_upper,
                    (mov.subtipo_salida or '').upper() if mov.tipo == 'salida' else '',
                    mov.lote.producto.descripcion if mov.lote and mov.lote.producto else 'N/A',
                    mov.lote.numero_lote if mov.lote else 'N/A',
                    mov.cantidad,
                    centro_display,
                    usuario_display,
                    mov.numero_expediente or '',
                    (mov.motivo or ''),
                ])
            
            # Ajustar anchos - actualizado para 11 columnas - Columnas más anchas para textos completos
            column_widths = [8, 18, 12, 15, 60, 18, 10, 40, 20, 18, 60]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
            
            # Respuesta
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="Movimientos_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
            wb.save(response)
            return response
            
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            return Response({
                'error': 'Error al generar Excel de movimientos',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['get'], url_path='recibo-salida')
    def recibo_salida(self, request, pk=None):
        """
        Genera PDF de recibo de salida para un movimiento específico.
        
        Parámetros opcionales:
        - finalizado: si es 'true', muestra sello ENTREGADO en lugar de firmas
        
        SEGURIDAD: Usuarios pueden generar recibos de movimientos que les correspondan.
        """
        from core.utils.pdf_reports import generar_recibo_salida_movimiento
        
        try:
            movimiento = self.get_object()
            
            # Verificar que es un movimiento de salida
            if movimiento.tipo != 'salida':
                return Response(
                    {'error': 'Solo se pueden generar recibos para movimientos de salida'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Verificar permisos - admin/farmacia pueden ver todos, otros solo sus centros
            user = request.user
            if not is_farmacia_or_admin(user):
                user_centro = get_user_centro(user)
                if user_centro:
                    # Usuario de centro puede ver si es origen o destino
                    if movimiento.centro_origen != user_centro and movimiento.centro_destino != user_centro:
                        if movimiento.lote and movimiento.lote.centro != user_centro:
                            return Response(
                                {'error': 'No tienes permiso para ver este movimiento'},
                                status=status.HTTP_403_FORBIDDEN
                            )
            
            finalizado = request.query_params.get('finalizado', 'false').lower() == 'true'
            
            # Construir datos del movimiento
            movimiento_data = {
                'folio': movimiento.id,
                'fecha': movimiento.fecha.strftime('%Y-%m-%d %H:%M') if movimiento.fecha else 'N/A',
                'tipo': movimiento.tipo,
                'subtipo_salida': movimiento.subtipo_salida or 'transferencia',
                'centro_origen': {
                    'id': movimiento.centro_origen.id if movimiento.centro_origen else None,
                    'nombre': movimiento.centro_origen.nombre if movimiento.centro_origen else 'Almacén Central'
                },
                'centro_destino': {
                    'id': movimiento.centro_destino.id if movimiento.centro_destino else None,
                    'nombre': movimiento.centro_destino.nombre if movimiento.centro_destino else ''
                },
                'cantidad': abs(movimiento.cantidad),  # ISS-FIX: Usar valor absoluto
                'observaciones': movimiento.motivo or '',
                'producto': movimiento.lote.producto.nombre if movimiento.lote and movimiento.lote.producto else 'N/A',
                'producto_clave': movimiento.lote.producto.clave if movimiento.lote and movimiento.lote.producto else 'N/A',
                'lote': movimiento.lote.numero_lote if movimiento.lote else 'N/A',
                'presentacion': movimiento.lote.producto.presentacion if movimiento.lote and movimiento.lote.producto else 'N/A',
                'usuario': movimiento.usuario.get_full_name() if movimiento.usuario else 'Sistema',
            }
            
            # Generar PDF usando la función específica para movimientos
            pdf_buffer = generar_recibo_salida_movimiento(
                movimiento_data,
                finalizado=finalizado
            )
            
            response = HttpResponse(
                pdf_buffer.getvalue(),
                content_type='application/pdf'
            )
            tipo_doc = 'Comprobante_Entrega' if finalizado else 'Recibo_Salida'
            response['Content-Disposition'] = f'attachment; filename="{tipo_doc}_{movimiento.id}_{timezone.now().strftime("%Y%m%d")}.pdf"'
            
            logger.info(f"Recibo de salida generado para movimiento {movimiento.id} por usuario {user.username}")
            return response
            
        except Movimiento.DoesNotExist:
            return Response(
                {'error': 'Movimiento no encontrado'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            logger.error(f"Error generando recibo de salida: {str(e)}")
            return Response({
                'error': 'Error al generar recibo de salida',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['post'], url_path='confirmar-entrega')
    def confirmar_entrega(self, request, pk=None):
        """
        Confirma la entrega física de un movimiento de salida individual.
        Marca el movimiento como confirmado agregando [CONFIRMADO] al motivo.
        
        Returns:
            - 200: Entrega confirmada exitosamente
            - 404: Movimiento no encontrado
            - 400: No es movimiento de salida o ya está confirmado
        """
        try:
            movimiento = self.get_object()
            
            # Verificar que es un movimiento de salida
            if movimiento.tipo != 'salida':
                return Response({
                    'error': True,
                    'message': 'Solo se pueden confirmar entregas de movimientos de salida'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Verificar permisos - admin/farmacia o usuario del centro destino
            user = request.user
            if not is_farmacia_or_admin(user):
                user_centro = get_user_centro(user)
                if user_centro and movimiento.centro_destino:
                    if movimiento.centro_destino.id != user_centro.id:
                        return Response({
                            'error': True,
                            'message': 'No tienes permiso para confirmar esta entrega'
                        }, status=status.HTTP_403_FORBIDDEN)
            
            # Verificar si ya está confirmado
            motivo_actual = movimiento.motivo or ''
            if '[CONFIRMADO]' in motivo_actual:
                return Response({
                    'error': True,
                    'message': 'Esta entrega ya fue confirmada anteriormente'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Marcar como confirmado
            movimiento.motivo = f'[CONFIRMADO] {motivo_actual}'.strip()
            movimiento.save(update_fields=['motivo'])
            
            logger.info(
                f'Entrega de movimiento {movimiento.id} confirmada por {request.user.username}'
            )
            
            return Response({
                'success': True,
                'message': 'Entrega confirmada exitosamente',
                'movimiento_id': movimiento.id
            })
            
        except Movimiento.DoesNotExist:
            return Response({
                'error': True,
                'message': 'Movimiento no encontrado'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f'Error confirmando entrega individual: {str(e)}')
            return Response({
                'error': True,
                'message': f'Error al confirmar entrega: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class RequisicionViewSet(CentroPermissionMixin, viewsets.ModelViewSet):
    """
    ISS-030: CRUD y flujo de requisiciones con control de acceso por centro.
    
    Estados en minúsculas:
    (borrador -> enviada -> autorizada/parcial -> surtida o rechazada/cancelada)
    
    ISS-011, ISS-021: El método surtir() usa RequisicionService para transacciones atómicas.
    ISS-014: Bloqueo optimista de lotes durante el surtido.
    ISS-030: Validación de acceso por centro en todas las operaciones.
    
    OPTIMIZACIÓN: Usa RequisicionListSerializer para listados (más ligero).
    """
    queryset = Requisicion.objects.select_related('centro_origen', 'centro_destino', 'solicitante', 'autorizador').prefetch_related('detalles__producto').all()
    serializer_class = RequisicionSerializer
    permission_classes = [IsCentroRole]
    pagination_class = CustomPagination
    
    def get_serializer_class(self):
        """OPTIMIZACIÓN: Usa serializer ligero para listados."""
        if self.action == 'list':
            return RequisicionListSerializer
        return RequisicionSerializer

    def _user_centro(self, user):
        return getattr(user, 'centro', None) or getattr(getattr(user, 'profile', None), 'centro', None)

    def _validar_stock_items(self, items, centro=None, validar_farmacia_central=True, modo='informativo'):
        """
        ISS-001, ISS-004, ISS-005 FIX: Valida stock disponible con lógica clara por ubicación.
        
        ISS-004 FIX (audit17): Optimizado para evitar N+1 queries.
        Usa agregación en bloque en lugar de consultas individuales por producto.
        
        ISS-005 FIX: Dos modos de validación:
        - 'informativo': Solo reporta problemas sin bloquear (para creación)
        - 'estricto': Bloquea si no hay stock suficiente (para autorización/surtido)
        
        LÓGICA DE VALIDACIÓN:
        - Para requisiciones (validar_farmacia_central=True): 
          Valida contra stock de FARMACIA CENTRAL, ya que las requisiciones
          solicitan medicamentos que serán surtidos desde farmacia.
          ISS-004: Solo considera lotes vigentes (no vencidos).
        
        - Para operaciones internas de centro (validar_farmacia_central=False):
          Valida contra stock del centro específico.
        
        Args:
            items: Lista de items con producto y cantidad
            centro: Centro para contexto (usado para info, no para validación de requisiciones)
            validar_farmacia_central: Si True, valida contra farmacia central
            modo: 'informativo' (default) o 'estricto'
            
        Returns:
            list: Lista de errores/advertencias de stock
        """
        from django.utils import timezone
        from django.db.models import Sum
        from core.models import Lote
        
        errores = []
        today = timezone.now().date()
        
        # ISS-004 FIX: Recopilar todos los producto_ids primero
        producto_ids = []
        items_validos = []
        
        for item_data in items:
            producto_id = item_data.get('producto')
            if not producto_id:
                continue
            try:
                cantidad = int(item_data.get('cantidad_autorizada') or item_data.get('cantidad_solicitada') or 0)
            except (TypeError, ValueError):
                continue
            
            if cantidad <= 0:
                continue
            
            producto_ids.append(producto_id)
            items_validos.append({
                'producto_id': producto_id,
                'cantidad': cantidad,
                'item_data': item_data
            })
        
        if not producto_ids:
            return errores
        
        # ISS-004 FIX: Prefetch de productos en una sola consulta
        productos = {p.id: p for p in Producto.objects.filter(id__in=producto_ids)}
        
        # ISS-004 FIX: Calcular stock en bloque usando agregación
        # Filtro base para lotes vigentes
        lotes_filter = {
            'activo': True,
            'fecha_caducidad__gte': today,
            'producto_id__in': producto_ids,
        }
        
        if validar_farmacia_central:
            # Stock de farmacia central = lotes sin centro
            lotes_filter['centro__isnull'] = True
        elif centro:
            # Stock de centro específico
            lotes_filter['centro'] = centro
        else:
            # Fallback a farmacia central
            lotes_filter['centro__isnull'] = True
        
        # Agregar stock por producto en una sola consulta
        stock_por_producto = dict(
            Lote.objects.filter(**lotes_filter)
            .values('producto_id')
            .annotate(total=Sum('cantidad_actual'))
            .values_list('producto_id', 'total')
        )
        
        # Validar cada item
        for item in items_validos:
            producto_id = item['producto_id']
            cantidad = item['cantidad']
            producto = productos.get(producto_id)
            
            if not producto:
                continue
            
            disponible = stock_por_producto.get(producto_id, 0) or 0
            
            if cantidad > disponible:
                error_item = {
                    'producto': producto.clave,
                    'descripcion': producto.descripcion[:50] if producto.descripcion else '',
                    'disponible': disponible,
                    'solicitado': cantidad,
                    'deficit': cantidad - disponible,
                    'ubicacion': 'farmacia_central' if validar_farmacia_central else (
                        centro.nombre if centro else 'farmacia_central'
                    ),
                    'nota': 'Solo se considera stock de lotes vigentes (no vencidos)',
                    # ISS-005: Indicar tipo de error según modo
                    'tipo': 'advertencia' if modo == 'informativo' else 'error',
                    'bloquea': modo == 'estricto'
                }
                errores.append(error_item)
        
        return errores

    def get_queryset(self):
        """
        FLUJO V2: Filtros de seguridad a nivel de fila (Row Level Security lógica).
        
        Cada rol ve solo las requisiciones que le corresponden según su posición
        en el flujo jerárquico.
        
        ISS-006 FIX (audit17): Registro de auditoría para accesos privilegiados.
        ISS-DIRECTOR FIX: Usa _get_rol_efectivo para consistencia.
        
        OPTIMIZACIÓN: Para listados (action='list'), usa queryset ligero sin prefetch
        de detalles y anota el conteo de productos para evitar N+1.
        """
        from core.validators import AuditLogger
        
        # OPTIMIZACIÓN: Queryset base según la acción
        if self.action == 'list':
            # Para listados: select_related mínimo + anotación de conteo
            queryset = Requisicion.objects.select_related(
                'centro_origen', 'solicitante'
            ).annotate(
                total_productos=Count('detalles')
            )
        else:
            # Para detalle/acciones: queryset completo con prefetch
            queryset = Requisicion.objects.select_related(
                'centro_origen', 'centro_destino', 'solicitante', 'autorizador',
                # FLUJO V2: Actores del flujo
                'administrador_centro', 'director_centro', 
                'receptor_farmacia', 'autorizador_farmacia', 'surtidor'
            ).prefetch_related('detalles__producto')
        
        user = getattr(self.request, 'user', None)
        if not user or not user.is_authenticated:
            return Requisicion.objects.none()
        
        filter_applied = False  # ISS-006: Rastrear si se aplica filtro de centro
        
        # ISS-DIRECTOR FIX: Usar rol efectivo para consistencia con frontend
        rol = _get_rol_efectivo(user)
        
        # 1. Superusuario o Admin Global: Ve todo
        if user.is_superuser or rol == 'admin':
            # ISS-006 FIX: Registrar acceso privilegiado sin filtro
            AuditLogger.log_privileged_access(
                user, 
                'requisicion', 
                action='list_all',
                details={'sin_filtro_centro': True}
            )
            pass  # Sin filtro de centro
        
        # 2. Personal de Farmacia Central: Ve solo lo que ha sido enviado por Director
        elif rol == 'farmacia':
            # Farmacia NO debe ver:
            # - borrador: aún no enviada por médico
            # - pendiente_admin: pendiente de admin del centro
            # - pendiente_director: pendiente de director del centro
            # - devuelta: fue devuelta al centro para correcciones
            queryset = queryset.exclude(estado__in=['borrador', 'pendiente_admin', 'pendiente_director', 'devuelta'])
            filter_applied = True
        
        # ISS-019 FIX: 2.5 Rol Vista (Auditoría/Control): Lectura global para trazabilidad
        elif rol == 'vista':
            # Vista puede ver todo para auditoría, sin requerir centro asignado
            # Excluir borradores (no son documentos oficiales aún)
            queryset = queryset.exclude(estado='borrador')
            AuditLogger.log_privileged_access(
                user,
                'requisicion',
                action='list_audit',
                details={'rol': 'vista', 'acceso_global': True}
            )
            filter_applied = True  # Marca como filtrado (excluyó borradores)
        
        # 3. Usuarios de Centros Penitenciarios
        else:
            user_centro = self._user_centro(user)
            if not user_centro:
                # ISS-DIRECTOR FIX: Marcar que no hay centro para validación posterior
                # El método list() validará esto y retornará 403
                self._missing_centro = True
                return Requisicion.objects.none()
            
            # Filtrar por centro del usuario
            queryset = queryset.filter(
                Q(centro_origen=user_centro) | Q(centro_destino=user_centro)
            )
            
            # 3.1 Médico: Solo sus propias requisiciones (las que creó)
            if rol == 'medico':
                queryset = queryset.filter(solicitante=user)
            
            # 3.2 Administrador Centro: Ve desde pendiente_admin en adelante
            # FLUJO V2: Admin NO ve borradores de otros ni pendiente_director (eso es del director)
            # ISS-ROL-FIX: Incluir alias 'admin_centro' para compatibilidad
            elif rol in ['administrador_centro', 'admin_centro']:
                # Puede ver: pendiente_admin (para autorizar) + todo lo que ya pasó esa etapa
                # NO ve: borradores de otros
                queryset = queryset.exclude(
                    Q(estado='borrador') & ~Q(solicitante=user)
                )
            
            # 3.3 Director Centro: Ve SOLO desde pendiente_director en adelante
            # FLUJO V2: Director NO debe ver pendiente_admin (eso es del Admin)
            # ISS-ROL-FIX: Incluir alias 'director' para compatibilidad
            elif rol in ['director_centro', 'director']:
                # Estados que el director puede ver:
                # - pendiente_director: para autorizar
                # - enviada, autorizada, parcial, surtida, entregada: ya autorizados
                # - rechazada, cancelada, vencida: finalizados
                # NO ve: borrador, pendiente_admin (aún no llegó a su etapa)
                queryset = queryset.exclude(estado__in=['borrador', 'pendiente_admin'])
                queryset = queryset.exclude(estado__in=['borrador', 'pendiente_admin'])
            
            # 3.4 Centro genérico: Ve todo del centro (lectura)
            # else: ya está filtrado por centro
            filter_applied = True
        
        # Aplicar filtros adicionales de query params
        # ISS-019 FIX: Admin/farmacia/vista pueden filtrar por centro específico
        if user.is_superuser or rol in ['admin', 'farmacia', 'vista']:
            centro_param = self.request.query_params.get('centro')
            if centro_param:
                queryset = queryset.filter(
                    Q(centro_origen_id=centro_param) | Q(centro_destino_id=centro_param)
                )
                filter_applied = True  # ISS-006: Ya filtraron por centro

        estado = self.request.query_params.get('estado')
        if estado:
            queryset = queryset.filter(estado=estado.lower())

        grupo = self.request.query_params.get('grupo_estado')
        if grupo and grupo in REQUISICION_GRUPOS_ESTADO:
            queryset = queryset.filter(estado__in=REQUISICION_GRUPOS_ESTADO[grupo])

        search = self.request.query_params.get('search')
        if search and search.strip():
            queryset = queryset.filter(numero__icontains=search.strip())

        # Filtros de fecha
        fecha_desde = self.request.query_params.get('fecha_desde')
        if fecha_desde:
            queryset = queryset.filter(fecha_solicitud__date__gte=fecha_desde)
        
        fecha_hasta = self.request.query_params.get('fecha_hasta')
        if fecha_hasta:
            queryset = queryset.filter(fecha_solicitud__date__lte=fecha_hasta)

        # ISS-006 FIX: Registrar consultas globales sin filtro de centro
        if not filter_applied:
            AuditLogger.log_global_query(user, queryset, filter_applied=False)

        return queryset.order_by('-fecha_solicitud')
    
    def list(self, request, *args, **kwargs):
        """
        ISS-DIRECTOR FIX: Override de list para retornar 403 explícito
        cuando un usuario de centro no tiene centro asignado.
        
        Esto evita que se devuelva lista vacía silenciosa, lo que puede
        ocultar errores de configuración o intentos de acceso indebido.
        """
        # Resetear flag
        self._missing_centro = False
        
        # Llamar get_queryset que puede setear _missing_centro
        queryset = self.filter_queryset(self.get_queryset())
        
        # Verificar si falta centro
        if getattr(self, '_missing_centro', False):
            rol = _get_rol_efectivo(request.user)
            logger.warning(
                f"ISS-DIRECTOR: Usuario {request.user.username} (rol={rol}) "
                f"intentó listar requisiciones sin centro asignado"
            )
            return Response({
                'error': 'Usuario de centro sin centro asignado',
                'detail': 'Su cuenta está configurada como usuario de centro pero no tiene un centro asignado. Contacte al administrador.',
                'rol': rol
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Continuar con el list normal
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)
    
    @transaction.atomic  # HALLAZGO #10 FIX: Garantizar atomicidad requisición + detalles
    def create(self, request, *args, **kwargs):
        """
        ISS-001 FIX: Crea requisicion SIEMPRE en estado borrador.
        
        SEGURIDAD:
        - El estado enviado por el cliente es IGNORADO
        - Usuarios de centro: siempre 'borrador'
        - Farmacia/Admin: siempre 'borrador' (deben usar acción 'enviar' para cambiar)
        - Se valida contra máquina de estados antes de persistir
        """
        try:
            data = request.data.copy()
            fecha = timezone.now()
            numero = f"REQ-{fecha.strftime('%Y%m%d')}-{random.randint(1000, 9999)}"
            while Requisicion.objects.filter(numero=numero).exists():
                numero = f"REQ-{fecha.strftime('%Y%m%d')}-{random.randint(1000, 9999)}"

            # ISS-001 FIX: IGNORAR estado enviado por cliente - siempre borrador
            # El cliente NO puede crear requisiciones en estados avanzados
            estado_solicitado = str(data.get('estado', '')).lower()
            if estado_solicitado and estado_solicitado != 'borrador':
                logger.warning(
                    f"ISS-001: Intento de crear requisición en estado '{estado_solicitado}' "
                    f"por usuario {request.user.username}. Forzando a 'borrador'."
                )
            data['estado'] = ESTADO_INICIAL_CENTRO  # Siempre 'borrador'
            data['numero'] = numero

            solicitante = request.user if request.user.is_authenticated else None
            es_privilegiado = False
            
            # ISS-FIX-CENTRO: Corregir semántica de centros
            # centro_origen = centro del usuario que SOLICITA (de donde sale la requisición)
            # centro_destino = farmacia central (NULL) - a donde va la requisición
            # El frontend envía 'centro' que es el centro del solicitante → debe ir en centro_origen
            if 'centro' in data:
                # El centro del frontend es el centro ORIGEN (quien solicita)
                data['centro_origen'] = data.pop('centro')
            
            # ISS-FIX: Mapear 'comentario' del frontend a 'notas' para consistencia
            if 'comentario' in data and 'notas' not in data:
                data['notas'] = data.pop('comentario')
            
            if solicitante:
                data['solicitante'] = getattr(solicitante, 'id', None)
                centro_user = self._user_centro(solicitante)
                es_privilegiado = is_farmacia_or_admin(solicitante)
                if centro_user and not es_privilegiado:
                    # Validar que el centro enviado sea el del usuario
                    if data.get('centro_origen') and int(data.get('centro_origen')) != centro_user.id and not solicitante.is_superuser:
                        return Response({'error': 'No puedes crear requisiciones para otro centro'}, status=status.HTTP_403_FORBIDDEN)
                    # Asignar centro_origen = centro del usuario solicitante
                    data['centro_origen'] = centro_user.id
                    # centro_destino = NULL (farmacia central, no tiene centro)
                    data['centro_destino'] = None
                elif not centro_user and not es_privilegiado and not solicitante.is_superuser:
                    return Response({'error': 'El usuario no tiene centro asignado'}, status=status.HTTP_403_FORBIDDEN)
                elif not data.get('centro_origen') and centro_user:
                    data['centro_origen'] = centro_user.id
                    data['centro_destino'] = None

            items_data = request.data.get('items', []) or request.data.get('detalles', []) or []
            # ISS-FIX-CENTRO: Usar centro_origen para obtener info del centro solicitante
            centro_origen = Centro.objects.filter(id=data.get('centro_origen')).first() if data.get('centro_origen') else None
            if not centro_origen and solicitante and not solicitante.is_superuser and not es_privilegiado:
                return Response({'error': 'No se encontró el centro origen para la requisición'}, status=status.HTTP_400_BAD_REQUEST)
            
            # ISS-001, ISS-004, ISS-005 FIX: Validación INFORMATIVA de stock en creación
            # Las requisiciones solicitan medicamentos que serán surtidos desde farmacia central
            # En creación solo advertimos, no bloqueamos (modo='informativo')
            advertencias_stock = self._validar_stock_items(
                items_data, 
                centro=centro_origen,  # Pasamos centro_origen para referencia
                validar_farmacia_central=True,
                modo='informativo'  # ISS-005: Solo advertir en creación
            )

            # Agregar detalles a data para que el serializer los procese
            data['detalles'] = items_data
            
            serializer = self.get_serializer(data=data)
            serializer.is_valid(raise_exception=True)
            requisicion = serializer.save()

            response_data = {
                'mensaje': 'Requisicion creada exitosamente',
                'requisicion': RequisicionSerializer(requisicion).data
            }
            
            # ISS-005: Incluir advertencias de stock si las hay
            if advertencias_stock:
                response_data['advertencias_stock'] = advertencias_stock
                response_data['nota'] = (
                    'Hay productos con stock insuficiente en farmacia central. '
                    'La requisición se creó pero puede ser rechazada en autorización.'
                )

            return Response(response_data, status=status.HTTP_201_CREATED)
        except serializers.ValidationError as exc:
            logger.error(f"[RequisicionViewSet.create] ValidationError: {exc.detail}")
            return Response({'error': 'Error de validacion', 'detalles': exc.detail}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as exc:
            # ISS-FIX: Log completo del error para diagnóstico
            import traceback
            logger.error(f"[RequisicionViewSet.create] Exception: {str(exc)}")
            logger.error(f"[RequisicionViewSet.create] Traceback: {traceback.format_exc()}")
            return Response({'error': 'Error al crear requisicion', 'mensaje': str(exc)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def update(self, request, *args, **kwargs):
        """Solo permite editar si sigue en borrador o está devuelta."""
        try:
            requisicion = self.get_object()
            estado_actual = (requisicion.estado or '').lower()
            # Permitir editar en borrador o devuelta (para que el médico corrija después de devolución)
            if estado_actual not in ['borrador', 'devuelta']:
                return Response({'error': 'Solo se pueden editar requisiciones en estado BORRADOR o DEVUELTA', 'estado_actual': requisicion.estado}, status=status.HTTP_400_BAD_REQUEST)

            centro_user = self._user_centro(request.user)
            if not request.user.is_superuser and centro_user and requisicion.centro_id != centro_user.id:
                return Response({'error': 'No puedes editar requisiciones de otro centro'}, status=status.HTTP_403_FORBIDDEN)

            # ISS-FIX: Extraer items/detalles ANTES de pasar al serializer
            # para evitar errores de validación del nested serializer
            items_data = request.data.get('items') or request.data.get('detalles') or []
            
            # Preparar datos sin 'detalles' para el serializer de la requisición
            serializer_data = {k: v for k, v in request.data.items() if k not in ['items', 'detalles']}
            
            partial = kwargs.pop('partial', False)
            serializer = self.get_serializer(requisicion, data=serializer_data, partial=True)  # Siempre partial
            serializer.is_valid(raise_exception=True)
            requisicion = serializer.save()

            if items_data:
                # ISS-001, ISS-004, ISS-005: Validar stock en modo informativo (solo advertir)
                advertencias_stock = self._validar_stock_items(
                    items_data, 
                    centro=requisicion.centro,
                    validar_farmacia_central=True,
                    modo='informativo'  # ISS-005: Solo advertir en edición
                )
                requisicion.detalles.all().delete()
                for item_data in items_data:
                    producto_id = item_data.get('producto')
                    cant = item_data.get('cantidad_solicitada')
                    if not producto_id or cant in [None, '']:
                        continue
                    # ISS-FIX-LOTE: Incluir el lote específico si viene en el request
                    lote_id = item_data.get('lote') or item_data.get('lote_id')
                    # ISS-FIX: No establecer cantidad_autorizada en borradores
                    # para evitar validación de motivo_ajuste (solo aplica en autorización)
                    detalle_data = {
                        'requisicion': requisicion,
                        'producto_id': producto_id,
                        'lote_id': lote_id,
                        'cantidad_solicitada': int(cant),
                        'notas': item_data.get('observaciones') or item_data.get('notas') or ''
                    }
                    # Solo incluir cantidad_autorizada si viene explícitamente en el request
                    if item_data.get('cantidad_autorizada') is not None:
                        detalle_data['cantidad_autorizada'] = int(item_data.get('cantidad_autorizada'))
                    DetalleRequisicion.objects.create(**detalle_data)
                
                # ISS-005: Devolver resultado con advertencias si las hay
                response_data = {'mensaje': 'Requisicion actualizada exitosamente', 'requisicion': RequisicionSerializer(requisicion).data}
                if advertencias_stock:
                    response_data['advertencias_stock'] = advertencias_stock
                return Response(response_data)

            return Response({'mensaje': 'Requisicion actualizada exitosamente', 'requisicion': RequisicionSerializer(requisicion).data})
        except ValidationError as e:
            logger.error(f"ValidationError en update requisición: {e}")
            return Response({'error': str(e.message_dict if hasattr(e, 'message_dict') else e)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            logger.error(f"Error inesperado en update requisición: {type(e).__name__}: {e}", exc_info=True)
            return Response({'error': f'Error interno: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def destroy(self, request, *args, **kwargs):
        requisicion = self.get_object()
        if (requisicion.estado or '').lower() != 'borrador':
            return Response({'error': 'Solo se pueden eliminar requisiciones en estado BORRADOR', 'estado_actual': requisicion.estado}, status=status.HTTP_400_BAD_REQUEST)
        centro_user = self._user_centro(request.user)
        if not request.user.is_superuser:
            if not centro_user or requisicion.centro_id != centro_user.id:
                return Response({'error': 'No puedes eliminar requisiciones de otro centro'}, status=status.HTTP_403_FORBIDDEN)
        folio = requisicion.folio
        requisicion.delete()
        return Response({'mensaje': 'Requisicion eliminada', 'folio_eliminado': folio}, status=status.HTTP_204_NO_CONTENT)

    @action(detail=True, methods=['post'])
    def enviar(self, request, pk=None):
        """
        ISS-002/ISS-007 FIX (audit17): Envía requisición con revalidación de stock.
        
        Antes de enviar, se revalida el stock disponible en farmacia central.
        Si hay productos sin stock suficiente, se advierte pero permite enviar
        (el bloqueo ocurre en autorización/surtido).
        """
        requisicion = self.get_object()
        if (requisicion.estado or '').lower() != 'borrador':
            return Response({'error': 'Solo se pueden enviar requisiciones en estado BORRADOR', 'estado_actual': requisicion.estado}, status=status.HTTP_400_BAD_REQUEST)
        centro_user = self._user_centro(request.user)
        es_privilegiado = is_farmacia_or_admin(request.user)
        if not request.user.is_superuser and not es_privilegiado:
            if not centro_user or requisicion.centro_id != centro_user.id:
                return Response({'error': 'No puedes enviar requisiciones de otro centro'}, status=status.HTTP_403_FORBIDDEN)
        if not requisicion.detalles.exists():
            return Response({'error': 'La requisicion debe tener al menos un producto'}, status=status.HTTP_400_BAD_REQUEST)
        
        # ISS-002/ISS-007 FIX: Revalidar stock antes de enviar (modo informativo)
        # ISS-FIX-CENTRO: usar centro_origen (el centro que solicita)
        items_para_validar = [
            {'producto': d.producto_id, 'cantidad_solicitada': d.cantidad_solicitada}
            for d in requisicion.detalles.all()
        ]
        advertencias_stock = self._validar_stock_items(
            items_para_validar,
            centro=requisicion.centro_origen,
            validar_farmacia_central=True,
            modo='informativo'
        )
        
        # ISS-DB-002: Usar 'enviada' (valor en BD Supabase)
        requisicion.estado = 'enviada'
        requisicion.save(update_fields=['estado'])
        
        response_data = {
            'mensaje': 'Requisicion enviada', 
            'requisicion': RequisicionSerializer(requisicion).data
        }
        
        # ISS-002: Incluir advertencias de stock si las hay
        if advertencias_stock:
            response_data['advertencias_stock'] = advertencias_stock
            response_data['nota'] = (
                'Algunos productos tienen stock insuficiente en farmacia central. '
                'La requisición puede ser ajustada o rechazada en autorización.'
            )
        
        return Response(response_data)

    @action(detail=True, methods=['post'])
    def autorizar(self, request, pk=None):
        """
        ISS-004 FIX (audit18): Autorizar requisición con bloqueo para prevenir race conditions.
        
        Usa transaction.atomic() y select_for_update() para evitar que múltiples
        usuarios autoricen la misma requisición simultáneamente.
        
        ISS-FLUJO-FIX: Solo permite autorizar desde 'en_revision' (después de recibir)
        El flujo correcto es: enviada → recibir → en_revision → autorizar
        """
        from django.db import transaction
        
        requisicion = self.get_object()
        # ISS-FLUJO-FIX: Solo aceptar 'en_revision' - el flujo requiere recibir primero
        estados_autorizables = ['en_revision']
        estado_actual = (requisicion.estado or '').lower()
        if estado_actual not in estados_autorizables:
            return Response({
                'error': f'Solo se pueden autorizar requisiciones en estado EN_REVISION. Debe recibir la requisición primero.',
                'estado_actual': requisicion.estado,
                'estados_permitidos': estados_autorizables,
                'flujo_correcto': 'enviada → recibir → en_revision → autorizar'
            }, status=status.HTTP_400_BAD_REQUEST)

        centro_user = self._user_centro(request.user)
        es_privilegiado = is_farmacia_or_admin(request.user)
        if not request.user.is_superuser and not es_privilegiado:
            if not centro_user:
                return Response({'error': 'El usuario no tiene centro asignado'}, status=status.HTTP_403_FORBIDDEN)
            if requisicion.centro_id != centro_user.id:
                return Response({'error': 'No puedes autorizar requisiciones de otro centro'}, status=status.HTTP_403_FORBIDDEN)
        # ISS-DIRECTOR FIX: Usar rol efectivo en lugar de rol campo directo
        rol_efectivo = _get_rol_efectivo(request.user)
        if not request.user.is_superuser and rol_efectivo.lower() not in ['admin_sistema', 'farmacia', 'admin_farmacia', 'superusuario'] and not request.user.is_staff:
            return Response({'error': 'No tienes permiso para autorizar requisiciones', 'rol_actual': rol_efectivo}, status=status.HTTP_403_FORBIDDEN)

        items_data = request.data.get('items') or request.data.get('detalles') or []
        
        # ISS-004 FIX (audit18): Ejecutar autorización dentro de transacción atómica
        try:
            with transaction.atomic():
                # Bloquear requisición para evitar modificaciones concurrentes
                requisicion_bloqueada = Requisicion.objects.select_for_update(nowait=False).get(pk=requisicion.pk)
                
                # ISS-FLUJO-FIX: Re-verificar estado después del bloqueo (pudo cambiar)
                estado_bloqueado = (requisicion_bloqueada.estado or '').lower()
                if estado_bloqueado not in estados_autorizables:
                    return Response({
                        'error': 'La requisición ya no está en estado EN_REVISION (modificada concurrentemente)',
                        'estado_actual': requisicion_bloqueada.estado,
                        'estados_permitidos': estados_autorizables,
                        'flujo_correcto': 'enviada → recibir → en_revision → autorizar'
                    }, status=status.HTTP_409_CONFLICT)
                
                return self._autorizar_con_bloqueo(request, requisicion_bloqueada, items_data)
                
        except Exception as e:
            logger.error(f"ISS-004: Error en autorización atómica: {e}")
            return Response({
                'error': f'Error al autorizar: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def _autorizar_con_bloqueo(self, request, requisicion, items_data):
        """
        ISS-004 FIX (audit18): Lógica de autorización ejecutada dentro de transacción.
        
        Este método se llama desde autorizar() dentro de un bloque transaction.atomic()
        con la requisición ya bloqueada via select_for_update().
        """
        # ISS-003: Revalidar stock disponible en FARMACIA CENTRAL antes de autorizar
        # El stock pudo haber cambiado desde que se creó/envió la requisición
        errores_stock = []
        advertencias_stock = []
        
        # ISS-004 FIX: Bloquear detalles también para evitar modificaciones
        detalles_bloqueados = DetalleRequisicion.objects.select_for_update().filter(
            requisicion=requisicion
        ).select_related('producto')
        
        # Crear mapa de detalles por ID para acceso rápido
        detalles_map = {d.id: d for d in detalles_bloqueados}
        
        for item_data in items_data:
            item_id = item_data.get('id')
            cant_autorizada = item_data.get('cantidad_autorizada')
            if item_id is None or cant_autorizada is None:
                continue
            
            # ISS-004 FIX: Usar detalle bloqueado del mapa
            item = detalles_map.get(item_id)
            if item is None:
                continue
            
            cant_autorizada = max(0, int(cant_autorizada))
            if cant_autorizada > 0:
                # ISS-001: Usar stock de FARMACIA CENTRAL (no global)
                stock_farmacia = item.producto.get_stock_farmacia_central()
                
                if stock_farmacia < cant_autorizada:
                    if stock_farmacia == 0:
                        errores_stock.append({
                            'producto': item.producto.clave,
                            'descripcion': item.producto.descripcion[:50],
                            'solicitado': cant_autorizada,
                            'disponible_farmacia': stock_farmacia,
                            'mensaje': 'Sin stock en farmacia central'
                        })
                    else:
                        advertencias_stock.append({
                            'producto': item.producto.clave,
                            'descripcion': item.producto.descripcion[:50],
                            'solicitado': cant_autorizada,
                            'disponible_farmacia': stock_farmacia,
                            'sugerido': stock_farmacia,
                            'mensaje': f'Stock insuficiente, disponible: {stock_farmacia}'
                        })
            
            item.cantidad_autorizada = cant_autorizada
            
            # MEJORA FLUJO 3: Guardar motivo_ajuste si se autoriza menos de lo solicitado
            if cant_autorizada < item.cantidad_solicitada:
                motivo_ajuste = (item_data.get('motivo_ajuste') or '').strip()
                if len(motivo_ajuste) < 10:
                    return Response({
                        'error': f'Debe indicar el motivo del ajuste (mínimo 10 caracteres) para {item.producto.clave}',
                        'producto': item.producto.clave,
                        'cantidad_solicitada': item.cantidad_solicitada,
                        'cantidad_autorizada': cant_autorizada
                    }, status=status.HTTP_400_BAD_REQUEST)
                item.motivo_ajuste = motivo_ajuste
            else:
                item.motivo_ajuste = None
            
            item.save()
        
        # ISS-003: Si hay productos sin stock, rechazar autorización o advertir
        if errores_stock:
            # Productos con 0 stock - no permitir autorización
            return Response({
                'error': 'No se puede autorizar: productos sin stock en farmacia central',
                'detalles': errores_stock,
                'advertencias': advertencias_stock
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Determinar estado final
        # Si todas las cantidades autorizadas son 0 o menores a lo solicitado -> parcial
        # IMPORTANTE: Usar consulta directa a DB para ver valores actualizados (evitar cache de QuerySet)
        from django.db.models import Sum
        totales = DetalleRequisicion.objects.filter(
            requisicion=requisicion
        ).aggregate(
            total_sol=Sum('cantidad_solicitada'),
            total_aut=Sum('cantidad_autorizada')
        )
        total_solicitado = totales['total_sol'] or 0
        total_autorizado = totales['total_aut'] or 0
        
        if total_autorizado == 0:
            return Response({
                'error': 'Debe autorizar al menos un producto',
                'total_solicitado': total_solicitado,
                'total_autorizado': total_autorizado
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # ISS-TRIGGER-FIX: El trigger de Supabase solo permite en_revision -> autorizada
        # El estado 'parcial' es para SURTIDO parcial, no autorización parcial
        # La autorización siempre va a 'autorizada', aunque las cantidades sean menores
        nuevo_estado = 'autorizada'
        es_autorizacion_parcial = total_autorizado < total_solicitado
        
        requisicion.estado = nuevo_estado
        requisicion.fecha_autorizacion = timezone.now()
        if request.user and request.user.is_authenticated:
            requisicion.autorizador = request.user
        requisicion.save(update_fields=['estado', 'fecha_autorizacion', 'autorizador_id'])

        response_data = {
            'mensaje': f'Requisición autorizada' + (' (cantidades ajustadas)' if es_autorizacion_parcial else ''),
            'requisicion': RequisicionSerializer(requisicion).data,
            'es_autorizacion_parcial': es_autorizacion_parcial,
            'total_solicitado': total_solicitado,
            'total_autorizado': total_autorizado
        }
        
        # ISS-003: Incluir advertencias si hay stock parcial
        if advertencias_stock:
            response_data['advertencias'] = advertencias_stock
            response_data['mensaje'] += ' (con advertencias de stock)'
        
        return Response(response_data)

    @action(detail=True, methods=['post'])
    def rechazar(self, request, pk=None):
        requisicion = self.get_object()
        # ISS-DB-002: Usar 'enviada' (valor en BD Supabase)
        if (requisicion.estado or '').lower() != 'enviada':
            return Response({'error': 'Solo se pueden rechazar requisiciones en estado ENVIADA', 'estado_actual': requisicion.estado}, status=status.HTTP_400_BAD_REQUEST)
        centro_user = self._user_centro(request.user)
        es_privilegiado = is_farmacia_or_admin(request.user)
        if not request.user.is_superuser and not es_privilegiado:
            if not centro_user or requisicion.centro_id != centro_user.id:
                return Response({'error': 'No puedes rechazar requisiciones de otro centro'}, status=status.HTTP_403_FORBIDDEN)
        motivo = request.data.get('observaciones') or request.data.get('comentario') or ''
        if not motivo.strip():
            return Response({'error': 'Debe proporcionar un motivo de rechazo'}, status=status.HTTP_400_BAD_REQUEST)
        requisicion.estado = 'rechazada'
        requisicion.notas = motivo
        requisicion.save(update_fields=['estado', 'notas'])
        return Response({'mensaje': 'Requisicion rechazada', 'requisicion': RequisicionSerializer(requisicion).data})

    @action(detail=True, methods=['post'])
    def cancelar(self, request, pk=None):
        """
        ISS-RES-002 FIX (audit-final): Cancelar requisición con transacción atómica.
        
        Usa transaction.atomic() y select_for_update() para prevenir
        modificaciones concurrentes durante la cancelación.
        """
        from django.db import transaction
        
        try:
            with transaction.atomic():
                # ISS-RES-002: Bloquear requisición para evitar modificaciones concurrentes
                requisicion = Requisicion.objects.select_for_update(nowait=False).get(pk=pk)
                
                estado_actual = (requisicion.estado or '').lower()
                
                # Verificar estado después del bloqueo (pudo cambiar)
                if estado_actual in ['surtida', 'cancelada', 'rechazada', 'entregada']:
                    return Response({
                        'error': f'No se puede cancelar una requisición en estado {requisicion.estado}',
                        'estado_actual': requisicion.estado
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                # Validar permisos de centro
                centro_user = self._user_centro(request.user)
                if not request.user.is_superuser:
                    if not centro_user or requisicion.centro_id != centro_user.id:
                        return Response({
                            'error': 'No puedes cancelar requisiciones de otro centro'
                        }, status=status.HTTP_403_FORBIDDEN)
                
                # ISS-RES-002: Usar cambiar_estado para validaciones
                # ISS-FIX: Motivo obligatorio para cancelación (mínimo 10 caracteres)
                motivo = request.data.get('motivo') or request.data.get('observaciones') or request.data.get('comentario') or ''
                motivo = motivo.strip() if motivo else ''
                
                if not motivo or len(motivo) < 10:
                    return Response({
                        'error': 'Debe proporcionar un motivo de cancelación (mínimo 10 caracteres)'
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                estado_anterior = requisicion.estado
                
                try:
                    requisicion.cambiar_estado(
                        'cancelada',
                        usuario=request.user,
                        motivo=motivo,
                        validar=True
                    )
                    # Guardar motivo en notas
                    requisicion.notas = f"[CANCELADA] {motivo}"
                    requisicion.save(update_fields=['estado', 'notas', 'updated_at'])
                    
                    # ISS-FIX: Registrar en historial con usuario y motivo
                    from core.models import RequisicionHistorialEstados
                    RequisicionHistorialEstados.objects.create(
                        requisicion=requisicion,
                        estado_anterior=estado_anterior,
                        estado_nuevo='cancelada',
                        usuario=request.user,
                        accion='cancelar',
                        motivo=motivo,
                        observaciones=f'Cancelada por {request.user.get_full_name() or request.user.username}',
                        ip_address=request.META.get('REMOTE_ADDR', ''),
                        datos_adicionales={
                            'cancelado_por_id': request.user.id,
                            'cancelado_por_nombre': request.user.get_full_name() or request.user.username,
                            'cancelado_por_rol': getattr(request.user, 'rol', 'N/A'),
                        }
                    )
                except ValidationError as e:
                    return Response({
                        'error': str(e.message_dict if hasattr(e, 'message_dict') else e)
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                # Log de auditoría
                logger.info(
                    f"ISS-RES-002: Requisición {requisicion.numero} cancelada por "
                    f"usuario {request.user.username}. Motivo: {motivo}"
                )
                
                return Response({
                    'mensaje': 'Requisición cancelada',
                    'requisicion': RequisicionSerializer(requisicion).data
                })
                
        except Requisicion.DoesNotExist:
            return Response({
                'error': 'Requisición no encontrada'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"ISS-RES-002: Error cancelando requisición {pk}: {e}")
            return Response({
                'error': f'Error al cancelar: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['get'], url_path='diagnostico-surtido')
    def diagnostico_surtido(self, request, pk=None):
        """
        ISS-DEBUG: Endpoint de diagnóstico para verificar si una requisición puede ser surtida.
        
        Retorna información detallada sobre:
        - Estado de la requisición
        - Detalles y sus productos
        - Stock disponible en farmacia central
        - Problemas detectados
        
        Este endpoint es de solo lectura, no modifica datos.
        """
        from django.utils import timezone
        from django.db.models import Sum
        
        hoy = timezone.now().date()
        
        try:
            requisicion = self.get_object()
        except Exception as e:
            return Response({
                'error': f'No se pudo obtener la requisición: {str(e)}'
            }, status=status.HTTP_404_NOT_FOUND)
        
        ESTADOS_SURTIBLES = {'autorizada', 'autorizada_farmacia', 'en_surtido', 'parcial'}
        estado_actual = (requisicion.estado or '').lower()
        
        diagnostico = {
            'requisicion': {
                'id': requisicion.pk,
                'numero': requisicion.numero,
                'folio': getattr(requisicion, 'folio', requisicion.numero),
                'estado': requisicion.estado,
                'centro': requisicion.centro.nombre if requisicion.centro else None,
                'centro_id': requisicion.centro_id,
                'solicitante': requisicion.solicitante.username if requisicion.solicitante else None,
                'fecha_creacion': requisicion.created_at.isoformat() if requisicion.created_at else None,
            },
            'estado_surtible': estado_actual in ESTADOS_SURTIBLES,
            'estados_validos': list(ESTADOS_SURTIBLES),
            'detalles': [],
            'errores': [],
            'advertencias': [],
            'puede_surtirse': True,
        }
        
        if estado_actual not in ESTADOS_SURTIBLES:
            diagnostico['errores'].append(
                f"Estado '{requisicion.estado}' no es surtible. Estados válidos: {ESTADOS_SURTIBLES}"
            )
            diagnostico['puede_surtirse'] = False
        
        detalles = requisicion.detalles.select_related('producto', 'lote').all()
        
        if not detalles:
            diagnostico['errores'].append("La requisición no tiene detalles")
            diagnostico['puede_surtirse'] = False
            return Response(diagnostico)
        
        for det in detalles:
            detalle_info = {
                'id': det.pk,
                'producto_id': det.producto_id,
                'producto_clave': det.producto.clave if det.producto else None,
                'producto_nombre': (det.producto.nombre[:50] + '...') if det.producto and len(det.producto.nombre) > 50 else (det.producto.nombre if det.producto else None),
                'cantidad_solicitada': det.cantidad_solicitada,
                'cantidad_autorizada': det.cantidad_autorizada,
                'cantidad_surtida': det.cantidad_surtida,
                'pendiente': 0,
                'lote_especifico': None,
                'stock_farmacia': 0,
                'lotes_disponibles': [],
                'puede_surtirse': True,
                'errores': [],
            }
            
            if not det.producto_id:
                detalle_info['errores'].append("Sin producto asignado")
                detalle_info['puede_surtirse'] = False
                diagnostico['puede_surtirse'] = False
                diagnostico['detalles'].append(detalle_info)
                continue
            
            # Calcular pendiente
            pendiente = (det.cantidad_autorizada or det.cantidad_solicitada or 0) - (det.cantidad_surtida or 0)
            detalle_info['pendiente'] = pendiente
            
            if pendiente <= 0:
                detalle_info['puede_surtirse'] = True  # Ya surtido
                diagnostico['detalles'].append(detalle_info)
                continue
            
            # Verificar cantidad autorizada
            if det.cantidad_autorizada is None or det.cantidad_autorizada == 0:
                diagnostico['advertencias'].append(
                    f"Detalle {det.pk} ({det.producto.clave}): cantidad_autorizada es NULL/0, "
                    f"se usará cantidad_solicitada={det.cantidad_solicitada}"
                )
            
            # Verificar lote específico vs FEFO
            if det.lote_id is not None:
                lote = det.lote
                detalle_info['lote_especifico'] = {
                    'id': lote.pk if lote else det.lote_id,
                    'numero': lote.numero_lote if lote else 'N/A',
                    'stock': lote.cantidad_actual if lote else 0,
                    'activo': lote.activo if lote else False,
                    'caducidad': str(lote.fecha_caducidad) if lote and lote.fecha_caducidad else None,
                    'centro': lote.centro.nombre if lote and lote.centro else 'Farmacia Central',
                }
                
                if not lote:
                    detalle_info['errores'].append(f"Lote referenciado {det.lote_id} no existe")
                    detalle_info['puede_surtirse'] = False
                elif not lote.activo:
                    detalle_info['errores'].append(f"Lote {lote.numero_lote} está inactivo")
                    detalle_info['puede_surtirse'] = False
                elif lote.cantidad_actual < pendiente:
                    detalle_info['errores'].append(
                        f"Stock insuficiente en lote {lote.numero_lote}: "
                        f"disponible={lote.cantidad_actual}, requerido={pendiente}"
                    )
                    detalle_info['puede_surtirse'] = False
                elif lote.fecha_caducidad and lote.fecha_caducidad < hoy:
                    detalle_info['errores'].append(f"Lote {lote.numero_lote} está vencido")
                    detalle_info['puede_surtirse'] = False
            else:
                # FEFO automático - buscar lotes en farmacia central
                lotes_disponibles = Lote.objects.filter(
                    centro__isnull=True,  # Farmacia central
                    producto=det.producto,
                    activo=True,
                    cantidad_actual__gt=0,
                    fecha_caducidad__gte=hoy
                ).order_by('fecha_caducidad')[:5]  # Limitar a 5 para respuesta
                
                stock_total = Lote.objects.filter(
                    centro__isnull=True,
                    producto=det.producto,
                    activo=True,
                    cantidad_actual__gt=0,
                    fecha_caducidad__gte=hoy
                ).aggregate(total=Sum('cantidad_actual'))['total'] or 0
                
                detalle_info['stock_farmacia'] = stock_total
                detalle_info['lotes_disponibles'] = [
                    {
                        'id': l.pk,
                        'numero': l.numero_lote,
                        'stock': l.cantidad_actual,
                        'caducidad': str(l.fecha_caducidad) if l.fecha_caducidad else None,
                    }
                    for l in lotes_disponibles
                ]
                
                if stock_total < pendiente:
                    detalle_info['errores'].append(
                        f"Stock insuficiente en farmacia: disponible={stock_total}, requerido={pendiente}"
                    )
                    detalle_info['puede_surtirse'] = False
            
            if not detalle_info['puede_surtirse']:
                diagnostico['puede_surtirse'] = False
            
            diagnostico['detalles'].append(detalle_info)
        
        # Resumen
        total_detalles = len(diagnostico['detalles'])
        detalles_ok = sum(1 for d in diagnostico['detalles'] if d['puede_surtirse'])
        diagnostico['resumen'] = {
            'total_detalles': total_detalles,
            'detalles_listos': detalles_ok,
            'detalles_con_problemas': total_detalles - detalles_ok,
        }
        
        return Response(diagnostico)

    @action(detail=True, methods=['post'])
    def surtir(self, request, pk=None):
        """
        ISS-011, ISS-021: Surte una requisición autorizada de forma ATÓMICA.
        
        Usa RequisicionService para garantizar:
        - Transacción atómica con rollback completo si falla cualquier paso
        - Bloqueo optimista de lotes (select_for_update)
        - Validación de permisos por centro
        
        PERMISOS:
        - Superuser: puede surtir cualquier requisición
        - Farmacia (admin_farmacia, farmacia): puede surtir cualquier requisición
        - Centro: solo puede surtir requisiciones de su propio centro
        
        LÓGICA DE STOCK:
        - Primero usa lotes de farmacia central (centro=NULL) → crea entrada en centro destino
        - Si no hay en farmacia central, usa lotes del centro solicitante (salida interna)
        
        NUEVO: Soporta foto de firma de surtido vía multipart/form-data
        """
        import traceback
        from inventario.services import (
            RequisicionService,
            StockInsuficienteError,
            EstadoInvalidoError,
            PermisoRequisicionError
        )
        
        # DEBUG: Log inicio de surtido
        logger.info(f"SURTIR: Iniciando surtido para requisición pk={pk}, usuario={request.user.username}")
        
        try:
            requisicion = self.get_object()
            logger.info(f"SURTIR: Requisición obtenida: {requisicion.folio}, estado={requisicion.estado}")
        except Exception as e:
            logger.exception(f"SURTIR: Error al obtener requisición: {e}")
            return Response({
                'error': 'No se pudo obtener la requisición',
                'detalle': str(e),
                'tipo_error': type(e).__name__
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
        # ISS-FIX-FECHA-VENCIDA: Verificación adicional de fecha límite antes de surtir
        # Si la fecha ya venció, marcar como vencida y retornar error
        fecha_limite = getattr(requisicion, 'fecha_recoleccion_limite', None)
        if fecha_limite and timezone.now() > fecha_limite:
            logger.warning(
                f"SURTIR: Requisición {requisicion.folio} tiene fecha límite vencida: {fecha_limite}"
            )
            # Marcar como vencida
            estado_anterior = requisicion.estado
            requisicion.estado = 'vencida'
            requisicion.fecha_vencimiento = timezone.now()
            requisicion.motivo_vencimiento = (
                f"Fecha límite de recolección vencida: {fecha_limite.strftime('%Y-%m-%d %H:%M')}"
            )
            requisicion.save(update_fields=['estado', 'fecha_vencimiento', 'motivo_vencimiento', 'updated_at'])
            
            # Registrar en historial
            self._registrar_historial(
                requisicion, estado_anterior, 'vencida',
                request.user, 'vencer_automatico', request,
                datos_adicionales={'motivo': requisicion.motivo_vencimiento}
            )
            
            return Response({
                'error': 'No se puede surtir: La fecha límite de recolección ya venció',
                'fecha_limite': fecha_limite.isoformat(),
                'estado_actual': 'vencida',
                'mensaje': 'La requisición ha sido marcada como vencida automáticamente'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # ISS-FIX-SURTIR: Auto-asignación de cantidad_autorizada movida al servicio
        # para que ocurra dentro de la transacción atómica
        
        try:
            # ISS-011, ISS-021: Usar servicio transaccional
            logger.info(f"SURTIR: Creando servicio para {requisicion.folio}")
            service = RequisicionService(requisicion, request.user)
            
            logger.info(f"SURTIR: Llamando service.surtir() para {requisicion.folio}")
            resultado = service.surtir(
                is_farmacia_or_admin_fn=is_farmacia_or_admin,
                get_user_centro_fn=get_user_centro
            )
            logger.info(f"SURTIR: Surtido exitoso para {requisicion.folio}")
            
            # Refrescar requisición para serializer
            requisicion.refresh_from_db()
            
            # ISS-004 FIX (audit21): Validar imagen completa (MIME, magic bytes, extensión)
            foto_firma = request.FILES.get('foto_firma_surtido') or request.FILES.get('foto_firma')
            if foto_firma:
                es_valido, error_msg = validar_archivo_imagen(foto_firma, max_size_mb=2)
                if es_valido:
                    requisicion.foto_firma_surtido = foto_firma
                    requisicion.fecha_firma_surtido = timezone.now()
                    requisicion.usuario_firma_surtido = request.user
                    requisicion.save(update_fields=['foto_firma_surtido', 'fecha_firma_surtido', 'usuario_firma_surtido'])
                else:
                    logger.warning(f"ISS-004: Firma rechazada en surtir {requisicion.folio}: {error_msg}")
            
            return Response({
                'mensaje': 'Requisición surtida exitosamente',
                'requisicion': RequisicionSerializer(requisicion, context={'request': request}).data,
                'detalles_surtido': resultado
            })
            
        except EstadoInvalidoError as e:
            return Response({
                'error': str(e),
                'codigo': e.code,
                'estado_actual': e.estado_actual
            }, status=status.HTTP_400_BAD_REQUEST)
            
        except PermisoRequisicionError as e:
            return Response({
                'error': str(e),
                'codigo': e.code
            }, status=status.HTTP_403_FORBIDDEN)
            
        except StockInsuficienteError as e:
            return Response({
                'error': str(e),
                'codigo': e.code,
                'detalles': e.detalles_stock
            }, status=status.HTTP_400_BAD_REQUEST)
            
        except Exception as e:
            logger.exception(f"Error inesperado al surtir requisición {requisicion.folio}: {e}")
            # ISS-DEBUG: Retornar error detallado para diagnóstico
            import traceback
            error_detalle = traceback.format_exc()
            return Response({
                'error': 'Error interno al procesar el surtido. Por favor intente nuevamente.',
                'codigo': 'error_interno',
                'detalle': str(e),
                'tipo_error': type(e).__name__
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['post'], url_path='marcar-recibida')
    def marcar_recibida(self, request, pk=None):
        """
        ISS-002 FIX: Marca una requisición surtida como recibida con CONCILIACIÓN.
        
        PERMISOS:
        - Solo usuarios del centro receptor pueden marcar como recibida
        - Solo requisiciones en estado 'surtida' o 'parcial' pueden marcarse
        
        DATOS REQUERIDOS:
        - lugar_entrega: Lugar donde se recibió
        - items_recibidos: Lista de {detalle_id, cantidad_recibida} para conciliación
        - observaciones_recepcion: Observaciones de la recepción (opcional)
        
        CONCILIACIÓN (ISS-002):
        - Valida cantidades recibidas vs surtidas
        - Registra divergencias (faltantes/daños)
        - Crea movimientos de recepción para trazabilidad
        """
        from django.utils import timezone
        from django.db import transaction
        
        requisicion = self.get_object()
        estado_actual = (requisicion.estado or '').lower()
        
        # ISS-002: También permitir 'parcial'
        if estado_actual not in ['surtida', 'parcial']:
            return Response({
                'error': 'Solo se pueden marcar como recibidas las requisiciones surtidas o parciales',
                'estado_actual': requisicion.estado
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # PERMISOS: Solo usuarios del centro receptor pueden confirmar recepción
        user = request.user
        if not user.is_superuser:
            centro_user = self._user_centro(user)
            if not centro_user or requisicion.centro_id != centro_user.id:
                return Response({
                    'error': 'Solo el centro receptor puede confirmar la recepción'
                }, status=status.HTTP_403_FORBIDDEN)
        
        # Obtener datos del request
        lugar_entrega = request.data.get('lugar_entrega', '')
        observaciones_recepcion = request.data.get('observaciones_recepcion', '')
        items_recibidos = request.data.get('items_recibidos', [])
        
        if not lugar_entrega:
            return Response({
                'error': 'El lugar de entrega es requerido'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # ISS-002 FIX: Conciliación de cantidades recibidas vs surtidas
        divergencias = []
        detalles_conciliados = []
        
        with transaction.atomic():
            for detalle in requisicion.detalles.select_related('producto'):
                cantidad_surtida = detalle.cantidad_surtida or 0
                
                # Buscar item recibido correspondiente
                item_recibido = next(
                    (i for i in items_recibidos if str(i.get('detalle_id')) == str(detalle.id)),
                    None
                )
                
                if item_recibido:
                    try:
                        cantidad_recibida = int(item_recibido.get('cantidad_recibida', cantidad_surtida))
                    except (ValueError, TypeError):
                        cantidad_recibida = cantidad_surtida
                    
                    observacion_item = item_recibido.get('observaciones', '')
                else:
                    # Si no se especifica, asumir que recibió lo surtido
                    cantidad_recibida = cantidad_surtida
                    observacion_item = ''
                
                # Detectar divergencia
                diferencia = cantidad_surtida - cantidad_recibida
                if diferencia != 0:
                    divergencias.append({
                        'producto': detalle.producto.clave,
                        'producto_nombre': (detalle.producto.nombre or '')[:50],
                        'cantidad_surtida': cantidad_surtida,
                        'cantidad_recibida': cantidad_recibida,
                        'diferencia': diferencia,
                        'tipo': 'faltante' if diferencia > 0 else 'excedente',
                        'observaciones': observacion_item
                    })
                    
                    # ISS-002: Registrar movimiento de ajuste por divergencia
                    # ISS-FIX: Usar centro_origen (quien hizo la requisición), no centro (alias de centro_destino)
                    if diferencia > 0 and requisicion.centro_origen:
                        # Faltante: registrar ajuste negativo en centro
                        lotes_centro = Lote.objects.filter(
                            producto=detalle.producto,
                            centro=requisicion.centro_origen,
                            activo=True,
                            cantidad_actual__gt=0
                        ).order_by('fecha_caducidad')
                        
                        faltante_pendiente = diferencia
                        for lote in lotes_centro:
                            if faltante_pendiente <= 0:
                                break
                            
                            ajustar = min(faltante_pendiente, lote.cantidad_actual)
                            registrar_movimiento_stock(
                                lote=lote,
                                tipo='ajuste',
                                cantidad=-ajustar,
                                usuario=user,
                                centro=requisicion.centro_origen,
                                requisicion=requisicion,
                                observaciones=f'AJUSTE_RECEPCION: Faltante en recepción REQ-{requisicion.numero}. {observacion_item}',
                                skip_centro_check=True
                            )
                            faltante_pendiente -= ajustar
                
                # Guardar cantidad recibida en detalle
                detalle.cantidad_recibida = cantidad_recibida
                detalle.save(update_fields=['cantidad_recibida'])
                
                detalles_conciliados.append({
                    'detalle_id': detalle.id,
                    'producto': detalle.producto.clave,
                    'surtida': cantidad_surtida,
                    'recibida': cantidad_recibida
                })
            
            # Actualizar la requisición
            requisicion.estado = 'entregada'
            requisicion.fecha_entrega = timezone.now()
            requisicion.lugar_entrega = lugar_entrega
            requisicion.usuario_firma_recepcion = user
            requisicion.fecha_firma_recepcion = timezone.now()
            
            # ISS-002: Registrar divergencias y responsable en notas
            notas_recepcion = f"[Recepción {timezone.now().strftime('%Y-%m-%d %H:%M')}] "
            notas_recepcion += f"Recibido por: {user.username}. "
            if observaciones_recepcion:
                notas_recepcion += f"Obs: {observaciones_recepcion}. "
            if divergencias:
                notas_recepcion += f"DIVERGENCIAS: {len(divergencias)} productos con diferencias. "
            
            notas_actual = requisicion.notas or ''
            requisicion.notas = f"{notas_actual}\n{notas_recepcion}".strip()
            
            # Validar foto de firma si se incluye
            foto_firma = request.FILES.get('foto_firma_recepcion') or request.FILES.get('foto_firma')
            if foto_firma:
                # ISS-006: Validar imagen antes de guardar
                es_valido, error_msg = validar_archivo_imagen(foto_firma, max_size_mb=2)
                if not es_valido:
                    return Response({
                        'error': f'Error en foto de firma: {error_msg}'
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                requisicion.foto_firma_recepcion = foto_firma
            
            update_fields = [
                'estado', 'fecha_entrega', 'lugar_entrega', 'notas',
                'usuario_firma_recepcion_id', 'fecha_firma_recepcion'
            ]
            if foto_firma:
                update_fields.append('foto_firma_recepcion')
            
            requisicion.save(update_fields=update_fields)
        
        response_data = {
            'mensaje': 'Requisición marcada como recibida',
            'requisicion': RequisicionSerializer(requisicion, context={'request': request}).data,
            'conciliacion': {
                'detalles': detalles_conciliados,
                'total_items': len(detalles_conciliados)
            }
        }
        
        # ISS-002: Incluir divergencias si las hay
        if divergencias:
            response_data['divergencias'] = divergencias
            response_data['mensaje'] += f' (con {len(divergencias)} divergencias registradas)'
        
        return Response(response_data)

    @action(detail=True, methods=['post'], url_path='subir-firma-surtido')
    def subir_firma_surtido(self, request, pk=None):
        """
        ISS-006 FIX: Sube la foto de firma de surtido con validación de imagen.
        
        ISS-DB-002: 'recibida' -> 'entregada' para alinearse con BD Supabase.
        Solo disponible para requisiciones en estado 'surtida' o 'entregada'.
        Solo usuarios de farmacia/admin pueden subir esta firma.
        
        Validaciones de seguridad (ISS-006):
        - Extensión permitida (.jpg, .jpeg, .png, .gif, .webp)
        - MIME type correcto
        - Magic bytes válidos (contenido real = imagen)
        - Tamaño máximo 2MB
        """
        from django.utils import timezone
        
        requisicion = self.get_object()
        estado_actual = (requisicion.estado or '').lower()
        
        # ISS-DB-002: 'recibida' -> 'entregada'
        if estado_actual not in ['surtida', 'entregada']:
            return Response({
                'error': 'Solo se puede subir firma de surtido para requisiciones surtidas o entregadas',
                'estado_actual': requisicion.estado
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # PERMISOS: Solo farmacia/admin puede subir firma de surtido
        user = request.user
        if not user.is_superuser and not is_farmacia_or_admin(user):
            return Response({
                'error': 'Solo farmacia o administradores pueden subir la firma de surtido'
            }, status=status.HTTP_403_FORBIDDEN)
        
        foto_firma = request.FILES.get('foto_firma_surtido') or request.FILES.get('foto_firma')
        if not foto_firma:
            return Response({
                'error': 'No se proporcionó la foto de firma'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # ISS-006 FIX: Validar imagen con magic bytes, MIME y extensión
        es_valido, error_msg = validar_archivo_imagen(foto_firma, max_size_mb=2)
        if not es_valido:
            return Response({
                'error': f'Error en foto de firma: {error_msg}'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        requisicion.foto_firma_surtido = foto_firma
        requisicion.fecha_firma_surtido = timezone.now()
        requisicion.usuario_firma_surtido = user
        requisicion.save(update_fields=['foto_firma_surtido', 'fecha_firma_surtido', 'usuario_firma_surtido'])
        
        return Response({
            'mensaje': 'Firma de surtido subida correctamente',
            'requisicion': RequisicionSerializer(requisicion, context={'request': request}).data
        })

    @action(detail=True, methods=['post'], url_path='subir-firma-recepcion')
    def subir_firma_recepcion(self, request, pk=None):
        """
        ISS-006 FIX: Sube la foto de firma de recepción con validación de imagen.
        
        ISS-DB-002: 'recibida' -> 'entregada' para alinearse con BD Supabase.
        Solo disponible para requisiciones en estado 'entregada'.
        Solo usuarios del centro receptor pueden subir esta firma.
        
        Validaciones de seguridad (ISS-006):
        - Extensión permitida (.jpg, .jpeg, .png, .gif, .webp)
        - MIME type correcto
        - Magic bytes válidos (contenido real = imagen)
        - Tamaño máximo 2MB
        """
        from django.utils import timezone
        
        requisicion = self.get_object()
        estado_actual = (requisicion.estado or '').lower()
        
        # ISS-DB-002: 'recibida' -> 'entregada'
        if estado_actual != 'entregada':
            return Response({
                'error': 'Solo se puede subir firma de recepción para requisiciones entregadas',
                'estado_actual': requisicion.estado
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # PERMISOS: Solo usuarios del centro receptor pueden subir firma de recepción
        user = request.user
        if not user.is_superuser:
            centro_user = self._user_centro(user)
            if not centro_user or requisicion.centro_id != centro_user.id:
                return Response({
                    'error': 'Solo el centro receptor puede subir la firma de recepción'
                }, status=status.HTTP_403_FORBIDDEN)
        
        foto_firma = request.FILES.get('foto_firma_recepcion') or request.FILES.get('foto_firma')
        if not foto_firma:
            return Response({
                'error': 'No se proporcionó la foto de firma'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # ISS-006 FIX: Validar imagen con magic bytes, MIME y extensión
        es_valido, error_msg = validar_archivo_imagen(foto_firma, max_size_mb=2)
        if not es_valido:
            return Response({
                'error': f'Error en foto de firma: {error_msg}'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        requisicion.foto_firma_recepcion = foto_firma
        requisicion.fecha_firma_recepcion = timezone.now()
        requisicion.usuario_firma_recepcion = user
        requisicion.save(update_fields=['foto_firma_recepcion', 'fecha_firma_recepcion', 'usuario_firma_recepcion'])
        
        return Response({
            'mensaje': 'Firma de recepción subida correctamente',
            'requisicion': RequisicionSerializer(requisicion, context={'request': request}).data
        })
        
        return Response({
            'mensaje': 'Firma de recepción subida correctamente',
            'requisicion': RequisicionSerializer(requisicion, context={'request': request}).data
        })

    @action(detail=True, methods=['get'], url_path='hoja-recoleccion')
    def hoja_recoleccion(self, request, pk=None):
        """
        Genera y descarga el PDF de la hoja de recolección para una requisición.
        ISS-DB-002: Estados alineados con BD Supabase.
        Solo disponible para requisiciones autorizadas, en_surtido, parcial o surtidas.
        """
        from core.utils.pdf_generator import generar_hoja_recoleccion
        
        requisicion = self.get_object()
        estado = (requisicion.estado or '').lower()
        
        # ISS-DB-002: Validar estados de BD Supabase
        if estado not in ['autorizada', 'en_surtido', 'parcial', 'surtida']:
            return Response({
                'error': 'Solo se pueden generar hojas para requisiciones autorizadas, en surtido, parciales o surtidas',
                'estado_actual': requisicion.estado
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            # Generar el PDF
            pdf_buffer = generar_hoja_recoleccion(requisicion)
            
            # Crear respuesta HTTP con el PDF
            response = HttpResponse(
                pdf_buffer.getvalue(),
                content_type='application/pdf'
            )
            folio_safe = (requisicion.folio or f'REQ-{requisicion.id}').replace('/', '-')
            response['Content-Disposition'] = f'attachment; filename="Hoja_Recoleccion_{folio_safe}.pdf"'
            
            return response
            
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            return Response({
                'error': 'Error al generar la hoja de recolección',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['get'], url_path='hoja-consulta')
    def hoja_consulta(self, request, pk=None):
        """
        Genera y descarga el PDF de hoja de consulta simplificada para centros.
        Esta versión tiene un sello "SURTIDA" y NO muestra firmas completas.
        Solo disponible para requisiciones surtidas o entregadas.
        Para uso exclusivo de roles de centro (médico, usuario_centro, etc.)
        """
        from core.utils.pdf_generator import generar_hoja_consulta
        
        requisicion = self.get_object()
        estado = (requisicion.estado or '').lower()
        
        # Solo para estados surtida o entregada
        if estado not in ['surtida', 'entregada']:
            return Response({
                'error': 'La hoja de consulta solo está disponible para requisiciones surtidas o entregadas',
                'estado_actual': requisicion.estado
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            # Generar el PDF de consulta con sello SURTIDA
            pdf_buffer = generar_hoja_consulta(requisicion)
            
            # Crear respuesta HTTP con el PDF
            response = HttpResponse(
                pdf_buffer.getvalue(),
                content_type='application/pdf'
            )
            folio_safe = (requisicion.folio or f'REQ-{requisicion.id}').replace('/', '-')
            response['Content-Disposition'] = f'attachment; filename="Consulta_Requisicion_{folio_safe}.pdf"'
            
            return response
            
        except Exception as e:
            import traceback
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Error generando hoja consulta req {pk}: {str(e)}")
            logger.error(traceback.format_exc())
            return Response({
                'error': 'Error al generar la hoja de consulta',
                'mensaje': str(e),
                'detalle': traceback.format_exc()
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['get'], url_path='pdf-rechazo')
    def pdf_rechazo(self, request, pk=None):
        """
        Genera y descarga el PDF de rechazo para una requisicin rechazada.
        """
        from core.utils.pdf_generator import generar_pdf_rechazo
        
        requisicion = self.get_object()
        estado = (requisicion.estado or '').lower()
        
        if estado != 'rechazada':
            return Response({
                'error': 'Solo se pueden generar PDFs de rechazo para requisiciones rechazadas',
                'estado_actual': requisicion.estado
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            pdf_buffer = generar_pdf_rechazo(requisicion)
            
            response = HttpResponse(
                pdf_buffer.getvalue(),
                content_type='application/pdf'
            )
            folio_safe = (requisicion.folio or f'REQ-{requisicion.id}').replace('/', '-')
            response['Content-Disposition'] = f'attachment; filename="Requisicion_Rechazada_{folio_safe}.pdf"'
            
            return response
            
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            return Response({
                'error': 'Error al generar el PDF de rechazo',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    # ==========================================================================
    # FLUJO V2: ENDPOINTS DE TRANSICIÓN DE ESTADOS JERÁRQUICOS
    # ==========================================================================
    
    def _get_client_ip(self, request):
        """Obtiene la IP del cliente para auditoría."""
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            return x_forwarded_for.split(',')[0].strip()
        return request.META.get('REMOTE_ADDR', '')
    
    def _registrar_historial(self, requisicion, estado_anterior, estado_nuevo, 
                              usuario, accion, request, motivo=None, datos_adicionales=None):
        """
        Registra un cambio de estado en el historial inmutable.
        """
        from core.models import RequisicionHistorialEstados
        
        try:
            RequisicionHistorialEstados.objects.create(
                requisicion=requisicion,
                estado_anterior=estado_anterior,
                estado_nuevo=estado_nuevo,
                usuario=usuario,
                accion=accion,
                motivo=motivo,
                ip_address=self._get_client_ip(request),
                user_agent=request.META.get('HTTP_USER_AGENT', '')[:500] if request else None,
                datos_adicionales=datos_adicionales
            )
        except Exception as e:
            logger.warning(f"No se pudo registrar historial: {e}")
    
    def _validar_transicion(self, estado_actual, estado_nuevo):
        """Valida que la transición de estado sea permitida."""
        transiciones_validas = TRANSICIONES_REQUISICION.get(estado_actual, [])
        return estado_nuevo in transiciones_validas
    
    def _validar_permiso_flujo(self, user, accion):
        """
        Valida que el usuario tenga permiso para ejecutar una acción del flujo.
        
        ISS-DIRECTOR FIX: Usa _get_rol_efectivo para consistencia con frontend.
        Antes usaba user.rol directamente, lo que fallaba si el campo estaba vacío.
        
        Args:
            user: Usuario que intenta ejecutar la acción
            accion: Clave del permiso (ej: 'puede_autorizar_admin')
            
        Returns:
            bool: True si tiene permiso
        """
        if user.is_superuser:
            return True
        
        # ISS-DIRECTOR FIX: Usar rol efectivo en lugar de user.rol directo
        # Esto garantiza que si el campo rol está vacío, se infiera correctamente
        rol = _get_rol_efectivo(user)
        permisos_rol = PERMISOS_FLUJO_REQUISICION.get(rol, {})
        
        tiene_permiso = permisos_rol.get(accion, False)
        
        # Log para debugging si falla
        if not tiene_permiso:
            logger.debug(
                f"_validar_permiso_flujo: Usuario {user.username} (rol={rol}, "
                f"rol_campo={getattr(user, 'rol', 'vacío')}) no tiene permiso '{accion}'"
            )
        
        return tiene_permiso

    @action(detail=True, methods=['post'], url_path='enviar-admin')
    @transaction.atomic
    def enviar_admin(self, request, pk=None):
        """
        FLUJO V2: Médico envía requisición al Administrador del Centro.
        
        Transición: borrador → pendiente_admin
        Permiso requerido: puede_enviar_admin (rol: medico)
        """
        try:
            requisicion = self.get_object()
            estado_actual = (requisicion.estado or '').lower()
            
            # Validar estado
            if estado_actual != 'borrador':
                return Response({
                    'error': 'Solo se pueden enviar a administrador las requisiciones en BORRADOR',
                    'estado_actual': requisicion.estado
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Validar transición
            if not self._validar_transicion(estado_actual, 'pendiente_admin'):
                return Response({
                    'error': 'Transición de estado no permitida',
                    'estado_actual': estado_actual,
                    'estado_destino': 'pendiente_admin'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Validar permiso
            if not self._validar_permiso_flujo(request.user, 'puede_enviar_admin'):
                # ISS-DIRECTOR FIX: Mostrar rol efectivo en mensaje de error
                rol_efectivo = _get_rol_efectivo(request.user)
                return Response({
                    'error': 'No tiene permiso para enviar requisiciones al administrador',
                    'rol_actual': rol_efectivo,
                    'detalle': f'El rol "{rol_efectivo}" no tiene el permiso puede_enviar_admin'
                }, status=status.HTTP_403_FORBIDDEN)
            
            # Validar que tenga detalles
            if not requisicion.detalles.exists():
                return Response({
                    'error': 'La requisición debe tener al menos un producto'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Validar centro - ISS-FIX-CENTRO: usar centro_origen (el centro que solicita)
            # FALLBACK: si centro_origen es NULL (datos viejos), usar centro_destino
            centro_user = self._user_centro(request.user)
            requisicion_centro_id = requisicion.centro_origen_id or requisicion.centro_destino_id
            if not request.user.is_superuser and centro_user:
                if requisicion_centro_id != centro_user.id:
                    return Response({
                        'error': 'No puede enviar requisiciones de otro centro'
                    }, status=status.HTTP_403_FORBIDDEN)
            
            # Ejecutar transición
            estado_anterior = requisicion.estado
            requisicion.estado = 'pendiente_admin'
            requisicion.fecha_envio_admin = timezone.now()
            requisicion.save(update_fields=['estado', 'fecha_envio_admin', 'updated_at'])
            
            # Registrar en historial (no crítico, no debe fallar la transición)
            try:
                self._registrar_historial(
                    requisicion, estado_anterior, 'pendiente_admin',
                    request.user, 'enviar_a_administrador', request,
                    datos_adicionales={'centro_id': centro_user.id if centro_user else None}
                )
            except Exception as hist_err:
                logger.warning(f"Error registrando historial (no crítico): {hist_err}")
            
            return Response({
                'mensaje': 'Requisición enviada al administrador',
                'requisicion': RequisicionSerializer(requisicion, context={'request': request}).data
            })
        except Exception as e:
            import traceback
            logger.error(f"[enviar_admin] Error: {str(e)}")
            logger.error(f"[enviar_admin] Traceback: {traceback.format_exc()}")
            return Response({
                'error': 'Error al enviar requisición',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['post'], url_path='autorizar-admin')
    @transaction.atomic
    def autorizar_admin(self, request, pk=None):
        """
        FLUJO V2: Administrador del Centro autoriza la requisición.
        
        Transición: pendiente_admin → pendiente_director
        Permiso requerido: puede_autorizar_admin (rol: administrador_centro)
        """
        requisicion = self.get_object()
        estado_actual = (requisicion.estado or '').lower()
        
        if estado_actual != 'pendiente_admin':
            return Response({
                'error': 'Solo se pueden autorizar requisiciones en PENDIENTE_ADMIN',
                'estado_actual': requisicion.estado
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if not self._validar_transicion(estado_actual, 'pendiente_director'):
            return Response({
                'error': 'Transición de estado no permitida'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if not self._validar_permiso_flujo(request.user, 'puede_autorizar_admin'):
            # ISS-DIRECTOR FIX: Mostrar rol efectivo en mensaje de error
            rol_efectivo = _get_rol_efectivo(request.user)
            return Response({
                'error': 'No tiene permiso para autorizar como administrador del centro',
                'rol_actual': rol_efectivo,
                'detalle': f'El rol "{rol_efectivo}" no tiene el permiso puede_autorizar_admin'
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Validar centro - ISS-FIX-CENTRO: usar centro_origen (el centro que solicita)
        # FALLBACK: si centro_origen es NULL (datos viejos), usar centro_destino
        centro_user = self._user_centro(request.user)
        requisicion_centro_id = requisicion.centro_origen_id or requisicion.centro_destino_id
        if not request.user.is_superuser and centro_user:
            if requisicion_centro_id != centro_user.id:
                return Response({
                    'error': 'No puede autorizar requisiciones de otro centro'
                }, status=status.HTTP_403_FORBIDDEN)
        
        observaciones = request.data.get('observaciones', '')
        
        estado_anterior = requisicion.estado
        requisicion.estado = 'pendiente_director'
        requisicion.fecha_autorizacion_admin = timezone.now()
        requisicion.administrador_centro = request.user
        if observaciones:
            requisicion.notas = f"{requisicion.notas or ''}\n[Admin] {observaciones}".strip()
        requisicion.save(update_fields=['estado', 'fecha_autorizacion_admin', 'administrador_centro', 'notas'])
        
        self._registrar_historial(
            requisicion, estado_anterior, 'pendiente_director',
            request.user, 'autorizar_administrador', request,
            datos_adicionales={'observaciones': observaciones}
        )
        
        return Response({
            'mensaje': 'Requisición autorizada por administrador, pendiente de director',
            'requisicion': RequisicionSerializer(requisicion, context={'request': request}).data
        })

    @action(detail=True, methods=['post'], url_path='autorizar-director')
    @transaction.atomic
    def autorizar_director(self, request, pk=None):
        """
        FLUJO V2: Director del Centro autoriza la requisición.
        
        Transición: pendiente_director → enviada (a farmacia central)
        Permiso requerido: puede_autorizar_director (rol: director_centro)
        """
        requisicion = self.get_object()
        estado_actual = (requisicion.estado or '').lower()
        
        if estado_actual != 'pendiente_director':
            return Response({
                'error': 'Solo se pueden autorizar requisiciones en PENDIENTE_DIRECTOR',
                'estado_actual': requisicion.estado
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if not self._validar_transicion(estado_actual, 'enviada'):
            return Response({
                'error': 'Transición de estado no permitida'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if not self._validar_permiso_flujo(request.user, 'puede_autorizar_director'):
            # ISS-DIRECTOR FIX: Mostrar rol efectivo en mensaje de error
            rol_efectivo = _get_rol_efectivo(request.user)
            return Response({
                'error': 'No tiene permiso para autorizar como director del centro',
                'rol_actual': rol_efectivo,
                'rol_campo': getattr(request.user, 'rol', '') or '(vacío)',
                'detalle': f'El rol "{rol_efectivo}" no tiene el permiso puede_autorizar_director'
            }, status=status.HTTP_403_FORBIDDEN)
        
        # ISS-FIX-CENTRO: usar centro_origen (el centro que solicita)
        # FALLBACK: si centro_origen es NULL (datos viejos), usar centro_destino
        centro_user = self._user_centro(request.user)
        requisicion_centro_id = requisicion.centro_origen_id or requisicion.centro_destino_id
        if not request.user.is_superuser and centro_user:
            if requisicion_centro_id != centro_user.id:
                return Response({
                    'error': 'No puede autorizar requisiciones de otro centro'
                }, status=status.HTTP_403_FORBIDDEN)
        
        observaciones = request.data.get('observaciones', '')
        
        estado_anterior = requisicion.estado
        requisicion.estado = 'enviada'
        requisicion.fecha_autorizacion_director = timezone.now()
        requisicion.fecha_envio_farmacia = timezone.now()
        requisicion.director_centro = request.user
        if observaciones:
            requisicion.notas = f"{requisicion.notas or ''}\n[Director] {observaciones}".strip()
        requisicion.save(update_fields=[
            'estado', 'fecha_autorizacion_director', 'fecha_envio_farmacia', 
            'director_centro', 'notas'
        ])
        
        self._registrar_historial(
            requisicion, estado_anterior, 'enviada',
            request.user, 'autorizar_director', request,
            datos_adicionales={'observaciones': observaciones}
        )
        
        return Response({
            'mensaje': 'Requisición autorizada por director, enviada a farmacia central',
            'requisicion': RequisicionSerializer(requisicion, context={'request': request}).data
        })

    @action(detail=True, methods=['post'], url_path='recibir-farmacia')
    @transaction.atomic
    def recibir_farmacia(self, request, pk=None):
        """
        FLUJO V2: Farmacia Central recibe la requisición para revisión.
        
        Transición: enviada → en_revision
        Permiso requerido: puede_recibir_farmacia (rol: farmacia)
        """
        requisicion = self.get_object()
        estado_actual = (requisicion.estado or '').lower()
        
        if estado_actual != 'enviada':
            return Response({
                'error': 'Solo se pueden recibir requisiciones en estado ENVIADA',
                'estado_actual': requisicion.estado
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if not self._validar_transicion(estado_actual, 'en_revision'):
            return Response({
                'error': 'Transición de estado no permitida'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if not self._validar_permiso_flujo(request.user, 'puede_recibir_farmacia'):
            if not is_farmacia_or_admin(request.user):
                # ISS-DIRECTOR FIX: Mostrar rol efectivo en mensaje de error
                rol_efectivo = _get_rol_efectivo(request.user)
                return Response({
                    'error': 'Solo personal de farmacia puede recibir requisiciones',
                    'rol_actual': rol_efectivo,
                    'detalle': f'El rol "{rol_efectivo}" no tiene el permiso puede_recibir_farmacia'
                }, status=status.HTTP_403_FORBIDDEN)
        
        observaciones = request.data.get('observaciones', '')
        
        estado_anterior = requisicion.estado
        requisicion.estado = 'en_revision'
        requisicion.fecha_recepcion_farmacia = timezone.now()
        requisicion.receptor_farmacia = request.user
        if observaciones:
            requisicion.observaciones_farmacia = observaciones
        requisicion.save(update_fields=[
            'estado', 'fecha_recepcion_farmacia', 'receptor_farmacia', 'observaciones_farmacia'
        ])
        
        self._registrar_historial(
            requisicion, estado_anterior, 'en_revision',
            request.user, 'recibir_farmacia', request,
            datos_adicionales={'observaciones': observaciones}
        )
        
        return Response({
            'mensaje': 'Requisición recibida en farmacia, en revisión',
            'requisicion': RequisicionSerializer(requisicion, context={'request': request}).data
        })

    @action(detail=True, methods=['post'], url_path='autorizar-farmacia')
    @transaction.atomic
    def autorizar_farmacia(self, request, pk=None):
        """
        FLUJO V2: Farmacia Central autoriza la requisición y asigna fecha límite de recolección.
        
        Transición: en_revision → autorizada
        Permiso requerido: puede_autorizar_farmacia (rol: farmacia)
        
        FLUJO CORRECTO: enviada → recibir → en_revision → autorizar
        
        IMPORTANTE: Debe incluir 'fecha_recoleccion_limite' en el request.
        """
        requisicion = self.get_object()
        estado_actual = (requisicion.estado or '').lower()
        
        # ISS-FLUJO-FIX: Solo permitir autorizar desde 'en_revision'
        # El flujo correcto es: enviada → recibir → en_revision → autorizar
        if estado_actual != 'en_revision':
            return Response({
                'error': 'Solo se pueden autorizar requisiciones en EN_REVISION. Debe recibir la requisición primero.',
                'estado_actual': requisicion.estado,
                'flujo_correcto': 'enviada → recibir → en_revision → autorizar'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if not self._validar_transicion(estado_actual, 'autorizada'):
            return Response({
                'error': 'Transición de estado no permitida'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if not self._validar_permiso_flujo(request.user, 'puede_autorizar_farmacia'):
            if not is_farmacia_or_admin(request.user):
                # ISS-DIRECTOR FIX: Mostrar rol efectivo en mensaje de error
                rol_efectivo = _get_rol_efectivo(request.user)
                return Response({
                    'error': 'Solo personal de farmacia puede autorizar requisiciones',
                    'rol_actual': rol_efectivo,
                    'detalle': f'El rol "{rol_efectivo}" no tiene el permiso puede_autorizar_farmacia'
                }, status=status.HTTP_403_FORBIDDEN)
        
        # CRÍTICO: Validar fecha límite de recolección
        fecha_recoleccion_str = request.data.get('fecha_recoleccion_limite')
        if not fecha_recoleccion_str:
            return Response({
                'error': 'Debe especificar la fecha límite de recolección',
                'campo_requerido': 'fecha_recoleccion_limite'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            from django.utils.dateparse import parse_datetime, parse_date
            fecha_recoleccion = parse_datetime(fecha_recoleccion_str)
            if not fecha_recoleccion:
                fecha_date = parse_date(fecha_recoleccion_str)
                if fecha_date:
                    # Si solo es fecha, agregar hora 17:00
                    from datetime import time
                    fecha_recoleccion = timezone.make_aware(
                        datetime.combine(fecha_date, time(17, 0, 0))
                    )
            # ISS-FIX: Asegurar que la fecha siempre sea timezone-aware
            elif timezone.is_naive(fecha_recoleccion):
                fecha_recoleccion = timezone.make_aware(fecha_recoleccion)
        except Exception:
            return Response({
                'error': 'Formato de fecha inválido. Use YYYY-MM-DD o YYYY-MM-DDTHH:MM:SS'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if not fecha_recoleccion:
            return Response({
                'error': 'No se pudo parsear la fecha límite de recolección'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Validar que la fecha sea futura
        if fecha_recoleccion <= timezone.now():
            return Response({
                'error': 'La fecha límite de recolección debe ser futura'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        observaciones = request.data.get('observaciones', '')
        
        # Procesar ajustes de cantidades si se envían
        items_data = request.data.get('items') or request.data.get('detalles') or []
        items_procesados = set()
        
        for item_data in items_data:
            item_id = item_data.get('id')
            cant_autorizada = item_data.get('cantidad_autorizada')
            if item_id is None or cant_autorizada is None:
                continue
            try:
                item = requisicion.detalles.get(id=item_id)
                item.cantidad_autorizada = max(0, int(cant_autorizada))
                motivo_ajuste = item_data.get('motivo_ajuste', '')
                if item.cantidad_autorizada < item.cantidad_solicitada and not motivo_ajuste:
                    return Response({
                        'error': f'Debe indicar motivo de ajuste para {item.producto.clave}',
                        'producto': item.producto.clave
                    }, status=status.HTTP_400_BAD_REQUEST)
                item.motivo_ajuste = motivo_ajuste
                item.save()
                items_procesados.add(item_id)
            except DetalleRequisicion.DoesNotExist:
                continue
        
        # ISS-SURTIR-FIX: Si NO se enviaron items explícitos, autorizar TODOS los detalles
        # con cantidad_autorizada = cantidad_solicitada (aprobación total)
        if not items_procesados:
            logger.info(f"autorizar_farmacia: No se enviaron items, autorizando todos los detalles con cantidad solicitada")
            for detalle in requisicion.detalles.all():
                if detalle.cantidad_autorizada is None or detalle.cantidad_autorizada == 0:
                    detalle.cantidad_autorizada = detalle.cantidad_solicitada
                    detalle.save()
                    logger.info(f"  - Detalle {detalle.id}: autorizado {detalle.cantidad_autorizada} unidades de {detalle.producto.clave}")
        
        estado_anterior = requisicion.estado
        requisicion.estado = 'autorizada'
        
        # ISS-FIX: Asignar campos directamente sin try/except individual
        # Los campos existen en el modelo y en la BD según el schema
        requisicion.fecha_autorizacion_farmacia = timezone.now()
        requisicion.fecha_autorizacion = timezone.now()  # Campo legacy
        requisicion.fecha_recoleccion_limite = fecha_recoleccion
        requisicion.autorizador_farmacia = request.user
        requisicion.autorizador = request.user  # Campo legacy
        
        if observaciones:
            requisicion.observaciones_farmacia = f"{requisicion.observaciones_farmacia or ''}\n{observaciones}".strip()
        
        # ISS-FIX: Guardar SIN update_fields para evitar problemas con managed=False
        # Django manejará qué campos actualizar basándose en los cambios detectados
        try:
            requisicion.save()
            logger.info(f"Requisición {requisicion.folio} guardada exitosamente con fecha_recoleccion_limite={fecha_recoleccion}")
        except Exception as save_error:
            logger.error(f"Error al guardar requisición autorizada: {save_error}")
            # Si falla, intentar guardar con update() directo incluyendo TODOS los campos importantes
            try:
                Requisicion.objects.filter(pk=requisicion.pk).update(
                    estado='autorizada',
                    fecha_autorizacion=timezone.now(),
                    fecha_autorizacion_farmacia=timezone.now(),
                    fecha_recoleccion_limite=fecha_recoleccion,
                    autorizador_farmacia=request.user,
                    autorizador=request.user
                )
                requisicion.refresh_from_db()
                logger.info(f"Guardado exitoso via update() directo con fecha_recoleccion_limite={fecha_recoleccion}")
            except Exception as fallback_error:
                logger.error(f"Error en fallback update: {fallback_error}")
                raise
        
        self._registrar_historial(
            requisicion, estado_anterior, 'autorizada',
            request.user, 'autorizar_farmacia', request,
            datos_adicionales={
                'fecha_recoleccion_limite': fecha_recoleccion.isoformat(),
                'observaciones': observaciones
            }
        )
        
        return Response({
            'mensaje': f'Requisición autorizada. Fecha límite de recolección: {fecha_recoleccion.strftime("%Y-%m-%d %H:%M")}',
            'requisicion': RequisicionSerializer(requisicion, context={'request': request}).data,
            'fecha_recoleccion_limite': fecha_recoleccion.isoformat()
        })

    @action(detail=True, methods=['post'], url_path='devolver')
    @transaction.atomic
    def devolver(self, request, pk=None):
        """
        FLUJO V2: Devuelve una requisición al centro para correcciones.
        
        Transiciones permitidas:
        - pendiente_admin → devuelta (por administrador)
        - pendiente_director → devuelta (por director)
        - en_revision → devuelta (por farmacia)
        
        Requiere motivo obligatorio.
        """
        requisicion = self.get_object()
        estado_actual = (requisicion.estado or '').lower()
        
        estados_devolvibles = ['pendiente_admin', 'pendiente_director', 'en_revision']
        if estado_actual not in estados_devolvibles:
            return Response({
                'error': f'Solo se pueden devolver requisiciones en: {", ".join(estados_devolvibles)}',
                'estado_actual': requisicion.estado
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if not self._validar_transicion(estado_actual, 'devuelta'):
            return Response({
                'error': 'Transición de estado no permitida'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        motivo = request.data.get('motivo') or request.data.get('observaciones', '')
        if not motivo or len(motivo.strip()) < 10:
            return Response({
                'error': 'Debe proporcionar un motivo de devolución (mínimo 10 caracteres)'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Validar permisos según estado
        tiene_permiso = False
        if estado_actual == 'pendiente_admin':
            tiene_permiso = self._validar_permiso_flujo(request.user, 'puede_autorizar_admin')
        elif estado_actual == 'pendiente_director':
            tiene_permiso = self._validar_permiso_flujo(request.user, 'puede_autorizar_director')
        elif estado_actual == 'en_revision':
            tiene_permiso = is_farmacia_or_admin(request.user)
        
        if not request.user.is_superuser and not tiene_permiso:
            return Response({
                'error': 'No tiene permiso para devolver esta requisición',
                'estado_actual': estado_actual
            }, status=status.HTTP_403_FORBIDDEN)
        
        estado_anterior = requisicion.estado
        requisicion.estado = 'devuelta'
        requisicion.motivo_devolucion = motivo.strip()
        requisicion.save(update_fields=['estado', 'motivo_devolucion'])
        
        self._registrar_historial(
            requisicion, estado_anterior, 'devuelta',
            request.user, 'devolver_centro', request,
            motivo=motivo.strip()
        )
        
        return Response({
            'mensaje': 'Requisición devuelta al médico para correcciones',
            'requisicion': RequisicionSerializer(requisicion, context={'request': request}).data,
            'motivo_devolucion': motivo.strip()
        })

    @action(detail=True, methods=['post'], url_path='reenviar')
    @transaction.atomic
    def reenviar(self, request, pk=None):
        """
        FLUJO V2: Reenvía una requisición devuelta al proceso de autorización.
        
        Transición: devuelta → pendiente_admin
        Permiso: médico del centro (creador original)
        """
        requisicion = self.get_object()
        estado_actual = (requisicion.estado or '').lower()
        
        if estado_actual != 'devuelta':
            return Response({
                'error': 'Solo se pueden reenviar requisiciones en estado DEVUELTA',
                'estado_actual': requisicion.estado
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if not self._validar_transicion(estado_actual, 'pendiente_admin'):
            return Response({
                'error': 'Transición de estado no permitida'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Validar que sea el solicitante original o admin
        # ISS-FIX-CENTRO: usar centro_origen (el centro que solicita)
        # FALLBACK: si centro_origen es NULL (datos viejos), usar centro_destino
        if not request.user.is_superuser:
            if requisicion.solicitante_id != request.user.id:
                centro_user = self._user_centro(request.user)
                requisicion_centro_id = requisicion.centro_origen_id or requisicion.centro_destino_id
                if not centro_user or requisicion_centro_id != centro_user.id:
                    return Response({
                        'error': 'Solo el solicitante original puede reenviar la requisición'
                    }, status=status.HTTP_403_FORBIDDEN)
        
        observaciones = request.data.get('observaciones', '')
        
        estado_anterior = requisicion.estado
        requisicion.estado = 'pendiente_admin'
        requisicion.fecha_envio_admin = timezone.now()
        requisicion.motivo_devolucion = None  # Limpiar motivo anterior
        if observaciones:
            requisicion.notas = f"{requisicion.notas or ''}\n[Reenvío] {observaciones}".strip()
        requisicion.save(update_fields=['estado', 'fecha_envio_admin', 'motivo_devolucion', 'notas'])
        
        self._registrar_historial(
            requisicion, estado_anterior, 'pendiente_admin',
            request.user, 'enviar_a_administrador', request,
            datos_adicionales={'es_reenvio': True, 'observaciones': observaciones}
        )
        
        return Response({
            'mensaje': 'Requisición reenviada para autorización',
            'requisicion': RequisicionSerializer(requisicion, context={'request': request}).data
        })

    @action(detail=True, methods=['post'], url_path='confirmar-entrega')
    @transaction.atomic
    def confirmar_entrega(self, request, pk=None):
        """
        FLUJO V2: Farmacia confirma la entrega de los medicamentos al centro.
        
        Transición: surtida → entregada
        Permiso: Solo FARMACIA puede confirmar (ellos entregan físicamente)
        
        IMPORTANTE: Debe confirmar antes de fecha_recoleccion_limite
        """
        requisicion = self.get_object()
        estado_actual = (requisicion.estado or '').lower()
        
        if estado_actual != 'surtida':
            return Response({
                'error': 'Solo se pueden confirmar entregas de requisiciones SURTIDAS',
                'estado_actual': requisicion.estado
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if not self._validar_transicion(estado_actual, 'entregada'):
            return Response({
                'error': 'Transición de estado no permitida'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # ISS-FIX: Solo FARMACIA puede confirmar entrega (ellos entregan físicamente al centro)
        user_rol = getattr(request.user, 'rol', '') or ''
        roles_permitidos = ['farmacia', 'admin', 'admin_sistema', 'superusuario']
        if user_rol.lower() not in roles_permitidos and not request.user.is_superuser:
            return Response({
                'error': 'Solo el personal de Farmacia puede confirmar la entrega de medicamentos',
                'detalle': 'El centro recibirá los medicamentos cuando Farmacia confirme la entrega'
            }, status=status.HTTP_403_FORBIDDEN)
        
        # Verificar fecha límite
        if requisicion.fecha_recoleccion_limite:
            if timezone.now() > requisicion.fecha_recoleccion_limite:
                return Response({
                    'error': 'La fecha límite de recolección ha expirado. La requisición será marcada como vencida.',
                    'fecha_limite': requisicion.fecha_recoleccion_limite.isoformat()
                }, status=status.HTTP_400_BAD_REQUEST)
        
        lugar_entrega = request.data.get('lugar_entrega', '')
        observaciones = request.data.get('observaciones', '')
        
        if not lugar_entrega:
            return Response({
                'error': 'Debe especificar el lugar de entrega'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        estado_anterior = requisicion.estado
        requisicion.estado = 'entregada'
        requisicion.fecha_entrega = timezone.now()
        requisicion.lugar_entrega = lugar_entrega
        if observaciones:
            requisicion.notas = f"{requisicion.notas or ''}\n[Recepción] {observaciones}".strip()
        
        # ISS-004 FIX (audit21): Validar imagen completa (MIME, magic bytes, extensión)
        foto_firma = request.FILES.get('foto_firma_recepcion') or request.FILES.get('foto_firma')
        update_fields = ['estado', 'fecha_entrega', 'lugar_entrega', 'notas']
        
        if foto_firma:
            es_valido, error_msg = validar_archivo_imagen(foto_firma, max_size_mb=2)
            if es_valido:
                requisicion.foto_firma_recepcion = foto_firma
                requisicion.fecha_firma_recepcion = timezone.now()
                requisicion.usuario_firma_recepcion = request.user
                update_fields.extend(['foto_firma_recepcion', 'fecha_firma_recepcion', 'usuario_firma_recepcion'])
            else:
                logger.warning(f"ISS-004: Firma rechazada en confirmar_entrega {requisicion.folio}: {error_msg}")
        
        requisicion.save(update_fields=update_fields)
        
        self._registrar_historial(
            requisicion, estado_anterior, 'entregada',
            request.user, 'confirmar_entrega', request,
            datos_adicionales={
                'lugar_entrega': lugar_entrega,
                'observaciones': observaciones
            }
        )
        
        return Response({
            'mensaje': 'Entrega confirmada exitosamente',
            'requisicion': RequisicionSerializer(requisicion, context={'request': request}).data
        })

    @action(detail=True, methods=['post'], url_path='marcar-vencida')
    @transaction.atomic
    def marcar_vencida(self, request, pk=None):
        """
        FLUJO V2: Marca una requisición como vencida (manualmente por admin).
        
        Transición: surtida → vencida
        Permiso: Solo admin/superuser
        
        NOTA: Normalmente esto lo hace un cron automáticamente.
        """
        requisicion = self.get_object()
        estado_actual = (requisicion.estado or '').lower()
        
        if estado_actual != 'surtida':
            return Response({
                'error': 'Solo se pueden marcar como vencidas requisiciones SURTIDAS',
                'estado_actual': requisicion.estado
            }, status=status.HTTP_400_BAD_REQUEST)
        
        if not request.user.is_superuser and not is_farmacia_or_admin(request.user):
            return Response({
                'error': 'Solo administradores pueden marcar requisiciones como vencidas'
            }, status=status.HTTP_403_FORBIDDEN)
        
        motivo = request.data.get('motivo', 'Vencimiento manual por administrador')
        
        estado_anterior = requisicion.estado
        requisicion.estado = 'vencida'
        requisicion.fecha_vencimiento = timezone.now()
        requisicion.motivo_vencimiento = motivo
        requisicion.save(update_fields=['estado', 'fecha_vencimiento', 'motivo_vencimiento'])
        
        self._registrar_historial(
            requisicion, estado_anterior, 'vencida',
            request.user, 'vencer', request,
            motivo=motivo
        )
        
        return Response({
            'mensaje': 'Requisición marcada como vencida',
            'requisicion': RequisicionSerializer(requisicion, context={'request': request}).data
        })

    @action(detail=True, methods=['get'], url_path='historial')
    def historial_estados(self, request, pk=None):
        """
        FLUJO V2: Obtiene el historial de cambios de estado de una requisición.
        
        Retorna lista ordenada de todos los cambios para auditoría.
        """
        from core.models import RequisicionHistorialEstados
        from core.serializers import RequisicionHistorialEstadosSerializer
        
        requisicion = self.get_object()
        
        # Validar acceso - ISS-FIX-CENTRO: usar centro_origen (el centro que solicita)
        # FALLBACK: si centro_origen es NULL (datos viejos), usar centro_destino
        if not request.user.is_superuser and not is_farmacia_or_admin(request.user):
            centro_user = self._user_centro(request.user)
            requisicion_centro_id = requisicion.centro_origen_id or requisicion.centro_destino_id
            if not centro_user or requisicion_centro_id != centro_user.id:
                return Response({
                    'error': 'No tiene acceso al historial de esta requisición'
                }, status=status.HTTP_403_FORBIDDEN)
        
        historial = RequisicionHistorialEstados.objects.filter(
            requisicion=requisicion
        ).select_related('usuario').order_by('fecha_cambio')
        
        # Usar el serializer formal para consistencia
        serializer = RequisicionHistorialEstadosSerializer(historial, many=True)
        
        return Response({
            'requisicion_id': requisicion.id,
            'folio': requisicion.folio,
            'estado_actual': requisicion.estado,
            'total_cambios': historial.count(),
            'historial': serializer.data
        })

    @action(detail=False, methods=['post'], url_path='verificar-vencidas')
    def verificar_vencidas(self, request):
        """
        FLUJO V2: Verifica y marca como vencidas las requisiciones que superaron 
        su fecha límite de recolección.
        
        Este endpoint simula lo que haría el cron diario.
        Solo disponible para admin/superuser.
        """
        if not request.user.is_superuser and not is_farmacia_or_admin(request.user):
            return Response({
                'error': 'Solo administradores pueden ejecutar esta operación'
            }, status=status.HTTP_403_FORBIDDEN)
        
        from core.models import RequisicionHistorialEstados
        
        ahora = timezone.now()
        requisiciones_vencidas = Requisicion.objects.filter(
            estado='surtida',
            fecha_recoleccion_limite__lt=ahora
        )
        
        vencidas = []
        for req in requisiciones_vencidas:
            estado_anterior = req.estado
            req.estado = 'vencida'
            req.fecha_vencimiento = ahora
            req.motivo_vencimiento = f'Vencimiento automático. Fecha límite: {req.fecha_recoleccion_limite.strftime("%Y-%m-%d %H:%M")}'
            req.save(update_fields=['estado', 'fecha_vencimiento', 'motivo_vencimiento'])
            
            # Registrar en historial
            try:
                RequisicionHistorialEstados.objects.create(
                    requisicion=req,
                    estado_anterior=estado_anterior,
                    estado_nuevo='vencida',
                    usuario=request.user,
                    accion='vencer_automatico',
                    motivo=req.motivo_vencimiento,
                    ip_address=self._get_client_ip(request)
                )
            except Exception as e:
                logger.warning(f"No se pudo registrar historial para REQ-{req.numero}: {e}")
            
            vencidas.append({
                'id': req.id,
                'folio': req.folio,
                'centro': req.centro_destino.nombre if req.centro_destino else None,
                'fecha_limite': req.fecha_recoleccion_limite.isoformat()
            })
        
        return Response({
            'mensaje': f'Se marcaron {len(vencidas)} requisiciones como vencidas',
            'total_vencidas': len(vencidas),
            'requisiciones': vencidas
        })

    @action(detail=False, methods=['get'], url_path='transiciones-disponibles')
    def transiciones_disponibles(self, request):
        """
        FLUJO V2: Retorna las transiciones de estado disponibles según el rol del usuario.
        
        Útil para que el frontend sepa qué acciones mostrar.
        """
        user = request.user
        rol = (getattr(user, 'rol', '') or '').lower()
        permisos_rol = PERMISOS_FLUJO_REQUISICION.get(rol, {})
        
        if user.is_superuser:
            permisos_rol = {k: True for k in [
                'puede_crear', 'puede_enviar_admin', 'puede_autorizar_admin',
                'puede_autorizar_director', 'puede_recibir_farmacia',
                'puede_autorizar_farmacia', 'puede_surtir', 'puede_confirmar_entrega'
            ]}
        
        acciones_disponibles = []
        
        if permisos_rol.get('puede_crear'):
            acciones_disponibles.append({
                'accion': 'crear',
                'endpoint': '/api/requisiciones/',
                'metodo': 'POST',
                'descripcion': 'Crear nueva requisición'
            })
        
        if permisos_rol.get('puede_enviar_admin'):
            acciones_disponibles.append({
                'accion': 'enviar_admin',
                'endpoint': '/api/requisiciones/{id}/enviar-admin/',
                'metodo': 'POST',
                'estado_requerido': 'borrador',
                'estado_resultante': 'pendiente_admin',
                'descripcion': 'Enviar a administrador del centro'
            })
        
        if permisos_rol.get('puede_autorizar_admin'):
            acciones_disponibles.append({
                'accion': 'autorizar_admin',
                'endpoint': '/api/requisiciones/{id}/autorizar-admin/',
                'metodo': 'POST',
                'estado_requerido': 'pendiente_admin',
                'estado_resultante': 'pendiente_director',
                'descripcion': 'Autorizar como administrador'
            })
        
        if permisos_rol.get('puede_autorizar_director'):
            acciones_disponibles.append({
                'accion': 'autorizar_director',
                'endpoint': '/api/requisiciones/{id}/autorizar-director/',
                'metodo': 'POST',
                'estado_requerido': 'pendiente_director',
                'estado_resultante': 'enviada',
                'descripcion': 'Autorizar como director'
            })
        
        if permisos_rol.get('puede_recibir_farmacia'):
            acciones_disponibles.append({
                'accion': 'recibir_farmacia',
                'endpoint': '/api/requisiciones/{id}/recibir-farmacia/',
                'metodo': 'POST',
                'estado_requerido': 'enviada',
                'estado_resultante': 'en_revision',
                'descripcion': 'Recibir en farmacia para revisión'
            })
        
        if permisos_rol.get('puede_autorizar_farmacia'):
            acciones_disponibles.append({
                'accion': 'autorizar_farmacia',
                'endpoint': '/api/requisiciones/{id}/autorizar-farmacia/',
                'metodo': 'POST',
                'estado_requerido': ['en_revision', 'enviada'],
                'estado_resultante': 'autorizada',
                'descripcion': 'Autorizar y asignar fecha de recolección',
                'campos_requeridos': ['fecha_recoleccion_limite']
            })
        
        if permisos_rol.get('puede_surtir'):
            acciones_disponibles.append({
                'accion': 'surtir',
                'endpoint': '/api/requisiciones/{id}/surtir/',
                'metodo': 'POST',
                'estado_requerido': 'autorizada',
                'estado_resultante': 'surtida',
                'descripcion': 'Surtir requisición'
            })
        
        if permisos_rol.get('puede_confirmar_entrega'):
            acciones_disponibles.append({
                'accion': 'confirmar_entrega',
                'endpoint': '/api/requisiciones/{id}/confirmar-entrega/',
                'metodo': 'POST',
                'estado_requerido': 'surtida',
                'estado_resultante': 'entregada',
                'descripcion': 'Confirmar recepción de medicamentos',
                'campos_requeridos': ['lugar_entrega']
            })
        
        return Response({
            'usuario': user.username,
            'rol': rol,
            'es_superuser': user.is_superuser,
            'es_farmacia': is_farmacia_or_admin(user),
            'acciones_disponibles': acciones_disponibles,
            'transiciones': TRANSICIONES_REQUISICION
        })

    @action(detail=False, methods=['get'])
    def estadisticas(self, request):
        """Estadsticas de requisiciones filtradas por centro si aplica."""
        try:
            # SEGURIDAD: Filtrar por centro del usuario si no es admin/farmacia
            user = request.user
            base_queryset = Requisicion.objects.all()
            
            if not is_farmacia_or_admin(user):
                user_centro = get_user_centro(user)
                if user_centro:
                    base_queryset = base_queryset.filter(centro_destino=user_centro)
                else:
                    base_queryset = Requisicion.objects.none()
            
            # OPTIMIZACIÓN: Una sola consulta agregada para conteos por estado
            conteos_raw = base_queryset.values('estado').annotate(count=Count('id'))
            conteos_dict = {item['estado']: item['count'] for item in conteos_raw}
            total = sum(conteos_dict.values())
            por_estado = {estado: conteos_dict.get(estado, 0) for estado, _ in ESTADOS_REQUISICION}
            
            # Top centros solo para admin/farmacia
            por_centro = []
            if is_farmacia_or_admin(user):
                # Usar requisiciones_destino (related_name correcto del FK centro_destino)
                centros = Centro.objects.annotate(total_requisiciones=Count('requisiciones_destino')).filter(total_requisiciones__gt=0).order_by('-total_requisiciones')[:10]
                for centro in centros:
                    por_centro.append({'centro': centro.nombre, 'total': centro.total_requisiciones})
            
            return Response({'total': total, 'por_estado': por_estado, 'top_centros': por_centro})
        except Exception as exc:
            # traceback removido por seguridad (ISS-008)
            return Response({'error': 'Error al obtener estadisticas', 'mensaje': str(exc)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path='resumen_estados')
    def resumen_estados(self, request):
        """
        Resumen de conteos por estado y por grupo lógico.
        
        FIX: Ahora aplica TODOS los filtros del frontend para que los contadores
        de las tabs reflejen correctamente los datos filtrados.
        
        Filtros soportados: centro, estado, grupo_estado, search, fecha_desde, fecha_hasta
        """
        try:
            user = request.user
            base_queryset = Requisicion.objects.all()
            
            # SEGURIDAD: Filtrar por centro del usuario si no es admin/farmacia
            if not is_farmacia_or_admin(user):
                user_centro = get_user_centro(user)
                if user_centro:
                    base_queryset = base_queryset.filter(
                        Q(centro_origen=user_centro) | Q(centro_destino=user_centro)
                    )
                else:
                    base_queryset = Requisicion.objects.none()
            else:
                # Admin/farmacia pueden filtrar por centro específico
                centro_param = request.query_params.get('centro')
                if centro_param and centro_param not in ['', 'null', 'undefined', 'todos', 'PENDING']:
                    try:
                        centro_id = int(centro_param)
                        base_queryset = base_queryset.filter(
                            Q(centro_origen_id=centro_id) | Q(centro_destino_id=centro_id)
                        )
                    except (ValueError, TypeError):
                        pass
            
            # FIX: Aplicar filtro de búsqueda (search)
            # ISS-FIX-CENTRO: buscar en centro_origen (datos nuevos) y centro_destino (datos viejos)
            search = request.query_params.get('search', '').strip()
            if search:
                base_queryset = base_queryset.filter(
                    Q(numero__icontains=search) |
                    Q(solicitante__first_name__icontains=search) |
                    Q(solicitante__last_name__icontains=search) |
                    Q(centro_origen__nombre__icontains=search) |
                    Q(centro_destino__nombre__icontains=search)
                )
            
            # FIX: Aplicar filtro de fechas
            fecha_desde = request.query_params.get('fecha_desde')
            if fecha_desde:
                base_queryset = base_queryset.filter(fecha_solicitud__date__gte=fecha_desde)
            
            fecha_hasta = request.query_params.get('fecha_hasta')
            if fecha_hasta:
                base_queryset = base_queryset.filter(fecha_solicitud__date__lte=fecha_hasta)
            
            # OPTIMIZACIÓN: Una sola consulta agregada en lugar de N consultas individuales
            # Esto reduce de ~20 queries a 1 sola query
            conteos_raw = base_queryset.values('estado').annotate(count=Count('id'))
            conteos_dict = {item['estado']: item['count'] for item in conteos_raw}
            
            # Mapear a formato esperado por el frontend
            por_estado = {estado.upper(): conteos_dict.get(estado, 0) for estado, _ in ESTADOS_REQUISICION}
            
            # Calcular grupos a partir de los conteos ya obtenidos (sin queries adicionales)
            por_grupo = {}
            for nombre, estados in REQUISICION_GRUPOS_ESTADO.items():
                por_grupo[nombre] = sum(conteos_dict.get(e, 0) for e in estados)
            
            return Response({'por_estado': por_estado, 'por_grupo': por_grupo})
        except Exception as exc:
            # traceback removido por seguridad (ISS-008)
            return Response({'error': 'Error al obtener resumen de estados', 'mensaje': str(exc)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def dashboard_resumen(request):
    """
    ISS-005: Resumen del dashboard con KPIs y últimos movimientos.
    Implementa caché para mejorar rendimiento.
    
    SEGURIDAD:
    - Usuarios de centro: solo ven datos de su centro
    - Admin/farmacia/vista: ven datos globales por defecto
    - Admin/farmacia/vista pueden usar ?centro=ID para filtrar por centro específico
    
    PARÁMETROS:
    - centro: ID del centro para filtrar (opcional, solo admin/farmacia/vista)
    - refresh: Si es 'true', fuerza recarga sin usar caché (útil después de importar datos)
    """
    try:
        # SEGURIDAD: Filtrar por centro si el usuario no es admin/farmacia/vista
        user = request.user
        filtrar_por_centro = not is_farmacia_or_admin(user)
        user_centro = get_user_centro(user) if filtrar_por_centro else None
        
        # Admin/farmacia/vista puede filtrar por centro especfico
        centro_param = request.query_params.get('centro')
        if centro_param and centro_param not in ['', 'null', 'undefined', 'todos']:
            if is_farmacia_or_admin(user):
                try:
                    user_centro = Centro.objects.get(pk=int(centro_param))
                    filtrar_por_centro = True
                except (Centro.DoesNotExist, ValueError, TypeError):
                    pass
        
        # ISS-005: Generar clave de caché única por usuario/centro
        centro_id = user_centro.id if user_centro else 'global'
        cache_key = f'dashboard_resumen_{centro_id}'
        
        # ISS-FIX: Parámetro refresh para forzar recarga sin caché
        force_refresh = request.query_params.get('refresh', '').lower() == 'true'
        
        # ISS-005: Intentar obtener del caché (excepto si se fuerza refresh)
        cached_kpi = None if force_refresh else cache.get(cache_key)
        
        if cached_kpi is None:
            # === PRODUCTOS ===
            # ISS-FIX: Para usuarios de centro, contar solo productos que tienen lotes en SU centro
            if filtrar_por_centro and user_centro:
                # Contar productos distintos que tienen lotes activos con stock en el centro
                total_productos = Producto.objects.filter(
                    activo=True,
                    lotes__centro=user_centro,
                    lotes__activo=True,
                    lotes__cantidad_actual__gt=0
                ).distinct().count()
            else:
                total_productos = Producto.objects.filter(activo=True).count()
            
            # === LOTES ===
            lotes_query = Lote.objects.filter(
                activo=True,
                cantidad_actual__gt=0
            )
            
            if filtrar_por_centro and user_centro:
                lotes_query = lotes_query.filter(centro=user_centro)
            
            stock_total = lotes_query.aggregate(
                total=Coalesce(Sum('cantidad_actual'), 0, output_field=IntegerField())
            )['total']
            # ISS-FIX: Contar lotes ÚNICOS por numero_lote (no registros duplicados por centro)
            lotes_activos = lotes_query.values('numero_lote').distinct().count()

            # === MOVIMIENTOS DEL MES ===
            inicio_mes = timezone.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            
            movimientos_base = Movimiento.objects.all()
            if filtrar_por_centro and user_centro:
                # ISS-CENTRO FIX v2: Solo movimientos donde el lote pertenece al centro
                # o donde el centro es el origen (no incluir salidas de otros centros hacia este)
                movimientos_base = movimientos_base.filter(
                    Q(lote__centro=user_centro) | Q(centro_origen=user_centro)
                )
            
            # Contar movimientos del mes actual
            movimientos_mes = movimientos_base.filter(fecha__gte=inicio_mes).count()
            
            # Si no hay movimientos este mes, mostrar total general como referencia
            total_movimientos = movimientos_base.count() if movimientos_mes == 0 else movimientos_mes
            
            cached_kpi = {
                'total_productos': total_productos,
                'stock_total': max(0, stock_total),
                'lotes_activos': lotes_activos,
                'movimientos_mes': total_movimientos
            }
            
            # ISS-005: Guardar en caché con TTL configurable
            cache_ttl = getattr(settings, 'CACHE_TTL_DASHBOARD', 60)
            cache.set(cache_key, cached_kpi, cache_ttl)
        
        # === ÚLTIMOS MOVIMIENTOS (siempre frescos, no cacheados) ===
        movimientos_base = Movimiento.objects.all()
        if filtrar_por_centro and user_centro:
            # ISS-CENTRO FIX v2: Solo movimientos donde el lote pertenece al centro
            # o donde el centro es el origen (no incluir salidas de otros centros hacia este)
            movimientos_base = movimientos_base.filter(
                Q(lote__centro=user_centro) | Q(centro_origen=user_centro)
            )
        
        ultimos_movimientos = movimientos_base.select_related(
            'lote__producto', 'lote__centro', 'centro_origen', 'centro_destino', 'requisicion', 'usuario'
        ).order_by('-fecha')[:10]
        
        movimientos_data = []
        for mov in ultimos_movimientos:
            lote = mov.lote
            producto_desc = 'N/A'
            producto_clave = 'N/A'
            lote_numero = 'N/A'
            
            if lote:
                lote_numero = lote.numero_lote or 'N/A'
                if lote.producto:
                    # ISS-FIX: Usar nombre (NOT NULL) como principal, descripcion como fallback
                    producto_desc = lote.producto.nombre or lote.producto.descripcion or 'N/A'
                    producto_clave = lote.producto.clave or 'N/A'
            
            # Determinar origen/destino
            lote_centro = lote.centro if lote else None
            mov_centro = mov.centro_destino
            
            if mov.tipo == 'entrada':
                origen = mov.referencia or 'Proveedor'
                destino = lote_centro.nombre if lote_centro else 'Farmacia Central'
            else:
                origen = lote_centro.nombre if lote_centro else 'Farmacia Central'
                destino = mov_centro.nombre if mov_centro else 'Consumo/Ajuste'
            
            movimientos_data.append({
                'id': mov.id,
                'tipo_movimiento': mov.tipo.upper(),
                'producto__descripcion': producto_desc,
                'producto__clave': producto_clave,
                'lote__codigo_lote': lote_numero,
                'cantidad': abs(mov.cantidad),
                'fecha_movimiento': mov.fecha.isoformat() if mov.fecha else None,
                'observaciones': mov.motivo or '',
                'origen': origen,
                'destino': destino,
                'requisicion_folio': mov.requisicion.numero if mov.requisicion else None,
                'referencia': mov.referencia or None,
                'usuario': (mov.usuario.get_full_name() or mov.usuario.username) if mov.usuario else 'Sistema',
            })

        return Response({
            'kpi': cached_kpi,
            'ultimos_movimientos': movimientos_data
        })
        
    except Exception as exc:
        # traceback removido por seguridad (ISS-008)
        logger.exception('Error en dashboard_resumen')
        return Response({
            'kpi': {'total_productos': 0, 'stock_total': 0, 'lotes_activos': 0, 'movimientos_mes': 0},
            'ultimos_movimientos': [],
            'error': 'Error interno del servidor'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def dashboard_graficas(request):
    """
    ISS-005: Datos para gráficas del dashboard con caché.
    Retorna consumo_mensual, stock_por_centro y requisiciones_por_estado.
    
    SEGURIDAD: Filtra por centro del usuario si no es admin/farmacia.
    Admin/farmacia puede usar ?centro=ID para filtrar por centro específico.
    
    PARÁMETROS:
    - centro: ID del centro para filtrar (opcional, solo admin/farmacia/vista)
    - refresh: Si es 'true', fuerza recarga sin usar caché
    """
    try:
        from dateutil.relativedelta import relativedelta
        from django.db.models import Q  # ISS-FIX: Importar Q al inicio de la función
        
        # SEGURIDAD: Determinar filtro de centro
        user = request.user
        filtrar_por_centro = not is_farmacia_or_admin(user)
        user_centro = get_user_centro(user) if filtrar_por_centro else None
        
        # Admin/farmacia puede filtrar por centro específico
        centro_param = request.query_params.get('centro')
        if centro_param and centro_param not in ['', 'null', 'undefined', 'todos']:
            if is_farmacia_or_admin(user):
                try:
                    user_centro = Centro.objects.get(pk=int(centro_param))
                    filtrar_por_centro = True
                except (Centro.DoesNotExist, ValueError, TypeError):
                    pass
        
        # ISS-005: Generar clave de caché única por centro
        centro_id = user_centro.id if user_centro else 'global'
        cache_key = f'dashboard_graficas_{centro_id}'
        
        # ISS-FIX: Parámetro refresh para forzar recarga sin caché
        force_refresh = request.query_params.get('refresh', '').lower() == 'true'
        
        # ISS-005: Intentar obtener del caché (excepto si se fuerza refresh)
        cached_data = None if force_refresh else cache.get(cache_key)
        if cached_data is not None:
            return Response(cached_data)
        
        hoy = timezone.now().date()
        
        # =========================================
        # 1. CONSUMO MENSUAL (LTIMOS 6 MESES)
        # =========================================
        consumo_mensual = []
        
        for i in range(6):
            # Calcular el mes (hace 5, 4, 3, 2, 1, 0 meses)
            fecha_mes = hoy - relativedelta(months=5-i)
            mes_inicio = fecha_mes.replace(day=1)
            
            # Calcular fin del mes
            if fecha_mes.month == 12:
                mes_fin = fecha_mes.replace(year=fecha_mes.year + 1, month=1, day=1)
            else:
                mes_fin = fecha_mes.replace(month=fecha_mes.month + 1, day=1)
            
            # Base query
            mov_mes = Movimiento.objects.filter(
                fecha__date__gte=mes_inicio,
                fecha__date__lt=mes_fin
            )
            
            if filtrar_por_centro and user_centro:
                # ISS-CENTRO FIX v2: Solo movimientos donde el lote pertenece al centro
                # o donde el centro es el origen (no incluir salidas de otros centros hacia este)
                mov_mes = mov_mes.filter(
                    Q(lote__centro=user_centro) | Q(centro_origen=user_centro)
                )
            elif not filtrar_por_centro:
                # ISS-FIX: Farmacia Central sin filtro de centro específico
                # Mostrar solo movimientos de lotes de Farmacia Central (sin centro asignado)
                # o lotes de "Almacén Central"
                mov_mes = mov_mes.filter(
                    Q(lote__centro__isnull=True) | 
                    Q(lote__centro__nombre__icontains='almacén central') |
                    Q(lote__centro__nombre__icontains='almacen central')
                )
            
            # Calcular entradas y salidas
            entradas = mov_mes.filter(tipo='entrada').aggregate(
                total=Coalesce(Sum('cantidad'), 0, output_field=IntegerField())
            )['total']
            
            salidas = mov_mes.filter(tipo='salida').aggregate(
                total=Coalesce(Sum('cantidad'), 0, output_field=IntegerField())
            )['total']
            
            consumo_mensual.append({
                'mes': mes_inicio.strftime('%b'),
                'entradas': max(0, abs(entradas)),
                'salidas': max(0, abs(salidas)),
            })
        
        # =========================================
        # 2. STOCK POR CENTRO (TODOS LOS CENTROS CON STOCK > 0)
        # =========================================
        stock_por_centro = []
        
        if not filtrar_por_centro:
            # Farmacia Central: consolidar lotes sin centro + lotes de "Almacén Central"
            # (Son conceptualmente el mismo lugar)
            stock_farmacia = Lote.objects.filter(
                Q(centro__isnull=True) | Q(centro__nombre__icontains='almacén central') | Q(centro__nombre__icontains='almacen central'),
                activo=True,
                cantidad_actual__gt=0
            ).aggregate(
                total=Coalesce(Sum('cantidad_actual'), 0, output_field=IntegerField())
            )['total']
            
            # Solo agregar Farmacia Central si tiene stock
            if stock_farmacia > 0:
                stock_por_centro.append({
                    'centro': 'Farmacia Central',
                    'stock': stock_farmacia
                })
            
            # Todos los centros activos CON STOCK (excluyendo Almacén Central que ya está consolidado)
            for centro in Centro.objects.filter(activo=True).exclude(
                Q(nombre__icontains='almacén central') | Q(nombre__icontains='almacen central')
            ).order_by('nombre'):
                stock = Lote.objects.filter(
                    centro=centro,
                    activo=True,
                    cantidad_actual__gt=0
                ).aggregate(
                    total=Coalesce(Sum('cantidad_actual'), 0, output_field=IntegerField())
                )['total']
                
                # Solo agregar centros con stock > 0
                if stock > 0:
                    # Truncar nombre largo
                    nombre = centro.nombre
                    if len(nombre) > 20:
                        nombre = nombre[:17] + '...'
                    
                    stock_por_centro.append({
                        'centro': nombre,
                        'stock': stock
                    })
        else:
            # Usuario de centro: solo su stock
            if user_centro:
                stock = Lote.objects.filter(
                    centro=user_centro,
                    activo=True,
                    cantidad_actual__gt=0
                ).aggregate(
                    total=Coalesce(Sum('cantidad_actual'), 0, output_field=IntegerField())
                )['total']
                
                stock_por_centro.append({
                    'centro': user_centro.nombre,
                    'stock': max(0, stock)
                })
        
        # =========================================
        # 3. REQUISICIONES POR ESTADO
        # =========================================
        requisiciones_qs = Requisicion.objects.all()
        if filtrar_por_centro and user_centro:
            requisiciones_qs = requisiciones_qs.filter(centro_destino=user_centro)
        
        estados_agg = requisiciones_qs.values('estado').annotate(
            cantidad=Count('id')
        ).order_by('estado')
        
        requisiciones_por_estado = []
        for item in estados_agg:
            if item['cantidad'] > 0:
                requisiciones_por_estado.append({
                    'estado': item['estado'].upper(),
                    'cantidad': item['cantidad']
                })
        
        # ISS-005: Preparar respuesta y guardar en caché
        response_data = {
            'consumo_mensual': consumo_mensual,
            'stock_por_centro': stock_por_centro,
            'requisiciones_por_estado': requisiciones_por_estado
        }
        
        # Guardar en caché con TTL de estadísticas (5 minutos por defecto)
        cache_ttl = getattr(settings, 'CACHE_TTL_ESTADISTICAS', 300)
        cache.set(cache_key, response_data, cache_ttl)
        
        return Response(response_data)
        
    except Exception as exc:
        # traceback removido por seguridad (ISS-008)
        logger.exception('Error en dashboard_graficas')
        return Response({
            'consumo_mensual': [],
            'stock_por_centro': [],
            'requisiciones_por_estado': [],
            'error': 'Error interno del servidor'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def trazabilidad_producto(request, clave):
    """
    Trazabilidad de un producto identificado por clave (case-insensitive).
    Retorna lotes, movimientos y alertas de stock/caducidad.
    
    Filtros soportados:
    - centro: ID del centro (solo admin/farmacia)
    - fecha_inicio: Fecha inicio (YYYY-MM-DD) para filtrar movimientos
    - fecha_fin: Fecha fin (YYYY-MM-DD) para filtrar movimientos
    - tipo: Tipo de movimiento (entrada, salida, ajuste)
    
    SEGURIDAD: Filtra por centro del usuario si no es admin/farmacia.
    """
    from datetime import datetime, date, timedelta
    
    try:
        # SEGURIDAD: Determinar filtro de centro
        user = request.user
        if not user or not user.is_authenticated:
            return Response({'error': 'Autenticacion requerida'}, status=status.HTTP_403_FORBIDDEN)
        rol_usuario = (getattr(user, 'rol', '') or '').lower()
        if rol_usuario == 'vista':
            return Response({'error': 'No tienes permiso para trazabilidad'}, status=status.HTTP_403_FORBIDDEN)
        filtrar_por_centro = not is_farmacia_or_admin(user)
        user_centro = get_user_centro(user) if filtrar_por_centro else None
        
        # Obtener parámetros de filtro
        centro_param = request.query_params.get('centro')
        fecha_inicio = request.query_params.get('fecha_inicio')
        fecha_fin = request.query_params.get('fecha_fin')
        tipo_movimiento = request.query_params.get('tipo')
        
        # ISS-FIX: Manejar valores especiales de centro
        if centro_param and is_farmacia_or_admin(user):
            if centro_param.lower() in ('central', 'todos'):
                # 'central' = Farmacia Central (sin filtro extra)
                # 'todos' = ver todo
                if centro_param.lower() == 'central':
                    user_centro = None
                else:
                    filtrar_por_centro = False
            else:
                try:
                    if centro_param.isdigit():
                        user_centro = Centro.objects.get(pk=centro_param)
                    else:
                        user_centro = Centro.objects.get(nombre__iexact=centro_param)
                    filtrar_por_centro = True
                except Centro.DoesNotExist:
                    pass
        
        producto = Producto.objects.filter(
            Q(clave__iexact=clave) | Q(descripcion__iexact=clave)
        ).first()
        if not producto:
            return Response({'error': 'Producto no encontrado', 'clave_buscada': clave}, status=status.HTTP_404_NOT_FOUND)

        # ISS-FIX: Filtrar lotes según rol del usuario
        # - Farmacia/Admin: Ver todos los lotes (para trazabilidad histórica completa)
        # - Centro: Solo lotes activos con stock
        es_admin_farmacia = is_farmacia_or_admin(user)
        
        if es_admin_farmacia:
            # Farmacia/Admin: ver todos los lotes incluyendo históricos
            lotes = Lote.objects.filter(producto=producto)
        else:
            # Centro: solo lotes activos con stock disponible
            lotes = Lote.objects.filter(producto=producto, activo=True, cantidad_actual__gt=0)
        
        # Aplicar filtro de centro
        if filtrar_por_centro and user_centro:
            lotes = lotes.filter(centro=user_centro)
        
        lotes = lotes.order_by('-created_at')
        
        # TRAZABILIDAD: Consolidar lotes por numero_lote (mismo lote puede estar en varios centros)
        # Agrupamos por numero_lote para mostrar información consolidada
        lotes_consolidados = {}
        for lote in lotes:
            key = lote.numero_lote
            if key not in lotes_consolidados:
                lotes_consolidados[key] = {
                    'lote_principal': lote,
                    'cantidad_total': 0,
                    'lotes_ids': [],
                    'centros': [],
                }
            lotes_consolidados[key]['cantidad_total'] += lote.cantidad_actual or 0
            lotes_consolidados[key]['lotes_ids'].append(lote.id)
            if lote.centro:
                centro_nombre = lote.centro.nombre
                if centro_nombre not in lotes_consolidados[key]['centros']:
                    lotes_consolidados[key]['centros'].append(centro_nombre)
            elif 'Farmacia Central' not in lotes_consolidados[key]['centros']:
                lotes_consolidados[key]['centros'].append('Farmacia Central')
        
        lotes_data = []
        for key, data in lotes_consolidados.items():
            lote = data['lote_principal']
            dias_caducidad = (lote.fecha_caducidad - date.today()).days if lote.fecha_caducidad else None
            if dias_caducidad is None:
                estado_caducidad = 'DESCONOCIDO'
            elif dias_caducidad < 0:
                estado_caducidad = 'VENCIDO'
            elif dias_caducidad <= 7:
                estado_caducidad = 'CRITICO'
            elif dias_caducidad <= 30:
                estado_caducidad = 'PROXIMO'
            else:
                estado_caducidad = 'NORMAL'

            # Calcular totales de movimientos para todos los lotes con este numero_lote
            movimientos_lote = Movimiento.objects.filter(lote_id__in=data['lotes_ids'])
            total_entradas = movimientos_lote.filter(tipo='entrada').aggregate(total=Sum('cantidad'))['total'] or 0
            total_salidas = movimientos_lote.filter(tipo='salida').aggregate(total=Sum('cantidad'))['total'] or 0

            lotes_data.append({
                'id': lote.id,  # ID del lote principal
                'numero_lote': lote.numero_lote,
                'fecha_caducidad': lote.fecha_caducidad.isoformat() if lote.fecha_caducidad else None,
                'dias_para_caducar': dias_caducidad,
                'estado_caducidad': estado_caducidad,
                'estado': estado_caducidad,  # Alias para frontend
                'cantidad_actual': data['cantidad_total'],  # Cantidad consolidada
                'cantidad_inicial': lote.cantidad_inicial,
                'total_entradas': total_entradas,
                'total_salidas': total_salidas,
                'marca': lote.marca or 'N/A',
                'precio_unitario': str(lote.precio_unitario) if lote.precio_unitario else None,
                # Campos de trazabilidad de contratos (solo para ADMIN/FARMACIA)
                'numero_contrato': (lote.numero_contrato or '') if is_farmacia_or_admin(user) else None,
                'activo': getattr(lote, 'activo', True),
                'created_at': lote.created_at.isoformat(),
                # NUEVO: Ubicaciones donde está distribuido el lote
                'ubicaciones': data['centros'],
                'distribuido_en': len(data['centros']) if data['centros'] else 0,
            })

        movimientos = Movimiento.objects.filter(lote__producto=producto).select_related('lote', 'centro_origen', 'centro_destino')
        
        # ISS-FIX: Aplicar filtro de centro de forma ESTRICTA a movimientos
        # Solo donde el centro es origen O destino (no por lote__centro)
        if filtrar_por_centro and user_centro:
            movimientos = movimientos.filter(
                Q(centro_origen=user_centro) | Q(centro_destino=user_centro)
            )
        
        # Aplicar filtros de fecha a movimientos
        if fecha_inicio:
            try:
                fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d')
                movimientos = movimientos.filter(fecha__gte=fecha_inicio_dt)
            except ValueError:
                pass
        
        if fecha_fin:
            try:
                fecha_fin_dt = datetime.strptime(fecha_fin, '%Y-%m-%d')
                fecha_fin_dt = fecha_fin_dt.replace(hour=23, minute=59, second=59)
                movimientos = movimientos.filter(fecha__lte=fecha_fin_dt)
            except ValueError:
                pass
        
        # Aplicar filtro de tipo de movimiento
        if tipo_movimiento:
            movimientos = movimientos.filter(tipo=tipo_movimiento.lower())
        
        movimientos = movimientos.order_by('-fecha')[:100]
        movimientos_data = []
        for mov in movimientos:
            movimientos_data.append({
                'id': mov.id,
                'tipo_movimiento': mov.tipo.upper(),
                'tipo': mov.tipo.upper(),
                'lote': mov.lote.numero_lote if mov.lote else 'N/A',
                'cantidad': mov.cantidad,
                'fecha_movimiento': mov.fecha.isoformat(),
                'observaciones': mov.motivo or ''
            })

        stock_total = lotes.filter(activo=True).aggregate(total=Sum('cantidad_actual'))['total'] or 0
        # TRAZABILIDAD: Contar lotes ÚNICOS (consolidados por numero_lote)
        lotes_activos = len([k for k, v in lotes_consolidados.items() if v['cantidad_total'] > 0])
        total_lotes = len(lotes_consolidados)  # Lotes únicos, no registros duplicados
        
        # Totales de entradas/salidas filtrados por centro si aplica
        mov_query = Movimiento.objects.filter(lote__producto=producto)
        if filtrar_por_centro and user_centro:
            mov_query = mov_query.filter(lote__centro=user_centro)
        total_entradas_prod = mov_query.filter(tipo='entrada').aggregate(total=Sum('cantidad'))['total'] or 0
        total_salidas_prod = mov_query.filter(tipo='salida').aggregate(total=Sum('cantidad'))['total'] or 0

        fecha_limite = date.today() + timedelta(days=30)
        lotes_proximos_vencer = lotes.filter(cantidad_actual__gt=0, fecha_caducidad__lte=fecha_limite, fecha_caducidad__gte=date.today()).count()
        lotes_vencidos = lotes.filter(cantidad_actual__gt=0, fecha_caducidad__lt=date.today()).count()

        alertas = []
        if stock_total < producto.stock_minimo:
            alertas.append({'tipo': 'STOCK_BAJO', 'mensaje': f'Stock actual ({stock_total}) por debajo del minimo ({producto.stock_minimo})', 'nivel': 'CRITICO'})
        if lotes_vencidos > 0:
            alertas.append({'tipo': 'LOTES_VENCIDOS', 'mensaje': f'{lotes_vencidos} lote(s) vencido(s) con stock', 'nivel': 'CRITICO'})
        if lotes_proximos_vencer > 0:
            alertas.append({'tipo': 'PROXIMOS_VENCER', 'mensaje': f'{lotes_proximos_vencer} lote(s) proximo(s) a vencer (30 dias)', 'nivel': 'ADVERTENCIA'})

        return Response({
            'codigo': producto.clave,
            'producto': {
                'id': producto.id,
                'clave': producto.clave,
                'nombre': producto.nombre,  # Campo nombre explícito
                'descripcion': producto.descripcion or producto.nombre,  # Descripción o nombre como fallback
                'unidad_medida': producto.unidad_medida,
                'stock_minimo': producto.stock_minimo,
                'precio_unitario': None,  # precio_unitario está en Lote, no en Producto
                'activo': producto.activo
            },
            'estadisticas': {
                'stock_total': stock_total,
                'total_lotes': total_lotes,
                'lotes_activos': lotes_activos,
                'total_entradas': total_entradas_prod,
                'total_salidas': total_salidas_prod,
                'diferencia': total_entradas_prod - total_salidas_prod,
                'lotes_proximos_vencer': lotes_proximos_vencer,
                'lotes_vencidos': lotes_vencidos,
                'bajo_minimo': stock_total < producto.stock_minimo
            },
            'lotes': lotes_data,
            'movimientos': movimientos_data,
            'total_movimientos': Movimiento.objects.filter(lote__producto=producto).count(),
            'alertas': alertas
        })
    except Exception as exc:
        # traceback removido por seguridad (ISS-008)
        return Response({'error': 'Error al obtener trazabilidad del producto', 'mensaje': str(exc)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
def trazabilidad_lote(request, codigo):
    """
    Trazabilidad completa de un lote por su numero.
    
    IMPORTANTE: Busca movimientos de TODOS los lotes con el mismo numero_lote,
    incluyendo el lote original en Farmacia Central y los lotes espejo en centros.
    Esto permite ver la trazabilidad completa: desde la entrada en farmacia,
    las transferencias a centros, y las dispensaciones/salidas en cada centro.
    
    Filtros soportados:
    - fecha_inicio: Fecha inicio (YYYY-MM-DD) para filtrar movimientos
    - fecha_fin: Fecha fin (YYYY-MM-DD) para filtrar movimientos
    - tipo: Tipo de movimiento (entrada, salida, ajuste)
    
    SEGURIDAD: Filtra por centro del usuario si no es admin/farmacia.
    """
    from datetime import datetime
    
    try:
        if not request.user or not request.user.is_authenticated or not is_farmacia_or_admin(request.user):
            return Response({'error': 'Solo usuarios de farmacia o administradores pueden acceder a reportes'}, status=status.HTTP_403_FORBIDDEN)

        # SEGURIDAD: Determinar filtro de centro
        user = request.user
        filtrar_por_centro = not is_farmacia_or_admin(user)
        user_centro = get_user_centro(user) if filtrar_por_centro else None
        
        # Obtener parámetros de filtro
        fecha_inicio = request.query_params.get('fecha_inicio')
        fecha_fin = request.query_params.get('fecha_fin')
        tipo_movimiento = request.query_params.get('tipo')
        
        # TRAZABILIDAD COMPLETA: Buscar TODOS los lotes con el mismo numero_lote
        # Esto incluye el lote original en Farmacia Central y los lotes espejo en centros
        lotes_con_mismo_numero = Lote.objects.select_related('producto', 'centro').filter(numero_lote__iexact=codigo)
        
        if not lotes_con_mismo_numero.exists():
            return Response({'error': 'Lote no encontrado', 'codigo_buscado': codigo}, status=status.HTTP_404_NOT_FOUND)
        
        # Usar el lote de Farmacia Central como principal (centro=NULL) o el primero disponible
        lote_principal = lotes_con_mismo_numero.filter(centro__isnull=True).first()
        if not lote_principal:
            lote_principal = lotes_con_mismo_numero.first()
        
        # Obtener todos los IDs de lotes con el mismo numero_lote
        lotes_ids = list(lotes_con_mismo_numero.values_list('id', flat=True))
        
        # Calcular cantidad total consolidada (suma de todos los lotes espejo)
        cantidad_total_consolidada = lotes_con_mismo_numero.aggregate(total=Sum('cantidad_actual'))['total'] or 0
        
        # Obtener centros donde está distribuido el lote
        centros_distribucion = []
        for lote_item in lotes_con_mismo_numero:
            centro_nombre = lote_item.centro.nombre if lote_item.centro else 'Farmacia Central'
            if centro_nombre not in centros_distribucion:
                centros_distribucion.append({
                    'nombre': centro_nombre,
                    'cantidad': lote_item.cantidad_actual
                })
        
        # TRAZABILIDAD COMPLETA: Buscar movimientos de TODOS los lotes con el mismo numero_lote
        movimientos = Movimiento.objects.select_related(
            'centro_origen', 'centro_destino', 'usuario', 'lote', 'lote__centro'
        ).filter(lote_id__in=lotes_ids)
        
        # Aplicar filtros de fecha
        if fecha_inicio:
            try:
                fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d')
                movimientos = movimientos.filter(fecha__gte=fecha_inicio_dt)
            except ValueError:
                pass
        
        if fecha_fin:
            try:
                fecha_fin_dt = datetime.strptime(fecha_fin, '%Y-%m-%d')
                fecha_fin_dt = fecha_fin_dt.replace(hour=23, minute=59, second=59)
                movimientos = movimientos.filter(fecha__lte=fecha_fin_dt)
            except ValueError:
                pass
        
        # Aplicar filtro de tipo
        if tipo_movimiento:
            movimientos = movimientos.filter(tipo=tipo_movimiento.lower())
        
        movimientos = movimientos.order_by('fecha')
        historial = []
        saldo = 0
        for mov in movimientos:
            saldo += mov.cantidad
            # Determinar el centro del movimiento de forma más precisa
            if mov.centro_destino:
                centro_mov = mov.centro_destino.nombre
            elif mov.centro_origen:
                centro_mov = mov.centro_origen.nombre
            elif mov.lote and mov.lote.centro:
                centro_mov = mov.lote.centro.nombre
            else:
                centro_mov = 'Farmacia Central'
            
            historial.append({
                'id': mov.id,
                'fecha': mov.fecha.isoformat(),
                'tipo': mov.tipo.upper(),
                'cantidad': mov.cantidad,
                'saldo': saldo,
                'centro': centro_mov,
                'usuario': mov.usuario.username if mov.usuario else '-',
                'lote': mov.lote.numero_lote if mov.lote else '-',
                'observaciones': mov.motivo or ''
            })

        total_entradas = movimientos.filter(tipo='entrada').aggregate(total=Sum('cantidad'))['total'] or 0
        total_salidas = abs(movimientos.filter(tipo='salida').aggregate(total=Sum('cantidad'))['total'] or 0)

        from datetime import date
        lote = lote_principal  # Usar lote principal para datos generales
        dias_caducidad = (lote.fecha_caducidad - date.today()).days if lote.fecha_caducidad else None
        if dias_caducidad is None:
            estado_caducidad = 'DESCONOCIDO'
        elif dias_caducidad < 0:
            estado_caducidad = 'VENCIDO'
        elif dias_caducidad <= 7:
            estado_caducidad = 'CRITICO'
        elif dias_caducidad <= 30:
            estado_caducidad = 'PROXIMO'
        else:
            estado_caducidad = 'NORMAL'

        alertas = []
        if dias_caducidad is not None:
            if dias_caducidad < 0:
                alertas.append({'tipo': 'VENCIDO', 'mensaje': f'Lote vencido hace {abs(dias_caducidad)} dias', 'nivel': 'CRITICO'})
            elif dias_caducidad <= 7:
                alertas.append({'tipo': 'CRITICO', 'mensaje': f'Caduca en {dias_caducidad} dias', 'nivel': 'CRITICO'})
            elif dias_caducidad <= 30:
                alertas.append({'tipo': 'PROXIMO', 'mensaje': f'Caduca en {dias_caducidad} dias', 'nivel': 'ADVERTENCIA'})

        return Response({
            'id': lote.id,
            'numero_lote': lote.numero_lote,
            'producto': lote.producto.clave,
            'lote': {
                'id': lote.id,
                'numero_lote': lote.numero_lote,
                'producto': lote.producto.clave,
                'producto_nombre': lote.producto.nombre,  # Campo para frontend
                'producto_descripcion': lote.producto.descripcion or lote.producto.nombre,
                'cantidad_actual': cantidad_total_consolidada,  # Cantidad consolidada de todos los centros
                'cantidad_inicial': lote.cantidad_inicial,
                'activo': lote.activo,
                'fecha_caducidad': lote.fecha_caducidad.isoformat() if lote.fecha_caducidad else None,
                'dias_para_caducar': dias_caducidad,
                'estado_caducidad': estado_caducidad,
                'marca': lote.marca,
                'proveedor': None,  # Campo no existe en modelo, dejarlo null
                # ISS-FIX: Agregar centro (nombre o 'Farmacia Central' si es null)
                'centro': lote.centro.nombre if lote.centro else 'Farmacia Central',
                'centro_id': lote.centro.id if lote.centro else None,
                # Campos de trazabilidad de contratos (solo para ADMIN/FARMACIA)
                'numero_contrato': lote.numero_contrato if is_farmacia_or_admin(user) else None,
                # NUEVO: Distribución del lote en centros
                'distribucion': centros_distribucion,
            },
            'estadisticas': {
                'total_entradas': total_entradas,
                'total_salidas': total_salidas,
                'diferencia': total_entradas - total_salidas,
                'cantidad_actual': cantidad_total_consolidada,
                'saldo_calculado': saldo,
                'diferencia_stock': saldo - cantidad_total_consolidada,
                'consistente': saldo == cantidad_total_consolidada
            },
            'movimientos': historial,
            'historial': historial,  # compatibilidad hacia atr?s
            'total_movimientos': movimientos.count(),
            'alertas': alertas
        })
    except Exception as exc:
        # traceback removido por seguridad (ISS-008)
        return Response({'error': 'Error al obtener trazabilidad del lote', 'mensaje': str(exc)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
def reporte_inventario(request):
    """
    Genera reporte de inventario actual.
    Por defecto devuelve JSON. Con ?formato=excel devuelve Excel.
    
    SEGURIDAD: Filtra por centro del usuario si no es admin/farmacia.
    Admin/farmacia puede usar ?centro=ID para filtrar.
    
    Parmetros:
    - centro: ID del centro o 'central' para farmacia central
    - nivel_stock: alto, bajo, normal, sin_stock
    - formato: json (default), excel, pdf
    """
    try:
        if not request.user or not request.user.is_authenticated or not is_farmacia_or_admin(request.user):
            return Response({'error': 'Solo usuarios de farmacia o administradores pueden acceder a reportes'}, status=status.HTTP_403_FORBIDDEN)
        # SEGURIDAD: Determinar filtro de centro
        user = request.user
        
        # Por defecto: admin/farmacia ven solo Farmacia Central
        # Usuarios de centro ven solo su centro
        filtrar_por_centro = True  # Siempre filtrar
        user_centro = None  # NULL = Farmacia Central por defecto
        
        if not is_farmacia_or_admin(user):
            # Usuario de centro: filtrar por su centro
            user_centro = get_user_centro(user)
        
        # Admin/farmacia puede filtrar por centro específico
        centro_param = request.query_params.get('centro')
        if centro_param and is_farmacia_or_admin(user):
            if centro_param.lower() == 'central':
                # Filtrar solo farmacia central (ya es el default)
                user_centro = None
            elif centro_param.lower() == 'todos':
                # Ver todo (sin filtro de centro)
                filtrar_por_centro = False
            else:
                try:
                    # ISS-FIX: Buscar por ID o nombre
                    if centro_param.isdigit():
                        user_centro = Centro.objects.get(pk=centro_param)
                    else:
                        user_centro = Centro.objects.get(nombre__iexact=centro_param)
                    filtrar_por_centro = True
                except Centro.DoesNotExist:
                    pass
        
        formato = request.query_params.get('formato', 'json')
        nivel_stock_filtro = request.query_params.get('nivel_stock', '').lower().strip()
        productos = Producto.objects.filter(activo=True).order_by('clave')
        
        # Construir datos
        datos = []
        total_productos = 0
        total_stock = 0
        productos_bajo_minimo = 0
        productos_sin_stock = 0
        
        for idx, producto in enumerate(productos, 1):
            # Aplicar filtro de centro si corresponde
            lotes_query = producto.lotes.filter(
                activo=True
            )
            if filtrar_por_centro:
                if user_centro:
                    lotes_query = lotes_query.filter(centro=user_centro)
                else:
                    # centro=NULL significa farmacia central
                    lotes_query = lotes_query.filter(centro__isnull=True)
            
            stock_total = lotes_query.aggregate(total=Sum('cantidad_actual'))['total'] or 0
            lotes_activos = lotes_query.filter(cantidad_actual__gt=0).count()
            
            # Calcular precio promedio desde lotes (precio_unitario está en Lote, no en Producto)
            from django.db.models import Avg
            precio_promedio = lotes_query.filter(cantidad_actual__gt=0).aggregate(avg=Avg('precio_unitario'))['avg'] or 0.0

            nivel = 'alto'
            if stock_total == 0:
                nivel = 'sin_stock'
                productos_sin_stock += 1
            elif stock_total < producto.stock_minimo:
                nivel = 'bajo'
                productos_bajo_minimo += 1
            elif stock_total < producto.stock_minimo * 1.5:
                nivel = 'normal'
            
            # Filtrar por nivel_stock si se especific
            if nivel_stock_filtro and nivel != nivel_stock_filtro:
                continue

            # Obtener marca del lote con más stock (o el más reciente)
            marca_lote = ''
            lote_con_marca = lotes_query.filter(
                cantidad_actual__gt=0, 
                marca__isnull=False
            ).exclude(marca='').order_by('-cantidad_actual').first()
            if lote_con_marca:
                marca_lote = lote_con_marca.marca or ''

            # Si hay filtro de centro activo, NO mostrar productos sin stock en ese centro
            # (evita confusión de ver 76 productos cuando solo 4 tienen stock)
            if filtrar_por_centro and stock_total == 0:
                continue
            
            # Se incluye 'nivel_stock' para compatibilidad con el frontend
            # Usar 'nombre' del producto como descripción principal
            datos.append({
                '#': idx,
                'clave': producto.clave,
                'descripcion': producto.nombre,  # nombre es el campo principal del producto
                'presentacion': producto.presentacion or '-',  # Presentación del producto
                'unidad': producto.unidad_medida,
                'unidad_medida': producto.unidad_medida,  # alias esperado por frontend
                'stock_minimo': producto.stock_minimo,
                'stock_actual': stock_total,
                'lotes_activos': lotes_activos,
                'nivel': nivel,
                'nivel_stock': nivel,
                'precio_unitario': float(precio_promedio),
                'marca': marca_lote,  # Marca del lote con más stock
            })
            total_productos += 1
            total_stock += stock_total
        
        resumen = {
            'total_productos': total_productos,
            'total_stock': total_stock,
            'stock_total': total_stock,  # alias para compatibilidad frontend
            'productos_bajo_minimo': productos_bajo_minimo,
            'productos_sin_stock': productos_sin_stock,
        }
        
        # Si formato es JSON, devolver datos
        if formato == 'json':
            return Response({
                'datos': datos,
                'resumen': resumen
            })
        
        # =====================================================
        # LOTES PARA PDF/EXCEL: Respeta filtro de centro
        # Por defecto muestra solo Farmacia Central (centro=NULL)
        # Si se pasa centro=todos, muestra consolidado de todos
        # =====================================================
        from core.models import Lote
        from collections import defaultdict
        
        # Aplicar MISMO filtro de centro que el JSON
        lotes_query = Lote.objects.select_related('producto', 'centro').filter(
            activo=True,
            cantidad_actual__gt=0
        )
        
        # Aplicar filtro de centro (igual que arriba)
        if filtrar_por_centro:
            if user_centro:
                lotes_query = lotes_query.filter(centro=user_centro)
            else:
                # centro=NULL significa farmacia central (por defecto)
                lotes_query = lotes_query.filter(centro__isnull=True)
        
        lotes_query = lotes_query.order_by('producto__clave', 'numero_lote', 'fecha_caducidad')
        
        # Determinar si consolidar (solo si no hay filtro de centro)
        debe_consolidar = not filtrar_por_centro or centro_param == 'todos'
        
        if debe_consolidar:
            # Consolidar lotes por (producto_id, numero_lote)
            lotes_consolidados = defaultdict(lambda: {
                'producto': None,
                'numero_lote': '',
                'cantidad_total': 0,
                'fecha_caducidad': None,
                'precio_unitario': 0,
                'marca': '-',
                'centros': [],
                'ubicacion': '-',
            })
            
            for lote in lotes_query:
                key = (lote.producto_id, lote.numero_lote)
                consolidado = lotes_consolidados[key]
                
                if consolidado['producto'] is None:
                    consolidado['producto'] = lote.producto
                    consolidado['numero_lote'] = lote.numero_lote
                    consolidado['fecha_caducidad'] = lote.fecha_caducidad
                    consolidado['precio_unitario'] = float(lote.precio_unitario or 0)
                    consolidado['marca'] = lote.marca or '-'
                
                consolidado['cantidad_total'] += lote.cantidad_actual
                centro_nombre = lote.centro.nombre if lote.centro else 'Almacén Central'
                if centro_nombre not in consolidado['centros']:
                    consolidado['centros'].append(centro_nombre)
            
            lotes_lista = sorted(
                lotes_consolidados.values(),
                key=lambda x: (x['producto'].clave, x['numero_lote'])
            )
        else:
            # Sin consolidación: lista normal de lotes
            lotes_lista = []
            for lote in lotes_query:
                lotes_lista.append({
                    'producto': lote.producto,
                    'numero_lote': lote.numero_lote,
                    'cantidad_total': lote.cantidad_actual,
                    'fecha_caducidad': lote.fecha_caducidad,
                    'precio_unitario': float(lote.precio_unitario or 0),
                    'marca': lote.marca or '-',
                    'centros': [lote.centro.nombre if lote.centro else 'Almacén Central'],
                    'ubicacion': lote.ubicacion or '-',
                })
        
        # Determinar título según filtro
        if not filtrar_por_centro or centro_param == 'todos':
            titulo_reporte = 'REPORTE DE INVENTARIO - LOTES CONSOLIDADOS'
            subtitulo_reporte = f"Generado el {timezone.now().strftime('%d/%m/%Y %H:%M:%S')} - Todos los centros"
        elif user_centro:
            titulo_reporte = f'INVENTARIO DE {user_centro.nombre.upper()}'
            subtitulo_reporte = f"Generado el {timezone.now().strftime('%d/%m/%Y %H:%M:%S')}"
        else:
            titulo_reporte = 'INVENTARIO DE FARMACIA CENTRAL'
            subtitulo_reporte = f"Generado el {timezone.now().strftime('%d/%m/%Y %H:%M:%S')} - Almacén Central"
        
        # Formato PDF - LOTES (con o sin consolidación según filtro)
        if formato == 'pdf':
            from core.utils.pdf_reports import generar_reporte_inventario_lotes
            
            lotes_data = []
            for lote_cons in lotes_lista:
                producto = lote_cons['producto']
                lotes_data.append({
                    'clave': producto.clave,
                    'producto': producto.nombre or producto.descripcion or '',
                    'presentacion': producto.presentacion or '-',
                    'numero_lote': lote_cons['numero_lote'],
                    'fecha_caducidad': lote_cons['fecha_caducidad'].strftime('%d/%m/%Y') if lote_cons['fecha_caducidad'] else '-',
                    'cantidad': lote_cons['cantidad_total'],
                    'precio_unitario': lote_cons['precio_unitario'],
                    'ubicacion': ', '.join(lote_cons['centros'][:2]) + ('...' if len(lote_cons['centros']) > 2 else ''),
                    'marca': lote_cons['marca'],
                })
            
            filtros = {
                'fecha_generacion': timezone.now().strftime('%d/%m/%Y %H:%M'),
                'total_lotes': len(lotes_data),
                'total_productos': resumen['total_productos'],
                'titulo': titulo_reporte,
            }
            
            pdf_buffer = generar_reporte_inventario_lotes(lotes_data, filtros=filtros)
            
            response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = f"attachment; filename=Inventario_Lotes_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            return response
        
        # Formato Excel - LOTES (con o sin consolidación según filtro)
        from openpyxl.styles import Border, Side
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Inventario por Lotes'

        ws.merge_cells('A1:K1')
        titulo = ws['A1']
        titulo.value = titulo_reporte
        titulo.font = Font(bold=True, size=14, color='632842')
        titulo.alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells('A2:K2')
        subtitulo = ws['A2']
        subtitulo.value = subtitulo_reporte
        subtitulo.alignment = Alignment(horizontal='center')
        subtitulo.font = Font(size=10, italic=True)

        ws.append([])
        headers = ['#', 'Clave', 'Producto', 'Presentación', 'Lote', 'Caducidad', 'Stock Total', 'Precio Unit.', 'Nivel', 'Distribuido en', 'Marca / Laboratorio']
        ws.append(headers)

        header_fill = PatternFill(start_color='632842', end_color='632842', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=10)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        for cell in ws[4]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        row_num = 5
        total_lotes = 0
        total_unidades = 0
        
        for idx, lote_cons in enumerate(lotes_lista, 1):
            producto = lote_cons['producto']
            stock_total = lote_cons['cantidad_total']
            
            # Determinar nivel de stock del producto
            nivel = 'ALTO'
            if stock_total == 0:
                nivel = 'SIN STOCK'
            elif stock_total < producto.stock_minimo:
                nivel = 'BAJO'
            elif stock_total < producto.stock_minimo * 1.5:
                nivel = 'NORMAL'
            
            # Mostrar centros donde está distribuido (máx 3)
            centros_texto = ', '.join(lote_cons['centros'][:3])
            if len(lote_cons['centros']) > 3:
                centros_texto += f' (+{len(lote_cons["centros"]) - 3})'
            
            ws.append([
                idx,
                producto.clave,
                (producto.nombre or producto.descripcion or '')[:50],
                producto.presentacion or '-',
                lote_cons['numero_lote'],
                lote_cons['fecha_caducidad'].strftime('%d/%m/%Y') if lote_cons['fecha_caducidad'] else '-',
                stock_total,
                lote_cons['precio_unitario'],
                nivel,
                centros_texto,
                lote_cons['marca']
            ])
            
            for cell in ws[row_num]:
                cell.border = thin_border
            
            # Colorear nivel
            nivel_cell = ws.cell(row=row_num, column=9)
            if nivel == 'ALTO':
                nivel_cell.fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
                nivel_cell.font = Font(color='155724', bold=True)
            elif nivel == 'BAJO':
                nivel_cell.fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
                nivel_cell.font = Font(color='856404', bold=True)
            elif nivel == 'SIN STOCK':
                nivel_cell.fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')
                nivel_cell.font = Font(color='721C24', bold=True)
            
            row_num += 1
            total_lotes += 1
            total_unidades += stock_total

        # Resumen
        ws.append([])
        resumen_row = ws.max_row + 1
        ws[f'A{resumen_row}'] = 'RESUMEN:'
        ws[f'A{resumen_row}'].font = Font(bold=True, size=11)
        ws[f'B{resumen_row}'] = f'Lotes únicos: {total_lotes}'
        ws[f'D{resumen_row}'] = f'Total Unidades: {total_unidades:,}'
        ws[f'F{resumen_row}'] = f'Productos únicos: {resumen["total_productos"]}'

        for col, width in zip(['A','B','C','D','E','F','G','H','I','J','K'], [5,10,35,25,15,12,12,12,10,25,20]):
            ws.column_dimensions[col].width = width

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f"attachment; filename=Inventario_Lotes_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(response)
        return response

    except Exception as e:
        # traceback removido por seguridad (ISS-008)
        return Response({'error': 'Error al generar reporte', 'mensaje': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
@api_view(['GET'])
def reporte_movimientos(request):
    """
    Genera reporte de movimientos con filtros.
    
    SEGURIDAD: Filtra por centro del usuario si no es admin/farmacia.
    Admin/farmacia puede usar ?centro=ID para filtrar.
    
    Parametros:
    - fecha_inicio: Fecha inicial (YYYY-MM-DD)
    - fecha_fin: Fecha final (YYYY-MM-DD)
    - tipo: ENTRADA o SALIDA
    - centro: ID del centro (solo admin/farmacia)
    - formato: excel o pdf
    """
    try:
        logger.info(f"Generando reporte de movimientos. Params: {dict(request.query_params)}")
        
        # SEGURIDAD: Determinar filtro de centro
        user = request.user
        filtrar_por_centro = not is_farmacia_or_admin(user)
        user_centro = get_user_centro(user) if filtrar_por_centro else None
        
        # Admin/farmacia puede filtrar por centro específico
        centro_param = request.query_params.get('centro')
        # Ignorar 'todos' como parámetro de centro
        if centro_param and centro_param.lower() == 'todos':
            centro_param = None
        # ISS-FIX: También ignorar 'central' ya que es Farmacia Central (sin filtro)
        if centro_param and centro_param.lower() == 'central':
            centro_param = None
        if centro_param and is_farmacia_or_admin(user):
            try:
                # ISS-FIX: Buscar por ID numérico o por nombre
                if centro_param.isdigit():
                    user_centro = Centro.objects.get(pk=centro_param)
                else:
                    user_centro = Centro.objects.get(nombre__iexact=centro_param)
                filtrar_por_centro = True
            except Centro.DoesNotExist:
                # Si no encuentra el centro, no filtrar (mostrar todos)
                pass
        
        # Obtener parametros
        fecha_inicio = request.query_params.get('fecha_inicio')
        fecha_fin = request.query_params.get('fecha_fin')
        tipo = request.query_params.get('tipo')
        formato = request.query_params.get('formato', 'excel')
        
        # Filtrar movimientos
        movimientos = Movimiento.objects.select_related('lote__producto', 'centro_origen', 'centro_destino').all()
        
        # FIX: Usar fecha__date para comparar solo la fecha (ignorar hora)
        if fecha_inicio:
            movimientos = movimientos.filter(fecha__date__gte=fecha_inicio)
        if fecha_fin:
            movimientos = movimientos.filter(fecha__date__lte=fecha_fin)
        if tipo:
            movimientos = movimientos.filter(tipo=tipo.lower())
        
        # ISS-FIX: Aplicar filtro de centro de forma ESTRICTA según el tipo de movimiento
        # Cuando admin/farmacia filtra por un centro específico, solo mostrar movimientos
        # directamente relacionados con ese centro (origen o destino), NO por lote__centro
        if filtrar_por_centro and user_centro:
            tipo_lower = (tipo or '').lower()
            if tipo_lower == 'salida':
                # Para SALIDAS: solo mostrar donde el centro es ORIGEN (salidas DESDE ese centro)
                # Excluimos lote__centro para evitar traer movimientos no relacionados
                movimientos = movimientos.filter(centro_origen=user_centro)
            elif tipo_lower == 'entrada':
                # Para ENTRADAS: solo mostrar donde el centro es DESTINO (entradas HACIA ese centro)
                movimientos = movimientos.filter(centro_destino=user_centro)
            else:
                # Sin tipo especificado: mostrar movimientos donde el centro es origen O destino
                # Se incluye lote__centro SOLO cuando no hay tipo para que usuarios de centro 
                # puedan ver sus propios movimientos internos (dispensaciones)
                movimientos = movimientos.filter(
                    Q(centro_origen=user_centro) | Q(centro_destino=user_centro) | Q(lote__centro=user_centro)
                )
        
        movimientos = movimientos.order_by('-fecha')
        
        # Agrupar movimientos por referencia/transacción
        # ISS-FIX: Agrupar por referencia + tipo para separar salidas y entradas de requisiciones
        transacciones = {}
        
        # ISS-FIX: Contadores globales para el resumen (por movimiento individual, no por transacción)
        total_unidades_entrada = 0
        total_unidades_salida = 0
        total_movs_entrada = 0
        total_movs_salida = 0
        
        for mov in movimientos:
            tipo_mov = mov.tipo.lower()
            amount = abs(mov.cantidad)  # Siempre positivo para mostrar
            ref = mov.referencia or f"MOV-{mov.id}"
            
            # Clasificar tipo: entrada, ajuste_positivo, devolucion = ENTRADA, resto = SALIDA
            es_entrada = tipo_mov in ['entrada', 'ajuste_positivo', 'devolucion']
            
            # ISS-FIX: Clave de agrupación incluye tipo para separar entradas y salidas
            # de la misma requisición/referencia
            tipo_grupo = 'ENTRADA' if es_entrada else 'SALIDA'
            grupo_key = f"{ref}|{tipo_grupo}"
            
            # ISS-FIX: Acumular métricas por cada movimiento individual
            if es_entrada:
                total_unidades_entrada += amount
                total_movs_entrada += 1
            else:
                total_unidades_salida += amount
                total_movs_salida += 1
            
            # Formatear subtipo de salida
            subtipo_display = ''
            if mov.subtipo_salida:
                subtipos_label = {
                    'receta': 'Receta Médica',
                    'consumo_interno': 'Consumo Interno',
                    'merma': 'Merma',
                    'caducidad': 'Caducidad',
                    'transferencia': 'Transferencia',
                }
                subtipo_display = subtipos_label.get(mov.subtipo_salida.lower(), mov.subtipo_salida.title())
            
            if grupo_key not in transacciones:
                # Crear nueva transacción agrupada
                transacciones[grupo_key] = {
                    'referencia': ref,
                    'fecha': mov.fecha.strftime('%d/%m/%Y %H:%M'),
                    'fecha_raw': mov.fecha,
                    'tipo': tipo_grupo,
                    'tipo_original': tipo_mov.upper(),
                    'subtipo_salida': mov.subtipo_salida or '',
                    'subtipo_display': subtipo_display,
                    'numero_expediente': mov.numero_expediente or '',
                    'centro_origen': mov.centro_origen.nombre if mov.centro_origen else 'Farmacia Central',
                    'centro_destino': mov.centro_destino.nombre if mov.centro_destino else 'Farmacia Central',
                    'total_productos': 0,
                    'total_cantidad': 0,
                    'observaciones': mov.motivo or '',
                    'detalles': [],
                    '_tipo_transaccion': tipo_grupo
                }
            
            # Agregar detalle a la transacción
            producto_info = 'N/A'
            presentacion = ''
            if mov.lote and mov.lote.producto:
                nombre = mov.lote.producto.nombre or mov.lote.producto.descripcion or mov.lote.producto.clave
                producto_info = f"{mov.lote.producto.clave} - {nombre[:50]}"
                presentacion = mov.lote.producto.presentacion or ''
            
            transacciones[grupo_key]['detalles'].append({
                'producto': producto_info,
                'presentacion': presentacion,
                'lote': mov.lote.numero_lote if mov.lote else 'N/A',
                'cantidad': amount,
                'subtipo_salida': mov.subtipo_salida or '',
                'numero_expediente': mov.numero_expediente or '',
                'tipo_mov': tipo_grupo,  # ISS-FIX: tipo del movimiento individual
            })
            transacciones[grupo_key]['total_productos'] += 1
            transacciones[grupo_key]['total_cantidad'] += amount
        
        # Convertir a lista ordenada por fecha
        datos = list(transacciones.values())
        datos.sort(key=lambda x: x['fecha_raw'], reverse=True)
        
        # ISS-FIX: Contar transacciones por tipo (basado en el primer movimiento de cada grupo)
        trans_entradas = sum(1 for t in datos if t['_tipo_transaccion'] == 'ENTRADA')
        trans_salidas = sum(1 for t in datos if t['_tipo_transaccion'] == 'SALIDA')
        
        # Limpiar campos internos antes de enviar
        for item in datos:
            del item['fecha_raw']
            del item['_tipo_transaccion']
        
        # ISS-FIX: El resumen usa los contadores calculados por MOVIMIENTO INDIVIDUAL
        # NO por tipo de transacción, para que los totales cuadren con la suma de la tabla
        resumen = {
            'total_transacciones': len(datos),
            'total_movimientos': total_movs_entrada + total_movs_salida,
            'trans_entradas': trans_entradas,  # Transacciones tipo entrada
            'trans_salidas': trans_salidas,    # Transacciones tipo salida
            'total_entradas': total_unidades_entrada,  # Unidades de entrada (suma de cantidades)
            'total_salidas': total_unidades_salida,    # Unidades de salida (suma de cantidades)
            'diferencia': total_unidades_entrada - total_unidades_salida,
        }
        
        # Formato JSON
        if formato == 'json':
            return Response({
                'datos': datos,
                'resumen': resumen
            })
        
        # Formato PDF
        if formato == 'pdf':
            from core.utils.pdf_reports import generar_reporte_movimientos
            
            # Los datos ya están agrupados por transacción, pasarlos directamente
            filtros = {
                'fecha_generacion': timezone.now().strftime('%d/%m/%Y %H:%M')
            }
            if fecha_inicio:
                filtros['fecha_inicio'] = fecha_inicio
            if fecha_fin:
                filtros['fecha_fin'] = fecha_fin
            if tipo:
                filtros['tipo'] = tipo
            if filtrar_por_centro and user_centro:
                filtros['centro'] = user_centro.nombre
            
            # Pasar resumen y datos agrupados al generador PDF
            pdf_buffer = generar_reporte_movimientos(datos, filtros=filtros, resumen=resumen)
            
            response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = f"attachment; filename=Movimientos_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            return response
        
        if formato == 'excel':
            # Importar estilos adicionales para bordes
            from openpyxl.styles import Border, Side
            
            # Generar Excel con detalle de productos (una fila por producto)
            wb = openpyxl.Workbook()
            
            # === HOJA PRINCIPAL: DETALLE POR PRODUCTO ===
            ws = wb.active
            ws.title = 'Movimientos Detalle'
            
            # Titulo - actualizado para 11 columnas
            ws.merge_cells('A1:K1')
            titulo_cell = ws['A1']
            titulo_cell.value = 'REPORTE DE MOVIMIENTOS CON DETALLE'
            titulo_cell.font = Font(bold=True, size=14, color='632842')
            titulo_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Filtros aplicados
            filtros_text = []
            if fecha_inicio:
                filtros_text.append(f'Desde: {fecha_inicio}')
            if fecha_fin:
                filtros_text.append(f'Hasta: {fecha_fin}')
            if tipo:
                filtros_text.append(f'Tipo: {tipo}')
            
            ws.merge_cells('A2:K2')
            filtros_cell = ws['A2']
            filtros_cell.value = ' | '.join(filtros_text) if filtros_text else 'Sin filtros'
            filtros_cell.font = Font(size=10, italic=True)
            filtros_cell.alignment = Alignment(horizontal='center')
            
            # Resumen
            ws['A3'] = f"Total Transacciones: {resumen['total_transacciones']}"
            ws['D3'] = f"Total Entradas: {resumen['total_entradas']}"
            ws['G3'] = f"Total Salidas: {resumen['total_salidas']}"
            ws['J3'] = f"Diferencia: {resumen['diferencia']}"
            for col in ['A', 'D', 'G', 'J']:
                ws[f'{col}3'].font = Font(bold=True, size=10)
            
            ws.append([])  # Linea en blanco
            
            # Encabezados - Una fila por producto
            headers = ['#', 'Referencia', 'Fecha', 'Tipo', 'Subtipo', 'Clave', 'Producto', 'Lote', 'Cantidad', 'Centro', 'No. Expediente']
            ws.append(headers)
            
            # Estilo encabezados
            header_fill = PatternFill(start_color='632842', end_color='632842', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=10)
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for cell in ws[5]:  # Fila 5 tiene los encabezados
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
            
            # Datos - Una fila por cada producto en cada transacción
            row_num = 6
            global_idx = 1
            for trans in datos:
                detalles = trans.get('detalles', [])
                
                if not detalles:
                    # Transacción sin detalles
                    ws.append([
                        global_idx,
                        trans['referencia'],
                        trans['fecha'],
                        trans['tipo'],
                        trans.get('subtipo_display', ''),
                        '-',
                        'Sin productos',
                        '-',
                        0,
                        trans['centro_destino'] if trans['tipo'] == 'SALIDA' else trans['centro_origen'],
                        trans.get('numero_expediente', '')
                    ])
                    for cell in ws[row_num]:
                        cell.border = thin_border
                    row_num += 1
                    global_idx += 1
                else:
                    # Una fila por cada producto
                    first_row = True
                    for det in detalles:
                        # Extraer clave y nombre del producto (formato: "CLAVE - NOMBRE")
                        producto_full = det.get('producto', 'N/A')
                        if ' - ' in producto_full:
                            clave, nombre = producto_full.split(' - ', 1)
                        else:
                            clave = 'N/A'
                            nombre = producto_full
                        
                        ws.append([
                            global_idx if first_row else '',
                            trans['referencia'] if first_row else '',
                            trans['fecha'] if first_row else '',
                            trans['tipo'] if first_row else '',
                            trans.get('subtipo_display', '') if first_row else '',
                            clave,
                            nombre[:40],
                            det.get('lote', 'N/A'),
                            det.get('cantidad', 0),
                            trans['centro_destino'] if trans['tipo'] == 'SALIDA' else trans['centro_origen'] if first_row else '',
                            det.get('numero_expediente', '') or (trans.get('numero_expediente', '') if first_row else '')
                        ])
                        
                        # Colorear tipo
                        if first_row:
                            tipo_cell = ws.cell(row=row_num, column=4)
                            if trans['tipo'].upper() == 'ENTRADA':
                                tipo_cell.fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
                                tipo_cell.font = Font(color='155724', bold=True)
                            else:
                                tipo_cell.fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')
                                tipo_cell.font = Font(color='721C24', bold=True)
                        
                        for cell in ws[row_num]:
                            cell.border = thin_border
                        
                        first_row = False
                        row_num += 1
                    global_idx += 1
            
            # Ajustar anchos - 11 columnas
            column_widths = {'A': 5, 'B': 20, 'C': 16, 'D': 10, 'E': 15, 'F': 12, 'G': 35, 'H': 14, 'I': 10, 'J': 22, 'K': 14}
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename=Movimientos_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            wb.save(response)
            
            logger.info(f"Reporte Excel generado: {len(datos)} transacciones")
            
            return response
            
        # Formato no soportado
        return Response({
            'error': 'Formato no soportado',
            'formatos_disponibles': ['json', 'pdf', 'excel']
        }, status=status.HTTP_400_BAD_REQUEST)
            
    except Exception as e:
        logger.error(f"Error generando reporte: {str(e)}")
        # traceback removido por seguridad (ISS-008)
        return Response({
            'error': 'Error al generar reporte de movimientos',
            'mensaje': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
def reporte_caducidades(request):
    """
    Genera reporte de lotes proximos a caducar.
    Por defecto devuelve JSON. Con ?formato=excel devuelve Excel.
    
    SEGURIDAD: Filtra por centro del usuario si no es admin/farmacia.
    Admin/farmacia puede usar ?centro=ID para filtrar.
    
    Parmetros:
    - dias: Nmero de das de anticipacin (default: 30)
    - centro: ID del centro o 'central' para farmacia central
    - estado: vencido, critico, proximo (filtra por estado especfico)
    - formato: json (default), excel, pdf
    """
    try:
        # SEGURIDAD: Determinar filtro de centro
        user = request.user
        filtrar_por_centro = not is_farmacia_or_admin(user)
        user_centro = get_user_centro(user) if filtrar_por_centro else None
        
        # Admin/farmacia puede filtrar por centro específico
        centro_param = request.query_params.get('centro')
        if centro_param and is_farmacia_or_admin(user):
            if centro_param == 'todos':
                # ISS-FIX: 'todos' significa NO filtrar por centro (ver todos)
                filtrar_por_centro = False
                user_centro = None
            elif centro_param == 'central':
                filtrar_por_centro = True
                user_centro = None
            else:
                try:
                    user_centro = Centro.objects.get(pk=centro_param)
                    filtrar_por_centro = True
                except (Centro.DoesNotExist, ValueError):
                    # ISS-FIX: ValueError para manejar IDs no numéricos
                    pass
        
        dias = int(request.query_params.get('dias', 30))
        formato = request.query_params.get('formato', 'json')
        estado_filtro = request.query_params.get('estado', '').lower().strip()
        
        fecha_limite = date.today() + timedelta(days=dias)
        
        # Obtener lotes proximos a vencer (solo lotes activos)
        lotes = Lote.objects.filter(
            activo=True,
            cantidad_actual__gt=0,
            fecha_caducidad__lte=fecha_limite
        ).select_related('producto')
        
        # Aplicar filtro de centro
        if filtrar_por_centro:
            if user_centro:
                lotes = lotes.filter(centro=user_centro)
            else:
                lotes = lotes.filter(centro__isnull=True)
        
        lotes = lotes.order_by('fecha_caducidad')
        
        # Construir datos
        datos = []
        vencidos = 0
        criticos = 0
        proximos = 0
        
        for lote in lotes:
            dias_restantes = (lote.fecha_caducidad - date.today()).days
            
            if dias_restantes < 0:
                estado = 'vencido'
            elif dias_restantes <= 7:
                estado = 'critico'
            else:
                estado = 'proximo'
            
            # Filtrar por estado si se especificó
            if estado_filtro and estado != estado_filtro:
                continue
            
            # Contadores DESPUÉS del filtro para reflejar datos reales
            if estado == 'vencido':
                vencidos += 1
            elif estado == 'critico':
                criticos += 1
            else:
                proximos += 1
            
            datos.append({
                'producto': f"{lote.producto.clave} - {lote.producto.nombre}",
                'lote': lote.numero_lote,
                'caducidad': lote.fecha_caducidad.isoformat(),
                'dias_restantes': dias_restantes,
                'stock': lote.cantidad_actual,
                'estado': estado,
            })
        
        resumen = {
            'total': len(datos),
            'vencidos': vencidos,
            'criticos': criticos,
            'proximos': proximos,
            'dias_filtro': dias,
        }
        
        # Si formato es JSON, devolver datos
        if formato == 'json':
            return Response({
                'datos': datos,
                'resumen': resumen
            })
        
        # Formato PDF
        if formato == 'pdf':
            from core.utils.pdf_reports import generar_reporte_caducidades
            
            # Preparar datos para el generador PDF
            lotes_data = []
            for item in datos:
                lotes_data.append({
                    'producto': item['producto'],
                    'numero_lote': item['lote'],
                    'fecha_caducidad': item['caducidad'],
                    'dias_restantes': item['dias_restantes'],
                    'cantidad_actual': item['stock'],
                    'estado': item['estado']
                })
            
            filtros = {
                'dias_anticipacion': dias,
                'fecha_generacion': timezone.now().strftime('%d/%m/%Y %H:%M')
            }
            if filtrar_por_centro and user_centro:
                filtros['centro'] = user_centro.nombre
            
            pdf_buffer = generar_reporte_caducidades(lotes_data, dias=dias, filtros=filtros)
            
            response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = f"attachment; filename=Caducidades_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            return response
        
        # Generar Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Caducidades'
        
        ws.merge_cells('A1:G1')
        titulo_cell = ws['A1']
        titulo_cell.value = f'REPORTE DE LOTES PROXIMOS A CADUCAR ({dias} DIAS)'
        titulo_cell.font = Font(bold=True, size=14, color='632842')
        titulo_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        ws.merge_cells('A2:G2')
        fecha_cell = ws['A2']
        fecha_cell.value = f'Generado el {timezone.now().strftime("%d/%m/%Y %H:%M")}'
        fecha_cell.font = Font(size=10, italic=True)
        fecha_cell.alignment = Alignment(horizontal='center')
        
        ws.append([])
        
        headers = ['#', 'Producto', 'Lote', 'Caducidad', 'Das Restantes', 'Stock', 'Estado']
        ws.append(headers)
        
        header_fill = PatternFill(start_color='632842', end_color='632842', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        
        for cell in ws[4]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for idx, item in enumerate(datos, 1):
            ws.append([
                idx,
                item['producto'][:50],
                item['lote'],
                item['caducidad'],
                item['dias_restantes'],
                item['stock'],
                item['estado'].upper()
            ])
        
        ws.append([])
        resumen_row = ws.max_row + 1
        ws[f'B{resumen_row}'] = 'Total:'
        ws[f'C{resumen_row}'] = resumen['total']
        ws[f'B{resumen_row + 1}'] = 'Vencidos:'
        ws[f'C{resumen_row + 1}'] = resumen['vencidos']
        ws[f'B{resumen_row + 2}'] = 'Crticos:'
        ws[f'C{resumen_row + 2}'] = resumen['criticos']
        ws[f'B{resumen_row + 3}'] = 'Prximos:'
        ws[f'C{resumen_row + 3}'] = resumen['proximos']
        
        for col, width in zip(['A','B','C','D','E','F','G'], [8,50,20,15,15,12,12]):
            ws.column_dimensions[col].width = width
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=Caducidades_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        wb.save(response)
        
        return response
        
    except Exception as e:
        # traceback removido por seguridad (ISS-008)
        return Response({
            'error': 'Error al generar reporte de caducidades',
            'mensaje': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def reporte_requisiciones(request):
    """
    Genera reporte de requisiciones con filtros.
    Por defecto devuelve JSON. Con ?formato=excel devuelve Excel.
    
    SEGURIDAD: Filtra por centro del usuario si no es admin/farmacia.
    Admin/farmacia puede usar ?centro=ID para filtrar.
    
    Parametros:
    - fecha_inicio: Fecha inicial (YYYY-MM-DD)
    - fecha_fin: Fecha final (YYYY-MM-DD)
    - estado: Estado de la requisicion
    - centro: ID del centro (solo admin/farmacia)
    - formato: json/excel
    """
    try:
        # SEGURIDAD: Determinar filtro de centro
        user = request.user
        filtrar_por_centro = not is_farmacia_or_admin(user)
        user_centro = get_user_centro(user) if filtrar_por_centro else None
        
        fecha_inicio = request.query_params.get('fecha_inicio')
        fecha_fin = request.query_params.get('fecha_fin')
        estado = request.query_params.get('estado')
        centro_param = request.query_params.get('centro') or request.query_params.get('centro_id')
        # Ignorar 'todos' como parámetro de centro
        if centro_param and centro_param.lower() == 'todos':
            centro_param = None
        formato = request.query_params.get('formato', 'json')
        
        requisiciones = Requisicion.objects.select_related('centro_origen', 'centro_destino', 'solicitante').all()
        
        # Aplicar filtro de centro obligatorio para usuarios de centro
        if filtrar_por_centro and user_centro:
            requisiciones = requisiciones.filter(Q(centro_origen=user_centro) | Q(centro_destino=user_centro))
        elif centro_param and is_farmacia_or_admin(user):
            requisiciones = requisiciones.filter(Q(centro_origen_id=centro_param) | Q(centro_destino_id=centro_param))
        
        if fecha_inicio:
            requisiciones = requisiciones.filter(fecha_solicitud__gte=fecha_inicio)
        if fecha_fin:
            requisiciones = requisiciones.filter(fecha_solicitud__lte=fecha_fin)
        if estado:
            requisiciones = requisiciones.filter(estado__iexact=estado)
        
        requisiciones = requisiciones.order_by('-fecha_solicitud')
        
        # Construir datos
        datos = []
        estados_count = {}
        
        for req in requisiciones:
            estado_req = req.estado.upper()
            estados_count[estado_req] = estados_count.get(estado_req, 0) + 1
            
            # Determinar centro (preferir destino, luego origen)
            centro_nombre = 'N/A'
            if req.centro_destino:
                centro_nombre = req.centro_destino.nombre
            elif req.centro_origen:
                centro_nombre = req.centro_origen.nombre
            
            # Obtener detalle de productos
            detalles_productos = []
            for detalle in req.detalles.select_related('producto').all():
                detalles_productos.append({
                    'clave': detalle.producto.clave if detalle.producto else 'N/A',
                    'nombre': detalle.producto.nombre if detalle.producto else 'N/A',
                    'cantidad_solicitada': detalle.cantidad_solicitada or 0,
                    'cantidad_autorizada': detalle.cantidad_autorizada or 0,
                    'cantidad_surtida': detalle.cantidad_surtida or 0,
                })
            
            datos.append({
                'id': req.id,
                'folio': req.folio or f'REQ-{req.id}',
                'centro': centro_nombre,
                'estado': estado_req,
                'fecha_solicitud': req.fecha_solicitud.isoformat() if req.fecha_solicitud else None,
                'total_productos': req.detalles.count(),
                'solicitante': req.solicitante.get_full_name() if req.solicitante else 'N/A',
                'productos': detalles_productos,  # Agregar detalle de productos
            })
        
        resumen = {
            'total': len(datos),
            'por_estado': estados_count,
        }
        
        # Si formato es JSON, devolver datos
        if formato == 'json':
            return Response({
                'datos': datos,
                'resumen': resumen
            })
        
        # Formato PDF
        if formato == 'pdf':
            from core.utils.pdf_reports import generar_reporte_requisiciones
            
            # Preparar datos para el generador PDF - INCLUIR PRODUCTOS
            requisiciones_data = []
            for item in datos:
                requisiciones_data.append({
                    'folio': item['folio'],
                    'centro': item['centro'],
                    'estado': item['estado'],
                    'fecha_solicitud': item['fecha_solicitud'],
                    'total_productos': item['total_productos'],
                    'solicitante': item['solicitante'],
                    'productos': item.get('productos', []),  # INCLUIR detalle de productos
                })
            
            filtros = {
                'fecha_generacion': timezone.now().strftime('%d/%m/%Y %H:%M')
            }
            if fecha_inicio:
                filtros['fecha_inicio'] = fecha_inicio
            if fecha_fin:
                filtros['fecha_fin'] = fecha_fin
            if estado:
                filtros['estado'] = estado
            if filtrar_por_centro and user_centro:
                filtros['centro'] = user_centro.nombre
            
            pdf_buffer = generar_reporte_requisiciones(requisiciones_data, filtros=filtros)
            
            response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = f"attachment; filename=Requisiciones_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            return response
        
        # Importar estilos adicionales para bordes
        from openpyxl.styles import Border, Side
        
        # Generar Excel con detalles de productos
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Requisiciones'
        
        ws.merge_cells('A1:J1')
        titulo_cell = ws['A1']
        titulo_cell.value = 'REPORTE DE REQUISICIONES CON DETALLE'
        titulo_cell.font = Font(bold=True, size=14, color='632842')
        titulo_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        ws.merge_cells('A2:J2')
        fecha_cell = ws['A2']
        fecha_cell.value = f'Generado el {timezone.now().strftime("%d/%m/%Y %H:%M")}'
        fecha_cell.font = Font(size=10, italic=True)
        fecha_cell.alignment = Alignment(horizontal='center')
        
        ws.append([])
        
        # Encabezados con columnas de detalle de producto
        headers = ['#', 'Folio', 'Centro', 'Fecha', 'Estado', 'Solicitante', 'Clave Producto', 'Producto', 'Cant. Solicitada', 'Cant. Autorizada']
        ws.append(headers)
        
        header_fill = PatternFill(start_color='632842', end_color='632842', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=10)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for cell in ws[4]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        row_num = 5
        for idx, item in enumerate(datos, 1):
            productos = item.get('productos', [])
            
            if not productos:
                # Requisición sin productos
                ws.append([
                    idx,
                    item['folio'],
                    item['centro'],
                    item['fecha_solicitud'][:10] if item['fecha_solicitud'] else 'N/A',
                    item['estado'],
                    item['solicitante'],
                    '-',
                    'Sin productos',
                    0,
                    0
                ])
                for cell in ws[row_num]:
                    cell.border = thin_border
                row_num += 1
            else:
                # Una fila por cada producto
                first_row = True
                for prod in productos:
                    ws.append([
                        idx if first_row else '',
                        item['folio'] if first_row else '',
                        item['centro'] if first_row else '',
                        item['fecha_solicitud'][:10] if first_row and item['fecha_solicitud'] else ('' if not first_row else 'N/A'),
                        item['estado'] if first_row else '',
                        item['solicitante'] if first_row else '',
                        prod.get('clave', 'N/A'),
                        prod.get('nombre', 'N/A'),
                        prod.get('cantidad_solicitada', 0),
                        prod.get('cantidad_autorizada', 0)
                    ])
                    for cell in ws[row_num]:
                        cell.border = thin_border
                    first_row = False
                    row_num += 1
        
        # Ajustar anchos de columna
        for col, width in zip(['A','B','C','D','E','F','G','H','I','J'], [5,20,30,12,12,30,15,40,15,15]):
            ws.column_dimensions[col].width = width
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=Requisiciones_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        wb.save(response)
        
        return response
        
    except Exception as e:
        # traceback removido por seguridad (ISS-008)
        return Response({
            'error': 'Error al generar reporte de requisiciones',
            'mensaje': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def reporte_medicamentos_por_caducar(request):
    """Resumen JSON de productos con lotes proximos a caducar.
    
    SEGURIDAD: Filtra por centro del usuario si no es admin/farmacia.
    """
    try:
        # SEGURIDAD: Determinar filtro de centro
        user = request.user
        filtrar_por_centro = not is_farmacia_or_admin(user)
        user_centro = get_user_centro(user) if filtrar_por_centro else None
        
        centro_param = request.query_params.get('centro')
        if centro_param and is_farmacia_or_admin(user):
            if centro_param.lower() in ('central', 'todos'):
                if centro_param.lower() == 'central':
                    user_centro = None
                else:
                    filtrar_por_centro = False
            else:
                try:
                    # ISS-FIX: Buscar por ID o nombre
                    if centro_param.isdigit():
                        user_centro = Centro.objects.get(pk=centro_param)
                    else:
                        user_centro = Centro.objects.get(nombre__iexact=centro_param)
                    filtrar_por_centro = True
                except Centro.DoesNotExist:
                    pass
        
        dias = int(request.query_params.get('dias', 30))
        hoy = date.today()
        limite = hoy + timedelta(days=dias)
        lotes = Lote.objects.filter(
            cantidad_actual__gt=0,
            fecha_caducidad__gt=hoy,
            fecha_caducidad__lte=limite
        ).select_related('producto')
        
        # Aplicar filtro de centro
        if filtrar_por_centro and user_centro:
            lotes = lotes.filter(centro=user_centro)

        agregados = {}
        for lote in lotes:
            prod = lote.producto
            key = prod.id
            entry = agregados.setdefault(key, {
                'producto_id': prod.id,
                'clave': prod.clave,
                'descripcion': prod.descripcion,
                'stock_total': 0,
                'lotes': 0,
                'primer_vencimiento': None,
            })
            entry['lotes'] += 1
            entry['stock_total'] += lote.cantidad_actual
            fecha = lote.fecha_caducidad
            if entry['primer_vencimiento'] is None or fecha < entry['primer_vencimiento']:
                entry['primer_vencimiento'] = fecha

        resultados = sorted(
            agregados.values(),
            key=lambda x: x.get('primer_vencimiento') or limite
        )
        for res in resultados:
            fv = res['primer_vencimiento']
            res['primer_vencimiento'] = fv.isoformat() if fv else None

        return Response({
            'total_productos': len(resultados),
            'total_lotes': lotes.count(),
            'dias_configurados': dias,
            'resultados': resultados
        })
    except Exception as exc:
        # traceback removido por seguridad (ISS-008)
        return Response({'error': 'Error al obtener medicamentos por caducar', 'mensaje': str(exc)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def reporte_bajo_stock(request):
    """Productos con stock por debajo del minimo.
    
    SEGURIDAD: Filtra por centro del usuario si no es admin/farmacia.
    """
    try:
        # SEGURIDAD: Determinar filtro de centro
        user = request.user
        filtrar_por_centro = not is_farmacia_or_admin(user)
        user_centro = get_user_centro(user) if filtrar_por_centro else None
        
        centro_param = request.query_params.get('centro')
        if centro_param and is_farmacia_or_admin(user):
            if centro_param.lower() in ('central', 'todos'):
                if centro_param.lower() == 'central':
                    user_centro = None
                else:
                    filtrar_por_centro = False
            else:
                try:
                    if centro_param.isdigit():
                        user_centro = Centro.objects.get(pk=centro_param)
                    else:
                        user_centro = Centro.objects.get(nombre__iexact=centro_param)
                    filtrar_por_centro = True
                except Centro.DoesNotExist:
                    pass
        
        productos = Producto.objects.filter(activo=True)
        resultados = []
        for prod in productos:
            lotes_query = prod.lotes.filter(
                activo=True
            )
            # Aplicar filtro de centro
            if filtrar_por_centro and user_centro:
                lotes_query = lotes_query.filter(centro=user_centro)
            
            stock = lotes_query.aggregate(total=Sum('cantidad_actual'))['total'] or 0
            if stock < prod.stock_minimo:
                resultados.append({
                    'producto_id': prod.id,
                    'clave': prod.clave,
                    'descripcion': prod.descripcion,
                    'stock_actual': stock,
                    'stock_minimo': prod.stock_minimo,
                    'diferencia': prod.stock_minimo - stock
                })
        resultados = sorted(resultados, key=lambda x: x['diferencia'], reverse=True)
        return Response({'total': len(resultados), 'resultados': resultados})
    except Exception as exc:
        # traceback removido por seguridad (ISS-008)
        return Response({'error': 'Error al obtener productos en bajo stock', 'mensaje': str(exc)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def reporte_consumo(request):
    """
    Consumo (salidas) por producto en un rango de fechas.
    
    SEGURIDAD: Filtra por centro del usuario si no es admin/farmacia.
    """
    try:
        # SEGURIDAD: Determinar filtro de centro
        user = request.user
        filtrar_por_centro = not is_farmacia_or_admin(user)
        user_centro = get_user_centro(user) if filtrar_por_centro else None
        
        centro_param = request.query_params.get('centro')
        if centro_param and centro_param.lower() != 'todos' and is_farmacia_or_admin(user):
            try:
                user_centro = Centro.objects.get(pk=centro_param)
                filtrar_por_centro = True
            except Centro.DoesNotExist:
                pass
        
        fecha_inicio = request.query_params.get('fecha_inicio')
        fecha_fin = request.query_params.get('fecha_fin')
        movimientos = Movimiento.objects.select_related('lote__producto', 'centro_origen', 'centro_destino').filter(tipo='salida')
        
        # ISS-FIX: Aplicar filtro de centro de forma ESTRICTA
        # Para salidas: solo donde el centro es ORIGEN (salidas DESDE ese centro)
        if filtrar_por_centro and user_centro:
            movimientos = movimientos.filter(centro_origen=user_centro)
        
        if fecha_inicio:
            movimientos = movimientos.filter(fecha__gte=fecha_inicio)
        if fecha_fin:
            movimientos = movimientos.filter(fecha__lte=fecha_fin)

        agregados = {}
        for mov in movimientos:
            prod = getattr(mov.lote, 'producto', None)
            if not prod:
                continue
            key = prod.id
            entry = agregados.setdefault(key, {
                'producto_id': prod.id,
                'clave': prod.clave,
                'descripcion': prod.descripcion,
                'total_salidas': 0
            })
            entry['total_salidas'] += abs(mov.cantidad or 0)

        resultados = sorted(agregados.values(), key=lambda x: x['total_salidas'], reverse=True)
        return Response({'total_productos': len(resultados), 'resultados': resultados})
    except Exception as exc:
        # traceback removido por seguridad (ISS-008)
        return Response({'error': 'Error al obtener consumo', 'mensaje': str(exc)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@api_view(['GET'])
def reportes_precarga(request):
    """
    Obtiene datos para precargar formularios de reportes.
    
    SEGURIDAD: Filtra centros y lotes segn el rol del usuario.
    - Admin/farmacia: ven todos los centros
    - Usuario de centro: solo ve su centro
    
    Retorna:
    - Lista de productos activos
    - Lista de centros activos (filtrada segn rol)
    - Tipos de movimiento disponibles
    """
    try:
        # SEGURIDAD: Determinar filtro de centro
        user = request.user
        es_admin_farmacia = is_farmacia_or_admin(user)
        user_centro = get_user_centro(user) if not es_admin_farmacia else None
        
        # Usar clave como identificador principal del producto
        productos_qs = Producto.objects.filter(activo=True).values('id', 'clave', 'descripcion').order_by('clave')
        productos = [{'id': p['id'], 'clave': p['clave'], 'descripcion': p['descripcion']} for p in productos_qs]
        
        # Filtrar centros segun rol
        if es_admin_farmacia:
            # Centro no tiene campo 'clave', usar id como identificador
            centros = list(Centro.objects.filter(activo=True).values('id', 'nombre').order_by('nombre'))
            # Agregar 'clave' basado en id para compatibilidad con frontend
            for c in centros:
                c['clave'] = str(c['id'])
        elif user_centro:
            centros = [{'id': user_centro.id, 'clave': user_centro.clave, 'nombre': user_centro.nombre}]
        else:
            centros = []
        
        # Filtrar lotes segun rol
        lotes_query = Lote.objects.filter(activo=True)
        if not es_admin_farmacia and user_centro:
            lotes_query = lotes_query.filter(centro=user_centro)
        lotes = list(lotes_query.values('id', 'numero_lote', 'producto_id'))
        
        return Response({
            'productos': productos,
            'centros': centros,
            'lotes': lotes,
            'tipos_movimiento': ['ENTRADA', 'SALIDA']
        })
        
    except Exception as e:
        # traceback removido por seguridad (ISS-008)
        return Response({
            'error': 'Error al obtener datos de precarga',
            'mensaje': str(e)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class HojaRecoleccionViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet de solo lectura para Hojas de Recoleccion.
    Las hojas se generan automaticamente al autorizar requisiciones.
    """
    queryset = HojaRecoleccion.objects.select_related(
        'centro', 'responsable'
    ).prefetch_related('detalles')
    serializer_class = HojaRecoleccionSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = CustomPagination

    def get_queryset(self):
        """Filtra por centro si el usuario no es admin/farmacia."""
        queryset = super().get_queryset().order_by('-created_at')
        user = self.request.user
        if not user or not user.is_authenticated:
            return HojaRecoleccion.objects.none()
        if user.is_superuser:
            return queryset
        rol = getattr(user, 'rol', '').lower()
        if rol in ('admin', 'farmacia', 'administrador', 'usuario_farmacia'):
            return queryset
        # Filtrar por centro del usuario
        user_centro = getattr(user, 'centro', None)
        if user_centro:
            return queryset.filter(centro=user_centro)
        return HojaRecoleccion.objects.none()

    @action(detail=True, methods=['get'])
    def verificar_integridad(self, request, pk=None):
        """Verifica que el hash de la hoja coincida con su contenido."""
        import hashlib
        import json
        hoja = self.get_object()
        contenido = json.dumps(hoja.contenido_json, sort_keys=True, ensure_ascii=False)
        hash_calculado = hashlib.sha256(contenido.encode('utf-8')).hexdigest()
        return Response({
            'folio': hoja.numero,
            'hash_almacenado': hoja.hash_contenido,
            'hash_calculado': hash_calculado,
            'integridad_ok': hash_calculado == hoja.hash_contenido
        })

    @action(detail=True, methods=['get'])
    def pdf(self, request, pk=None):
        """Genera y descarga el PDF de la hoja de recolección."""
        from core.utils.pdf_generator import generar_hoja_recoleccion
        
        hoja = self.get_object()
        # ISS-FIX: Obtener requisición a través de detalles ya que no hay relación directa
        detalle = hoja.detalles.select_related('requisicion').first()
        if not detalle or not detalle.requisicion:
            return Response({
                'error': 'No se encontró requisición asociada a esta hoja',
            }, status=status.HTTP_404_NOT_FOUND)
        requisicion = detalle.requisicion
        
        try:
            pdf_buffer = generar_hoja_recoleccion(requisicion)
            
            response = HttpResponse(
                pdf_buffer.getvalue(),
                content_type='application/pdf'
            )
            folio_safe = (hoja.numero or f'HOJA-{hoja.id}').replace('/', '-')
            response['Content-Disposition'] = f'attachment; filename="Hoja_Recoleccion_{folio_safe}.pdf"'
            
            return response
            
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            return Response({
                'error': 'Error al generar PDF',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path='por-requisicion/(?P<requisicion_id>[^/.]+)')
    def por_requisicion(self, request, requisicion_id=None):
        """
        Obtiene la hoja de recolección asociada a una requisición específica.
        ISS-FIX: Buscar a través de detalles__requisicion__id (FK a Requisicion)
        ISS-FIX-500: Siempre devolver 200 con existe=false si hay error
        """
        try:
            requisicion_id = int(requisicion_id)
        except (ValueError, TypeError):
            return Response({
                'existe': False,
                'requisicion_id': requisicion_id,
                'mensaje': 'ID de requisición inválido'
            }, status=status.HTTP_200_OK)
            
        try:
            hoja = self.get_queryset().filter(detalles__requisicion__id=requisicion_id).distinct().first()
            if hoja:
                return Response({
                    'existe': True,
                    'hoja': HojaRecoleccionSerializer(hoja, context={'request': request}).data
                })
            return Response({
                'existe': False,
                'requisicion_id': requisicion_id,
                'mensaje': 'No existe hoja de recolección para esta requisición'
            }, status=status.HTTP_200_OK)
        except Exception as e:
            logger.error(f"Error obteniendo hoja por requisición {requisicion_id}: {e}")
            return Response({
                'existe': False,
                'requisicion_id': requisicion_id,
                'mensaje': 'Error al buscar hoja de recolección'
            }, status=status.HTTP_200_OK)  # ISS-FIX-500: Siempre 200


# ============================================
# BÚSQUEDA UNIFICADA DE TRAZABILIDAD
# ============================================

@api_view(['GET'])
def trazabilidad_buscar(request):
    """
    Búsqueda unificada de trazabilidad: busca por lote, producto o clave.
    
    Primero intenta buscar como lote, luego como producto.
    Retorna el tipo encontrado y los datos correspondientes.
    
    Query params:
    - q: término de búsqueda (requerido)
    - centro: filtro de centro (opcional, solo admin/farmacia)
    
    SEGURIDAD: Filtra por centro del usuario si no es admin/farmacia.
    """
    try:
        user = request.user
        if not user or not user.is_authenticated:
            return Response({'error': 'Autenticación requerida'}, status=status.HTTP_403_FORBIDDEN)
        
        rol_usuario = (getattr(user, 'rol', '') or '').lower()
        if rol_usuario == 'vista':
            return Response({'error': 'No tienes permiso para trazabilidad'}, status=status.HTTP_403_FORBIDDEN)
        
        query = request.query_params.get('q', '').strip()
        if not query:
            return Response({'error': 'Parámetro de búsqueda requerido'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Determinar filtro de centro
        es_admin_farmacia = is_farmacia_or_admin(user)
        filtrar_por_centro = not es_admin_farmacia
        user_centro = get_user_centro(user) if filtrar_por_centro else None
        
        centro_param = request.query_params.get('centro')
        if centro_param and es_admin_farmacia:
            if centro_param != 'todos':
                try:
                    user_centro = Centro.objects.get(pk=centro_param)
                    filtrar_por_centro = True
                except Centro.DoesNotExist:
                    pass
        
        # 1. Intentar buscar como lote (solo admin/farmacia)
        if es_admin_farmacia:
            lote_query = Lote.objects.select_related('producto', 'centro').filter(
                Q(numero_lote__iexact=query) |
                Q(numero_lote__icontains=query)
            )
            if filtrar_por_centro and user_centro:
                lote_query = lote_query.filter(centro=user_centro)
            
            lote = lote_query.first()
            if lote:
                # Encontrado como lote - retornar tipo 'lote'
                return Response({
                    'tipo': 'lote',
                    'encontrado': True,
                    'identificador': lote.numero_lote,
                    'id': lote.id,
                    'datos': {
                        'numero_lote': lote.numero_lote,
                        'producto_clave': lote.producto.clave,
                        'producto_nombre': lote.producto.nombre,
                    }
                })
        
        # 2. Buscar como producto (clave o nombre)
        producto = Producto.objects.filter(
            Q(clave__iexact=query) |
            Q(clave__icontains=query) |
            Q(nombre__icontains=query)
        ).first()
        
        if producto:
            # Verificar si hay lotes accesibles para este usuario
            lotes_accesibles = Lote.objects.filter(producto=producto, activo=True)
            if filtrar_por_centro and user_centro:
                lotes_accesibles = lotes_accesibles.filter(centro=user_centro)
            
            if lotes_accesibles.exists() or es_admin_farmacia:
                return Response({
                    'tipo': 'producto',
                    'encontrado': True,
                    'identificador': producto.clave,
                    'id': producto.id,
                    'datos': {
                        'clave': producto.clave,
                        'nombre': producto.nombre,
                        'descripcion': producto.descripcion,
                    }
                })
        
        # 3. No encontrado
        return Response({
            'tipo': None,
            'encontrado': False,
            'identificador': query,
            'mensaje': 'No se encontró producto ni lote con ese término'
        }, status=status.HTTP_404_NOT_FOUND)
        
    except Exception as exc:
        logger.exception('Error en trazabilidad_buscar')
        return Response({
            'error': 'Error al buscar',
            'mensaje': str(exc)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def trazabilidad_autocomplete(request):
    """
    Autocompletado unificado para búsqueda de trazabilidad.
    Retorna productos y lotes que coincidan con el término de búsqueda.
    
    Query params:
    - search: término de búsqueda (mínimo 2 caracteres)
    - centro: filtro de centro (opcional)
    
    SEGURIDAD: 
    - Usuarios de centro solo ven productos con lotes en su centro
    - Solo admin/farmacia ven lotes directamente
    """
    try:
        user = request.user
        if not user or not user.is_authenticated:
            return Response({'results': []})
        
        search = request.query_params.get('search', '').strip()
        if len(search) < 2:
            return Response({'results': []})
        
        es_admin_farmacia = is_farmacia_or_admin(user)
        filtrar_por_centro = not es_admin_farmacia
        user_centro = get_user_centro(user) if filtrar_por_centro else None
        
        centro_param = request.query_params.get('centro')
        if centro_param and es_admin_farmacia and centro_param != 'todos':
            try:
                user_centro = Centro.objects.get(pk=centro_param)
                filtrar_por_centro = True
            except Centro.DoesNotExist:
                pass
        
        results = []
        
        # 1. Buscar productos
        productos_query = Producto.objects.filter(
            activo=True
        ).filter(
            Q(clave__icontains=search) |
            Q(nombre__icontains=search)
        )
        
        # ISS-FIX: Filtrar productos según rol del usuario
        # - Farmacia/Admin: Ver todos los productos (para trazabilidad histórica)
        # - Centro: Solo productos con lotes que tengan stock en su centro
        if filtrar_por_centro and user_centro:
            # Usuarios de Centro: solo productos con lotes CON stock
            productos_con_lotes = Lote.objects.filter(
                centro=user_centro, activo=True, cantidad_actual__gt=0
            ).values_list('producto_id', flat=True).distinct()
            productos_query = productos_query.filter(id__in=productos_con_lotes)
        
        for producto in productos_query[:5]:
            results.append({
                'tipo': 'producto',
                'id': producto.id,
                'identificador': producto.clave,
                'display': f"📦 {producto.clave}",
                'secundario': producto.nombre[:50] + ('...' if len(producto.nombre) > 50 else ''),
            })
        
        # 2. Buscar lotes (solo admin/farmacia para trazabilidad completa)
        # Farmacia/Admin pueden ver lotes históricos sin stock
        if es_admin_farmacia:
            lotes_query = Lote.objects.select_related('producto').filter(
                Q(numero_lote__icontains=search) |
                Q(producto__clave__icontains=search) |
                Q(producto__nombre__icontains=search)
            )
            
            if filtrar_por_centro and user_centro:
                lotes_query = lotes_query.filter(centro=user_centro)
            
            for lote in lotes_query[:5]:
                # Indicar si el lote está agotado
                stock_info = f" (Stock: {lote.cantidad_actual})" if lote.cantidad_actual == 0 else ""
                results.append({
                    'tipo': 'lote',
                    'id': lote.id,
                    'identificador': lote.numero_lote,
                    'display': f"🏷️ {lote.numero_lote}{stock_info}",
                    'secundario': f"{lote.producto.clave} - {lote.producto.nombre[:30]}",
                })
        
        return Response({
            'results': results,
            'count': len(results)
        })
        
    except Exception as exc:
        logger.exception('Error en trazabilidad_autocomplete')
        return Response({'results': [], 'error': str(exc)})


# ============================================
# TRAZABILIDAD GLOBAL Y CON FILTROS DE FECHA
# ============================================

@api_view(['GET'])
def trazabilidad_global(request):
    """
    Reporte global de trazabilidad de todos los lotes.
    
    Filtros soportados:
    - fecha_inicio: Fecha inicio (YYYY-MM-DD)
    - fecha_fin: Fecha fin (YYYY-MM-DD)
    - centro: Filtro de centro:
        * vacío o 'central': Solo movimientos de Farmacia Central (por defecto)
        * 'todos': Todos los movimientos de todos los centros
        * ID numérico: Solo movimientos del centro específico
    - tipo: tipo de movimiento (entrada, salida, ajuste)
    - producto: ID del producto
    - formato: json (default), excel, pdf
    
    SEGURIDAD: Solo admin/farmacia pueden acceder.
    
    IMPORTANTE: Para evitar confusiones en reportes:
    - Por defecto se muestran SOLO movimientos de Farmacia Central
    - Para ver todos los centros juntos, usar centro='todos' explícitamente
    """
    from datetime import datetime, timedelta
    from django.utils import timezone
    
    try:
        user = request.user
        if not user or not user.is_authenticated:
            return Response({'error': 'Autenticación requerida'}, status=status.HTTP_403_FORBIDDEN)
        
        if not is_farmacia_or_admin(user):
            return Response({'error': 'Solo administradores y farmacia pueden acceder a trazabilidad global'}, status=status.HTTP_403_FORBIDDEN)
        
        # Obtener parámetros de filtro
        fecha_inicio = request.query_params.get('fecha_inicio')
        fecha_fin = request.query_params.get('fecha_fin')
        centro_param = request.query_params.get('centro')
        tipo_movimiento = request.query_params.get('tipo')
        producto_param = request.query_params.get('producto')
        formato = request.query_params.get('formato', 'json')
        
        # Construir query base de lotes
        # TRAZABILIDAD: Incluir TODOS los lotes (activos e inactivos) para trazabilidad completa
        # Solo admin/farmacia acceden a este endpoint, necesitan ver historial completo
        lotes_query = Lote.objects.select_related('producto', 'centro')
        
        # Filtrar por centro
        if centro_param:
            if centro_param == 'central':
                lotes_query = lotes_query.filter(centro__isnull=True)
            else:
                try:
                    lotes_query = lotes_query.filter(centro_id=int(centro_param))
                except ValueError:
                    pass
        
        # Filtrar por producto
        if producto_param:
            try:
                lotes_query = lotes_query.filter(producto_id=int(producto_param))
            except ValueError:
                lotes_query = lotes_query.filter(
                    Q(producto__clave__icontains=producto_param) |
                    Q(producto__nombre__icontains=producto_param)
                )
        
        # Construir query de movimientos
        # TRAZABILIDAD: Incluir movimientos de TODOS los lotes (activos e inactivos)
        movimientos_query = Movimiento.objects.select_related(
            'lote', 'lote__producto', 'lote__centro',
            'centro_origen', 'centro_destino', 'usuario'
        )
        
        # FILTRO DE CENTRO - Lógica clara para evitar confusiones:
        # - vacío o 'central': Solo Farmacia Central (centro_id=NULL)
        # - 'todos': Todos los movimientos sin filtrar por centro
        # - ID numérico: Solo ese centro específico
        if not centro_param or centro_param == 'central':
            # Por defecto: Solo movimientos de Farmacia Central
            movimientos_query = movimientos_query.filter(
                Q(lote__centro__isnull=True) |
                Q(centro_origen__isnull=True, centro_destino__isnull=True)
            )
        elif centro_param != 'todos':
            # Centro específico por ID
            try:
                centro_id = int(centro_param)
                movimientos_query = movimientos_query.filter(
                    Q(lote__centro_id=centro_id) |
                    Q(centro_origen_id=centro_id) |
                    Q(centro_destino_id=centro_id)
                )
            except (ValueError, TypeError):
                pass
        # Si centro_param == 'todos': No aplicar filtro de centro (mostrar todo)
        
        # FIX: Filtrar movimientos por fecha usando fecha__date para ignorar hora
        if fecha_inicio:
            try:
                fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
                movimientos_query = movimientos_query.filter(fecha__date__gte=fecha_inicio_dt)
            except ValueError:
                pass
        
        if fecha_fin:
            try:
                fecha_fin_dt = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
                # Usar fecha__date__lte para incluir todo el día automáticamente
                movimientos_query = movimientos_query.filter(fecha__date__lte=fecha_fin_dt)
            except ValueError:
                pass
        
        # NOTA: El filtro de centro ya se aplicó arriba con la nueva lógica
        # (vacío/central = Farmacia Central, todos = sin filtro, ID = centro específico)
        
        # Filtrar por tipo de movimiento
        if tipo_movimiento:
            movimientos_query = movimientos_query.filter(tipo=tipo_movimiento.lower())
        
        # Filtrar por producto en movimientos
        if producto_param:
            try:
                movimientos_query = movimientos_query.filter(lote__producto_id=int(producto_param))
            except ValueError:
                movimientos_query = movimientos_query.filter(
                    Q(lote__producto__clave__icontains=producto_param) |
                    Q(lote__producto__nombre__icontains=producto_param)
                )
        
        movimientos_query = movimientos_query.order_by('-fecha')
        
        # Limitar resultados según formato
        if formato == 'json':
            movimientos_query = movimientos_query[:500]
        else:
            movimientos_query = movimientos_query[:2000]
        
        # Preparar datos de movimientos
        movimientos_data = []
        for mov in movimientos_query:
            # ISS-FIX: Lógica clara para mostrar el centro relevante
            # Para SALIDA: mostrar destino (a dónde va)
            # Para ENTRADA: mostrar origen (de dónde viene)
            tipo_upper = mov.tipo.upper()
            if tipo_upper == 'SALIDA':
                centro_display = mov.centro_destino.nombre if mov.centro_destino else 'Farmacia Central'
            elif tipo_upper == 'ENTRADA':
                centro_display = mov.centro_origen.nombre if mov.centro_origen else 'Farmacia Central'
            else:
                # Ajuste u otro: mostrar cualquiera que tenga
                centro_display = (
                    mov.centro_destino.nombre if mov.centro_destino else 
                    (mov.centro_origen.nombre if mov.centro_origen else 'Farmacia Central')
                )
            
            # ISS-FIX: Mostrar nombre completo del usuario o username
            if mov.usuario:
                usuario_display = mov.usuario.get_full_name()
                if not usuario_display or usuario_display.strip() == '':
                    usuario_display = mov.usuario.username
            else:
                usuario_display = 'Sistema'
            
            movimientos_data.append({
                'id': mov.id,
                'fecha': mov.fecha.isoformat() if mov.fecha else None,
                'fecha_str': mov.fecha.strftime('%d/%m/%Y %H:%M') if mov.fecha else 'N/A',
                'tipo': tipo_upper,
                'cantidad': mov.cantidad,
                'lote': mov.lote.numero_lote if mov.lote else 'N/A',
                'producto_clave': mov.lote.producto.clave if mov.lote and mov.lote.producto else 'N/A',
                'producto_nombre': mov.lote.producto.nombre if mov.lote and mov.lote.producto else 'N/A',
                'centro': centro_display,
                'usuario': usuario_display,
                'observaciones': mov.motivo or '',
                'numero_contrato': mov.lote.numero_contrato if mov.lote else None,
                # ISS-FIX: Campos de trazabilidad para salidas de centro
                'subtipo_salida': getattr(mov, 'subtipo_salida', None) or '',
                'numero_expediente': getattr(mov, 'numero_expediente', None) or '',
            })
        
        # Estadísticas - usar abs() para cantidades porque salidas pueden ser negativas
        total_entradas = sum(abs(m['cantidad']) for m in movimientos_data if m['tipo'] == 'ENTRADA')
        total_salidas = sum(abs(m['cantidad']) for m in movimientos_data if m['tipo'] == 'SALIDA')
        total_ajustes = sum(m['cantidad'] for m in movimientos_data if m['tipo'] == 'AJUSTE')
        
        # Contar lotes únicos
        lotes_unicos = set(m['lote'] for m in movimientos_data if m['lote'] != 'N/A')
        productos_unicos = set(m['producto_clave'] for m in movimientos_data if m['producto_clave'] != 'N/A')
        
        response_data = {
            'movimientos': movimientos_data,
            'total_movimientos': len(movimientos_data),
            'estadisticas': {
                'total_entradas': total_entradas,
                'total_salidas': total_salidas,
                'total_ajustes': total_ajustes,
                'lotes_unicos': len(lotes_unicos),
                'productos_unicos': len(productos_unicos),
            },
            'filtros_aplicados': {
                'fecha_inicio': fecha_inicio,
                'fecha_fin': fecha_fin,
                'centro': centro_param,
                'tipo': tipo_movimiento,
                'producto': producto_param,
            }
        }
        
        # Exportar según formato
        if formato == 'excel':
            return _exportar_trazabilidad_global_excel(movimientos_data, response_data['filtros_aplicados'])
        elif formato == 'pdf':
            return _exportar_trazabilidad_global_pdf(movimientos_data, response_data)
        
        return Response(response_data)
        
    except Exception as exc:
        logger.exception('Error en trazabilidad_global')
        return Response({
            'error': 'Error al obtener trazabilidad global',
            'mensaje': str(exc)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


def _exportar_trazabilidad_global_excel(movimientos, filtros):
    """Genera Excel de trazabilidad global."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trazabilidad"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="9F2241", end_color="9F2241", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Determinar título según el filtro de centro
    centro_param = filtros.get('centro', '')
    if centro_param == 'todos':
        titulo_centro = "TODOS LOS CENTROS (CONSOLIDADO)"
    elif centro_param and centro_param != 'central':
        try:
            centro_obj = Centro.objects.get(pk=int(centro_param))
            titulo_centro = centro_obj.nombre.upper()
        except (Centro.DoesNotExist, ValueError):
            titulo_centro = f"CENTRO {centro_param}"
    else:
        titulo_centro = "FARMACIA CENTRAL"
    
    # Título dinámico según el centro
    ws.merge_cells('A1:K1')
    ws['A1'] = f"REPORTE DE TRAZABILIDAD - {titulo_centro}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Filtros aplicados
    row = 3
    ws[f'A{row}'] = "Filtros aplicados:"
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    
    # Siempre mostrar el centro para claridad
    ws[f'A{row}'] = f"Centro: {titulo_centro}"
    row += 1
    
    if filtros.get('fecha_inicio'):
        ws[f'A{row}'] = f"Fecha inicio: {filtros['fecha_inicio']}"
        row += 1
    if filtros.get('fecha_fin'):
        ws[f'A{row}'] = f"Fecha fin: {filtros['fecha_fin']}"
        row += 1
    if filtros.get('tipo'):
        ws[f'A{row}'] = f"Tipo de movimiento: {filtros['tipo'].upper()}"
        row += 1
    
    ws[f'A{row}'] = f"Total movimientos: {len(movimientos)}"
    row += 2
    
    # Encabezados - ISS-FIX: Agregado Subtipo Salida y No. Expediente para trazabilidad completa
    headers = ['Fecha', 'Tipo', 'Subtipo', 'Producto', 'Nombre Producto', 'Lote', 'Cantidad', 'Centro', 'Usuario', 'No. Expediente', 'Observaciones']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Datos - ISS-FIX: Incluye subtipo_salida y numero_expediente
    # NO truncar texto - Excel maneja el ajuste automático
    for mov in movimientos:
        row += 1
        data = [
            mov['fecha_str'],
            mov['tipo'],
            mov.get('subtipo_salida', '') or '-',
            mov['producto_clave'],
            mov['producto_nombre'] or '',  # Sin truncar
            mov['lote'],
            mov['cantidad'],
            mov['centro'] or '',  # Sin truncar
            mov['usuario'],
            mov.get('numero_expediente', '') or '-',
            mov['observaciones'] or ''  # Sin truncar
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
            # Habilitar ajuste de texto en celdas con contenido largo
            if col in [5, 8, 11]:  # Nombre Producto, Centro, Observaciones
                cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # Ajustar anchos - Columnas más anchas para mejor legibilidad
    ws.column_dimensions['A'].width = 18  # Fecha
    ws.column_dimensions['B'].width = 10  # Tipo
    ws.column_dimensions['C'].width = 14  # Subtipo
    ws.column_dimensions['D'].width = 12  # Producto (clave)
    ws.column_dimensions['E'].width = 45  # Nombre Producto - MÁS ANCHO
    ws.column_dimensions['F'].width = 14  # Lote
    ws.column_dimensions['G'].width = 10  # Cantidad
    ws.column_dimensions['H'].width = 35  # Centro - MÁS ANCHO
    ws.column_dimensions['I'].width = 18  # Usuario
    ws.column_dimensions['J'].width = 14  # No. Expediente
    ws.column_dimensions['K'].width = 50  # Observaciones - MÁS ANCHO
    
    # Guardar
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    from django.utils import timezone
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Trazabilidad_Global_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
    return response


def _exportar_trazabilidad_global_pdf(movimientos, data):
    """Genera PDF de trazabilidad global."""
    from core.utils.pdf_reports import generar_reporte_trazabilidad
    from io import BytesIO
    from django.utils import timezone
    
    # Determinar el nombre del centro para el título del PDF
    centro_param = data.get('filtros_aplicados', {}).get('centro', '')
    if centro_param == 'todos':
        titulo_centro = "TODOS LOS CENTROS (CONSOLIDADO)"
    elif centro_param and centro_param != 'central':
        try:
            centro_obj = Centro.objects.get(pk=int(centro_param))
            titulo_centro = centro_obj.nombre.upper()
        except (Centro.DoesNotExist, ValueError):
            titulo_centro = f"CENTRO {centro_param}"
    else:
        titulo_centro = "FARMACIA CENTRAL"
    
    # Adaptar datos al formato esperado por generar_reporte_trazabilidad
    # ISS-FIX: Incluir subtipo_salida y numero_expediente para trazabilidad completa
    trazabilidad_data = []
    for mov in movimientos[:200]:  # Limitar para PDF
        trazabilidad_data.append({
            'fecha': mov['fecha_str'],
            'tipo': mov['tipo'],
            'subtipo_salida': mov.get('subtipo_salida', '') or '',
            'lote': mov['lote'],
            'cantidad': mov['cantidad'],
            'centro': mov['centro'],
            'usuario': mov['usuario'],
            'numero_expediente': mov.get('numero_expediente', '') or '',
            'observaciones': mov['observaciones'] or '',
            'producto': f"{mov['producto_clave']} - {mov['producto_nombre'][:30]}"
        })
    
    # Info del reporte con título dinámico
    producto_info = {
        'clave': 'GLOBAL',
        'descripcion': f'Reporte de Trazabilidad - {titulo_centro}',
        'unidad_medida': '-',
        # ISS-FIX: Nunca mostrar stock negativo - usar max(0, ...)
        'stock_actual': max(0, data['estadisticas']['total_entradas'] - data['estadisticas']['total_salidas']),
        'stock_minimo': 0,
        'es_global': True,
        'filtros': data['filtros_aplicados'],
        'estadisticas': data['estadisticas'],
        'titulo_centro': titulo_centro,  # Para el PDF
    }
    
    pdf_buffer = generar_reporte_trazabilidad(trazabilidad_data, producto_info=producto_info)
    
    # Nombre de archivo dinámico según el centro
    nombre_centro_archivo = titulo_centro.replace(' ', '_').replace('(', '').replace(')', '')
    response = HttpResponse(
        pdf_buffer.getvalue(),
        content_type='application/pdf'
    )
    response['Content-Disposition'] = f'attachment; filename="Trazabilidad_{nombre_centro_archivo}_{timezone.now().strftime("%Y%m%d_%H%M")}.pdf"'
    return response


@api_view(['GET'])
def trazabilidad_producto_exportar(request, clave):
    """
    Exportar trazabilidad de un producto con filtros de fecha.
    
    Parámetros:
    - fecha_inicio: Fecha inicio (YYYY-MM-DD)
    - fecha_fin: Fecha fin (YYYY-MM-DD)
    - formato: excel o pdf
    """
    from datetime import datetime
    from django.utils import timezone
    
    try:
        user = request.user
        if not user or not user.is_authenticated:
            return Response({'error': 'Autenticación requerida'}, status=status.HTTP_403_FORBIDDEN)
        
        rol_usuario = (getattr(user, 'rol', '') or '').lower()
        if rol_usuario == 'vista':
            return Response({'error': 'No tienes permiso para exportar'}, status=status.HTTP_403_FORBIDDEN)
        
        # Obtener producto
        producto = Producto.objects.filter(
            Q(clave__iexact=clave) | Q(descripcion__iexact=clave)
        ).first()
        if not producto:
            return Response({'error': 'Producto no encontrado'}, status=status.HTTP_404_NOT_FOUND)
        
        # Filtros
        fecha_inicio = request.query_params.get('fecha_inicio')
        fecha_fin = request.query_params.get('fecha_fin')
        formato = request.query_params.get('formato', 'pdf')
        centro_param = request.query_params.get('centro')
        
        # Determinar filtro de centro
        filtrar_por_centro = not is_farmacia_or_admin(user)
        user_centro = get_user_centro(user) if filtrar_por_centro else None
        
        if centro_param and is_farmacia_or_admin(user):
            if centro_param.lower() in ('central', 'todos'):
                if centro_param.lower() == 'central':
                    user_centro = None
                else:
                    filtrar_por_centro = False
            else:
                try:
                    if centro_param.isdigit():
                        user_centro = Centro.objects.get(pk=centro_param)
                    else:
                        user_centro = Centro.objects.get(nombre__iexact=centro_param)
                    filtrar_por_centro = True
                except Centro.DoesNotExist:
                    pass
        
        # Obtener movimientos
        movimientos = Movimiento.objects.filter(
            lote__producto=producto
        ).select_related('lote', 'centro_origen', 'centro_destino', 'usuario')
        
        # ISS-FIX: Aplicar filtro de centro de forma ESTRICTA
        # Para trazabilidad: solo donde el centro es origen O destino (no por lote__centro)
        if filtrar_por_centro and user_centro:
            movimientos = movimientos.filter(
                Q(centro_origen=user_centro) | Q(centro_destino=user_centro)
            )
        
        # Aplicar filtros de fecha
        if fecha_inicio:
            try:
                fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d')
                movimientos = movimientos.filter(fecha__gte=fecha_inicio_dt)
            except ValueError:
                pass
        
        if fecha_fin:
            try:
                fecha_fin_dt = datetime.strptime(fecha_fin, '%Y-%m-%d')
                fecha_fin_dt = fecha_fin_dt.replace(hour=23, minute=59, second=59)
                movimientos = movimientos.filter(fecha__lte=fecha_fin_dt)
            except ValueError:
                pass
        
        movimientos = movimientos.order_by('-fecha')[:500]
        
        # Preparar datos de movimientos
        trazabilidad_data = []
        for mov in movimientos:
            # ISS-FIX: Lógica clara para centro según tipo de movimiento
            tipo_upper = mov.tipo.upper()
            if tipo_upper == 'SALIDA':
                centro_display = mov.centro_destino.nombre if mov.centro_destino else 'Farmacia Central'
            elif tipo_upper == 'ENTRADA':
                centro_display = mov.centro_origen.nombre if mov.centro_origen else 'Farmacia Central'
            else:
                centro_display = mov.centro_destino.nombre if mov.centro_destino else (mov.centro_origen.nombre if mov.centro_origen else 'Farmacia Central')
            
            # ISS-FIX: Mostrar username si no hay nombre completo
            if mov.usuario:
                usuario_display = mov.usuario.get_full_name()
                if not usuario_display or usuario_display.strip() == '':
                    usuario_display = mov.usuario.username
            else:
                usuario_display = 'Sistema'
            
            trazabilidad_data.append({
                'fecha': mov.fecha.strftime('%d/%m/%Y %H:%M') if mov.fecha else 'N/A',
                'tipo': tipo_upper,
                'lote': mov.lote.numero_lote if mov.lote else 'N/A',
                'cantidad': mov.cantidad,
                'centro': centro_display,
                'usuario': usuario_display,
                'observaciones': mov.motivo or ''
            })
        
        # Obtener TODOS los lotes del producto para incluir en el reporte
        lotes_query = Lote.objects.filter(producto=producto, activo=True)
        if filtrar_por_centro and user_centro:
            lotes_query = lotes_query.filter(centro=user_centro)
        lotes_query = lotes_query.select_related('centro').order_by('-fecha_caducidad')
        
        lotes_data = []
        for lote in lotes_query:
            lotes_data.append({
                'numero_lote': lote.numero_lote,
                'numero_contrato': lote.numero_contrato or 'N/A',
                'fecha_caducidad': lote.fecha_caducidad.strftime('%d/%m/%Y') if lote.fecha_caducidad else 'N/A',
                'cantidad_actual': lote.cantidad_actual,
                'cantidad_inicial': lote.cantidad_inicial,
                'marca': lote.marca or 'N/A',
                'centro': lote.centro.nombre if lote.centro else 'Farmacia Central',
                'precio_unitario': float(lote.precio_unitario) if lote.precio_unitario else 0,
            })
        
        # Calcular stock total
        stock_total = sum(l['cantidad_actual'] for l in lotes_data)
        
        producto_info = {
            'clave': producto.clave,
            'descripcion': producto.nombre,
            'unidad_medida': producto.unidad_medida,
            'stock_actual': stock_total,
            'stock_minimo': producto.stock_minimo,
            'lotes': lotes_data,  # Incluir lotes en el reporte
            'total_lotes': len(lotes_data),
            'filtros': {
                'fecha_inicio': fecha_inicio,
                'fecha_fin': fecha_fin,
                'centro': centro_param
            }
        }
        
        if formato == 'excel':
            return _exportar_producto_excel(trazabilidad_data, producto_info)
        else:
            from core.utils.pdf_reports import generar_reporte_trazabilidad
            pdf_buffer = generar_reporte_trazabilidad(trazabilidad_data, producto_info=producto_info)
            
            response = HttpResponse(
                pdf_buffer.getvalue(),
                content_type='application/pdf'
            )
            response['Content-Disposition'] = f'attachment; filename="Trazabilidad_{clave}_{timezone.now().strftime("%Y%m%d")}.pdf"'
            return response
            
    except Exception as exc:
        logger.exception('Error en trazabilidad_producto_exportar')
        return Response({
            'error': 'Error al exportar trazabilidad',
            'mensaje': str(exc)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


def _exportar_producto_excel(movimientos, producto_info):
    """Genera Excel de trazabilidad de producto."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO
    from django.utils import timezone
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trazabilidad"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="9F2241", end_color="9F2241", fill_type="solid")
    subheader_fill = PatternFill(start_color="BC955C", end_color="BC955C", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    bold_font = Font(bold=True)
    
    # Título
    ws.merge_cells('A1:G1')
    ws['A1'] = f"TRAZABILIDAD DE PRODUCTO: {producto_info['clave']}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Info del producto
    ws['A3'] = "Producto:"
    ws['A3'].font = bold_font
    ws['B3'] = producto_info['descripcion']
    ws['A4'] = "Unidad:"
    ws['A4'].font = bold_font
    ws['B4'] = producto_info['unidad_medida']
    ws['A5'] = "Stock Total:"
    ws['A5'].font = bold_font
    ws['B5'] = producto_info.get('stock_actual', 0)
    ws['C5'] = "Stock Mínimo:"
    ws['C5'].font = bold_font
    ws['D5'] = producto_info.get('stock_minimo', 0)
    
    row = 7
    
    # ========== SECCIÓN DE LOTES ==========
    lotes = producto_info.get('lotes', [])
    if lotes:
        ws.merge_cells(f'A{row}:G{row}')
        ws[f'A{row}'] = "LOTES DEL PRODUCTO"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].fill = subheader_fill
        row += 1
        
        # Encabezados de lotes
        lotes_headers = ['No. Lote', 'No. Contrato', 'Caducidad', 'Stock', 'Marca', 'Centro', 'Precio Unit.']
        for col, header in enumerate(lotes_headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        # Datos de lotes
        for lote in lotes:
            row += 1
            lote_data = [
                lote.get('numero_lote', 'N/A'),
                lote.get('numero_contrato', 'N/A'),
                lote.get('fecha_caducidad', 'N/A'),
                lote.get('cantidad_actual', 0),
                lote.get('marca', 'N/A'),
                lote.get('centro', 'N/A'),
                f"${lote.get('precio_unitario', 0):.2f}" if lote.get('precio_unitario') else 'N/A',
            ]
            for col, value in enumerate(lote_data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = thin_border
        
        row += 2  # Espacio entre secciones
    
    # Filtros
    filtros = producto_info.get('filtros', {})
    if any(filtros.values()):
        if filtros.get('fecha_inicio'):
            ws[f'A{row}'] = f"Desde: {filtros['fecha_inicio']}"
            row += 1
        if filtros.get('fecha_fin'):
            ws[f'A{row}'] = f"Hasta: {filtros['fecha_fin']}"
            row += 1
        row += 1
    
    # ========== SECCIÓN DE MOVIMIENTOS ==========
    ws.merge_cells(f'A{row}:G{row}')
    ws[f'A{row}'] = "HISTORIAL DE MOVIMIENTOS"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'A{row}'].fill = subheader_fill
    row += 1
    
    # Encabezados de movimientos
    headers = ['Fecha', 'Tipo', 'Lote', 'Cantidad', 'Centro', 'Usuario', 'Observaciones']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    # Datos de movimientos
    if movimientos:
        for mov in movimientos:
            row += 1
            data = [
                mov['fecha'],
                mov['tipo'],
                mov['lote'],
                mov['cantidad'],
                mov['centro'],
                mov['usuario'],
                mov['observaciones'] if mov['observaciones'] else ''
            ]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = thin_border
    else:
        row += 1
        ws.merge_cells(f'A{row}:G{row}')
        ws[f'A{row}'] = "No hay movimientos registrados"
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
    
    # Ajustar anchos - Columnas más anchas para textos completos
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 40  # Centro más ancho para nombres largos
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 60  # Observaciones mucho más ancho para texto completo
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Trazabilidad_{producto_info["clave"]}_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    return response


@api_view(['GET'])
def trazabilidad_lote_exportar(request, codigo):
    """
    Exportar trazabilidad COMPLETA de un lote con filtros de fecha.
    
    IMPORTANTE: Busca movimientos de TODOS los lotes con el mismo numero_lote,
    incluyendo el lote original en Farmacia Central y los lotes espejo en centros.
    Esto permite ver la trazabilidad completa del lote físico.
    
    Parámetros:
    - fecha_inicio: Fecha inicio (YYYY-MM-DD)
    - fecha_fin: Fecha fin (YYYY-MM-DD)
    - formato: excel o pdf
    """
    from datetime import datetime
    from django.utils import timezone
    
    try:
        user = request.user
        if not user or not user.is_authenticated:
            return Response({'error': 'Autenticación requerida'}, status=status.HTTP_403_FORBIDDEN)
        
        if not is_farmacia_or_admin(user):
            return Response({'error': 'Solo administradores y farmacia pueden exportar trazabilidad de lotes'}, status=status.HTTP_403_FORBIDDEN)
        
        # TRAZABILIDAD COMPLETA: Obtener TODOS los lotes con el mismo numero_lote
        lotes_con_mismo_numero = Lote.objects.select_related('producto', 'centro').filter(numero_lote__iexact=codigo)
        
        if not lotes_con_mismo_numero.exists():
            return Response({'error': 'Lote no encontrado'}, status=status.HTTP_404_NOT_FOUND)
        
        # Usar el lote de Farmacia Central como principal (centro=NULL) o el primero disponible
        lote_principal = lotes_con_mismo_numero.filter(centro__isnull=True).first()
        if not lote_principal:
            lote_principal = lotes_con_mismo_numero.first()
        lote = lote_principal
        
        # Obtener todos los IDs de lotes con el mismo numero_lote
        lotes_ids = list(lotes_con_mismo_numero.values_list('id', flat=True))
        
        # Calcular cantidad total consolidada
        cantidad_total_consolidada = lotes_con_mismo_numero.aggregate(total=Sum('cantidad_actual'))['total'] or 0
        
        # Filtros
        fecha_inicio = request.query_params.get('fecha_inicio')
        fecha_fin = request.query_params.get('fecha_fin')
        formato = request.query_params.get('formato', 'pdf')
        
        # TRAZABILIDAD COMPLETA: Obtener movimientos de TODOS los lotes con el mismo numero_lote
        movimientos = Movimiento.objects.filter(
            lote_id__in=lotes_ids
        ).select_related('centro_origen', 'centro_destino', 'usuario', 'lote', 'lote__centro')
        
        # Aplicar filtros de fecha
        if fecha_inicio:
            try:
                fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d')
                movimientos = movimientos.filter(fecha__gte=fecha_inicio_dt)
            except ValueError:
                pass
        
        if fecha_fin:
            try:
                fecha_fin_dt = datetime.strptime(fecha_fin, '%Y-%m-%d')
                fecha_fin_dt = fecha_fin_dt.replace(hour=23, minute=59, second=59)
                movimientos = movimientos.filter(fecha__lte=fecha_fin_dt)
            except ValueError:
                pass
        
        movimientos = movimientos.order_by('fecha')
        
        # Preparar datos con saldo
        trazabilidad_data = []
        saldo = 0
        for mov in movimientos:
            saldo += mov.cantidad
            # Determinar el centro del movimiento de forma más precisa
            if mov.centro_destino:
                centro_mov = mov.centro_destino.nombre
            elif mov.centro_origen:
                centro_mov = mov.centro_origen.nombre
            elif mov.lote and mov.lote.centro:
                centro_mov = mov.lote.centro.nombre
            else:
                centro_mov = 'Farmacia Central'
            
            trazabilidad_data.append({
                'fecha': mov.fecha.strftime('%d/%m/%Y %H:%M') if mov.fecha else 'N/A',
                'tipo': mov.tipo.upper(),
                'lote': lote.numero_lote,
                'cantidad': mov.cantidad,
                'saldo': saldo,
                'centro': centro_mov,
                'usuario': mov.usuario.get_full_name() if mov.usuario else 'Sistema',
                'observaciones': mov.motivo or ''
            })
        
        producto_info = {
            'clave': lote.producto.clave if lote.producto else 'N/A',
            'descripcion': lote.producto.nombre if lote.producto else 'N/A',
            'unidad_medida': lote.producto.unidad_medida if lote.producto else 'N/A',
            'stock_actual': cantidad_total_consolidada,  # Cantidad consolidada de todos los centros
            'stock_minimo': lote.producto.stock_minimo if lote.producto else 0,
            'numero_lote': lote.numero_lote,
            'fecha_caducidad': lote.fecha_caducidad.strftime('%d/%m/%Y') if lote.fecha_caducidad else 'N/A',
            'proveedor': lote.marca or 'No especificado',
            'numero_contrato': lote.numero_contrato or 'N/A',
            'precio_unitario': float(lote.precio_unitario) if lote.precio_unitario else 0,
            'filtros': {
                'fecha_inicio': fecha_inicio,
                'fecha_fin': fecha_fin
            }
        }
        
        if formato == 'excel':
            return _exportar_lote_excel(trazabilidad_data, producto_info)
        else:
            from core.utils.pdf_reports import generar_reporte_trazabilidad
            pdf_buffer = generar_reporte_trazabilidad(trazabilidad_data, producto_info=producto_info)
            
            response = HttpResponse(
                pdf_buffer.getvalue(),
                content_type='application/pdf'
            )
            response['Content-Disposition'] = f'attachment; filename="Trazabilidad_Lote_{codigo}_{timezone.now().strftime("%Y%m%d")}.pdf"'
            return response
            
    except Exception as exc:
        logger.exception('Error en trazabilidad_lote_exportar')
        return Response({
            'error': 'Error al exportar trazabilidad del lote',
            'mensaje': str(exc)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def exportar_control_inventarios(request):
    """
    Exporta el inventario en formato "Control de Inventarios del Almacén Central de Medicamentos"
    IDÉNTICO al archivo de referencia de licitación.
    
    Características:
    - Fila separadora vacía entre cada producto diferente
    - Columna A con borde medium (grueso)
    - Columnas B-M con borde thin
    - Fondo amarillo SOLO en columna M (evidencia)
    - IconSet (semáforo con circulitos) en columnas H e I
    - Sin colores de fondo excepto evidencia y headers
    """
    from datetime import datetime
    from dateutil.relativedelta import relativedelta
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.formatting.rule import IconSetRule
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    
    user = request.user
    if not user or not user.is_authenticated:
        return Response({'error': 'Autenticación requerida'}, status=status.HTTP_403_FORBIDDEN)
    
    if not is_farmacia_or_admin(user):
        return Response({'error': 'Solo administradores y farmacia pueden exportar este formato'}, status=status.HTTP_403_FORBIDDEN)
    
    try:
        # Obtener todos los lotes activos con stock, agrupados por producto
        lotes = Lote.objects.select_related('producto').filter(
            activo=True,
            cantidad_actual__gt=0,
            centro__isnull=True  # Solo farmacia central (Almacén Central)
        ).order_by('producto__clave', 'numero_lote')
        
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Hoja1"
        
        # Estilos EXACTOS del formato de referencia
        # ISS-FIX: Cambio de verde (#C4D79B) a gris (#D9D9D9) según especificación
        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        
        # Bordes - Columna A tiene medium, las demás thin
        medium_border_left = Border(
            left=Side(style='medium'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Folio en esquina superior derecha (I2:L2)
        ws.merge_cells('I2:L2')
        ws['I2'] = "Folio:_____________________"
        ws['I2'].alignment = Alignment(horizontal='right')
        
        # Título principal (B4:L4)
        ws.merge_cells('B4:L4')
        ws['B4'] = "CONTROL DE INVENTARIOS DEL ALMACÉN CENTRAL DE MEDICAMENTOS"
        ws['B4'].font = Font(bold=True, size=12)
        ws['B4'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[4].height = 42.75
        
        # Encabezados en fila 6 (12 columnas, sin evidencia por ahora)
        headers = [
            'NO. PARTIDA',
            'CLAVE',
            'ARTÍCULO',
            'LOTE',
            'NOMBRE COMERCIAL O GENÉRICO',
            'CONCENTRACIÓN',
            'PRESENTACIÓN',
            'MESES',
            'VENCIMIENTO (SEMAFORIZACIÓN) / FECHA DE CADUCIDAD',
            'CANTIDAD',
            'FECHA DE INGRESO',
            'FECHA DE SALIDA (ULTIMA)'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col, value=header)
            cell.font = Font(bold=True, size=8)
            cell.fill = header_fill
            cell.border = medium_border_left if col == 1 else thin_border
            cell.alignment = header_align
        
        ws.row_dimensions[6].height = 36
        
        # Datos con filas separadoras
        row = 7
        partida_actual = 0
        producto_anterior_id = None
        filas_con_iconset = []  # Para aplicar IconSet después
        
        for lote in lotes:
            producto = lote.producto
            
            # Si cambia el producto, agregar fila vacía separadora (excepto el primero)
            if producto_anterior_id is not None and producto.id != producto_anterior_id:
                row += 1  # Fila vacía sin bordes ni nada
            
            # Nueva partida si cambia el producto
            if producto.id != producto_anterior_id:
                partida_actual += 1
                producto_anterior_id = producto.id
            
            # Obtener fecha de ingreso
            fecha_ingreso = Movimiento.objects.filter(
                lote=lote,
                tipo='entrada'
            ).order_by('fecha').values_list('fecha', flat=True).first()
            
            if not fecha_ingreso and hasattr(lote, 'created_at') and lote.created_at:
                fecha_ingreso = lote.created_at
            
            # Obtener última fecha de salida
            ultima_salida = Movimiento.objects.filter(
                lote=lote,
                tipo='salida'
            ).order_by('-fecha').values_list('fecha', flat=True).first()
            
            # Preparar fechas como objetos date
            fecha_ingreso_date = None
            if fecha_ingreso:
                fecha_ingreso_date = fecha_ingreso.date() if hasattr(fecha_ingreso, 'date') else fecha_ingreso
            
            fecha_salida_date = None
            if ultima_salida:
                fecha_salida_date = ultima_salida.date() if hasattr(ultima_salida, 'date') else ultima_salida
            
            # Datos de la fila (12 columnas, sin evidencia por ahora)
            data = [
                partida_actual,  # A
                producto.clave,  # B
                producto.nombre or producto.descripcion or '',  # C
                lote.numero_lote,  # D
                producto.nombre_comercial or '',  # E - Nombre comercial del producto (vacío si no tiene)
                producto.concentracion or '',  # F
                producto.presentacion or '',  # G
                f'=ROUND((I{row}-K{row})/30,0)',  # H - Fórmula MESES (redondeado a entero)
                lote.fecha_caducidad,  # I - Fecha caducidad
                max(0, lote.cantidad_actual),  # J - ISS-FIX: Nunca mostrar stock negativo
                fecha_ingreso_date,  # K
                fecha_salida_date,  # L
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = Font(size=9)
                
                # Borde: columna A con medium, resto thin
                cell.border = medium_border_left if col == 1 else thin_border
                
                # Alineación
                if col in [1, 8, 10]:  # Partida, Meses, Cantidad
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align
                
                # Formato fecha para columnas I, K, L
                if col in [9, 11, 12] and value:
                    cell.number_format = 'DD/MM/YYYY'
            
            # Guardar fila para IconSet
            filas_con_iconset.append(row)
            
            row += 1
        
        # Aplicar IconSet (semáforo con circulitos) SOLO a columna H (MESES)
        # Columna I solo muestra la fecha de caducidad, SIN semáforo
        for fila in filas_con_iconset:
            # Columna H - semáforo basado en valor numérico (meses)
            rule_h = IconSetRule(
                '3TrafficLights1',
                'num',
                [0, 6, 12],
                showValue=True,
                reverse=False
            )
            ws.conditional_formatting.add(f'H{fila}', rule_h)
        
        # Anchos de columna EXACTOS (12 columnas, A-L)
        column_widths = {
            'A': 3.71,
            'B': 6.71,
            'C': 26.71,
            'D': 9.43,
            'E': 10.14,
            'F': 15.14,
            'G': 19.0,
            'H': 5.43,
            'I': 10.0,
            'J': 6.14,
            'K': 9.71,
            'L': 8.29,
        }
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
        
        # Guardar
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"Control_Inventarios_Almacen_Central_{timezone.now().strftime('%Y%m%d')}.xlsx"
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
        
    except Exception as exc:
        logger.exception('Error al exportar control de inventarios')
        return Response({
            'error': 'Error al exportar control de inventarios',
            'mensaje': str(exc)
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


def _exportar_lote_excel(movimientos, producto_info):
    """Genera Excel de trazabilidad de lote."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO
    from django.utils import timezone
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trazabilidad Lote"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="9F2241", end_color="9F2241", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws.merge_cells('A1:H1')
    ws['A1'] = f"TRAZABILIDAD DE LOTE: {producto_info['numero_lote']}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Info del lote
    ws['A3'] = "Producto:"
    ws['B3'] = f"{producto_info['clave']} - {producto_info['descripcion']}"
    ws['A4'] = "Caducidad:"
    ws['B4'] = producto_info['fecha_caducidad']
    ws['A5'] = "Contrato:"
    ws['B5'] = producto_info['numero_contrato']
    
    # Filtros
    filtros = producto_info.get('filtros', {})
    row = 7
    if any(filtros.values()):
        if filtros.get('fecha_inicio'):
            ws[f'A{row}'] = f"Desde: {filtros['fecha_inicio']}"
            row += 1
        if filtros.get('fecha_fin'):
            ws[f'A{row}'] = f"Hasta: {filtros['fecha_fin']}"
            row += 1
        row += 1
    
    # Encabezados
    headers = ['Fecha', 'Tipo', 'Cantidad', 'Saldo', 'Centro', 'Usuario', 'Observaciones']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    # Datos
    for mov in movimientos:
        row += 1
        data = [
            mov['fecha'],
            mov['tipo'],
            mov['cantidad'],
            mov.get('saldo', ''),
            mov['centro'],
            mov['usuario'],
            mov['observaciones'] if mov['observaciones'] else ''
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
    
    # Ajustar anchos - Columnas más anchas para textos completos
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 40  # Centro más ancho para nombres largos
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 60  # Observaciones mucho más ancho para texto completo
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Trazabilidad_Lote_{producto_info["numero_lote"]}_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    return response





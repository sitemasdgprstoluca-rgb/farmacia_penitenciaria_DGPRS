# -*- coding: utf-8 -*-
"""
Módulo de Reportes para inventario.

Funciones de reportes extraídas de views_legacy.py:
- reporte_inventario: Reporte de inventario actual por centro/producto
- reporte_movimientos: Reporte de movimientos con filtros de fecha/tipo
- reporte_caducidades: Reporte de lotes próximos a caducar
- reporte_requisiciones: Reporte de requisiciones por estado/centro
- reporte_medicamentos_por_caducar: Medicamentos próximos a vencer
- reporte_bajo_stock: Productos con stock bajo mínimo
- reporte_consumo: Análisis de consumo por periodo
- reporte_contratos: Reporte de lotes por contrato con consumo
- reportes_precarga: Datos de precarga para filtros de reportes
- reporte_medicamentos_controlados: Análisis de medicamentos controlados vs no controlados
- reporte_auditoria_productos: Auditoría de cambios en campo es_controlado

Nota: Por ahora se re-exporta desde views_legacy.py para mantener compatibilidad.
La migración completa del código se hará de forma incremental.
"""

from io import BytesIO

from django.db.models import Sum, Q, Avg
from django.http import HttpResponse
from django.utils import timezone
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from core.models import Producto, AuditoriaLogs
from inventario.views.base import is_farmacia_or_admin, has_global_read_access, get_user_centro

from inventario.views_legacy import (
    reporte_inventario,
    reporte_movimientos,
    reporte_caducidades,
    reporte_requisiciones,
    reporte_medicamentos_por_caducar,
    reporte_bajo_stock,
    reporte_consumo,
    reporte_contratos,
    reporte_parcialidades,
    reportes_precarga,
)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def reporte_medicamentos_controlados(request):
    """
    Análisis de Medicamentos Controlados.
    Retorna resumen y listado de productos controlados vs no controlados.
    Filtros opcionales: ?centro=ID&formato=json|excel|pdf
    """
    try:
        user = request.user
        if not has_global_read_access(user):
            return Response(
                {'error': 'No tiene permisos para ver este reporte.'},
                status=status.HTTP_403_FORBIDDEN,
            )

        formato = request.query_params.get('formato', 'json').lower()

        filtrar_por_centro = not is_farmacia_or_admin(user)
        user_centro = get_user_centro(user) if filtrar_por_centro else None

        centro_param = request.query_params.get('centro')
        centro_filtro = None
        if centro_param and is_farmacia_or_admin(user):
            if centro_param not in ('central', 'todos'):
                try:
                    centro_filtro = int(centro_param)
                except (ValueError, TypeError):
                    pass

        productos = Producto.objects.filter(activo=True)
        controlados = productos.filter(es_controlado=True)
        no_controlados = productos.filter(es_controlado=False)

        def _build_producto_row(prod):
            
            lotes_q = prod.lotes.filter(activo=True, cantidad_actual__gt=0)
            if filtrar_por_centro and user_centro:
                lotes_q = lotes_q.filter(centro=user_centro)
            elif centro_filtro:
                lotes_q = lotes_q.filter(centro_id=centro_filtro)
            
            stock = lotes_q.aggregate(total=Sum('cantidad_actual'))['total'] or 0
            precio_promedio = lotes_q.aggregate(promedio=Avg('precio_unitario'))['promedio'] or 0
            valor_total = float(stock) * float(precio_promedio)
            
            # Obtener lote más próximo a caducar
            lote_proximo = lotes_q.order_by('fecha_caducidad').first()
            lote_info = f"{lote_proximo.numero_lote[:15]} - {lote_proximo.fecha_caducidad.strftime('%d/%m/%Y')}" if lote_proximo else '—'
            
            return {
                'id': prod.id,
                'clave': prod.clave,
                'nombre': prod.nombre,
                'presentacion': prod.presentacion or '',
                'precio_promedio': round(precio_promedio, 2),
                'es_controlado': prod.es_controlado,
                'stock_actual': stock,
                'lote_caducidad': lote_info,
                'valor_total': round(valor_total, 2),
                'requiere_receta': prod.requiere_receta,
            }

        lista_controlados = [_build_producto_row(p) for p in controlados.order_by('clave')]
        lista_no_controlados = [_build_producto_row(p) for p in no_controlados.order_by('clave')]

        total = productos.count()
        total_ctrl = controlados.count()
        total_no_ctrl = no_controlados.count()
        
        # Calcular valores económicos globales
        valor_total_controlados = sum(item['valor_total'] for item in lista_controlados)
        valor_total_no_controlados = sum(item['valor_total'] for item in lista_no_controlados)
        valor_total_global = valor_total_controlados + valor_total_no_controlados
        
        # Análisis económico por centro (solo controlados)
        from core.models import Centro
        analisis_por_centro = []
        
        if not centro_filtro and not filtrar_por_centro:  # Solo si vemos todos los centros
            centros = Centro.objects.filter(activo=True).order_by('nombre')
            
            for centro in centros:
                centro_data = {
                    'centro_id': centro.id,
                    'centro_nombre': centro.nombre,
                    'total_productos_controlados': 0,
                    'total_stock_controlados': 0,
                    'valor_total_controlados': 0,
                }
                
                for prod in controlados:
                    lotes_centro = prod.lotes.filter(activo=True, cantidad_actual__gt=0, centro=centro)
                    if lotes_centro.exists():
                        stock_centro = lotes_centro.aggregate(total=Sum('cantidad_actual'))['total'] or 0
                        precio_prom = lotes_centro.aggregate(promedio=Avg('precio_unitario'))['promedio'] or 0
                        valor = stock_centro * float(precio_prom)
                        
                        centro_data['total_productos_controlados'] += 1
                        centro_data['total_stock_controlados'] += stock_centro
                        centro_data['valor_total_controlados'] += valor
                
                # Solo agregar si tiene stock de controlados
                if centro_data['total_stock_controlados'] > 0:
                    centro_data['valor_total_controlados'] = round(centro_data['valor_total_controlados'], 2)
                    analisis_por_centro.append(centro_data)
            
            # Ordenar por valor descendente
            analisis_por_centro.sort(key=lambda x: x['valor_total_controlados'], reverse=True)

        resumen = {
            'total_productos': total,
            'total_controlados': total_ctrl,
            'total_no_controlados': total_no_ctrl,
            'porcentaje_controlados': round((total_ctrl / total * 100), 1) if total else 0,
            'porcentaje_no_controlados': round((total_no_ctrl / total * 100), 1) if total else 0,
            'valor_total_controlados': round(valor_total_controlados, 2),
            'valor_total_no_controlados': round(valor_total_no_controlados, 2),
            'valor_total_global': round(valor_total_global, 2),
        }

        # ── JSON ──
        if formato == 'json':
            return Response({
                'resumen': resumen,
                'controlados': lista_controlados,
                'no_controlados': lista_no_controlados,
                'analisis_por_centro': analisis_por_centro,
            })

        # ── PDF ──
        if formato == 'pdf':
            from core.utils.pdf_reports import generar_reporte_medicamentos_controlados
            todos = lista_controlados + lista_no_controlados
            filtros = {
                'fecha_generacion': timezone.now().strftime('%d/%m/%Y %H:%M'),
                'total_productos': total,
                'total_controlados': total_ctrl,
                'total_no_controlados': total_no_ctrl,
                'valor_total_controlados': valor_total_controlados,
                'valor_total_no_controlados': valor_total_no_controlados,
                'valor_total_global': valor_total_global,
            }
            pdf_buffer = generar_reporte_medicamentos_controlados(todos, filtros=filtros)
            response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = (
                f"attachment; filename=Medicamentos_Controlados_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            )
            return response

        # ── Excel (2 hojas: Controlados + No Controlados) ──
        wb = openpyxl.Workbook()

        headers = ['#', 'Clave', 'Nombre', 'Presentación', 'Precio Prom.', 'Inventario', 'Lote/Caducidad', 'Valor Total']
        col_widths = [6, 10, 40, 30, 14, 12, 25, 14]

        # Hoja 1: Controlados
        ws1 = wb.active
        ws1.title = 'Controlados'

        ws1.merge_cells('A1:H1')
        c = ws1['A1']
        c.value = f'MEDICAMENTOS CONTROLADOS ({total_ctrl})'
        c.font = Font(bold=True, size=14, color='FFFFFF')
        c.fill = PatternFill(start_color='DC2626', end_color='DC2626', fill_type='solid')
        c.alignment = Alignment(horizontal='center', vertical='center')

        ws1.merge_cells('A2:H2')
        c2 = ws1['A2']
        c2.value = f'Generado el {timezone.now().strftime("%d/%m/%Y %H:%M")} — Análisis económico y de inventario'
        c2.font = Font(size=10, italic=True)
        c2.alignment = Alignment(horizontal='center')

        ws1.append([])
        ws1.append(headers)

        header_fill_red = PatternFill(start_color='991B1B', end_color='991B1B', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        for cell in ws1[4]:
            cell.fill = header_fill_red
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for idx, item in enumerate(lista_controlados, 1):
            ws1.append([
                idx, item['clave'], item['nombre'][:50],
                item['presentacion'][:40], 
                f"${item['precio_promedio']:,.2f}",
                item['stock_actual'], 
                item['lote_caducidad'],
                f"${item['valor_total']:,.2f}",
            ])
        for col_letter, width in zip('ABCDEFGH', col_widths):
            ws1.column_dimensions[col_letter].width = width
        
        # Agregar fila de totales para controlados
        ws1.append([])
        total_row = ws1.max_row + 1
        ws1.append(['', '', '', '', 'TOTAL CONTROLADOS:', '', '', f"${valor_total_controlados:,.2f}"])
        for cell in ws1[total_row]:
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')

        # Hoja 2: No Controlados
        ws2 = wb.create_sheet('No Controlados')

        ws2.merge_cells('A1:H1')
        c3 = ws2['A1']
        c3.value = f'MEDICAMENTOS NO CONTROLADOS ({total_no_ctrl})'
        c3.font = Font(bold=True, size=14, color='FFFFFF')
        c3.fill = PatternFill(start_color='16A34A', end_color='16A34A', fill_type='solid')
        c3.alignment = Alignment(horizontal='center', vertical='center')

        ws2.merge_cells('A2:H2')
        c4 = ws2['A2']
        c4.value = f'Generado el {timezone.now().strftime("%d/%m/%Y %H:%M")} — Medicamentos de uso general'
        c4.font = Font(size=10, italic=True)
        c4.alignment = Alignment(horizontal='center')

        ws2.append([])
        ws2.append(headers)

        header_fill_green = PatternFill(start_color='15803D', end_color='15803D', fill_type='solid')
        for cell in ws2[4]:
            cell.fill = header_fill_green
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for idx, item in enumerate(lista_no_controlados, 1):
            ws2.append([
                idx, item['clave'], item['nombre'][:50],
                item['presentacion'][:40], 
                f"${item['precio_promedio']:,.2f}",
                item['stock_actual'], 
                item['lote_caducidad'],
                f"${item['valor_total']:,.2f}",
            ])
        for col_letter, width in zip('ABCDEFGH', col_widths):
            ws2.column_dimensions[col_letter].width = width
        
        # Agregar fila de totales para no controlados
        ws2.append([])
        total_row2 = ws2.max_row + 1
        ws2.append(['', '', '', '', 'TOTAL NO CONTROLADOS:', '', '', f"${valor_total_no_controlados:,.2f}"])
        for cell in ws2[total_row2]:
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
        
        # Hoja 3: Análisis por Centro (solo si hay datos)
        if analisis_por_centro:
            ws3 = wb.create_sheet('Análisis por Centro')
            
            ws3.merge_cells('A1:E1')
            c5 = ws3['A1']
            c5.value = 'ANÁLISIS ECONÓMICO DE MEDICAMENTOS CONTROLADOS POR CENTRO'
            c5.font = Font(bold=True, size=14, color='FFFFFF')
            c5.fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
            c5.alignment = Alignment(horizontal='center', vertical='center')
            
            ws3.merge_cells('A2:E2')
            c6 = ws3['A2']
            c6.value = f'Distribución de inventario controlado — Generado el {timezone.now().strftime("%d/%m/%Y %H:%M")}'
            c6.font = Font(size=10, italic=True)
            c6.alignment = Alignment(horizontal='center')
            
            ws3.append([])
            headers_centro = ['#', 'Centro', 'Productos', 'Stock Total', 'Valor Total']
            ws3.append(headers_centro)
            
            header_fill_blue = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
            for cell in ws3[4]:
                cell.fill = header_fill_blue
                cell.font = Font(bold=True, color='FFFFFF', size=11)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            for idx, centro in enumerate(analisis_por_centro, 1):
                ws3.append([
                    idx,
                    centro['centro_nombre'][:40],
                    centro['total_productos_controlados'],
                    centro['total_stock_controlados'],
                    f"${centro['valor_total_controlados']:,.2f}",
                ])
            
            # Totales generales
            ws3.append([])
            ws3.append([
                '', 'TOTAL GENERAL:', 
                sum(c['total_productos_controlados'] for c in analisis_por_centro),
                sum(c['total_stock_controlados'] for c in analisis_por_centro),
                f"${valor_total_controlados:,.2f}",
            ])
            total_row3 = ws3.max_row
            for cell in ws3[total_row3]:
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color='BFDBFE', end_color='BFDBFE', fill_type='solid')
            
            # Anchos de columna
            col_widths_centro = [6, 40, 14, 14, 16]
            for col_letter, width in zip('ABCDE', col_widths_centro):
                ws3.column_dimensions[col_letter].width = width

        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        response = HttpResponse(
            excel_buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = (
            f'attachment; filename=Medicamentos_Controlados_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
        return response

    except Exception as exc:
        return Response(
            {'error': 'Error al generar reporte de medicamentos controlados.', 'mensaje': str(exc)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def reporte_auditoria_productos(request):
    """
    Auditoría de cambios en productos - enfocado en campo es_controlado.
    Filtros: ?fecha_inicio=YYYY-MM-DD&fecha_fin=YYYY-MM-DD&producto_id=N&campo=es_controlado&formato=json|excel|pdf
    """
    # Mapeo de nombres técnicos → legibles para personas no técnicas
    CAMPO_LABELS = {
        'nombre': 'Nombre del producto',
        'nombre_comercial': 'Nombre comercial',
        'clave': 'Clave del producto',
        'descripcion': 'Descripción',
        'unidad_medida': 'Unidad de medida',
        'categoria': 'Categoría',
        'presentacion': 'Presentación',
        'sustancia_activa': 'Sustancia activa',
        'concentracion': 'Concentración',
        'via_administracion': 'Vía de administración',
        'stock_minimo': 'Stock mínimo',
        'stock_actual': 'Stock actual',
        'precio_unitario': 'Precio unitario',
        'es_controlado': 'Medicamento controlado',
        'requiere_receta': 'Requiere receta',
        'activo': 'Estado del producto',
        'tiene_lotes': 'Tiene lotes registrados',
        'tiene_movimientos': 'Tiene movimientos',
        'marca': 'Marca',
        'laboratorio': 'Laboratorio',
        'created_at': 'Fecha de creación',
        'updated_at': 'Última actualización',
        'unidad_minima': 'Unidad mínima de dispensación',
        'factor_conversion': 'Factor de conversión',
        'imagen': 'Imagen del producto',
    }
    VALOR_LABELS = {
        True: 'Sí',
        False: 'No',
        'true': 'Sí',
        'false': 'No',
        'medicamento': 'Medicamento',
        'material_curacion': 'Material de curación',
        'insumo': 'Insumo',
        'equipo': 'Equipo',
        'otro': 'Otro',
    }
    ROL_LABELS = {
        'farmacia': 'Farmacia',
        'admin': 'Administrador',
        'admin_sistema': 'Administrador del Sistema',
        'medico': 'Médico',
        'director': 'Director',
        'centro': 'Centro',
        'superuser': 'Super Administrador',
    }

    # Campos internos que no aportan al usuario final
    CAMPOS_OCULTOS = {
        'id', 'created_by', 'updated_by', 'created_by_id', 'updated_by_id',
        'modified_by', 'modified_by_id',
    }

    def humanizar_valor(val):
        """Convierte un valor técnico a texto legible."""
        if val is None:
            return 'Sin dato'
        if isinstance(val, bool) or val in ('true', 'false', True, False):
            return VALOR_LABELS.get(val, str(val))
        if isinstance(val, str) and val in VALOR_LABELS:
            return VALOR_LABELS[val]
        if isinstance(val, str) and 'T' in val and len(val) > 19:
            # ISO datetime → formato legible
            try:
                from datetime import datetime as dt
                parsed = dt.fromisoformat(val.replace('+00:00', '+00:00').split('+')[0])
                return parsed.strftime('%d/%m/%Y %H:%M')
            except Exception:
                pass
        return str(val)

    def generar_narrativa(cambios_humanizados):
        """Genera una descripción narrativa de los cambios."""
        if not cambios_humanizados:
            return 'Sin cambios detectados'
        partes = []
        for ch in cambios_humanizados:
            campo = ch['campo_label']
            anterior = ch['valor_anterior_label']
            nuevo = ch['valor_nuevo_label']
            if anterior == 'Sin dato':
                partes.append(f'Se estableció {campo} como "{nuevo}"')
            else:
                partes.append(f'Se cambió {campo} de "{anterior}" a "{nuevo}"')
        return '. '.join(partes) + '.'
    try:
        user = request.user
        if not has_global_read_access(user):
            return Response(
                {'error': 'No tiene permisos para ver este reporte.'},
                status=status.HTTP_403_FORBIDDEN,
            )

        formato = request.query_params.get('formato', 'json').lower()

        logs = AuditoriaLogs.objects.filter(
            modelo__icontains='Producto',
            accion__in=[
                'ACTUALIZAR', 'MODIFICAR',           # middleware (uppercase)
                'actualizar', 'modificar',           # signals (lowercase)
                'update', 'update:INFO',             # audit_log service
                'PUT', 'PATCH',                      # fallback
            ],
        ).select_related('usuario').order_by('-timestamp')

        # Filtros opcionales
        fecha_inicio = request.query_params.get('fecha_inicio')
        fecha_fin = request.query_params.get('fecha_fin')
        producto_id = request.query_params.get('producto_id')
        campo = request.query_params.get('campo')

        if fecha_inicio:
            logs = logs.filter(timestamp__date__gte=fecha_inicio)
        if fecha_fin:
            logs = logs.filter(timestamp__date__lte=fecha_fin)
        if producto_id:
            logs = logs.filter(objeto_id=str(producto_id))

        # Pre-cargar nombres de productos para resolución rápida
        producto_ids = set()
        for log in logs[:500]:
            if log.objeto_id:
                try:
                    producto_ids.add(int(log.objeto_id))
                except (ValueError, TypeError):
                    pass
        producto_nombres = {}
        if producto_ids:
            from core.models import Producto
            for p in Producto.objects.filter(id__in=producto_ids).only('id', 'nombre', 'clave'):
                producto_nombres[str(p.id)] = f"{p.nombre} ({p.clave})" if p.clave else p.nombre

        resultados = []
        for log in logs[:500]:
            datos_ant = log.datos_anteriores or {}
            datos_new = log.datos_nuevos or {}

            # Detectar cambios campo por campo
            cambios = []
            cambios_humanizados = []
            campos_a_revisar = set(list(datos_ant.keys()) + list(datos_new.keys()))
            for c in campos_a_revisar:
                if c in CAMPOS_OCULTOS:
                    continue
                val_ant = datos_ant.get(c)
                val_new = datos_new.get(c)
                if val_ant != val_new and val_new is not None:
                    cambio = {
                        'campo': c,
                        'valor_anterior': val_ant,
                        'valor_nuevo': val_new,
                    }
                    cambios.append(cambio)
                    cambios_humanizados.append({
                        'campo': c,
                        'campo_label': CAMPO_LABELS.get(c, c.replace('_', ' ').capitalize()),
                        'valor_anterior_label': humanizar_valor(val_ant),
                        'valor_nuevo_label': humanizar_valor(val_new),
                    })

            # Si se pide filtrar por un campo específico, solo incluir si ese campo cambió
            if campo:
                cambios = [ch for ch in cambios if ch['campo'] == campo]
                cambios_humanizados = [ch for ch in cambios_humanizados if ch['campo'] == campo]
                if not cambios:
                    continue

            usuario_nombre = ''
            if log.usuario:
                usuario_nombre = log.usuario.get_full_name() or log.usuario.username

            # Formato de fecha legible
            fecha_legible = log.timestamp.strftime('%d/%m/%Y %H:%M') if log.timestamp else ''

            # Nombre del producto
            producto_nombre = producto_nombres.get(log.objeto_id, f'Producto #{log.objeto_id}')

            # Rol legible
            rol_legible = ROL_LABELS.get(log.rol_usuario, log.rol_usuario or '')

            resultados.append({
                'id': log.id,
                'fecha': log.timestamp.isoformat(),
                'fecha_legible': fecha_legible,
                'usuario': usuario_nombre,
                'rol': log.rol_usuario or '',
                'rol_legible': rol_legible,
                'producto_id': log.objeto_id,
                'producto_nombre': producto_nombre,
                'accion': log.accion,
                'cambios': cambios,
                'cambios_humanizados': cambios_humanizados,
                'descripcion': generar_narrativa(cambios_humanizados),
                'ip': log.ip_address or '',
            })

        # ── JSON ──
        if formato == 'json':
            return Response({
                'total': len(resultados),
                'resultados': resultados,
            })

        # ── PDF ──
        if formato == 'pdf':
            from core.utils.pdf_reports import generar_reporte_auditoria_productos
            filtros = {
                'fecha_generacion': timezone.now().strftime('%d/%m/%Y %H:%M'),
                'total': len(resultados),
            }
            if fecha_inicio:
                filtros['fecha_inicio'] = fecha_inicio
            if fecha_fin:
                filtros['fecha_fin'] = fecha_fin
            if campo:
                filtros['campo'] = campo
            pdf_buffer = generar_reporte_auditoria_productos(resultados, filtros=filtros)
            response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = (
                f"attachment; filename=Auditoria_Productos_{timezone.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            )
            return response

        # ── Excel ──
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Auditoría Productos'

        ws.merge_cells('A1:G1')
        c = ws['A1']
        c.value = 'AUDITORÍA DE CAMBIOS EN PRODUCTOS'
        c.font = Font(bold=True, size=14, color='632842')
        c.alignment = Alignment(horizontal='center', vertical='center')

        ws.merge_cells('A2:G2')
        c2 = ws['A2']
        c2.value = f'Generado el {timezone.now().strftime("%d/%m/%Y %H:%M")} — Total de registros: {len(resultados)}'
        c2.font = Font(size=10, italic=True)
        c2.alignment = Alignment(horizontal='center')

        ws.append([])
        headers = ['#', 'Fecha', 'Usuario', 'Rol', 'Producto', 'Descripción del cambio']
        ws.append(headers)

        header_fill = PatternFill(start_color='632842', end_color='632842', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        for cell in ws[4]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for idx, item in enumerate(resultados, 1):
            ws.append([
                idx,
                item.get('fecha_legible', item['fecha'][:19].replace('T', ' ')),
                item['usuario'],
                item.get('rol_legible', item['rol']),
                item.get('producto_nombre', item['producto_id']),
                item.get('descripcion', ''),
            ])

        for col, width in zip('ABCDEF', [6, 20, 22, 18, 30, 70]):
            ws.column_dimensions[col].width = width

        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        response = HttpResponse(
            excel_buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = (
            f'attachment; filename=Auditoria_Productos_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
        return response

    except Exception as exc:
        return Response(
            {'error': 'Error al generar auditoría de productos.', 'mensaje': str(exc)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


__all__ = [
    'reporte_inventario',
    'reporte_movimientos',
    'reporte_caducidades',
    'reporte_requisiciones',
    'reporte_medicamentos_por_caducar',
    'reporte_bajo_stock',
    'reporte_consumo',
    'reporte_contratos',
    'reporte_parcialidades',
    'reportes_precarga',
    'reporte_medicamentos_controlados',
    'reporte_auditoria_productos',
]

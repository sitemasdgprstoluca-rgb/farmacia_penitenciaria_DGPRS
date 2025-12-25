# -*- coding: utf-8 -*-
"""
Módulo base para views de inventario.

Contiene imports, constantes, helpers y clases base compartidas
por todos los ViewSets del módulo inventario.

Refactorización audit34: Separación del monolítico views.py (7654 líneas)
en módulos especializados por recurso.
"""
from rest_framework import viewsets, status, serializers, permissions, mixins
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.pagination import PageNumberPagination
from django.core.paginator import InvalidPage
from django.core.cache import cache
from django.http import HttpResponse
from django.db import transaction
from django.db.models import Q, Sum, Count, F, IntegerField
from django.db.models.functions import Coalesce
from django.utils import timezone
from django.conf import settings
from django.contrib.auth.models import Group
from datetime import datetime, timedelta, date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os
import logging
from io import BytesIO

# ISS-004 FIX (audit9): Pillow para validación profunda de imágenes
try:
    from PIL import Image
    PILLOW_AVAILABLE = True
except ImportError:
    PILLOW_AVAILABLE = False
    Image = None

# ISS-011, ISS-021, ISS-030: Import de servicios transaccionales
from inventario.services import CentroPermissionMixin

logger = logging.getLogger(__name__)


# -----------------------------------------------------------
# VALIDADORES DE IMPORTACIÓN EXCEL
# -----------------------------------------------------------

# Magic bytes para formatos Excel válidos
# XLSX/XLSM son archivos ZIP (PK\x03\x04)
# XLS antiguo usa formato OLE2 (Compound File Binary Format)
# ISS-008 FIX (audit7): XLSM eliminado - archivos con macros NO permitidos
EXCEL_MAGIC_BYTES = {
    '.xlsx': b'PK\x03\x04',  # ZIP archive (Office Open XML)
    # '.xlsm': REMOVIDO - ISS-008 FIX: No permitir archivos con macros
    '.xls': b'\xD0\xCF\x11\xE0',  # OLE2 Compound Document
}

# ISS-008 FIX (audit7): Extensiones explícitamente prohibidas por seguridad
EXCEL_EXTENSIONS_BLOCKED = {'.xlsm', '.xlsb', '.xltm', '.xla', '.xlam'}  # Formatos con macros

# ISS-006: Magic bytes y MIME types para imágenes permitidas en firmas
IMAGE_MAGIC_BYTES = {
    '.jpg': [b'\xff\xd8\xff'],  # JPEG
    '.jpeg': [b'\xff\xd8\xff'],  # JPEG
    '.png': [b'\x89PNG\r\n\x1a\n'],  # PNG
    '.gif': [b'GIF87a', b'GIF89a'],  # GIF
    '.webp': [b'RIFF'],  # WebP (RIFF container)
}

ALLOWED_IMAGE_EXTENSIONS = {'.jpg', '.jpeg', '.png', '.gif', '.webp'}
ALLOWED_IMAGE_MIMES = {
    'image/jpeg', 'image/png', 'image/gif', 'image/webp'
}

# ISS-005 FIX (audit7): Magic bytes para validación de PDF
PDF_MAGIC_BYTES = b'%PDF-'  # Todos los PDFs válidos empiezan con este header
PDF_MAX_SIZE_MB = 10  # Tamaño máximo en MB
PDF_MAX_SIZE_BYTES = PDF_MAX_SIZE_MB * 1024 * 1024

# ISS-002 FIX (audit9): Content-Types válidos para PDF (validación estricta)
PDF_VALID_CONTENT_TYPES = {'application/pdf', 'application/x-pdf'}

# ISS-001 FIX (audit9): Estado inicial SOLO borrador para TODOS los usuarios
# El flujo jerárquico OBLIGA: borrador → pendiente_admin → pendiente_director → enviada
# NUNCA se puede crear una requisición directamente en 'enviada'
ESTADO_INICIAL_UNICO = 'borrador'
# Mantener por compatibilidad pero DEPRECADO - usar ESTADO_INICIAL_UNICO
ESTADOS_INICIALES_VALIDOS = {'borrador'}  # ISS-001 FIX: Eliminado 'enviada'
ESTADO_INICIAL_CENTRO = 'borrador'
ESTADO_INICIAL_FARMACIA = 'borrador'


# Imports de modelos y serializers
from core.models import Producto, Lote, Movimiento, Centro, Requisicion, DetalleRequisicion, HojaRecoleccion, LoteDocumento
from core.serializers import (
    ProductoSerializer, LoteSerializer, MovimientoSerializer, 
    CentroSerializer, RequisicionSerializer, DetalleRequisicionSerializer,
    HojaRecoleccionSerializer, LoteDocumentoSerializer
)

from django.contrib.auth import get_user_model
from core.permissions import (
    IsAdminRole, IsFarmaciaRole, IsCentroRole, IsVistaRole,
    IsFarmaciaAdminOrReadOnly, CanAuthorizeRequisicion
)
from core.constants import (
    ESTADOS_REQUISICION,
    PAGINATION_DEFAULT_PAGE_SIZE,
    PAGINATION_MAX_PAGE_SIZE,
    UNIDADES_MEDIDA,
    REQUISICION_GRUPOS_ESTADO,
    TRANSICIONES_REQUISICION,
    PERMISOS_FLUJO_REQUISICION,
)

User = get_user_model()


# ============================================================================
# FUNCIONES DE VALIDACIÓN DE ARCHIVOS
# ============================================================================

def leer_archivo_con_limite(file, max_bytes):
    """
    ISS-001: Lee un archivo con límite estricto de bytes.
    
    Previene DoS por archivos enormes que no declaran tamaño correcto.
    Lee en chunks y aborta si se excede el límite.
    
    Args:
        file: Archivo a leer (UploadedFile o similar)
        max_bytes: Máximo de bytes permitidos
    
    Returns:
        tuple: (BytesIO con contenido, bytes_leidos) o (None, error_message)
    """
    CHUNK_SIZE = 64 * 1024  # 64KB chunks
    buffer = BytesIO()
    bytes_leidos = 0
    
    try:
        # Asegurar que estamos al inicio del archivo
        if hasattr(file, 'seek'):
            file.seek(0)
        
        while True:
            chunk = file.read(CHUNK_SIZE)
            if not chunk:
                break
            
            bytes_leidos += len(chunk)
            
            # ISS-001: Corte estricto si excede el límite
            if bytes_leidos > max_bytes:
                logger.warning(
                    f"Archivo rechazado: tamaño real ({bytes_leidos} bytes) "
                    f"excede límite ({max_bytes} bytes)"
                )
                return None, f'El archivo excede el tamaño máximo permitido ({max_bytes / 1024 / 1024:.1f}MB)'
            
            buffer.write(chunk)
        
        buffer.seek(0)
        return buffer, bytes_leidos
        
    except Exception as e:
        logger.error(f"Error leyendo archivo: {e}")
        return None, f'Error al leer el archivo: {str(e)}'


def validar_archivo_pdf(file, max_size_mb=PDF_MAX_SIZE_MB):
    """
    ISS-005 FIX (audit7): Valida archivo PDF antes de guardarlo.
    
    Validaciones:
    1. Archivo presente y con nombre válido
    2. Extensión .pdf
    3. Tamaño dentro de límites
    4. Magic bytes correctos (%PDF-)
    5. Content-Type MIME válido (si disponible)
    
    Retorna: (es_valido, mensaje_error)
    """
    # 1. Validar que hay archivo
    if not file:
        return False, 'No se recibió archivo PDF'
    
    # 2. Validar nombre obligatorio
    nombre = getattr(file, 'name', None)
    if not nombre or not nombre.strip():
        return False, 'El archivo debe tener un nombre válido'
    
    # 3. Validar extensión
    extension = ('.' + nombre.split('.')[-1].lower()) if '.' in nombre else ''
    if extension != '.pdf':
        return False, f'Solo se permiten archivos PDF (.pdf). Extensión recibida: {extension}'
    
    # 4. Validar tamaño
    max_size_bytes = max_size_mb * 1024 * 1024
    if hasattr(file, 'size') and file.size > max_size_bytes:
        return False, f'El archivo excede el tamaño máximo de {max_size_mb}MB ({file.size / 1024 / 1024:.2f}MB)'
    
    # 5. Validar magic bytes - leer primeros bytes con límite
    try:
        buffer_result, bytes_leidos = leer_archivo_con_limite(file, max_size_bytes + 1024)
        
        if buffer_result is None:
            return False, bytes_leidos  # bytes_leidos contiene el mensaje de error
        
        # Validar tamaño real
        if bytes_leidos > max_size_bytes:
            return False, f'El archivo excede el tamaño máximo de {max_size_mb}MB (tamaño real: {bytes_leidos / 1024 / 1024:.2f}MB)'
        
        # Verificar magic bytes de PDF
        header = buffer_result.read(8)
        buffer_result.seek(0)  # Restaurar posición
        
        if not header.startswith(PDF_MAGIC_BYTES):
            logger.warning(
                f"ISS-005: Archivo {nombre} rechazado - magic bytes incorrectos: {header[:8]!r}"
            )
            return False, 'El archivo no es un PDF válido (magic bytes incorrectos)'
        
    except Exception as e:
        logger.error(f"ISS-005: Error validando PDF {nombre}: {e}")
        return False, f'Error al validar el archivo PDF: {str(e)}'
    
    # 6. ISS-002 FIX (audit9): Validar Content-Type ESTRICTAMENTE
    # No solo advertir, BLOQUEAR archivos con Content-Type sospechoso
    content_type = getattr(file, 'content_type', None)
    if content_type:
        if content_type not in PDF_VALID_CONTENT_TYPES:
            logger.error(
                f"ISS-002: Archivo {nombre} BLOQUEADO - content-type inválido: {content_type}. "
                f"Solo se permiten: {PDF_VALID_CONTENT_TYPES}"
            )
            return False, (
                f'Content-Type inválido: {content_type}. '
                f'Solo se permiten archivos PDF (application/pdf)'
            )
    
    return True, None


def validar_archivo_imagen(file, max_size_mb=2):
    """
    ISS-006: Valida archivo de imagen antes de guardarlo.
    
    Validaciones:
    1. Archivo presente y con nombre válido
    2. Extensión permitida (.jpg, .jpeg, .png, .gif, .webp)
    3. Tamaño dentro de límites (default 2MB)
    4. Magic bytes correctos (contenido real corresponde a imagen)
    5. Content-Type MIME válido (si disponible)
    
    Retorna: (es_valido, mensaje_error)
    """
    # 1. Validar que hay archivo
    if not file:
        return False, 'No se recibió archivo de imagen'
    
    # 2. Validar nombre obligatorio
    nombre = getattr(file, 'name', None)
    if not nombre or not nombre.strip():
        return False, 'El archivo debe tener un nombre válido'
    
    nombre_lower = nombre.lower().strip()
    ext = os.path.splitext(nombre_lower)[1]
    
    # 3. Validar extensión
    if not ext:
        return False, 'El archivo debe tener una extensión (.jpg, .png, etc.)'
    
    if ext not in ALLOWED_IMAGE_EXTENSIONS:
        return False, f'Extensión no permitida: {ext}. Use: {", ".join(ALLOWED_IMAGE_EXTENSIONS)}'
    
    # 4. Validar tamaño
    max_size_bytes = max_size_mb * 1024 * 1024
    file_size = getattr(file, 'size', None)
    if file_size is not None and file_size > max_size_bytes:
        return False, f'Imagen demasiado grande: {file_size / 1024 / 1024:.1f}MB. Máximo: {max_size_mb}MB'
    
    # 5. Validar Content-Type MIME si está disponible
    content_type = getattr(file, 'content_type', None)
    if content_type and content_type.lower() not in ALLOWED_IMAGE_MIMES:
        logger.warning(
            f"Archivo rechazado: content_type {content_type} no permitido. "
            f"Nombre: {nombre}"
        )
        return False, f'Tipo de archivo no permitido: {content_type}. Use imágenes JPG, PNG, GIF o WebP'
    
    # 6. Validar magic bytes - contenido real del archivo
    try:
        pos = file.tell() if hasattr(file, 'tell') else 0
        header = file.read(16)  # Leer suficientes bytes
        
        if hasattr(file, 'seek'):
            file.seek(pos)
        
        if not header or len(header) < 4:
            return False, 'Archivo vacío o corrupto'
        
        # Verificar magic bytes según extensión
        expected_magics = IMAGE_MAGIC_BYTES.get(ext, [])
        if expected_magics:
            match = False
            for magic in expected_magics:
                if header.startswith(magic):
                    match = True
                    break
            
            if not match:
                logger.warning(
                    f"Archivo imagen rechazado: extensión {ext} pero magic bytes incorrectos. "
                    f"Header: {header[:8].hex()}, Nombre: {nombre}"
                )
                return False, f'El contenido del archivo no corresponde a una imagen {ext} válida'
    
    except Exception as e:
        logger.error(f"Error validando magic bytes de imagen: {e}")
        return False, 'Error al validar el contenido del archivo de imagen'
    
    # 7. ISS-004 FIX (audit9): Validación profunda con Pillow
    # Intenta abrir la imagen para detectar archivos corruptos o maliciosos
    if PILLOW_AVAILABLE and Image is not None:
        try:
            pos = file.tell() if hasattr(file, 'tell') else 0
            
            # Leer contenido completo para Pillow
            if hasattr(file, 'seek'):
                file.seek(0)
            content = file.read()
            
            # Restaurar posición
            if hasattr(file, 'seek'):
                file.seek(pos)
            
            # Intentar abrir con Pillow
            img_buffer = BytesIO(content)
            with Image.open(img_buffer) as img:
                # Verificar que realmente se puede decodificar
                img.verify()
                
                # Reabrir para validar dimensiones (verify() invalida el handle)
                img_buffer.seek(0)
                with Image.open(img_buffer) as img_check:
                    width, height = img_check.size
                    
                    # Límites de dimensiones para evitar DoS
                    MAX_IMAGE_DIMENSION = 4096  # 4K máximo
                    if width > MAX_IMAGE_DIMENSION or height > MAX_IMAGE_DIMENSION:
                        logger.warning(
                            f"Imagen rechazada por dimensiones excesivas: {width}x{height}, "
                            f"máximo: {MAX_IMAGE_DIMENSION}x{MAX_IMAGE_DIMENSION}"
                        )
                        return False, f'Imagen demasiado grande: {width}x{height}. Máximo: {MAX_IMAGE_DIMENSION}x{MAX_IMAGE_DIMENSION} píxeles'
                    
                    # Verificar formato reportado por Pillow
                    pillow_format = img_check.format
                    expected_formats = {
                        '.jpg': ['JPEG'],
                        '.jpeg': ['JPEG'],
                        '.png': ['PNG'],
                        '.gif': ['GIF'],
                        '.webp': ['WEBP'],
                    }
                    
                    if ext in expected_formats and pillow_format not in expected_formats[ext]:
                        logger.warning(
                            f"Imagen rechazada: extensión {ext} pero Pillow detectó {pillow_format}. "
                            f"Nombre: {nombre}"
                        )
                        return False, f'El contenido del archivo ({pillow_format}) no corresponde a la extensión {ext}'
            
            # ISS-006 FIX (audit11): Asegurar que el stream quede en posición original
            if hasattr(file, 'seek'):
                file.seek(0)
                    
        except Exception as e:
            logger.warning(
                f"Imagen rechazada: Pillow no pudo abrir/verificar el archivo. "
                f"Error: {e}, Nombre: {nombre}"
            )
            return False, 'Imagen corrupta o formato no válido - no se puede procesar'
    else:
        # ISS-006 FIX (audit11): Rechazar imágenes si Pillow no está disponible
        # La validación solo con magic bytes no es suficiente para seguridad
        logger.error(
            "Pillow NO disponible - rechazando imagen por seguridad. "
            "Instale Pillow con: pip install Pillow"
        )
        return False, 'Validación de imágenes no disponible. Contacte al administrador.'
    
    # ISS-006 FIX (audit11): Asegurar stream en posición 0 para el caller
    if hasattr(file, 'seek'):
        file.seek(0)
    
    return True, None


def validar_archivo_excel(file):
    """
    ISS-008 FIX (audit7): Valida archivo Excel antes de procesarlo con múltiples capas de seguridad.
    
    Validaciones:
    1. Archivo presente y con nombre válido
    2. Extensión permitida (.xlsx, .xls) - NO xlsm ni otros con macros
    3. Tamaño dentro de límites
    4. Magic bytes correctos (contenido real coincide con extensión)
    
    ISS-008 FIX (audit7): Archivos con macros (.xlsm, .xlsb, etc) están BLOQUEADOS
    por seguridad. Solo se permiten formatos sin código ejecutable.
    
    Retorna: (es_valido, mensaje_error)
    """
    # 1. Validar que hay archivo
    if not file:
        return False, 'No se recibió archivo'
    
    # 2. ISS-001: Validar nombre obligatorio - rechazar archivos sin nombre
    nombre = getattr(file, 'name', None)
    if not nombre or not nombre.strip():
        return False, 'El archivo debe tener un nombre válido'
    
    nombre_lower = nombre.lower().strip()
    ext = os.path.splitext(nombre_lower)[1]
    
    # 3. Validar extensión
    if not ext:
        return False, 'El archivo debe tener una extensión (.xlsx o .xls)'
    
    # ISS-008 FIX (audit7): Bloquear extensiones con macros ANTES de cualquier otra validación
    if ext in EXCEL_EXTENSIONS_BLOCKED:
        logger.warning(
            f"ISS-008: Archivo rechazado por extensión con macros: {nombre}. "
            f"Extensiones bloqueadas: {EXCEL_EXTENSIONS_BLOCKED}"
        )
        return False, (
            f'Extensión {ext} no permitida por seguridad. '
            f'Los archivos con macros (.xlsm, .xlsb, etc.) están bloqueados. '
            f'Por favor convierta a .xlsx (sin macros) antes de importar.'
        )
    
    extensiones_permitidas = getattr(settings, 'IMPORT_ALLOWED_EXTENSIONS', ['.xlsx', '.xls'])
    if ext not in extensiones_permitidas:
        return False, f'Extensión no permitida: {ext}. Use: {", ".join(extensiones_permitidas)}'
    
    # 4. Validar tamaño declarado ANTES de leer contenido (primera línea de defensa)
    max_size_mb = getattr(settings, 'IMPORT_MAX_FILE_SIZE_MB', 10)
    max_size_bytes = max_size_mb * 1024 * 1024
    
    file_size = getattr(file, 'size', None)
    if file_size is not None and file_size > max_size_bytes:
        return False, f'Archivo demasiado grande: {file_size / 1024 / 1024:.1f}MB. Máximo: {max_size_mb}MB'
    
    # 5. ISS-001: Validar magic bytes - contenido real del archivo
    try:
        # Guardar posición actual y leer primeros bytes
        pos = file.tell() if hasattr(file, 'tell') else 0
        header = file.read(8)  # Leer suficientes bytes para identificar formato
        
        # Restaurar posición para que el archivo pueda ser procesado después
        if hasattr(file, 'seek'):
            file.seek(pos)
        
        if not header or len(header) < 4:
            return False, 'Archivo vacío o corrupto'
        
        # Verificar magic bytes según extensión
        expected_magic = EXCEL_MAGIC_BYTES.get(ext)
        if expected_magic:
            if not header.startswith(expected_magic):
                # El contenido no coincide con la extensión declarada
                logger.warning(
                    f"Archivo rechazado: extensión {ext} pero magic bytes incorrectos. "
                    f"Header: {header[:8].hex()}"
                )
                return False, f'El contenido del archivo no corresponde a un archivo {ext} válido'
    except Exception as e:
        logger.error(f"Error validando magic bytes: {e}")
        return False, 'Error al validar el contenido del archivo'
    
    return True, None


def cargar_workbook_seguro(file):
    """
    ISS-001: Carga un workbook Excel con límites de seguridad estrictos.
    
    1. Lee el archivo con límite real de bytes (no confía en file.size)
    2. Usa read_only=True para streaming y menor uso de memoria
    3. Usa data_only=True para ignorar fórmulas (previene ataques por fórmulas)
    
    Returns:
        tuple: (workbook, error_message) - workbook es None si hay error
    """
    max_size_mb = getattr(settings, 'IMPORT_MAX_FILE_SIZE_MB', 10)
    max_size_bytes = max_size_mb * 1024 * 1024
    
    # Leer archivo con límite estricto de bytes
    buffer, resultado = leer_archivo_con_limite(file, max_size_bytes)
    
    if buffer is None:
        # resultado contiene el mensaje de error
        return None, resultado
    
    try:
        # ISS-001: Usar read_only=True para streaming y data_only=True para ignorar fórmulas
        # Esto reduce significativamente el uso de memoria en archivos grandes
        wb = openpyxl.load_workbook(
            buffer, 
            read_only=True,  # Streaming mode - no carga todo en memoria
            data_only=True   # Ignora fórmulas - solo valores (previene ataques)
        )
        return wb, None
    except Exception as e:
        logger.error(f"Error cargando workbook: {e}")
        return None, f'Error al procesar el archivo Excel: {str(e)}'


def validar_filas_excel(ws):
    """
    Valida número de filas en worksheet con streaming y corte temprano.
    
    ISS-003: Optimizado para no recorrer archivos enormes innecesariamente.
    Se detiene al superar el límite máximo permitido.
    
    Retorna: (es_valido, mensaje_error, num_filas)
    """
    max_rows = getattr(settings, 'IMPORT_MAX_ROWS', 5000)
    # Margen adicional para detener el conteo (evitar contar millones de filas)
    cutoff = max_rows + 1
    
    num_filas = 0
    # ISS-003: Streaming con corte temprano - detenerse apenas superamos el límite
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Solo contar filas que tienen al menos un valor no vacío
        if any(cell is not None and str(cell).strip() for cell in row):
            num_filas += 1
            if num_filas > max_rows:
                # Corte temprano: ya sabemos que excede el límite
                return False, f'El archivo excede el máximo de {max_rows} filas permitidas', num_filas
    
    return True, None, num_filas


# ============================================================================
# HELPERS DE SEGURIDAD - Filtrado por centro para roles no-admin
# ISS-AUDIT FIX: Separación de funciones para claridad y seguridad
# ============================================================================

def is_farmacia_or_admin(user):
    """
    ISS-AUDIT FIX: Verifica si el usuario es farmacia o admin (roles de escritura).
    
    IMPORTANTE: Esta función NO incluye 'vista' - usar has_global_read_access()
    para operaciones de solo lectura que incluyan vista.
    
    Roles con acceso de escritura:
    - admin: admin_sistema, superusuario, administrador
    - farmacia: farmacia, admin_farmacia, farmaceutico, usuario_farmacia
    """
    if not user or not user.is_authenticated:
        return False
    if user.is_superuser:
        return True
    
    rol = (getattr(user, 'rol', '') or '').lower()
    
    # ISS-AUDIT FIX: Solo roles admin y farmacia (NO vista)
    ROLES_ESCRITURA = {
        # Admin
        'admin', 'admin_sistema', 'superusuario', 'administrador',
        # Farmacia
        'farmacia', 'admin_farmacia', 'farmaceutico', 'usuario_farmacia',
    }
    
    if rol in ROLES_ESCRITURA:
        return True
    
    # Verificar grupos (por si el rol no esta en campo directo)
    group_names = {g.name.upper() for g in user.groups.all()}
    GRUPOS_ESCRITURA = {'FARMACIA_ADMIN', 'FARMACEUTICO'}
    
    return bool(group_names & GRUPOS_ESCRITURA)


def has_global_read_access(user):
    """
    ISS-AUDIT FIX: Verifica si el usuario puede leer datos globales (incluye vista).
    
    Usar esta función para operaciones de SOLO LECTURA donde el rol vista
    debe tener acceso.
    
    Roles con lectura global:
    - admin, farmacia (heredan de is_farmacia_or_admin)
    - vista: vista, usuario_vista (solo lectura)
    """
    # Admin y farmacia siempre tienen acceso de lectura
    if is_farmacia_or_admin(user):
        return True
    
    if not user or not user.is_authenticated:
        return False
    
    rol = (getattr(user, 'rol', '') or '').lower()
    
    # Roles de solo lectura global
    ROLES_VISTA = {'vista', 'usuario_vista'}
    
    if rol in ROLES_VISTA:
        return True
    
    # Verificar grupos
    group_names = {g.name.upper() for g in user.groups.all()}
    return 'VISTA_USER' in group_names


def get_user_centro(user):
    """Obtiene el centro asignado al usuario."""
    return getattr(user, 'centro', None) or getattr(getattr(user, 'profile', None), 'centro', None)


def invalidar_cache_dashboard(centro_id=None):
    """
    ISS-005: Invalida el caché del dashboard cuando hay cambios en datos.
    
    Se llama automáticamente al registrar movimientos, crear requisiciones, etc.
    
    Args:
        centro_id: ID del centro afectado (opcional). Si es None, invalida solo el global.
                   Si se pasa, invalida tanto el del centro como el global.
    """
    try:
        # Siempre invalidar el global
        cache.delete('dashboard_resumen_global')
        cache.delete('dashboard_graficas_global')
        
        # Si hay un centro específico, invalidar también ese
        if centro_id:
            cache.delete(f'dashboard_resumen_{centro_id}')
            cache.delete(f'dashboard_graficas_{centro_id}')
    except Exception as e:
        # No fallar si hay problemas con el caché
        logger.warning(f'Error al invalidar caché del dashboard: {e}')


def registrar_movimiento_stock(*, lote, tipo, cantidad, usuario=None, centro=None, requisicion=None, observaciones='', skip_centro_check=False, subtipo_salida=None, numero_expediente=None):
    """
    Helper central para registrar un movimiento y actualizar cantidad_actual del lote.
    
    ISS-002: Incluye validación de pertenencia de centro para prevenir manipulación
    de inventario entre centros.
    
    ISS-003 FIX: Los ajustes negativos requieren observaciones obligatorias con
    longitud mínima para auditoría y prevención de robo hormiga.
    
    MEJORA FLUJO 5: Soporte para subtipo_salida y numero_expediente para
    trazabilidad de pacientes en salidas por receta médica.
    
    Args:
        lote: Instancia del Lote a modificar
        tipo: 'entrada', 'salida' o 'ajuste'
        cantidad: Cantidad a mover (positivo)
        usuario: Usuario que realiza la operación (opcional)
        centro: Centro donde se registra el movimiento (opcional)
        requisicion: Requisición asociada (opcional)
        observaciones: Texto descriptivo (obligatorio para ajustes negativos)
        skip_centro_check: Si True, omite validación de pertenencia (solo para operaciones
                          de sistema/admin que ya validaron permisos)
        subtipo_salida: Tipo de salida (receta, consumo_interno, merma, etc.)
        numero_expediente: Número de expediente del paciente (para recetas)
    
    Raises:
        ValidationError: Si los datos son inválidos o hay problemas de autorización
    
    Returns:
        tuple: (movimiento_creado, lote_actualizado)
    """
    tipo_normalizado = (tipo or '').lower()
    if tipo_normalizado not in ('entrada', 'salida', 'ajuste'):
        raise serializers.ValidationError({'tipo': 'Tipo de movimiento no valido'})
    if cantidad is None:
        raise serializers.ValidationError({'cantidad': 'Cantidad requerida'})
    try:
        cantidad_int = int(cantidad)
    except (TypeError, ValueError):
        raise serializers.ValidationError({'cantidad': 'La cantidad debe ser un numero entero'})

    # =========================================================================
    # ISS-003 FIX: Validación de observaciones para ajustes negativos
    # =========================================================================
    # Los ajustes que reducen stock (negativos) requieren justificación obligatoria
    # para prevenir robo hormiga y asegurar trazabilidad de pérdidas
    LONGITUD_MINIMA_OBSERVACION = 10
    es_ajuste_negativo = tipo_normalizado == 'ajuste' and cantidad_int < 0
    
    if es_ajuste_negativo:
        if not observaciones or len(observaciones.strip()) < LONGITUD_MINIMA_OBSERVACION:
            raise serializers.ValidationError({
                'observaciones': f'Los ajustes negativos requieren una justificación de al menos {LONGITUD_MINIMA_OBSERVACION} caracteres. '
                                 f'Explique el motivo del ajuste (merma, caducidad, rotura, etc.)'
            })

    # =========================================================================
    # ISS-002: Validación de pertenencia de centro
    # =========================================================================
    # Si se pasa un centro explícito, verificar que el lote pertenezca a ese centro
    # Esto previene que usuarios de un centro manipulen stock de otros centros
    if not skip_centro_check and centro is not None:
        lote_centro = getattr(lote, 'centro', None)
        if lote_centro is not None and lote_centro.pk != centro.pk:
            logger.warning(
                f"Intento de operación de stock rechazado: "
                f"usuario={getattr(usuario, 'username', 'N/A')}, "
                f"lote_centro={lote_centro.pk}, centro_solicitado={centro.pk}"
            )
            raise serializers.ValidationError({
                'centro': 'El lote no pertenece al centro especificado. Operación no autorizada.'
            })
    
    # Si el usuario no es admin/farmacia, verificar que tenga acceso al centro del lote
    if not skip_centro_check and usuario is not None and hasattr(usuario, 'is_authenticated') and usuario.is_authenticated:
        if not is_farmacia_or_admin(usuario):
            user_centro = get_user_centro(usuario)
            lote_centro = getattr(lote, 'centro', None)
            if lote_centro is not None and user_centro is not None:
                if user_centro.pk != lote_centro.pk:
                    logger.warning(
                        f"Acceso no autorizado a lote de otro centro: "
                        f"usuario={usuario.username}, user_centro={user_centro.pk}, "
                        f"lote_centro={lote_centro.pk}"
                    )
                    raise serializers.ValidationError({
                        'lote': 'No tiene permiso para operar sobre lotes de otros centros.'
                    })

    delta = cantidad_int
    if tipo_normalizado == 'salida' and delta > 0:
        delta = -delta
    if tipo_normalizado == 'entrada' and delta < 0:
        delta = abs(delta)

    with transaction.atomic():
        lote_ref = Lote.objects.select_for_update().get(pk=lote.pk)
        stock_disponible = lote_ref.cantidad_actual

        if delta < 0 and abs(delta) > stock_disponible:
            raise serializers.ValidationError({
                'cantidad': f'Stock insuficiente en el lote (disponible {stock_disponible}).'
            })

        nuevo_stock = stock_disponible + delta
        if nuevo_stock < 0:
            raise serializers.ValidationError({'cantidad': 'La operacion dejaria el lote con stock negativo'})

        # Actualizar stock y estado de disponibilidad
        lote_ref.cantidad_actual = nuevo_stock
        
        # Para entradas, tambien actualizar cantidad_inicial si es necesario
        # Esto permite recibir stock adicional en lotes existentes
        update_fields = ['cantidad_actual', 'activo', 'updated_at']
        if tipo_normalizado == 'entrada' and nuevo_stock > lote_ref.cantidad_inicial:
            lote_ref.cantidad_inicial = nuevo_stock
            update_fields.append('cantidad_inicial')
        
        if nuevo_stock == 0:
            lote_ref.activo = False
        elif not lote_ref.activo:
            lote_ref.activo = True
        lote_ref.save(update_fields=update_fields)

        # Crear movimiento con campos correctos de la BD
        # ISS-FIX: Para salidas (transferencias), centro_destino es el destino, centro_origen es el origen del lote
        # Para entradas, centro_destino es donde entra el stock
        if tipo_normalizado == 'salida':
            # Salida: el lote sale del centro del lote (o Farmacia Central si es None) hacia el centro especificado
            mov_centro_origen = lote_ref.centro  # None si es Farmacia Central
            mov_centro_destino = centro  # El centro destino de la transferencia
        elif tipo_normalizado == 'entrada':
            # Entrada: el stock entra al centro especificado
            mov_centro_origen = None  # No aplica para entradas
            mov_centro_destino = centro
        else:
            # Ajuste: se hace en el centro del lote
            mov_centro_origen = lote_ref.centro
            mov_centro_destino = None
        
        movimiento = Movimiento(
            tipo=tipo_normalizado,
            producto=lote_ref.producto,
            lote=lote_ref,
            centro_destino=mov_centro_destino,
            centro_origen=mov_centro_origen,
            requisicion=requisicion,
            usuario=usuario if usuario and getattr(usuario, 'is_authenticated', False) else None,
            cantidad=delta,
            motivo=observaciones or '',
            # MEJORA FLUJO 5: Campos de trazabilidad para pacientes
            subtipo_salida=subtipo_salida if tipo_normalizado == 'salida' else None,
            numero_expediente=numero_expediente if tipo_normalizado == 'salida' and subtipo_salida == 'receta' else None
        )
        # Guardar stock previo para evitar fallos de validacion al crear el movimiento
        movimiento._stock_pre_movimiento = stock_disponible
        movimiento.save()
        
        # ISS-005: Invalidar caché del dashboard al registrar movimientos
        centro_afectado = centro.id if centro else (lote_ref.centro.id if lote_ref.centro else None)
        invalidar_cache_dashboard(centro_afectado)

    return movimiento, lote_ref


# ============================================================================
# PAGINACIÓN UNIFICADA
# ============================================================================

class CustomPagination(PageNumberPagination):
    """Paginación unificada para todos los listados del API."""
    page_size = PAGINATION_DEFAULT_PAGE_SIZE
    page_size_query_param = 'page_size'
    max_page_size = PAGINATION_MAX_PAGE_SIZE

    def paginate_queryset(self, queryset, request, view=None):
        try:
            return super().paginate_queryset(queryset, request, view=view)
        except InvalidPage:
            # Si la página solicitada no existe, devolver la última disponible sin 404
            self.page = self.paginator.page(self.paginator.num_pages or 1)
            self.request = request
            return list(self.page)

# -*- coding: utf-8 -*-
"""
ViewSet de Centros Penitenciarios.

Gestión completa de centros incluyendo:
- CRUD de centros
- Importación/exportación Excel
- Inventario por centro
- Requisiciones por centro
"""
from rest_framework import status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.db.models import Q, Sum
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import logging

from .base import (
    viewsets,
    serializers,
    Centro,
    Lote,
    Movimiento,
    Producto,
    CentroSerializer,
    CustomPagination,
    is_farmacia_or_admin,
    validar_archivo_excel,
    cargar_workbook_seguro,
    validar_filas_excel,
    IsFarmaciaAdminOrReadOnly,
)

logger = logging.getLogger(__name__)


class CentroViewSet(viewsets.ModelViewSet):
    """
    ViewSet para gestionar centros penitenciarios.
    
    Funcionalidades:
    - CRUD completo
    - Búsqueda por clave, nombre y dirección
    - Filtrado por estado activo/inactivo
    - Exportar a Excel con formato profesional
    - Importar desde Excel con validaciones
    - Obtener requisiciones por centro
    """
    queryset = Centro.objects.all()
    serializer_class = CentroSerializer
    permission_classes = [IsFarmaciaAdminOrReadOnly]
    pagination_class = CustomPagination

    def _user_centro(self, user):
        return getattr(user, 'centro', None) or getattr(getattr(user, 'profile', None), 'centro', None)

    def get_queryset(self):
        """Filtra centros según parámetros"""
        queryset = Centro.objects.all()
        user = getattr(self.request, 'user', None)
        
        # Admin, Farmacia y Superusuarios pueden ver todos los centros
        # Otros usuarios solo ven su propio centro
        if user and not user.is_superuser and not is_farmacia_or_admin(user):
            user_centro = self._user_centro(user)
            if user_centro:
                queryset = queryset.filter(id=user_centro.id)
            else:
                return Centro.objects.none()
        
        # Filtro por búsqueda
        search = self.request.query_params.get('search')
        if search and search.strip():
            queryset = queryset.filter(
                Q(nombre__icontains=search) | 
                Q(direccion__icontains=search) |
                Q(email__icontains=search) |
                Q(telefono__icontains=search)
            )
        
        # Filtro por estado activo
        activo = self.request.query_params.get('activo')
        if activo == 'true':
            queryset = queryset.filter(activo=True)
        elif activo == 'false':
            queryset = queryset.filter(activo=False)
        
        # Ordenamiento (respeta parámetro del frontend)
        ordering = self.request.query_params.get('ordering', '-created_at')
        valid_orderings = ['nombre', '-nombre', 'created_at', '-created_at', 'activo', '-activo']
        if ordering in valid_orderings:
            queryset = queryset.order_by(ordering)
        else:
            queryset = queryset.order_by('-created_at')
        
        return queryset
    
    def create(self, request, *args, **kwargs):
        """Crea un nuevo centro"""
        try:
            logger.debug(f"CREAR CENTRO - Body: {request.data}")
            
            serializer = self.get_serializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            centro = serializer.save()
            
            logger.info(f"Centro creado: {centro.clave} - {centro.nombre}")
            
            return Response({
                'mensaje': 'Centro creado exitosamente',
                'centro': CentroSerializer(centro).data
            }, status=status.HTTP_201_CREATED)
            
        except serializers.ValidationError as e:
            logger.warning(f"Error de validación al crear centro: {e.detail}")
            return Response({
                'error': 'Error de validación',
                'detalles': e.detail
            }, status=status.HTTP_400_BAD_REQUEST)
            
        except Exception as e:
            logger.exception(f"Error inesperado al crear centro: {e}")
            # traceback removido por seguridad (ISS-008)
            return Response({
                'error': 'Error al crear centro',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def update(self, request, *args, **kwargs):
        """Actualiza un centro existente"""
        try:
            partial = kwargs.pop('partial', False)
            instance = self.get_object()
            serializer = self.get_serializer(instance, data=request.data, partial=partial)
            serializer.is_valid(raise_exception=True)
            self.perform_update(serializer)
            
            return Response({
                'mensaje': 'Centro actualizado exitosamente',
                'centro': serializer.data
            })
        except serializers.ValidationError as e:
            return Response(
                {'error': 'Error de validación', 'detalles': e.detail},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            return Response(
                {'error': 'Error al actualizar centro', 'mensaje': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def destroy(self, request, *args, **kwargs):
        """
        Elimina un centro.
        
        Validaciones:
        - No puede eliminarse si tiene requisiciones asociadas (como origen o destino)
        - No puede eliminarse si tiene usuarios asignados
        """
        instance = self.get_object()
        
        try:
            # Verificar requisiciones (como origen O destino)
            has_req_origen = hasattr(instance, 'requisiciones_origen') and instance.requisiciones_origen.exists()
            has_req_destino = hasattr(instance, 'requisiciones_destino') and instance.requisiciones_destino.exists()
            
            if has_req_origen or has_req_destino:
                total_origen = instance.requisiciones_origen.count() if has_req_origen else 0
                total_destino = instance.requisiciones_destino.count() if has_req_destino else 0
                total_requisiciones = total_origen + total_destino
                
                # Contar requisiciones activas
                activas_origen = instance.requisiciones_origen.exclude(
                    estado__in=['CANCELADA', 'SURTIDA']
                ).count() if has_req_origen else 0
                activas_destino = instance.requisiciones_destino.exclude(
                    estado__in=['CANCELADA', 'SURTIDA']
                ).count() if has_req_destino else 0
                requisiciones_activas = activas_origen + activas_destino
                
                return Response({
                    'error': 'No se puede eliminar el centro',
                    'razon': 'Tiene requisiciones asociadas',
                    'total_requisiciones': total_requisiciones,
                    'como_origen': total_origen,
                    'como_destino': total_destino,
                    'requisiciones_activas': requisiciones_activas,
                    'sugerencia': 'Marque el centro como inactivo en lugar de eliminarlo'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Verificar usuarios asignados (solo activos, consistente con serializer)
            if hasattr(instance, 'usuarios'):
                usuarios_activos = instance.usuarios.filter(is_active=True)
                if usuarios_activos.exists():
                    total_usuarios = usuarios_activos.count()
                    
                    return Response({
                        'error': 'No se puede eliminar el centro',
                        'razon': 'Tiene usuarios activos asignados',
                        'total_usuarios': total_usuarios,
                        'sugerencia': 'Reasigne los usuarios a otro centro o marque el centro como inactivo'
                    }, status=status.HTTP_400_BAD_REQUEST)
            
            # Verificar lotes con stock activo
            if hasattr(instance, 'lotes'):
                lotes_con_stock = instance.lotes.filter(activo=True, cantidad_actual__gt=0).count()
                if lotes_con_stock > 0:
                    return Response({
                        'error': 'No se puede eliminar el centro',
                        'razon': 'Tiene lotes con stock disponible',
                        'lotes_con_stock': lotes_con_stock,
                        'sugerencia': 'Transfiera el inventario a otro centro o marque el centro como inactivo'
                    }, status=status.HTTP_400_BAD_REQUEST)
            
            # Si no tiene relaciones, se puede eliminar
            clave_eliminada = instance.clave
            nombre_eliminado = instance.nombre
            instance.delete()
            
            return Response({
                'mensaje': 'Centro eliminado exitosamente',
                'centro_eliminado': f"{clave_eliminada} - {nombre_eliminado}"
            }, status=status.HTTP_204_NO_CONTENT)
            
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            return Response({
                'error': 'Error al eliminar centro',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['post'], url_path='toggle-activo')
    def toggle_activo(self, request, pk=None):
        """
        Activa o desactiva un centro.
        POST /api/centros/{id}/toggle-activo/
        
        Usa update() directo para evitar validación de otros campos.
        """
        try:
            centro = self.get_object()
            nuevo_estado = not centro.activo
            
            # Usar update() directo para evitar validación de otros campos
            Centro.objects.filter(pk=centro.pk).update(activo=nuevo_estado)
            
            estado = 'activado' if nuevo_estado else 'desactivado'
            return Response({
                'mensaje': f'Centro {estado} exitosamente',
                'activo': nuevo_estado,
                'id': centro.id,
                'clave': centro.clave,
                'nombre': centro.nombre
            }, status=status.HTTP_200_OK)
        except Centro.DoesNotExist:
            return Response({'error': 'Centro no encontrado'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"Error en toggle_activo centro: {str(e)}", exc_info=True)
            return Response({'error': f'Error interno: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['get'])
    def inventario(self, request, pk=None):
        """Devuelve inventario resumido del centro a partir de lotes asociados a movimientos del centro."""
        centro = self.get_object()
        user_centro = self._user_centro(request.user)
        if not request.user.is_superuser:
            if not user_centro or user_centro.id != centro.id:
                return Response({'error': 'Solo puedes ver inventario de tu centro'}, status=status.HTTP_403_FORBIDDEN)

        # Lotes que han tenido movimientos en este centro
        lote_ids = Movimiento.objects.filter(
            Q(centro_origen=centro) | Q(centro_destino=centro)
        ).values_list('lote_id', flat=True)
        lotes = Lote.objects.filter(
            Q(id__in=lote_ids) | Q(centro=centro),
            activo=True,
            cantidad_actual__gt=0
        ).select_related('producto')

        inventario_dict = {}
        for lote in lotes:
            prod = lote.producto
            item = inventario_dict.setdefault(prod.id, {
                'producto_id': prod.id,
                'clave': prod.clave,
                'producto': prod.nombre,
                'cantidad_disponible': 0,
                'lote_proximo_caducar': None,
                'fecha_caducidad': None,
            })
            item['cantidad_disponible'] += lote.cantidad_actual
            if lote.fecha_caducidad:
                fecha_actual = item['fecha_caducidad']
                if fecha_actual is None or lote.fecha_caducidad < fecha_actual:
                    item['lote_proximo_caducar'] = lote.numero_lote
                    item['fecha_caducidad'] = lote.fecha_caducidad

        # Si no hay lotes asociados, caer al agregado por movimientos para no dejar vacío
        inventario = list(inventario_dict.values())
        if not inventario:
            movimientos = Movimiento.objects.filter(
                Q(centro_origen=centro) | Q(centro_destino=centro)
            )
            agregados = movimientos.values('lote__producto').annotate(cantidad=Coalesce(Sum('cantidad'), 0))
            for item in agregados:
                producto = Producto.objects.filter(id=item['lote__producto']).first()
                if not producto:
                    continue
                inventario.append({
                    'producto_id': producto.id,
                    'clave': producto.clave,
                    'producto': producto.nombre,
                    'cantidad_disponible': max(0, item['cantidad']),
                    'lote_proximo_caducar': None,
                    'fecha_caducidad': None,
                })

        return Response({
            'centro': centro.nombre,
            'centro_id': centro.id,
            'total_productos': len(inventario),
            'inventario': inventario
        })
    
    @action(detail=False, methods=['get'], url_path='exportar-excel')
    def exportar_excel(self, request):
        """
        Exporta todos los centros a Excel con formato profesional.
        
        Columnas (alineadas con BD):
        - #, Nombre, Dirección, Teléfono, Email, Total Requisiciones, Estado
        """
        try:
            centros = self.get_queryset()
            
            # Crear libro de Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Centros Penitenciarios'
            
            # Título del reporte
            ws.merge_cells('A1:G1')
            titulo_cell = ws['A1']
            titulo_cell.value = 'REPORTE DE CENTROS PENITENCIARIOS'
            titulo_cell.font = Font(bold=True, size=14, color='632842')
            titulo_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Fecha de generación
            ws.merge_cells('A2:G2')
            fecha_cell = ws['A2']
            fecha_cell.value = f'Generado el {timezone.now().strftime("%d/%m/%Y %H:%M")}'
            fecha_cell.font = Font(size=10, italic=True)
            fecha_cell.alignment = Alignment(horizontal='center')
            
            # Espacio
            ws.append([])
            
            # Encabezados (sin Clave - campo no existe en BD)
            headers = ['#', 'Nombre', 'Dirección', 'Teléfono', 'Email', 'Total Requisiciones', 'Estado']
            ws.append(headers)
            
            # Estilo de encabezados
            header_fill = PatternFill(start_color='632842', end_color='632842', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            for col_num, cell in enumerate(ws[4], 1):
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # Datos de centros
            for idx, centro in enumerate(centros, start=1):
                # Calcular total de requisiciones (origen + destino, consistente con serializer)
                total_requisiciones = 0
                if hasattr(centro, 'requisiciones_origen'):
                    total_requisiciones += centro.requisiciones_origen.count()
                if hasattr(centro, 'requisiciones_destino'):
                    total_requisiciones += centro.requisiciones_destino.count()
                
                ws.append([
                    idx,
                    centro.nombre,
                    centro.direccion or '',
                    centro.telefono or '',
                    getattr(centro, 'email', '') or '',
                    total_requisiciones,
                    'Activo' if centro.activo else 'Inactivo'
                ])
                
                # Estilo para filas
                row_num = idx + 4
                for cell in ws[row_num]:
                    cell.alignment = Alignment(vertical='center')
                    
                # Colorear estado
                estado_cell = ws.cell(row=row_num, column=7)
                if centro.activo:
                    estado_cell.fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
                    estado_cell.font = Font(color='155724', bold=True)
                else:
                    estado_cell.fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')
                    estado_cell.font = Font(color='721C24', bold=True)
            
            # Ajustar anchos de columna
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 50
            ws.column_dimensions['C'].width = 40
            ws.column_dimensions['D'].width = 18
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 20
            ws.column_dimensions['G'].width = 12
            
            # Agregar bordes
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=1, max_col=7):
                for cell in row:
                    cell.border = thin_border
            
            # Resumen al final
            ws.append([])
            resumen_row = ws.max_row + 1
            ws.merge_cells(f'A{resumen_row}:C{resumen_row}')
            resumen_cell = ws[f'A{resumen_row}']
            resumen_cell.value = f'TOTAL DE CENTROS: {centros.count()}'
            resumen_cell.font = Font(bold=True, size=11)
            resumen_cell.alignment = Alignment(horizontal='left')
            
            # Generar respuesta
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename=Centros_Penitenciarios_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            wb.save(response)
            
            return response
            
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            return Response({
                'error': 'Error al exportar centros',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['post'], url_path='importar')
    def importar_excel(self, request):
        """
        Importa centros desde Excel.
        
        Formato esperado (columnas en orden):
        1. Nombre (REQUERIDO, único) - Nombre del centro penitenciario
        2. Dirección (opcional) - Dirección física
        3. Teléfono (opcional) - Número de teléfono
        4. Email (opcional) - Correo electrónico
        5. Estado (opcional, default: Activo) - 'Activo' o 'Inactivo'
        
        Límites de seguridad:
        - Tamaño máximo: configurado en IMPORT_MAX_FILE_SIZE_MB (default 10MB)
        - Filas máximas: configurado en IMPORT_MAX_ROWS (default 5000)
        - Extensiones: .xlsx, .xls
        
        Nota: Si el nombre ya existe, se actualizan los demás campos.
        """
        file = request.FILES.get('file')
        
        # Validar archivo
        es_valido, error_msg = validar_archivo_excel(file)
        if not es_valido:
            return Response({
                'error': 'Archivo inválido',
                'mensaje': error_msg
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            # ISS-001: Usar carga segura con límite real de bytes
            wb, error_carga = cargar_workbook_seguro(file)
            if wb is None:
                return Response({
                    'error': 'Error al procesar archivo',
                    'mensaje': error_carga
                }, status=status.HTTP_400_BAD_REQUEST)
            
            ws = wb.active
            
            # Validar número de filas
            filas_validas, error_filas, num_filas = validar_filas_excel(ws)
            if not filas_validas:
                wb.close()  # Liberar recursos en modo read_only
                return Response({
                    'error': 'Archivo demasiado grande',
                    'mensaje': error_filas
                }, status=status.HTTP_400_BAD_REQUEST)
            
            creados = 0
            actualizados = 0
            errores = []
            
            # Procesar filas
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Saltar filas vacías
                    if not row or all(cell is None or str(cell).strip() == '' for cell in row):
                        continue
                    
                    # Extraer datos - Formato: Nombre, Dirección, Teléfono, Email, Estado
                    # (mínimo 1 columna: nombre)
                    nombre = row[0] if len(row) > 0 else None
                    direccion = row[1] if len(row) > 1 else None
                    telefono = row[2] if len(row) > 2 else None
                    email = row[3] if len(row) > 3 else None
                    estado = row[4] if len(row) > 4 else 'Activo'
                    
                    # Validar requeridos
                    if not nombre or str(nombre).strip() == '':
                        errores.append({'fila': row_idx, 'error': 'Nombre es requerido'})
                        continue
                    
                    nombre_limpio = str(nombre).strip()
                    
                    # Preparar datos
                    datos = {
                        'direccion': str(direccion).strip() if direccion else '',
                        'telefono': str(telefono).strip() if telefono else '',
                        'email': str(email).strip() if email else '',
                        'activo': str(estado).lower() in ['activo', 'si', 'sí', 'true', '1', 'yes'] if estado else True
                    }
                    
                    # Crear o actualizar usando nombre como identificador único
                    centro, created = Centro.objects.update_or_create(
                        nombre=nombre_limpio,
                        defaults=datos
                    )
                    
                    if created:
                        creados += 1
                    else:
                        actualizados += 1
                        
                except Exception as e:
                    errores.append(f'Fila {row_idx}: {str(e)}')
            
            return Response({
                'mensaje': 'Importación completada',
                'resumen': {
                    'creados': creados,
                    'actualizados': actualizados,
                    'total_procesados': creados + actualizados,
                    'errores_encontrados': len(errores)
                },
                'errores': errores[:10] if errores else [],  # Máximo 10 errores
                'tiene_mas_errores': len(errores) > 10,
                'exito': len(errores) == 0
            }, status=status.HTTP_200_OK if len(errores) == 0 else status.HTTP_207_MULTI_STATUS)
            
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            logger.exception(f'Error en importación de centros: {str(e)}')
            return Response({
                'error': 'Error al procesar el archivo',
                'mensaje': str(e),
                'sugerencia': 'Verifique que el archivo tenga el formato correcto: Nombre, Dirección, Teléfono, Email, Estado'
            }, status=status.HTTP_400_BAD_REQUEST)
    
    @action(detail=True, methods=['get'])
    def requisiciones(self, request, pk=None):
        """Obtiene todas las requisiciones de un centro"""
        try:
            centro = self.get_object()
            
            if not hasattr(centro, 'requisiciones'):
                return Response({
                    'centro': {
                        'id': centro.id,
                        'clave': centro.clave,
                        'nombre': centro.nombre
                    },
                    'requisiciones': [],
                    'total': 0,
                    'mensaje': 'No hay requisiciones disponibles'
                })
            
            requisiciones = centro.requisiciones_destino.all().order_by('-fecha_solicitud')
            
            # Agrupar por estado
            por_estado = {}
            for req in requisiciones:
                estado = req.estado
                if estado not in por_estado:
                    por_estado[estado] = 0
                por_estado[estado] += 1
            
            requisiciones_data = []
            for req in requisiciones:
                requisiciones_data.append({
                    'id': req.id,
                    'folio': req.folio,
                    'estado': req.estado,
                    'fecha_solicitud': req.fecha_solicitud,
                    'total_items': req.items.count() if hasattr(req, 'items') else 0
                })
            
            return Response({
                'centro': {
                    'id': centro.id,
                    'clave': centro.clave,
                    'nombre': centro.nombre
                },
                'estadisticas': {
                    'total': requisiciones.count(),
                    'por_estado': por_estado
                },
                'requisiciones': requisiciones_data
            })
            
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            return Response({
                'error': 'Error al obtener requisiciones',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['get'], url_path='plantilla')
    def plantilla_centros(self, request):
        """Descarga plantilla de Excel para importación de centros."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Centros'
        
        # Headers que coinciden con el modelo Centro
        headers = ['Nombre', 'Dirección', 'Teléfono', 'Email', 'Estado']
        ws.append(headers)
        
        # Fila de ejemplo
        ws.append(['CENTRO PENITENCIARIO EJEMPLO', 'Av. Principal 123, Ciudad', '(555) 123-4567', 'centro@ejemplo.gob.mx', 'Activo'])
        
        # Aplicar formato a headers
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='9F2241', end_color='9F2241', fill_type='solid')
        for col in range(1, 6):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 45  # Nombre
        ws.column_dimensions['B'].width = 40  # Dirección
        ws.column_dimensions['C'].width = 18  # Teléfono
        ws.column_dimensions['D'].width = 30  # Email
        ws.column_dimensions['E'].width = 12  # Estado
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=Plantilla_Centros.xlsx'
        wb.save(response)
        return response

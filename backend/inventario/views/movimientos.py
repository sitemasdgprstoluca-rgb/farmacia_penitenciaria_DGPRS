# -*- coding: utf-8 -*-
"""
MÃ³dulo MovimientoViewSet para gestiÃ³n de movimientos de inventario.

Contiene el ViewSet para operaciones CRUD sobre movimientos de stock,
incluyendo exportaciÃ³n a PDF/Excel y trazabilidad de productos y lotes.

RefactorizaciÃ³n audit34: ExtraÃ­do del monolÃ­tico views.py (7654 lÃ­neas)
para mejorar mantenibilidad y separaciÃ³n de responsabilidades.
"""
import logging
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from django.db.models import Q
from django.http import HttpResponse
from django.utils import timezone

from rest_framework import viewsets, mixins, status, serializers
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated

from core.models import Producto, Lote, Movimiento, Centro
from core.serializers import MovimientoSerializer
# ISS-MEDICO FIX: Usar IsCentroCanManageInventory para excluir mÃ©dico
from core.permissions import IsFarmaciaRole, IsCentroRole, IsCentroCanManageInventory, RoleHelper

from .base import (
    CustomPagination,
    is_farmacia_or_admin,
    get_user_centro,
    registrar_movimiento_stock,
)

from inventario.services.requisicion_service import CentroPermissionMixin

logger = logging.getLogger(__name__)


class MovimientoViewSet(
    mixins.ListModelMixin,
    mixins.RetrieveModelMixin,
    mixins.CreateModelMixin,
    viewsets.GenericViewSet
):
    """
    ViewSet para gestionar movimientos de inventario.
    
    PERMISOS:
    - Admin/Farmacia: acceso completo a todos los movimientos
    - Centro (administrador_centro, director_centro): puede VER y CREAR movimientos
    - Medico: NO puede crear movimientos (ISS-MEDICO FIX)
    - Vista: solo lectura
    
    FILTROS (alineados con exportación):
    - tipo: entrada/salida/ajuste
    - centro: ID del centro origen (lote.centro)
    - centro_destino: ID del centro destino (para ver transferencias a un centro)
    - producto: ID del producto
    - lote: ID del lote o número de lote
    - subtipo_salida: receta/consumo_interno/merma/transferencia/etc
    - referencia: código de grupo de salida (ej: SAL-0107-1913-8)
    - fecha_inicio: YYYY-MM-DD
    - fecha_fin: YYYY-MM-DD
    - search: búsqueda en motivo, número de lote, producto, referencia
    
    Esto permite auditoría completa de consumos y transferencias en cada centro.
    """
    queryset = Movimiento.objects.select_related('lote__producto', 'centro_origen', 'centro_destino', 'usuario').all()
    serializer_class = MovimientoSerializer
    # ISS-MEDICO FIX: Usar permiso que excluye mÃ©dico de operaciones de escritura
    permission_classes = [IsCentroCanManageInventory]
    pagination_class = CustomPagination
    http_method_names = ['get', 'post', 'head', 'options']

    def _get_producto_display(self, mov):
        """Helper para obtener display de producto de forma segura - SIN TRUNCAR."""
        if not mov.lote or not mov.lote.producto:
            return 'N/A'
        producto = mov.lote.producto
        nombre = producto.nombre or producto.descripcion or ''
        return f"{producto.clave} - {nombre}"

    def get_queryset(self):
        """
        Filtra movimientos segun parametros.
        
        Parametros (alineados con exportacin):
        - tipo: entrada/salida/ajuste
        - centro: ID del centro (solo admin/farmacia/vista)
        - producto: ID del producto
        - lote: ID del lote
        - fecha_inicio: fecha mnima (YYYY-MM-DD)
        - fecha_fin: fecha mxima (YYYY-MM-DD)
        - search: bsqueda en observaciones, lote, producto
        
        Seguridad: Usuarios de centro solo ven movimientos de su centro.
        Admin/farmacia/vista ven todo por defecto, pueden filtrar con ?centro=.
        """
        queryset = Movimiento.objects.select_related('lote__producto', 'centro_origen', 'centro_destino', 'usuario')
        
        # SEGURIDAD: Filtrar por centro segun rol
        user = self.request.user
        
        # Obtener tipo de filtro (para aplicar filtro de centro inteligente)
        tipo_param = self.request.query_params.get('tipo', '').lower()
        
        if not is_farmacia_or_admin(user):
            # Usuario de centro: solo ve movimientos relevantes a su centro
            # ISS-CENTRO FIX v2: Solo mostrar:
            # 1. Movimientos donde el lote pertenece a su centro (entradas/salidas de su inventario)
            # 2. Movimientos donde su centro es el ORIGEN (salidas que hizo su centro)
            # NO mostrar: Salidas de Farmacia Central aunque centro_destino=su centro
            # (para esas, se muestra la ENTRADA correspondiente, no la salida de farmacia)
            user_centro = get_user_centro(user)
            if user_centro:
                queryset = queryset.filter(
                    Q(lote__centro=user_centro) | 
                    Q(centro_origen=user_centro)
                )
            else:
                return Movimiento.objects.none()
        else:
            # Admin/farmacia/vista: pueden filtrar por centro especifico
            centro_param = self.request.query_params.get('centro')
            if centro_param:
                # ISS-FIX: Filtrar centro de forma inteligente según el tipo para evitar duplicados
                if tipo_param == 'salida':
                    # Para salidas: solo mostrar donde el centro es ORIGEN
                    queryset = queryset.filter(
                        Q(lote__centro_id=centro_param) | 
                        Q(centro_origen_id=centro_param)
                    )
                elif tipo_param == 'entrada':
                    # Para entradas: solo mostrar donde el centro es DESTINO
                    queryset = queryset.filter(
                        Q(lote__centro_id=centro_param) | 
                        Q(centro_destino_id=centro_param)
                    )
                else:
                    # Sin tipo: mostrar todos relacionados
                    queryset = queryset.filter(
                        Q(lote__centro_id=centro_param) | 
                        Q(centro_origen_id=centro_param) | 
                        Q(centro_destino_id=centro_param)
                    )
        
        # Filtro por tipo
        tipo = self.request.query_params.get('tipo')
        if tipo:
            queryset = queryset.filter(tipo=tipo.lower())
        
        # Filtro por producto
        producto = self.request.query_params.get('producto')
        if producto:
            queryset = queryset.filter(lote__producto_id=producto)
        
        # Filtro por lote (acepta ID numÃ©rico o nÃºmero de lote como texto)
        lote = self.request.query_params.get('lote')
        if lote:
            if lote.isdigit():
                # Si es un nÃºmero, buscar por ID
                queryset = queryset.filter(lote_id=lote)
            else:
                # Si es texto, buscar por nÃºmero de lote (coincidencia parcial)
                queryset = queryset.filter(lote__numero_lote__icontains=lote)
        
        # Filtro por subtipo de salida (receta, consumo_interno, merma, etc.)
        subtipo_salida = self.request.query_params.get('subtipo_salida')
        if subtipo_salida:
            queryset = queryset.filter(subtipo_salida__iexact=subtipo_salida)
        
        # Filtro por referencia (grupo_salida para transferencias)
        referencia = self.request.query_params.get('referencia')
        if referencia:
            queryset = queryset.filter(referencia__icontains=referencia)
        
        # Filtro por centro_destino (para ver transferencias a un centro específico)
        centro_destino = self.request.query_params.get('centro_destino')
        if centro_destino:
            queryset = queryset.filter(centro_destino_id=centro_destino)
        
        # Filtro por rango de fechas
        fecha_inicio = self.request.query_params.get('fecha_inicio')
        if fecha_inicio:
            queryset = queryset.filter(fecha__date__gte=fecha_inicio)
        
        fecha_fin = self.request.query_params.get('fecha_fin')
        if fecha_fin:
            queryset = queryset.filter(fecha__date__lte=fecha_fin)
        
        # Busqueda en motivo, lote y producto (incluye referencia en búsqueda)
        search = self.request.query_params.get('search')
        if search and search.strip():
            search_term = search.strip()
            queryset = queryset.filter(
                Q(motivo__icontains=search_term) |
                Q(lote__numero_lote__icontains=search_term) |
                Q(lote__producto__clave__icontains=search_term) |
                Q(lote__producto__descripcion__icontains=search_term) |
                Q(numero_expediente__icontains=search_term) |
                Q(referencia__icontains=search_term)
            )
        
        # ISS-FIX: Filtro por estado de confirmación (entrega)
        # - confirmado: salidas con [CONFIRMADO] en el motivo
        # - pendiente: salidas SIN [CONFIRMADO] en el motivo
        # NOTA: Este filtro se omite para la vista agrupada (se aplica post-agrupación)
        # El flag _skip_estado_confirmacion se usa internamente
        skip_estado_filter = getattr(self, '_skip_estado_confirmacion', False)
        estado_confirmacion = self.request.query_params.get('estado_confirmacion')
        if estado_confirmacion and not skip_estado_filter:
            estado_lower = estado_confirmacion.lower()
            if estado_lower == 'confirmado':
                # Confirmados: solo salidas que tienen [CONFIRMADO]
                queryset = queryset.filter(
                    tipo='salida',
                    motivo__icontains='[CONFIRMADO]'
                )
            elif estado_lower == 'pendiente':
                # Pendientes: solo salidas que NO tienen [CONFIRMADO]
                queryset = queryset.filter(tipo='salida').exclude(motivo__icontains='[CONFIRMADO]')
        
        return queryset.order_by('-fecha')

    def perform_create(self, serializer):
        """
        Crea un movimiento validando permisos por centro.
        
        SEGURIDAD:
        - Admin/farmacia: pueden crear cualquier movimiento en cualquier lote
        - Usuario de centro (administrador/director): pueden crear movimientos en lotes de su centro
          y solo ciertos tipos: 'salida' (consumo), 'ajuste' (inventario físico)
        - Médico: SOLO puede crear salidas con subtipo 'receta' (dispensación por receta)
        - Usuario de centro NO puede crear 'entrada' (solo vía surtido de requisición)
        """
        user = self.request.user
        
        lote = serializer.validated_data.get('lote')
        tipo = serializer.validated_data.get('tipo', '').lower()
        subtipo_salida_raw = serializer.validated_data.get('subtipo_salida')
        subtipo_salida = subtipo_salida_raw.lower() if subtipo_salida_raw else ''
        numero_expediente = serializer.validated_data.get('numero_expediente')
        
        # ISS-MEDICO FIX v3: Médicos SOLO pueden crear salidas por receta
        if RoleHelper.is_medico(user):
            if tipo != 'salida' or subtipo_salida != 'receta':
                raise serializers.ValidationError({
                    'detail': 'Como médico, solo puede registrar dispensaciones por receta médica.'
                })
            # Validar que el lote pertenece al centro del médico
            user_centro = get_user_centro(user)
            if lote and lote.centro != user_centro:
                raise serializers.ValidationError({
                    'lote': 'Solo puede dispensar medicamentos de los lotes de su centro.'
                })
        # Validar que usuario de centro (no médico) solo opere con sus lotes
        elif not is_farmacia_or_admin(user):
            user_centro = get_user_centro(user)
            
            # Validar que el lote pertenece al centro del usuario
            if lote and lote.centro != user_centro:
                raise serializers.ValidationError({
                    'lote': 'Solo puedes registrar movimientos en lotes de tu centro'
                })
            
            # Validar tipos de movimiento permitidos para centros
            # Centros pueden: salida (consumo), ajuste (inventario físico)
            # Centros NO pueden: entrada (solo vía surtido automático)
            tipos_permitidos_centro = ['salida', 'ajuste']
            if tipo not in tipos_permitidos_centro:
                raise serializers.ValidationError({
                    'tipo': f'Los centros solo pueden registrar: {", ".join(tipos_permitidos_centro)}. Las entradas se generan automáticamente al surtir requisiciones.'
                })
        
        # ISS-FIX: Para transferencias desde Almacén Central a Centro,
        # el lote es de Almacén Central (centro=None) pero el destino es un Centro específico.
        # Debemos permitir esto para admin/farmacia usando skip_centro_check=True
        
        # ISS-FIX-500: Convertir centro_id a objeto Centro si se pasa un ID
        centro_destino_raw = serializer.validated_data.get('centro')
        centro_destino = None
        if centro_destino_raw:
            if isinstance(centro_destino_raw, Centro):
                centro_destino = centro_destino_raw
            else:
                try:
                    centro_destino = Centro.objects.get(pk=int(centro_destino_raw))
                except (Centro.DoesNotExist, ValueError, TypeError):
                    raise serializers.ValidationError({
                        'centro': f'Centro con ID {centro_destino_raw} no encontrado'
                    })
        
        es_transferencia_farmacia = is_farmacia_or_admin(user) and centro_destino and lote and lote.centro is None
        
        movimiento, _ = registrar_movimiento_stock(
            lote=lote,
            tipo=serializer.validated_data.get('tipo'),
            cantidad=serializer.validated_data.get('cantidad'),
            usuario=user,
            centro=centro_destino or (lote.centro if lote else None),
            requisicion=serializer.validated_data.get('requisicion'),
            # FIX: El serializer mapea 'observaciones' del frontend a 'motivo' via to_internal_value
            observaciones=serializer.validated_data.get('motivo', ''),
            subtipo_salida=subtipo_salida_raw,  # Usar valor original, no lowercase
            numero_expediente=numero_expediente,
            # ISS-FIX: Saltear validación de centro para transferencias de Almacén Central
            skip_centro_check=es_transferencia_farmacia
        )
        # Dejar instancia lista para serializer.data
        serializer.instance = movimiento

    @action(detail=False, methods=['get'], url_path='trazabilidad-pdf')
    def trazabilidad_pdf(self, request):
        """
        Genera PDF de trazabilidad de un producto.
        Parmetros: ?producto_clave=XXX
        
        SEGURIDAD: Filtra por centro del usuario si no es admin/farmacia.
        """
        from core.utils.pdf_reports import generar_reporte_trazabilidad
        
        # SEGURIDAD: Verificar permisos y determinar filtro de centro
        user = request.user
        filtrar_por_centro = not is_farmacia_or_admin(user)
        user_centro = get_user_centro(user) if filtrar_por_centro else None
        
        clave = request.query_params.get('producto_clave')
        if not clave:
            return Response({'error': 'Se requiere producto_clave'}, status=status.HTTP_400_BAD_REQUEST)
        
        producto = Producto.objects.filter(
            Q(clave__iexact=clave) | Q(descripcion__iexact=clave)
        ).first()
        if not producto:
            return Response({'error': 'Producto no encontrado'}, status=status.HTTP_404_NOT_FOUND)
        
        try:
            # Obtener movimientos del producto
            movimientos = Movimiento.objects.filter(
                lote__producto=producto
            ).select_related('lote', 'centro_origen', 'centro_destino', 'usuario')
            
            # Aplicar filtro de centro si corresponde
            if filtrar_por_centro and user_centro:
                movimientos = movimientos.filter(
                    Q(centro_origen=user_centro) | Q(centro_destino=user_centro) | Q(lote__centro=user_centro)
                )
            
            movimientos = movimientos.order_by('-fecha')[:100]
            
            trazabilidad_data = []
            for mov in movimientos:
                # ISS-FIX: Lógica clara para centro según tipo de movimiento
                tipo_upper = mov.tipo.upper()
                if tipo_upper == 'SALIDA':
                    centro_display = mov.centro_destino.nombre if mov.centro_destino else 'Farmacia Central'
                elif tipo_upper == 'ENTRADA':
                    centro_display = mov.centro_origen.nombre if mov.centro_origen else 'Farmacia Central'
                else:
                    centro_display = mov.centro_destino.nombre if mov.centro_destino else (mov.centro_origen.nombre if mov.centro_origen else 'Farmacia Central')
                
                # ISS-FIX: Mostrar username si no hay nombre completo
                if mov.usuario:
                    usuario_display = mov.usuario.get_full_name()
                    if not usuario_display or usuario_display.strip() == '':
                        usuario_display = mov.usuario.username
                else:
                    usuario_display = 'Sistema'
                
                trazabilidad_data.append({
                    'fecha': mov.fecha.strftime('%d/%m/%Y %H:%M') if mov.fecha else 'N/A',
                    'tipo': tipo_upper,
                    'lote': mov.lote.numero_lote if mov.lote else 'N/A',
                    'cantidad': mov.cantidad,
                    'centro': centro_display,
                    'usuario': usuario_display,
                    'observaciones': mov.motivo or ''
                })
            
            # Obtener primer lote para info de contrato y precio
            lote_principal = Lote.objects.filter(producto=producto, activo=True).order_by('-cantidad_actual').first()
            
            producto_info = {
                'clave': producto.clave,
                'descripcion': producto.nombre,  # Usar nombre como descripciÃ³n principal
                'unidad_medida': producto.unidad_medida,
                'stock_actual': producto.get_stock_actual() if hasattr(producto, 'get_stock_actual') else 0,
                'stock_minimo': producto.stock_minimo,
                'precio_unitario': lote_principal.precio_unitario if lote_principal else 0,
                'numero_contrato': lote_principal.numero_contrato if lote_principal else 'N/A',
            }
            
            pdf_buffer = generar_reporte_trazabilidad(trazabilidad_data, producto_info=producto_info)
            
            response = HttpResponse(
                pdf_buffer.getvalue(),
                content_type='application/pdf'
            )
            response['Content-Disposition'] = f'attachment; filename="Trazabilidad_{clave}_{timezone.now().strftime("%Y%m%d")}.pdf"'
            
            return response
            
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            return Response({
                'error': 'Error al generar PDF de trazabilidad',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path='trazabilidad-lote-pdf')
    def trazabilidad_lote_pdf(self, request):
        """
        Genera PDF de trazabilidad de un lote especfico.
        Parmetros: ?numero_lote=XXX
        
        SEGURIDAD: Solo admin/farmacia pueden acceder.
        """
        from core.utils.pdf_reports import generar_reporte_trazabilidad
        
        # SEGURIDAD: Solo admin/farmacia pueden exportar trazabilidad de lotes
        if not is_farmacia_or_admin(request.user):
            return Response({'error': 'Solo administradores y farmacia pueden exportar trazabilidad de lotes'}, status=status.HTTP_403_FORBIDDEN)
        
        numero_lote = request.query_params.get('numero_lote')
        lote_id = request.query_params.get('lote_id')
        
        if not numero_lote and not lote_id:
            return Response({'error': 'Se requiere numero_lote o lote_id'}, status=status.HTTP_400_BAD_REQUEST)
        
        # ISS-FIX: Preferir lote_id para evitar ambigÃ¼edad con lotes duplicados en diferentes centros
        if lote_id:
            lote = Lote.objects.filter(id=lote_id).select_related('producto', 'centro').first()
        else:
            lote = Lote.objects.filter(numero_lote__iexact=numero_lote).select_related('producto', 'centro').first()
        
        if not lote:
            return Response({'error': 'Lote no encontrado'}, status=status.HTTP_404_NOT_FOUND)
        
        try:
            # Obtener movimientos del lote
            movimientos = Movimiento.objects.filter(
                lote=lote
            ).select_related('lote', 'centro_origen', 'centro_destino', 'usuario').order_by('-fecha')[:100]
            
            trazabilidad_data = []
            for mov in movimientos:
                # ISS-FIX: Lógica clara para centro según tipo de movimiento
                tipo_upper = mov.tipo.upper()
                if tipo_upper == 'SALIDA':
                    centro_display = mov.centro_destino.nombre if mov.centro_destino else 'Farmacia Central'
                elif tipo_upper == 'ENTRADA':
                    centro_display = mov.centro_origen.nombre if mov.centro_origen else 'Farmacia Central'
                else:
                    centro_display = mov.centro_destino.nombre if mov.centro_destino else (mov.centro_origen.nombre if mov.centro_origen else 'Farmacia Central')
                
                # ISS-FIX: Mostrar username si no hay nombre completo
                if mov.usuario:
                    usuario_display = mov.usuario.get_full_name()
                    if not usuario_display or usuario_display.strip() == '':
                        usuario_display = mov.usuario.username
                else:
                    usuario_display = 'Sistema'
                
                trazabilidad_data.append({
                    'fecha': mov.fecha.strftime('%d/%m/%Y %H:%M') if mov.fecha else 'N/A',
                    'tipo': tipo_upper,
                    'lote': mov.lote.numero_lote if mov.lote else 'N/A',
                    'cantidad': mov.cantidad,
                    'centro': centro_display,
                    'usuario': usuario_display,
                    'observaciones': mov.motivo or ''
                })
            
            # ISS-FIX: Usar nombre como fallback para descripcion, manejar None correctamente
            descripcion_producto = 'N/A'
            if lote.producto:
                descripcion_producto = lote.producto.nombre or lote.producto.descripcion or 'N/A'
            
            producto_info = {
                'clave': lote.producto.clave if lote.producto else 'N/A',
                'descripcion': descripcion_producto,
                'unidad_medida': lote.producto.unidad_medida if lote.producto else 'N/A',
                'stock_actual': lote.cantidad_actual,
                'stock_minimo': lote.producto.stock_minimo if lote.producto else 0,
                'numero_lote': lote.numero_lote,
                'fecha_caducidad': lote.fecha_caducidad.strftime('%d/%m/%Y') if lote.fecha_caducidad else 'N/A',
                'proveedor': lote.marca or 'No especificado',
                'numero_contrato': lote.numero_contrato if lote.numero_contrato else 'N/A',
                'precio_unitario': float(lote.precio_unitario) if lote.precio_unitario else 0,
            }
            
            pdf_buffer = generar_reporte_trazabilidad(trazabilidad_data, producto_info=producto_info)
            
            response = HttpResponse(
                pdf_buffer.getvalue(),
                content_type='application/pdf'
            )
            # ISS-FIX: Usar numero_lote del objeto lote (no del query param que puede ser None)
            nombre_archivo = lote.numero_lote or f"ID_{lote.id}"
            response['Content-Disposition'] = f'attachment; filename="Trazabilidad_Lote_{nombre_archivo}_{timezone.now().strftime("%Y%m%d")}.pdf"'
            
            return response
            
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            return Response({
                'error': 'Error al generar PDF de trazabilidad del lote',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path='exportar-pdf')
    def exportar_pdf(self, request):
        """
        Genera PDF de movimientos con filtros opcionales.
        Filtros soportados: tipo, fecha_inicio, fecha_fin, producto, centro, lote, subtipo_salida, search
        """
        from core.utils.pdf_reports import generar_reporte_movimientos
        
        try:
            # Aplicar filtros (get_queryset ya aplica filtros base, aquÃ­ se duplican por consistencia explÃ­cita)
            queryset = self.get_queryset()
            
            tipo = request.query_params.get('tipo')
            if tipo:
                queryset = queryset.filter(tipo=tipo.lower())
            
            # FIX: Usar fecha__date para comparar solo la fecha (ignorar hora)
            fecha_inicio = request.query_params.get('fecha_inicio')
            if fecha_inicio:
                queryset = queryset.filter(fecha__date__gte=fecha_inicio)
            
            fecha_fin = request.query_params.get('fecha_fin')
            if fecha_fin:
                queryset = queryset.filter(fecha__date__lte=fecha_fin)
            
            producto = request.query_params.get('producto')
            if producto:
                queryset = queryset.filter(lote__producto_id=producto)
            
            # ISS-FIX: Filtrar centro de forma inteligente según el tipo para evitar duplicados
            centro = request.query_params.get('centro')
            if centro:
                tipo_lower = (tipo or '').lower()
                if tipo_lower == 'salida':
                    queryset = queryset.filter(Q(centro_origen_id=centro) | Q(lote__centro_id=centro))
                elif tipo_lower == 'entrada':
                    queryset = queryset.filter(Q(centro_destino_id=centro) | Q(lote__centro_id=centro))
                else:
                    queryset = queryset.filter(Q(centro_origen_id=centro) | Q(centro_destino_id=centro) | Q(lote__centro_id=centro))
            
            lote = request.query_params.get('lote')
            if lote:
                if lote.isdigit():
                    queryset = queryset.filter(lote_id=lote)
                else:
                    queryset = queryset.filter(lote__numero_lote__icontains=lote)
            
            subtipo_salida = request.query_params.get('subtipo_salida')
            if subtipo_salida:
                queryset = queryset.filter(subtipo_salida__iexact=subtipo_salida)
            
            search = request.query_params.get('search')
            if search:
                queryset = queryset.filter(
                    Q(lote__numero_lote__icontains=search) |
                    Q(lote__producto__nombre__icontains=search) |
                    Q(lote__producto__descripcion__icontains=search) |
                    Q(motivo__icontains=search) |
                    Q(numero_expediente__icontains=search)
                )
            
            movimientos = queryset[:200]  # Limitar para PDF
            
            # Agrupar movimientos por referencia/transacciÃ³n para PDF
            transacciones = {}
            total_entradas = 0
            total_salidas = 0
            
            for mov in movimientos:
                amount = abs(mov.cantidad) if mov.tipo == 'salida' else mov.cantidad
                ref = mov.referencia or f"MOV-{mov.id}"
                
                if ref not in transacciones:
                    transacciones[ref] = {
                        'referencia': ref,
                        'fecha': mov.fecha.strftime('%d/%m/%Y %H:%M') if mov.fecha else 'N/A',
                        'tipo': mov.tipo.upper(),
                        'centro_origen': mov.centro_origen.nombre if mov.centro_origen else 'Almacén Central',
                        'centro_destino': mov.centro_destino.nombre if mov.centro_destino else 'Almacén Central',
                        'total_productos': 0,
                        'total_cantidad': 0,
                        'detalles': []
                    }
                
                transacciones[ref]['detalles'].append({
                    'producto': self._get_producto_display(mov),
                    'lote': mov.lote.numero_lote if mov.lote else 'N/A',
                    'cantidad': amount
                })
                transacciones[ref]['total_productos'] += 1
                transacciones[ref]['total_cantidad'] += amount
                
                if mov.tipo == 'entrada':
                    total_entradas += amount
                else:
                    total_salidas += amount
            
            movimientos_data = list(transacciones.values())
            resumen_data = {
                'total_transacciones': len(movimientos_data),
                'total_movimientos': sum(t['total_productos'] for t in movimientos_data),
                'total_entradas': total_entradas,
                'total_salidas': total_salidas,
                'diferencia': total_entradas - total_salidas
            }
            
            pdf_buffer = generar_reporte_movimientos(movimientos_data, resumen=resumen_data)
            
            response = HttpResponse(
                pdf_buffer.getvalue(),
                content_type='application/pdf'
            )
            response['Content-Disposition'] = f'attachment; filename="Movimientos_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
            
            return response
            
        except Exception as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f'Error generando PDF de movimientos: {str(e)}', exc_info=True)
            return Response({
                'error': 'Error al generar PDF de movimientos',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path='exportar-excel')
    def exportar_excel(self, request):
        """
        Genera Excel de movimientos con filtros opcionales.
        Filtros soportados: tipo, fecha_inicio, fecha_fin, producto, centro, lote, subtipo_salida, search
        """
        try:
            # Aplicar filtros (get_queryset ya aplica filtros base, aquÃ­ se duplican por consistencia explÃ­cita)
            queryset = self.get_queryset()
            
            tipo = request.query_params.get('tipo')
            if tipo:
                queryset = queryset.filter(tipo=tipo.lower())
            
            # FIX: Usar fecha__date para comparar solo la fecha (ignorar hora)
            fecha_inicio = request.query_params.get('fecha_inicio')
            if fecha_inicio:
                queryset = queryset.filter(fecha__date__gte=fecha_inicio)
            
            fecha_fin = request.query_params.get('fecha_fin')
            if fecha_fin:
                queryset = queryset.filter(fecha__date__lte=fecha_fin)
            
            producto = request.query_params.get('producto')
            if producto:
                queryset = queryset.filter(lote__producto_id=producto)
            
            # ISS-FIX: Filtrar centro de forma inteligente según el tipo para evitar duplicados
            centro = request.query_params.get('centro')
            if centro:
                tipo_lower = (tipo or '').lower()
                if tipo_lower == 'salida':
                    queryset = queryset.filter(Q(centro_origen_id=centro) | Q(lote__centro_id=centro))
                elif tipo_lower == 'entrada':
                    queryset = queryset.filter(Q(centro_destino_id=centro) | Q(lote__centro_id=centro))
                else:
                    queryset = queryset.filter(Q(centro_origen_id=centro) | Q(centro_destino_id=centro) | Q(lote__centro_id=centro))
            
            lote = request.query_params.get('lote')
            if lote:
                if lote.isdigit():
                    queryset = queryset.filter(lote_id=lote)
                else:
                    queryset = queryset.filter(lote__numero_lote__icontains=lote)
            
            subtipo_salida = request.query_params.get('subtipo_salida')
            if subtipo_salida:
                queryset = queryset.filter(subtipo_salida__iexact=subtipo_salida)
            
            search = request.query_params.get('search')
            if search:
                queryset = queryset.filter(
                    Q(lote__numero_lote__icontains=search) |
                    Q(lote__producto__nombre__icontains=search) |
                    Q(lote__producto__descripcion__icontains=search) |
                    Q(motivo__icontains=search) |
                    Q(numero_expediente__icontains=search)
                )
            
            movimientos = queryset[:1000]  # Limitar para Excel
            
            # Crear libro de Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Movimientos'
            
            # TÃ­tulo - Con columnas de centro origen y destino
            ws.merge_cells('A1:L1')
            ws['A1'] = 'REPORTE DE MOVIMIENTOS'
            ws['A1'].font = Font(bold=True, size=14, color='632842')
            ws['A1'].alignment = Alignment(horizontal='center')
            
            # Fecha
            ws.merge_cells('A2:L2')
            ws['A2'] = f'Generado el {timezone.now().strftime("%d/%m/%Y %H:%M")}'
            ws['A2'].alignment = Alignment(horizontal='center')
            
            # Encabezados - Con Centro Origen, Centro Destino, Subtipo y Expediente
            headers = ['#', 'Fecha', 'Tipo', 'Subtipo', 'Producto', 'Lote', 'Cantidad', 'Centro Origen', 'Centro Destino', 'No. Expediente', 'Usuario', 'Observaciones']
            ws.append([])
            ws.append(headers)
            
            header_fill = PatternFill(start_color='632842', end_color='632842', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            for cell in ws[4]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Datos - Con centro origen, destino, subtipo_salida y numero_expediente
            for idx, mov in enumerate(movimientos, 1):
                # ISS-FIX: Usar nombre o descripcion, manejar None correctamente - SIN TRUNCAR
                producto_nombre = ''
                if mov.lote and mov.lote.producto:
                    producto_nombre = (mov.lote.producto.nombre or mov.lote.producto.descripcion or '')
                
                # Formatear subtipo de salida para mostrar
                subtipo_display = ''
                if mov.subtipo_salida:
                    subtipos_label = {
                        'receta': 'Receta Médica',
                        'consumo_interno': 'Consumo Interno',
                        'merma': 'Merma',
                        'caducidad': 'Caducidad',
                        'transferencia': 'Transferencia',
                    }
                    subtipo_display = subtipos_label.get(mov.subtipo_salida.lower(), mov.subtipo_salida.title())
                
                ws.append([
                    idx,
                    mov.fecha.strftime('%d/%m/%Y %H:%M') if mov.fecha else 'N/A',
                    mov.tipo.upper(),
                    subtipo_display,
                    producto_nombre or 'N/A',
                    mov.lote.numero_lote if mov.lote else 'N/A',
                    mov.cantidad,
                    mov.centro_origen.nombre if mov.centro_origen else 'Almacén Central',
                    mov.centro_destino.nombre if mov.centro_destino else 'Almacén Central',
                    mov.numero_expediente or '',
                    mov.usuario.get_full_name() or mov.usuario.username if mov.usuario else 'Sistema',
                    (mov.motivo or ''),  # SIN TRUNCAR
                ])
            
            # Ajustar anchos - Columnas más anchas para textos completos
            column_widths = [6, 16, 10, 16, 60, 18, 10, 40, 40, 14, 18, 60]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
            
            # Respuesta
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="Movimientos_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
            wb.save(response)
            return response
            
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            return Response({
                'error': 'Error al generar Excel de movimientos',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['get'], url_path='recibo-salida')
    def recibo_salida(self, request, pk=None):
        """
        Genera PDF de recibo de salida para un movimiento específico.
        
        Parámetros opcionales:
        - finalizado: si es 'true', muestra sello ENTREGADO en lugar de firmas
        
        SEGURIDAD: Usuarios pueden generar recibos de movimientos que les correspondan.
        """
        from core.utils.pdf_reports import generar_recibo_salida_movimiento
        
        try:
            movimiento = self.get_object()
            
            # Verificar que es un movimiento de salida
            if movimiento.tipo != 'salida':
                return Response(
                    {'error': 'Solo se pueden generar recibos para movimientos de salida'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Verificar permisos - admin/farmacia pueden ver todos, otros solo sus centros
            user = request.user
            if not is_farmacia_or_admin(user):
                user_centro = get_user_centro(user)
                if user_centro:
                    # Usuario de centro puede ver si es origen o destino
                    if movimiento.centro_origen != user_centro and movimiento.centro_destino != user_centro:
                        if movimiento.lote and movimiento.lote.centro != user_centro:
                            return Response(
                                {'error': 'No tienes permiso para ver este movimiento'},
                                status=status.HTTP_403_FORBIDDEN
                            )
            
            finalizado = request.query_params.get('finalizado', 'false').lower() == 'true'
            
            # Construir datos del movimiento
            movimiento_data = {
                'folio': movimiento.id,
                'fecha': movimiento.fecha.strftime('%Y-%m-%d %H:%M') if movimiento.fecha else 'N/A',
                'tipo': movimiento.tipo,
                'subtipo_salida': movimiento.subtipo_salida or 'transferencia',
                'centro_origen': {
                    'id': movimiento.centro_origen.id if movimiento.centro_origen else None,
                    'nombre': movimiento.centro_origen.nombre if movimiento.centro_origen else 'Almacén Central'
                },
                'centro_destino': {
                    'id': movimiento.centro_destino.id if movimiento.centro_destino else None,
                    'nombre': movimiento.centro_destino.nombre if movimiento.centro_destino else ''
                },
                'cantidad': abs(movimiento.cantidad),  # ISS-FIX: Usar valor absoluto
                'observaciones': movimiento.motivo or '',
                'producto': movimiento.lote.producto.nombre if movimiento.lote and movimiento.lote.producto else 'N/A',
                'producto_clave': movimiento.lote.producto.clave if movimiento.lote and movimiento.lote.producto else 'N/A',
                'lote': movimiento.lote.numero_lote if movimiento.lote else 'N/A',
                'presentacion': movimiento.lote.producto.presentacion if movimiento.lote and movimiento.lote.producto else 'N/A',
                'usuario': movimiento.usuario.get_full_name() if movimiento.usuario else 'Sistema',
            }
            
            # Generar PDF usando la función específica para movimientos
            pdf_buffer = generar_recibo_salida_movimiento(
                movimiento_data,
                finalizado=finalizado
            )
            
            response = HttpResponse(
                pdf_buffer.getvalue(),
                content_type='application/pdf'
            )
            tipo_doc = 'Comprobante_Entrega' if finalizado else 'Recibo_Salida'
            response['Content-Disposition'] = f'attachment; filename="{tipo_doc}_{movimiento.id}_{timezone.now().strftime("%Y%m%d")}.pdf"'
            
            logger.info(f"Recibo de salida generado para movimiento {movimiento.id} por usuario {user.username}")
            return response
            
        except Movimiento.DoesNotExist:
            return Response(
                {'error': 'Movimiento no encontrado'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            logger.error(f"Error generando recibo de salida: {str(e)}")
            return Response({
                'error': 'Error al generar recibo de salida',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path='agrupados')
    def movimientos_agrupados(self, request):
        """
        Devuelve movimientos agrupados por grupo de salida o requisición.
        
        ISS-FIX: Esta vista agrupa TODOS los movimientos relacionados antes de paginar,
        evitando que un mismo grupo se divida entre páginas.
        
        Query params:
            - page: Página actual (default 1)
            - page_size: Grupos por página (default 15)
            - tipo, centro, fecha_inicio, fecha_fin, search: Filtros estándar
            - estado_confirmacion: 'confirmado' o 'pendiente' - filtra grupos/movimientos
            
        Returns:
            - grupos: Lista de grupos paginados
            - sin_grupo: Movimientos individuales sin grupo
            - total_grupos: Total de grupos
            - total_elementos: Total de grupos + sin_grupo
        """
        import re
        from collections import defaultdict
        
        # ISS-FIX: Para vista agrupada, capturamos el filtro de estado_confirmacion
        # y lo aplicamos DESPUÉS de agrupar (para no romper la agrupación)
        estado_confirmacion = request.query_params.get('estado_confirmacion', '').lower()
        
        # Omitir el filtro de estado_confirmacion en get_queryset
        # para obtener todos los movimientos relacionados y agrupar correctamente
        self._skip_estado_confirmacion = True
        queryset = self.get_queryset()
        self._skip_estado_confirmacion = False
        
        # Obtener TODOS los movimientos (sin paginar) para agrupar correctamente
        # Limitamos a un máximo razonable para evitar problemas de memoria
        MAX_MOVIMIENTOS = 5000
        movimientos = list(queryset.order_by('-fecha')[:MAX_MOVIMIENTOS])
        
        # Agrupar movimientos
        grupos = defaultdict(lambda: {
            'id': None,
            'tipo_grupo': None,
            'items': [],
            'salidas': [],  # ISS-FIX: Lista separada de salidas para el frontend
            'entradas': [],  # ISS-FIX: Lista separada de entradas para el frontend
            'centro_nombre': 'Almacén Central',
            'fecha': None,
            'usuario_nombre': 'Sistema',
            'total_cantidad': 0,
            'cantidad_salidas': 0,
            'cantidad_entradas': 0,
            'num_salidas': 0,
            'num_entradas': 0,
            'confirmado': False,
            'pendiente': False,
        })
        sin_grupo = []
        
        def extraer_grupo(mov):
            """Extrae el ID del grupo de un movimiento."""
            motivo = mov.motivo or ''
            
            # Patrón 1: Salidas masivas [SAL-xxx]
            match_sal = re.search(r'\[(SAL-[^\]]+)\]', motivo)
            if match_sal:
                return match_sal.group(1), 'salida_masiva'
            
            # Patrón 2: Movimientos por requisición
            match_req = re.search(r'(SALIDA|ENTRADA)_POR_REQUISICION\s+(REQ-[\w-]+)', motivo, re.IGNORECASE)
            if match_req:
                return match_req.group(2), 'requisicion'
            
            # Patrón 3: Campo requisicion directo
            if hasattr(mov, 'requisicion_id') and mov.requisicion_id:
                return f'REQ-{mov.requisicion_id}', 'requisicion'
            
            # Patrón 4: ISS-FIX - Agrupar salidas del MISMO CENTRO, MISMO DÍA, MISMA HORA Y MINUTO
            # Esto detecta salidas masivas de centro que no tienen etiqueta [SAL-xxx]
            if mov.tipo == 'salida' and mov.fecha:
                # Obtener centro destino (para salidas a un centro específico)
                centro_id = None
                if mov.centro_destino:
                    centro_id = mov.centro_destino.id
                elif mov.lote and mov.lote.centro:
                    centro_id = mov.lote.centro.id
                
                if centro_id:
                    # Crear ID de grupo basado en: centro + fecha + hora + minuto
                    fecha_str = mov.fecha.strftime('%Y%m%d-%H%M')
                    grupo_auto = f'AUTO-{centro_id}-{fecha_str}'
                    return grupo_auto, 'salida_centro'
            
            return None, None
        
        def es_confirmado(mov):
            return '[CONFIRMADO]' in (mov.motivo or '')
        
        def es_pendiente(mov):
            return '[PENDIENTE]' in (mov.motivo or '')
        
        for mov in movimientos:
            grupo_id, tipo_grupo_detectado = extraer_grupo(mov)
            motivo = mov.motivo or ''
            
            # ISS-FIX: Si se detectó un grupo, procesarlo
            if grupo_id:
                grupo = grupos[grupo_id]
                
                if grupo['id'] is None:
                    # Inicializar grupo
                    grupo['id'] = grupo_id
                    grupo['tipo_grupo'] = tipo_grupo_detectado
                    grupo['fecha'] = mov.fecha.isoformat() if mov.fecha else None
                    if mov.usuario:
                        grupo['usuario_nombre'] = mov.usuario.get_full_name() or mov.usuario.username
                    grupo['confirmado'] = es_confirmado(mov)
                    grupo['pendiente'] = es_pendiente(mov)
                
                # ISS-FIX: Determinar centro del movimiento correctamente
                # SALIDAS: salen de Farmacia Central (Almacén Central)
                # ENTRADAS: entran al centro destino (el centro del lote)
                if mov.tipo == 'entrada':
                    # Entradas: el centro es donde se recibe (centro del lote o centro_destino)
                    item_centro = mov.lote.centro.nombre if mov.lote and mov.lote.centro else (
                        mov.centro_destino.nombre if mov.centro_destino else 'Centro destino'
                    )
                else:
                    # Salidas: salen de Farmacia Central (Almacén Central)
                    item_centro = 'Farmacia Central'
                
                # Agregar item al grupo
                item_data = {
                    'id': mov.id,
                    'tipo': mov.tipo,
                    'cantidad': mov.cantidad,
                    'fecha': mov.fecha.isoformat() if mov.fecha else None,
                    'motivo': mov.motivo,
                    'lote_numero': mov.lote.numero_lote if mov.lote else None,
                    'producto_clave': mov.lote.producto.clave if mov.lote and mov.lote.producto else None,
                    'producto_nombre': mov.lote.producto.nombre if mov.lote and mov.lote.producto else None,
                    'centro_nombre': item_centro,  # ISS-FIX: Agregar centro del item
                }
                grupo['items'].append(item_data)
                
                # ISS-FIX: Agregar a lista de salidas o entradas según el tipo
                if mov.tipo == 'salida':
                    grupo['salidas'].append(item_data)
                elif mov.tipo == 'entrada':
                    grupo['entradas'].append(item_data)
                
                # Contabilizar
                if mov.tipo == 'salida':
                    grupo['cantidad_salidas'] += abs(mov.cantidad or 0)
                    grupo['num_salidas'] += 1
                    # ISS-FIX: Para salidas automáticas (salida_centro), el centro es el destino
                    if tipo_grupo_detectado == 'salida_centro' and mov.centro_destino:
                        grupo['centro_nombre'] = mov.centro_destino.nombre
                    elif tipo_grupo_detectado == 'salida_centro' and mov.lote and mov.lote.centro:
                        grupo['centro_nombre'] = mov.lote.centro.nombre
                elif mov.tipo == 'entrada':
                    grupo['cantidad_entradas'] += abs(mov.cantidad or 0)
                    grupo['num_entradas'] += 1
                    # Centro destino viene de las entradas
                    if mov.centro_destino:
                        grupo['centro_nombre'] = mov.centro_destino.nombre
                    elif mov.lote and mov.lote.centro:
                        grupo['centro_nombre'] = mov.lote.centro.nombre
                
                # Total = salidas (o entradas si no hay salidas)
                grupo['total_cantidad'] = grupo['cantidad_salidas'] or grupo['cantidad_entradas']
                
                # Actualizar fecha con la más reciente
                mov_fecha = mov.fecha
                grupo_fecha = grupo['fecha']
                if mov_fecha and grupo_fecha:
                    if mov_fecha.isoformat() > grupo_fecha:
                        grupo['fecha'] = mov_fecha.isoformat()
            else:
                # Movimiento individual sin grupo
                sin_grupo.append({
                    'id': mov.id,
                    'tipo': mov.tipo,
                    'subtipo_salida': mov.subtipo_salida,
                    'cantidad': mov.cantidad,
                    'fecha': mov.fecha.isoformat() if mov.fecha else None,
                    'motivo': mov.motivo,
                    'confirmado': es_confirmado(mov),
                    'pendiente': es_pendiente(mov),
                    'lote_numero': mov.lote.numero_lote if mov.lote else None,
                    'producto_clave': mov.lote.producto.clave if mov.lote and mov.lote.producto else None,
                    'producto_nombre': mov.lote.producto.nombre if mov.lote and mov.lote.producto else None,
                    'centro_nombre': (
                        mov.centro_destino.nombre if mov.centro_destino 
                        else (mov.lote.centro.nombre if mov.lote and mov.lote.centro else 'Almacén Central')
                    ),
                    'usuario_nombre': (mov.usuario.get_full_name() or mov.usuario.username) if mov.usuario else 'Sistema',
                })
        
        # Convertir grupos a lista y ordenar por fecha
        grupos_list = sorted(
            grupos.values(), 
            key=lambda g: g['fecha'] or '', 
            reverse=True
        )
        
        # ISS-FIX: Aplicar filtro de estado_confirmacion a grupos y movimientos individuales
        # Para grupos: filtrar basándose en si el grupo está confirmado o pendiente
        # Para sin_grupo: filtrar basándose en el estado del movimiento individual (solo salidas)
        if estado_confirmacion == 'confirmado':
            # Filtrar grupos confirmados
            grupos_list = [g for g in grupos_list if g.get('confirmado', False)]
            # Filtrar movimientos individuales: solo salidas confirmadas
            sin_grupo = [m for m in sin_grupo if m.get('tipo') == 'salida' and m.get('confirmado', False)]
        elif estado_confirmacion == 'pendiente':
            # Filtrar grupos pendientes (no confirmados)
            grupos_list = [g for g in grupos_list if not g.get('confirmado', False)]
            # Filtrar movimientos individuales: solo salidas pendientes (no confirmadas)
            sin_grupo = [m for m in sin_grupo if m.get('tipo') == 'salida' and not m.get('confirmado', False)]
        
        # Calcular totales (después de filtrar por estado)
        total_grupos = len(grupos_list)
        total_sin_grupo = len(sin_grupo)
        total_elementos = total_grupos + total_sin_grupo
        
        # Paginar (grupos + sin_grupo juntos)
        page = int(request.query_params.get('page', 1))
        page_size = int(request.query_params.get('page_size', 15))
        
        todos_elementos = (
            [{'tipo_elem': 'grupo', 'data': g} for g in grupos_list] +
            [{'tipo_elem': 'individual', 'data': m} for m in sin_grupo]
        )
        
        start_idx = (page - 1) * page_size
        end_idx = start_idx + page_size
        elementos_paginados = todos_elementos[start_idx:end_idx]
        
        # Separar de nuevo
        grupos_paginados = [e['data'] for e in elementos_paginados if e['tipo_elem'] == 'grupo']
        sin_grupo_paginados = [e['data'] for e in elementos_paginados if e['tipo_elem'] == 'individual']
        
        return Response({
            'grupos': grupos_paginados,
            'sin_grupo': sin_grupo_paginados,
            'total_grupos': total_grupos,
            'total_sin_grupo': total_sin_grupo,
            'total_elementos': total_elementos,
            'page': page,
            'page_size': page_size,
            'total_pages': (total_elementos + page_size - 1) // page_size,
        })

    @action(detail=True, methods=['post'], url_path='confirmar-entrega')
    def confirmar_entrega(self, request, pk=None):
        """
        Confirma la entrega física de un movimiento de salida individual.
        Marca el movimiento como confirmado agregando [CONFIRMADO] al motivo.
        
        NOTA: El stock ya fue descontado al crear el movimiento.
        Esta acción solo marca el movimiento como "confirmado/entregado".
        
        Returns:
            - 200: Entrega confirmada exitosamente
            - 404: Movimiento no encontrado
            - 400: No es movimiento de salida o ya está confirmado
        """
        try:
            movimiento = self.get_object()
            
            # Verificar que es un movimiento de salida
            if movimiento.tipo != 'salida':
                return Response({
                    'error': True,
                    'message': 'Solo se pueden confirmar entregas de movimientos de salida'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Verificar permisos - admin/farmacia o usuario del centro del lote
            user = request.user
            if not is_farmacia_or_admin(user):
                user_centro = get_user_centro(user)
                # Verificar por centro_destino O por centro del lote
                lote_centro = movimiento.lote.centro if movimiento.lote else None
                puede_confirmar = False
                if user_centro:
                    if movimiento.centro_destino and movimiento.centro_destino.id == user_centro.id:
                        puede_confirmar = True
                    elif lote_centro and lote_centro.id == user_centro.id:
                        puede_confirmar = True
                if not puede_confirmar:
                    return Response({
                        'error': True,
                        'message': 'No tienes permiso para confirmar esta entrega'
                    }, status=status.HTTP_403_FORBIDDEN)
            
            # Verificar si ya está confirmado
            motivo_actual = movimiento.motivo or ''
            if '[CONFIRMADO]' in motivo_actual:
                return Response({
                    'error': True,
                    'message': 'Esta entrega ya fue confirmada anteriormente'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Marcar como confirmado usando update directo para evitar validación clean()
            # El stock ya fue descontado al crear el movimiento
            from core.models import Movimiento as MovimientoModel
            MovimientoModel.objects.filter(pk=movimiento.pk).update(
                motivo=f'[CONFIRMADO] {motivo_actual}'.strip()
            )
            
            logger.info(
                f'Entrega de movimiento {movimiento.id} confirmada por {request.user.username}'
            )
            
            return Response({
                'success': True,
                'message': 'Entrega confirmada exitosamente',
                'movimiento_id': movimiento.id
            })
            
        except Movimiento.DoesNotExist:
            return Response({
                'error': True,
                'message': 'Movimiento no encontrado'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f'Error confirmando entrega individual: {str(e)}')
            return Response({
                'error': True,
                'message': f'Error al confirmar entrega: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

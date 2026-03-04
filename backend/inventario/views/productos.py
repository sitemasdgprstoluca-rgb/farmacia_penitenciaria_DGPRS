# -*- coding: utf-8 -*-
"""
ViewSet de Productos.

Gestión completa de productos farmacéuticos incluyendo:
- CRUD de productos
- Importación/exportación Excel
- Auditoría de cambios
- Toggle de estado activo/inactivo
"""
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.db import transaction
from django.db.models import Q, Sum, F, Subquery, OuterRef, Value, CharField
from django.db.models.expressions import RawSQL
from django.db.models.functions import Coalesce, NullIf, Trim, Concat
from django.http import HttpResponse
from django.utils import timezone
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import logging

from core.models import Producto, AuditoriaLogs
from core.serializers import ProductoSerializer
from core.constants import UNIDADES_MEDIDA
from core.permissions import HasProductosPermission, IsFarmaciaRole
from inventario.utils.idempotency import check_idempotency, save_idempotency, _get_key
from core.mixins import ConfirmationRequiredMixin
from .base import (
    CustomPagination,
    is_farmacia_or_admin,
    has_global_read_access,
    get_user_centro,
    validar_archivo_excel,
    cargar_workbook_seguro,
    validar_filas_excel,
)

logger = logging.getLogger(__name__)


class ProductoViewSet(ConfirmationRequiredMixin, viewsets.ModelViewSet):
    """
    ViewSet para gestionar productos farmacéuticos.
    
    ISS-SEC: Requiere confirmación para operaciones de eliminación
    """
    queryset = Producto.objects.all()
    serializer_class = ProductoSerializer
    permission_classes = [IsAuthenticated, HasProductosPermission]
    pagination_class = CustomPagination
    
    # ISS-SEC: Configuración de confirmación
    require_delete_confirmation = True

    def get_permissions(self):
        """
        Permisos personalizados por acción:
        - list, retrieve: IsAuthenticated + HasProductosPermission (validar perm_productos)
        - create, update, destroy, toggle_activo, importar_excel, exportar_excel, auditoria: IsFarmaciaRole
        """
        acciones_farmacia = [
            'create', 'update', 'partial_update', 'destroy', 
            'toggle_activo', 'importar_excel', 'exportar_excel', 'auditoria'
        ]
        if self.action in acciones_farmacia:
            return [IsAuthenticated(), IsFarmaciaRole()]
        # ISS-SEC-FIX: list y retrieve requieren HasProductosPermission
        return [IsAuthenticated(), HasProductosPermission()]

    def get_queryset(self):
        # ── Subqueries para auditoría de creación / modificación ──────────────
        _user_display = Coalesce(
            NullIf(Trim(Concat('usuario__first_name', Value(' '), 'usuario__last_name')), Value('')),
            'usuario__username',
            output_field=CharField(),
        )
        creado_por_sub = Subquery(
            AuditoriaLogs.objects.filter(
                modelo='Producto',
                objeto_id=RawSQL('"productos"."id"::text', [], output_field=CharField()),
                accion='crear',
                usuario__isnull=False,
            ).order_by('timestamp').annotate(display=_user_display).values('display')[:1],
            output_field=CharField(),
        )
        modificado_por_sub = Subquery(
            AuditoriaLogs.objects.filter(
                modelo='Producto',
                objeto_id=RawSQL('"productos"."id"::text', [], output_field=CharField()),
                accion='actualizar',
                usuario__isnull=False,
            ).order_by('-timestamp').annotate(display=_user_display).values('display')[:1],
            output_field=CharField(),
        )
        queryset = Producto.objects.annotate(
            _creado_por_nombre=creado_por_sub,
            _modificado_por_nombre=modificado_por_sub,
        )
        user = self.request.user
        
        # ISS-SEC-FIX: Determinar el centro para filtrar stock
        # Usuarios CENTRO solo ven productos con stock en SU centro (lo que farmacia les ha surtido)
        # Admin/Farmacia/Vista ven stock de farmacia central por defecto
        centro_param = self.request.query_params.get('centro')
        
        # ISS-SEC-FIX: Flag para determinar si el usuario es de centro
        # Usar has_global_read_access() para incluir rol VISTA correctamente
        tiene_acceso_global = has_global_read_access(user)
        es_usuario_centro = not tiene_acceso_global
        
        if es_usuario_centro:
            # Usuario de centro - filtrar stock solo por su centro
            user_centro = get_user_centro(user)
            if user_centro:
                # Anotar stock_calculado basado SOLO en lotes de su centro
                queryset = queryset.annotate(
                    stock_calculado=Coalesce(
                        Sum(
                            'lotes__cantidad_actual',
                            filter=Q(
                                lotes__activo=True,
                                lotes__cantidad_actual__gt=0,
                                lotes__centro=user_centro
                            )
                        ),
                        0
                    )
                )
                # ISS-FIX: Usuarios de centro SOLO ven productos con stock > 0 en su centro
                # Ya no ven todos los 76 productos con stock = 0
                queryset = queryset.filter(stock_calculado__gt=0)
            else:
                # Usuario sin centro asignado - no ve ningún producto
                queryset = queryset.annotate(
                    stock_calculado=Coalesce(Sum('lotes__cantidad_actual', filter=Q(pk__isnull=True)), 0)
                ).filter(stock_calculado__gt=0)  # Filtro imposible = 0 resultados
        else:
            # Admin/Farmacia/Vista - pueden ver stock global o por centro específico
            # ISS-SEC-FIX: Validar que centro=todos solo lo usen roles con acceso global
            if centro_param == 'todos' and not tiene_acceso_global:
                # Usuario sin acceso global intenta ver todos - denegar
                centro_param = None  # Ignorar parámetro inválido
            
            if centro_param:
                if centro_param == 'central':
                    # Solo stock de farmacia central (centro=NULL)
                    queryset = queryset.annotate(
                        stock_calculado=Coalesce(
                            Sum(
                                'lotes__cantidad_actual',
                                filter=Q(
                                    lotes__activo=True,
                                    lotes__cantidad_actual__gt=0,
                                    lotes__centro__isnull=True
                                )
                            ),
                            0
                        )
                    )
                else:
                    # Stock de un centro específico
                    queryset = queryset.annotate(
                        stock_calculado=Coalesce(
                            Sum(
                                'lotes__cantidad_actual',
                                filter=Q(
                                    lotes__activo=True,
                                    lotes__cantidad_actual__gt=0,
                                    lotes__centro_id=centro_param
                                )
                            ),
                            0
                        )
                    )
            else:
                # Por defecto: stock de farmacia central (donde está el inventario principal)
                queryset = queryset.annotate(
                    stock_calculado=Coalesce(
                        Sum(
                            'lotes__cantidad_actual',
                            filter=Q(
                                lotes__activo=True,
                                lotes__cantidad_actual__gt=0,
                                lotes__centro__isnull=True
                            )
                        ),
                        0
                    )
                )
        
        activo = self.request.query_params.get('activo')
        if activo == 'true':
            queryset = queryset.filter(activo=True)
        elif activo == 'false':
            queryset = queryset.filter(activo=False)
        
        # ISS-FIX: Filtro de unidad_medida flexible
        # La BD puede tener valores compuestos ("CAJA CON 20 TABLETAS")
        # El frontend envía valores simples ("CAJA")
        # Buscamos productos cuya unidad_medida INICIE con el valor o lo CONTENGA
        unidad = self.request.query_params.get('unidad_medida')
        if unidad and unidad != '':
            unidad_upper = unidad.upper()
            # Buscar coincidencia exacta, que empiece con, o que contenga la palabra
            queryset = queryset.filter(
                Q(unidad_medida__iexact=unidad_upper) |
                Q(unidad_medida__istartswith=unidad_upper + ' ') |
                Q(unidad_medida__istartswith=unidad_upper + '/') |
                Q(unidad_medida__icontains=' ' + unidad_upper + ' ') |
                Q(unidad_medida__icontains=' ' + unidad_upper) |
                Q(unidad_medida__iendswith=' ' + unidad_upper)
            )
        
        search = self.request.query_params.get('search')
        if search and search.strip():
            queryset = queryset.filter(
                Q(clave__icontains=search) | 
                Q(nombre__icontains=search)
            )

        stock_status = self.request.query_params.get('stock_status')
        if stock_status:
            # ISS-FIX: Usar stock_calculado ya anotado (respeta el centro del usuario)
            # Ya no es necesario crear stock_total_calc separado
            status_val = stock_status.lower()
            if status_val == 'sin_stock':
                queryset = queryset.filter(stock_calculado__lte=0)
            elif status_val == 'critico':
                queryset = queryset.filter(
                    stock_calculado__gt=0,
                    stock_minimo__gt=0,
                    stock_calculado__lt=F('stock_minimo') * 0.5
                )
            elif status_val == 'bajo':
                queryset = queryset.filter(
                    Q(
                        stock_minimo__gt=0,
                        stock_calculado__gte=F('stock_minimo') * 0.5,
                        stock_calculado__lt=F('stock_minimo')
                    ) | Q(
                        stock_minimo__lte=0,
                        stock_calculado__gt=0,
                        stock_calculado__lt=25
                    )
                )
            elif status_val == 'normal':
                queryset = queryset.filter(
                    Q(
                        stock_minimo__gt=0,
                        stock_calculado__gte=F('stock_minimo'),
                        stock_calculado__lte=F('stock_minimo') * 2
                    ) | Q(
                        stock_minimo__lte=0,
                        stock_calculado__gte=25,
                        stock_calculado__lt=100
                    )
                )
            elif status_val == 'alto':
                queryset = queryset.filter(
                    Q(
                        stock_minimo__gt=0,
                        stock_calculado__gt=F('stock_minimo') * 2
                    ) | Q(
                        stock_minimo__lte=0,
                        stock_calculado__gte=100
                    )
                )
        
        return queryset.order_by('-created_at')
    
    def create(self, request, *args, **kwargs):
        """
        Crea un nuevo producto con soporte de variantes por presentación (ISS-PROD-VAR).

        Si ya existe un producto con el mismo código base pero distinta presentación,
        se asigna automáticamente el siguiente sufijo disponible (ej. 663 → 663.2).

        Responde con `variante_info` indicando el código asignado y si es variante.
        """
        from core.utils.producto_variante import obtener_o_crear_variante
        # Idempotencia transversal
        idem_hit, idem_response = check_idempotency(request, 'productos')
        if idem_hit:
            return idem_response

        try:
            data = request.data

            # Validar todos los campos usando el serializer, pero quitar
            # el UniqueValidator de 'clave' porque nosotros manejamos la unicidad
            serializer = self.get_serializer(data=data)
            # Eliminar UniqueValidator del campo clave antes de is_valid()
            if 'clave' in serializer.fields:
                clave_field = serializer.fields['clave']
                clave_field.validators = [
                    v for v in clave_field.validators
                    if type(v).__name__ != 'UniqueValidator'
                ]
            serializer.is_valid(raise_exception=True)

            validated = serializer.validated_data
            clave_input = str(data.get('clave', '')).strip().upper()[:50]
            nombre = str(validated.get('nombre', '')).strip()
            presentacion = str(validated.get('presentacion', '') or '').strip()

            # Campos que van en defaults (todo menos clave, nombre y presentacion)
            campos_excluidos = {'clave', 'nombre', 'presentacion'}
            defaults_prod = {
                k: v for k, v in validated.items()
                if k not in campos_excluidos
            }

            with transaction.atomic():
                producto, created, var_info = obtener_o_crear_variante(
                    clave_input=clave_input,
                    nombre=nombre,
                    presentacion=presentacion,
                    defaults=defaults_prod,
                    usuario=request.user,
                )

            response_data = dict(self.get_serializer(producto).data)
            response_data['variante_info'] = var_info

            http_status = status.HTTP_201_CREATED if created else status.HTTP_200_OK
            headers = self.get_success_headers(response_data)
            # Idempotencia: solo para creates (201)
            if created:
                _idem_key = _get_key(request)
                if _idem_key:
                    save_idempotency(request, 'productos', _idem_key, response_data, 201)
            return Response(response_data, status=http_status, headers=headers)

        except ValueError as e:
            # Conflicto de presentación (usuario envió clave sufijada incorrecta)
            logger.warning(f"ISS-PROD-VAR: conflicto de variante: {e}")
            return Response(
                {'error': str(e)},
                status=status.HTTP_409_CONFLICT
            )
        except Exception as e:
            logger.error(f"Error al crear producto: {str(e)}", exc_info=True)
            return Response(
                {'error': 'Error al crear producto. Verifique los datos ingresados.'},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    def update(self, request, *args, **kwargs):
        """
        Actualiza un producto existente.
        
        Validaciones:
        - Clave única (si se modifica)
        - Datos válidos
        """
        try:
            partial = kwargs.pop('partial', False)
            instance = self.get_object()
            serializer = self.get_serializer(instance, data=request.data, partial=partial)
            serializer.is_valid(raise_exception=True)
            self.perform_update(serializer)
            
            return Response(serializer.data)
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            logger.error(f"Error al actualizar producto: {str(e)}", exc_info=True)
            return Response({'error': 'Error al actualizar producto. Verifique los datos ingresados.'}, status=status.HTTP_400_BAD_REQUEST)
    
    def destroy(self, request, *args, **kwargs):
        """
        Elimina un producto.
        
        Validaciones:
        - No puede eliminarse si tiene lotes asociados
        - Confirmación de eliminación
        - AUDIT-OWN: Solo el creador o admin puede eliminar
        """
        instance = self.get_object()
        user = request.user

        # AUDIT-OWN: Solo el creador o admin puede eliminar
        rol = (getattr(user, 'rol', '') or '').lower()
        ROLES_ADMIN = {'admin', 'admin_sistema', 'superusuario', 'administrador'}
        if not user.is_superuser and rol not in ROLES_ADMIN:
            log_crear = AuditoriaLogs.objects.filter(
                modelo='Producto', objeto_id=str(instance.pk), accion='crear'
            ).order_by('timestamp').first()
            if log_crear and log_crear.usuario and log_crear.usuario.pk != user.pk:
                logger.warning(
                    f"AUDIT-OWN: {user.username} intentó eliminar producto #{instance.pk} "
                    f"(creado por {log_crear.usuario.username}) - DENEGADO"
                )
                return Response(
                    {'error': 'Solo el usuario que registró este producto o un administrador puede eliminarlo',
                     'creado_por': log_crear.usuario.get_full_name() or log_crear.usuario.username},
                    status=status.HTTP_403_FORBIDDEN
                )

        try:
            if instance.lotes.exists():
                return Response({
                    'error': 'No se puede eliminar el producto',
                    'razon': 'Tiene lotes asociados',
                    'sugerencia': 'Elimine primero los lotes o marque el producto como inactivo'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            instance.delete()
            return Response({'mensaje': 'Producto eliminado exitosamente'}, status=status.HTTP_204_NO_CONTENT)
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            logger.error(f"Error al eliminar producto: {str(e)}", exc_info=True)
            return Response({'error': 'Error al eliminar producto'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=True, methods=['post'], url_path='toggle-activo')
    def toggle_activo(self, request, pk=None):
        """
        Activa o desactiva un producto.
        POST /api/productos/{id}/toggle-activo/
        
        Reglas:
        - No se puede desactivar un producto con stock disponible > 0
        - Usa update() directo para evitar validación de otros campos
        """
        try:
            producto = self.get_object()
            nuevo_estado = not producto.activo
            
            # Si se va a desactivar, verificar que no tenga stock disponible
            if not nuevo_estado:  # Desactivando
                stock_disponible = producto.lotes.filter(
                    activo=True,
                    cantidad_actual__gt=0
                ).aggregate(total=Sum('cantidad_actual'))['total'] or 0
                
                if stock_disponible > 0:
                    return Response({
                        'error': 'No se puede desactivar el producto',
                        'razon': f'Tiene {stock_disponible} unidades en stock disponible',
                        'sugerencia': 'Transfiera o agote el inventario antes de desactivar'
                    }, status=status.HTTP_400_BAD_REQUEST)
            
            # Usar update() directo para evitar validación de otros campos
            Producto.objects.filter(pk=producto.pk).update(activo=nuevo_estado)
            
            estado = 'activado' if nuevo_estado else 'desactivado'
            return Response({
                'mensaje': f'Producto {estado} exitosamente',
                'activo': nuevo_estado,
                'id': producto.id
            }, status=status.HTTP_200_OK)
        except Producto.DoesNotExist:
            return Response({'error': 'Producto no encontrado'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"Error en toggle_activo: {str(e)}", exc_info=True)
            return Response({'error': f'Error interno: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=True, methods=['get'], url_path='lotes')
    def lotes(self, request, pk=None):
        """
        Obtiene los lotes de un producto.
        GET /api/productos/{id}/lotes/
        
        TRAZABILIDAD: Para admin/farmacia sin filtro de centro, devuelve lotes CONSOLIDADOS
        (un lote físico = un registro, con cantidades sumadas de todos los centros).
        
        Parámetros opcionales:
        - centro: ID del centro o 'central' para farmacia central (desactiva consolidación)
        - activo: 'true' para solo lotes activos
        - con_stock: 'true' para solo lotes con cantidad > 0
        - consolidar: 'false' para desactivar consolidación (admin/farmacia)
        """
        try:
            from core.models import Lote
            from core.serializers import LoteSerializer
            from collections import defaultdict
            from datetime import date
            
            producto = self.get_object()
            user = request.user
            
            # Filtro base: lotes del producto
            lotes_qs = Lote.objects.select_related('centro').filter(producto=producto)
            
            # Determinar si consolidar (usado en Trazabilidad)
            centro_param = request.query_params.get('centro')
            consolidar_param = request.query_params.get('consolidar', 'true').lower()
            
            # ISS-SEC-FIX: Usar has_global_read_access para validar acceso global
            # Esto incluye admin, farmacia, superuser y rol VISTA
            tiene_acceso_global = has_global_read_access(user)
            
            # ISS-FIX: Por defecto para admin/farmacia, solo mostrar Almacén Central
            # Para ver todos los centros (Trazabilidad), pasar ?centro=todos o ?consolidar=true
            todos_centros = centro_param == 'todos'
            
            # ISS-SEC-FIX: VALIDACIÓN CRÍTICA - Bloquear centro=todos para usuarios sin acceso global
            # Esto previene IDOR/BOLA donde usuarios de centro ven lotes de otros centros
            if todos_centros and not tiene_acceso_global:
                logger.warning(f"ISS-SEC: Usuario {user.username} (rol={getattr(user, 'rol', 'N/A')}) intentó acceder a lotes de todos los centros - DENEGADO")
                # Forzar filtro por centro del usuario
                user_centro = get_user_centro(user)
                if user_centro:
                    lotes_qs = lotes_qs.filter(centro=user_centro)
                else:
                    return Response({'error': 'No tiene acceso a lotes de otros centros'}, status=status.HTTP_403_FORBIDDEN)
                centro_param = None  # Resetear para evitar bypass
                todos_centros = False
            
            # TRAZABILIDAD: Consolidar solo cuando se pide explícitamente todos los centros
            debe_consolidar = tiene_acceso_global and todos_centros and consolidar_param != 'false'
            
            # Filtro por centro
            if centro_param:
                if centro_param == 'central':
                    lotes_qs = lotes_qs.filter(centro__isnull=True)
                elif centro_param == 'todos' and tiene_acceso_global:
                    # ISS-SEC-FIX: Solo con acceso global se permite ver todos
                    pass  # No filtrar
                elif centro_param != 'todos':
                    # Centro específico por ID
                    # ISS-SEC-FIX: Usuarios sin acceso global solo pueden ver su propio centro
                    if not tiene_acceso_global:
                        user_centro = get_user_centro(user)
                        if user_centro and str(user_centro.id) == str(centro_param):
                            lotes_qs = lotes_qs.filter(centro_id=centro_param)
                        else:
                            # Intentando acceder a centro ajeno - forzar su centro
                            logger.warning(f"ISS-SEC: Usuario {user.username} intentó acceder a centro {centro_param} - DENEGADO")
                            if user_centro:
                                lotes_qs = lotes_qs.filter(centro=user_centro)
                            else:
                                return Response({'error': 'No tiene acceso a lotes de otros centros'}, status=status.HTTP_403_FORBIDDEN)
                    else:
                        lotes_qs = lotes_qs.filter(centro_id=centro_param)
            elif tiene_acceso_global:
                # ISS-SEC-FIX: Admin/farmacia/vista por defecto solo ve Almacén Central
                # Los lotes de otros centros se ven en Trazabilidad con ?centro=todos
                lotes_qs = lotes_qs.filter(centro__isnull=True)
            else:
                # Usuarios de centro solo ven lotes de su centro
                user_centro = get_user_centro(user)
                if user_centro:
                    lotes_qs = lotes_qs.filter(centro=user_centro)
                else:
                    lotes_qs = Lote.objects.none()
            
            # Filtro por activo
            activo_param = request.query_params.get('activo')
            if activo_param == 'true':
                lotes_qs = lotes_qs.filter(activo=True)
            elif activo_param == 'false':
                lotes_qs = lotes_qs.filter(activo=False)
            
            # Filtro por stock
            con_stock = request.query_params.get('con_stock')
            if con_stock == 'true':
                lotes_qs = lotes_qs.filter(cantidad_actual__gt=0)
            
            # Ordenar por fecha de caducidad
            lotes_qs = lotes_qs.order_by('fecha_caducidad')
            
            if debe_consolidar:
                # TRAZABILIDAD: Consolidar lotes por numero_lote
                hoy = date.today()
                lotes_consolidados = defaultdict(lambda: {
                    'id': None,
                    'producto': producto.id,
                    'producto_nombre': producto.nombre,
                    'producto_clave': producto.clave,
                    'producto_descripcion': producto.nombre,
                    'producto_info': {
                        'presentacion': producto.presentacion or '',
                        'unidad_medida': producto.unidad_medida or 'PIEZA',
                    },
                    'numero_lote': '',
                    'fecha_caducidad': None,
                    'fecha_fabricacion': None,
                    'cantidad_inicial': 0,
                    'cantidad_actual': 0,
                    'precio_unitario': '0.00',
                    'precio_compra': '0.00',
                    'numero_contrato': None,
                    'marca': '-',
                    'ubicacion': '',
                    'activo': True,
                    'estado': 'disponible',
                    'dias_para_caducar': 999,
                    'alerta_caducidad': 'normal',
                    'porcentaje_consumido': 0,
                    'documentos': [],
                    'tiene_documentos': False,
                    'centros': [],
                    'centros_detalle': [],
                    'centro': None,
                    'centro_nombre': '',
                    'created_at': None,
                    'updated_at': None,
                })
                
                for lote in lotes_qs:
                    key = lote.numero_lote
                    cons = lotes_consolidados[key]
                    
                    if cons['id'] is None:
                        cons['id'] = lote.id
                        cons['numero_lote'] = lote.numero_lote
                        cons['fecha_caducidad'] = lote.fecha_caducidad.isoformat() if lote.fecha_caducidad else None
                        cons['fecha_fabricacion'] = lote.fecha_fabricacion.isoformat() if lote.fecha_fabricacion else None
                        cons['precio_unitario'] = str(lote.precio_unitario or 0)
                        cons['precio_compra'] = str(lote.precio_unitario or 0)
                        cons['numero_contrato'] = lote.numero_contrato
                        cons['marca'] = lote.marca or '-'
                        cons['created_at'] = lote.created_at.isoformat() if lote.created_at else None
                        cons['updated_at'] = lote.updated_at.isoformat() if lote.updated_at else None
                        
                        # Calcular días para caducar y alerta
                        if lote.fecha_caducidad:
                            dias = (lote.fecha_caducidad - hoy).days
                            cons['dias_para_caducar'] = dias
                            if dias < 0:
                                cons['alerta_caducidad'] = 'vencido'
                                cons['estado'] = 'vencido'
                            elif dias < 90:
                                cons['alerta_caducidad'] = 'critico'
                            elif dias < 180:
                                cons['alerta_caducidad'] = 'proximo'
                    
                    cons['cantidad_inicial'] += lote.cantidad_inicial
                    cons['cantidad_actual'] += lote.cantidad_actual
                    
                    # Registrar centro donde está
                    centro_nombre = lote.centro.nombre if lote.centro else 'Almacén Central'
                    if centro_nombre not in cons['centros']:
                        cons['centros'].append(centro_nombre)
                    
                    cons['centros_detalle'].append({
                        'centro_id': lote.centro_id,
                        'centro_nombre': centro_nombre,
                        'cantidad': lote.cantidad_actual,
                        'ubicacion': lote.ubicacion or '-',
                    })
                
                # Construir lista final
                lotes_data = []
                for key, cons in lotes_consolidados.items():
                    # Calcular porcentaje consumido
                    if cons['cantidad_inicial'] > 0:
                        cons['porcentaje_consumido'] = round(
                            (1 - cons['cantidad_actual'] / cons['cantidad_inicial']) * 100
                        )
                    
                    # Ubicación muestra centros
                    cons['ubicacion'] = ', '.join(cons['centros'][:2])
                    if len(cons['centros']) > 2:
                        cons['ubicacion'] += f' (+{len(cons["centros"]) - 2})'
                    
                    cons['centro_nombre'] = cons['ubicacion']
                    lotes_data.append(cons)
                
                # Ordenar por fecha de caducidad
                lotes_data.sort(key=lambda x: x['fecha_caducidad'] or '9999-12-31')
                
                total_lotes = len(lotes_data)
                total_stock = sum(l['cantidad_actual'] for l in lotes_data)
            else:
                # Sin consolidación: serializar normalmente
                serializer = LoteSerializer(lotes_qs, many=True)
                lotes_data = serializer.data
                total_lotes = len(lotes_data)
                total_stock = sum(l['cantidad_actual'] for l in lotes_data)
            
            return Response({
                'producto_id': producto.id,
                'producto_nombre': producto.nombre,
                'producto_clave': producto.clave,
                'producto': {
                    'id': producto.id,
                    'clave': producto.clave,
                    'nombre': producto.nombre,
                    'unidad_medida': producto.unidad_medida,
                },
                'lotes': lotes_data,
                'total': total_lotes,
                'total_lotes': total_lotes,
                'total_stock': total_stock,
                'consolidado': debe_consolidar,  # Indica si los lotes están consolidados
            })
            
        except Producto.DoesNotExist:
            return Response({'error': 'Producto no encontrado'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"Error al obtener lotes del producto: {str(e)}", exc_info=True)
            return Response({'error': 'Error al obtener lotes'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=True, methods=['get'], url_path='auditoria')
    def auditoria(self, request, pk=None):
        """
        Obtiene el historial de cambios de un producto.
        GET /api/productos/{id}/auditoria/
        """
        try:
            from core.models import AuditoriaLogs
            producto = self.get_object()

            # Buscar logs de auditoría relacionados con este producto
            logs = AuditoriaLogs.objects.filter(
                Q(modelo='producto') | Q(modelo='core_producto') | Q(modelo='Producto'),
                objeto_id=str(producto.id)
            ).select_related('usuario').order_by('-timestamp')[:50]

            ACCIONES_DISPLAY = {
                'crear':      'Registro inicial',
                'actualizar': 'Actualización de datos',
                'eliminar':   'Eliminación',
                'importar':   'Importación masiva',
                'activar':    'Activación',
                'desactivar': 'Desactivación',
                'restaurar':  'Restauración',
            }
            ACCIONES_ICONO = {
                'crear':      'crear',
                'actualizar': 'actualizar',
                'eliminar':   'eliminar',
                'importar':   'importar',
                'activar':    'activar',
                'desactivar': 'desactivar',
            }

            historial = []
            for log in logs:
                historial.append({
                    'id': log.id,
                    'fecha': log.timestamp.isoformat() if log.timestamp else None,
                    'usuario_nombre': (
                        log.usuario.get_full_name() or log.usuario.username
                        if log.usuario else 'Sistema'
                    ),
                    'accion': log.accion,
                    'accion_display': ACCIONES_DISPLAY.get(
                        (log.accion or '').lower(),
                        (log.accion or '').replace('_', ' ').title()
                    ),
                    'accion_tipo': ACCIONES_ICONO.get((log.accion or '').lower(), 'otro'),
                    'datos_anteriores': log.datos_anteriores,
                    'datos_nuevos': log.datos_nuevos,
                    'ip': log.ip_address if log.ip_address else None,
                })
            
            return Response({
                'producto': {
                    'id': producto.id,
                    'clave': producto.clave,
                    'nombre': producto.nombre,
                },
                'historial': historial,
                'total': len(historial)
            })
        except Producto.DoesNotExist:
            return Response({'error': 'Producto no encontrado'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"Error en auditoria: {str(e)}", exc_info=True)
            return Response({'error': 'Error al obtener auditoría'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['get'], url_path='exportar-excel')
    def exportar_excel(self, request):
        """
        Exporta productos aplicando filtros de la interfaz.
        
        Columnas (basadas en formulario de Productos):
        - Clave, Nombre, Nombre Comercial, Unidad Medida, Stock Minimo, Categoria, Presentacion
        - Descripcion, Sustancia Activa, Concentracion, Via Admin, Requiere Receta, Controlado
        """
        try:
            # Reutilizar el metodo get_queryset que ya aplica TODOS los filtros
            productos = self.get_queryset()
            
            # Crear libro de Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Productos'
            
            # Encabezados exactos según formulario solicitado
            headers = [
                'Clave', 'Nombre', 'Nombre Comercial', 
                'Unidad de Medida', 'Stock Mínimo', 'Categoría', 'Presentación',
                'Descripción Adicional', 'Sustancia Activa', 'Concentración', 'Vía de Administración',
                'Requiere Receta', 'Es Controlado'
            ]
            ws.append(headers)
            
            # Estilo de encabezados
            header_fill = PatternFill(start_color='632842', end_color='632842', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=12)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Datos de productos
            for idx, producto in enumerate(productos, start=1):
                ws.append([
                    producto.clave or '',
                    producto.nombre,
                    producto.nombre_comercial or '',
                    producto.unidad_medida,
                    producto.stock_minimo,
                    producto.categoria or '',
                    producto.presentacion or '',
                    producto.descripcion or '',
                    producto.sustancia_activa or '',
                    producto.concentracion or '',
                    producto.via_administracion or '',
                    'Sí' if producto.requiere_receta else 'No',
                    'Sí' if producto.es_controlado else 'No'
                ])
                
            # Ajustar anchos de columna
            # Clave(15), Nombre(40), NomCom(30), Unidad(15), Min(10), Cat(15), Pres(20), Desc(30), Sust(20), Conc(15), Via(15), Receta(12), Control(12)
            column_widths = [15, 40, 30, 15, 12, 18, 25, 30, 20, 15, 20, 15, 15]
            for i, width in enumerate(column_widths, start=1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
                
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename=Productos_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            wb.save(response)
            return response

        except Exception as e:
            logger.error(f"Error al exportar productos: {str(e)}")
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            ws.column_dimensions['D'].width = 25   # Nombre Comercial
            ws.column_dimensions['E'].width = 18   # Categoria
            ws.column_dimensions['F'].width = 18   # Unidad Medida
            ws.column_dimensions['G'].width = 12   # Stock Minimo
            ws.column_dimensions['H'].width = 12   # Stock Actual
            ws.column_dimensions['I'].width = 20   # Sustancia Activa
            ws.column_dimensions['J'].width = 18   # Presentacion
            ws.column_dimensions['K'].width = 14   # Concentracion
            ws.column_dimensions['L'].width = 14   # Via Admin
            ws.column_dimensions['M'].width = 14   # Requiere Receta
            ws.column_dimensions['N'].width = 12   # Controlado
            ws.column_dimensions['O'].width = 12   # Lotes Activos
            ws.column_dimensions['P'].width = 10   # Estado
            
            # Generar respuesta
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename=Productos_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            wb.save(response)
            
            return response
            
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            return Response({
                'error': 'Error al exportar productos',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['post'], url_path='importar-excel')
    def importar_excel(self, request):
        """
        Importa productos desde archivo Excel usando el importador actualizado.
        
        Usa core.utils.excel_importer que maneja el esquema real de la BD.
        Formato esperado alineado con la plantilla descargable.
        
        Límites de seguridad:
        - Tamaño máximo: 10MB
        - Filas máximas: 10,000
        - Extensiones: .xlsx
        """
        from core.utils.excel_importer import importar_productos_desde_excel, crear_log_importacion
        from django.core.cache import cache
        from core.models import Centro
        
        file = request.FILES.get('file')
        
        # Validar archivo
        es_valido, error_msg = validar_archivo_excel(file)
        if not es_valido:
            return Response({
                'error': 'Archivo inválido',
                'mensaje': error_msg
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            # Ejecutar importación
            resultado = importar_productos_desde_excel(file, request.user)
            
            # ISS-FIX: Invalidar caché del dashboard después de importar productos
            cache.delete('dashboard_resumen_global')
            cache.delete('dashboard_graficas_global')
            for centro in Centro.objects.all():
                cache.delete(f'dashboard_resumen_{centro.id}')
                cache.delete(f'dashboard_graficas_{centro.id}')
            
            # Crear log de importación
            crear_log_importacion(
                usuario=request.user,
                tipo='Producto',
                archivo_nombre=file.name,
                resultado_dict=resultado
            )
            
            # Determinar status code
            if resultado['exitosa']:
                status_code = status.HTTP_200_OK
            elif resultado['registros_exitosos'] > 0:
                status_code = status.HTTP_206_PARTIAL_CONTENT
            else:
                status_code = status.HTTP_400_BAD_REQUEST
            
            return Response(resultado, status=status_code)
            
        except Exception as e:
            logger.exception(f"Error en importación de productos: {e}")
            return Response({
                'error': 'Error al procesar archivo',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path='plantilla')
    def plantilla_productos(self, request):
        """
        Descarga plantilla Excel actualizada para importación de productos.
        
        Usa el generador estandarizado con el esquema real de la base de datos.
        """
        # HALLAZGO #5: Manejo robusto de errores en generación de plantilla
        try:
            from core.utils.excel_templates import generar_plantilla_productos
            return generar_plantilla_productos()
        except ImportError as exc:
            logger.error(f'Error al importar generador de plantilla: {exc}')
            return Response(
                {'error': 'Módulo de generación de plantillas no disponible'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        except Exception as exc:
            logger.exception(f'Error al generar plantilla de productos: {exc}')
            return Response(
                {'error': 'No se pudo generar la plantilla', 'mensaje': str(exc)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )



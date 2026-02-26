# -*- coding: utf-8 -*-
"""
Módulo de ViewSet para Lotes.

Contiene LoteViewSet para gestionar el CRUD completo de lotes farmacéuticos,
incluyendo:
- Listado con filtros por producto, caducidad, stock y centro
- Exportación a PDF y Excel
- Importación masiva desde Excel
- Control de vencimientos y alertas
- Gestión de documentos asociados (facturas, contratos, remisiones)
- Trazabilidad y historial de movimientos

Refactorización audit34: Extraído del monolítico views.py (7654 líneas)
para mejorar mantenibilidad y organización del código.
"""
from datetime import datetime, date, timedelta
import logging
import uuid
import os
from io import BytesIO

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from rest_framework import viewsets, serializers, status
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.core.exceptions import ValidationError as DjangoValidationError
from django.db import transaction, models
from django.db.models import Q, Sum, Subquery, OuterRef, Value, CharField, IntegerField
from django.db.models.expressions import RawSQL
from django.db.models.functions import Coalesce, NullIf, Trim, Concat, Cast
from django.contrib.auth import get_user_model
from django.http import HttpResponse
from django.utils import timezone

# Imports desde el módulo base
from .base import (
    # Clases base y utilidades
    CustomPagination,
    # Helpers de seguridad
    is_farmacia_or_admin,
    has_global_read_access,  # ISS-FIX: Para validar lectura global (incluye vista)
    get_user_centro,
    # Validadores de archivos
    validar_archivo_excel,
    cargar_workbook_seguro,
    validar_filas_excel,
    validar_archivo_pdf,
    # Helper de movimientos
    registrar_movimiento_stock,
    # Constantes
    logger,
)

# Modelos
from core.models import Producto, Lote, Movimiento, Centro, LoteDocumento, LoteParcialidad, AuditoriaLogs

# Serializers
from core.serializers import LoteSerializer, LoteDocumentoSerializer, LoteParcialidadSerializer

# Permisos
from core.permissions import IsFarmaciaAdminOrReadOnly, IsFarmaciaRole, IsCentroRole

# ISS-SEC: Mixin de confirmación
from core.mixins import ConfirmationRequiredMixin


class LoteViewSet(ConfirmationRequiredMixin, viewsets.ModelViewSet):
    """
    ViewSet para gestionar lotes.
    
    Funcionalidades:
    - CRUD completo
    - Filtrado por producto
    - Filtrado por estado de caducidad
    - Busqueda por numero de lote
    - Validaciones de integridad
    
    ISS-SEC: Requiere confirmación para operaciones destructivas y de escritura.
    
    Reglas de negocio blindadas:
    - cantidad_inicial: Solo en CREATE, NUNCA editable vía UPDATE
    - cantidad_actual: SOLO se modifica vía Movimiento (read_only en API)
    - cantidad_contrato: Solo editable por Farmacia/Admin, con auditoría
    - Doble confirmación obligatoria para crear y actualizar lotes
    """
    queryset = Lote.objects.select_related('producto').all()
    serializer_class = LoteSerializer
    permission_classes = [IsFarmaciaAdminOrReadOnly]
    pagination_class = CustomPagination
    
    # ISS-SEC: Configuración de confirmación obligatoria
    require_delete_confirmation = True
    require_update_confirmation = True  # Doble confirmación para edición

    def perform_create(self, serializer):
        """Guarda el usuario autenticado en created_by_id vía SQL directo."""
        instance = serializer.save()
        user = self.request.user
        if user and user.is_authenticated:
            try:
                from django.db import connection
                with connection.cursor() as cursor:
                    cursor.execute(
                        "UPDATE lotes SET created_by_id = %s WHERE id = %s AND created_by_id IS NULL",
                        [user.pk, instance.pk]
                    )
            except Exception:
                pass

    def get_queryset(self):
        """
        Filtra lotes segun parametros.
        
        Parametros:
        - producto: ID del producto
        - activo: true/false
        - caducidad: vencido/critico/proximo/normal
        - search: busqueda por numero de lote
        - centro: ID del centro o 'central' para farmacia (solo admin/farmacia/vista)
        
        Seguridad: Usuarios de centro solo ven lotes de su centro.
        Admin/farmacia/vista ven todo por defecto, pueden filtrar con ?centro=.
        """
        User = get_user_model()

        # ── Subquery: nombre del usuario que creó el lote vía created_by_id ──
        # `created_by_id` existe en la DB pero no está declarado en el modelo Django
        # (managed=False), por eso usamos RawSQL para referenciar la columna.
        _user_display = Coalesce(
            NullIf(Trim(Concat('first_name', Value(' '), 'last_name')), Value('')),
            'username',
            output_field=CharField(),
        )
        # FIX: Referencia directa a la columna sin comillas para evitar problemas de escape
        creado_por_created_by = Subquery(
            User.objects.filter(
                pk=Cast(RawSQL('lotes.created_by_id', []), output_field=IntegerField())
            ).annotate(display=_user_display).values('display')[:1],
            output_field=CharField(),
        )
        # Fallback: primer registro en lote_parcialidades (lotes importados sin created_by)
        creado_por_parcialidad = Subquery(
            LoteParcialidad.objects.filter(
                lote_id=OuterRef('pk')
            ).order_by('created_at').annotate(
                display=Coalesce(
                    NullIf(Trim(Concat('usuario__first_name', Value(' '), 'usuario__last_name')), Value('')),
                    'usuario__username',
                    output_field=CharField(),
                )
            ).values('display')[:1],
            output_field=CharField(),
        )
        # ── Subquery: nombre del último usuario que actualizó el lote ──
        modificado_por_sub = Subquery(
            AuditoriaLogs.objects.filter(
                modelo='Lote',
                objeto_id=RawSQL('"lotes"."id"::text', [], output_field=CharField()),
                accion='actualizar',
                usuario__isnull=False,
            ).order_by('-timestamp').annotate(
                display=Coalesce(
                    NullIf(Trim(Concat('usuario__first_name', Value(' '), 'usuario__last_name')), Value('')),
                    'usuario__username',
                    output_field=CharField(),
                )
            ).values('display')[:1],
            output_field=CharField(),
        )

        queryset = (
            Lote.objects
            .select_related('producto', 'centro')
            .prefetch_related('parcialidades')
            .annotate(
                _creado_por_nombre=Coalesce(creado_por_created_by, creado_por_parcialidad),
                _modificado_por_nombre=modificado_por_sub,
            )
        )
        
        # SEGURIDAD: Filtrar por centro segun rol
        user = self.request.user
        
        # ISS-FIX: Usuarios de centro pueden ver lotes de farmacia central
        # cuando están creando requisiciones (para_requisicion=true)
        # PERO solo si tienen rol MEDICO (único rol que puede crear requisiciones)
        para_requisicion = self.request.query_params.get('para_requisicion', '').lower() == 'true'
        
        if not has_global_read_access(user):
            # Usuario de centro - validar acceso
            user_centro = get_user_centro(user)
            if not user_centro:
                return Lote.objects.none()
            
            # ISS-FIX: Solo MEDICO puede usar para_requisicion=true
            rol = (getattr(user, 'rol', '') or '').lower()
            if para_requisicion and rol == 'medico':
                # ISS-FIX: Para crear requisiciones, mostrar lotes de FARMACIA CENTRAL
                # porque las requisiciones se surten desde farmacia central
                queryset = queryset.filter(centro__isnull=True)
                logger.debug(
                    f"ISS-FIX: Médico {user.username} consultando lotes de farmacia central "
                    f"para requisición"
                )
            elif para_requisicion:
                # Rol no autorizado para para_requisicion
                logger.warning(
                    f"ISS-FIX: Usuario {user.username} con rol={rol} intentó usar "
                    f"para_requisicion=true sin permiso. Denegando acceso."
                )
                return Lote.objects.none()
            else:
                # Por defecto: solo lotes de SU centro
                queryset = queryset.filter(centro=user_centro)
        else:
            # Admin/farmacia/vista: pueden filtrar por centro especifico
            centro_param = self.request.query_params.get('centro')
            if centro_param:
                if centro_param == 'central':
                    # Filtrar solo farmacia central (centro=NULL)
                    queryset = queryset.filter(centro__isnull=True)
                else:
                    queryset = queryset.filter(centro_id=centro_param)
        
        # Filtrar por producto
        producto = self.request.query_params.get('producto')
        if producto:
            queryset = queryset.filter(producto_id=producto)
        
        # Filtrar por activo (el campo real en la BD)
        activo = self.request.query_params.get('activo')
        if activo is not None:
            if activo.lower() in ['true', '1', 'si']:
                queryset = queryset.filter(activo=True)
            elif activo.lower() in ['false', '0', 'no']:
                queryset = queryset.filter(activo=False)
        
        # Busqueda por numero de lote, clave o nombre producto (ISS-003)
        # MEJORADA: Búsqueda inteligente con prioridad a coincidencias exactas
        search = self.request.query_params.get('search')
        if search and search.strip():
            search_term = search.strip()
            
            # Estrategia inteligente: Si el término parece ser una clave (numérico corto),
            # priorizar coincidencia exacta en clave
            is_numeric = search_term.isdigit()
            
            if is_numeric:
                # Primero: buscar coincidencia EXACTA en clave de producto
                exact_match = queryset.filter(producto__clave__iexact=search_term)
                
                if exact_match.exists():
                    # Si hay coincidencia exacta en clave, usar solo esos resultados
                    queryset = exact_match
                else:
                    # No hay coincidencia exacta, buscar parcial incluyendo lotes
                    queryset = queryset.filter(
                        Q(numero_lote__icontains=search_term) |
                        Q(producto__clave__icontains=search_term) |
                        Q(producto__nombre__icontains=search_term)
                    ).distinct()
            else:
                # Búsqueda de texto: estrategia de precisión
                # 1. Primero buscar SOLO en nombre del producto (más específico)
                nombre_match = queryset.filter(producto__nombre__icontains=search_term)
                
                if nombre_match.exists():
                    # Si hay coincidencias en nombre, usar solo esas
                    queryset = nombre_match
                else:
                    # Si no hay en nombre, buscar en lote y clave
                    queryset = queryset.filter(
                        Q(numero_lote__icontains=search_term) |
                        Q(producto__clave__icontains=search_term)
                    ).distinct()
        
        # Filtrar por estado de caducidad segun especificacion SIFP:
        # Normal: > 6 meses (180 dias)
        # Proximo: 3-6 meses (90-180 dias)
        # Critico: < 3 meses (90 dias)
        # Vencido: < 0 dias
        caducidad = self.request.query_params.get('caducidad')
        if caducidad:
            hoy = date.today()
            
            if caducidad == 'vencido':
                queryset = queryset.filter(fecha_caducidad__lt=hoy)
            elif caducidad == 'critico':
                # Menos de 3 meses (< 90 dias) pero no vencido
                queryset = queryset.filter(
                    fecha_caducidad__gte=hoy,
                    fecha_caducidad__lt=hoy + timedelta(days=90)
                )
            elif caducidad == 'proximo':
                # Entre 3 y 6 meses (90-180 dias)
                queryset = queryset.filter(
                    fecha_caducidad__gte=hoy + timedelta(days=90),
                    fecha_caducidad__lt=hoy + timedelta(days=180)
                )
            elif caducidad == 'normal':
                # Mas de 6 meses (> 180 dias)
                queryset = queryset.filter(fecha_caducidad__gte=hoy + timedelta(days=180))
        
        # Filtrar por stock minimo (para catalogo de requisiciones)
        stock_min = self.request.query_params.get('stock_min')
        if stock_min:
            try:
                queryset = queryset.filter(cantidad_actual__gte=int(stock_min))
            except (ValueError, TypeError):
                pass
        
        # Filtrar por existencia de stock (con_stock/sin_stock)
        con_stock = self.request.query_params.get('con_stock')
        if con_stock in ['con_stock', 'true', '1', True]:
            queryset = queryset.filter(cantidad_actual__gt=0)
        elif con_stock in ['sin_stock', 'false', '0', False]:
            queryset = queryset.filter(cantidad_actual=0)
        
        # Filtrar solo lotes disponibles (no vencidos) para el catalogo
        solo_disponibles = self.request.query_params.get('solo_disponibles')
        if solo_disponibles == 'true':
            queryset = queryset.filter(
                activo=True,
                fecha_caducidad__gt=date.today()
            )
        
        # Filtrar por numero_contrato (para validación CCG en frontend)
        numero_contrato_param = self.request.query_params.get('numero_contrato')
        if numero_contrato_param:
            queryset = queryset.filter(numero_contrato=numero_contrato_param)

        # ISS-FIX: Filtrar por rango de fechas (para exportación de trazabilidad)
        fecha_desde = self.request.query_params.get('fecha_desde')
        if fecha_desde:
            try:
                from datetime import datetime
                fecha_desde_dt = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
                queryset = queryset.filter(created_at__date__gte=fecha_desde_dt)
            except (ValueError, TypeError):
                pass
        
        fecha_hasta = self.request.query_params.get('fecha_hasta')
        if fecha_hasta:
            try:
                from datetime import datetime
                fecha_hasta_dt = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
                queryset = queryset.filter(created_at__date__lte=fecha_hasta_dt)
            except (ValueError, TypeError):
                pass
        
        return queryset.order_by('-created_at')
    
    def retrieve(self, request, *args, **kwargs):
        """
        Obtiene un lote específico por ID.
        
        ISS-FIX: Usuarios de centro pueden consultar lotes de farmacia central
        cuando están verificando stock para requisiciones (para_requisicion=true).
        Esto es necesario porque las requisiciones se surten desde farmacia central.
        PERO solo si tienen rol MEDICO (único rol que puede crear requisiciones).
        """
        para_requisicion = request.query_params.get('para_requisicion', '').lower() == 'true'
        user = request.user
        
        # Si es para requisición y el usuario es de centro, buscar el lote sin filtro de centro
        if para_requisicion and not has_global_read_access(user):
            # ISS-FIX: Solo MEDICO puede usar para_requisicion
            rol = (getattr(user, 'rol', '') or '').lower()
            if rol != 'medico':
                logger.warning(
                    f"ISS-FIX: Usuario {user.username} con rol={rol} intentó usar "
                    f"para_requisicion=true en retrieve. Denegando acceso."
                )
                return Response(
                    {'error': 'No tiene permisos para consultar lotes para requisiciones'},
                    status=status.HTTP_403_FORBIDDEN
                )
            
            try:
                # Buscar el lote directamente (sin filtro de centro)
                lote = Lote.objects.select_related('producto', 'centro').get(pk=kwargs['pk'])
                
                # Solo permitir ver lotes de farmacia central (para requisiciones)
                if lote.centro is not None:
                    return Response(
                        {'error': 'Solo puede consultar lotes de farmacia central para requisiciones'},
                        status=status.HTTP_403_FORBIDDEN
                    )
                
                serializer = self.get_serializer(lote)
                return Response(serializer.data)
            except Lote.DoesNotExist:
                return Response(
                    {'error': 'Lote no encontrado'},
                    status=status.HTTP_404_NOT_FOUND
                )
        
        # Comportamiento normal: usa get_queryset() con filtros de seguridad
        return super().retrieve(request, *args, **kwargs)
    
    def create(self, request, *args, **kwargs):
        """
        Crea un nuevo lote con validaciones y alerta de contrato global.
        
        PARCIALIDADES: La cantidad_inicial se registra automáticamente como
        la primera entrega (parcialidad inicial). Esto permite trazabilidad
        completa desde la primera carga, ya sea manual o por importación.
        """
        try:
            with transaction.atomic():
                serializer = self.get_serializer(data=request.data)
                serializer.is_valid(raise_exception=True)
                self.perform_create(serializer)
                lote = serializer.instance
                
                # Verificar si fue MERGE (el serializer ya creó la parcialidad)
                merge_realizado = getattr(serializer, '_lote_merge_realizado', None)
                
                # PARCIALIDADES: Crear parcialidad inicial automáticamente
                # Solo si NO fue merge (en merge ya se crea en el serializer)
                parcialidad_creada = False
                if not merge_realizado and lote.cantidad_inicial and lote.cantidad_inicial > 0:
                    # Verificar que no exista ya (idempotencia)
                    if not LoteParcialidad.objects.filter(lote=lote).exists():
                        fecha_entrega = lote.fecha_fabricacion or date.today()
                        LoteParcialidad.objects.create(
                            lote=lote,
                            fecha_entrega=fecha_entrega,
                            cantidad=lote.cantidad_inicial,
                            notas='Entrega inicial - creación de lote',
                            usuario=request.user if request.user.is_authenticated else None
                        )
                        parcialidad_creada = True
                        logger.info(
                            f"Parcialidad inicial creada - Lote: {lote.numero_lote}, "
                            f"Cantidad: {lote.cantidad_inicial}, Usuario: {request.user.username}"
                        )
                
                headers = self.get_success_headers(serializer.data)
                response_data = dict(serializer.data)
                
                # Informar sobre parcialidad inicial
                if parcialidad_creada:
                    response_data['parcialidad_inicial'] = {
                        'creada': True,
                        'cantidad': lote.cantidad_inicial,
                        'mensaje': 'Se registró la entrega inicial automáticamente'
                    }

                # ISS-INV-001: Incluir alerta de contrato del lote si existe (sobreentrega)
                alerta_lote = getattr(serializer, '_alerta_contrato_lote', None)
                if alerta_lote:
                    response_data['alerta_contrato_lote'] = alerta_lote

                # ISS-INV-003: Incluir alerta de contrato global si existe
                alerta = getattr(serializer, '_alerta_contrato_global', None)
                if alerta:
                    response_data['alerta_contrato_global'] = alerta

                # MERGE: Informar al frontend si el lote fue fusionado con existente
                merge_info = getattr(serializer, '_lote_merge_realizado', None)
                if merge_info:
                    response_data['merge_realizado'] = merge_info
                    response_data['mensaje_informativo'] = (
                        f'El lote "{merge_info["numero_lote"]}" ya existía. '
                        f'Se agregaron {merge_info["cantidad_agregada"]} unidades. '
                        f'Nueva cantidad total: {merge_info["nueva_cantidad_total"]}.'
                    )

                # AUTO-SUFIJO: Informar al frontend si el número de lote fue renombrado
                auto_renombrado = getattr(serializer, '_numero_lote_auto_renombrado', None)
                if auto_renombrado:
                    response_data['numero_lote_auto_asignado'] = auto_renombrado
                    response_data['mensaje_informativo'] = (
                        f'El número de lote "{auto_renombrado["original"]}" ya existe para este producto. '
                        f'Se asignó automáticamente: "{auto_renombrado["asignado"]}".'
                    )

                return Response(
                    response_data,
                    status=status.HTTP_201_CREATED,
                    headers=headers
                )
        except serializers.ValidationError as e:
            return Response(
                {'error': 'Error de validacion', 'detalles': e.detail}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        except DjangoValidationError as e:
            # Model-level clean() raises Django ValidationError (not DRF's)
            detalles = e.message_dict if hasattr(e, 'message_dict') else {'__all__': e.messages}
            return Response(
                {'error': 'Error de validacion', 'detalles': detalles},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            return Response(
                {'error': 'Error al crear lote', 'mensaje': str(e)}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def update(self, request, *args, **kwargs):
        """Actualiza un lote existente.
        
        ISS-010 FIX: Validación explícita de permisos para evitar IDOR.
        Solo admin/farmacia pueden modificar lotes de farmacia central.
        Usuarios de centro solo pueden modificar lotes de SU centro.
        """
        try:
            partial = kwargs.pop('partial', False)
            instance = self.get_object()
            
            # ISS-010: Validar permisos de escritura sobre este lote específico
            user = request.user
            if not is_farmacia_or_admin(user):
                user_centro = get_user_centro(user)
                lote_centro = instance.centro
                
                # Si el lote es de farmacia central o de otro centro, denegar
                if lote_centro is None or (user_centro and lote_centro.pk != user_centro.pk):
                    logger.warning(
                        f"ISS-010: Intento de modificación no autorizada de lote. "
                        f"Usuario={user.username}, lote={instance.numero_lote}, "
                        f"lote_centro={lote_centro}, user_centro={user_centro}"
                    )
                    return Response(
                        {'error': 'No tiene permisos para modificar este lote'},
                        status=status.HTTP_403_FORBIDDEN
                    )
            
            serializer = self.get_serializer(instance, data=request.data, partial=partial)
            serializer.is_valid(raise_exception=True)
            self.perform_update(serializer)
            
            response_data = dict(serializer.data)
            # ISS-INV-003: Incluir alerta de contrato global si existe
            alerta = getattr(serializer, '_alerta_contrato_global', None)
            if alerta:
                response_data['alerta_contrato_global'] = alerta
            
            return Response(response_data)
        except serializers.ValidationError as e:
            return Response(
                {'error': 'Error de validacion', 'detalles': e.detail}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        except DjangoValidationError as e:
            detalles = e.message_dict if hasattr(e, 'message_dict') else {'__all__': e.messages}
            return Response(
                {'error': 'Error de validacion', 'detalles': detalles},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            return Response(
                {'error': 'Error al actualizar lote', 'mensaje': str(e)}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def destroy(self, request, *args, **kwargs):
        """
        Elimina un lote.
        
        ISS-010 FIX: Validación explícita de permisos para evitar IDOR.
        AUDIT-OWN: Solo el usuario que registró el lote o un admin puede eliminarlo.
        
        Validaciones:
        - Permisos de escritura sobre el lote
        - No puede eliminarse si tiene movimientos asociados
        """
        instance = self.get_object()
        
        # ISS-010: Validar permisos de escritura sobre este lote específico
        user = request.user
        if not is_farmacia_or_admin(user):
            user_centro = get_user_centro(user)
            lote_centro = instance.centro
            
            # Si el lote es de farmacia central o de otro centro, denegar
            if lote_centro is None or (user_centro and lote_centro.pk != user_centro.pk):
                logger.warning(
                    f"ISS-010: Intento de eliminación no autorizada de lote. "
                    f"Usuario={user.username}, lote={instance.numero_lote}, "
                    f"lote_centro={lote_centro}, user_centro={user_centro}"
                )
                return Response(
                    {'error': 'No tiene permisos para eliminar este lote'},
                    status=status.HTTP_403_FORBIDDEN
                )

        # AUDIT-OWN: Solo el creador o admin puede eliminar
        rol = (getattr(user, 'rol', '') or '').lower()
        ROLES_ADMIN = {'admin', 'admin_sistema', 'superusuario', 'administrador'}
        if not user.is_superuser and rol not in ROLES_ADMIN:
            log_crear = AuditoriaLogs.objects.filter(
                modelo='Lote', objeto_id=str(instance.pk), accion='crear'
            ).order_by('timestamp').first()
            if log_crear and log_crear.usuario and log_crear.usuario.pk != user.pk:
                logger.warning(
                    f"AUDIT-OWN: {user.username} intentó eliminar lote #{instance.pk} "
                    f"(creado por {log_crear.usuario.username}) - DENEGADO"
                )
                return Response(
                    {'error': 'Solo el usuario que registró este lote o un administrador puede eliminarlo',
                     'creado_por': log_crear.usuario.get_full_name() or log_crear.usuario.username},
                    status=status.HTTP_403_FORBIDDEN
                )
        
        try:
            # Verificar si tiene movimientos
            if Movimiento.objects.filter(lote=instance).exists():
                total_movimientos = Movimiento.objects.filter(lote=instance).count()
                
                return Response({
                    'error': 'No se puede eliminar el lote',
                    'razon': 'Tiene movimientos asociados',
                    'total_movimientos': total_movimientos,
                    'sugerencia': 'Marque el lote como inactivo en lugar de eliminarlo'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Si no tiene movimientos, se puede eliminar
            numero_lote = instance.numero_lote
            producto_clave = instance.producto.clave
            instance.delete()
            
            return Response({
                'mensaje': 'Lote eliminado exitosamente',
                'lote_eliminado': numero_lote,
                'producto': producto_clave
            }, status=status.HTTP_204_NO_CONTENT)
            
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            return Response({
                'error': 'Error al eliminar lote',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path='consolidados')
    def consolidados(self, request):
        """
        TRAZABILIDAD: Obtiene lotes consolidados (únicos por numero_lote + producto).
        
        Los lotes que se distribuyen a múltiples centros se consolidan sumando
        las cantidades de todos los centros. Esto representa la trazabilidad real:
        un lote físico = un registro, aunque esté distribuido en varios centros.
        
        GET /api/lotes/consolidados/
        
        Parámetros opcionales:
        - search: Buscar por número de lote o clave/nombre de producto
        - producto: Filtrar por ID de producto
        - centro: ID del centro o 'central' para farmacia central
        - activo: true/false
        - page, page_size: Paginación
        
        ISS-FIX SEGURIDAD: Usuarios de centro solo ven lotes de su propio centro.
        Admin/farmacia/vista ven todos los centros.
        
        Returns:
            Lista de lotes únicos con cantidad_total sumada de todos los centros
            y lista de centros donde está distribuido cada lote.
        """
        from collections import defaultdict
        
        # Obtener lotes con filtro opcional de centro
        # Optimización: usar only() para traer solo campos necesarios
        queryset = Lote.objects.select_related('producto', 'centro').only(
            'id', 'numero_lote', 'fecha_caducidad', 'fecha_fabricacion',
            'precio_unitario', 'marca', 'numero_contrato', 'ubicacion',
            'cantidad_inicial', 'cantidad_actual', 'cantidad_contrato', 
            'cantidad_contrato_global',  # ISS-INV-003: Incluir CCG en consolidados
            'activo',
            'producto__id', 'producto__clave', 'producto__nombre', 
            'producto__descripcion', 'producto__presentacion', 'producto__unidad_medida',
            'centro__id', 'centro__nombre'
        )
        
        # Determinar rol del usuario PRIMERO
        user = request.user
        tiene_acceso_global = has_global_read_access(user)
        
        # ISS-FIX SEGURIDAD: Usuarios de centro solo ven SU centro
        # Esto previene IDOR donde un usuario de centro vería lotes de otros centros
        if not tiene_acceso_global:
            user_centro = get_user_centro(user)
            if not user_centro:
                logger.warning(
                    f"ISS-FIX: Usuario {user.username} sin centro intentó acceder a consolidados. Denegando."
                )
                return Response({
                    'count': 0,
                    'total_pages': 0,
                    'current_page': 1,
                    'results': [],
                    'warning': 'No tiene un centro asignado'
                })
            # Filtrar SOLO lotes de su centro
            queryset = queryset.filter(centro=user_centro)
            logger.debug(
                f"ISS-FIX: Usuario {user.username} de centro {user_centro.nombre} "
                f"accediendo a consolidados filtrados"
            )
        
        # Verificar si hay búsqueda específica por número de lote
        search = request.query_params.get('search', '').strip()
        busqueda_especifica = bool(search and len(search) >= 3)
        
        # Filtrar por activo según rol:
        # - Farmacia/Admin: Ven TODOS los lotes (activos e inactivos) para trazabilidad completa
        # - Centro: Solo ven lotes activos
        activo_param = request.query_params.get('activo')
        if activo_param is not None:
            if activo_param.lower() in ['true', '1', 'activo']:
                queryset = queryset.filter(activo=True)
            elif activo_param.lower() in ['false', '0', 'inactivo']:
                queryset = queryset.filter(activo=False)
            # Si es vacío o 'todos', no filtrar por activo
        elif not tiene_acceso_global:
            # Usuarios de Centro: por defecto solo ver lotes activos
            queryset = queryset.filter(activo=True)
        # Farmacia/Admin sin filtro: ver TODOS los lotes (activos e inactivos) para trazabilidad
        
        # Filtrar por con_stock según rol del usuario:
        # - Farmacia/Admin: Por defecto ven TODOS los lotes (con y sin stock) para trazabilidad
        # - Centro: Por defecto solo ven lotes CON stock disponible
        
        con_stock_param = request.query_params.get('con_stock')
        if con_stock_param:
            if con_stock_param.lower() in ['con_stock', 'true', '1']:
                queryset = queryset.filter(cantidad_actual__gt=0)
            elif con_stock_param.lower() in ['sin_stock', 'false', '0']:
                queryset = queryset.filter(cantidad_actual=0)
        elif not tiene_acceso_global:
            # Usuarios de Centro: por defecto solo ver lotes con stock
            queryset = queryset.filter(cantidad_actual__gt=0)
        # Farmacia/Admin sin filtro: mostrar todos los lotes
        
        # ISS-FIX: Filtrar por centro SOLO si usuario tiene acceso global
        # Los usuarios de centro ya están filtrados arriba por su propio centro
        centro_param = request.query_params.get('centro')
        if centro_param and tiene_acceso_global:
            if centro_param == 'central':
                # Solo farmacia central (centro=NULL)
                queryset = queryset.filter(centro__isnull=True)
            elif centro_param != 'todos':
                # Centro específico por ID
                try:
                    queryset = queryset.filter(centro_id=int(centro_param))
                except (ValueError, TypeError):
                    pass
        
        # Aplicar filtros de búsqueda
        # MEJORADA: Búsqueda inteligente con prioridad a coincidencias exactas
        search = request.query_params.get('search')
        if search and search.strip():
            search_term = search.strip()
            
            # Estrategia inteligente: Si el término parece ser una clave (numérico corto),
            # priorizar coincidencia exacta en clave
            is_numeric = search_term.isdigit()
            
            if is_numeric:
                # Primero: buscar coincidencia EXACTA en clave de producto
                exact_match = queryset.filter(producto__clave__iexact=search_term)
                
                if exact_match.exists():
                    # Si hay coincidencia exacta en clave, usar solo esos resultados
                    queryset = exact_match
                else:
                    # No hay coincidencia exacta, buscar parcial incluyendo lotes
                    queryset = queryset.filter(
                        Q(numero_lote__icontains=search_term) |
                        Q(producto__clave__icontains=search_term) |
                        Q(producto__nombre__icontains=search_term)
                    ).distinct()
            else:
                # Búsqueda de texto: estrategia de precisión
                # 1. Primero buscar SOLO en nombre del producto (más específico)
                nombre_match = queryset.filter(producto__nombre__icontains=search_term)
                
                if nombre_match.exists():
                    # Si hay coincidencias en nombre, usar solo esas
                    queryset = nombre_match
                else:
                    # Si no hay en nombre, buscar en lote y clave
                    queryset = queryset.filter(
                        Q(numero_lote__icontains=search_term) |
                        Q(producto__clave__icontains=search_term)
                    ).distinct()
        
        producto = request.query_params.get('producto')
        if producto:
            queryset = queryset.filter(producto_id=producto)
        
        # Filtrar por estado de caducidad (mismo que en get_queryset)
        caducidad = request.query_params.get('caducidad')
        if caducidad:
            hoy = date.today()
            if caducidad == 'vencido':
                queryset = queryset.filter(fecha_caducidad__lt=hoy)
            elif caducidad == 'critico':
                queryset = queryset.filter(
                    fecha_caducidad__gte=hoy,
                    fecha_caducidad__lt=hoy + timedelta(days=90)
                )
            elif caducidad == 'proximo':
                queryset = queryset.filter(
                    fecha_caducidad__gte=hoy + timedelta(days=90),
                    fecha_caducidad__lt=hoy + timedelta(days=180)
                )
            elif caducidad == 'normal':
                queryset = queryset.filter(fecha_caducidad__gte=hoy + timedelta(days=180))
        
        queryset = queryset.order_by('producto__clave', 'numero_lote', 'fecha_caducidad')
        
        # Consolidar por (producto_id, numero_lote)
        lotes_consolidados = defaultdict(lambda: {
            'id': None,  # ID del primer lote encontrado (para referencia)
            'producto': None,
            'producto_id': None,
            'producto_clave': '',
            'producto_nombre': '',
            'producto_info': None,
            'numero_lote': '',
            'cantidad_total': 0,
            'cantidad_inicial_total': 0,
            'fecha_caducidad': None,
            'precio_unitario': 0,
            'marca': '-',
            'numero_contrato': None,
            'centros': [],
            'centros_detalle': [],  # Lista con nombre y cantidad por centro
            'cantidad_contrato': None,  # Total de contrato (agregado de todos los lotes del grupo)
            'cantidad_contrato_total': 0,  # Suma de cantidades contrato
            'cantidad_contrato_global': None,  # ISS-INV-003: Contrato global compartido
            'activo': True,
            'dias_para_caducar': 999,
            'alerta_caducidad': 'normal',
            'lotes_ids': [],  # IDs de todos los lotes consolidados para verificar movimientos
        })
        
        # Obtener IDs de lotes que tienen movimientos (consulta eficiente)
        lotes_con_movimientos = set(
            Movimiento.objects.filter(
                lote__in=queryset
            ).values_list('lote_id', flat=True).distinct()
        )
        
        hoy = date.today()
        
        for lote in queryset:
            key = (lote.producto_id, lote.numero_lote)
            cons = lotes_consolidados[key]
            
            if cons['id'] is None:
                cons['id'] = lote.id
                cons['producto'] = lote.producto_id
                cons['producto_id'] = lote.producto_id
                cons['producto_clave'] = lote.producto.clave
                cons['producto_nombre'] = lote.producto.nombre or lote.producto.descripcion or ''
                cons['producto_info'] = {
                    'presentacion': lote.producto.presentacion or '',
                    'unidad_medida': lote.producto.unidad_medida or 'PIEZA',
                }
                cons['numero_lote'] = lote.numero_lote
                cons['fecha_caducidad'] = lote.fecha_caducidad.isoformat() if lote.fecha_caducidad else None
                cons['precio_unitario'] = float(lote.precio_unitario or 0)
                cons['marca'] = lote.marca or '-'
                cons['numero_contrato'] = lote.numero_contrato
                # ISS-INV-003: Incluir cantidad_contrato_global
                cons['cantidad_contrato_global'] = lote.cantidad_contrato_global
                
                # Calcular días para caducar y alerta
                if lote.fecha_caducidad:
                    dias = (lote.fecha_caducidad - hoy).days
                    cons['dias_para_caducar'] = dias
                    if dias < 0:
                        cons['alerta_caducidad'] = 'vencido'
                    elif dias < 90:
                        cons['alerta_caducidad'] = 'critico'
                    elif dias < 180:
                        cons['alerta_caducidad'] = 'proximo'
                    else:
                        cons['alerta_caducidad'] = 'normal'
            
            cons['cantidad_total'] += lote.cantidad_actual
            cons['cantidad_inicial_total'] += lote.cantidad_inicial
            # Acumular cantidad_contrato de todos los lotes del grupo
            if lote.cantidad_contrato:
                cons['cantidad_contrato_total'] += lote.cantidad_contrato
            cons['lotes_ids'].append(lote.id)  # Agregar ID para verificar movimientos
            
            # Registrar centro
            centro_nombre = lote.centro.nombre if lote.centro else 'Almacén Central'
            if centro_nombre not in cons['centros']:
                cons['centros'].append(centro_nombre)
            
            cons['centros_detalle'].append({
                'centro_id': lote.centro_id,
                'centro_nombre': centro_nombre,
                'cantidad': lote.cantidad_actual,
                'ubicacion': lote.ubicacion or '-',
            })
        
        # Convertir a lista y agregar campos calculados
        resultados = []
        for key, cons in lotes_consolidados.items():
            # Calcular porcentaje consumido
            if cons['cantidad_inicial_total'] > 0:
                porcentaje = round((1 - cons['cantidad_total'] / cons['cantidad_inicial_total']) * 100, 1)
            else:
                porcentaje = 0
            cons['porcentaje_consumido'] = porcentaje
            cons['cantidad_actual'] = cons['cantidad_total']  # Alias para compatibilidad
            cons['cantidad_inicial'] = cons['cantidad_inicial_total']
            # Exponer cantidad_contrato solo si algún lote tenía contrato
            cons['cantidad_contrato'] = cons['cantidad_contrato_total'] if cons['cantidad_contrato_total'] > 0 else None
            # Calcular pendiente respecto al contrato
            if cons['cantidad_contrato']:
                cons['cantidad_pendiente'] = max(0, cons['cantidad_contrato'] - cons['cantidad_inicial_total'])
            else:
                cons['cantidad_pendiente'] = 0
            
            # ISS-INV-003: Calcular cantidad_pendiente_global si hay CCG
            if cons['cantidad_contrato_global'] and cons['numero_contrato']:
                # Sumar TODAS las entregas del mismo producto+contrato en la BD
                # (no solo las del queryset actual que pueden estar filtradas)
                # CRÍTICO: Usa cantidad_inicial (recibido), NO cantidad_actual (disponible)
                # Las salidas NO afectan el contrato. Ej: Contrato 500, recibido 200, salió 100 → falta 300 (no 400)
                total_recibido_global = Lote.objects.filter(
                    producto_id=cons['producto_id'],
                    numero_contrato=cons['numero_contrato'],
                    cantidad_contrato_global__isnull=False
                ).aggregate(total=Sum('cantidad_inicial'))['total'] or 0
                
                # Total en INVENTARIO actual (cantidad_actual): Lo que REALMENTE hay en stock
                total_inventario_global = Lote.objects.filter(
                    producto_id=cons['producto_id'],
                    numero_contrato=cons['numero_contrato'],
                    cantidad_contrato_global__isnull=False,
                    activo=True
                ).aggregate(total=Sum('cantidad_actual'))['total'] or 0
                
                cons['cantidad_recibido_global'] = total_recibido_global  # Total recibido (suma cantidad_inicial)
                cons['total_inventario_global'] = total_inventario_global  # Total en inventario (suma cantidad_actual)
                cons['cantidad_pendiente_global'] = cons['cantidad_contrato_global'] - total_recibido_global
            else:
                cons['cantidad_recibido_global'] = None
                cons['total_inventario_global'] = None
                cons['cantidad_pendiente_global'] = None
            del cons['cantidad_contrato_total']  # No exponer campo auxiliar
            cons['centro_nombre'] = ', '.join(cons['centros'][:2]) + ('...' if len(cons['centros']) > 2 else '')
            
            # ISS-TRAZ: Indicar si el lote tiene movimientos (para bloquear edición de campos críticos)
            # Si CUALQUIERA de los lotes consolidados tiene movimientos, se bloquea la edición
            cons['tiene_movimientos'] = any(lid in lotes_con_movimientos for lid in cons['lotes_ids'])
            del cons['lotes_ids']  # No exponer los IDs internos
            
            resultados.append(cons)
        
        # Ordenar por clave de producto y número de lote
        resultados.sort(key=lambda x: (x['producto_clave'], x['numero_lote']))
        
        # Paginación simple
        page_size = int(request.query_params.get('page_size', 25))
        page = int(request.query_params.get('page', 1))
        total = len(resultados)
        start = (page - 1) * page_size
        end = start + page_size
        
        return Response({
            'count': total,
            'total_pages': (total + page_size - 1) // page_size,
            'current_page': page,
            'results': resultados[start:end],
        })

    @action(detail=False, methods=['get'], url_path='diagnostico-centro')
    def diagnostico_centro(self, request):
        """
        ISS-FIX (lotes-centro): Endpoint de diagnóstico de lotes por centro.
        GET /api/lotes/diagnostico-centro/?producto_id={id}
        
        Solo accesible para admin/farmacia. Muestra distribución de lotes
        de un producto específico entre todos los centros.
        """
        # Solo admin/farmacia puede usar este endpoint
        if not is_farmacia_or_admin(request.user) and not request.user.is_superuser:
            return Response({
                'error': 'Este endpoint solo está disponible para administradores'
            }, status=status.HTTP_403_FORBIDDEN)
        
        producto_id = request.query_params.get('producto_id')
        if not producto_id:
            return Response({
                'error': 'Se requiere el parámetro producto_id'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            producto = Producto.objects.get(pk=producto_id)
        except Producto.DoesNotExist:
            return Response({
                'error': f'Producto con ID {producto_id} no encontrado'
            }, status=status.HTTP_404_NOT_FOUND)
        
        try:
            # Obtener TODOS los lotes del producto (sin filtros de centro)
            todos_lotes = Lote.objects.filter(
                producto=producto
            ).select_related('centro').order_by('centro_id', 'numero_lote')
            
            # Agrupar por centro
            centros_info = {}
            for lote in todos_lotes:
                centro_key = lote.centro_id if lote.centro else 'central'
                centro_nombre = lote.centro.nombre if lote.centro else 'Farmacia Central'
                
                if centro_key not in centros_info:
                    centros_info[centro_key] = {
                        'centro_id': centro_key,
                        'centro_nombre': centro_nombre,
                        'lotes': [],
                        'total_stock': 0,
                        'total_lotes': 0,
                        'lotes_activos': 0,
                        'lotes_con_stock': 0,
                    }
                
                lote_info = {
                    'id': lote.id,
                    'numero_lote': lote.numero_lote,
                    'cantidad_inicial': lote.cantidad_inicial,
                    'cantidad_actual': lote.cantidad_actual,
                    'fecha_caducidad': lote.fecha_caducidad.isoformat() if lote.fecha_caducidad else None,
                    'activo': lote.activo,
                    'created_at': lote.created_at.isoformat() if hasattr(lote, 'created_at') and lote.created_at else None,
                }
                
                centros_info[centro_key]['lotes'].append(lote_info)
                centros_info[centro_key]['total_stock'] += lote.cantidad_actual
                centros_info[centro_key]['total_lotes'] += 1
                if lote.activo:
                    centros_info[centro_key]['lotes_activos'] += 1
                if lote.cantidad_actual > 0:
                    centros_info[centro_key]['lotes_con_stock'] += 1
            
            # Listar todos los centros del sistema para ver cuáles NO tienen lotes
            from core.models import Centro
            all_centros = list(Centro.objects.values('id', 'nombre'))
            centros_sin_lotes = [
                c for c in all_centros 
                if c['id'] not in centros_info and 'central' not in centros_info
            ]
            
            return Response({
                'producto': {
                    'id': producto.id,
                    'clave': producto.clave,
                    'nombre': producto.nombre,
                },
                'resumen_global': {
                    'total_lotes': todos_lotes.count(),
                    'total_stock_global': sum(l.cantidad_actual for l in todos_lotes),
                    'centros_con_lotes': len(centros_info),
                    'centros_sin_lotes': len(centros_sin_lotes),
                },
                'distribucion_por_centro': centros_info,
                'centros_sin_lotes': centros_sin_lotes,
            })
            
        except Exception as e:
            logger.error(f"Error en diagnostico_centro: {str(e)}", exc_info=True)
            return Response({
                'error': f'Error: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path='exportar-pdf')
    def exportar_pdf(self, request):
        """
        Genera PDF de inventario de lotes con filtros opcionales.
        
        Usa los mismos filtros que get_queryset para consistencia:
        - producto: ID del producto
        - activo: true/false
        - search: búsqueda en número de lote o producto
        - caducidad: vencido/critico/proximo/normal
        - con_stock: con_stock/sin_stock
        - centro: ID del centro o 'central'
        
        Respeta permisos de usuario:
        - Usuarios de centro solo ven lotes de su centro
        - Admin/Farmacia/Vista ven todo
        """
        from core.utils.pdf_reports import generar_reporte_lotes
        
        try:
            # Usar get_queryset que ya aplica filtros y permisos
            queryset = self.get_queryset()
            
            # Limitar a 500 lotes para PDF
            lotes = queryset[:500]
            
            # Preparar datos para el PDF
            lotes_data = []
            for lote in lotes:
                lotes_data.append({
                    'producto_clave': getattr(lote.producto, 'clave', '') if lote.producto else '',
                    'producto_nombre': getattr(lote.producto, 'nombre', '') if lote.producto else '',
                    'numero_lote': lote.numero_lote or '',
                    'fecha_fabricacion': lote.fecha_fabricacion.strftime('%Y-%m-%d') if lote.fecha_fabricacion else '',
                    'fecha_caducidad': lote.fecha_caducidad.strftime('%Y-%m-%d') if lote.fecha_caducidad else '',
                    'fecha_caducidad_raw': lote.fecha_caducidad,
                    'cantidad_contrato': lote.cantidad_contrato,
                    'cantidad_contrato_global': lote.cantidad_contrato_global,
                    'cantidad_inicial': lote.cantidad_inicial,
                    'cantidad_actual': lote.cantidad_actual,
                    'numero_contrato': lote.numero_contrato or '',
                    'centro_nombre': getattr(lote.centro, 'nombre', 'Farmacia Central') if lote.centro else 'Farmacia Central',
                    'activo': lote.activo,
                })
            
            # Preparar filtros para mostrar en el PDF
            filtros = {}
            if request.query_params.get('producto'):
                try:
                    producto = Producto.objects.get(pk=request.query_params.get('producto'))
                    filtros['producto'] = f"{producto.clave} - {producto.nombre}"
                except Producto.DoesNotExist:
                    pass
            if request.query_params.get('centro'):
                centro_param = request.query_params.get('centro')
                if centro_param == 'central':
                    filtros['centro'] = 'Farmacia Central'
                else:
                    try:
                        centro = Centro.objects.get(pk=centro_param)
                        filtros['centro'] = centro.nombre
                    except Centro.DoesNotExist:
                        pass
            if request.query_params.get('caducidad'):
                filtros['caducidad'] = request.query_params.get('caducidad')
            if request.query_params.get('con_stock'):
                filtros['con_stock'] = request.query_params.get('con_stock')
            if request.query_params.get('activo'):
                filtros['activo'] = request.query_params.get('activo')
            if request.query_params.get('search'):
                filtros['busqueda'] = request.query_params.get('search')
            
            pdf_buffer = generar_reporte_lotes(lotes_data, filtros)
            
            response = HttpResponse(
                pdf_buffer.getvalue(),
                content_type='application/pdf'
            )
            response['Content-Disposition'] = f'attachment; filename="Inventario_Lotes_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
            
            return response
            
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            return Response({
                'error': 'Error al generar PDF de lotes',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path='exportar-excel')
    def exportar_excel(self, request):
        """
        Exporta lotes aplicando los filtros de listado.
        Incluye campos de contrato global calculados de forma eficiente.
        """
        try:
            # Reutilizar el queryset que ya aplica todos los filtros
            lotes_queryset = self.get_queryset().select_related('producto', 'centro')
            lista_exportar = list(lotes_queryset)
            
            # Precalcular totales globales por (producto_id, numero_contrato) para eficiencia
            from collections import defaultdict
            from django.db.models import Sum
            
            # Agrupar lotes por producto+contrato para calcular totales globales
            contratos_data = {}
            for lote in lista_exportar:
                if lote.cantidad_contrato_global and lote.numero_contrato:
                    key = (lote.producto_id, lote.numero_contrato)
                    if key not in contratos_data:
                        # Calcular una sola vez por grupo
                        totales = Lote.objects.filter(
                            producto_id=lote.producto_id,
                            numero_contrato=lote.numero_contrato,
                            cantidad_contrato_global__isnull=False,
                            activo=True
                        ).aggregate(
                            total_inicial=Sum('cantidad_inicial'),
                            total_actual=Sum('cantidad_actual')
                        )
                        total_recibido = totales['total_inicial'] or 0
                        total_stock = totales['total_actual'] or 0
                        pendiente = lote.cantidad_contrato_global - total_recibido
                        contratos_data[key] = {
                            'total_stock': total_stock,
                            'pendiente': pendiente
                        }

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Lotes'

            # Headers claros y profesionales
            headers = [
                'Clave', 'Producto', 'Presentación', 'Código de Lote', 'Fecha de Caducidad',
                'Contrato Lote', 'Recibido Lote', 'Inventario Lote',
                'Contrato Global', 'Inventario Global', 'Pendiente Global',
                'Precio Unitario', 'Fecha de Entrega',
                'Ubicación', 'Número de Contrato', 'Marca / Laboratorio', 'Activo'
            ]
            ws.append(headers)
            
            header_fill = PatternFill(start_color='632842', end_color='632842', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            for lote in lista_exportar:
                # Datos del producto
                clave = lote.producto.clave if lote.producto else ''
                nom_prod = lote.producto.nombre if lote.producto else 'Producto Desconocido'
                presentacion = lote.producto.presentacion if lote.producto else ''

                # Datos del lote individual
                cant_contrato_lote = lote.cantidad_contrato or ''
                cant_recibida_lote = lote.cantidad_inicial or 0
                cant_stock_lote = lote.cantidad_actual or 0
                
                # Datos del contrato global (precalculados)
                cant_contrato_global = ''
                stock_global = ''
                pendiente_global = ''
                
                if lote.cantidad_contrato_global and lote.numero_contrato:
                    key = (lote.producto_id, lote.numero_contrato)
                    cant_contrato_global = lote.cantidad_contrato_global
                    if key in contratos_data:
                        stock_global = contratos_data[key]['total_stock']
                        pendiente_global = contratos_data[key]['pendiente']
                
                activo_str = 'Sí' if lote.activo else 'No'
                ubicacion_str = lote.centro.nombre if lote.centro else 'Almacén Central'
                
                fecha_cad_str = lote.fecha_caducidad.strftime('%d/%m/%Y') if lote.fecha_caducidad else ''
                fecha_fab_str = lote.fecha_fabricacion.strftime('%d/%m/%Y') if lote.fecha_fabricacion else ''

                ws.append([
                    clave,
                    nom_prod,
                    presentacion,
                    lote.numero_lote or '',
                    fecha_cad_str,
                    cant_contrato_lote,
                    cant_recibida_lote,
                    cant_stock_lote,
                    cant_contrato_global,
                    stock_global,
                    pendiente_global,
                    float(lote.precio_unitario) if lote.precio_unitario else 0.00,
                    fecha_fab_str,
                    ubicacion_str,
                    lote.numero_contrato or '',
                    lote.marca or '',
                    activo_str
                ])

            # Ajustar anchos de columna
            column_widths = [10, 40, 25, 15, 15, 15, 15, 12, 15, 12, 15, 15, 15, 25, 20, 20, 10]
            for col_idx, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = width

            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename=Lotes_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            wb.save(response)
            return response
        except Exception as exc:
            import traceback
            logger.error(f"Error exportando Excel: {exc}\n{traceback.format_exc()}")
            return Response({'error': 'Error al exportar lotes', 'mensaje': str(exc)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['post'], url_path='importar-excel')
    def importar_excel(self, request):
        """
        Importa lotes desde Excel usando el importador estandarizado.
        
        Soporta múltiples formatos de columnas:
        - Clave Producto (OBLIGATORIO)
        - Nombre Producto (OBLIGATORIO - debe coincidir con clave)
        - Número Lote (OBLIGATORIO)
        - Cantidad Inicial (OBLIGATORIO) - unidades recibidas/surtidas
        - Fecha Caducidad (OBLIGATORIO)
        - Cantidad Contrato (opcional) - ISS-INV-001: total según contrato
        - Fecha Fabricación (opcional)
        - Precio Unitario (opcional, default 0)
        - Número Contrato (opcional)
        - Marca (opcional)
        - Ubicación (opcional)
        - Centro (opcional - nombre del centro)
        - Activo (opcional, default Activo)
        
        ISS-INV-001: ENTREGAS PARCIALES
        Si el contrato dice 100 pero llegaron 80:
        - Cantidad Inicial = 80 (lo que llegó)
        - Cantidad Contrato = 100 (lo contratado)
        - El sistema calculará Pendiente = 20
        
        REIMPORTACIÓN: Si reimporta con misma clave/lote/contrato/marca/fecha,
        el sistema SUMARÁ las cantidades (no duplica lotes).
        
        Detecta automáticamente la fila de encabezados.
        Soporta archivos con encabezados en fila 1, 2 o 3.
        
        Límites de seguridad:
        - Tamaño máximo: 10MB
        - Extensiones: .xlsx, .xls
        """
        from core.utils.excel_importer import importar_lotes_desde_excel, crear_log_importacion
        from django.core.cache import cache
        from core.models import Centro
        
        file = request.FILES.get('file')
        
        # Validar archivo
        es_valido, error_msg = validar_archivo_excel(file)
        if not es_valido:
            return Response({
                'error': 'Archivo inválido',
                'mensaje': error_msg
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            # Obtener centro del usuario si aplica
            centro_id = None
            if hasattr(request.user, 'centro') and request.user.centro:
                centro_id = request.user.centro.id
            
            # Ejecutar importación
            resultado = importar_lotes_desde_excel(file, request.user, centro_id=centro_id)
            
            # ISS-FIX: Invalidar caché del dashboard después de importar lotes
            cache.delete('dashboard_resumen_global')
            cache.delete('dashboard_graficas_global')
            for centro in Centro.objects.all():
                cache.delete(f'dashboard_resumen_{centro.id}')
                cache.delete(f'dashboard_graficas_{centro.id}')
            
            # Crear log de importación
            crear_log_importacion(
                usuario=request.user,
                tipo='Lote',
                archivo_nombre=file.name,
                resultado_dict=resultado
            )
            
            # Determinar status code
            if resultado['exitosa']:
                status_code = status.HTTP_200_OK
            elif resultado['registros_exitosos'] > 0:
                status_code = status.HTTP_206_PARTIAL_CONTENT
            else:
                status_code = status.HTTP_400_BAD_REQUEST
                # FIX: Agregar campo 'error' al nivel superior para que el frontend lo muestre
                errores = resultado.get('errores', [])
                if errores:
                    primer_error = errores[0].get('error', 'Error desconocido')
                    total_errores = len(errores)
                    resultado['error'] = (
                        f'{total_errores} error(es) al importar. '
                        f'Primer error (fila {errores[0].get("fila", "?")}): {primer_error}'
                    )
                else:
                    resultado['error'] = 'No se pudo importar ningún lote del archivo'
            
            return Response(resultado, status=status_code)
            
        except ValueError as e:
            # Errores de validación de negocio (ej. contrato global excedido)
            logger.warning(f"Importación rechazada por validación: {e}")
            return Response({
                'error': 'Importación rechazada',
                'mensaje': str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            logger.exception(f"Error en importación de lotes: {e}")
            return Response({
                'error': 'Error al procesar archivo',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path='plantilla')
    def plantilla_lotes(self, request):
        """
        Descarga plantilla Excel actualizada para importación de lotes.
        
        Usa el generador estandarizado con el esquema real de la base de datos.
        """
        # HALLAZGO #5: Manejo robusto de errores en generación de plantilla
        try:
            from core.utils.excel_templates import generar_plantilla_lotes
            
            # Obtener centro del usuario si aplica
            user = request.user
            centro = None
            if not is_farmacia_or_admin(user):
                centro = get_user_centro(user)
            
            return generar_plantilla_lotes(centro=centro)
        except ImportError as exc:
            logger.error(f'Error al importar generador de plantilla: {exc}')
            return Response(
                {'error': 'Módulo de generación de plantillas no disponible'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        except Exception as exc:
            logger.exception(f'Error al generar plantilla de lotes: {exc}')
            return Response(
                {'error': 'No se pudo generar la plantilla', 'mensaje': str(exc)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'])
    def por_vencer(self, request):
        """
        Obtiene lotes proximos a vencer.
        
        ISS-FIX SEGURIDAD: Usuarios de centro solo ven lotes de su centro.
        Admin/farmacia/vista ven todos los centros.
        
        Parametros:
        - dias: numero de dias (default: 30)
        """
        try:
            dias = int(request.query_params.get('dias', 30))
            fecha_limite = date.today() + timedelta(days=dias)
            
            lotes = Lote.objects.select_related('producto', 'centro').filter(
                activo=True,
                cantidad_actual__gt=0,
                fecha_caducidad__lte=fecha_limite
            )
            
            # ISS-FIX SEGURIDAD: Filtrar por centro según rol
            user = request.user
            if not has_global_read_access(user):
                user_centro = get_user_centro(user)
                if not user_centro:
                    logger.warning(
                        f"ISS-FIX: Usuario {user.username} sin centro intentó acceder a por_vencer"
                    )
                    return Response({
                        'total': 0,
                        'dias_configurados': dias,
                        'fecha_limite': fecha_limite,
                        'lotes': [],
                        'warning': 'No tiene un centro asignado'
                    })
                lotes = lotes.filter(centro=user_centro)
            
            lotes = lotes.order_by('fecha_caducidad')
            serializer = self.get_serializer(lotes, many=True)
            
            return Response({
                'total': lotes.count(),
                'dias_configurados': dias,
                'fecha_limite': fecha_limite,
                'lotes': serializer.data
            })
            
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            return Response({
                'error': 'Error al obtener lotes por vencer',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path='por-caducar')
    def por_caducar(self, request):
        """
        Alias compatible para el frontend: lotes proximos a vencer.
        
        ISS-FIX SEGURIDAD: Usuarios de centro solo ven lotes de su centro.
        Admin/farmacia/vista ven todos los centros.
        """
        try:
            dias = int(request.query_params.get('dias', 90))
            hoy = date.today()
            fecha_limite = hoy + timedelta(days=dias)
            lotes = Lote.objects.select_related('producto', 'centro').filter(
                cantidad_actual__gt=0,
                fecha_caducidad__gt=hoy,
                fecha_caducidad__lte=fecha_limite
            )
            
            # ISS-FIX SEGURIDAD: Filtrar por centro según rol
            user = request.user
            if not has_global_read_access(user):
                user_centro = get_user_centro(user)
                if not user_centro:
                    logger.warning(
                        f"ISS-FIX: Usuario {user.username} sin centro intentó acceder a por_caducar"
                    )
                    return Response([])  # Retornar lista vacía para compatibilidad
                lotes = lotes.filter(centro=user_centro)
            
            lotes = lotes.order_by('fecha_caducidad')

            page = self.paginate_queryset(lotes)
            if page is not None:
                serializer = self.get_serializer(page, many=True)
                return self.get_paginated_response(serializer.data)

            serializer = self.get_serializer(lotes, many=True)
            return Response(serializer.data)
        except Exception as exc:
            # traceback removido por seguridad (ISS-008)
            return Response({'error': 'Error al obtener lotes por caducar', 'mensaje': str(exc)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'])
    def vencidos(self, request):
        """
        Lotes con caducidad vencida y stock disponible.
        
        ISS-FIX SEGURIDAD: Usuarios de centro solo ven lotes de su centro.
        Admin/farmacia/vista ven todos los centros.
        """
        try:
            hoy = date.today()
            lotes = Lote.objects.select_related('producto', 'centro').filter(
                cantidad_actual__gt=0,
                fecha_caducidad__lt=hoy
            )
            
            # ISS-FIX SEGURIDAD: Filtrar por centro según rol
            user = request.user
            if not has_global_read_access(user):
                user_centro = get_user_centro(user)
                if not user_centro:
                    logger.warning(
                        f"ISS-FIX: Usuario {user.username} sin centro intentó acceder a vencidos"
                    )
                    return Response([])  # Retornar lista vacía para compatibilidad
                lotes = lotes.filter(centro=user_centro)
            
            lotes = lotes.order_by('fecha_caducidad')

            page = self.paginate_queryset(lotes)
            if page is not None:
                serializer = self.get_serializer(page, many=True)
                return self.get_paginated_response(serializer.data)

            serializer = self.get_serializer(lotes, many=True)
            return Response(serializer.data)
        except Exception as exc:
            # traceback removido por seguridad (ISS-008)
            return Response({'error': 'Error al obtener lotes vencidos', 'mensaje': str(exc)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=True, methods=['get'])
    def historial(self, request, pk=None):
        """Obtiene el historial de movimientos de un lote"""
        try:
            lote = self.get_object()
            
            movimientos = Movimiento.objects.filter(lote=lote).select_related(
                'lote__producto'
            ).order_by('-fecha')
            
            total_entradas = movimientos.filter(tipo='entrada').aggregate(
                total=Sum('cantidad')
            )['total'] or 0
            
            total_salidas = movimientos.filter(tipo='salida').aggregate(
                total=Sum('cantidad')
            )['total'] or 0
            
            movimientos_data = []
            for mov in movimientos:
                movimientos_data.append({
                    'id': mov.id,
                    'tipo': mov.tipo,
                    'cantidad': mov.cantidad,
                    'fecha': mov.fecha,
                    'observaciones': mov.observaciones or ''
                })
            
            return Response({
                'lote': {
                    'id': lote.id,
                    'numero_lote': lote.numero_lote,
                    'producto': lote.producto.clave,
                    'cantidad_actual': lote.cantidad_actual
                },
                'estadisticas': {
                    'total_entradas': total_entradas,
                    'total_salidas': total_salidas,
                    'diferencia': total_entradas - total_salidas
                },
                'movimientos': movimientos_data,
                'total_movimientos': len(movimientos_data)
            })
            
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            return Response({
                'error': 'Error al obtener historial',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['post'])
    def ajustar_stock(self, request, pk=None):
        """
        Ajusta stock del lote y crea movimiento asociado (entrada/salida/ajuste).
        """
        lote = self.get_object()
        tipo = request.data.get('tipo', 'ajuste')
        cantidad = request.data.get('cantidad')
        observaciones = request.data.get('observaciones', '')

        try:
            movimiento, lote_actualizado = registrar_movimiento_stock(
                lote=lote,
                tipo=tipo,
                cantidad=cantidad,
                usuario=request.user if request.user.is_authenticated else None,
                centro=None,
                requisicion=None,
                observaciones=observaciones
            )
            response_data = {
                'mensaje': 'Stock ajustado correctamente',
                'lote': self.get_serializer(lote_actualizado).data,
                'movimiento_id': movimiento.id
            }
            
            return Response(response_data)
        except serializers.ValidationError as exc:
            return Response({'error': 'Error de validacion', 'detalles': exc.detail}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as exc:
            # traceback removido por seguridad (ISS-008)
            return Response({'error': 'Error al ajustar stock', 'mensaje': str(exc)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['get'], url_path='lotes-derivados')
    def lotes_derivados(self, request, pk=None):
        """
        Obtiene los lotes derivados de un lote de farmacia (vinculados a centros).
        
        Solo aplica para lotes de farmacia central (centro=NULL).
        Muestra todos los centros que tienen stock de este lote.
        """
        try:
            lote = self.get_object()
            
            # Verificar que es un lote de farmacia
            if lote.centro is not None:
                return Response({
                    'error': 'Solo los lotes de farmacia central tienen lotes derivados',
                    'lote_id': lote.id,
                    'centro': lote.centro.nombre if lote.centro else None
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Nota: En Supabase no hay lote_origen, se simplifica
            # Los lotes derivados se manejan diferente
            derivados = Lote.objects.none()
            
            # Calcular totales
            total_derivados = 0
            stock_total_centros = 0
            
            derivados_data = []
            # Código original removido - lote_origen no existe en Supabase
            
            return Response({
                'lote_farmacia': {
                    'id': lote.id,
                    'numero_lote': lote.numero_lote,
                    'producto_clave': lote.producto.clave,
                    'producto_nombre': lote.producto.nombre,
                    'cantidad_actual': lote.cantidad_actual,
                    'fecha_caducidad': lote.fecha_caducidad
                },
                'resumen': {
                    'total_centros_con_stock': total_derivados,
                    'stock_total_en_centros': stock_total_centros,
                    'stock_farmacia': lote.cantidad_actual,
                    'stock_total_sistema': lote.cantidad_actual + stock_total_centros
                },
                'lotes_derivados': derivados_data
            })
            
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            return Response({
                'error': 'Error al obtener lotes derivados',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['get'], url_path='trazabilidad')
    def trazabilidad_lote(self, request, pk=None):
        """
        Obtiene la trazabilidad completa de un lote:
        - Si es lote de farmacia: muestra derivados en centros
        - Si es lote de centro: muestra origen en farmacia
        """
        try:
            lote = self.get_object()
            
            result = {
                'lote': {
                    'id': lote.id,
                    'numero_lote': lote.numero_lote,
                    'producto_clave': lote.producto.clave,
                    'producto_descripcion': lote.producto.descripcion,
                    'cantidad_actual': lote.cantidad_actual,
                    'fecha_caducidad': lote.fecha_caducidad,
                    'es_lote_farmacia': lote.centro is None,
                    'ubicacion': lote.centro.nombre if lote.centro else 'Farmacia Central'
                },
                'origen': None,
                'derivados': []
            }
            
            # En Supabase no hay lote_origen - trazabilidad simplificada
            # Los lotes de cada centro son independientes
            
            # Movimientos relacionados
            movimientos = Movimiento.objects.filter(
                lote=lote
            ).select_related('requisicion', 'usuario', 'centro_origen', 'centro_destino').order_by('-fecha')[:20]
            
            result['movimientos'] = [{
                'id': m.id,
                'tipo': m.tipo,
                'cantidad': m.cantidad,
                'fecha': m.fecha,
                'requisicion_folio': m.requisicion.folio if m.requisicion else None,
                'observaciones': m.observaciones
            } for m in movimientos]
            
            return Response(result)
            
        except Exception as e:
            # traceback removido por seguridad (ISS-008)
            return Response({
                'error': 'Error al obtener trazabilidad',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    # =========================================================================
    # ACCIONES DE DOCUMENTOS (facturas, contratos, remisiones)
    # =========================================================================
    
    @action(detail=True, methods=['get'], url_path='documentos')
    def listar_documentos(self, request, pk=None):
        """
        Lista todos los documentos asociados a un lote.
        """
        try:
            lote = self.get_object()
            documentos = LoteDocumento.objects.filter(lote=lote).order_by('-created_at')
            serializer = LoteDocumentoSerializer(documentos, many=True)
            return Response({
                'lote_id': lote.id,
                'numero_lote': lote.numero_lote,
                'total_documentos': documentos.count(),
                'documentos': serializer.data
            })
        except Exception as e:
            return Response({
                'error': 'Error al obtener documentos',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['post'], url_path='subir-documento')
    def subir_documento(self, request, pk=None):
        """
        Sube un documento (PDF) asociado al lote.
        
        Campos requeridos:
        - documento: archivo PDF (multipart)
        - tipo_documento: factura/contrato/remision/otro
        
        Campos opcionales:
        - numero_documento: número del documento
        - fecha_documento: fecha del documento (YYYY-MM-DD)
        - notas: notas adicionales
        """
        try:
            lote = self.get_object()
            
            # Validar permisos de escritura
            user = request.user
            if not is_farmacia_or_admin(user):
                user_centro = get_user_centro(user)
                lote_centro = lote.centro
                if lote_centro is None or (user_centro and lote_centro.pk != user_centro.pk):
                    return Response(
                        {'error': 'No tiene permisos para subir documentos a este lote'},
                        status=status.HTTP_403_FORBIDDEN
                    )
            
            # ISS-005 FIX (audit7): Validar archivo PDF con función centralizada
            archivo = request.FILES.get('documento')
            if not archivo:
                return Response(
                    {'error': 'Debe proporcionar un archivo'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Usar validador centralizado que verifica extensión, tamaño Y magic bytes
            es_valido, error_msg = validar_archivo_pdf(archivo)
            if not es_valido:
                return Response(
                    {'error': error_msg},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Validar tipo de documento
            tipo_documento = request.data.get('tipo_documento', 'otro')
            tipos_validos = ['factura', 'contrato', 'remision', 'otro']
            if tipo_documento not in tipos_validos:
                return Response(
                    {'error': f'Tipo de documento inválido. Valores permitidos: {tipos_validos}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Generar path único para el archivo
            timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
            unique_name = f"{tipo_documento}_{timestamp}_{uuid.uuid4().hex[:8]}.pdf"
            archivo_path = f"lotes/documentos/{lote.id}/{unique_name}"
            
            # Guardar nombre del archivo original
            nombre_archivo = getattr(archivo, 'name', unique_name)
            
            # ISS-001 FIX: Subir archivo al almacenamiento ANTES de crear registro
            from inventario.services.storage_service import get_storage_service, StorageError
            
            storage = get_storage_service()
            upload_result = storage.upload_file(
                file_content=archivo,
                file_path=archivo_path,
                content_type='application/pdf',
                metadata={
                    'lote_id': lote.id,
                    'tipo_documento': tipo_documento,
                    'uploaded_by': user.username if user.is_authenticated else 'anonymous'
                }
            )
            
            # ISS-001 FIX: Si falla la subida, NO crear registro (rollback)
            if not upload_result.get('success'):
                logger.error(
                    f"ISS-001: Fallo subida documento lote {lote.id}: {upload_result.get('error')}"
                )
                return Response({
                    'error': 'Error al guardar archivo en almacenamiento',
                    'detalle': upload_result.get('error', 'Error desconocido')
                }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
            # Parsear fecha si viene
            fecha_documento = None
            if request.data.get('fecha_documento'):
                try:
                    fecha_documento = datetime.strptime(
                        request.data.get('fecha_documento'), '%Y-%m-%d'
                    ).date()
                except ValueError:
                    pass
            
            # ISS-001 FIX: Solo crear registro si la subida fue exitosa
            try:
                documento = LoteDocumento.objects.create(
                    lote=lote,
                    tipo_documento=tipo_documento,
                    numero_documento=request.data.get('numero_documento', ''),
                    archivo=archivo_path,
                    nombre_archivo=nombre_archivo,
                    fecha_documento=fecha_documento,
                    notas=request.data.get('notas', ''),
                    created_by=user if user.is_authenticated else None
                )
            except Exception as db_error:
                # ISS-001 FIX: Si falla crear registro, eliminar archivo subido (rollback)
                logger.error(f"ISS-001: Error BD, revirtiendo subida: {db_error}")
                storage.delete_file(archivo_path)
                raise
            
            logger.info(
                f"ISS-001: Documento subido exitosamente - Lote: {lote.id}, "
                f"Path: {archivo_path}, Storage: {upload_result.get('storage')}"
            )
            
            serializer = LoteDocumentoSerializer(documento)
            return Response({
                'mensaje': 'Documento subido correctamente',
                'documento': serializer.data,
                'storage_info': {
                    'path': archivo_path,
                    'url': upload_result.get('url'),
                    'storage_type': upload_result.get('storage')
                }
            }, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            return Response({
                'error': 'Error al subir documento',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['delete'], url_path='eliminar-documento/(?P<doc_id>[0-9]+)')
    def eliminar_documento(self, request, pk=None, doc_id=None):
        """
        Elimina un documento específico del lote.
        """
        try:
            lote = self.get_object()
            
            # Validar permisos de escritura
            user = request.user
            if not is_farmacia_or_admin(user):
                user_centro = get_user_centro(user)
                lote_centro = lote.centro
                if lote_centro is None or (user_centro and lote_centro.pk != user_centro.pk):
                    return Response(
                        {'error': 'No tiene permisos para eliminar documentos de este lote'},
                        status=status.HTTP_403_FORBIDDEN
                    )
            
            # Buscar documento
            try:
                documento = LoteDocumento.objects.get(id=doc_id, lote=lote)
            except LoteDocumento.DoesNotExist:
                return Response(
                    {'error': 'Documento no encontrado'},
                    status=status.HTTP_404_NOT_FOUND
                )
            
            # ISS-002 FIX: Eliminar archivo del almacenamiento ANTES de borrar registro
            from inventario.services.storage_service import get_storage_service
            
            archivo_path = documento.archivo
            nombre = documento.nombre_archivo
            
            storage = get_storage_service()
            delete_result = storage.delete_file(archivo_path)
            
            # ISS-002 FIX: Registrar resultado de eliminación de storage
            if not delete_result.get('success'):
                logger.warning(
                    f"ISS-002: No se pudo eliminar archivo '{archivo_path}' del storage: "
                    f"{delete_result.get('error')}. Se eliminará el registro de BD igualmente."
                )
            else:
                logger.info(
                    f"ISS-002: Archivo eliminado de storage: {archivo_path}"
                )
            
            # Eliminar registro de BD
            documento.delete()
            
            logger.info(
                f"ISS-002: Documento eliminado - Lote: {lote.id}, "
                f"Archivo: {nombre}, Path: {archivo_path}"
            )
            
            return Response({
                'mensaje': 'Documento eliminado correctamente',
                'documento_eliminado': nombre,
                'storage_cleanup': delete_result.get('success', False)
            })
            
        except Exception as e:
            return Response({
                'error': 'Error al eliminar documento',
                'mensaje': str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    # =========================================================================
    # PARCIALIDADES - Historial de entregas parciales
    # =========================================================================
    
    @action(detail=True, methods=['get'], url_path='parcialidades')
    def listar_parcialidades(self, request, pk=None):
        """
        Lista el historial de entregas parciales del lote.
        
        Estados:
        - PENDIENTE: total_entregado = 0
        - PARCIAL: 0 < total_entregado < cantidad_contrato
        - CUMPLIDO: total_entregado == cantidad_contrato
        - SOBREENTREGA: total_entregado > cantidad_contrato
        - SIN_CONTRATO: cantidad_contrato no definida
        
        Retorna:
        - Lista de parcialidades ordenadas por fecha (más reciente primero)
        - Total acumulado
        - Comparación con contratos (lote y global)
        
        AUTO-SINCRONIZACIÓN:
        - Si fecha_fabricacion es NULL y hay parcialidades, se sincroniza automáticamente
        """
        lote = self.get_object()
        
        parcialidades = LoteParcialidad.objects.filter(lote=lote).order_by('-fecha_entrega')
        serializer = LoteParcialidadSerializer(parcialidades, many=True)
        
        # Calcular totales (una sola query)
        from django.db.models import Sum, Count, Min, Max
        stats = parcialidades.aggregate(
            total_cantidad=Sum('cantidad'),
            num_entregas=Count('id'),
            primera_entrega=Min('fecha_entrega'),
            ultima_entrega=Max('fecha_entrega')
        )
        
        # =================================================================
        # AUTO-SINCRONIZACIÓN DE FECHA_FABRICACION
        # Si el lote no tiene fecha pero hay parcialidades, sincronizar
        # =================================================================
        if lote.fecha_fabricacion is None and stats['primera_entrega']:
            try:
                lote.fecha_fabricacion = stats['primera_entrega']
                lote.save(update_fields=['fecha_fabricacion'])
                logger.info(f"[AUTO-SYNC] Lote {lote.numero_lote}: fecha_fabricacion = {stats['primera_entrega']}")
            except Exception as e:
                logger.warning(f"[AUTO-SYNC] Error sincronizando fecha para lote {lote.pk}: {e}")
        
        # Calcular estado del contrato de lote
        total_parcialidades = stats['total_cantidad'] or 0
        cantidad_contrato_lote = lote.cantidad_contrato or 0
        
        # Usar función centralizada para calcular estado
        estado_lote = self._calcular_estado_contrato(total_parcialidades, cantidad_contrato_lote)
        
        pendiente_lote = max(0, cantidad_contrato_lote - total_parcialidades) if cantidad_contrato_lote > 0 else None
        porcentaje_lote = min(100, (total_parcialidades / cantidad_contrato_lote) * 100) if cantidad_contrato_lote > 0 else 0
        excedente_lote = max(0, total_parcialidades - cantidad_contrato_lote) if cantidad_contrato_lote > 0 and total_parcialidades > cantidad_contrato_lote else None
        
        # Información del contrato global (usar función centralizada)
        contrato_global_info = self._calcular_estado_contrato_global(lote)
        
        # Fecha de entrega a devolver (priorizar fecha del lote, fallback a primera parcialidad)
        fecha_entrega_display = lote.fecha_fabricacion or stats['primera_entrega']
        
        return Response({
            'parcialidades': serializer.data,
            'resumen': {
                'total_cantidad': total_parcialidades,
                'num_entregas': stats['num_entregas'] or 0,
                'primera_entrega': stats['primera_entrega'],
                'ultima_entrega': stats['ultima_entrega'],
                'fecha_entrega': fecha_entrega_display,  # Campo explícito para el frontend
            },
            'contrato_lote': {
                'cantidad_contrato': cantidad_contrato_lote or None,
                'total_entregado': total_parcialidades,
                'pendiente': pendiente_lote,
                'porcentaje': round(porcentaje_lote, 2),
                'estado': estado_lote,
                'excedente': excedente_lote,
            },
            'contrato_global': contrato_global_info,
            # Información adicional del lote para el modal
            'lote_info': {
                'fecha_fabricacion': lote.fecha_fabricacion,
                'numero_contrato': lote.numero_contrato,
                'cantidad_contrato_global': lote.cantidad_contrato_global,
            },
        })

    @action(detail=True, methods=['get'], url_path='exportar-entregas-pdf')
    def exportar_entregas_pdf(self, request, pk=None):
        """
        Exporta el historial de entregas del lote a PDF con diseño profesional.
        Incluye información del producto, lote, contrato y todas las parcialidades.
        """
        lote = self.get_object()
        parcialidades = LoteParcialidad.objects.filter(lote=lote).order_by('-fecha_entrega')
        
        # Calcular totales
        from django.db.models import Sum, Count, Min, Max
        stats = parcialidades.aggregate(
            total_cantidad=Sum('cantidad'),
            num_entregas=Count('id'),
            primera_entrega=Min('fecha_entrega'),
            ultima_entrega=Max('fecha_entrega')
        )
        
        # Info del contrato global
        contrato_global_info = self._calcular_estado_contrato_global(lote)
        
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch, cm
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
            from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
            
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter, 
                                   leftMargin=0.5*inch, rightMargin=0.5*inch,
                                   topMargin=0.5*inch, bottomMargin=0.5*inch)
            elements = []
            styles = getSampleStyleSheet()
            
            # Estilos personalizados
            titulo_style = ParagraphStyle('TituloReporte', parent=styles['Heading1'],
                                         fontSize=16, textColor=colors.HexColor('#9F2241'),
                                         alignment=TA_CENTER, spaceAfter=12)
            subtitulo_style = ParagraphStyle('Subtitulo', parent=styles['Heading2'],
                                            fontSize=12, textColor=colors.HexColor('#235B4E'),
                                            alignment=TA_CENTER, spaceAfter=6)
            normal_style = ParagraphStyle('Normal', parent=styles['Normal'],
                                         fontSize=10, spaceAfter=4)
            
            # Título
            elements.append(Paragraph("HISTORIAL DE ENTREGAS", titulo_style))
            elements.append(Paragraph(f"Lote: {lote.numero_lote}", subtitulo_style))
            elements.append(Spacer(1, 0.2*inch))
            
            # Información del producto
            producto = lote.producto
            info_producto = [
                ['INFORMACIÓN DEL PRODUCTO', '', '', ''],
                ['Clave:', producto.clave if producto else 'N/A', 
                 'Nombre:', producto.nombre if producto else 'N/A'],
                ['Número de Lote:', lote.numero_lote, 
                 'Centro:', lote.centro.nombre if lote.centro else 'Farmacia Central'],
                ['Fecha Entrega:', lote.fecha_fabricacion.strftime('%d/%m/%Y') if lote.fecha_fabricacion else 'No registrada',
                 'Fecha Caducidad:', lote.fecha_caducidad.strftime('%d/%m/%Y') if lote.fecha_caducidad else 'N/A'],
                ['Número Contrato:', lote.numero_contrato or 'Sin contrato', 
                 'Marca:', lote.marca or 'N/A'],
            ]
            
            t_producto = Table(info_producto, colWidths=[1.5*inch, 2*inch, 1.5*inch, 2.5*inch])
            t_producto.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#9F2241')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('SPAN', (0, 0), (-1, 0)),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#F5F5F5')),
                ('BACKGROUND', (2, 1), (2, -1), colors.HexColor('#F5F5F5')),
                ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
                ('FONTNAME', (2, 1), (2, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            elements.append(t_producto)
            elements.append(Spacer(1, 0.2*inch))
            
            # Resumen del contrato
            total_parcialidades = stats['total_cantidad'] or 0
            cantidad_contrato_lote = lote.cantidad_contrato or 0
            
            resumen_data = [
                ['RESUMEN DE ENTREGAS', '', '', ''],
                ['Total Entregas:', str(stats['num_entregas'] or 0),
                 'Total Entregado:', f"{total_parcialidades:,} uds"],
                ['Primera Entrega:', stats['primera_entrega'].strftime('%d/%m/%Y') if stats['primera_entrega'] else 'N/A',
                 'Última Entrega:', stats['ultima_entrega'].strftime('%d/%m/%Y') if stats['ultima_entrega'] else 'N/A'],
            ]
            
            # Info contrato lote
            if cantidad_contrato_lote > 0:
                pendiente_lote = max(0, cantidad_contrato_lote - total_parcialidades)
                porcentaje_lote = min(100, (total_parcialidades / cantidad_contrato_lote) * 100)
                resumen_data.append(['Contrato Lote:', f"{cantidad_contrato_lote:,} uds",
                                    'Pendiente Lote:', f"{pendiente_lote:,} uds ({100-porcentaje_lote:.1f}%)"])
            
            # Info contrato global
            if contrato_global_info and contrato_global_info.get('cantidad_contrato_global'):
                ccg = contrato_global_info['cantidad_contrato_global']
                te = contrato_global_info['total_entregado']
                pend = contrato_global_info['pendiente']
                porc = contrato_global_info['porcentaje']
                resumen_data.append(['Contrato Global:', f"{ccg:,} uds", 
                                    'Entregado Global:', f"{te:,} uds ({porc:.1f}%)"])
                resumen_data.append(['Pendiente Global:', f"{pend:,} uds",
                                    'Estado:', contrato_global_info['estado']])
            
            t_resumen = Table(resumen_data, colWidths=[1.5*inch, 2*inch, 1.5*inch, 2.5*inch])
            t_resumen.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#235B4E')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('SPAN', (0, 0), (-1, 0)),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#F5F5F5')),
                ('BACKGROUND', (2, 1), (2, -1), colors.HexColor('#F5F5F5')),
                ('FONTNAME', (0, 1), (0, -1), 'Helvetica-Bold'),
                ('FONTNAME', (2, 1), (2, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ]))
            elements.append(t_resumen)
            elements.append(Spacer(1, 0.3*inch))
            
            # Tabla de parcialidades (detalle de entregas)
            elements.append(Paragraph("DETALLE DE ENTREGAS", titulo_style))
            elements.append(Spacer(1, 0.1*inch))
            
            if parcialidades.exists():
                # Encabezados
                tabla_data = [['#', 'Fecha', 'Cantidad', 'Factura', 'Remisión', 'Proveedor', 'Notas']]
                
                # Filas de datos
                acumulado = 0
                for i, p in enumerate(parcialidades.order_by('fecha_entrega'), 1):
                    acumulado += p.cantidad
                    tabla_data.append([
                        str(i),
                        p.fecha_entrega.strftime('%d/%m/%Y') if p.fecha_entrega else '',
                        f"{p.cantidad:,}",
                        p.numero_factura or '-',
                        p.numero_remision or '-',
                        p.proveedor or '-',
                        (p.notas[:30] + '...') if p.notas and len(p.notas) > 30 else (p.notas or '-'),
                    ])
                
                # Fila de total
                tabla_data.append(['', 'TOTAL', f"{total_parcialidades:,}", '', '', '', ''])
                
                t_parcialidades = Table(tabla_data, colWidths=[0.4*inch, 0.9*inch, 0.8*inch, 1*inch, 1*inch, 1.2*inch, 2.2*inch])
                t_parcialidades.setStyle(TableStyle([
                    # Encabezado
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#9F2241')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    # Filas de datos
                    ('FONTSIZE', (0, 1), (-1, -2), 8),
                    ('ALIGN', (0, 1), (0, -1), 'CENTER'),
                    ('ALIGN', (2, 1), (2, -1), 'RIGHT'),
                    # Fila total
                    ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#235B4E')),
                    ('TEXTCOLOR', (0, -1), (-1, -1), colors.white),
                    ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                    # Bordes
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
                    ('TOPPADDING', (0, 0), (-1, -1), 4),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                    # Alternar colores de fila
                    ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F9F9F9')]),
                ]))
                elements.append(t_parcialidades)
            else:
                elements.append(Paragraph("No hay entregas registradas para este lote.", normal_style))
            
            # Pie de página
            elements.append(Spacer(1, 0.5*inch))
            fecha_gen = timezone.now().strftime('%d/%m/%Y %H:%M')
            elements.append(Paragraph(f"Generado: {fecha_gen} | Usuario: {request.user.username}", 
                                     ParagraphStyle('Pie', parent=styles['Normal'], 
                                                   fontSize=8, textColor=colors.gray,
                                                   alignment=TA_RIGHT)))
            
            doc.build(elements)
            buffer.seek(0)
            
            response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="Entregas_{lote.numero_lote}_{timezone.now().strftime("%Y%m%d")}.pdf"'
            return response
            
        except Exception as e:
            logger.exception(f"Error generando PDF de entregas para lote {pk}: {e}")
            return Response({'error': f'Error generando PDF: {str(e)}'}, 
                          status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['get'], url_path='exportar-entregas-excel')
    def exportar_entregas_excel(self, request, pk=None):
        """
        Exporta el historial de entregas del lote a Excel con formato profesional.
        """
        lote = self.get_object()
        parcialidades = LoteParcialidad.objects.filter(lote=lote).order_by('fecha_entrega')
        
        # Calcular totales
        from django.db.models import Sum, Count, Min, Max
        stats = parcialidades.aggregate(
            total_cantidad=Sum('cantidad'),
            num_entregas=Count('id'),
            primera_entrega=Min('fecha_entrega'),
            ultima_entrega=Max('fecha_entrega')
        )
        
        # Info contrato global
        contrato_global_info = self._calcular_estado_contrato_global(lote)
        
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Historial Entregas"
            
            # Estilos
            header_fill = PatternFill(start_color="9F2241", end_color="9F2241", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            subheader_fill = PatternFill(start_color="235B4E", end_color="235B4E", fill_type="solid")
            subheader_font = Font(bold=True, color="FFFFFF", size=10)
            label_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
            label_font = Font(bold=True, size=9)
            total_fill = PatternFill(start_color="235B4E", end_color="235B4E", fill_type="solid")
            total_font = Font(bold=True, color="FFFFFF")
            
            row = 1
            
            # Título
            ws.merge_cells(f'A{row}:G{row}')
            ws[f'A{row}'] = "HISTORIAL DE ENTREGAS"
            ws[f'A{row}'].fill = header_fill
            ws[f'A{row}'].font = Font(bold=True, color="FFFFFF", size=14)
            ws[f'A{row}'].alignment = Alignment(horizontal='center')
            row += 1
            
            ws.merge_cells(f'A{row}:G{row}')
            ws[f'A{row}'] = f"Lote: {lote.numero_lote}"
            ws[f'A{row}'].alignment = Alignment(horizontal='center')
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 2
            
            # Información del producto
            producto = lote.producto
            ws.merge_cells(f'A{row}:G{row}')
            ws[f'A{row}'] = "INFORMACIÓN DEL PRODUCTO"
            ws[f'A{row}'].fill = subheader_fill
            ws[f'A{row}'].font = subheader_font
            row += 1
            
            info_rows = [
                ('Clave:', producto.clave if producto else 'N/A', 'Nombre:', producto.nombre if producto else 'N/A'),
                ('Número de Lote:', lote.numero_lote, 'Centro:', lote.centro.nombre if lote.centro else 'Farmacia Central'),
                ('Fecha Entrega:', lote.fecha_fabricacion.strftime('%d/%m/%Y') if lote.fecha_fabricacion else 'No registrada',
                 'Fecha Caducidad:', lote.fecha_caducidad.strftime('%d/%m/%Y') if lote.fecha_caducidad else 'N/A'),
                ('Número Contrato:', lote.numero_contrato or 'Sin contrato', 'Marca:', lote.marca or 'N/A'),
            ]
            
            for label1, val1, label2, val2 in info_rows:
                ws[f'A{row}'] = label1
                ws[f'A{row}'].fill = label_fill
                ws[f'A{row}'].font = label_font
                ws[f'B{row}'] = val1
                ws[f'C{row}'] = label2
                ws[f'C{row}'].fill = label_fill
                ws[f'C{row}'].font = label_font
                ws[f'D{row}'] = val2
                row += 1
            
            row += 1
            
            # Resumen
            ws.merge_cells(f'A{row}:G{row}')
            ws[f'A{row}'] = "RESUMEN DE ENTREGAS"
            ws[f'A{row}'].fill = subheader_fill
            ws[f'A{row}'].font = subheader_font
            row += 1
            
            total_parcialidades = stats['total_cantidad'] or 0
            cantidad_contrato_lote = lote.cantidad_contrato or 0
            
            resumen_rows = [
                ('Total Entregas:', str(stats['num_entregas'] or 0), 'Total Entregado:', f"{total_parcialidades:,} uds"),
                ('Primera Entrega:', stats['primera_entrega'].strftime('%d/%m/%Y') if stats['primera_entrega'] else 'N/A',
                 'Última Entrega:', stats['ultima_entrega'].strftime('%d/%m/%Y') if stats['ultima_entrega'] else 'N/A'),
            ]
            
            if cantidad_contrato_lote > 0:
                pendiente_lote = max(0, cantidad_contrato_lote - total_parcialidades)
                porcentaje_lote = min(100, (total_parcialidades / cantidad_contrato_lote) * 100)
                resumen_rows.append(('Contrato Lote:', f"{cantidad_contrato_lote:,} uds",
                                    'Pendiente Lote:', f"{pendiente_lote:,} uds ({100-porcentaje_lote:.1f}%)"))
            
            if contrato_global_info and contrato_global_info.get('cantidad_contrato_global'):
                ccg = contrato_global_info['cantidad_contrato_global']
                te = contrato_global_info['total_entregado']
                pend = contrato_global_info['pendiente']
                porc = contrato_global_info['porcentaje']
                resumen_rows.append(('Contrato Global:', f"{ccg:,} uds", 
                                    'Entregado Global:', f"{te:,} uds ({porc:.1f}%)"))
                resumen_rows.append(('Pendiente Global:', f"{pend:,} uds",
                                    'Estado:', contrato_global_info['estado']))
            
            for label1, val1, label2, val2 in resumen_rows:
                ws[f'A{row}'] = label1
                ws[f'A{row}'].fill = label_fill
                ws[f'A{row}'].font = label_font
                ws[f'B{row}'] = val1
                ws[f'C{row}'] = label2
                ws[f'C{row}'].fill = label_fill
                ws[f'C{row}'].font = label_font
                ws[f'D{row}'] = val2
                row += 1
            
            row += 1
            
            # Tabla de parcialidades
            ws.merge_cells(f'A{row}:G{row}')
            ws[f'A{row}'] = "DETALLE DE ENTREGAS"
            ws[f'A{row}'].fill = header_fill
            ws[f'A{row}'].font = header_font
            ws[f'A{row}'].alignment = Alignment(horizontal='center')
            row += 1
            
            # Encabezados de tabla
            headers = ['#', 'Fecha', 'Cantidad', 'Factura', 'Remisión', 'Proveedor', 'Notas']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            row += 1
            
            # Datos
            if parcialidades.exists():
                for i, p in enumerate(parcialidades, 1):
                    ws.cell(row=row, column=1, value=i)
                    ws.cell(row=row, column=2, value=p.fecha_entrega.strftime('%d/%m/%Y') if p.fecha_entrega else '')
                    ws.cell(row=row, column=3, value=p.cantidad)
                    ws.cell(row=row, column=4, value=p.numero_factura or '-')
                    ws.cell(row=row, column=5, value=p.numero_remision or '-')
                    ws.cell(row=row, column=6, value=p.proveedor or '-')
                    ws.cell(row=row, column=7, value=p.notas or '-')
                    row += 1
                
                # Fila total
                ws.cell(row=row, column=1, value='')
                ws.cell(row=row, column=2, value='TOTAL')
                ws.cell(row=row, column=2).font = total_font
                ws.cell(row=row, column=2).fill = total_fill
                ws.cell(row=row, column=3, value=total_parcialidades)
                ws.cell(row=row, column=3).font = total_font
                ws.cell(row=row, column=3).fill = total_fill
                for col in [1, 4, 5, 6, 7]:
                    ws.cell(row=row, column=col).fill = total_fill
            else:
                ws.merge_cells(f'A{row}:G{row}')
                ws[f'A{row}'] = "No hay entregas registradas"
                ws[f'A{row}'].alignment = Alignment(horizontal='center')
            
            # Ajustar anchos de columna
            ws.column_dimensions['A'].width = 6
            ws.column_dimensions['B'].width = 14
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 18
            ws.column_dimensions['G'].width = 35
            
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            response = HttpResponse(buffer.getvalue(), 
                                   content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="Entregas_{lote.numero_lote}_{timezone.now().strftime("%Y%m%d")}.xlsx"'
            return response
            
        except Exception as e:
            logger.exception(f"Error generando Excel de entregas para lote {pk}: {e}")
            return Response({'error': f'Error generando Excel: {str(e)}'}, 
                          status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=True, methods=['post'], url_path='agregar-parcialidad')
    def agregar_parcialidad(self, request, pk=None):
        """
        Registra una nueva entrega parcial para el lote.
        
        FLUJO ROBUSTO DE PARCIALIDADES:
        1. Validar permisos de escritura
        2. Verificar estado actual del contrato (ANTES de agregar)
        3. Si CUMPLIDO/SOBREENTREGA: requerir override (admin/supervisor + motivo)
        4. Crear parcialidad con transaction.atomic + select_for_update
        5. Recalcular estados y retornar información completa
        
        Estados posibles:
        - PENDIENTE: total_entregado = 0
        - PARCIAL: 0 < total_entregado < cantidad_contrato
        - CUMPLIDO: total_entregado == cantidad_contrato
        - SOBREENTREGA: total_entregado > cantidad_contrato
        - SIN_CONTRATO: cantidad_contrato no definida
        
        Campos esperados:
        - fecha_entrega: date (obligatorio)
        - cantidad: integer > 0 (obligatorio)
        - numero_factura: string (opcional)
        - numero_remision: string (opcional)
        - proveedor: string (opcional)
        - notas: string (opcional)
        - override: boolean (requerido si contrato ya cumplido)
        - motivo_override: string (requerido si override=true)
        """
        lote = self.get_object()
        user = request.user
        
        # Validar permisos de escritura
        if not is_farmacia_or_admin(user):
            user_centro = get_user_centro(user)
            lote_centro = lote.centro
            if lote_centro is None or (user_centro and lote_centro.pk != user_centro.pk):
                return Response(
                    {'error': 'No tiene permisos para registrar entregas de este lote'},
                    status=status.HTTP_403_FORBIDDEN
                )
        
        # Usar transacción atómica con select_for_update para evitar condiciones de carrera
        with transaction.atomic():
            # Bloquear el lote para evitar actualizaciones concurrentes
            lote_locked = Lote.objects.select_for_update().get(pk=lote.pk)
            
            # Calcular estado ACTUAL del contrato (ANTES de agregar)
            total_actual = LoteParcialidad.objects.filter(lote=lote_locked).aggregate(
                total=Sum('cantidad')
            )['total'] or 0
            
            cantidad_contrato = lote_locked.cantidad_contrato or 0
            estado_previo = self._calcular_estado_contrato(total_actual, cantidad_contrato)
            
            # VALIDACIÓN DE SOBRE-ENTREGA
            # Si el contrato ya está CUMPLIDO o en SOBREENTREGA, requerir override
            if estado_previo in ['CUMPLIDO', 'SOBREENTREGA'] and cantidad_contrato > 0:
                override = request.data.get('override', False)
                motivo_override = request.data.get('motivo_override', '').strip()
                
                # Solo admin/farmacia pueden hacer override
                puede_override = is_farmacia_or_admin(user)
                
                if not override:
                    return Response({
                        'error': 'Contrato ya cumplido',
                        'estado_actual': estado_previo,
                        'total_entregado': total_actual,
                        'cantidad_contrato': cantidad_contrato,
                        'requiere_override': True,
                        'mensaje': (
                            f'El contrato del lote ya está {estado_previo}. '
                            f'Para agregar más entregas, active "override" y proporcione un motivo.'
                        )
                    }, status=status.HTTP_409_CONFLICT)
                
                if not puede_override:
                    return Response({
                        'error': 'Sin permisos para override',
                        'mensaje': 'Solo administradores o supervisores pueden agregar entregas a contratos cumplidos.'
                    }, status=status.HTTP_403_FORBIDDEN)
                
                if not motivo_override or len(motivo_override) < 10:
                    return Response({
                        'error': 'Motivo de override requerido',
                        'mensaje': 'Debe proporcionar un motivo detallado (mínimo 10 caracteres) para sobre-entrega.'
                    }, status=status.HTTP_400_BAD_REQUEST)
            
            # Crear o MERGE parcialidad (evita duplicados de entregas equivalentes)
            from core.utils.parcialidad_merge import merge_or_create_parcialidad
            
            # Extraer datos del request
            fecha_entrega = request.data.get('fecha_entrega')
            cantidad = request.data.get('cantidad')
            
            # Validar campos obligatorios
            if not fecha_entrega:
                return Response({
                    'error': 'fecha_entrega es requerida',
                }, status=status.HTTP_400_BAD_REQUEST)
            
            try:
                cantidad = int(cantidad)
                if cantidad <= 0:
                    raise ValueError()
            except (TypeError, ValueError):
                return Response({
                    'error': 'cantidad debe ser un número entero positivo',
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Datos opcionales
            factura = request.data.get('numero_factura') or None
            numero_remision = request.data.get('numero_remision') or None
            proveedor = request.data.get('proveedor') or None
            notas = request.data.get('notas') or None
            
            # Determinar si es sobreentrega y si hay override
            es_sobreentrega = estado_previo in ['CUMPLIDO', 'SOBREENTREGA'] and request.data.get('override', False)
            motivo_override = request.data.get('motivo_override', '').strip() if es_sobreentrega else None
            
            # Ejecutar MERGE o CREATE
            resultado_merge = merge_or_create_parcialidad(
                lote=lote_locked,
                fecha_entrega=fecha_entrega,
                cantidad=cantidad,
                usuario=user,
                factura=factura,
                proveedor=proveedor,
                numero_remision=numero_remision,
                notas=notas,
                es_sobreentrega=es_sobreentrega,
                motivo_override=motivo_override,
                request=request,
            )
            
            parcialidad = resultado_merge['parcialidad']
            fue_merge = resultado_merge['merged']
            
            # Recalcular estado DESPUÉS de agregar (sin N+1, una sola query)
            total_nuevo = LoteParcialidad.objects.filter(lote=lote_locked).aggregate(
                total=Sum('cantidad')
            )['total'] or 0
            
            estado_nuevo = self._calcular_estado_contrato(total_nuevo, cantidad_contrato)
            
            # Calcular métricas del lote
            pendiente_lote = max(0, cantidad_contrato - total_nuevo) if cantidad_contrato > 0 else None
            porcentaje_lote = min(100, (total_nuevo / cantidad_contrato) * 100) if cantidad_contrato > 0 else 0
            excedente_lote = max(0, total_nuevo - cantidad_contrato) if cantidad_contrato > 0 and total_nuevo > cantidad_contrato else None
            
            # Calcular estado del contrato GLOBAL (si aplica)
            estado_global_info = self._calcular_estado_contrato_global(lote_locked)
            
            # Registrar en AuditLog PERSISTENTE si es sobre-entrega (SOX/ISO 27001)
            audit_info = None
            if estado_previo in ['CUMPLIDO', 'SOBREENTREGA'] and request.data.get('override'):
                motivo = request.data.get('motivo_override', 'No especificado')
                audit_info = {
                    'accion': 'OVERRIDE_SOBREENTREGA',
                    'usuario': user.username,
                    'lote': lote_locked.numero_lote,
                    'lote_id': lote_locked.pk,
                    'parcialidad_id': parcialidad.pk,
                    'cantidad_agregada': parcialidad.cantidad,
                    'fecha_entrega': str(parcialidad.fecha_entrega),
                    'total_antes': total_actual,
                    'total_despues': total_nuevo,
                    'cantidad_contrato': cantidad_contrato,
                    'cantidad_contrato_global': lote_locked.cantidad_contrato_global,
                    'excedente': excedente_lote,
                    'motivo': motivo,
                    'estado_previo': estado_previo,
                    'estado_nuevo': estado_nuevo,
                }
                
                # P0-1 FIX: Persistir AuditLog en BD (obligatorio SOX/ISO 27001)
                # Dentro de transaction.atomic para garantizar consistencia
                try:
                    from core.models import AuditoriaLogs
                    
                    # Extraer IP del request
                    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
                    ip_address = x_forwarded_for.split(',')[0].strip() if x_forwarded_for else request.META.get('REMOTE_ADDR', '')
                    
                    AuditoriaLogs.objects.create(
                        usuario=user,
                        accion='OVERRIDE_SOBREENTREGA',
                        modelo='LoteParcialidad',
                        objeto_id=str(parcialidad.pk),
                        datos_anteriores={
                            'total_entregado': total_actual,
                            'estado': estado_previo,
                        },
                        datos_nuevos={
                            'total_entregado': total_nuevo,
                            'estado': estado_nuevo,
                            'parcialidad_cantidad': parcialidad.cantidad,
                            'es_sobreentrega': True,
                        },
                        ip_address=ip_address[:45] if ip_address else None,
                        user_agent=request.META.get('HTTP_USER_AGENT', '')[:200] if request.META.get('HTTP_USER_AGENT') else None,
                        detalles={
                            'motivo_override': motivo,
                            'lote_id': lote_locked.pk,
                            'lote_numero': lote_locked.numero_lote,
                            'cantidad_contrato': cantidad_contrato,
                            'cantidad_contrato_global': lote_locked.cantidad_contrato_global,
                            'excedente': excedente_lote,
                            'fecha_entrega': str(parcialidad.fecha_entrega),
                        }
                    )
                    logger.info(
                        f"[AUDIT-DB] OVERRIDE_SOBREENTREGA persistido - Lote: {lote_locked.numero_lote}, "
                        f"Parcialidad: {parcialidad.pk}, Usuario: {user.username}"
                    )
                except Exception as audit_error:
                    # Si falla el audit log, hacer rollback de toda la transacción
                    # Esto garantiza que no queden parcialidades sin auditoría
                    logger.error(
                        f"[AUDIT-ERROR] Fallo al persistir AuditLog - Lote: {lote_locked.numero_lote}, "
                        f"Error: {audit_error}. Revirtiendo transacción."
                    )
                    raise Exception(f"Error crítico de auditoría: {audit_error}. Operación cancelada.")
                
                # Log adicional a Python logger (no reemplaza BD, es complemento)
                logger.warning(
                    f"[AUDIT] SOBREENTREGA AUTORIZADA - Lote: {lote_locked.numero_lote}, "
                    f"Usuario: {user.username}, Cantidad: {parcialidad.cantidad}, "
                    f"Excedente: {excedente_lote}, Motivo: {motivo}"
                )
            
            logger.info(
                f"Parcialidad {'MERGED' if fue_merge else 'CREATED'} - Lote: {lote_locked.numero_lote}, "
                f"Cantidad: {parcialidad.cantidad}, Fecha: {parcialidad.fecha_entrega}, "
                f"Usuario: {user.username}, Estado: {estado_previo} -> {estado_nuevo}"
            )
            
            # Construir datos de la parcialidad para la respuesta
            parcialidad_data = {
                'id': parcialidad.pk,
                'lote': lote_locked.pk,
                'fecha_entrega': str(parcialidad.fecha_entrega),
                'cantidad': parcialidad.cantidad,
                'numero_factura': parcialidad.numero_factura,
                'numero_remision': parcialidad.numero_remision,
                'proveedor': parcialidad.proveedor,
                'notas': parcialidad.notas,
                'es_sobreentrega': parcialidad.es_sobreentrega,
            }
            
            # Construir respuesta completa
            mensaje_base = 'Entrega consolidada con registro existente' if fue_merge else 'Entrega registrada correctamente'
            response_data = {
                'mensaje': mensaje_base,
                'parcialidad': parcialidad_data,
                'merge_info': {
                    'fue_merge': fue_merge,
                    'nivel_match': resultado_merge.get('nivel_match', ''),
                    'cantidad_antes': resultado_merge.get('cantidad_antes', 0),
                    'campos_rellenados': resultado_merge.get('campos_rellenados', []),
                } if fue_merge else None,
                'estado_contrato': {
                    'lote': {
                        'cantidad_contrato': cantidad_contrato or None,
                        'total_entregado': total_nuevo,
                        'pendiente': pendiente_lote,
                        'porcentaje': round(porcentaje_lote, 2),
                        'estado': estado_nuevo,
                        'estado_anterior': estado_previo,
                        'excedente': excedente_lote,
                    },
                    'global': estado_global_info
                }
            }
            
            # Agregar advertencias según el estado
            if estado_nuevo == 'SOBREENTREGA':
                response_data['advertencia'] = (
                    f'SOBREENTREGA: El total entregado ({total_nuevo}) excede el contrato ({cantidad_contrato}) '
                    f'por {excedente_lote} unidades.'
                )
                response_data['audit_log'] = audit_info
            elif estado_nuevo == 'CUMPLIDO' and estado_previo != 'CUMPLIDO':
                response_data['info'] = '¡Contrato CUMPLIDO! Se ha alcanzado la cantidad contratada.'
            
            return Response(response_data, status=status.HTTP_201_CREATED)
    
    def _calcular_estado_contrato(self, total_entregado, cantidad_contrato):
        """
        Calcula el estado del contrato basado en entregas vs contrato.
        
        Estados:
        - SIN_CONTRATO: No hay cantidad_contrato definida
        - PENDIENTE: total = 0
        - PARCIAL: 0 < total < contrato
        - CUMPLIDO: total == contrato
        - SOBREENTREGA: total > contrato
        """
        if not cantidad_contrato or cantidad_contrato <= 0:
            return 'SIN_CONTRATO'
        if total_entregado <= 0:
            return 'PENDIENTE'
        if total_entregado < cantidad_contrato:
            return 'PARCIAL'
        if total_entregado == cantidad_contrato:
            return 'CUMPLIDO'
        return 'SOBREENTREGA'
    
    def _calcular_estado_contrato_global(self, lote):
        """
        Calcula el estado del contrato global sumando parcialidades de todos los lotes.
        Usa numero_contrato + producto para agrupar lotes del mismo contrato global.
        
        IMPORTANTE: El contrato global es por CLAVE DE PRODUCTO + NÚMERO DE CONTRATO.
        Busca cantidad_contrato_global de CUALQUIER lote del mismo grupo.
        
        MEJORA: Si no hay numero_contrato pero sí tiene cantidad_contrato_global,
        usar solo el campo del lote actual (para casos de lote único).
        """
        # Caso 1: El lote tiene cantidad_contrato_global directamente
        if lote.cantidad_contrato_global and lote.cantidad_contrato_global > 0:
            # Calcular total entregado solo para este lote si no tiene contrato
            if not lote.numero_contrato:
                total_entregado = LoteParcialidad.objects.filter(
                    lote=lote
                ).aggregate(total=Sum('cantidad'))['total'] or 0
                
                estado = self._calcular_estado_contrato(total_entregado, lote.cantidad_contrato_global)
                pendiente = max(0, lote.cantidad_contrato_global - total_entregado)
                porcentaje = min(100, (total_entregado / lote.cantidad_contrato_global) * 100) if lote.cantidad_contrato_global > 0 else 0
                excedente = max(0, total_entregado - lote.cantidad_contrato_global) if total_entregado > lote.cantidad_contrato_global else None
                
                return {
                    'numero_contrato': 'Sin número de contrato',
                    'cantidad_contrato_global': lote.cantidad_contrato_global,
                    'total_entregado': total_entregado,
                    'pendiente': pendiente,
                    'porcentaje': round(porcentaje, 2),
                    'estado': estado,
                    'excedente': excedente,
                    'num_lotes': 1,
                }
        
        # Caso 2: Buscar por numero_contrato + producto
        if not lote.numero_contrato:
            return None
        
        # Buscar todos los lotes con el mismo contrato y producto (contrato global)
        lotes_ccg = Lote.objects.filter(
            numero_contrato=lote.numero_contrato,
            producto=lote.producto,
            activo=True
        )
        
        if not lotes_ccg.exists():
            return None
        
        # Buscar cantidad_contrato_global de CUALQUIER lote del grupo
        # Puede que este lote específico no lo tenga pero otro del mismo contrato sí
        cantidad_contrato_global = max(
            (l.cantidad_contrato_global or 0) for l in lotes_ccg
        )
        
        # Si ningún lote tiene CCG definido, retornar info básica con 0
        if cantidad_contrato_global <= 0:
            # Aún retornar info del contrato para mostrar al usuario
            total_entregado = LoteParcialidad.objects.filter(
                lote__numero_contrato=lote.numero_contrato,
                lote__producto=lote.producto,
                lote__activo=True
            ).aggregate(total=Sum('cantidad'))['total'] or 0
            
            return {
                'numero_contrato': lote.numero_contrato,
                'cantidad_contrato_global': None,  # No definido
                'total_entregado': total_entregado,
                'pendiente': None,
                'porcentaje': None,
                'estado': 'SIN_CONTRATO_GLOBAL',
                'excedente': None,
                'num_lotes': lotes_ccg.count(),
            }
        
        # Total entregado = suma de TODAS las parcialidades de lotes del contrato
        total_entregado_global = LoteParcialidad.objects.filter(
            lote__numero_contrato=lote.numero_contrato,
            lote__producto=lote.producto,
            lote__activo=True
        ).aggregate(total=Sum('cantidad'))['total'] or 0
        
        estado_global = self._calcular_estado_contrato(total_entregado_global, cantidad_contrato_global)
        pendiente_global = max(0, cantidad_contrato_global - total_entregado_global)
        porcentaje_global = min(100, (total_entregado_global / cantidad_contrato_global) * 100)
        excedente_global = max(0, total_entregado_global - cantidad_contrato_global) if total_entregado_global > cantidad_contrato_global else None
        
        return {
            'numero_contrato': lote.numero_contrato,
            'cantidad_contrato_global': cantidad_contrato_global,
            'total_entregado': total_entregado_global,
            'pendiente': pendiente_global,
            'porcentaje': round(porcentaje_global, 2),
            'estado': estado_global,
            'excedente': excedente_global,
            'num_lotes': lotes_ccg.count(),
        }
    
    @action(detail=True, methods=['delete'], url_path='eliminar-parcialidad/(?P<parcialidad_id>[0-9]+)')
    def eliminar_parcialidad(self, request, pk=None, parcialidad_id=None):
        """
        Elimina una parcialidad específica del lote.
        Solo permite eliminar parcialidades recientes (< 7 días).
        Retorna estado actualizado del contrato.
        """
        lote = self.get_object()
        
        # Validar permisos de escritura
        user = request.user
        if not is_farmacia_or_admin(user):
            user_centro = get_user_centro(user)
            lote_centro = lote.centro
            if lote_centro is None or (user_centro and lote_centro.pk != user_centro.pk):
                return Response(
                    {'error': 'No tiene permisos para eliminar entregas de este lote'},
                    status=status.HTTP_403_FORBIDDEN
                )
        
        # Buscar parcialidad
        try:
            parcialidad = LoteParcialidad.objects.get(id=parcialidad_id, lote=lote)
        except LoteParcialidad.DoesNotExist:
            return Response(
                {'error': 'Parcialidad no encontrada'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Validar que sea reciente (< 7 días)
        from django.utils import timezone
        dias_desde_creacion = (timezone.now() - parcialidad.created_at).days
        if dias_desde_creacion > 7:
            return Response({
                'error': 'No se puede eliminar una parcialidad con más de 7 días de antigüedad',
                'dias_desde_creacion': dias_desde_creacion
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Usar transacción atómica
        with transaction.atomic():
            cantidad = parcialidad.cantidad
            fecha = parcialidad.fecha_entrega
            parcialidad.delete()
            
            # Calcular estado actualizado del contrato
            from django.db.models import Sum
            total_parcialidades = LoteParcialidad.objects.filter(lote=lote).aggregate(
                total=Sum('cantidad')
            )['total'] or 0
            
            # Estado del contrato de lote
            estado_lote = 'sin_contrato'
            pendiente_lote = None
            if lote.cantidad_contrato:
                pendiente_lote = max(0, lote.cantidad_contrato - total_parcialidades)
                if total_parcialidades >= lote.cantidad_contrato:
                    estado_lote = 'cumplido'
                elif total_parcialidades > 0:
                    estado_lote = 'parcial'
                else:
                    estado_lote = 'pendiente'
            
            logger.info(
                f"Parcialidad eliminada - Lote: {lote.numero_lote}, "
                f"Cantidad: {cantidad}, Fecha: {fecha}, Usuario: {user.username}, "
                f"Nuevo estado contrato: {estado_lote}"
            )
            
            return Response({
                'mensaje': 'Entrega eliminada correctamente',
                'estado_contrato': {
                    'lote': {
                        'cantidad_contrato': lote.cantidad_contrato,
                        'total_recibido': total_parcialidades,
                        'pendiente': pendiente_lote,
                        'estado': estado_lote,
                    }
                }
            })

    # =========================================================================
    # ADMIN: Crear parcialidades retroactivas para lotes sin historial
    # =========================================================================
    @action(detail=False, methods=['post'], url_path='crear-parcialidades-retroactivas')
    def crear_parcialidades_retroactivas(self, request):
        """
        ADMIN: Crea parcialidades iniciales para lotes que no tienen ninguna.
        
        Este endpoint es necesario para lotes importados antes del sistema de
        entregas parciales. Crea una parcialidad inicial con la cantidad_inicial
        del lote.
        
        También actualiza fecha_fabricacion de lotes que la tengan vacía,
        usando created_at como referencia.
        
        Solo usuarios admin/farmacia pueden ejecutar esta acción.
        
        Returns:
            - lotes_procesados: número de lotes que recibieron parcialidad
            - lotes_ya_tenian: número de lotes que ya tenían parcialidad
            - lotes_sin_cantidad: lotes con cantidad_inicial = 0 (ignorados)
            - lotes_fecha_actualizada: lotes con fecha_fabricacion actualizada
            - errores: lista de errores si los hubo
        """
        user = request.user
        
        # Solo admin/farmacia pueden ejecutar
        if not is_farmacia_or_admin(user):
            return Response(
                {'error': 'Solo administradores pueden ejecutar esta acción'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        try:
            from django.db import transaction
            from django.utils import timezone
            
            # Estadísticas
            lotes_procesados = 0
            lotes_ya_tenian = 0
            lotes_sin_cantidad = 0
            lotes_fecha_actualizada = 0
            errores = []
            detalles = []
            
            # Obtener todos los lotes activos
            lotes = Lote.objects.filter(activo=True)
            
            with transaction.atomic():
                for lote in lotes:
                    # Verificar si ya tiene parcialidades
                    ya_tiene_parcialidades = LoteParcialidad.objects.filter(lote=lote).exists()
                    
                    if ya_tiene_parcialidades:
                        lotes_ya_tenian += 1
                        # Aun así, actualizar fecha_fabricacion si está vacía
                        if not lote.fecha_fabricacion and lote.created_at:
                            lote.fecha_fabricacion = lote.created_at.date()
                            lote.save(update_fields=['fecha_fabricacion'])
                            lotes_fecha_actualizada += 1
                        continue
                    
                    # Ignorar lotes sin cantidad inicial
                    if not lote.cantidad_inicial or lote.cantidad_inicial <= 0:
                        lotes_sin_cantidad += 1
                        continue
                    
                    try:
                        # Usar fecha_fabricacion o created_at o fecha actual
                        fecha_entrega = lote.fecha_fabricacion
                        actualizar_fecha_fab = False
                        
                        if not fecha_entrega:
                            # Si no hay fecha_fabricacion, usar created_at o hoy
                            if lote.created_at:
                                fecha_entrega = lote.created_at.date()
                            else:
                                fecha_entrega = timezone.now().date()
                            actualizar_fecha_fab = True
                        
                        # Crear parcialidad inicial
                        LoteParcialidad.objects.create(
                            lote=lote,
                            fecha_entrega=fecha_entrega,
                            cantidad=lote.cantidad_inicial,
                            notas='Carga inicial retroactiva (lote preexistente sin historial)',
                            usuario=user,
                            es_sobreentrega=False
                        )
                        
                        # Actualizar fecha_fabricacion si estaba vacía
                        if actualizar_fecha_fab:
                            lote.fecha_fabricacion = fecha_entrega
                            lote.save(update_fields=['fecha_fabricacion'])
                            lotes_fecha_actualizada += 1
                        
                        lotes_procesados += 1
                        detalles.append({
                            'lote_id': lote.id,
                            'numero_lote': lote.numero_lote,
                            'producto': lote.producto.clave if lote.producto else 'N/A',
                            'cantidad': lote.cantidad_inicial,
                            'fecha': str(fecha_entrega),
                            'fecha_actualizada': actualizar_fecha_fab
                        })
                        
                    except Exception as e:
                        errores.append({
                            'lote_id': lote.id,
                            'numero_lote': lote.numero_lote,
                            'error': str(e)
                        })
            
            logger.info(
                f"[PARCIALIDADES RETROACTIVAS] Usuario {user.username}: "
                f"procesados={lotes_procesados}, ya_tenian={lotes_ya_tenian}, "
                f"sin_cantidad={lotes_sin_cantidad}, fechas_actualizadas={lotes_fecha_actualizada}, "
                f"errores={len(errores)}"
            )
            
            return Response({
                'mensaje': f'Proceso completado: {lotes_procesados} lotes con parcialidades, {lotes_fecha_actualizada} fechas actualizadas',
                'resumen': {
                    'lotes_procesados': lotes_procesados,
                    'lotes_ya_tenian': lotes_ya_tenian,
                    'lotes_sin_cantidad': lotes_sin_cantidad,
                    'lotes_fecha_actualizada': lotes_fecha_actualizada,
                    'total_errores': len(errores)
                },
                'detalles': detalles[:50],  # Limitar detalles mostrados
                'errores': errores[:10] if errores else []
            })
            
        except Exception as e:
            logger.exception(f"Error en crear_parcialidades_retroactivas: {e}")
            return Response(
                {'error': 'Error al crear parcialidades', 'mensaje': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
